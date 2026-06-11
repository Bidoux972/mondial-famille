"""
DISTRICARB HUB v0.5 — Cockpit de pilotage station B. DISTRICARB SARL
=====================================================================
Bugs corrigés, contenu enrichi, autonomie intelligente.
"""
import os,sys,json,traceback,shutil,tempfile,warnings,time,subprocess,re,threading
import tkinter as tk
from datetime import datetime,date,timedelta,time as dt_time
from pathlib import Path
from tkinter import filedialog,messagebox
from http.server import HTTPServer,BaseHTTPRequestHandler
from socketserver import ThreadingMixIn

# =============================================================================
# CŒUR MÉTIER COMMUN — districarb_core
# =============================================================================
# Le HUB Python (cette interface customtkinter) ET le futur backend FastAPI du
# HUB Web partagent le même cœur métier `districarb_core` (Python pur, sans UI).
# Toute évolution métier se fait dans le core d'abord, puis les deux interfaces
# en bénéficient. Voir architecture validée Bidou 26/05/2026.
#
# Extraction progressive : ce que le core fournit aujourd'hui (Étape 3,
# 27/05/2026) :
#   - martinique : helpers calendrier Martinique (jours fériés, livrables, …)
#   - trous      : détection + qualification Pont/Weekend des trous de livraison
# Plus de modules suivront aux phases ultérieures (readers, projections, …).
# =============================================================================
from districarb_core.martinique import (
    _easter,
    get_feries_martinique,
    is_ferie,
    jour_de_commande,
    nb_jours_livrables_avant,
    nom_ferie as _nom_ferie,  # alias compatibilité ancien nom interne
)
from districarb_core.trous import detecter_trous, qualifier_trou


# =============================================================================
# LOGGER SILENT ERRORS (introduit le 11/05/2026)
# =============================================================================
# Remplace les `except: pass` qui rendaient les bugs invisibles. Le hub continue
# son flux après l'erreur (continuité préservée), MAIS on garde la trace dans
# `errors.log` avec timestamp + localisation (fichier:ligne:fonction). Permet de
# debugger des comportements bizarres a posteriori.
# =============================================================================
_ERRORS_LOG_PATH = None  # initialisé tardivement, voir _init_errors_log
_ERRORS_LOG_MAX_BYTES = 5 * 1024 * 1024  # 5 Mo : seuil au-delà duquel on rotate

def _init_errors_log():
    """Initialise le chemin de errors.log à côté du script. Appelé tardivement
    (au premier _log_silent_err) pour ne pas dépendre de l'ordre d'init.
    
    FIX 21/05/2026 : ROTATION automatique. Si errors.log dépasse 5 Mo, on
    archive l'ancien en errors.log.old (écrasement de l'ancien archive s'il
    existe) et on repart d'un fichier vide. Évite l'effet boule de neige
    constaté : 99 Mo sur le PC de Bidou, ralentissait le PC + saturait OneDrive."""
    global _ERRORS_LOG_PATH
    if _ERRORS_LOG_PATH is not None: return
    try:
        script_dir = Path(__file__).parent.resolve()
        _ERRORS_LOG_PATH = script_dir / "errors.log"
        # Rotation si fichier trop gros
        try:
            if _ERRORS_LOG_PATH.exists() and _ERRORS_LOG_PATH.stat().st_size > _ERRORS_LOG_MAX_BYTES:
                old=_ERRORS_LOG_PATH.with_suffix(".log.old")
                try:
                    if old.exists(): old.unlink()
                except Exception: pass
                try: _ERRORS_LOG_PATH.rename(old)
                except Exception: pass  # si rename échoue, on continue, le fichier restera gros temporairement
        except Exception: pass
    except Exception:
        _ERRORS_LOG_PATH = Path("errors.log")

def _log_silent_err(context="", exc=None):
    """Trace silencieuse d'une exception attrapée par un except: pass historique.
    Écrit dans `errors.log` à côté du hub, sans casser le flux (ne lève rien).
    
    Args:
        context: étiquette manuelle (ex. "calc_autonomie") ou vide pour auto.
        exc: l'exception attrapée (transmise via `as e`).
    """
    try:
        _init_errors_log()
        import traceback as _tb
        # Remonter la pile pour trouver la frame de l'appelant (skip _log_silent_err lui-même)
        stack = _tb.extract_stack(limit=4)
        frame = stack[-2] if len(stack) >= 2 else None
        loc = f"{Path(frame.filename).name}:{frame.lineno}:{frame.name}" if frame else "?"
        ctx = f" [{context}]" if context else ""
        exc_part = f"{type(exc).__name__}: {exc}" if exc else "(no exc)"
        msg = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {loc}{ctx} - {exc_part}\n"
        try:
            with open(_ERRORS_LOG_PATH, "a", encoding="utf-8") as f:
                f.write(msg)
        except Exception: pass  # log d'erreur du log : on abandonne silencieusement
    except Exception: pass  # garde-fou ultime : ne JAMAIS casser le flux du hub

try:
    from PIL import Image,ImageDraw,ImageFilter,ImageFont
    PIL_OK=True
except ImportError:
    PIL_OK=False
warnings.filterwarnings("ignore",category=UserWarning)
try: import customtkinter as ctk
except ImportError: print("pip install customtkinter openpyxl"); sys.exit(1)
try: import openpyxl
except ImportError: print("pip install customtkinter openpyxl"); sys.exit(1)

C={"bg":"#0D1219","panel":"#111820","card":"#171F2A","card_h":"#1E2836","border":"#252F3C","border2":"#344053",
   "t1":"#D4C8B0","t2":"#8A97A8","t3":"#56647A","red":"#E82530","gold":"#F2B530","green":"#2DA84A",
   "amber":"#E8963A","blue":"#4A9EFF","teal":"#2BB5A0","alert_bg":"#1A1215","alert_border":"#3A1520",
   # Couleurs d'identité des vignettes (désaturées, plus éditoriales, aucun doublon)
   "vig_red":"#C94A52","vig_blue":"#5B92D4","vig_gold":"#E4BC4D",
   "vig_green":"#4FAE5F","vig_teal":"#3FB5A3","vig_amber":"#D4934A"}

# Police pour les chiffres et valeurs : Space Grotesk si installée, sinon Segoe UI
# Space Grotesk se télécharge gratuitement sur https://fonts.google.com/specimen/Space+Grotesk
# Installation Windows : double-clic sur les fichiers .ttf téléchargés → "Installer"
FONT_NUM="Segoe UI"  # fallback par défaut, remplacé au lancement si Space Grotesk est détectée
def _detect_num_font():
    global FONT_NUM
    try:
        from tkinter import font as _tkfont
        # Nécessite qu'une fenêtre Tk existe déjà — appelé après la création de CTk()
        families=_tkfont.families()
        if "Space Grotesk" in families:
            FONT_NUM="Space Grotesk"
            print("[UI] Police Space Grotesk détectée — utilisée pour les chiffres.")
        else:
            print("[UI] Space Grotesk non détectée — chiffres en Segoe UI. Installation optionnelle : https://fonts.google.com/specimen/Space+Grotesk")
    except Exception as e:
        print(f"[UI] Détection police échouée : {e} — fallback Segoe UI")
HUB_FILES=[
    {"key":"gest_piste","label":"GEST PISTE","sub":"Boutique & caisse","icon":"\u26fd","color":C["vig_red"]},
    {"key":"cartes","label":"CARTES","sub":"T\u00e9l\u00e9collectes CB/CP","icon":"\U0001f4b3","color":C["vig_blue"]},
    {"key":"prevision","label":"PR\u00c9VISION","sub":"Stocks & commandes","icon":"\U0001f4ca","color":C["vig_gold"]},
    {"key":"objectif","label":"OBJECTIF","sub":"Alertes & pilotage","icon":"\U0001f6a8","color":C["vig_green"]},
    {"key":"litrage","label":"LITRAGE","sub":"Performance & historique","icon":"\U0001f4da","color":C["vig_teal"]}]

# ============================================================
# DOSSIER DE L'APP : ~/.districarb_hub (inchangé pour préserver tes données)
# Toutes tes données existantes sont là : livraisons, snapshots, acks, cycle, etc.
#
# CHANGEMENT D'EXTENSION : .json → .cfg
# ESET (antivirus) bloque l'écriture des .json par Python (protection ransomware).
# En passant en .cfg, ESET arrête de bloquer. Le contenu reste du JSON valide,
# seule l'extension change.
#
# RÉTROCOMPATIBILITÉ : pour chaque fichier de config, on lit en priorité le .cfg ;
# s'il n'existe pas, on lit l'ancien .json. À la première écriture, le .cfg est créé.
# Tu peux supprimer manuellement les anciens .json plus tard quand tout fonctionne.
# ============================================================
APP_DIR=Path.home()/".districarb_hub"
CONFIG_FILE=APP_DIR/"config.cfg"
LIVRAISON_FILE=APP_DIR/"livraisons.cfg"
LIVRAISONS_LOG_FILE=APP_DIR/"livraisons_log.cfg"
CYCLE_FILE=APP_DIR/"cycle.cfg"
ECARTS_FILE=APP_DIR/"ecarts_resolus.cfg"
ANTIRUPTURE_ACK_FILE=APP_DIR/"antirupture_ack.cfg"
TENDANCE_ACK_FILE=APP_DIR/"tendance_ack.cfg"
SAISIES_IRR_ACK_FILE=APP_DIR/"saisies_irrealistes_ack.cfg"
LIVR_REPORT_ACK_FILE=APP_DIR/"livraisons_reporter_ack.cfg"
POPUP_SILENCE_FILE=APP_DIR/"popup_silence.cfg"
RAPPORT_PROMPT_FILE=APP_DIR/"rapport_prompt.cfg"
# Options sélectionnées du rapport mensuel (cases à cocher mémorisées d'une fois sur l'autre)
RAPPORT_OPTIONS_FILE=APP_DIR/"rapport_options.cfg"
# Mapping {YYYY-MM : chemin du fichier Objectif de ce mois}, mémorisé quand l'utilisateur
# sélectionne manuellement un fichier Objectif pour un mois passé (cas où la déduction
# automatique a échoué). Évite de redemander à chaque génération.
OBJECTIF_PATHS_FILE=APP_DIR/"objectif_paths.cfg"
# Journal d'événements horodatés (ponts traversés, anomalies confirmées, ruptures effectives,
# livraisons reçues, acquittements). Un seul fichier permanent qui contient toute l'historique.
# Utilisé par le rapport mensuel pour les sections "Événements remarquables" et par le bouton
# "Journal des événements" du hub. Voir EvenementsManager pour la structure et les helpers.
EVENEMENTS_FILE=APP_DIR/"evenements.cfg"
# Livraisons sur jour non-livrable que Bidou a déclarées VOLONTAIRES et ASSUMÉES
# (ex : Lundi de Pentecôte → SARA fermée → livraison décalée au samedi, décision
# prise et validée par Bidou). Une date forcée n'est plus signalée comme
# incohérence par le moteur anti-rupture. Structure :
# {"2026-05-23": {"ts": "...", "note": "exception assumée", "vol": 32000}}
FORCAGE_FILE=APP_DIR/"forcages_livraison.cfg"
# Commandes carburant passées par Bidou chez TotalEnergies (saisies via le
# bouton "Commande du jour"). Le hub ne voit jamais la commande côté Total :
# cette saisie est sa seule trace. Lue plus tard par la fenêtre du matin
# (Brique 2). Clé = date cible de livraison ISO 'YYYY-MM-DD'. Structure :
# {"2026-05-20": {"sp":16000,"go":12000,"gnr":0,"tour":1,
#                 "premier_voyage":false,"ts":"..."}}
COMMANDE_FILE=APP_DIR/"commandes.cfg"
# Historique des prix mensuels (prix achat + prix vente + marge) pour calculer
# l'effet spéculation : chaque début de mois, on capture les nouveaux prix et on
# compare aux prix du mois précédent pour estimer le gain/perte sur le stock pivot.
# Structure : {"2026-04": {"pa_sp":..., "pv_sp":..., "marge":...}, "2026-05": {...}}
PRIX_HISTO_FILE=APP_DIR/"prix_historique.cfg"
# Fichier debug : placé à côté du script (HubDistricarb/) pour accès facile via Explorateur
DEBUG_LOG_FILE=Path(__file__).parent/"debug_antirupture.log"
# Snapshots mensuels (JSON contenu, mais l'extension peut rester .json car la protection
# ESET ne s'applique généralement pas dans les sous-dossiers ; à surveiller)
SNAPSHOTS_DIR=APP_DIR/"snapshots"
# Rapports PDF/HTML générés (visibles, imprimables, partageables)
RAPPORTS_DIR=Path.home()/"Documents"/"DISTRICARB Rapports"

def debug_log(msg):
    """Écrit un message dans le fichier debug_antirupture.log avec horodatage."""
    try:
        with open(DEBUG_LOG_FILE,"a",encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")
    except Exception as _e: _log_silent_err(exc=_e)
# =============================================================================
# GESTION DU SILENCE DES POPUPS (acquittement par snapshot d'alertes)
# =============================================================================
# Principe : quand l'utilisateur clique un bouton de fermeture (snooze, "C'est noté",
# "OK je gère"), on enregistre :
#   1. Un snapshot fingerprint de TOUTES les alertes visibles au moment du clic
#   2. Une date d'expiration (ex: lundi 6h, ou +N minutes)
# À chaque refresh, on calcule le snapshot actuel des alertes pour le type de popup,
# et on compare avec le snapshot acquitté :
#   - Si tout l'actuel est déjà dans l'acquitté ET pas expiré → silence
#   - Si nouvelle alerte ou aggravation détectée → popup s'ouvre (avec mention)
#   - Si expiration passée → popup s'ouvre normalement
# =============================================================================
def _load_silence():
    """Charge le fichier de silence global. Retourne {} si absent ou corrompu."""
    try: return load_json(POPUP_SILENCE_FILE) or {}
    except Exception as _e: _log_silent_err(exc=_e); return {}

def _save_silence(data):
    """Sauve le fichier de silence."""
    try: save_json(POPUP_SILENCE_FILE,data)
    except Exception as e: print(f"[silence popup] save error: {e}")

def silence_popup(popup_type,fingerprints,until_iso,extra=None):
    """Enregistre l'acquittement d'une popup avec son snapshot et son expiration.
    Args:
        popup_type: 'antirupture' | 'tendance' | 'saisies_irr'
        fingerprints: liste/set des fingerprints d'alertes acquittées (chaînes)
        until_iso: date ISO d'expiration ('2026-05-04T06:00:00')
        extra: dict optionnel (ex: pour anti-rupture, niveaux de manques par fingerprint)
    """
    data=_load_silence()
    data[popup_type]={
        "fingerprints":sorted(set(fingerprints)),
        "until_iso":until_iso,
        "ack_at_iso":datetime.now().isoformat(),
        "extra":extra or {},
    }
    _save_silence(data)

def is_popup_silenced(popup_type,current_fingerprints,current_extra=None,aggravation_check=None):
    """Détermine si la popup doit rester silencieuse au refresh.
    Args:
        popup_type: 'antirupture' | 'tendance' | 'saisies_irr'
        current_fingerprints: set/liste des fingerprints actuellement actifs
        current_extra: dict optionnel des valeurs actuelles (ex: manques par fingerprint)
        aggravation_check: callable optionnel (fp, old_extra, new_extra) -> bool ;
                           retourne True si aggravation détectée
    Retourne:
        True si silence (popup ne doit pas s'ouvrir),
        False sinon (nouvelle alerte / aggravation / expiration → popup s'ouvre)
    """
    data=_load_silence().get(popup_type)
    if not data: return False
    # Expiration ?
    try:
        until=datetime.fromisoformat(data["until_iso"])
        if datetime.now()>=until:
            return False  # expiré : popup peut s'ouvrir
    except Exception as _e: _log_silent_err(exc=_e); return False  # date corrompue : on ne silence pas
    # Nouveauté ?
    silenced_fps=set(data.get("fingerprints",[]))
    current_fps=set(current_fingerprints)
    new_alerts=current_fps-silenced_fps
    if new_alerts:
        return False  # une alerte non-acquittée → popup s'ouvre
    # Aggravation ?
    if aggravation_check is not None and current_extra is not None:
        old_extra=data.get("extra") or {}
        for fp in current_fps & silenced_fps:
            if aggravation_check(fp,old_extra,current_extra):
                return False  # aggravation → popup s'ouvre
    return True  # tout est OK et acquitté → silence

def clear_popup_silence(popup_type):
    """Efface le silence d'un type de popup (utilisé après changement intentionnel)."""
    data=_load_silence()
    if popup_type in data:
        del data[popup_type]
        _save_silence(data)

def _antirupture_fps_extra(ar):
    """Calcule (fingerprints, extra) pour le silence de la popup anti-rupture
    à partir de la structure de retour `ar` de read_all().
    fingerprints = liste des pont_id ; extra = {pont_id: {carb: manque_l, ...}}
    """
    fps=[];extra={}
    ruptures=ar.get("ruptures_dans_trou",[]) or []
    for r in ruptures:
        try:
            pont_id=f"pont_{r['trou_start'].strftime('%d%m%Y')}"
        except Exception as _e: _log_silent_err(exc=_e); continue
        if pont_id not in extra:
            extra[pont_id]={}
            fps.append(pont_id)
        carb=str(r.get("carburant","")).lower()
        manque=int(r.get("manque",0) or 0)
        # Cumuler par carburant si plusieurs ruptures du même pont/carb
        extra[pont_id][carb]=extra[pont_id].get(carb,0)+manque
    return fps,extra

# =============================================================================
# SECTIONS DU RAPPORT MENSUEL — liste des sections optionnelles à cocher dans le dialogue
# Chaque entrée = (clé interne, label affiché, valeur par défaut)
# La clé est utilisée pour stocker la sélection dans rapport_options.cfg
# =============================================================================
RAPPORT_SECTIONS=[
    ("synthese","Synth\u00e8se du mois (KPI litrage / CA piste / boutique / total)",True),
    ("carburants","R\u00e9partition par carburant",True),
    ("encaissements","Encaissements (CB / CP / Esp\u00e8ces)",True),
    ("admin","Pilotage administratif & alertes (Balance D/E, retards, impay\u00e9s)",False),
    ("detail_jours","D\u00e9tail jour par jour",True),
    ("top3_piste_meilleures","Top 3 meilleures journ\u00e9es (CA piste)",True),
    ("top3_piste_pires","Top 3 plus faibles journ\u00e9es (CA piste)",True),
    ("top3_bout_meilleures","Top 3 meilleures journ\u00e9es (CA boutique)",False),
    ("top3_bout_pires","Top 3 plus faibles journ\u00e9es (CA boutique)",False),
    ("anomalies","Anomalies de tendance d\u00e9tect\u00e9es dans le mois",False),
    ("ponts","Ponts travers\u00e9s avec statut",False),
]

def load_rapport_options():
    """Charge la dernière sélection de sections du rapport, ou les défauts si jamais sauvé."""
    saved=load_json(RAPPORT_OPTIONS_FILE) or {}
    return {key:saved.get(key,default) for key,label,default in RAPPORT_SECTIONS}

def save_rapport_options(opts):
    """Sauvegarde la sélection courante (dict {clé: bool}) pour rappel à la prochaine ouverture."""
    try: save_json(RAPPORT_OPTIONS_FILE,opts)
    except Exception as e: print(f"[rapport options] save error: {e}")

def load_rapport_format_pref():
    """Charge la préférence de format du rapport ('auto' par défaut = comportement historique).
    Valeurs possibles : 'auto' (PDF prioritaire, fallback HTML), 'html', 'pdf'.
    Stockée dans RAPPORT_OPTIONS_FILE sous la clé '_format_pref' (préfixée pour ne pas
    entrer en collision avec les clés de sections de RAPPORT_SECTIONS)."""
    saved=load_json(RAPPORT_OPTIONS_FILE) or {}
    pref=saved.get("_format_pref","auto")
    return pref if pref in ("auto","html","pdf") else "auto"

def save_rapport_format_pref(pref):
    """Sauvegarde la préférence de format ('auto'/'html'/'pdf')."""
    try:
        saved=load_json(RAPPORT_OPTIONS_FILE) or {}
        saved["_format_pref"]=pref if pref in ("auto","html","pdf") else "auto"
        save_json(RAPPORT_OPTIONS_FILE,saved)
    except Exception as e: print(f"[rapport format pref] save error: {e}")


def resolve_objectif_path_for_month(year, month, current_objectif_path):
    """Détermine le chemin du fichier Objectif d'un mois spécifique.
    
    Stratégie :
    1. Si un mapping a été mémorisé pour ce mois → l'utiliser (et vérifier qu'il existe)
    2. Sinon, déduire automatiquement à partir du path actuel en remplaçant le nom de mois
       (pattern : 'Objectif mensuel <Mois> (<Année>).xlsx')
    3. Si le fichier déduit existe → renvoyer son path
    4. Sinon → renvoyer None (l'appelant demandera à l'utilisateur via file picker)
    
    Returns:
        str|None: Path du fichier si trouvé, None sinon
    """
    # 1. Mapping mémorisé
    saved=load_json(OBJECTIF_PATHS_FILE) or {}
    key=f"{year:04d}-{month:02d}"
    if key in saved:
        p=saved[key]
        if p and os.path.exists(p):
            return p
        # Path mémorisé mais fichier disparu → on continue avec la déduction
    # 2. Déduction automatique à partir du path actuel
    if not current_objectif_path:
        return None
    mois_fr=["Janvier","F\u00e9vrier","Mars","Avril","Mai","Juin","Juillet","Ao\u00fbt","Septembre","Octobre","Novembre","D\u00e9cembre"]
    target_mois=mois_fr[month-1]
    target_year=year
    # On scan le path actuel pour identifier le segment "Mois (Année)" ou "Mois Année"
    # Pattern attendu : "Objectif mensuel Mai (2026).xlsx"
    src_path=Path(current_objectif_path)
    src_name=src_path.name
    src_dir=src_path.parent
    # Détecter les noms de mois dans le nom de fichier actuel et les remplacer
    new_name=src_name
    replaced=False
    for m_idx,m_name in enumerate(mois_fr,start=1):
        if m_name in new_name:
            new_name=new_name.replace(m_name,target_mois)
            replaced=True
            break
    # Remplacer aussi l'année si différente
    for yr_candidate in range(2020,2031):
        yr_str=str(yr_candidate)
        if yr_str in new_name and yr_candidate!=target_year:
            new_name=new_name.replace(yr_str,str(target_year))
            break
    if not replaced:
        return None
    candidate=src_dir/new_name
    if candidate.exists():
        return str(candidate)
    return None


def remember_objectif_path(year, month, path):
    """Mémorise le chemin du fichier Objectif sélectionné manuellement pour un mois donné."""
    try:
        saved=load_json(OBJECTIF_PATHS_FILE) or {}
        saved[f"{year:04d}-{month:02d}"]=str(path)
        save_json(OBJECTIF_PATHS_FILE,saved)
    except Exception as e: print(f"[objectif paths] save error: {e}")


# =============================================================================
# JOURNAL DES ÉVÉNEMENTS (Sujet E)
# Un fichier permanent unique qui archive tous les événements horodatés du hub :
# ponts traversés, anomalies de tendance confirmées (>= 1h), ruptures effectives,
# livraisons reçues, acquittements de popups avec commentaire optionnel.
#
# Structure du JSON :
#   {"events": [
#       {"id": "uuid", "ts": "2026-05-12T11:32:00", "type": "pont", "data": {...}, "commentaire": "..."},
#       ...
#   ]}
#
# Les types d'événements :
#   - "pont"          : pont traversé (a). data = {date_debut, date_fin, duree, carburants_concernes, ack_auto/manuel}
#   - "anomalie"      : anomalie de tendance confirmée >= 1h (b). data = {carburant, ecart_pct, jour, debut, fin}
#   - "rupture"       : carburant tombé sous plancher physique (d). data = {carburant, jour, niveau, plancher}
#   - "livraison"     : livraison reçue (e). data = {jour, sp, go, gnr}
#   - "ack"           : acquittement d'une popup avec commentaire (f).
# =============================================================================

def _new_event_id():
    """Génère un ID unique court pour un événement (timestamp + 4 hex)."""
    import secrets
    return datetime.now().strftime("%Y%m%d%H%M%S")+"_"+secrets.token_hex(2)

def load_forcages():
    """Retourne le dict des livraisons sur jour non-livrable assumées par Bidou.
    Clé = date ISO 'YYYY-MM-DD'. Vide si fichier absent/illisible."""
    try:
        return load_json(FORCAGE_FILE) or {}
    except Exception as _e:
        _log_silent_err(exc=_e); return {}

def add_forcage(d,note="exception assum\u00e9e",vol=0):
    """Déclare une date comme livraison exceptionnelle ASSUMÉE.
    d : date | datetime | str ISO. Idempotent (réécrit la même clé)."""
    try:
        if hasattr(d,"isoformat"):
            key=(d if isinstance(d,date) and not isinstance(d,datetime) else d).strftime("%Y-%m-%d") if not isinstance(d,str) else d
        else:
            key=str(d)
        key=key[:10]
        fc=load_forcages()
        fc[key]={"ts":datetime.now().isoformat(),"note":note,"vol":int(vol or 0)}
        save_json(FORCAGE_FILE,fc)
        return True
    except Exception as _e:
        _log_silent_err(exc=_e); return False

def is_date_forcee(d):
    """True si la date (date|datetime|str ISO) a été déclarée exception assumée."""
    try:
        if hasattr(d,"strftime") and not isinstance(d,str):
            key=d.strftime("%Y-%m-%d")
        else:
            key=str(d)[:10]
        return key in load_forcages()
    except Exception as _e:
        _log_silent_err(exc=_e); return False

def load_commandes():
    """Dict des commandes carburant saisies. Clé = date cible livraison ISO."""
    try:
        return load_json(COMMANDE_FILE) or {}
    except Exception as _e:
        _log_silent_err(exc=_e); return {}

def add_commande(date_cible,sp=0,go=0,gnr=0,tour=1,premier_voyage=False):
    """Enregistre la commande passée chez Total pour la livraison de date_cible.
    date_cible : date|datetime|str ISO. Idempotent (réécrit la même clé)."""
    def _n(x):
        try: return int(round(float(str(x).replace(",",".").strip() or 0)))
        except Exception: return 0
    try:
        if hasattr(date_cible,"strftime") and not isinstance(date_cible,str):
            key=date_cible.strftime("%Y-%m-%d")
        else:
            key=str(date_cible)[:10]
        cmds=load_commandes()
        cmds[key]={"sp":_n(sp),"go":_n(go),"gnr":_n(gnr),
                   "tour":int(tour) if str(tour).isdigit() else 1,
                   "premier_voyage":bool(premier_voyage),
                   "ts":datetime.now().isoformat()}
        save_json(COMMANDE_FILE,cmds)
        return True
    except Exception as _e:
        _log_silent_err(exc=_e); return False

def get_commande(date_cible):
    """Retourne la commande saisie pour date_cible, ou None."""
    try:
        if hasattr(date_cible,"strftime") and not isinstance(date_cible,str):
            key=date_cible.strftime("%Y-%m-%d")
        else:
            key=str(date_cible)[:10]
        return load_commandes().get(key)
    except Exception as _e:
        _log_silent_err(exc=_e); return None

def delete_commande(date_cible):
    """Supprime la commande de date_cible. Idempotent.
    Trace l'opération dans le journal des événements pour audit."""
    try:
        if hasattr(date_cible,"strftime") and not isinstance(date_cible,str):
            key=date_cible.strftime("%Y-%m-%d")
        else:
            key=str(date_cible)[:10]
        cmds=load_commandes()
        if key in cmds:
            removed=cmds.pop(key)
            save_json(COMMANDE_FILE,cmds)
            # Trace dans le journal pour audit (suppression manuelle)
            try:
                add_evenement("commande",{
                    "jour":key,
                    "sp":int(removed.get("sp",0)),
                    "go":int(removed.get("go",0)),
                    "gnr":int(removed.get("gnr",0)),
                    "tour":int(removed.get("tour",1)),
                    "premier_voyage":bool(removed.get("premier_voyage",False)),
                    "statut":"supprimee",
                },commentaire="\u2716 Commande supprim\u00e9e manuellement")
            except Exception as _e: _log_silent_err(exc=_e)
            return True
        return False
    except Exception as _e:
        _log_silent_err(exc=_e); return False

def add_evenement(type_evt, data, commentaire=None, ts=None):
    """Ajoute un événement au journal permanent.
    Args:
        type_evt: 'pont' / 'anomalie' / 'rupture' / 'livraison' / 'ack'
        data: dict spécifique au type
        commentaire: chaîne optionnelle
        ts: datetime ISO (défaut = maintenant)
    Idempotence : si un événement avec la même fingerprint (type+data clés principales) existe
    déjà sur le même jour, on ne le re-crée pas pour éviter doublons sur refresh."""
    try:
        all_evt=load_json(EVENEMENTS_FILE) or {}
        events=all_evt.get("events",[])
        ts_iso=ts.isoformat() if ts else datetime.now().isoformat()
        # Fingerprint pour idempotence : type + clés stables de data
        fp=_evt_fingerprint(type_evt,data)
        today_key=ts_iso[:10]  # YYYY-MM-DD
        for evt in events:
            if evt.get("type")==type_evt and evt.get("ts","")[:10]==today_key:
                if _evt_fingerprint(evt.get("type"),evt.get("data",{}))==fp:
                    # Même événement déjà présent aujourd'hui → on ne re-crée pas
                    return evt.get("id")
        new_evt={
            "id":_new_event_id(),
            "ts":ts_iso,
            "type":type_evt,
            "data":data,
        }
        if commentaire:
            new_evt["commentaire"]=commentaire
        events.append(new_evt)
        all_evt["events"]=events
        save_json(EVENEMENTS_FILE,all_evt)
        return new_evt["id"]
    except Exception as e:
        print(f"[evenements] add error : {e}")
        return None

def masquer_evenement(evt_id):
    """Marque un événement comme masqué (champ masque=true). N'efface pas la donnée,
    juste un flag pour que le rendu du journal le saute par défaut. Réversible via
    demasquer_evenement(). Sujet validé Bidou 21/05/2026 : bouton manuel par carte +
    toggle 'Afficher masqués' pour réversibilité."""
    try:
        all_evt=load_json(EVENEMENTS_FILE) or {}
        events=all_evt.get("events",[])
        for evt in events:
            if evt.get("id")==evt_id:
                evt["masque"]=True
                evt["masque_ts"]=datetime.now().isoformat()
                all_evt["events"]=events
                save_json(EVENEMENTS_FILE,all_evt)
                return True
        return False
    except Exception as e:
        print(f"[evenements] masquer error : {e}")
        return False

def demasquer_evenement(evt_id):
    """Retire le flag masque d'un événement (le rend à nouveau visible)."""
    try:
        all_evt=load_json(EVENEMENTS_FILE) or {}
        events=all_evt.get("events",[])
        for evt in events:
            if evt.get("id")==evt_id:
                evt.pop("masque",None)
                evt.pop("masque_ts",None)
                all_evt["events"]=events
                save_json(EVENEMENTS_FILE,all_evt)
                return True
        return False
    except Exception as e:
        print(f"[evenements] demasquer error : {e}")
        return False

def delete_evenement(evt_id):
    """Supprime DÉFINITIVEMENT un événement de evenements.cfg. Action IRRÉVERSIBLE.
    Validée Bidou 25/05/2026 — différent de masquer (qui flag) ou marquer résolu (qui change statut).
    Effacement réel pour nettoyer la pollution du journal (résidus Pre_vision, fausses alertes
    héritées d'anciens bugs, événements qu'on ne veut PAS garder en trace même masquée).
    Retourne True si supprimé, False si pas trouvé ou erreur."""
    try:
        all_evt=load_json(EVENEMENTS_FILE) or {}
        events=all_evt.get("events",[])
        before=len(events)
        events=[e for e in events if e.get("id")!=evt_id]
        if len(events)==before: return False  # pas trouvé
        all_evt["events"]=events
        save_json(EVENEMENTS_FILE,all_evt)
        return True
    except Exception as e:
        print(f"[evenements] delete error : {e}")
        return False

def _evt_fingerprint(type_evt,data):
    """Calcule une fingerprint stable pour idempotence (évite doublons sur refresh).

    Important : le fingerprint doit être basé sur les CLÉS NATURELLES de l'alerte
    (type + carburant + jour ou pont concerné), PAS sur l'état courant (statut, snooze_until,
    commentaire, etc.). Sinon chaque changement d'état crée un nouvel événement → pollution
    du journal. Avec un fingerprint stable : 1 alerte unique = 1 entrée, statut évolue par
    update via EventActionDlg._update_evt.
    """
    if type_evt=="pont":
        return f"pont:{data.get('date_debut','')}-{data.get('date_fin','')}"
    if type_evt=="anomalie":
        return f"anom:{data.get('carburant','')}:{data.get('jour','')}"
    if type_evt=="rupture":
        return f"rupt:{data.get('carburant','')}:{data.get('jour','')}"
    if type_evt=="livraison":
        return f"livr:{data.get('jour','')}:{data.get('sp',0)}-{data.get('go',0)}-{data.get('gnr',0)}"
    if type_evt=="livraison_reporter":
        # 1 alerte par (carburant × date de livraison concernée) — ignore snooze_until/statut
        return f"livr_rep:{data.get('carburant','')}:{data.get('date','')}"
    if type_evt=="tendance":
        # 1 alerte par (carburant × pont concerné OU date), peu importe le snooze
        return f"tend:{data.get('carburant','')}:{data.get('pont_id','') or data.get('date','')}"
    if type_evt=="saisies_irregulieres":
        # 1 alerte par (carburant × date), peu importe les snoozes successifs
        return f"saisies:{data.get('carburant','')}:{data.get('date','')}"
    if type_evt=="ack":
        return f"ack:{data.get('popup','')}:{data.get('jour','')}"
    if type_evt=="commande":
        # 1 entrée par (date cible × volumes) — une commande corrigée (volumes
        # changés) crée une nouvelle ligne = traçabilité de l'ajustement.
        return f"cmd:{data.get('jour','')}:{data.get('sp',0)}-{data.get('go',0)}-{data.get('gnr',0)}"
    if type_evt=="marge_tendue":
        # 1 alerte par (carburant × date livraison). IGNORE marge_restante / stock_matin
        # qui varient légèrement à chaque refresh (ventes nuit), ce qui créait
        # des doublons : 3 cartes "Marge tendue SP samedi 23/05" avec marges
        # différentes (3 626 L à 06h58, 3 456 L à 07h42, etc.). Bug signalé
        # Bidou 20/05 20h05.
        return f"marge:{data.get('carburant','')}:{data.get('date','')}"
    if type_evt=="ferie_isole":
        # 1 alerte par date fériée concernée (peu importe les détails carburants).
        return f"ferie:{data.get('date','')}"
    if type_evt=="livraison_attendue":
        # 1 alerte par jour : camion attendu non confirmé arrivé au tour dépassé.
        return f"livratt:{data.get('date','')}"
    return f"{type_evt}:{json.dumps(data,sort_keys=True)}"

JOURNAL_RECONSTITUE_FILE=APP_DIR/"journal_reconstitue.cfg"


def detect_recent_events(hist_data,window_days=15):
    """Détection LIVE des ruptures et anomalies sur les derniers jours.

    Appelée à chaque refresh du hub (manuel ou auto 15 min) pour journaliser en TEMPS RÉEL
    les événements qui se produisent pendant que le hub tourne en continu, au lieu d'attendre
    une reconstitution complète au démarrage (qui ne se fait qu'une fois grâce au drapeau
    journal_reconstitue.cfg).

    Logique identique à reconstitute_journal_2026 mais sur fenêtre glissante des
    `window_days` derniers jours. Dédoublonnage natif via add_evenement (fingerprint
    par carburant+jour), donc rescanner les mêmes jours toutes les 15 min ne crée
    aucun doublon — on ne fait que rajouter si quelque chose de nouveau apparaît.

    Args:
        hist_data: liste de dicts d'historique LITRAGE (sortie de _read_hist).
        window_days: nombre de jours à scanner depuis aujourd'hui (défaut 15).
    """
    if not hist_data: return
    today=date.today()
    cutoff=today-timedelta(days=window_days)
    # Construire la fenêtre de détection (jours complets dans window_days)
    jours_window=[]
    for h in hist_data:
        d=h.get("date")
        if isinstance(d,str):
            try: d=datetime.strptime(d,"%Y-%m-%d").date()
            except Exception as _e: _log_silent_err(exc=_e); continue
        if not d or h.get("en_cours"): continue
        if cutoff<=d<=today:
            jours_window.append((d,h))
    jours_window.sort(key=lambda x:x[0])
    if len(jours_window)<3: return
    # === RUPTURES par carburant (séquences à 0 encadrées par jours >100 L) ===
    for carb in ("sp","go","gnr"):
        i=0;n=len(jours_window)
        while i<n:
            if sf(jours_window[i][1].get(carb,0))>0:
                i+=1;continue
            start=i
            while i<n and sf(jours_window[i][1].get(carb,0))<=0: i+=1
            end=i-1
            veille_vol=sf(jours_window[start-1][1].get(carb,0)) if start>0 else 0
            lendemain_vol=sf(jours_window[end+1][1].get(carb,0)) if end+1<n else 0
            if veille_vol>100 and lendemain_vol>100:
                d_start=jours_window[start][0];d_end=jours_window[end][0]
                duree=(d_end-d_start).days+1
                ts=datetime.combine(d_start,datetime.min.time()).replace(hour=12)
                if duree==1:
                    libelle=f"Rupture de {carb.upper()} le {d_start.strftime('%d/%m/%Y')}"
                    commentaire=f"D\u00e9tection auto : 0 L vendu sur {carb.upper()} (veille {int(veille_vol)} L, lendemain {int(lendemain_vol)} L)"
                else:
                    libelle=f"Rupture de {carb.upper()} du {d_start.strftime('%d/%m')} au {d_end.strftime('%d/%m/%Y')} ({duree} jours)"
                    commentaire=f"D\u00e9tection auto : {duree} jours cons\u00e9cutifs \u00e0 0 L sur {carb.upper()} (veille {int(veille_vol)} L, lendemain {int(lendemain_vol)} L)"
                add_evenement("rupture",{
                    "carburant":carb.upper(),"jour":d_start.strftime("%Y-%m-%d"),
                    "jour_fin":d_end.strftime("%Y-%m-%d"),"duree":duree,"libelle":libelle,
                    "vol_veille":int(veille_vol),"vol_lendemain":int(lendemain_vol),
                },commentaire=commentaire,ts=ts)
    # === ANOMALIES (variation litrage TOTAL > 30% vs moyenne 7j glissante) ===
    # On a besoin de tout l'historique 2026 pour calculer la moyenne 7j glissante,
    # même si on ne crée d'événement que pour les jours dans la fenêtre window_days.
    all_jours=[]
    for h in hist_data:
        d=h.get("date")
        if isinstance(d,str):
            try: d=datetime.strptime(d,"%Y-%m-%d").date()
            except Exception as _e: _log_silent_err(exc=_e); continue
        if not d or h.get("en_cours"): continue
        if d.year==2026:
            all_jours.append((d,h))
    all_jours.sort(key=lambda x:x[0])
    for i,(d,h) in enumerate(all_jours):
        if d<cutoff: continue  # Hors fenêtre de détection (déjà journalisé par reconstitute)
        if i<7: continue
        vol_jour=sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))
        if vol_jour<=0: continue  # Rupture déjà couverte ci-dessus
        volumes_prec=[]
        for j in range(max(0,i-7),i):
            v=sf(all_jours[j][1].get("sp",0))+sf(all_jours[j][1].get("go",0))+sf(all_jours[j][1].get("gnr",0))
            if v>0: volumes_prec.append(v)
        if len(volumes_prec)<5: continue
        moyenne=sum(volumes_prec)/len(volumes_prec)
        if moyenne<=0: continue
        ecart_pct=(vol_jour-moyenne)/moyenne*100
        if abs(ecart_pct)>30:
            ts=datetime.combine(d,datetime.min.time()).replace(hour=12)
            signe="+" if ecart_pct>0 else ""
            add_evenement("anomalie",{
                "carburant":"TOTAL","jour":d.strftime("%Y-%m-%d"),
                "litrage_jour":int(vol_jour),"moyenne_7j":int(moyenne),
                "ecart_pct":round(ecart_pct,1),
            },commentaire=f"D\u00e9tection auto : {signe}{ecart_pct:.1f}% vs moyenne 7j ({int(moyenne)} L)",ts=ts)

def reconstitute_journal_2026():
    """Reconstitution unique et silencieuse du journal des événements pour 2026.
    Scanne LITRAGE et Achat_carburant.xlsx pour ajouter rétroactivement :
      - Ruptures (jours où un carburant est resté à 0 L vendu)
      - Anomalies (variations litrage > 30% vs moyenne 7j glissante)
      - Livraisons enrichies depuis Achat_carburant.xlsx (montant €, BL, transporteur)

    Drapeau dans journal_reconstitue.cfg pour ne pas re-tourner à chaque démarrage.
    Dédoublonnage géré nativement par add_evenement (fingerprint).
    """
    flag=load_json(JOURNAL_RECONSTITUE_FILE) or {}
    if flag.get("done_2026"):
        return  # Déjà fait
    cfg=load_json(CONFIG_FILE) or {}
    nb_added=0
    # === 1. LIVRAISONS depuis Achat carburant.xlsx ===
    achat_path=cfg.get("achat_carburant","")
    # Si pas configuré, chercher dans le dossier OneDrive Fichiers Total à côté des autres.
    # Essaie plusieurs variantes du nom car le nom réel est "Achat carburant.xlsx" (espace,
    # minuscule à 'c'), mais on tolère quelques fautes de frappe potentielles.
    if not achat_path:
        pv_path=cfg.get("prevision","")
        if pv_path:
            try:
                parent=Path(pv_path).parent
                candidats=[
                    "Achat carburant.xlsx",   # Vrai nom Bidou
                    "achat carburant.xlsx",
                    "Achat_carburant.xlsx",
                    "Achats carburant.xlsx",
                    "Achats carburants.xlsx",
                ]
                for nom in candidats:
                    cand=parent/nom
                    if cand.exists():
                        achat_path=str(cand);break
            except Exception as _e: _log_silent_err(exc=_e)
    if achat_path and Path(achat_path).exists():
        try:
            import openpyxl
            tmp=copy_to_temp(achat_path)
            wb=openpyxl.load_workbook(tmp,data_only=True)
            ws=wb["Feuil1"] if "Feuil1" in wb.sheetnames else wb[wb.sheetnames[0]]
            # Le fichier comporte DEUX colonnes parallèles avec le même pattern :
            #   - Bloc gauche : col A=date, col B='SP', col C=volume_SP, col D=PA, col E=montant
            #                   col A+1=BL, ..., col F (ligne+3)=total, col A+3=transporteur
            #   - Bloc droit  : col H=date, col I='SP', col J=volume_SP, col K=PA, col L=montant
            #                   col H+1=BL, ..., col M (ligne+3)=total, col H+3=transporteur
            # On scanne ligne par ligne et on traite chaque bloc indépendamment.
            for r in range(1,ws.max_row+1):
                # Pour chaque "départ" potentiel (col A=1 ou col H=8)
                for col_date in (1,8):
                    val=ws.cell(row=r,column=col_date).value
                    if isinstance(val,datetime) and val.year==2026:
                        col_total=col_date+5  # F si A=1, M si H=8
                        transporteur=ws.cell(row=r+3,column=col_date).value or ""
                        sp_l=sf(ws.cell(row=r,column=col_date+2).value or 0)
                        go_l=sf(ws.cell(row=r+1,column=col_date+2).value or 0)
                        gnr_l=sf(ws.cell(row=r+2,column=col_date+2).value or 0)
                        total_eur=sf(ws.cell(row=r+3,column=col_total).value or 0)
                        if (sp_l or go_l or gnr_l) and total_eur>0:
                            d=val.date()
                            ts=datetime.combine(d,datetime.min.time()).replace(hour=12)
                            data={
                                "jour":d.strftime("%Y-%m-%d"),
                                "sp":int(sp_l),"go":int(go_l),"gnr":int(gnr_l),
                                "total_eur":round(total_eur,2),
                                "transporteur":str(transporteur).strip(),
                            }
                            new_id=add_evenement("livraison",data,ts=ts)
                            if new_id: nb_added+=1
        except Exception as e: print(f"[reconstitute livraisons xlsx] {e}")
    # === 1bis. LIVRAISONS depuis livraisons.cfg (saisies en temps réel via le hub) ===
    # Permet de récupérer les livraisons saisies après la dernière mise à jour d'Achat carburant.xlsx
    # (en pratique, Véronique met à jour le fichier avec un délai). Dédoublonnage par fingerprint.
    try:
        livrs=load_json(LIVRAISON_FILE) or {}
        for date_key,info in livrs.items():
            if not isinstance(info,dict): continue
            if info.get("none"): continue  # marqueur "pas de livraison ce jour"
            try:
                d=datetime.strptime(date_key,"%d/%m/%y").date()
            except Exception as _e: _log_silent_err(exc=_e); continue
            if d.year!=2026: continue
            sp_l=sf(info.get("sp",0))
            go_l=sf(info.get("go",0))
            gnr_l=sf(info.get("gnr",0))
            if not (sp_l or go_l or gnr_l): continue
            ts=datetime.combine(d,datetime.min.time()).replace(hour=12)
            data={
                "jour":d.strftime("%Y-%m-%d"),
                "sp":int(sp_l),"go":int(go_l),"gnr":int(gnr_l),
            }
            new_id=add_evenement("livraison",data,ts=ts)
            if new_id: nb_added+=1
    except Exception as e: print(f"[reconstitute livraisons.cfg] {e}")
    # === 2. RUPTURES depuis LITRAGE (jours où un carburant a un litrage = 0 alors que normalement non) ===
    # On utilise hist_data via le DataReader qui a déjà toute la logique de lecture LITRAGE.
    # On NE relit pas LITRAGE manuellement ici pour éviter de réinventer la roue.
    try:
        reader=DataReader(cfg)
        hist=reader._read_hist(full=True) or []
        # Filtrer 2026 et complets
        jours_2026=[]
        for h in hist:
            d=parse_label_date(h.get("label",""))
            if d and d.year==2026 and not h.get("en_cours"):
                jours_2026.append((d,h))
        # Pour chaque carburant, repérer les SÉQUENCES consécutives à 0 L.
        # Logique : on parcourt jours_2026, on identifie les segments où vol=0, puis pour chaque
        # segment on vérifie qu'il est encadré par des jours normaux (>100 L) avant et après
        # (pour distinguer les vraies ruptures des jours de fermeture station ou bornes).
        # Un seul événement créé par séquence, daté du premier jour de la séquence.
        for carb in ("sp","go","gnr"):
            i=0
            n=len(jours_2026)
            while i<n:
                d_i,h_i=jours_2026[i]
                vol_i=sf(h_i.get(carb,0))
                if vol_i>0:
                    i+=1;continue
                # Début d'une séquence à 0. Trouver la fin.
                start=i
                while i<n and sf(jours_2026[i][1].get(carb,0))<=0:
                    i+=1
                end=i-1  # dernier jour à 0 inclus
                # Vérifier les jours encadrants
                veille_vol=sf(jours_2026[start-1][1].get(carb,0)) if start>0 else 0
                lendemain_vol=sf(jours_2026[end+1][1].get(carb,0)) if end+1<n else 0
                if veille_vol>100 and lendemain_vol>100:
                    d_start=jours_2026[start][0]
                    d_end=jours_2026[end][0]
                    duree=(d_end-d_start).days+1
                    ts=datetime.combine(d_start,datetime.min.time()).replace(hour=12)
                    if duree==1:
                        libelle=f"Rupture de {carb.upper()} le {d_start.strftime('%d/%m/%Y')}"
                        commentaire=f"Détection auto : 0 L vendu sur {carb.upper()} (veille {int(veille_vol)} L, lendemain {int(lendemain_vol)} L)"
                    else:
                        libelle=f"Rupture de {carb.upper()} du {d_start.strftime('%d/%m')} au {d_end.strftime('%d/%m/%Y')} ({duree} jours)"
                        commentaire=f"Détection auto : {duree} jours consécutifs à 0 L sur {carb.upper()} (veille {int(veille_vol)} L, lendemain {int(lendemain_vol)} L)"
                    data={
                        "carburant":carb.upper(),
                        "jour":d_start.strftime("%Y-%m-%d"),
                        "jour_fin":d_end.strftime("%Y-%m-%d"),
                        "duree":duree,
                        "libelle":libelle,
                        "vol_veille":int(veille_vol),
                        "vol_lendemain":int(lendemain_vol),
                    }
                    new_id=add_evenement("rupture",data,commentaire=commentaire,ts=ts)
                    if new_id: nb_added+=1
        # === 3. ANOMALIES (variation litrage TOTAL > 30% vs moyenne 7j glissante) ===
        # Note : on ne fait que sur le litrage total (pas par carburant) pour limiter le bruit.
        for i,(d,h) in enumerate(jours_2026):
            if i<7: continue  # besoin de 7 jours d'historique avant
            vol_jour=sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))
            if vol_jour<=0: continue  # rupture déjà couverte
            # Moyenne 7 jours glissants précédents (en excluant les ruptures à 0)
            volumes_prec=[]
            for j in range(max(0,i-7),i):
                v=sf(jours_2026[j][1].get("sp",0))+sf(jours_2026[j][1].get("go",0))+sf(jours_2026[j][1].get("gnr",0))
                if v>0: volumes_prec.append(v)
            if len(volumes_prec)<5: continue  # pas assez de données fiables
            moyenne=sum(volumes_prec)/len(volumes_prec)
            if moyenne<=0: continue
            ecart_pct=(vol_jour-moyenne)/moyenne*100
            if abs(ecart_pct)>30:
                ts=datetime.combine(d,datetime.min.time()).replace(hour=12)
                signe="+" if ecart_pct>0 else ""
                data={
                    "carburant":"TOTAL",
                    "jour":d.strftime("%Y-%m-%d"),
                    "litrage_jour":int(vol_jour),
                    "moyenne_7j":int(moyenne),
                    "ecart_pct":round(ecart_pct,1),
                }
                new_id=add_evenement("anomalie",data,
                                      commentaire=f"Détection auto : {signe}{ecart_pct:.1f}% vs moyenne 7j ({int(moyenne)} L)",
                                      ts=ts)
                if new_id: nb_added+=1
    except Exception as e: print(f"[reconstitute ruptures/anomalies] {e}")
    # === 4. PASSAGES DE MOIS depuis prix_historique.cfg ===
    # Pour chaque mois 2026 qui contient un stock_pivot et des ventes_avant_6h saisis,
    # on reconstitue l'événement journal "passage de mois" avec le calcul de l'effet.
    try:
        histo=_load_prix_histo()
        prix=histo.get("prix",{}) or {}
        mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin",
                   "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        for key in sorted(prix.keys()):
            if not key.startswith("2026-"): continue
            data_mois=prix[key]
            if not isinstance(data_mois,dict): continue
            # Vérifier qu'un passage de mois a bien été saisi (champ phare = stock_pivot_sp)
            if "stock_pivot_sp" not in data_mois: continue
            saisi_at=data_mois.get("passage_mois_saisi_at")
            if not saisi_at: continue
            try:
                year,month=int(key[:4]),int(key[5:7])
                stock_pivot={
                    "sp":sf(data_mois.get("stock_pivot_sp",0)),
                    "go":sf(data_mois.get("stock_pivot_go",0)),
                    "gnr":sf(data_mois.get("stock_pivot_gnr",0)),
                }
                ventes_avant_6h={
                    "sp":sf(data_mois.get("ventes_avant_6h_sp",0)),
                    "go":sf(data_mois.get("ventes_avant_6h_go",0)),
                    "gnr":sf(data_mois.get("ventes_avant_6h_gnr",0)),
                }
                eff=calc_effet_speculation(year,month,
                    stock_pivot["sp"],stock_pivot["go"],stock_pivot["gnr"],
                    ventes_avant_6h["sp"],ventes_avant_6h["go"],ventes_avant_6h["gnr"])
                # calc_effet_speculation retourne {"total": ...}, pas "effet_total"
                effet_total=sf(eff.get("total",0)) if isinstance(eff,dict) else 0
                # Mois précédent pour le label
                prev_month=12 if month==1 else month-1
                prev_year=year-1 if month==1 else year
                label=f"Passage {mois_noms[prev_month-1]} {prev_year} \u2192 {mois_noms[month-1]} {year}"
                # Date de l'événement = 1er du mois à 00h05 (juste après le basculement de prix)
                ts=datetime(year,month,1,0,5)
                evt_data={
                    "year":year,"month":month,
                    "stock_pivot":stock_pivot,
                    "ventes_avant_6h":ventes_avant_6h,
                    "effet_total":round(effet_total,2),
                    "label":label,
                }
                # Force update : supprimer les anciens passage_mois pour ce (year,month)
                # avant d'en créer un nouveau. Évite les soucis de fingerprint qui
                # empêcheraient la mise à jour si un événement avec un mauvais effet_total
                # existait déjà (cas du bug "effet_total" vs "total").
                try:
                    all_evt=load_json(EVENEMENTS_FILE) or {}
                    events=all_evt.get("events",[])
                    cleaned=[e for e in events
                             if not (e.get("type")=="passage_mois"
                                     and (e.get("data",{}) or {}).get("year")==year
                                     and (e.get("data",{}) or {}).get("month")==month)]
                    if len(cleaned)!=len(events):
                        all_evt["events"]=cleaned
                        save_json(EVENEMENTS_FILE,all_evt)
                except Exception as e: print(f"[reconstitute passage cleanup] {e}")
                new_id=add_evenement("passage_mois",evt_data,ts=ts)
                if new_id: nb_added+=1
            except Exception as e: print(f"[reconstitute passage {key}] {e}")
    except Exception as e: print(f"[reconstitute passages] {e}")
    # Marquer comme fait
    flag["done_2026"]=True
    flag["done_at"]=datetime.now().isoformat()
    flag["nb_events_added"]=nb_added
    save_json(JOURNAL_RECONSTITUE_FILE,flag)
    print(f"[journal] Reconstitution 2026 terminée : {nb_added} événements ajoutés")


def update_evenement_commentaire(evt_id,commentaire):
    """Met à jour le commentaire d'un événement existant."""
    try:
        all_evt=load_json(EVENEMENTS_FILE) or {}
        events=all_evt.get("events",[])
        for evt in events:
            if evt.get("id")==evt_id:
                if commentaire:
                    evt["commentaire"]=commentaire
                else:
                    evt.pop("commentaire",None)
                all_evt["events"]=events
                save_json(EVENEMENTS_FILE,all_evt)
                return True
        return False
    except Exception as e:
        print(f"[evenements] update error : {e}")
        return False

def load_evenements_period(start_date,end_date):
    """Retourne la liste des événements dont le timestamp est dans [start_date, end_date] inclus.
    Args :
        start_date, end_date : objets date (pas datetime).
    Trié par ordre chronologique."""
    try:
        all_evt=load_json(EVENEMENTS_FILE) or {}
        events=all_evt.get("events",[])
        result=[]
        for evt in events:
            ts=evt.get("ts","")
            if not ts: continue
            try:
                evt_date=datetime.fromisoformat(ts).date()
            except Exception as _e: _log_silent_err(exc=_e); continue
            if start_date<=evt_date<=end_date:
                result.append(evt)
        result.sort(key=lambda e:e.get("ts",""))
        return result
    except Exception as e:
        print(f"[evenements] load_period error : {e}")
        return []

def load_evenements_month(year,month):
    """Helper : retourne les événements d'un mois donné."""
    import calendar
    last_day=calendar.monthrange(year,month)[1]
    return load_evenements_period(date(year,month,1),date(year,month,last_day))


# =============================================================================
# HISTORIQUE DES PRIX MENSUELS (pour calcul effet spéculation)
# À chaque refresh, on enregistre les prix lus du mois en cours dans le fichier
# `prix_historique.cfg`. Si le mois change (ex: passage avril→mai au 1er du mois)
# on gèle les prix du mois précédent et on capture les nouveaux. Cela permet
# ensuite de calculer la "spéculation" : au passage 30/04→01/05, le stock cuve
# du matin du 1er a basculé du prix d'avril au prix de mai → gain ou perte.
# =============================================================================

def _load_prix_histo():
    """Charge le fichier prix_historique.cfg en gérant 2 formats compatibles :
    - Nouveau (depuis import) : {"prix": {"YYYY-MM": {...}}}
    - Ancien (créé par update_prix_courant) : {"YYYY-MM": {...}}
    Retourne toujours la structure {"prix": {...}}."""
    raw=load_json(PRIX_HISTO_FILE) or {}
    if "prix" in raw and isinstance(raw["prix"],dict):
        return raw
    # Convertir ancien format → nouveau
    new_prix={}
    for k,v in raw.items():
        if isinstance(v,dict) and ("pv_sp" in v or "pa_sp" in v or "marge_boutique_taux" in v):
            new_prix[k]=v
    return {"prix":new_prix}

def _save_prix_histo(data):
    """Sauvegarde au format {"prix": {...}}."""
    if "prix" not in data: data={"prix":data}
    save_json(PRIX_HISTO_FILE,data)


def update_prix_courant(pa_sp,pa_go,pa_gnr,pv_sp,pv_go,pv_gnr,marge_unit):
    """Met à jour la valeur courante des prix pour le mois en cours.
    Idempotent : écrase juste les valeurs du mois courant. Ne touche pas aux mois antérieurs.
    Préserve TOUS les champs déjà présents (marge_boutique_taux, mais aussi
    stock_pivot_sp/go/gnr et ventes_avant_6h_sp/go/gnr / passage_mois_saisi_at
    écrits par save_passage_mois) en mutant le dict au lieu de le remplacer."""
    try:
        today=date.today()
        key=f"{today.year:04d}-{today.month:02d}"
        histo=_load_prix_histo()
        prix=histo.setdefault("prix",{})
        existing=prix.get(key,{})
        # Calcul des marges détaillées (palier réglementaire après sept 2024 = uniforme)
        m_sp=marge_unit if marge_unit else (pv_sp-pa_sp if pv_sp and pa_sp else 0)
        m_go=marge_unit if marge_unit else (pv_go-pa_go if pv_go and pa_go else 0)
        m_gnr=marge_unit if marge_unit else (pv_gnr-pa_gnr if pv_gnr and pa_gnr else 0)
        # MUTATION (pas remplacement) pour préserver les champs hors prix
        existing["pa_sp"]=round(sf(pa_sp),5)
        existing["pa_go"]=round(sf(pa_go),5)
        existing["pa_gnr"]=round(sf(pa_gnr),5)
        existing["pv_sp"]=round(sf(pv_sp),4)
        existing["pv_go"]=round(sf(pv_go),4)
        existing["pv_gnr"]=round(sf(pv_gnr),4)
        existing["marge_unit"]=round(sf(marge_unit),5)
        existing["marge_sp"]=round(sf(m_sp),5)
        existing["marge_go"]=round(sf(m_go),5)
        existing["marge_gnr"]=round(sf(m_gnr),5)
        # marge_boutique_taux : conservé si déjà défini, sinon 30% par défaut
        if "marge_boutique_taux" not in existing:
            existing["marge_boutique_taux"]=0.30
        existing["last_seen"]=datetime.now().isoformat()
        # source : conservée si déjà définie
        if "source" not in existing:
            existing["source"]="auto_pre_vision"
        prix[key]=existing
        _save_prix_histo(histo)
    except Exception as e: print(f"[prix histo] update error : {e}")


def get_prix_for_month(year,month):
    """Retourne les prix mémorisés pour un mois donné, ou None."""
    histo=_load_prix_histo()
    return histo.get("prix",{}).get(f"{year:04d}-{month:02d}")


def get_prix_previous_month(year,month):
    """Retourne les prix du mois précédent, ou None si pas mémorisés."""
    if month==1: pyr,pmo=year-1,12
    else: pyr,pmo=year,month-1
    return get_prix_for_month(pyr,pmo)


def save_passage_mois(year,month,stock_pivot,ventes_avant_6h):
    """Enregistre les données du passage de mois dans prix_historique.cfg.

    Args:
        year, month: mois CIBLE (ex: 2026, 6 = passage de mai → juin saisi le 1er juin)
        stock_pivot: dict {sp, go, gnr} = stock à 6h le 1er du mois (= matin du 1er)
        ventes_avant_6h: dict {sp, go, gnr} = ventes 0h-6h du 1er, à déduire
    """
    try:
        histo=_load_prix_histo()
        prix=histo.setdefault("prix",{})
        key=f"{year:04d}-{month:02d}"
        existing=prix.get(key,{})
        # On ne touche QUE aux champs passage de mois, le reste est préservé
        existing["stock_pivot_sp"]=round(sf(stock_pivot.get("sp",0)),0)
        existing["stock_pivot_go"]=round(sf(stock_pivot.get("go",0)),0)
        existing["stock_pivot_gnr"]=round(sf(stock_pivot.get("gnr",0)),0)
        existing["ventes_avant_6h_sp"]=round(sf(ventes_avant_6h.get("sp",0)),0)
        existing["ventes_avant_6h_go"]=round(sf(ventes_avant_6h.get("go",0)),0)
        existing["ventes_avant_6h_gnr"]=round(sf(ventes_avant_6h.get("gnr",0)),0)
        existing["passage_mois_saisi_at"]=datetime.now().isoformat()
        prix[key]=existing
        _save_prix_histo(histo)
        return True
    except Exception as e:
        print(f"[passage mois] save error : {e}")
        return False


def get_passage_mois(year,month):
    """Retourne les données du passage de mois si saisies, sinon None."""
    p=get_prix_for_month(year,month) or {}
    if not p.get("passage_mois_saisi_at"):
        return None
    return {
        "stock_pivot":{"sp":sf(p.get("stock_pivot_sp",0)),
                       "go":sf(p.get("stock_pivot_go",0)),
                       "gnr":sf(p.get("stock_pivot_gnr",0))},
        "ventes_avant_6h":{"sp":sf(p.get("ventes_avant_6h_sp",0)),
                           "go":sf(p.get("ventes_avant_6h_go",0)),
                           "gnr":sf(p.get("ventes_avant_6h_gnr",0))},
        "saisi_at":p.get("passage_mois_saisi_at"),
    }


def calc_effet_speculation(year,month,stock_pivot_sp,stock_pivot_go,stock_pivot_gnr,
                            ventes_avant_6h_sp=0,ventes_avant_6h_go=0,ventes_avant_6h_gnr=0):
    """Calcule l'effet spéculation pour un mois donné.

    Principe : le prix de vente change à MINUIT le 1er du mois (arrêté préfectoral).
    Tout ce qui est en cuve à minuit a été acheté au prix du mois précédent (PA),
    mais sera vendu au nouveau prix (PV mois en cours). Différence = gain (si prix
    monté) ou perte (si prix baissé).

    Le stock à minuit n'est pas mesurable directement (la jauge cuve se lit à 6h
    après la C3). On le RECONSTITUE :
        stock_minuit = stock_6h + ventes_0h_6h
    Les ventes 0h-6h ont déjà été vendues au nouveau prix (le prix a basculé à 0h),
    donc elles ont AUSSI bénéficié du changement → on les RÉINTÈGRE au stock pivot.

    Args:
        year, month: mois pour lequel on calcule (ex: 2026, 5 → mai 2026)
        stock_pivot_*: stocks en litres lus en cuve à 6h le 1er du mois
        ventes_avant_6h_*: litres écoulés entre 0h et 6h du 1er (déjà au nouveau prix)

    Returns:
        dict {sp, go, gnr, total, prix_avant, prix_apres, hausse_pct} ou None si pas de données.
    """
    prix_curr=get_prix_for_month(year,month)
    prix_prev=get_prix_previous_month(year,month)
    if not prix_curr or not prix_prev: return None
    # Stock pivot effectif = stock 6h + ventes 0h-6h (= stock à minuit reconstitué)
    pivot_sp=sf(stock_pivot_sp)+sf(ventes_avant_6h_sp)
    pivot_go=sf(stock_pivot_go)+sf(ventes_avant_6h_go)
    pivot_gnr=sf(stock_pivot_gnr)+sf(ventes_avant_6h_gnr)
    # Effet pour chaque carburant = pivot × (PV_nouveau - PV_ancien)
    eff_sp=pivot_sp*(sf(prix_curr.get("pv_sp",0))-sf(prix_prev.get("pv_sp",0)))
    eff_go=pivot_go*(sf(prix_curr.get("pv_go",0))-sf(prix_prev.get("pv_go",0)))
    eff_gnr=pivot_gnr*(sf(prix_curr.get("pv_gnr",0))-sf(prix_prev.get("pv_gnr",0)))
    total=eff_sp+eff_go+eff_gnr
    return {
        "sp":round(eff_sp,2),"go":round(eff_go,2),"gnr":round(eff_gnr,2),
        "total":round(total,2),
        "prix_avant":{"sp":prix_prev.get("pv_sp",0),"go":prix_prev.get("pv_go",0),"gnr":prix_prev.get("pv_gnr",0)},
        "prix_apres":{"sp":prix_curr.get("pv_sp",0),"go":prix_curr.get("pv_go",0),"gnr":prix_curr.get("pv_gnr",0)},
        "stocks_pivot":{"sp":pivot_sp,"go":pivot_go,"gnr":pivot_gnr},
        "ventes_avant_6h":{"sp":sf(ventes_avant_6h_sp),"go":sf(ventes_avant_6h_go),"gnr":sf(ventes_avant_6h_gnr)},
    }


# =============================================================================
# SNAPSHOT MENSUEL
# Capture les chiffres clés d'un mois donné dans un JSON dans ~/.districarb_hub/snapshots/
# Permanent et empilable : peut être généré à n'importe quel moment, écrase la version précédente.
# Sert de source pour la génération de rapports mensuels.
# =============================================================================
def _format_anomalies_from_journal(year,month):
    """Lit le journal d'événements et formate les anomalies du mois pour affichage rapport."""
    try:
        events=load_evenements_month(year,month)
        result=[]
        for evt in events:
            if evt.get("type")!="anomalie": continue
            d=evt.get("data",{}) or {}
            jour=d.get("jour","");carb=d.get("carburant","")
            ecart=d.get("ecart_pct",0)
            sens="hausse" if ecart>0 else "baisse"
            txt=f"{jour} \u2014 {carb} : {sens} de {abs(ecart):.0f}%"
            comm=evt.get("commentaire")
            if comm: txt+=f" ({comm})"
            result.append(txt)
        return result
    except Exception as _e: _log_silent_err(exc=_e); return []

def _format_ponts_from_journal(year,month):
    """Lit le journal d'événements et formate les ponts traversés du mois pour rapport."""
    try:
        events=load_evenements_month(year,month)
        result=[]
        for evt in events:
            if evt.get("type")!="pont": continue
            d=evt.get("data",{}) or {}
            d_deb=d.get("date_debut","");d_fin=d.get("date_fin","")
            duree=d.get("duree","?")
            txt=f"{d_deb} \u2192 {d_fin} ({duree} jour{'s' if duree!=1 else ''}) \u2014 sous contr\u00f4le"
            comm=evt.get("commentaire")
            if comm: txt+=f" \u2014 {comm}"
            result.append(txt)
        return result
    except Exception as _e: _log_silent_err(exc=_e); return []


def _periode_label(start_date, end_date):
    """Génère un libellé humain pour une plage : 'Avril 2026', 'Q2 2025', 'Année 2024',
    'Mai 2021 → Décembre 2025', etc."""
    mois_noms=["janvier","février","mars","avril","mai","juin",
               "juillet","août","septembre","octobre","novembre","décembre"]
    s,e=start_date,end_date
    # Mois unique
    if s.year==e.year and s.month==e.month:
        # Si plage couvre tout le mois → libellé mois
        last_day=(date(s.year,s.month%12+1,1)-timedelta(days=1)) if s.month<12 else date(s.year,12,31)
        if s.day==1 and e==last_day:
            return f"{mois_noms[s.month-1].capitalize()} {s.year}"
    # Année complète
    if s.month==1 and s.day==1 and e.month==12 and e.day==31 and s.year==e.year:
        return f"Année {s.year}"
    # Trimestre
    if s.day==1 and s.year==e.year:
        for q,(m1,m2) in enumerate([(1,3),(4,6),(7,9),(10,12)],start=1):
            last_q=date(s.year,m2%12+1,1)-timedelta(days=1) if m2<12 else date(s.year,12,31)
            if s.month==m1 and e.month==m2 and e==last_q:
                return f"T{q} {s.year}"
    # Semestre
    if s.day==1 and s.year==e.year:
        if s.month==1 and e.month==6 and e==date(s.year,6,30):
            return f"S1 {s.year}"
        if s.month==7 and e.month==12 and e==date(s.year,12,31):
            return f"S2 {s.year}"
    # Plage générique
    return f"{mois_noms[s.month-1].capitalize()} {s.year} → {mois_noms[e.month-1].capitalize()} {e.year}"


def _aggregate_par_mois(jours_periode):
    """Agrège des jours en (year, month) → dict des totaux pour ce mois.
    Sert pour les bilans multi-mois : détail mensuel + top mois.
    Retourne une liste ordonnée chronologiquement."""
    par_mois={}
    for d,h in jours_periode:
        key=(d.year,d.month)
        if key not in par_mois:
            par_mois[key]={"year":d.year,"month":d.month,"nb_jours":0,
                           "sp":0,"go":0,"gnr":0,"piste":0,"bout":0,"cb":0,"cp":0,"esp":0}
        m=par_mois[key]
        m["nb_jours"]+=1
        m["sp"]+=sf(h.get("sp",0));m["go"]+=sf(h.get("go",0));m["gnr"]+=sf(h.get("gnr",0))
        m["piste"]+=sf(h.get("piste",0));m["bout"]+=sf(h.get("bout",0))
        m["cb"]+=sf(h.get("cb",0));m["cp"]+=sf(h.get("cp",0));m["esp"]+=sf(h.get("esp",0))
    mois_noms=["janvier","février","mars","avril","mai","juin",
               "juillet","août","septembre","octobre","novembre","décembre"]
    result=[]
    for key in sorted(par_mois.keys()):
        m=par_mois[key]
        m["litrage_total"]=int(m["sp"]+m["go"]+m["gnr"])
        m["ca_total"]=round(m["piste"]+m["bout"],2)
        m["label"]=f"{mois_noms[m['month']-1].capitalize()} {m['year']}"
        # Marge mensuelle (si prix dispo dans prix_historique)
        prix=get_prix_for_month(m["year"],m["month"]) or {}
        pv_sp=sf(prix.get("pv_sp",0));pa_sp=sf(prix.get("pa_sp",0))
        pv_go=sf(prix.get("pv_go",0));pa_go=sf(prix.get("pa_go",0))
        pv_gnr=sf(prix.get("pv_gnr",0));pa_gnr=sf(prix.get("pa_gnr",0))
        msp=(pv_sp-pa_sp) if (pv_sp and pa_sp) else 0
        mgo=(pv_go-pa_go) if (pv_go and pa_go) else 0
        mgnr=(pv_gnr-pa_gnr) if (pv_gnr and pa_gnr) else 0
        m["marge_carb"]=round(m["sp"]*msp+m["go"]*mgo+m["gnr"]*mgnr,2)
        # Marge boutique avec taux du mois (sinon 30%)
        taux=sf(prix.get("marge_boutique_taux",0.30)) if prix else 0.30
        if taux<=0 or taux>1: taux=0.30
        m["marge_bout"]=round(m["bout"]*taux,2)
        m["marge_total"]=round(m["marge_carb"]+m["marge_bout"],2)
        # Arrondis pour sortie
        for k in ("sp","go","gnr"): m[k]=int(m[k])
        for k in ("piste","bout","cb","cp","esp"): m[k]=round(m[k],2)
        result.append(m)
    return result


def build_period_snapshot(hist_data, start_date, end_date, anomalies=None, ponts=None,
                          alerts=None, objectif=None):
    """Construit le snapshot d'une plage de dates [start_date, end_date] (inclusif).

    Mode automatique selon les dates :
      - LITE : si start_date < 2025-01-01 → KPI de base seulement (CA piste, boutique, marge,
        litrage), détail mensuel agrégé, top mois, comparaison période précédente même durée.
      - COMPLET : si start_date >= 2025-01-01 → tout le LITE + marges détaillées par mois +
        top jours (mono-mois) ou top mois (multi-mois) + événements remarquables.

    Args:
        hist_data: liste de dicts (sortie de _read_hist).
        start_date, end_date: date objects inclusifs.
        anomalies, ponts, alerts, objectif: comme build_month_snapshot.
    Retourne: dict snapshot.
    """
    if start_date>end_date: start_date,end_date=end_date,start_date
    # Filtrer les jours dans la plage
    jours_periode=[]
    for h in hist_data:
        if h.get("en_cours"): continue
        d=parse_label_date(h.get("label",""))
        if d and start_date<=d<=end_date:
            jours_periode.append((d,h))
    jours_periode.sort(key=lambda x:x[0])
    # Détermination du niveau (Lite/Complet) selon dates
    niveau="complet" if start_date>=date(2025,1,1) else "lite"
    # Détermination mono-mois vs multi-mois (impacte le top : jours vs mois)
    mono_mois=(start_date.year==end_date.year and start_date.month==end_date.month)
    nb_mois_couverts=len({(d.year,d.month) for d,_ in jours_periode})
    # Agrégats globaux
    total_sp=sum(sf(h.get("sp",0)) for d,h in jours_periode)
    total_go=sum(sf(h.get("go",0)) for d,h in jours_periode)
    total_gnr=sum(sf(h.get("gnr",0)) for d,h in jours_periode)
    total_l=total_sp+total_go+total_gnr
    ca_piste=sum(sf(h.get("piste",0)) for d,h in jours_periode)
    ca_boutique=sum(sf(h.get("bout",0)) for d,h in jours_periode)
    cb=sum(sf(h.get("cb",0)) for d,h in jours_periode)
    cp=sum(sf(h.get("cp",0)) for d,h in jours_periode)
    esp=sum(sf(h.get("esp",0)) for d,h in jours_periode)
    ca_total=ca_piste+ca_boutique
    # MARGE : agrégation mois par mois (palier réglementaire change selon la date)
    par_mois=_aggregate_par_mois(jours_periode)
    marge_carb_total=round(sum(m["marge_carb"] for m in par_mois),2)
    marge_bout_total=round(sum(m["marge_bout"] for m in par_mois),2)
    marge_grand_total=round(marge_carb_total+marge_bout_total,2)
    # Top journées (mono-mois) ou top mois (multi-mois)
    top_meilleures=[];top_pires=[]
    top_bout_meilleures=[];top_bout_pires=[]
    if mono_mois:
        triees=sorted(jours_periode,key=lambda x:sf(x[1].get("piste",0)),reverse=True)
        for d,h in triees[:3]:
            top_meilleures.append({
                "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
                "litrage":int(sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))),
                "ca_piste":round(sf(h.get("piste",0)),2),
                "ca_boutique":round(sf(h.get("bout",0)),2),
            })
        for d,h in triees[-3:][::-1]:
            top_pires.append({
                "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
                "litrage":int(sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))),
                "ca_piste":round(sf(h.get("piste",0)),2),
                "ca_boutique":round(sf(h.get("bout",0)),2),
            })
        triees_b=sorted(jours_periode,key=lambda x:sf(x[1].get("bout",0)),reverse=True)
        for d,h in triees_b[:3]:
            top_bout_meilleures.append({
                "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
                "ca_boutique":round(sf(h.get("bout",0)),2),
                "ca_piste":round(sf(h.get("piste",0)),2),
            })
        for d,h in triees_b[-3:][::-1]:
            top_bout_pires.append({
                "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
                "ca_boutique":round(sf(h.get("bout",0)),2),
                "ca_piste":round(sf(h.get("piste",0)),2),
            })
    else:
        # Multi-mois : top par mois (sur CA piste)
        tri_m=sorted(par_mois,key=lambda m:m["piste"],reverse=True)
        for m in tri_m[:3]:
            top_meilleures.append({
                "date":f"{m['year']:04d}-{m['month']:02d}-01","label":m["label"],
                "litrage":m["litrage_total"],
                "ca_piste":m["piste"],"ca_boutique":m["bout"],
            })
        for m in tri_m[-3:][::-1]:
            top_pires.append({
                "date":f"{m['year']:04d}-{m['month']:02d}-01","label":m["label"],
                "litrage":m["litrage_total"],
                "ca_piste":m["piste"],"ca_boutique":m["bout"],
            })
        tri_mb=sorted(par_mois,key=lambda m:m["bout"],reverse=True)
        for m in tri_mb[:3]:
            top_bout_meilleures.append({
                "date":f"{m['year']:04d}-{m['month']:02d}-01","label":m["label"],
                "ca_boutique":m["bout"],"ca_piste":m["piste"],
            })
        for m in tri_mb[-3:][::-1]:
            top_bout_pires.append({
                "date":f"{m['year']:04d}-{m['month']:02d}-01","label":m["label"],
                "ca_boutique":m["bout"],"ca_piste":m["piste"],
            })
    # Comparaison période précédente même durée
    nb_jours_periode=(end_date-start_date).days+1
    prev_end=start_date-timedelta(days=1)
    prev_start=prev_end-timedelta(days=nb_jours_periode-1)
    jours_prev=[]
    for h in hist_data:
        if h.get("en_cours"): continue
        d=parse_label_date(h.get("label",""))
        if d and prev_start<=d<=prev_end:
            jours_prev.append((d,h))
    comparaison=None
    if jours_prev:
        prev_l=sum(sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0)) for _,h in jours_prev)
        prev_piste=sum(sf(h.get("piste",0)) for _,h in jours_prev)
        prev_bout=sum(sf(h.get("bout",0)) for _,h in jours_prev)
        def _delta_pct(curr,prev):
            return round((curr-prev)/prev*100,2) if prev>0 else None
        comparaison={
            "label":f"{prev_start.strftime('%d/%m/%Y')} → {prev_end.strftime('%d/%m/%Y')}",
            "litrage_l":int(prev_l),"ca_piste_eur":round(prev_piste,2),"ca_boutique_eur":round(prev_bout,2),
            "ca_total_eur":round(prev_piste+prev_bout,2),
            "delta_litrage_pct":_delta_pct(total_l,prev_l),
            "delta_ca_piste_pct":_delta_pct(ca_piste,prev_piste),
            "delta_ca_boutique_pct":_delta_pct(ca_boutique,prev_bout),
            "delta_ca_total_pct":_delta_pct(ca_total,prev_piste+prev_bout),
        }
    # Détail jour par jour : seulement en mode COMPLET et mono-mois (sinon 365 lignes illisibles)
    jours_detail=[]
    if niveau=="complet" and mono_mois:
        for d,h in jours_periode:
            jours_detail.append({
                "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),"weekday":d.weekday(),
                "ferie":is_ferie(d),
                "sp":int(sf(h.get("sp",0))),"go":int(sf(h.get("go",0))),"gnr":int(sf(h.get("gnr",0))),
                "litrage_total":int(sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))),
                "ca_piste":round(sf(h.get("piste",0)),2),
                "cb":round(sf(h.get("cb",0)),2),"cp":round(sf(h.get("cp",0)),2),
                "esp":round(sf(h.get("esp",0)),2),
                "ca_boutique":round(sf(h.get("bout",0)),2),
            })
    snapshot={
        "version":2,
        "type":"periode",
        "niveau":niveau,
        "mono_mois":mono_mois,
        "start_date":start_date.strftime("%Y-%m-%d"),
        "end_date":end_date.strftime("%Y-%m-%d"),
        "periode_label":_periode_label(start_date,end_date),
        "generated_at":datetime.now().isoformat(),
        "jours_complets":len(jours_periode),
        "nb_jours_periode":nb_jours_periode,
        "nb_mois_couverts":nb_mois_couverts,
        "totaux":{
            "litrage_l":int(total_l),
            "litrage_sp_l":int(total_sp),"litrage_go_l":int(total_go),"litrage_gnr_l":int(total_gnr),
            "ca_piste_eur":round(ca_piste,2),
            "ca_boutique_eur":round(ca_boutique,2),
            "ca_total_eur":round(ca_total,2),
            "encaiss_cb_eur":round(cb,2),
            "encaiss_cp_eur":round(cp,2),
            "encaiss_esp_eur":round(esp,2),
        },
        "moyennes":{
            "litrage_jour":int(total_l/len(jours_periode)) if jours_periode else 0,
            "ca_piste_jour":round(ca_piste/len(jours_periode),2) if jours_periode else 0,
            "ca_boutique_jour":round(ca_boutique/len(jours_periode),2) if jours_periode else 0,
        },
        "marge":{
            "carburant_eur":marge_carb_total,
            "boutique_eur":marge_bout_total,
            "total_eur":marge_grand_total,
            "moyen_unit":round(marge_carb_total/total_l,5) if total_l else 0,
        },
        "detail_mensuel":par_mois,  # toujours présent (utile en Lite ET Complet)
        "top_meilleures":top_meilleures,
        "top_pires":top_pires,
        "top_bout_meilleures":top_bout_meilleures,
        "top_bout_pires":top_bout_pires,
        "comparaison_prec":comparaison,
        "jours":jours_detail,
        "anomalies_tendance":anomalies if anomalies is not None else (
            _format_anomalies_from_journal(start_date.year,start_date.month) if mono_mois else []
        ),
        "ponts_traverses":ponts if ponts is not None else (
            _format_ponts_from_journal(start_date.year,start_date.month) if mono_mois else []
        ),
        "objectif":{
            "obj_ca_eur":round(sf((objectif or {}).get("obj_ca",0)),2),
            "enc_ca_eur":round(sf((objectif or {}).get("enc_ca",0)),2),
            "balance_de_eur":round(sf((alerts or {}).get("balance_de",0)),2),
            "taux_avancement":round(sf((objectif or {}).get("taux",0)),4),
            "cp_pending_count":len((alerts or {}).get("cp_pending",[])),
            "cp_pending_total":round(sum(c.get("montant",0) for c in (alerts or {}).get("cp_pending",[])),2),
            "cp_retard_count":len([c for c in (alerts or {}).get("cp_pending",[]) if c.get("retard",0)>0]),
            "cp_retard_total":round(sum(c.get("montant",0) for c in (alerts or {}).get("cp_pending",[]) if c.get("retard",0)>0),2),
            "dec_pending_count":len((alerts or {}).get("dec_pending",[])),
            "dec_pending_total":round(sum(d.get("montant",0) for d in (alerts or {}).get("dec_pending",[])),2),
            "dec_urgent_count":len([d for d in (alerts or {}).get("dec_pending",[]) if d.get("reste") is not None and d["reste"]<=7]),
            "enc_retard_count":len([e for e in (alerts or {}).get("enc_pending",[]) if e.get("reste") is not None and e["reste"]<0]),
            "enc_retard_total":round(sum(e.get("montant",0) for e in (alerts or {}).get("enc_pending",[]) if e.get("reste") is not None and e["reste"]<0),2),
            "clients_impayes_count":len((alerts or {}).get("clients_impayes",[])),
            "clients_impayes_total":round(sf((alerts or {}).get("cli_total",0)),2),
        } if mono_mois else {
            # Multi-mois : Objectif n'a pas de sens (état instantané), on neutralise
            "obj_ca_eur":0,"enc_ca_eur":0,"balance_de_eur":0,"taux_avancement":0,
            "cp_pending_count":0,"cp_pending_total":0,"cp_retard_count":0,"cp_retard_total":0,
            "dec_pending_count":0,"dec_pending_total":0,"dec_urgent_count":0,
            "enc_retard_count":0,"enc_retard_total":0,
            "clients_impayes_count":0,"clients_impayes_total":0,
        },
    }
    return snapshot


def build_month_snapshot(hist_data, year, month, anomalies=None, ponts=None, alerts=None, objectif=None, prix_data=None):
    """Construit le snapshot d'un mois donné à partir de hist_data.
    Args:
        hist_data: liste de dicts (sortie de _read_hist) avec sp/go/gnr/cb/cp/esp/bout/total/label/caisses
        year, month: int (ex: 2026, 4)
        anomalies: liste optionnelle d'anomalies de tendance détectées dans le mois
        ponts: liste optionnelle de ponts traversés avec statut/cause
        prix_data: dict optionnel {"marge_sp", "marge_go", "marge_gnr", "marge_unit", "pv_sp"...}
                   Si non fourni, on tire depuis le PRIX_HISTO_FILE pour le mois demandé.
    Retourne: dict du snapshot, prêt à être sérialisé en JSON.
    """
    # Filtrer les jours du mois (jours complets uniquement, pas en cours)
    jours_mois=[]
    for h in hist_data:
        if h.get("en_cours"): continue
        d=parse_label_date(h.get("label",""))
        if d and d.year==year and d.month==month:
            jours_mois.append((d,h))
    jours_mois.sort(key=lambda x:x[0])
    # Agrégats globaux
    total_sp=sum(sf(h.get("sp",0)) for d,h in jours_mois)
    total_go=sum(sf(h.get("go",0)) for d,h in jours_mois)
    total_gnr=sum(sf(h.get("gnr",0)) for d,h in jours_mois)
    total_l=total_sp+total_go+total_gnr
    ca_piste=sum(sf(h.get("piste",0)) for d,h in jours_mois)
    ca_boutique=sum(sf(h.get("bout",0)) for d,h in jours_mois)
    cb=sum(sf(h.get("cb",0)) for d,h in jours_mois)
    cp=sum(sf(h.get("cp",0)) for d,h in jours_mois)
    esp=sum(sf(h.get("esp",0)) for d,h in jours_mois)
    ca_total=ca_piste+ca_boutique
    # Top 3 meilleures et pires journées (par CA piste)
    triees=sorted(jours_mois,key=lambda x:sf(x[1].get("piste",0)),reverse=True)
    top3_meilleures=[]
    for d,h in triees[:3]:
        top3_meilleures.append({
            "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
            "litrage":int(sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))),
            "ca_piste":round(sf(h.get("piste",0)),2),
            "ca_boutique":round(sf(h.get("bout",0)),2),
        })
    top3_pires=[]
    for d,h in triees[-3:][::-1]:
        top3_pires.append({
            "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
            "litrage":int(sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))),
            "ca_piste":round(sf(h.get("piste",0)),2),
            "ca_boutique":round(sf(h.get("bout",0)),2),
        })
    # Top 3 par CA boutique (meilleures et pires journées boutique)
    triees_bout=sorted(jours_mois,key=lambda x:sf(x[1].get("bout",0)),reverse=True)
    top3_bout_meilleures=[]
    for d,h in triees_bout[:3]:
        top3_bout_meilleures.append({
            "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
            "ca_boutique":round(sf(h.get("bout",0)),2),
            "ca_piste":round(sf(h.get("piste",0)),2),
        })
    top3_bout_pires=[]
    for d,h in triees_bout[-3:][::-1]:
        top3_bout_pires.append({
            "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),
            "ca_boutique":round(sf(h.get("bout",0)),2),
            "ca_piste":round(sf(h.get("piste",0)),2),
        })
    # Détail jour par jour pour rapport
    jours_detail=[]
    for d,h in jours_mois:
        jours_detail.append({
            "date":d.strftime("%Y-%m-%d"),"label":h.get("label",""),"weekday":d.weekday(),
            "ferie":is_ferie(d),
            "sp":int(sf(h.get("sp",0))),"go":int(sf(h.get("go",0))),"gnr":int(sf(h.get("gnr",0))),
            "litrage_total":int(sf(h.get("sp",0))+sf(h.get("go",0))+sf(h.get("gnr",0))),
            "ca_piste":round(sf(h.get("piste",0)),2),
            "cb":round(sf(h.get("cb",0)),2),"cp":round(sf(h.get("cp",0)),2),
            "esp":round(sf(h.get("esp",0)),2),
            "ca_boutique":round(sf(h.get("bout",0)),2),
        })
    # ============================================================
    # MARGE DU MOIS
    # Source : prix_data fourni explicitement (lecture en cours du fichier),
    # ou défaut : prix_historique.cfg pour le mois concerné.
    # Calcul : marge_eur = ∑(litrage_carb × marge_unitaire_carb)
    # ============================================================
    if prix_data is None:
        prix_data=get_prix_for_month(year,month) or {}
    pa_sp_m=sf(prix_data.get("pa_sp",0))
    pa_go_m=sf(prix_data.get("pa_go",0))
    pa_gnr_m=sf(prix_data.get("pa_gnr",0))
    pv_sp_m=sf(prix_data.get("pv_sp",0))
    pv_go_m=sf(prix_data.get("pv_go",0))
    pv_gnr_m=sf(prix_data.get("pv_gnr",0))
    marge_sp_unit=round(pv_sp_m-pa_sp_m,5) if (pv_sp_m and pa_sp_m) else 0
    marge_go_unit=round(pv_go_m-pa_go_m,5) if (pv_go_m and pa_go_m) else 0
    marge_gnr_unit=round(pv_gnr_m-pa_gnr_m,5) if (pv_gnr_m and pa_gnr_m) else 0
    marge_sp_eur=round(total_sp*marge_sp_unit,2)
    marge_go_eur=round(total_go*marge_go_unit,2)
    marge_gnr_eur=round(total_gnr*marge_gnr_unit,2)
    marge_total_eur=round(marge_sp_eur+marge_go_eur+marge_gnr_eur,2)
    # Marge boutique : taux paramétré (lu depuis prix_data ou défaut 30%)
    marge_boutique_taux=sf(prix_data.get("marge_boutique_taux",0.30)) if prix_data else 0.30
    if marge_boutique_taux<=0 or marge_boutique_taux>1: marge_boutique_taux=0.30
    marge_boutique_eur=round(ca_boutique*marge_boutique_taux,2)
    # Effet spéculation : tente de lire le passage de mois saisi pour CE mois
    # (= passage de mois_precedent → ce mois). S'il n'a pas été saisi, effet=None.
    # Le passage est stocké dans prix_historique.cfg via la dialogue PassageMoisDlg.
    effet_specu=None
    try:
        passage=get_passage_mois(year,month)
        if passage:
            sp=passage["stock_pivot"];vb=passage["ventes_avant_6h"]
            effet_specu=calc_effet_speculation(
                year,month,
                sp["sp"],sp["go"],sp["gnr"],
                vb["sp"],vb["go"],vb["gnr"],
            )
    except Exception as _e: _log_silent_err(exc=_e)
    snapshot={
        "version":1,
        "type":"mensuel",
        "year":year,"month":month,
        "month_name":["janvier","f\u00e9vrier","mars","avril","mai","juin",
                      "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"][month-1],
        "generated_at":datetime.now().isoformat(),
        "jours_complets":len(jours_mois),
        "totaux":{
            "litrage_l":int(total_l),
            "litrage_sp_l":int(total_sp),"litrage_go_l":int(total_go),"litrage_gnr_l":int(total_gnr),
            "ca_piste_eur":round(ca_piste,2),
            "ca_boutique_eur":round(ca_boutique,2),
            "ca_total_eur":round(ca_total,2),
            "encaiss_cb_eur":round(cb,2),
            "encaiss_cp_eur":round(cp,2),
            "encaiss_esp_eur":round(esp,2),
        },
        "moyennes":{
            "litrage_jour":int(total_l/len(jours_mois)) if jours_mois else 0,
            "ca_piste_jour":round(ca_piste/len(jours_mois),2) if jours_mois else 0,
            "ca_boutique_jour":round(ca_boutique/len(jours_mois),2) if jours_mois else 0,
        },
        # Marge du mois — données stratégiques !
        "marge":{
            "total_eur":marge_total_eur,
            "sp_eur":marge_sp_eur,"go_eur":marge_go_eur,"gnr_eur":marge_gnr_eur,
            "sp_unit":marge_sp_unit,"go_unit":marge_go_unit,"gnr_unit":marge_gnr_unit,
            "moyen_unit":round(marge_total_eur/total_l,5) if total_l else 0,
            "prix_vente":{"sp":pv_sp_m,"go":pv_go_m,"gnr":pv_gnr_m},
            "prix_achat":{"sp":pa_sp_m,"go":pa_go_m,"gnr":pa_gnr_m},
            # Marge boutique : taux × CA boutique
            "boutique_taux":marge_boutique_taux,
            "boutique_eur":marge_boutique_eur,
            # Marge totale globale = carburant + boutique
            "grand_total_eur":round(marge_total_eur+marge_boutique_eur,2),
            # Effet spéculation : gain/perte au passage de mois sur stock pivot
            "effet_speculation":effet_specu,
        },
        "top3_meilleures":top3_meilleures,
        "top3_pires":top3_pires,
        "top3_bout_meilleures":top3_bout_meilleures,
        "top3_bout_pires":top3_bout_pires,
        "jours":jours_detail,
        # Anomalies et ponts : si non fournis explicitement, on tire depuis le journal d'événements
        "anomalies_tendance":anomalies if anomalies is not None else _format_anomalies_from_journal(year,month),
        "ponts_traverses":ponts if ponts is not None else _format_ponts_from_journal(year,month),
        # Données du fichier Objectif (état au moment de la génération du snapshot)
        "objectif":{
            "obj_ca_eur":round(sf((objectif or {}).get("obj_ca",0)),2),
            "enc_ca_eur":round(sf((objectif or {}).get("enc_ca",0)),2),
            "balance_de_eur":round(sf((alerts or {}).get("balance_de",0)),2),
            "taux_avancement":round(sf((objectif or {}).get("taux",0)),4),
            # Alertes administratives capturées
            "cp_pending_count":len((alerts or {}).get("cp_pending",[])),
            "cp_pending_total":round(sum(c.get("montant",0) for c in (alerts or {}).get("cp_pending",[])),2),
            "cp_retard_count":len([c for c in (alerts or {}).get("cp_pending",[]) if c.get("retard",0)>0]),
            "cp_retard_total":round(sum(c.get("montant",0) for c in (alerts or {}).get("cp_pending",[]) if c.get("retard",0)>0),2),
            "dec_pending_count":len((alerts or {}).get("dec_pending",[])),
            "dec_pending_total":round(sum(d.get("montant",0) for d in (alerts or {}).get("dec_pending",[])),2),
            "dec_urgent_count":len([d for d in (alerts or {}).get("dec_pending",[]) if d.get("reste") is not None and d["reste"]<=7]),
            "enc_retard_count":len([e for e in (alerts or {}).get("enc_pending",[]) if e.get("reste") is not None and e["reste"]<0]),
            "enc_retard_total":round(sum(e.get("montant",0) for e in (alerts or {}).get("enc_pending",[]) if e.get("reste") is not None and e["reste"]<0),2),
            "clients_impayes_count":len((alerts or {}).get("clients_impayes",[])),
            "clients_impayes_total":round(sf((alerts or {}).get("cli_total",0)),2),
        },
    }
    return snapshot

def save_month_snapshot(snapshot):
    """Écrit le snapshot dans ~/.districarb_hub/snapshots/YYYY-MM.json"""
    try:
        SNAPSHOTS_DIR.mkdir(parents=True,exist_ok=True)
        path=SNAPSHOTS_DIR/f"{snapshot['year']:04d}-{snapshot['month']:02d}.json"
        with open(path,"w",encoding="utf-8") as f:
            json.dump(snapshot,f,ensure_ascii=False,indent=2)
        return path
    except Exception as e:
        print(f"[snapshot] erreur écriture : {e}")
        return None

def load_month_snapshot(year,month):
    """Charge le snapshot d'un mois donné depuis le disque (None si absent)."""
    try:
        path=SNAPSHOTS_DIR/f"{year:04d}-{month:02d}.json"
        if not path.exists(): return None
        with open(path,"r",encoding="utf-8") as f:
            return json.load(f)
    except Exception as _e: _log_silent_err(exc=_e); return None

# =============================================================================
# GÉNÉRATEUR DE RAPPORT MENSUEL
# - Génère le HTML autonome (charte propre, imprimable depuis le navigateur via Ctrl+P)
# - PDF via Microsoft Edge en mode headless (--print-to-pdf) avec rendu IDENTIQUE
# Sortie : ~/Documents/DISTRICARB Rapports/<YYYY>/Rapport_mensuel_<YYYY>_<MM>_<MOIS>.{pdf|html}
# =============================================================================
def _format_eur(v):
    """Formate un nombre en euros avec espace fin comme séparateur de milliers."""
    try: return f"{int(round(float(v))):,} \u20ac".replace(",","\u202f")
    except Exception as _e: _log_silent_err(exc=_e); return "0 \u20ac"

def _format_l(v):
    try: return f"{int(round(float(v))):,} L".replace(",","\u202f")
    except Exception as _e: _log_silent_err(exc=_e); return "0 L"

def generate_monthly_report_html(snapshot,sections=None):
    """Génère un fichier HTML autonome compact (avec @media print pour impression).
    Args:
        sections: dict {clé: bool} ; si None, toutes les sections par défaut sont incluses.
    Retourne le path du fichier HTML écrit, ou None en cas d'échec."""
    # Résolution des sections
    if sections is None:
        sections={key:default for key,label,default in RAPPORT_SECTIONS}
    def want(key): return bool(sections.get(key,False))
    try:
        year=snapshot["year"];month=snapshot["month"]
        out_dir=RAPPORTS_DIR/f"{year:04d}"
        out_dir.mkdir(parents=True,exist_ok=True)
        out_path=out_dir/f"Rapport_mensuel_{year:04d}_{month:02d}_{snapshot['month_name'].capitalize()}.html"
        t=snapshot["totaux"];m=snapshot["moyennes"];nb_j=snapshot["jours_complets"]
        # Moyenne CA boutique par jour
        moy_bout_jour=int(round(t['ca_boutique_eur']/nb_j)) if nb_j else 0
        # Construction tableau jours détaillés
        rows_jours="".join(
            f"<tr class='{'fer' if j['ferie'] else ''}'>"
            f"<td>{j['label']}</td>"
            f"<td class='r'>{_format_l(j['sp'])}</td>"
            f"<td class='r'>{_format_l(j['go'])}</td>"
            f"<td class='r'>{_format_l(j['gnr'])}</td>"
            f"<td class='r b'>{_format_l(j['litrage_total'])}</td>"
            f"<td class='r'>{_format_eur(j['ca_piste'])}</td>"
            f"<td class='r'>{_format_eur(j['ca_boutique'])}</td>"
            f"</tr>" for j in snapshot["jours"]
        )
        # Top 3 meilleures
        rows_top3=""
        for j in snapshot["top3_meilleures"]:
            rows_top3+=f"<tr><td>{j['label']}</td><td class='r'>{_format_l(j['litrage'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
        rows_pires=""
        for j in snapshot["top3_pires"]:
            rows_pires+=f"<tr><td>{j['label']}</td><td class='r'>{_format_l(j['litrage'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
        # Top 3 par CA boutique
        rows_bout_top=""
        for j in snapshot.get("top3_bout_meilleures",[]):
            rows_bout_top+=f"<tr><td>{j['label']}</td><td class='r'>{_format_eur(j['ca_boutique'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
        rows_bout_pires=""
        for j in snapshot.get("top3_bout_pires",[]):
            rows_bout_pires+=f"<tr><td>{j['label']}</td><td class='r'>{_format_eur(j['ca_boutique'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
        # Section Objectif & alertes administratives
        ob=snapshot.get("objectif",{}) or {}
        objectif_rows=""
        # Balance D/E
        bal=ob.get("balance_de_eur",0)
        if bal:
            color_bal="#0e7c3a" if bal>0 else "#b91c1c"
            objectif_rows+=f"<tr><td>Balance D/E</td><td class='r' style='color:{color_bal};font-weight:bold;'>{_format_eur(bal)}</td><td>\u2014</td></tr>"
        # Avancement objectif CA
        taux=ob.get("taux_avancement",0)
        if taux:
            objectif_rows+=f"<tr><td>Avancement objectif CA</td><td class='r'>{taux*100:.1f}%</td><td>Sur objectif {_format_eur(ob.get('obj_ca_eur',0))}</td></tr>"
        # CP en retard
        if ob.get("cp_retard_count",0)>0:
            objectif_rows+=f"<tr><td>\u26a0 CP en retard</td><td class='r' style='color:#b91c1c;'>{_format_eur(ob['cp_retard_total'])}</td><td>{ob['cp_retard_count']} op\u00e9ration(s)</td></tr>"
        # Décaissements à venir
        if ob.get("dec_pending_count",0)>0:
            color_dec="#b91c1c" if ob.get("dec_urgent_count",0)>0 else "#9c5300"
            objectif_rows+=f"<tr><td>D\u00e9caissements \u00e0 venir</td><td class='r' style='color:{color_dec};'>{_format_eur(ob['dec_pending_total'])}</td><td>{ob['dec_pending_count']} pr\u00e9vu(s){', dont '+str(ob['dec_urgent_count'])+' urgent(s) (\u22647j)' if ob.get('dec_urgent_count',0)>0 else ''}</td></tr>"
        # Encaissements en retard
        if ob.get("enc_retard_count",0)>0:
            objectif_rows+=f"<tr><td>\u26a0 Encaissements en retard</td><td class='r' style='color:#b91c1c;'>{_format_eur(ob['enc_retard_total'])}</td><td>{ob['enc_retard_count']} op\u00e9ration(s)</td></tr>"
        # Clients impayés
        if ob.get("clients_impayes_count",0)>0:
            objectif_rows+=f"<tr><td>R\u00e8glements clients en attente</td><td class='r' style='color:#b91c1c;'>{_format_eur(ob['clients_impayes_total'])}</td><td>{ob['clients_impayes_count']} client(s)</td></tr>"
        objectif_section=""
        if objectif_rows:
            objectif_section=f"""
<h2>Pilotage administratif &amp; alertes</h2>
<p style='font-size:0.85em;color:#666;margin:0 0 6px 0;'>\u00c9tat enregistr\u00e9 au moment de la g\u00e9n\u00e9ration du rapport (source : Objectif mensuel).</p>
<table>
<tr><th>Indicateur</th><th style='text-align:right;'>Montant</th><th>D\u00e9tail</th></tr>
{objectif_rows}
</table>"""
        # Anomalies/ponts (si capturés)
        anomalies_section=""
        if want("anomalies") and snapshot.get("anomalies_tendance"):
            rows_a="".join(f"<li>{a}</li>" for a in snapshot["anomalies_tendance"])
            anomalies_section=f"<h3 style='color:#444;margin-top:14px;'>Anomalies de tendance</h3><ul>{rows_a}</ul>"
        ponts_section=""
        if want("ponts") and snapshot.get("ponts_traverses"):
            rows_p="".join(f"<li>{p}</li>" for p in snapshot["ponts_traverses"])
            ponts_section=f"<h3 style='color:#444;margin-top:14px;'>Ponts travers\u00e9s</h3><ul>{rows_p}</ul>"
        # ============================================================
        # SECTIONS CONDITIONNELLES (apparaissent seulement si la case est cochée)
        # ============================================================
        # Section : Synthèse du mois — 6 KPI inchangés (CA total RESTE)
        synthese_section=""
        if want("synthese"):
            synthese_section=f"""<h2>Synth\u00e8se du mois ({nb_j} jours saisis)</h2>
<div class="kpi-grid">
  <div class="kpi"><div class="kpi-label">Total litrage</div><div class="kpi-value">{_format_l(t['litrage_l'])}</div></div>
  <div class="kpi"><div class="kpi-label">CA piste</div><div class="kpi-value">{_format_eur(t['ca_piste_eur'])}</div></div>
  <div class="kpi"><div class="kpi-label">CA boutique</div><div class="kpi-value">{_format_eur(t['ca_boutique_eur'])}</div></div>
  <div class="kpi"><div class="kpi-label">Moy. CA boutique/j</div><div class="kpi-value">{_format_eur(moy_bout_jour)}</div></div>
  <div class="kpi"><div class="kpi-label">Moy. litrage/j</div><div class="kpi-value">{_format_l(m['litrage_jour'])}</div></div>
  <div class="kpi"><div class="kpi-label">CA total</div><div class="kpi-value">{_format_eur(t['ca_total_eur'])}</div></div>
</div>"""
        # Section : Marge totale et détaillée — nouvelle section dédiée, juste après synthèse
        marge_section=""
        if want("synthese"):
            marge_data=snapshot.get("marge",{}) or {}
            marge_carb_total=marge_data.get("total_eur",0)
            marge_bout_taux=marge_data.get("boutique_taux",0.30)
            marge_bout_eur=marge_data.get("boutique_eur",0)
            marge_grand_total=marge_carb_total+marge_bout_eur
            def _fmt_unit(v): return f"{v:.5f} \u20ac/L".replace(".",",") if v else "\u2014"
            marge_section=f"""<h2 style="color:#0e7c3a;">\U0001f4b0 Marge totale et d\u00e9taill\u00e9e</h2>
<div class="kpi-grid">
  <div class="kpi" style="background:#eaf5ec;border-left-color:#0e7c3a;"><div class="kpi-label">Marge carburant</div><div class="kpi-value" style="color:#0e7c3a;">{_format_eur(marge_carb_total)}</div></div>
  <div class="kpi" style="background:#eaf5ec;border-left-color:#0e7c3a;"><div class="kpi-label">Marge boutique ({marge_bout_taux*100:.2f}%)</div><div class="kpi-value" style="color:#0e7c3a;">{_format_eur(marge_bout_eur)}</div></div>
  <div class="kpi" style="background:#d4edda;border-left-color:#0e7c3a;"><div class="kpi-label">Marge totale</div><div class="kpi-value" style="color:#0e7c3a;font-size:1.3em;">{_format_eur(marge_grand_total)}</div></div>
</div>"""
            # Sous-tableau détail par carburant (uniquement si on a les marges unitaires)
            if marge_data.get("sp_unit") or marge_data.get("go_unit") or marge_data.get("gnr_unit"):
                # Détecter si la marge est uniforme (même taux pour SP/GO/GNR)
                m_sp=round(sf(marge_data.get("sp_unit",0)),5)
                m_go=round(sf(marge_data.get("go_unit",0)),5)
                m_gnr=round(sf(marge_data.get("gnr_unit",0)),5)
                # Carburants effectifs (un carburant non vendu n'est pas comparé)
                margins_present=[]
                if t['litrage_sp_l']>0: margins_present.append(m_sp)
                if t['litrage_go_l']>0: margins_present.append(m_go)
                if t['litrage_gnr_l']>0: margins_present.append(m_gnr)
                marges_uniformes=(len(set(margins_present))==1) if margins_present else True
                if marges_uniformes:
                    # Marge uniforme : mention au-dessus + tableau simplifié sans colonne unitaire
                    marge_uniforme=margins_present[0] if margins_present else 0
                    marge_uniforme_str=f"{marge_uniforme:.5f}".replace(".",",")
                    marge_section+=f"""
<h3>D\u00e9tail marge carburant par produit</h3>
<p style='font-size:0.9em;color:#555;margin:0 0 6px 0;'>
<em>Marge r\u00e9glementaire appliqu\u00e9e : {marge_uniforme_str} \u20ac/L (uniforme sur les 3 carburants)</em>
</p>
<table>
<tr><th>Carburant</th><th style='text-align:right;'>Volume vendu</th><th style='text-align:right;'>Marge totale</th></tr>
<tr><td>Sans Plomb (SP)</td><td class='r'>{_format_l(t['litrage_sp_l'])}</td><td class='r'>{_format_eur(marge_data.get('sp_eur',0))}</td></tr>
<tr><td>Gazole (GO)</td><td class='r'>{_format_l(t['litrage_go_l'])}</td><td class='r'>{_format_eur(marge_data.get('go_eur',0))}</td></tr>
<tr><td>GNR</td><td class='r'>{_format_l(t['litrage_gnr_l'])}</td><td class='r'>{_format_eur(marge_data.get('gnr_eur',0))}</td></tr>
<tr style='background:#f5f7fa;font-weight:bold;'><td>Total carburant</td><td class='r'>{_format_l(t['litrage_l'])}</td><td class='r'>{_format_eur(marge_carb_total)}</td></tr>
</table>"""
                else:
                    # Marges différenciées : on garde la colonne unitaire pour montrer la différence
                    marge_section+=f"""
<h3>D\u00e9tail marge carburant par produit</h3>
<p style='font-size:0.9em;color:#555;margin:0 0 6px 0;'>
<em>Marges r\u00e9glementaires diff\u00e9renci\u00e9es ce mois</em>
</p>
<table>
<tr><th>Carburant</th><th style='text-align:right;'>Volume vendu</th><th style='text-align:right;'>Marge unitaire</th><th style='text-align:right;'>Marge totale</th></tr>
<tr><td>Sans Plomb (SP)</td><td class='r'>{_format_l(t['litrage_sp_l'])}</td><td class='r'>{_fmt_unit(m_sp)}</td><td class='r'>{_format_eur(marge_data.get('sp_eur',0))}</td></tr>
<tr><td>Gazole (GO)</td><td class='r'>{_format_l(t['litrage_go_l'])}</td><td class='r'>{_fmt_unit(m_go)}</td><td class='r'>{_format_eur(marge_data.get('go_eur',0))}</td></tr>
<tr><td>GNR</td><td class='r'>{_format_l(t['litrage_gnr_l'])}</td><td class='r'>{_fmt_unit(m_gnr)}</td><td class='r'>{_format_eur(marge_data.get('gnr_eur',0))}</td></tr>
<tr style='background:#f5f7fa;font-weight:bold;'><td>Total carburant</td><td class='r'>{_format_l(t['litrage_l'])}</td><td class='r'>{_fmt_unit(marge_data.get('moyen_unit',0))}</td><td class='r'>{_format_eur(marge_carb_total)}</td></tr>
</table>"""
            # Effet spéculation
            eff=marge_data.get("effet_speculation")
            if eff and eff.get("total"):
                tot_eff=eff.get("total",0)
                color="#0e7c3a" if tot_eff>=0 else "#b91c1c"
                signe="+" if tot_eff>=0 else ""
                pa=eff.get("prix_avant",{});pap=eff.get("prix_apres",{})
                stocks=eff.get("stocks_pivot",{})
                def _delta(k):
                    a=pa.get(k,0);b=pap.get(k,0)
                    if not a: return "\u2014"
                    diff=b-a;sgn="+" if diff>=0 else ""
                    return f"{sgn}{diff:.3f}".replace(".",",")+" \u20ac/L"
                marge_section+=f"""
<h3 style="color:{color};">\U0001f3af Effet sp\u00e9culation au 1er du mois</h3>
<p style='font-size:0.9em;color:#555;margin:0 0 6px 0;'>
Le stock pr\u00e9sent en cuve au matin du 1er (= stock 6h + ventes 0h-6h) a \u00e9t\u00e9 achet\u00e9 au
prix du mois pr\u00e9c\u00e9dent mais vendu au prix de ce mois.
</p>
<table>
<tr><th>Carburant</th><th style='text-align:right;'>Stock pivot</th><th style='text-align:right;'>\u0394 prix vente</th><th style='text-align:right;'>Effet</th></tr>
<tr><td>SP</td><td class='r'>{_format_l(stocks.get('sp',0))}</td><td class='r'>{_delta('sp')}</td><td class='r'>{_format_eur(eff.get('sp',0))}</td></tr>
<tr><td>GO</td><td class='r'>{_format_l(stocks.get('go',0))}</td><td class='r'>{_delta('go')}</td><td class='r'>{_format_eur(eff.get('go',0))}</td></tr>
<tr><td>GNR</td><td class='r'>{_format_l(stocks.get('gnr',0))}</td><td class='r'>{_delta('gnr')}</td><td class='r'>{_format_eur(eff.get('gnr',0))}</td></tr>
<tr style='background:#f5f7fa;font-weight:bold;'><td colspan='3'>{("Gain" if tot_eff>=0 else "Perte")} sur sp\u00e9culation</td><td class='r' style='color:{color};'>{signe}{_format_eur(tot_eff)}</td></tr>
</table>"""
        # Section : Répartition par carburant
        carburants_section=""
        if want("carburants"):
            tot=t['litrage_l'] or 1
            carburants_section=f"""<h3>R\u00e9partition par carburant</h3>
<table>
<tr><th>Carburant</th><th style='text-align:right;'>Volume vendu</th><th style='text-align:right;'>Part</th></tr>
<tr><td>Sans Plomb (SP)</td><td class='r'>{_format_l(t['litrage_sp_l'])}</td><td class='r'>{(100*t['litrage_sp_l']/tot):.1f}%</td></tr>
<tr><td>Gazole (GO)</td><td class='r'>{_format_l(t['litrage_go_l'])}</td><td class='r'>{(100*t['litrage_go_l']/tot):.1f}%</td></tr>
<tr><td>GNR</td><td class='r'>{_format_l(t['litrage_gnr_l'])}</td><td class='r'>{(100*t['litrage_gnr_l']/tot):.1f}%</td></tr>
</table>"""
        # Section : Encaissements
        encaissements_section=""
        if want("encaissements"):
            encaissements_section=f"""<h3>Encaissements</h3>
<table>
<tr><th>Mode</th><th style='text-align:right;'>Montant</th></tr>
<tr><td>Carte Bancaire (CB)</td><td class='r'>{_format_eur(t['encaiss_cb_eur'])}</td></tr>
<tr><td>Carte Privative (CP)</td><td class='r'>{_format_eur(t['encaiss_cp_eur'])}</td></tr>
<tr><td>Esp\u00e8ces</td><td class='r'>{_format_eur(t['encaiss_esp_eur'])}</td></tr>
</table>"""
        # Désactiver objectif_section si pas coché
        if not want("admin"):
            objectif_section=""
        # Section : Détail jour par jour
        detail_jours_section=""
        if want("detail_jours"):
            detail_jours_section=f"""<div class='pagebreak'></div>
<h2>D\u00e9tail jour par jour</h2>
<table>
<tr><th>Jour</th><th style='text-align:right;'>SP</th><th style='text-align:right;'>GO</th><th style='text-align:right;'>GNR</th><th style='text-align:right;'>Total L</th><th style='text-align:right;'>CA piste</th><th style='text-align:right;'>CA boutique</th></tr>
{rows_jours}
</table>
<p style='font-size:8pt;color:#888;margin:4px 0 0 0;'>Lignes ambr\u00e9es : jours f\u00e9ri\u00e9s.</p>"""
        # Section : Événements remarquables (Top 3)
        evenements_parts=[]
        any_top3=any(want(k) for k in ("top3_piste_meilleures","top3_piste_pires","top3_bout_meilleures","top3_bout_pires"))
        if any_top3:
            evenements_parts.append("<h2 style='margin-top:18px;'>\u00c9v\u00e9nements remarquables</h2>")
        if want("top3_piste_meilleures"):
            evenements_parts.append(f"""<h3>Top 3 meilleures journ\u00e9es (CA piste)</h3>
<table><tr><th>Jour</th><th style='text-align:right;'>Litrage</th><th style='text-align:right;'>CA piste</th></tr>
{rows_top3}
</table>""")
        if want("top3_piste_pires"):
            evenements_parts.append(f"""<h3>Top 3 plus faibles journ\u00e9es (CA piste)</h3>
<table><tr><th>Jour</th><th style='text-align:right;'>Litrage</th><th style='text-align:right;'>CA piste</th></tr>
{rows_pires}
</table>""")
        if want("top3_bout_meilleures"):
            evenements_parts.append(f"""<h3>Top 3 meilleures journ\u00e9es (CA boutique)</h3>
<table><tr><th>Jour</th><th style='text-align:right;'>CA boutique</th><th style='text-align:right;'>CA piste</th></tr>
{rows_bout_top}
</table>""")
        if want("top3_bout_pires"):
            evenements_parts.append(f"""<h3>Top 3 plus faibles journ\u00e9es (CA boutique)</h3>
<table><tr><th>Jour</th><th style='text-align:right;'>CA boutique</th><th style='text-align:right;'>CA piste</th></tr>
{rows_bout_pires}
</table>""")
        evenements_section="\n".join(evenements_parts)
        html=f"""<!DOCTYPE html>
<html lang="fr"><head>
<meta charset="utf-8">
<title>Rapport mensuel {snapshot['month_name']} {year} \u2014 B. DISTRICARB SARL</title>
<style>
@page {{ size:A4; margin:1cm; }}
body {{ font-family:'Segoe UI',Arial,sans-serif; color:#222; max-width:21cm; margin:0 auto; padding:1cm; line-height:1.4; font-size:10pt; }}
h1 {{ color:#1a3a5c; border-bottom:2px solid #1a3a5c; padding-bottom:4px; margin:0 0 2px 0; font-size:18pt; }}
h2 {{ color:#1a3a5c; margin:22px 0 8px 0; border-bottom:1px solid #cfd8dc; padding-bottom:3px; font-size:13pt; }}
h3 {{ color:#444; margin:18px 0 6px 0; font-size:11pt; font-weight:600; }}
.subtitle {{ color:#666; font-size:9pt; margin-bottom:14px; }}
.kpi-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:8px; margin:10px 0 6px 0; }}
.kpi {{ background:#f5f7fa; border-left:3px solid #1a3a5c; padding:8px 12px; border-radius:3px; }}
.kpi-label {{ font-size:8.5pt; color:#555; text-transform:uppercase; letter-spacing:0.3px; }}
.kpi-value {{ font-size:13.5pt; font-weight:bold; color:#1a3a5c; margin-top:3px; }}
table {{ width:100%; border-collapse:collapse; margin:6px 0 12px 0; font-size:9.5pt; }}
th {{ background:#1a3a5c; color:#fff; text-align:left; padding:5px 8px; font-weight:600; font-size:9.5pt; }}
td {{ padding:4px 8px; border-bottom:1px solid #e0e6ed; }}
tr.fer td {{ background:#fff4e0; }}
td.r {{ text-align:right; }}
td.b {{ font-weight:bold; }}
.footer {{ margin-top:20px; font-size:8pt; color:#888; text-align:center; border-top:1px solid #ddd; padding-top:6px; }}
@media print {{ body {{ padding:0; max-width:none; }} .pagebreak {{ page-break-before:always; }} h2 {{ page-break-after:avoid; }} h3 {{ page-break-after:avoid; }} }}
</style></head>
<body>

<h1>Rapport mensuel \u2014 {snapshot['month_name'].capitalize()} {year}</h1>
<div class="subtitle">B. DISTRICARB SARL \u2014 Station-service Le Lamentin \u2014 G\u00e9n\u00e9r\u00e9 le {datetime.fromisoformat(snapshot['generated_at']).strftime('%d/%m/%Y \u00e0 %Hh%M')}</div>

{synthese_section}

{marge_section}

{carburants_section}

{encaissements_section}

{objectif_section}

{detail_jours_section}

{evenements_section}

{anomalies_section}
{ponts_section}

<div class="footer">
DISTRICARB HUB v0.5 \u2014 Rapport g\u00e9n\u00e9r\u00e9 automatiquement \u2014 Source : LITRAGE.xlsx + Pr\u00e9vision compte.xlsx + Objectif mensuel.xlsx
</div>
</body></html>"""
        with open(out_path,"w",encoding="utf-8") as f:
            f.write(html)
        return out_path
    except Exception as e:
        print(f"[rapport HTML] erreur : {e}")
        return None

def generate_period_report_html(snapshot,sections=None):
    """Génère un HTML autonome pour un snapshot période (multi-mois ou mono-mois plage libre).
    Reprend EXACTEMENT la charte visuelle du rapport mensuel : @page A4, palette bleu
    marine #1a3a5c, KPI grid, tables sobres. Adapte le contenu selon snapshot["niveau"]
    (lite/complet) et snapshot["mono_mois"]."""
    if sections is None:
        sections={key:default for key,label,default in RAPPORT_SECTIONS}
    def want(key): return bool(sections.get(key,False))
    try:
        sd=snapshot["start_date"];ed=snapshot["end_date"]
        safe_label=snapshot["periode_label"].replace(" ","_").replace("/","-").replace("\u2192","a")
        for src,dst in [("\u00e9","e"),("\u00e8","e"),("\u00ea","e"),("\u00e0","a"),("\u00e2","a"),("\u00f4","o"),("\u00fb","u"),("\u00e7","c"),("\u00c9","E"),("\u00c8","E")]:
            safe_label=safe_label.replace(src,dst)
        out_dir=RAPPORTS_DIR/sd[:4]
        out_dir.mkdir(parents=True,exist_ok=True)
        out_path=out_dir/f"Rapport_periode_{sd}_{ed}_{safe_label}.html"
        t=snapshot["totaux"];m=snapshot["moyennes"];nb_j=snapshot["jours_complets"]
        niveau=snapshot.get("niveau","complet")
        mono=snapshot.get("mono_mois",False)
        nb_mois=snapshot.get("nb_mois_couverts",0)
        ma=snapshot.get("marge",{}) or {}
        # Moyenne CA boutique par jour
        moy_bout_jour=int(round(t['ca_boutique_eur']/nb_j)) if nb_j else 0
        # ===== SECTION SYNTHÈSE (KPI grid 3 colonnes, identique mensuel) =====
        synthese_section=f"""<h2>Synth\u00e8se</h2>
<div class="kpi-grid">
<div class="kpi"><div class="kpi-label">Litrage total</div><div class="kpi-value">{_format_l(t['litrage_l'])}</div></div>
<div class="kpi"><div class="kpi-label">CA piste</div><div class="kpi-value">{_format_eur(t['ca_piste_eur'])}</div></div>
<div class="kpi"><div class="kpi-label">CA boutique</div><div class="kpi-value">{_format_eur(t['ca_boutique_eur'])}</div></div>
<div class="kpi"><div class="kpi-label">CA total</div><div class="kpi-value">{_format_eur(t['ca_total_eur'])}</div></div>
<div class="kpi"><div class="kpi-label">Moy. litrage / jour</div><div class="kpi-value">{_format_l(m['litrage_jour'])}</div></div>
<div class="kpi"><div class="kpi-label">Moy. CA piste / jour</div><div class="kpi-value">{_format_eur(m['ca_piste_jour'])}</div></div>
</div>
<p style='font-size:9pt;color:#666;margin:6px 0 0 0;'>{snapshot['nb_jours_periode']} jour(s) calendaire(s) sur la p\u00e9riode \u2014 {nb_j} jour(s) avec donn\u00e9es saisies{(" \u2014 "+str(nb_mois)+" mois couverts") if not mono else ""}.</p>
"""
        # ===== SECTION MARGE (vignettes vertes, identique mensuel) =====
        marge_carb=ma.get("carburant_eur",0)
        marge_bout=ma.get("boutique_eur",0)
        marge_tot=ma.get("total_eur",0)
        marge_unit=ma.get("moyen_unit",0)
        # Taux boutique appliqué : on regarde le détail mensuel pour voir s'il est uniforme
        detail=snapshot.get("detail_mensuel",[]) or []
        # Taux boutique effectif = marge_bout / bout (si bout > 0)
        taux_bout=(marge_bout/t['ca_boutique_eur']*100) if t['ca_boutique_eur']>0 else 0
        marge_section=f"""<h2 style="color:#0e7c3a;">\U0001f4b0 Marge totale</h2>
<div class="kpi-grid">
<div class="kpi marge"><div class="kpi-label">Marge carburant</div><div class="kpi-value">{_format_eur(marge_carb)}</div></div>
<div class="kpi marge"><div class="kpi-label">Marge boutique ({taux_bout:.1f}%)</div><div class="kpi-value">{_format_eur(marge_bout)}</div></div>
<div class="kpi marge"><div class="kpi-label">Marge totale</div><div class="kpi-value">{_format_eur(marge_tot)}</div></div>
</div>
<p style='font-size:9pt;color:#666;margin:6px 0 0 0;'>Marge moyenne carburant : {marge_unit:.5f} \u20ac/L sur l'ensemble de la p\u00e9riode. Calcul mois par mois selon le palier r\u00e9glementaire applicable, puis somm\u00e9.</p>
"""
        # ===== SECTION COMPARAISON PÉRIODE PRÉCÉDENTE =====
        comp=snapshot.get("comparaison_prec")
        comparaison_section=""
        if comp:
            def _delta(pct):
                if pct is None: return "<span style='color:#888;'>n/a</span>"
                color="#0e7c3a" if pct>=0 else "#b3261e"
                signe="+" if pct>=0 else ""
                return f"<span style='color:{color};font-weight:bold;'>{signe}{pct:.1f}%</span>"
            comparaison_section=f"""<h2>\u00c9volution vs p\u00e9riode pr\u00e9c\u00e9dente</h2>
<p style='font-size:9pt;color:#666;margin:0 0 6px 0;'>Plage pr\u00e9c\u00e9dente de dur\u00e9e \u00e9quivalente : {comp['label']}</p>
<table>
<tr><th>Indicateur</th><th style='text-align:right;'>P\u00e9riode actuelle</th><th style='text-align:right;'>P\u00e9riode pr\u00e9c\u00e9dente</th><th style='text-align:right;'>\u00c9volution</th></tr>
<tr><td>Litrage total</td><td class='r'>{_format_l(t['litrage_l'])}</td><td class='r'>{_format_l(comp['litrage_l'])}</td><td class='r'>{_delta(comp['delta_litrage_pct'])}</td></tr>
<tr><td>CA piste</td><td class='r'>{_format_eur(t['ca_piste_eur'])}</td><td class='r'>{_format_eur(comp['ca_piste_eur'])}</td><td class='r'>{_delta(comp['delta_ca_piste_pct'])}</td></tr>
<tr><td>CA boutique</td><td class='r'>{_format_eur(t['ca_boutique_eur'])}</td><td class='r'>{_format_eur(comp['ca_boutique_eur'])}</td><td class='r'>{_delta(comp['delta_ca_boutique_pct'])}</td></tr>
<tr><td><b>CA total</b></td><td class='r b'>{_format_eur(t['ca_total_eur'])}</td><td class='r b'>{_format_eur(comp['ca_total_eur'])}</td><td class='r'>{_delta(comp['delta_ca_total_pct'])}</td></tr>
</table>
"""
        # ===== SECTION DÉTAIL MENSUEL =====
        rows_mensuel="".join(
            f"<tr>"
            f"<td>{mm['label']}</td>"
            f"<td class='r'>{mm['nb_jours']}</td>"
            f"<td class='r'>{_format_l(mm['litrage_total'])}</td>"
            f"<td class='r'>{_format_eur(mm['piste'])}</td>"
            f"<td class='r'>{_format_eur(mm['bout'])}</td>"
            f"<td class='r b'>{_format_eur(mm['ca_total'])}</td>"
            f"<td class='r'>{_format_eur(mm['marge_total'])}</td>"
            f"</tr>" for mm in detail
        )
        detail_mensuel_section=f"""<h2>D\u00e9tail mensuel</h2>
<table>
<tr><th>Mois</th><th style='text-align:right;'>Jours</th><th style='text-align:right;'>Litrage</th><th style='text-align:right;'>CA piste</th><th style='text-align:right;'>CA boutique</th><th style='text-align:right;'>CA total</th><th style='text-align:right;'>Marge</th></tr>
{rows_mensuel}
</table>
"""
        # ===== DÉTAIL JOUR PAR JOUR (mono-mois COMPLET seulement) =====
        detail_jours_section=""
        if niveau=="complet" and mono and snapshot.get("jours"):
            rows_jours="".join(
                f"<tr class='{'fer' if j.get('ferie') else ''}'>"
                f"<td>{j['label']}</td>"
                f"<td class='r'>{_format_l(j['sp'])}</td>"
                f"<td class='r'>{_format_l(j['go'])}</td>"
                f"<td class='r'>{_format_l(j['gnr'])}</td>"
                f"<td class='r b'>{_format_l(j['litrage_total'])}</td>"
                f"<td class='r'>{_format_eur(j['ca_piste'])}</td>"
                f"<td class='r'>{_format_eur(j['ca_boutique'])}</td>"
                f"</tr>" for j in snapshot["jours"]
            )
            detail_jours_section=f"""<h2>D\u00e9tail jour par jour</h2>
<table>
<tr><th>Jour</th><th style='text-align:right;'>SP</th><th style='text-align:right;'>GO</th><th style='text-align:right;'>GNR</th><th style='text-align:right;'>Total L</th><th style='text-align:right;'>CA piste</th><th style='text-align:right;'>CA boutique</th></tr>
{rows_jours}
</table>
<p style='font-size:8pt;color:#888;margin:4px 0 0 0;'>Lignes ambr\u00e9es : jours f\u00e9ri\u00e9s.</p>
"""
        # ===== ÉVÉNEMENTS REMARQUABLES (Top 3) =====
        top_unit="journ\u00e9es" if mono else "mois"
        rows_top="".join(
            f"<tr><td>{j['label']}</td><td class='r'>{_format_l(j['litrage'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
            for j in snapshot.get("top_meilleures",[])
        )
        rows_pires="".join(
            f"<tr><td>{j['label']}</td><td class='r'>{_format_l(j['litrage'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
            for j in snapshot.get("top_pires",[])
        )
        rows_bout_top="".join(
            f"<tr><td>{j['label']}</td><td class='r'>{_format_eur(j['ca_boutique'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
            for j in snapshot.get("top_bout_meilleures",[])
        )
        rows_bout_pires="".join(
            f"<tr><td>{j['label']}</td><td class='r'>{_format_eur(j['ca_boutique'])}</td><td class='r'>{_format_eur(j['ca_piste'])}</td></tr>"
            for j in snapshot.get("top_bout_pires",[])
        )
        col_label="Jour" if mono else "Mois"
        evenements_section=f"""<h2>\u00c9v\u00e9nements remarquables</h2>
<h3>Top 3 meilleures {top_unit} (CA piste)</h3>
<table><tr><th>{col_label}</th><th style='text-align:right;'>Litrage</th><th style='text-align:right;'>CA piste</th></tr>
{rows_top}
</table>
<h3>Top 3 plus faibles {top_unit} (CA piste)</h3>
<table><tr><th>{col_label}</th><th style='text-align:right;'>Litrage</th><th style='text-align:right;'>CA piste</th></tr>
{rows_pires}
</table>
<h3>Top 3 meilleures {top_unit} (CA boutique)</h3>
<table><tr><th>{col_label}</th><th style='text-align:right;'>CA boutique</th><th style='text-align:right;'>CA piste</th></tr>
{rows_bout_top}
</table>
<h3>Top 3 plus faibles {top_unit} (CA boutique)</h3>
<table><tr><th>{col_label}</th><th style='text-align:right;'>CA boutique</th><th style='text-align:right;'>CA piste</th></tr>
{rows_bout_pires}
</table>
"""
        # ===== ANOMALIES & PONTS (mono-mois 2026+ seulement) =====
        anomalies_section=""
        anos=snapshot.get("anomalies_tendance",[]) or []
        if niveau=="complet" and anos:
            rows_anos="".join(
                f"<tr><td>{a.get('date','\u2014')}</td><td>{a.get('type','\u2014')}</td><td>{a.get('description','\u2014')}</td></tr>"
                for a in anos
            )
            anomalies_section=f"""<h2>Anomalies de tendance</h2>
<table><tr><th>Date</th><th>Type</th><th>Description</th></tr>
{rows_anos}
</table>
"""
        ponts_section=""
        ponts=snapshot.get("ponts_traverses",[]) or []
        if niveau=="complet" and ponts:
            rows_ponts="".join(
                f"<tr><td>{p.get('date','\u2014')}</td><td>{p.get('statut','\u2014')}</td><td>{p.get('cause','\u2014')}</td></tr>"
                for p in ponts
            )
            ponts_section=f"""<h2>Ponts travers\u00e9s</h2>
<table><tr><th>Date</th><th>Statut</th><th>Cause</th></tr>
{rows_ponts}
</table>
"""
        # ===== BANDEAU LITE/COMPLET (sobre, comme un sous-titre) =====
        niveau_note=""
        if niveau=="lite":
            niveau_note=" \u2014 Rapport Lite (p\u00e9riode antérieure \u00e0 2025)"
        # ===== ASSEMBLAGE final avec exactement la charte du mensuel =====
        from datetime import datetime as _dt
        gen_at=_dt.fromisoformat(snapshot["generated_at"]).strftime("%d/%m/%Y \u00e0 %Hh%M")
        # Convertir start_date / end_date string ISO → date pour affichage humain
        sd_d=date.fromisoformat(sd);ed_d=date.fromisoformat(ed)
        plage_h=f"du {sd_d.strftime('%d/%m/%Y')} au {ed_d.strftime('%d/%m/%Y')}"
        html=f"""<!DOCTYPE html>
<html lang="fr"><head>
<meta charset="utf-8">
<title>Rapport {snapshot['periode_label']} \u2014 B. DISTRICARB SARL</title>
<style>
@page {{ size:A4; margin:1cm; }}
body {{ font-family:'Segoe UI',Arial,sans-serif; color:#222; max-width:21cm; margin:0 auto; padding:1cm; line-height:1.4; font-size:10pt; }}
h1 {{ color:#1a3a5c; border-bottom:2px solid #1a3a5c; padding-bottom:4px; margin:0 0 2px 0; font-size:18pt; }}
h2 {{ color:#1a3a5c; margin:22px 0 8px 0; border-bottom:1px solid #cfd8dc; padding-bottom:3px; font-size:13pt; }}
h3 {{ color:#444; margin:18px 0 6px 0; font-size:11pt; font-weight:600; }}
.subtitle {{ color:#666; font-size:9pt; margin-bottom:14px; }}
.kpi-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:8px; margin:10px 0 6px 0; }}
.kpi {{ background:#f5f7fa; border-left:3px solid #1a3a5c; padding:8px 12px; border-radius:3px; }}
.kpi.marge {{ background:#eaf5ec; border-left-color:#0e7c3a; }}
.kpi-label {{ font-size:8.5pt; color:#555; text-transform:uppercase; letter-spacing:0.3px; }}
.kpi-value {{ font-size:13.5pt; font-weight:bold; color:#1a3a5c; margin-top:3px; }}
.kpi.marge .kpi-value {{ color:#0e7c3a; }}
table {{ width:100%; border-collapse:collapse; margin:6px 0 12px 0; font-size:9.5pt; }}
th {{ background:#1a3a5c; color:#fff; text-align:left; padding:5px 8px; font-weight:600; font-size:9.5pt; }}
td {{ padding:4px 8px; border-bottom:1px solid #e0e6ed; }}
tr.fer td {{ background:#fff4e0; }}
td.r {{ text-align:right; }}
td.b {{ font-weight:bold; }}
.footer {{ margin-top:20px; font-size:8pt; color:#888; text-align:center; border-top:1px solid #ddd; padding-top:6px; }}
@media print {{ body {{ padding:0; max-width:none; }} .pagebreak {{ page-break-before:always; }} h2 {{ page-break-after:avoid; }} h3 {{ page-break-after:avoid; }} }}
</style></head>
<body>

<h1>Rapport \u2014 {snapshot['periode_label']}</h1>
<div class="subtitle">B. DISTRICARB SARL \u2014 Station-service Le Lamentin \u2014 P\u00e9riode {plage_h}{niveau_note} \u2014 G\u00e9n\u00e9r\u00e9 le {gen_at}</div>

{synthese_section}

{marge_section}

{comparaison_section}

{detail_mensuel_section}

{detail_jours_section}

{evenements_section}

{anomalies_section}
{ponts_section}

<div class="footer">
DISTRICARB HUB v0.5 \u2014 Rapport g\u00e9n\u00e9r\u00e9 automatiquement \u2014 Source : LITRAGE.xlsx + Pr\u00e9vision compte.xlsx
</div>
</body></html>"""
        with open(out_path,"w",encoding="utf-8") as f:
            f.write(html)
        return out_path
    except Exception as e:
        print(f"[rapport HTML p\u00e9riode] erreur : {e}")
        import traceback;traceback.print_exc()
        return None


def generate_report_html(snapshot,sections=None):
    """Wrapper : aiguille vers generate_monthly_report_html (snapshots type 'mensuel')
    ou generate_period_report_html (snapshots type 'periode'). Garantit zéro régression
    sur les rapports mensuels existants — snapshot mensuel = code historique inchangé."""
    if snapshot.get("type")=="periode":
        return generate_period_report_html(snapshot,sections=sections)
    return generate_monthly_report_html(snapshot,sections=sections)


def _find_edge_exe():
    """Cherche msedge.exe dans les chemins d'installation standard Windows.
    Edge est imposé sur Windows 10/11, donc présent dans 99% des cas.
    Retourne le chemin si trouvé, None sinon."""
    candidates=[
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    ]
    # Variante user-install (rare)
    try:
        local_app=os.environ.get("LOCALAPPDATA","")
        if local_app:
            candidates.append(os.path.join(local_app,"Microsoft","Edge","Application","msedge.exe"))
    except Exception as _e: _log_silent_err(exc=_e)
    for p in candidates:
        if os.path.exists(p): return p
    return None

def generate_period_report_pdf(snapshot,sections=None):
    """Génère le PDF d'un rapport périodique via Microsoft Edge en mode headless.
    Garantit un rendu IDENTIQUE au HTML (Edge utilise le moteur Chromium pour le rendu).

    Pourquoi Edge et pas une lib Python (WeasyPrint/Playwright) :
      - WeasyPrint nécessite des DLL GTK natives absentes sur Windows par défaut.
      - Playwright télécharge un Chromium NON signé Microsoft que Smart App Control bloque.
      - Edge est DÉJÀ installé sur Windows 11 et signé Microsoft → zéro install, zéro blocage SAC.

    Stratégie : on génère le HTML, puis on appelle `msedge.exe --headless --print-to-pdf`.
    """
    try:
        html_path=generate_period_report_html(snapshot,sections=sections)
        if not html_path: return None
        pdf_path=html_path.with_suffix(".pdf")
        edge_exe=_find_edge_exe()
        if not edge_exe:
            print("[pdf] msedge.exe non trouv\u00e9 sur ce poste")
            return None
        # Edge en headless avec print-to-pdf : rendu identique à Ctrl+P depuis Edge
        result=subprocess.run([
            edge_exe,
            "--headless=new",
            "--disable-gpu",
            "--no-pdf-header-footer",
            "--print-to-pdf-no-header",
            f"--print-to-pdf={pdf_path}",
            f"file:///{html_path.as_posix()}",
        ],capture_output=True,timeout=90)
        if pdf_path.exists() and pdf_path.stat().st_size>0:
            return pdf_path
        stderr=result.stderr.decode("utf-8",errors="ignore") if result.stderr else "(pas de stderr)"
        print(f"[pdf] Edge n'a pas produit le PDF. stderr={stderr[:500]}")
        return None
    except Exception as e:
        print(f"[generate_period_report_pdf edge] {e}")
        return None

def generate_monthly_report_pdf(snapshot,sections=None):
    """Génère le PDF du rapport mensuel via Microsoft Edge en mode headless.
    Garantit un rendu IDENTIQUE au HTML (Edge utilise le moteur Chromium pour le rendu).

    Pourquoi Edge et pas une lib Python (WeasyPrint/Playwright) :
      - WeasyPrint nécessite des DLL GTK natives absentes sur Windows par défaut.
      - Playwright télécharge un Chromium NON signé Microsoft que Smart App Control bloque.
      - Edge est DÉJÀ installé sur Windows 11 et signé Microsoft → zéro install, zéro blocage SAC.

    Stratégie : on génère le HTML, puis on appelle `msedge.exe --headless --print-to-pdf`.

    Args:
        sections: dict {clé: bool} ; si None, toutes les sections par défaut sont incluses.
    """
    try:
        html_path=generate_monthly_report_html(snapshot,sections=sections)
        if not html_path: return None
        pdf_path=html_path.with_suffix(".pdf")
        edge_exe=_find_edge_exe()
        if not edge_exe:
            print("[pdf] msedge.exe non trouv\u00e9 sur ce poste")
            return None
        result=subprocess.run([
            edge_exe,
            "--headless=new",
            "--disable-gpu",
            "--no-pdf-header-footer",
            "--print-to-pdf-no-header",
            f"--print-to-pdf={pdf_path}",
            f"file:///{html_path.as_posix()}",
        ],capture_output=True,timeout=90)
        if pdf_path.exists() and pdf_path.stat().st_size>0:
            return pdf_path
        stderr=result.stderr.decode("utf-8",errors="ignore") if result.stderr else "(pas de stderr)"
        print(f"[pdf] Edge n'a pas produit le PDF. stderr={stderr[:500]}")
        return None
    except Exception as e:
        print(f"[generate_monthly_report_pdf edge] {e}")
        return None

def archive_existing_report(year,month):
    """Si un rapport existe déjà pour ce mois, le renomme avec suffixe _archive_YYYYMMDD_HHMM.
    Retourne le path archivé ou None si rien à archiver."""
    out_dir=RAPPORTS_DIR/f"{year:04d}"
    if not out_dir.exists(): return None
    # Chercher fichier existant non-archive (pdf ou html)
    mn_capit=["Janvier","F\u00e9vrier","Mars","Avril","Mai","Juin","Juillet","Ao\u00fbt","Septembre","Octobre","Novembre","D\u00e9cembre"][month-1]
    base=f"Rapport_mensuel_{year:04d}_{month:02d}_{mn_capit}"
    archived=[]
    for ext in ["pdf","html"]:
        existing=out_dir/f"{base}.{ext}"
        if existing.exists():
            ts=datetime.now().strftime("%Y%m%d_%H%M")
            archive_path=out_dir/f"{base}_archive_{ts}.{ext}"
            try:
                existing.rename(archive_path)
                archived.append(archive_path)
            except Exception as e:
                print(f"[archive] erreur : {e}")
    return archived if archived else None

def is_month_locked(year,month):
    """Retourne True si le mois est PASSÉ (donc rapport doit être figé après 1ère génération).
    Le mois en cours reste libre de re-génération."""
    today=date.today()
    return (year,month)<(today.year,today.month)

def generate_monthly_report(snapshot,force_regenerate=False,sections=None,format_pref="auto"):
    """Génère PDF ou HTML.
    Args:
        sections: dict {clé_section: bool} indiquant quelles sections inclure.
                  Si None, toutes les sections par défaut sont incluses.
        format_pref: 'auto' (PDF prioritaire, fallback HTML — comportement historique),
                     'html' (HTML uniquement, pas de tentative PDF),
                     'pdf'  (PDF uniquement, pas de fallback HTML si échec).
    Pour un mois PASSÉ : si rapport existe déjà ET force_regenerate=False → retourne (path_existant, fmt, "existing").
                        si rapport existe ET force_regenerate=True → archive ancien, génère neuf.
    Pour le mois EN COURS : génère toujours (écrase l'ancien).
    Retourne (path, format, status) où status ∈ {"new", "existing", "regenerated"}."""
    year=snapshot["year"];month=snapshot["month"]
    locked=is_month_locked(year,month)
    out_dir=RAPPORTS_DIR/f"{year:04d}"
    mn_capit=["Janvier","F\u00e9vrier","Mars","Avril","Mai","Juin","Juillet","Ao\u00fbt","Septembre","Octobre","Novembre","D\u00e9cembre"][month-1]
    base=f"Rapport_mensuel_{year:04d}_{month:02d}_{mn_capit}"
    # Existing check : selon format_pref on cherche le bon fichier en premier
    existing_pdf=out_dir/f"{base}.pdf"
    existing_html=out_dir/f"{base}.html"
    if locked and not force_regenerate:
        if format_pref=="html":
            if existing_html.exists(): return (existing_html,"html","existing")
        elif format_pref=="pdf":
            if existing_pdf.exists(): return (existing_pdf,"pdf","existing")
        else:  # auto : PDF prioritaire (comportement historique)
            if existing_pdf.exists(): return (existing_pdf,"pdf","existing")
            if existing_html.exists(): return (existing_html,"html","existing")
    if locked and force_regenerate:
        archive_existing_report(year,month)
    # Génération selon préférence
    if format_pref=="html":
        p=generate_monthly_report_html(snapshot,sections=sections)
        if p: return (p,"html","regenerated" if locked else "new")
        return (None,None,None)
    if format_pref=="pdf":
        p=generate_monthly_report_pdf(snapshot,sections=sections)
        if p: return (p,"pdf","regenerated" if locked else "new")
        return (None,None,None)
    # format_pref=="auto" : PDF prioritaire, fallback HTML si PDF rate (comportement historique)
    p=generate_monthly_report_pdf(snapshot,sections=sections)
    if p: return (p,"pdf","regenerated" if locked else "new")
    p=generate_monthly_report_html(snapshot,sections=sections)
    if p: return (p,"html","regenerated" if locked else "new")
    return (None,None,None)
# Plancher physique cuve : en dessous de ce seuil les pompes ne distribuent plus
# Donc les ventes max d'un jour = (stock matin + livraison) - plancher
PLANCHER_PHYSIQUE = {"sp": 500, "go": 500, "gnr": 250}
TEMP_DIR=Path(tempfile.gettempdir())/"districarb_hub"
REFRESH_MS=900000
JOURS_FR=["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
# Tours de livraison TEMAG (créneaux d'arrivée du camion). Bornes validées par
# Bidou : 1er 6h→9h (alerte si non arrivé à 9h), 2e 9h→12h, 3e 12h→fermeture
# SARA. "Premier voyage" = priorité de file demandée à la commande, PAS un tour.
TOURS_LIVRAISON={
    1:{"label":"1er tour","debut":6,"fin":9,"alerte":9,"plage":"6h\u20139h"},
    2:{"label":"2e tour","debut":9,"fin":12,"alerte":12,"plage":"9h\u201312h"},
    3:{"label":"3e tour","debut":12,"fin":15,"alerte":15,"plage":"12h\u2013fermeture SARA"},
}
# Mapping des onglets de Prévision compte
SHEETS_W1={"Lundi":"Lundi","Mardi":"Mardi","Mercredi":"Mercredi","Jeudi":"Jeudi","Vendredi":"Vendredi","Samedi":"Samedi","Dimanche":"Dimanche"}
SHEETS_W2={"Lundi":"Lundi2","Mardi":"Mardi2","Mercredi":"Merc2","Jeudi":"Jeudi 2","Vendredi":"Vend 2","Samedi":"Sam 2","Dimanche":"Dim2"}

def load_json(f):
    """Lit un fichier JSON. Si f est un .cfg et qu'il n'existe pas, tente automatiquement
    de lire l'ancien .json (rétrocompatibilité après le passage .json → .cfg pour ESET).
    Retourne {} si aucune des deux versions n'existe ou si le contenu est corrompu.
    
    FIX CRITIQUE 21/05/2026 : si le fichier existe mais que le JSON est corrompu (cause :
    écriture interrompue par crash/taskkill/Windows), on BACKUP le fichier corrompu avant
    de retourner {}. Sinon le prochain save_json écrase silencieusement les données perdues
    (cas Bidou 21/05 matin : commandes du 23/05 et 26/05 perdues sans alerte). Le backup
    permet de récupérer manuellement le contenu si besoin (texte tronqué mais lisible)."""
    if f.exists():
        try: return json.loads(f.read_text(encoding="utf-8"))
        except Exception as _e:
            # Backup du fichier corrompu pour audit/récupération manuelle
            try:
                ts=datetime.now().strftime("%Y%m%d_%H%M%S")
                backup=f.with_suffix(f.suffix+f".corrupted_{ts}")
                shutil.copy2(str(f),str(backup))
                print(f"[load_json] FICHIER CORROMPU : {f.name} \u2192 backup sauvegard\u00e9 dans {backup.name}")
            except Exception as _e2: _log_silent_err(exc=_e2)
            _log_silent_err(exc=_e); return {}
    # Fallback rétrocompat : tenter l'ancien .json si on lit un .cfg
    if str(f).endswith(".cfg"):
        legacy=Path(str(f)[:-4]+".json")
        if legacy.exists():
            try: return json.loads(legacy.read_text(encoding="utf-8"))
            except Exception as _e: _log_silent_err(exc=_e); return {}
    return {}
def save_json(f,data):
    """Écriture ATOMIQUE d'un JSON. Pattern : on écrit dans un fichier temporaire à côté,
    puis on fait un rename (atomique sur tous les OS modernes). Si le programme crashe
    pendant l'écriture, soit l'ancien fichier reste intact, soit le nouveau est complet.
    JAMAIS l'état intermédiaire tronqué.
    
    FIX CRITIQUE 21/05/2026 : avant, write_text direct → corruption sur taskkill/crash
    pendant écriture. Cas constaté : commandes.cfg de Bidou réduit à 1 entrée au lieu de 3
    après une journée avec multiples interruptions hub + explorer plantés. Cause racine
    de perte de données silencieuse pour TOUS les .cfg du hub."""
    f.parent.mkdir(parents=True,exist_ok=True)
    tmp=f.with_suffix(f.suffix+".tmp")
    try:
        tmp.write_text(json.dumps(data,indent=2,ensure_ascii=False),encoding="utf-8")
        # os.replace est atomique sur tous les OS modernes (Windows depuis Vista, POSIX nativement)
        os.replace(str(tmp),str(f))
    except Exception as _e:
        # Cleanup tmp si erreur (ne pas laisser de fichier .tmp orphelin)
        try:
            if tmp.exists(): tmp.unlink()
        except Exception: pass
        raise  # propage l'erreur originale
def copy_to_temp(src):
    if not src or not os.path.exists(src): return None
    TEMP_DIR.mkdir(parents=True,exist_ok=True)
    # Nom unique avec timestamp pour éviter les fichiers fantômes
    base=os.path.splitext(os.path.basename(src))
    dest=TEMP_DIR/f"{base[0]}_{int(time.time()*1000)}{base[1]}"
    try: shutil.copy2(src,dest); return str(dest)
    except:
        try:
            with open(src,'rb') as f: data=f.read()
            with open(dest,'wb') as f: f.write(data)
            return str(dest)
        except Exception as _e: _log_silent_err(exc=_e); return None
def feur(v,d=2):
    if v is None: return "\u2014"
    try:
        f=float(v)
        if d==0: return f"{f:,.0f} \u20ac".replace(","," ")
        return f"{f:,.2f} \u20ac".replace(","," ").replace(".",",")
    except Exception as _e: _log_silent_err(exc=_e); return "\u2014"
def fnum(v,s=""):
    if v is None: return "\u2014"
    try: return f"{float(v):,.0f} {s}".replace(","," ").strip()
    except Exception as _e: _log_silent_err(exc=_e); return "\u2014"
def fmt_autonomie(jours):
    """Formate une autonomie (en jours décimaux) pour affichage humain.
    Règles validées avec Bidou :
     - >= 1 jour : 'Xj YYh' (ex: '1j 19h', '5j 4h', '12j 0h')
     - 1h à 24h  : 'Xh' ou 'Xh30' arrondi à la demi-heure (ex: '18h', '12h30', '4h', '1h30')
     - < 1h      : '< 1h' (zone urgence, précision inutile)
     - <= 0      : '0h' (rupture)
    Le code couleur reste continu basé sur la durée réelle, géré séparément ailleurs."""
    if jours is None: return "\u2014"
    try: j=float(jours)
    except Exception as _e: _log_silent_err(exc=_e); return "\u2014"
    if j<=0: return "0h"
    total_minutes=int(round(j*24*60))
    # Cas < 1h
    if total_minutes<60:
        return "< 1h"
    # Convertir en jours/heures/minutes
    days=total_minutes//(24*60)
    rem_min=total_minutes-days*24*60
    hours=rem_min//60
    minutes=rem_min-hours*60
    # Cas >= 1 jour : on affiche jours + heures pleines (pas de demi-heure à cette échelle)
    if days>=1:
        # Arrondir à l'heure pleine la plus proche
        if minutes>=30: hours+=1
        if hours>=24:
            days+=1;hours=0
        return f"{days}j {hours}h"
    # Cas 1h à 24h : arrondir à la demi-heure la plus proche
    # minutes ∈ [0, 60), on arrondit à 0 ou 30
    if minutes<15:
        h_label=f"{hours}h"
    elif minutes<45:
        h_label=f"{hours}h30"
    else:
        # >= 45 min : arrondi à hours+1
        if hours+1>=24:
            return "1j 0h"
        h_label=f"{hours+1}h"
    return h_label
def fpct(v):
    if v is None: return "\u2014"
    try: return f"{float(v)*100:.1f}%".replace(".",",")
    except Exception as _e: _log_silent_err(exc=_e); return "\u2014"
def sf(v,d=0):
    if v is None: return d
    try: return float(v)
    except Exception as _e: _log_silent_err(exc=_e); return d
def trend(cur,prev):
    if cur is None or prev is None or prev==0: return ("",C["t3"])
    try:
        c,p=float(cur),float(prev)
        if c>p*1.02: return ("\u25b2",C["green"])
        elif c<p*0.98: return ("\u25bc",C["red"])
        else: return ("\u25cf",C["t3"])
    except Exception as _e: _log_silent_err(exc=_e); return ("",C["t3"])
def jour_fr():
    return JOURS_FR[date.today().weekday()]

def parse_label_date(lbl):
    """Extrait la date d'un label LITRAGE type 'lund 13/04/26'. Retourne None si échec.
    Tolère plusieurs typos courantes :
    - format normal : '13/04/26' ou '13/04/2026'
    - séparateur manquant : '13/0426' (sera lu comme 13/04/26)
    - séparateur manquant doublé : '130426'
    - tirets ou points au lieu de slashes : '13-04-26' ou '13.04.26'
    - espaces parasites : '13 / 04 / 26'"""
    if not lbl: return None
    s=str(lbl)
    # Normaliser : remplacer séparateurs alternatifs par /, supprimer espaces autour
    s_norm=re.sub(r'\s*[-./]\s*','/',s)
    s_norm=re.sub(r'\s+','',s_norm)  # plus aucun espace
    # Tentative 1 : format normal jj/mm/aa ou jj/mm/aaaa
    m=re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})',s_norm)
    if m:
        try:
            d,mo,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
            if y<100: y+=2000
            return date(y,mo,d)
        except Exception as _e: _log_silent_err(exc=_e)
    # Tentative 2 : un slash manquant entre mois et année (ex 24/0326 → 24/03/26)
    m=re.search(r'(\d{1,2})/(\d{2})(\d{2})\b',s_norm)
    if m:
        try:
            d,mo,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
            if y<100: y+=2000
            return date(y,mo,d)
        except Exception as _e: _log_silent_err(exc=_e)
    # Tentative 3 : un slash manquant entre jour et mois (ex 2403/26 → 24/03/26)
    m=re.search(r'(?<!\d)(\d{2})(\d{2})/(\d{2,4})',s_norm)
    if m:
        try:
            d,mo,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
            if y<100: y+=2000
            return date(y,mo,d)
        except Exception as _e: _log_silent_err(exc=_e)
    # Tentative 4 : aucun slash (ex 240326 → 24/03/26)
    m=re.search(r'(?<!\d)(\d{2})(\d{2})(\d{2})(?!\d)',s_norm)
    if m:
        try:
            d,mo,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
            if y<100: y+=2000
            return date(y,mo,d)
        except Exception as _e: _log_silent_err(exc=_e)
    return None

# ===== Helpers livraisons (multi-livraisons par jour supportées) =====
def normalize_livr_day(raw):
    """Convertit l'entrée d'un jour en LISTE de livraisons (compat ancien format).
    Ancien format jour: {"sp":X,"go":Y,"gnr":Z} ou {"none":True}
    Nouveau format jour: [{"sp":X,"go":Y,"gnr":Z,"transporteur":..,"note":..,"heure":..}, ...]
    Retourne toujours une liste (vide si "none" ou rien)."""
    if not raw: return []
    if isinstance(raw,list):
        return [x for x in raw if isinstance(x,dict) and not x.get("none")]
    if isinstance(raw,dict):
        if raw.get("none"): return []
        # Ancien format unique : retour comme liste à 1 élément
        return [raw]
    return []

def aggregate_livr_day(raw):
    """Somme SP/GO/GNR de toutes les livraisons d'un jour. Retourne {sp,go,gnr} ou {} si vide."""
    livrs=normalize_livr_day(raw)
    if not livrs: return {}
    return {
        "sp":sum(float(l.get("sp",0) or 0) for l in livrs),
        "go":sum(float(l.get("go",0) or 0) for l in livrs),
        "gnr":sum(float(l.get("gnr",0) or 0) for l in livrs),
    }

# ===== Jours fériés Martinique (fixes + mobiles calculés) =====
# Extraits dans districarb_core.martinique (Étape 3 — 27/05/2026).
# Les fonctions _easter, get_feries_martinique et is_ferie sont importées en
# haut de ce fichier. Comportement strictement identique.

# nb_jours_livrables_avant : extraite dans districarb_core.martinique (Étape 3 — 27/05/2026)

def fmt_rappel_dt(target_dt):
    """Formate un datetime cible en label humain selon la proximité avec maintenant.
    
    Exemples (si maintenant = mer. 13/05 10h20) :
      - target = 13/05 12h20 → "auj. 12h20"
      - target = 14/05 10h20 → "demain 10h20"
      - target = 16/05 10h20 → "ven. 16/05 10h20"
      - target = 20/05 10h20 → "mer. 20/05 10h20"
    
    Choix : pas d'année (toujours dans la semaine ou les 7 jours suivants) pour rester court.
    """
    JC_FR=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
    today_d=date.today()
    target_d=target_dt.date()
    hm=target_dt.strftime("%Hh%M")
    if target_d==today_d:
        return f"auj. {hm}"
    if target_d==today_d+timedelta(days=1):
        return f"demain {hm}"
    return f"{JC_FR[target_d.weekday()]} {target_d.strftime('%d/%m')} {hm}"


# jour_de_commande : extraite dans districarb_core.martinique (Étape 3 — 27/05/2026)


def compute_moments_cles(date_livraison):
    """Calcule la liste des MOMENTS-CLÉS métier auxquels une alerte (marge tendue ou
    livraison à reporter) doit se réveiller pour une livraison donnée.
    
    Règle métier (validée Bidou 20/05/2026) : Bidou transmet le message au transporteur
    le JOUR DE LA COMMANDE (= avant-veille livrable de la livraison) AVANT 11H (deadline
    TEMAG). Puis il vérifie au matin de la livraison à partir de 6h que les chiffres
    sont toujours cohérents.
    
    Retourne une liste de tuples (datetime, label, contexte) triée par chronologie :
      [
        (datetime_jour_commande_6h,     "jour_commande_matin",  "C'est aujourd'hui que tu préviens le transporteur. Deadline TEMAG : 11h."),
        (datetime_jour_commande_10h30,  "jour_commande_deadline","⚠ Deadline TEMAG dans 30 min. Transporteur prévenu ?"),
        (datetime_jour_livraison_6h,    "matin_livraison",      "Matin de la livraison. Vérifie que les chiffres sont toujours cohérents."),
      ]
    
    Args:
      date_livraison: date (pas datetime) de la livraison concernée
    Returns:
      Liste de 3 tuples (datetime, code, message). Si jour_de_commande est None
      (cas extrême), fallback sur la veille calendaire de la livraison.
    """
    if date_livraison is None: return []
    jc=jour_de_commande(date_livraison)
    if jc is None:
        # Fallback : veille calendaire (cas où aucun jour livrable trouvé dans les 14j)
        jc=date_livraison-timedelta(days=1)
    return [
        (datetime.combine(jc,dt_time(6,0)),
         "jour_commande_matin",
         "C'est aujourd'hui que tu pr\u00e9viens le transporteur. Deadline TEMAG : 11h."),
        (datetime.combine(jc,dt_time(10,30)),
         "jour_commande_deadline",
         "\u26a0 Deadline TEMAG dans 30 min. Transporteur pr\u00e9venu ?"),
        (datetime.combine(date_livraison,dt_time(6,0)),
         "matin_livraison",
         "Matin de la livraison. V\u00e9rifie que les chiffres sont toujours coh\u00e9rents."),
    ]


def prochain_moment_cle(date_livraison,now=None):
    """Retourne le PROCHAIN moment-clé STRICTEMENT futur pour cette livraison,
    ou None si tous les moments-clés sont passés.
    
    Sert au calcul du silence intelligent dans _cest_note / _cest_fait : "silence
    jusqu'au prochain moment-clé" plutôt qu'un délai arbitraire.
    
    Cap à 5h : si on est entre 5h et 6h du jour J, le moment-clé 6h est considéré
    comme "imminent" et est retourné (pour que la popup s'affiche dès que Bidou
    ouvre le hub à 5h45 par exemple).
    """
    if now is None: now=datetime.now()
    moments=compute_moments_cles(date_livraison)
    for dt_moment,code,msg in moments:
        # Considère comme "futur" si dt_moment > maintenant - 1h
        # (laisse une marge de 1h pour matcher le cas "Bidou ouvre à 5h45, le moment-clé 6h s'affiche")
        if dt_moment > now-timedelta(hours=1):
            return (dt_moment,code,msg)
    return None


def prochain_moment_cle_strict(date_livraison,now=None):
    """Variante STRICTE de prochain_moment_cle, sans la marge -1h.
    Retourne le 1er moment-clé STRICTEMENT > now, ou None si aucun.
    
    Bug fix 21/05/2026 : la marge -1h de prochain_moment_cle est utile pour la
    DÉTECTION du mode courant (afficher la popup à 5h45 dès l'ouverture du hub),
    mais elle est NÉFASTE pour le calcul du silence (silencer jusqu'à 10h30 alors
    qu'on est à 10h54 = silence immédiatement expiré → popup ré-ouverte au refresh
    suivant → harcèlement). Pour le silence, on veut un moment STRICTEMENT futur.
    """
    if now is None: now=datetime.now()
    moments=compute_moments_cles(date_livraison)
    for dt_moment,code,msg in moments:
        if dt_moment>now:
            return (dt_moment,code,msg)
    return None


def detecter_moment_courant(date_livraison,now=None):
    """Identifie DANS QUEL "mode" se trouve la popup au moment de son ouverture,
    pour adapter les boutons affichés.
    
    Retourne un code :
      - "anticipation"     : on est AVANT le jour de la commande (Bidou ne peut encore rien faire)
      - "jour_commande_avant_10h30" : jour J avant 10h30 (Bidou peut agir tranquillement)
      - "jour_commande_deadline"    : jour J entre 10h30 et 11h (URGENT, deadline imminente)
      - "jour_commande_apres_11h"   : jour J après 11h (deadline passée, action déjà manquée)
      - "entre_commande_et_livraison": entre J (commande) et la livraison (= surveillance passive)
      - "matin_livraison"  : jour de la livraison à partir de 5h (vérification finale avant camion)
      - "livraison_passee" : livraison passée (cas exceptionnel, alerte obsolète)
    """
    if now is None: now=datetime.now()
    today=now.date()
    jc=jour_de_commande(date_livraison)
    if jc is None: jc=date_livraison-timedelta(days=1)
    if date_livraison<today:
        return "livraison_passee"
    if today==date_livraison and now.hour>=5:
        return "matin_livraison"
    if today<jc:
        return "anticipation"
    if today==jc:
        if now.hour<10 or (now.hour==10 and now.minute<30):
            return "jour_commande_avant_10h30"
        elif now.hour<11:
            return "jour_commande_deadline"
        else:
            return "jour_commande_apres_11h"
    # today > jc et today < date_livraison → on attend le matin de la livraison
    return "entre_commande_et_livraison"


def make_snooze_options(durations_h):
    """Construit les options du dropdown 'Rappel dans...' avec des LABELS EN DATE ABSOLUE.
    
    Au lieu de "2 heures" / "1 jour" / "1 semaine" (relatif au moment du clic, devient illisible
    une fois consulté plus tard dans le journal), on génère des labels comme "auj. 12h20",
    "demain 10h20", "ven. 16/05 10h20".
    
    Args:
      durations_h: liste d'heures, ex. [1,2,4,6,24,72,168]
    Returns:
      (labels_list, label_to_hours): labels pour le dropdown + dict inverse pour récupérer h.
    """
    now=datetime.now()
    labels=[];mapping={}
    for h in durations_h:
        target=now+timedelta(hours=h)
        lbl=fmt_rappel_dt(target)
        labels.append(lbl)
        mapping[lbl]=h
    return labels,mapping

def get_jour_pivot(partial_today=None,now=None):
    """Détermine le 'jour pivot' = premier jour pour lequel la commande SARA est encore passable.

    Règles métier (validées par Bidou) :
    - Bascule J -> J+1 dès l'une de ces conditions :
       (a) heure courante >= 11h00 (deadline commande SARA pour livraison J+1)
       (b) C1 du jour J est saisie (signe que la journée est lancée, commande déjà passée ou pas)
    - Si J+1 tombe sur un jour non-livrable (weekend, férié), on glisse au prochain jour ouvré.

    Retourne la date à partir de laquelle le hub doit projeter.
    """
    now=now or datetime.now()
    today=now.date()
    # Condition (a) : 11h passées
    after_11h=now.hour>=11
    # Condition (b) : C1 du jour calendaire today est saisie
    # partial_today["nb_caisses"] = nb de caisses saisies sur le DERNIER jour comptable connu.
    # Si partial.label correspond à today ET nb_caisses >= 1 → C1 saisie.
    c1_saisie=False
    if partial_today and partial_today.get("nb_caisses",0)>=1:
        # Vérifier que le partial concerne bien aujourd'hui (sinon c'est la veille en cours de fermeture)
        # On n'a pas la date du partial directement ici, donc on l'infère : si nb_caisses=1,2 c'est probable today.
        # Pour être conservateur on considère : nb_caisses >= 1 signifie qu'au moins une caisse du jour
        # courant est lancée → bascule.
        c1_saisie=True
    bascule=after_11h or c1_saisie
    if not bascule: return today
    # Bascule : prochaine livraison utile = J+1 ouvré livrable
    # FIX 22/05/2026 (bug "Pas de tension détectée" popup Commande du jour) :
    # On respecte les forçages — un jour weekend/férié explicitement forcé via
    # FORCAGE_FILE (résolution alerte Prévision via AntiRuptureDlg) EST un jour
    # livrable du point de vue du hub. Sans ce respect, le jour_pivot sautait
    # par-dessus les samedis/dimanches/fériés forcés (cas Bidou ven 22/05 après
    # 11h : pivot = mar 26/05 au lieu de sam 23/05), et analyze_antirupture
    # filtrait alors toutes les analyses (marges tendues, livraisons à reporter,
    # saisies impossibles) avec `fc_d < jour_pivot` → samedi 23/05 disparaissait
    # complètement du moteur, alors que c'était la prochaine livraison réelle.
    # Aligné avec la même condition dans CommandeDialog._ctx() ligne 8284.
    pivot=today+timedelta(days=1)
    while (pivot.weekday()>=5 or is_ferie(pivot)) and not is_date_forcee(pivot):
        pivot+=timedelta(days=1)
    return pivot

def is_periode_observatoire(d):
    """Effet Observatoire des Prix Martinique : entre le 25 du mois et le 5 du mois suivant,
    les ventes sont anormalement variables (anticipation hausse/baisse de prix au 1er).
    Sur cette période on assouplit les seuils d'alerte 'ventes irréalistes'."""
    return d.day>=25 or d.day<=5

# _nom_ferie : extraite dans districarb_core.martinique (Étape 3 — 27/05/2026)
# Importée en haut du fichier via `from districarb_core.martinique import nom_ferie as _nom_ferie`

def jours_rupture_par_carburant(depuis,jusqu_a):
    """B1 — Lit le journal (evenements.cfg) et retourne l'ensemble des couples
    (date, CARBURANT) couverts par une rupture, sur la période [depuis, jusqu_a].

    Pourquoi : le hub journalise déjà les ruptures (add_evenement 'rupture' avec
    carburant + jour + jour_fin). _calc_autonomie doit EXCLURE ces jours du calcul
    de consommation (vente bridée par le stock = ne reflète pas la demande réelle).
    C'est l'or déjà présent dans la maison, enfin relu.

    Subtilité (vérifiée dans le code) : load_evenements_period filtre sur le ts de
    l'événement (= jour de DÉBUT de la rupture), pas sur jour_fin. Une rupture qui
    commence avant 'depuis' mais déborde dedans serait ratée si on chargeait pile
    la fenêtre. On élargit donc le chargement de 30 j en amont, puis on déplie
    chaque plage jour→jour_fin nous-mêmes et on ne garde que ce qui tombe dans
    [depuis, jusqu_a].

    Retour : set de tuples (datetime.date, "SP"|"GO"|"GNR").
    """
    out=set()
    try:
        evts=load_evenements_period(depuis-timedelta(days=30),jusqu_a)
        for evt in evts:
            if evt.get("type")!="rupture": continue
            data=evt.get("data",{}) or {}
            carb=str(data.get("carburant","")).upper().strip()
            if carb not in ("SP","GO","GNR"): continue
            j0=data.get("jour","");j1=data.get("jour_fin",j0)
            try:
                d0=datetime.fromisoformat(j0).date()
                d1=datetime.fromisoformat(j1).date() if j1 else d0
            except Exception as _e:
                _log_silent_err(exc=_e);continue
            if d1<d0: d0,d1=d1,d0
            d=d0
            while d<=d1:
                if depuis<=d<=jusqu_a: out.add((d,carb))
                d+=timedelta(days=1)
    except Exception as _e:
        _log_silent_err(exc=_e)
    return out

def conso_garde_fou_par_jour_semaine(hist_data,nb_derniers=8):
    """GARDE-FOU autonomie : consommation de référence par jour de semaine,
    calculée sur des jours SAINS uniquement.

    PRINCIPE (Point B — la mémoire enfin relue) :
    Le hub journalise déjà les ruptures dans evenements.cfg (carburant + dates).
    Un jour de rupture = vente bridée par le stock vide, PAS par la demande réelle.
    On EXCLUT donc ces jours du calcul, factuellement, via le journal — par carburant
    (un jour de rupture SP est retiré du calcul SP seulement ; GO/GNR ce jour-là,
    s'ils étaient normaux, restent comptés).

    Une fois les jours pourris retirés FACTUELLEMENT (plus par devinette statistique),
    on revient à une simple MOYENNE sur les jours sains — transparente, auditable,
    explicable. Plus de P75 : il n'était qu'un contournement pour absorber les jours
    à 0 qu'on ne savait pas identifier. Maintenant on les identifie par la mémoire.

    On conserve l'exclusion des fériés (régime de conso atypique, déjà géré ailleurs).

    Retourne {0:{"sp":X,"go":Y,"gnr":Z}, ...} (lun=0..dim=6)."""
    complete=[h for h in hist_data if not h.get("en_cours")]
    # Charger les jours de rupture connus du journal, par carburant, sur l'amplitude
    # réelle de l'historique exploité (large : on prend tout l'historique complet).
    dates_hist=[parse_label_date(h.get("label","")) for h in complete]
    dates_hist=[d for d in dates_hist if d]
    rupt=set()
    if dates_hist:
        rupt=jours_rupture_par_carburant(min(dates_hist),max(dates_hist))
    def _moy(vals):
        return (sum(vals)/len(vals)) if vals else 0.0
    result={}
    for wd in range(7):
        # Jours candidats : même jour de semaine, complet, hors férié
        cand=[h for h in complete
              if parse_label_date(h.get("label","")) and
              parse_label_date(h.get("label","")).weekday()==wd and
              not is_ferie(parse_label_date(h.get("label","")))]
        # Pour CHAQUE carburant, on filtre indépendamment les jours en rupture
        # de CE carburant, puis on prend les N derniers sains et on moyenne.
        res_wd={}
        for carb in ("sp","go","gnr"):
            CB=carb.upper()
            sains=[h for h in cand
                   if (parse_label_date(h.get("label","")),CB) not in rupt]
            sains=sains[-nb_derniers:] if len(sains)>=nb_derniers else sains
            if sains:
                res_wd[carb]=_moy([sf(h.get(carb,0)) for h in sains])
            else:
                # Fallback : moyenne sur les 30 derniers jours sains (ce carburant)
                last_sains=[h for h in complete
                            if (parse_label_date(h.get("label","")),CB) not in rupt]
                last_sains=last_sains[-30:] if len(last_sains)>=30 else last_sains
                if last_sains:
                    res_wd[carb]=_moy([sf(h.get(carb,0)) for h in last_sains])
                else:
                    res_wd[carb]={"sp":9000,"go":7000,"gnr":300}[carb]
        result[wd]=res_wd
    return result

def avg_ventes_par_jour_semaine(hist_data,nb_derniers=4):
    """Calcul la moyenne des ventes SP/GO/GNR par jour de semaine (lun=0..dim=6),
    en utilisant les N derniers mêmes jours de semaine COMPLETS et HORS fériés.
    Retourne dict {0:{"sp":X,"go":Y,"gnr":Z}, 1:{...}, ...}"""
    complete=[h for h in hist_data if not h.get("en_cours")]
    result={}
    for wd in range(7):
        jours=[h for h in complete if parse_label_date(h.get("label","")) and
               parse_label_date(h.get("label","")).weekday()==wd and
               not is_ferie(parse_label_date(h.get("label","")))]
        # Prendre les N derniers
        jours=jours[-nb_derniers:] if len(jours)>=nb_derniers else jours
        if jours:
            result[wd]={
                "sp":sum(sf(h.get("sp",0)) for h in jours)/len(jours),
                "go":sum(sf(h.get("go",0)) for h in jours)/len(jours),
                "gnr":sum(sf(h.get("gnr",0)) for h in jours)/len(jours),
            }
        else:
            # Fallback : moyenne globale des 30 derniers jours
            last30=complete[-30:] if len(complete)>=30 else complete
            if last30:
                result[wd]={
                    "sp":sum(sf(h.get("sp",0)) for h in last30)/len(last30),
                    "go":sum(sf(h.get("go",0)) for h in last30)/len(last30),
                    "gnr":sum(sf(h.get("gnr",0)) for h in last30)/len(last30),
                }
            else:
                result[wd]={"sp":9000,"go":7000,"gnr":300}  # fallback dur si aucune donnée
    return result

def avg_ventes_n_caisses(hist_data,wd,n_caisses,nb_derniers=4):
    """Moyenne des ventes SP/GO/GNR cumulées sur les N PREMIÈRES caisses,
    pour les derniers mêmes jours de semaine COMPLETS et hors fériés.
    Sert à comparer partial_today (ventes sur n_caisses saisies) à une base comparable,
    au lieu d'une proportion horaire bancale.
    Retourne {"sp":X,"go":Y,"gnr":Z} ou None si pas assez de données."""
    if not hist_data or n_caisses<1: return None
    complete=[h for h in hist_data if not h.get("en_cours")]
    jours=[h for h in complete if parse_label_date(h.get("label","")) and
           parse_label_date(h.get("label","")).weekday()==wd and
           not is_ferie(parse_label_date(h.get("label","")))]
    jours=jours[-nb_derniers:] if len(jours)>=nb_derniers else jours
    if not jours: return None
    sums=[]
    for h in jours:
        caisses=h.get("caisses",{}) or {}
        s={"sp":0,"go":0,"gnr":0}
        for i in range(1,n_caisses+1):
            c=caisses.get(str(i),{})
            s["sp"]+=sf(c.get("sp",0));s["go"]+=sf(c.get("go",0));s["gnr"]+=sf(c.get("gnr",0))
        sums.append(s)
    n=len(sums)
    return {"sp":sum(s["sp"] for s in sums)/n,
            "go":sum(s["go"] for s in sums)/n,
            "gnr":sum(s["gnr"] for s in sums)/n}

def get_current_partial(hist_data,max_age_days=3):
    """Retourne le jour en cours RÉCENT dans LITRAGE (filtre orphelins historiques).
    Règle : dernier jour en_cours dont la date >= aujourd'hui - max_age_days.
    Corrige le bug minuit : lundi 13/04 reste visible mardi 7h si C3 pas saisie.
    """
    if not hist_data: return None
    threshold=date.today()-timedelta(days=max_age_days)
    recents=[]
    for h in hist_data:
        if not h.get("en_cours"): continue
        d=parse_label_date(h.get("label",""))
        if d and d>=threshold:
            recents.append(h)
    return recents[-1] if recents else None

# ===== Générateur de bulles Pillow (raccourcis) =====
_BUBBLE_CACHE={}

def make_bubble(color_hex,icon_char,size=56,state="normal"):
    """Génère une bulle PNG avec dégradé bombé + reflet + ombre portée.
    state: 'normal', 'hover' (éclaircie), 'press' (assombrie + descendue)."""
    if not PIL_OK: return None
    key=(color_hex,icon_char,size,state)
    if key in _BUBBLE_CACHE: return _BUBBLE_CACHE[key]
    scale=3
    s=size*scale
    img=Image.new("RGBA",(s,s),(0,0,0,0))
    h=color_hex.lstrip('#')
    base_rgb=tuple(int(h[i:i+2],16) for i in (0,2,4))
    if state=="hover":
        base_rgb=tuple(min(255,int(c+(255-c)*0.15)) for c in base_rgb)
    elif state=="press":
        base_rgb=tuple(max(0,int(c*0.8)) for c in base_rgb)
    def lighten(rgb,factor):
        return tuple(min(255,int(c+(255-c)*factor)) for c in rgb)
    def darken(rgb,factor):
        return tuple(max(0,int(c*(1-factor))) for c in rgb)
    # Ombre portée (réduite si press = bulle enfoncée)
    shadow=Image.new("RGBA",(s,s),(0,0,0,0))
    sd=ImageDraw.Draw(shadow)
    shadow_offset_y=int(s*0.04 if state=="press" else s*0.08)
    bubble_size=int(s*0.82)
    margin=(s-bubble_size)//2
    sd.ellipse([margin,margin+shadow_offset_y,margin+bubble_size,margin+bubble_size+shadow_offset_y],fill=(0,0,0,120 if state=="press" else 140))
    shadow=shadow.filter(ImageFilter.GaussianBlur(radius=int(s*0.025)))
    img=Image.alpha_composite(img,shadow)
    # Masque circulaire
    mask=Image.new("L",(s,s),0)
    md=ImageDraw.Draw(mask)
    md.ellipse([margin,margin,margin+bubble_size,margin+bubble_size],fill=255)
    # Dégradé bombé
    grad=Image.new("RGB",(s,s),base_rgb)
    gd=ImageDraw.Draw(grad)
    for y in range(s):
        t=(y-margin)/bubble_size if bubble_size>0 else 0
        t=max(0,min(1,t))
        if t<0.45:
            f=t/0.45
            top=lighten(base_rgb,0.35)
            col=tuple(int(top[i]+(base_rgb[i]-top[i])*f) for i in range(3))
        else:
            f=(t-0.45)/0.55
            bot=darken(base_rgb,0.35)
            col=tuple(int(base_rgb[i]+(bot[i]-base_rgb[i])*f) for i in range(3))
        gd.line([(0,y),(s,y)],fill=col)
    bubble=Image.new("RGBA",(s,s),(0,0,0,0))
    bubble.paste(grad,(0,0),mask)
    img=Image.alpha_composite(img,bubble)
    # Reflet interne
    highlight=Image.new("RGBA",(s,s),(0,0,0,0))
    hd=ImageDraw.Draw(highlight)
    hl_w=int(bubble_size*0.6)
    hl_h=int(bubble_size*0.35)
    hl_x=margin+(bubble_size-hl_w)//2
    hl_y=margin+int(bubble_size*0.08)
    hd.ellipse([hl_x,hl_y,hl_x+hl_w,hl_y+hl_h],fill=(255,255,255,90))
    highlight=highlight.filter(ImageFilter.GaussianBlur(radius=int(s*0.02)))
    hl_masked=Image.new("RGBA",(s,s),(0,0,0,0))
    hl_masked.paste(highlight,(0,0),mask)
    img=Image.alpha_composite(img,hl_masked)
    # Icône centrale
    font=None
    for font_path in [r"C:\Windows\Fonts\seguisym.ttf",r"C:\Windows\Fonts\segoeuisymbol.ttf",
                       r"C:\Windows\Fonts\arial.ttf","/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]:
        try:
            font=ImageFont.truetype(font_path,int(s*0.42));break
        except Exception as _e: _log_silent_err(exc=_e); continue
    if font is None: font=ImageFont.load_default()
    txt_img=Image.new("RGBA",(s,s),(0,0,0,0))
    td=ImageDraw.Draw(txt_img)
    try:
        bbox=td.textbbox((0,0),icon_char,font=font)
        tw=bbox[2]-bbox[0];th=bbox[3]-bbox[1]
        tx=(s-tw)//2-bbox[0];ty=(s-th)//2-bbox[1]
    except Exception:
        tx=ty=s//3
    td.text((tx,ty),icon_char,font=font,fill=(255,255,255,255))
    img=Image.alpha_composite(img,txt_img)
    result=img.resize((size,size),Image.LANCZOS)
    _BUBBLE_CACHE[key]=result
    return result

def get_cycle_week():
    """Retourne 1 ou 2 selon la semaine du cycle de 14j actuel."""
    cycle=load_json(CYCLE_FILE)
    anchor=cycle.get("anchor_date")
    anchor_week=cycle.get("anchor_week",1)
    if not anchor: return None
    try:
        anchor_d=datetime.strptime(anchor,"%Y-%m-%d").date()
        # Trouver le lundi de la semaine de l'ancre
        anchor_monday=anchor_d-timedelta(days=anchor_d.weekday())
        # Trouver le lundi de la semaine actuelle
        today_monday=date.today()-timedelta(days=date.today().weekday())
        weeks_diff=(today_monday-anchor_monday).days//7
        # Si l'ancre était semaine 1, et qu'on a fait N semaines : 
        # weeks_diff pair → même semaine que ancre, impair → autre semaine
        if anchor_week==1:
            return 1 if weeks_diff%2==0 else 2
        else:
            return 2 if weeks_diff%2==0 else 1
    except Exception as _e: _log_silent_err(exc=_e); return None

def set_cycle_week(week):
    """Sauvegarde l'ancrage du cycle."""
    today=date.today()
    monday=today-timedelta(days=today.weekday())
    save_json(CYCLE_FILE,{"anchor_date":monday.strftime("%Y-%m-%d"),"anchor_week":week,"set_at":datetime.now().isoformat()})

def get_sheet_for_day(day_offset):
    """Retourne le nom de l'onglet Prévision pour aujourd'hui+offset.
    Le cycle Pre_vision tourne sur 14 jours = 2 semaines en alternance.
    Donc le mapping date→onglet utilise une rotation modulo 2 :
     - week_offset pair (0, 2, 4...) → même semaine du cycle qu'aujourd'hui
     - week_offset impair (1, 3, 5...) → semaine opposée du cycle"""
    target=date.today()+timedelta(days=day_offset)
    target_jour=JOURS_FR[target.weekday()]
    week=get_cycle_week()
    if not week: return None
    today_monday=date.today()-timedelta(days=date.today().weekday())
    target_monday=target-timedelta(days=target.weekday())
    week_offset=(target_monday-today_monday).days//7
    # Rotation modulo 2 (cycle 14 jours = 2 semaines en alternance)
    if week_offset%2==0:
        target_week=week  # même semaine du cycle
    else:
        target_week=3-week  # semaine opposée (1↔2)
    if target_week==1: return SHEETS_W1.get(target_jour)
    else: return SHEETS_W2.get(target_jour)

def get_pre_vision_horizon():
    """Retourne la dernière date couverte par le fichier Pre_vision actuel.

    Le fichier couvre toujours 14 jours consécutifs Lundi→Dim2 :
     - Si aujourd'hui = Sem 1 du cycle, le fichier a démarré au lundi de cette semaine,
       et se termine au dimanche de la semaine suivante (Dim2).
     - Si aujourd'hui = Sem 2 du cycle, le fichier a démarré au lundi de la semaine précédente
       (Sem 1), et se termine au dimanche de cette semaine (Dim2).

    Au-delà de cet horizon, les onglets ne représentent plus les bonnes dates : il faut donc
    ignorer col E (livraisons ponctuelles) tout en gardant col D (ventes prévues, pattern hebdo).

    Returns:
        date: dernier jour couvert par le fichier (un dimanche), ou None si cycle non configuré.
    """
    week=get_cycle_week()
    if not week: return None
    today=date.today()
    today_monday=today-timedelta(days=today.weekday())
    if week==1:
        # Sem 1 : fichier a démarré au lundi de cette semaine, se termine au dimanche +13
        file_start_monday=today_monday
    else:
        # Sem 2 : fichier a démarré au lundi de la semaine précédente, se termine au dimanche +6
        file_start_monday=today_monday-timedelta(days=7)
    return file_start_monday+timedelta(days=13)  # dernier jour = dimanche


# =============================================================================
class DataReader:
    def __init__(self,config): self.config=config
    def _open_detailed(self,key):
        """Comme _open mais distingue les causes d'échec. Retourne (wb, error_code) où error_code ∈
        None (OK) / 'notfound' (chemin manquant ou fichier absent) / 'locked' (verrou Excel
        ou copie impossible) / 'read' (lecture openpyxl plantée).
        Utilisé par les flux qui ont besoin d'afficher une bannière utilisateur explicite
        (ex: lecture Prévision dans read_all -> bannière en haut du hub si verrouillé)."""
        src=self.config.get(key)
        if not src: return (None,"notfound")
        if not os.path.exists(src): return (None,"notfound")
        tmp=copy_to_temp(src)
        if not tmp: return (None,"locked")
        try: return (openpyxl.load_workbook(tmp,data_only=True,read_only=True),None)
        except PermissionError: return (None,"locked")
        except Exception as _e: _log_silent_err(exc=_e); return (None,"read")
    def _open(self,key):
        wb,_=self._open_detailed(key)
        return wb
    def _open_styled(self,key):
        src=self.config.get(key)
        if not src: return None
        tmp=copy_to_temp(src)
        if not tmp: return None
        try: return openpyxl.load_workbook(tmp,data_only=True,read_only=False)
        except Exception as _e: _log_silent_err(exc=_e); return None
    def _c(self,ws,coord):
        try: return ws[coord].value
        except Exception as _e: _log_silent_err(exc=_e); return None
    def read_all(self):
        # Lire l'historique d'abord pour extraire les ventes partielles du jour en cours LITRAGE
        hist_data=self._read_hist()
        # POINT A — Écrire la mémoire AVANT les calculs (même cycle de refresh).
        # detect_recent_events journalise ruptures/anomalies dans evenements.cfg.
        # Il DOIT tourner ici (avant analyze_antirupture et _calc_autonomie) pour que
        # ces calculs puissent relire un journal à jour DANS LE MÊME refresh — sinon
        # ils liraient toujours l'état d'avant (retard d'un cycle de 15 min).
        # Dédoublonnage natif par fingerprint dans add_evenement → safe à chaque refresh.
        try: detect_recent_events(hist_data)
        except Exception as _e: _log_silent_err(exc=_e)
        _pc=get_current_partial(hist_data)
        partial_today={"sp":_pc.get("sp",0),"go":_pc.get("go",0),"gnr":_pc.get("gnr",0),"nb_caisses":int(_pc.get("nb_caisses",0))} if _pc else None
        # Charger la livraison correspondant au MÊME jour que le partial (pas le jour calendaire)
        # Important quand la C3 de la veille n'est pas encore saisie : les stocks temps réel
        # correspondent à la veille, donc la livraison à ajouter est celle de la veille aussi.
        if _pc:
            pc_date=parse_label_date(_pc.get("label",""))
            livr_day=pc_date.strftime("%d/%m/%y") if pc_date else date.today().strftime("%d/%m/%y")
        else:
            livr_day=date.today().strftime("%d/%m/%y")
        livr=load_json(LIVRAISON_FILE).get(livr_day,{})
        pv_data=self._read_pv(livr,partial_today,hist=hist_data)
        # Mémoriser les prix courants pour calcul effet spéculation au passage de mois
        try:
            if pv_data.get("st")=="ok":
                update_prix_courant(
                    pv_data.get("pa_sp",0),pv_data.get("pa_go",0),pv_data.get("pa_gnr",0),
                    pv_data.get("pv_sp",0),pv_data.get("pv_go",0),pv_data.get("pv_gnr",0),
                    pv_data.get("marge_unit",0),
                )
        except Exception as _e: _log_silent_err(exc=_e)
        proj14=self.projection_14j(pv_data,hist_data,partial_today)
        antirupture=self.analyze_antirupture(pv_data,proj14,hist_data)
        return {"gp":self._read_gp(),"ca":self._read_ca(),"pv":pv_data,
                "ob":self._read_ob(),"li":self._read_li(),"hist":hist_data,
                "alerts":self._read_alerts(),"auto":self._calc_autonomie(pv_data,hist_data),
                "partial":partial_today,"livr":livr,"proj14":proj14,
                "antirupture":antirupture,"ts":datetime.now().strftime("%H:%M:%S")}

    def _read_gp(self):
        wb=self._open("gest_piste")
        if not wb: return {"st":"miss"}
        o={"st":"ok"}
        try:
            ws=wb["Piste"]
            o["caisse"]=self._c(ws,"D3");o["date"]=self._c(ws,"H3");o["jour"]=self._c(ws,"H2")
            o["l_gnr"]=self._c(ws,"D14") or 0;o["l_sp"]=self._c(ws,"H14") or 0;o["l_go"]=self._c(ws,"L14") or 0
            o["e_gnr"]=self._c(ws,"F20") or 0;o["e_sp"]=self._c(ws,"F21") or 0;o["e_go"]=self._c(ws,"F22") or 0
            o["total_piste"]=self._c(ws,"F24") or 0
            o["cb"]=self._c(ws,"O23") or 0;o["cp"]=self._c(ws,"O24") or 0
            o["caisse_total"]=self._c(ws,"O28") or 0;o["theorique"]=self._c(ws,"O29") or 0;o["ecart"]=self._c(ws,"O30") or 0
            ws2=wb["Boutique"]
            o["b_esp"]=self._c(ws2,"D23") or 0;o["b_cb"]=self._c(ws2,"D26") or 0
            o["b_cp"]=self._c(ws2,"D27") or 0;o["b_chq"]=self._c(ws2,"D25") or 0
            o["b_total"]=self._c(ws2,"D29") or 0;o["b_ecart"]=self._c(ws2,"D32") or 0
        except Exception as e: print(f"[ERR gp] {e}");o["st"]="err"
        finally:
            try:wb.close()
            except Exception as _e: _log_silent_err(exc=_e)
        return o

    def _read_ca(self):
        wb2=self._open_styled("cartes")
        if not wb2: return {"st":"miss","critical":[],"rose":0,"jaune":0,"days":[]}
        o={"st":"ok","rose":0,"jaune":0,"critical":[],"days":[]}
        try:
            import re
            ws=wb2["Feuil1"]
            MONTHS={"janvier":1,"f\u00e9vrier":2,"fevrier":2,"mars":3,"avril":4,"mai":5,"juin":6,
                    "juillet":7,"ao\u00fbt":8,"aout":8,"septembre":9,"octobre":10,"novembre":11,"d\u00e9cembre":12,"decembre":12}
            today=date.today()
            def parse_title(s):
                if not s: return None
                m=re.match(r'(lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)\s+(\d{1,2})\s+(\w+)\s+(\d{4})',str(s).lower().strip())
                if not m: return None
                _,d_,mo,y=m.groups()
                if mo not in MONTHS: return None
                try: return date(int(y),MONTHS[mo],int(d_))
                except Exception as _e: _log_silent_err(exc=_e); return None
            def is_yellow(cell):
                try:
                    if not cell.fill or not cell.fill.fgColor: return False
                    rgb=cell.fill.fgColor.rgb
                    if not isinstance(rgb,str) or len(rgb)<8 or rgb=="00000000": return False
                    r,g,b=int(rgb[2:4],16),int(rgb[4:6],16),int(rgb[6:8],16)
                    return r>=200 and g>=200 and b<=150
                except Exception as _e: _log_silent_err(exc=_e); return False
            # Étape 1 : repérer titres jour 2026 en col A, rouge FFFF0000
            start=max(1,ws.max_row-4500)
            titres=[]
            for r in range(start,ws.max_row+1):
                cell=ws.cell(r,1)
                a=cell.value
                if not a: continue
                try:
                    col=cell.font.color.rgb if cell.font and cell.font.color else None
                except: col=None
                if col!="FFFF0000": continue
                d=parse_title(a)
                if d and d.year==2026:
                    titres.append((r,d,str(a)))
            # Étape 2 : pour chaque titre, lire col D entre titre+1 et "credit agricole" inclus
            for i,(rt,dt,lbl) in enumerate(titres):
                end=None
                for r in range(rt+1,min(rt+35,ws.max_row+1)):
                    av=ws.cell(r,1).value
                    if av and 'credit' in str(av).lower() and 'agricole' in str(av).lower():
                        end=r;break
                if not end: continue
                if end-rt-1>26: continue  # fail-safe parsing
                tcs=[]
                for r in range(rt+1,end+1):
                    dv=ws.cell(r,4).value
                    # FIX 21/05/2026 : cellule vide → skip SILENCIEUX (pas de log).
                    # Avant : float(None) levait TypeError, attrapé par _log_silent_err,
                    # qui loggait à CHAQUE cellule vide × 26 lignes × N chèques × refresh 15 min.
                    # → des milliers de lignes inutiles dans errors.log par jour (cause du
                    # fichier de 99 Mo). Maintenant on filtre avant la conversion.
                    if dv is None: continue
                    try: v=float(dv)
                    except (TypeError,ValueError): continue
                    if not (100<=v<=7000): continue
                    y=is_yellow(ws.cell(r,4))
                    tcs.append({"row":r,"montant":v,"jaune":y})
                    if y: o["jaune"]+=1
                    else: o["rose"]+=1
                if tcs:
                    o["days"].append({"date":dt,"label":lbl,"tcs":tcs,
                                      "total":sum(t["montant"] for t in tcs),
                                      "pointees":sum(1 for t in tcs if t["jaune"]),
                                      "en_attente":sum(1 for t in tcs if not t["jaune"])})
            # Étape 3 : alertes retard >3j sur non pointées >100€
            for day in o["days"]:
                age=(today-day["date"]).days
                if age<=3: continue
                for t in day["tcs"]:
                    if t["jaune"]: continue
                    if t["montant"]<=100: continue
                    o["critical"].append({"date":day["date"],"montant":t["montant"],"age":age})
        except Exception as e: print(f"[ERR ca] {e}");import traceback;traceback.print_exc();o["st"]="err"
        finally:
            try:wb2.close()
            except Exception as _e: _log_silent_err(exc=_e)
        return o

    def _read_pv(self,livr,partial_today=None,hist=None):
        # Lecture du classeur Prévision avec distinction des causes d'erreur, pour permettre
        # une bannière utilisateur explicite dans le hub. Retours possibles :
        #   {"st":"ok", ...}                            : lecture réussie
        #   {"st":"miss", "_error":"notfound"|"locked"|"read"} : pas de données utiles, cause connue
        wb,open_err=self._open_detailed("prevision")
        if not wb: return {"st":"miss","_error":open_err or "miss"}
        o={"st":"ok"}
        try:
            ws=wb["Bilan Mati\u00e8re"]
            # Lire le bilan (B16-B18 = stocks matin agrégés, C16-C18 = conso veille)
            sp_b=sf(self._c(ws,"B16"));go_b=sf(self._c(ws,"B17"));gnr_b=sf(self._c(ws,"B18"))
            o["sp_c"]=sf(self._c(ws,"C16"));o["go_c"]=sf(self._c(ws,"C17"));o["gnr_c"]=sf(self._c(ws,"C18"))
            o["valo_matin"]=sf(self._c(ws,"B20"))
            # Prix achat unitaires : lus depuis l'onglet Lundi (stables d'un onglet à l'autre)
            pa_sp=pa_go=pa_gnr=0
            # Prix de vente + marge : également depuis Lundi.
            #  - C21/C22/C23 = prix de vente SP/GO/GNR €/L (modifiés le 1er de chaque mois)
            #  - C25         = marge unitaire €/L (moyenne ; valeur de référence du mois)
            pv_sp=pv_go=pv_gnr=0;marge_unit=0
            try:
                ws_pa=wb["Lundi"]
                pa_sp=sf(self._c(ws_pa,"B21"));pa_go=sf(self._c(ws_pa,"B22"));pa_gnr=sf(self._c(ws_pa,"B23"))
                pv_sp=sf(self._c(ws_pa,"C21"));pv_go=sf(self._c(ws_pa,"C22"));pv_gnr=sf(self._c(ws_pa,"C23"))
                marge_unit=sf(self._c(ws_pa,"C25"))
            except Exception as _e: _log_silent_err(exc=_e)
            o["pa_sp"]=pa_sp;o["pa_go"]=pa_go;o["pa_gnr"]=pa_gnr
            o["pv_sp"]=pv_sp;o["pv_go"]=pv_go;o["pv_gnr"]=pv_gnr
            o["marge_unit"]=marge_unit
            # Marges par carburant (déduites de prix vente - prix achat)
            o["marge_sp"]=round(pv_sp-pa_sp,5) if (pv_sp and pa_sp) else 0
            o["marge_go"]=round(pv_go-pa_go,5) if (pv_go and pa_go) else 0
            o["marge_gnr"]=round(pv_gnr-pa_gnr,5) if (pv_gnr and pa_gnr) else 0
            # Livraison de la veille (E5/E6/E7 du bloc inventaire 6h)
            sp_livr_v=sf(self._c(ws,"E5"));go_livr_v=sf(self._c(ws,"E6"));gnr_livr_v=sf(self._c(ws,"E7"))
            # Détection fraîcheur : comparer C16/C17/C18 (conso veille saisie) avec ventes du
            # dernier jour complet LITRAGE. Match → bilan saisi ce matin. Sinon → obsolète.
            # Tolérance ±50L ou ±2% (arrondis pompistes). NE PAS utiliser C11 (=AUJOURDHUI()).
            bilan_freshness="stale"
            j1_complete=None
            if hist:
                j1_complete=next((h for h in reversed(hist) if not h.get("en_cours")),None)
            def _bm(cv,vt):
                if cv is None or vt is None: return False
                return abs(cv-vt)<=max(50,0.02*max(abs(cv),abs(vt)))
            if j1_complete:
                if _bm(o["sp_c"],sf(j1_complete.get("sp",0))) and _bm(o["go_c"],sf(j1_complete.get("go",0))) and _bm(o["gnr_c"],sf(j1_complete.get("gnr",0))):
                    bilan_freshness="today"
            o["bilan_freshness"]=bilan_freshness
            o["bilan_date_raw"]=j1_complete.get("label","") if j1_complete else ""
            # Stratégie selon fraîcheur
            if bilan_freshness=="today":
                # Bilan à jour : on l'utilise tel quel
                sp_matin=sp_b;go_matin=go_b;gnr_matin=gnr_b
                o["source"]="bilan_today"
            else:
                # Bilan obsolète : reconstituer
                # stock_aujourdhui_matin = B (stock veille matin) - ventes_veille + livraison_veille
                if j1_complete:
                    sp_matin=sp_b-sf(j1_complete.get("sp",0))+sp_livr_v
                    go_matin=go_b-sf(j1_complete.get("go",0))+go_livr_v
                    gnr_matin=gnr_b-sf(j1_complete.get("gnr",0))+gnr_livr_v
                    o["source"]="reconstructed"
                else:
                    sp_matin=sp_b;go_matin=go_b;gnr_matin=gnr_b
                    o["source"]="bilan_stale"
            o["sp_matin"]=sp_matin;o["go_matin"]=go_matin;o["gnr_matin"]=gnr_matin
            # Stocks temps réel : matin - ventes partielles + livraison
            sp=sp_matin;go=go_matin;gnr=gnr_matin
            if partial_today:
                sp-=sf(partial_today.get("sp"))
                go-=sf(partial_today.get("go"))
                gnr-=sf(partial_today.get("gnr"))
            # Livraisons : agrège toutes celles du jour (compat ancien format)
            livr_agg=aggregate_livr_day(livr)
            if livr_agg:
                sp+=sf(livr_agg.get("sp"));go+=sf(livr_agg.get("go"));gnr+=sf(livr_agg.get("gnr"))
                o["livr_recu"]=True
            else: o["livr_recu"]=False
            o["sp"]=sp;o["go"]=go;o["gnr"]=gnr
            # Valorisation temps réel = stocks RT × prix achat unitaires
            o["valo"]=sp*pa_sp+go*pa_go+gnr*pa_gnr
            # Calculer écarts du bilan si fait aujourd'hui
            # Écarts du bilan matière : déjà calculés par Excel en D16/D17/D18 (en litres)
            # NE PAS recalculer (sp_c - sp_b est faux, ce serait conso - stock matin)
            o["ecart_sp"]=sf(self._c(ws,"D16")) if bilan_freshness=="today" else None
            o["ecart_go"]=sf(self._c(ws,"D17")) if bilan_freshness=="today" else None
            o["ecart_gnr"]=sf(self._c(ws,"D18")) if bilan_freshness=="today" else None
            # Lire prévisions selon cycle 14j (jour+offset → onglet)
            # IMPORTANT : le fichier Pre_vision couvre 14 jours (Lundi → Dim2). Au-delà de
            # cet horizon, les onglets représentent des dates passées : col D (ventes prévues)
            # ET col E (livraisons) sont des résidus du cycle précédent. Aucune des deux n'est
            # fiable. Donc on ne crée AUCUN forecast pour les jours hors horizon. Les modules
            # en aval (saisies impossibles, anti-rupture, projection 14j) n'auront rien à
            # analyser pour ces jours → pas d'alertes basées sur des données fantômes.
            o["forecasts"]=[]
            horizon=get_pre_vision_horizon()
            for offset in range(14):
                target_day=date.today()+timedelta(days=offset)
                # Hors horizon → on n'inclut pas ce jour dans les forecasts
                if horizon is not None and target_day>horizon:
                    continue
                target_jour=JOURS_FR[target_day.weekday()]
                sheet_name=get_sheet_for_day(offset)
                ws_d=None
                if sheet_name and sheet_name in wb.sheetnames:
                    ws_d=wb[sheet_name]
                elif target_jour in wb.sheetnames:
                    ws_d=wb[target_jour];sheet_name=target_jour
                if ws_d:
                    # Lire D7/D8/D9 en gardant la distinction cellule vide (None) vs 0 saisi.
                    # Une cellule vide = "pas saisi, hub peut estimer". Une cellule = 0 = "je dis explicitement 0
                    # (rupture prévue, station fermée, pas de ventes attendues)" → respecter cette saisie.
                    raw_d7=self._c(ws_d,"D7");raw_d8=self._c(ws_d,"D8");raw_d9=self._c(ws_d,"D9")
                    o["forecasts"].append({
                        "day":target_jour,"date":target_day,"sheet":sheet_name,
                        "sp":sf(raw_d7),"go":sf(raw_d8),"gnr":sf(raw_d9),
                        # Flags : True si l'utilisateur a explicitement saisi une valeur (même 0), False si vide
                        "sp_saisi":raw_d7 is not None,"go_saisi":raw_d8 is not None,"gnr_saisi":raw_d9 is not None,
                        "livr_sp":sf(self._c(ws_d,"E7")),"livr_go":sf(self._c(ws_d,"E8")),"livr_gnr":sf(self._c(ws_d,"E9")),
                        # Stocks matin (06h) saisis dans Pre_vision = vérité officielle Bidou
                        "stm_sp":sf(self._c(ws_d,"C7")),"stm_go":sf(self._c(ws_d,"C8")),"stm_gnr":sf(self._c(ws_d,"C9")),
                    })
        except Exception as e: print(f"[ERR pv] {e}");o["st"]="err"
        finally:
            try:wb.close()
            except Exception as _e: _log_silent_err(exc=_e)
        return o

    def _calc_autonomie(self,pv,hist_data=None):
        """Calcul autonomie en jours avec le stock ACTUEL (sans livraisons futures).
        L'autonomie est calculée jusqu'au PLANCHER PHYSIQUE (rupture commerciale), pas
        jusqu'à zéro. En dessous de 500L SP/GO et 250L GNR, les pompes ne distribuent
        plus → c'est de fait une rupture du point de vue commercial.
        Utilise la MOYENNE HISTORIQUE par jour de semaine (hors fériés), INDÉPENDAMMENT
        du cycle Pre_vision visible : l'autonomie d'un stock physique ne doit jamais
        être bornée par l'horizon de prévision (qui peut être de 3-4 jours seulement
        en seconde moitié de cycle 14j)."""
        if pv.get("st")!="ok": return {"sp":0,"go":0,"gnr":0}
        # GARDE-FOU : conso PRUDENTE (P75) par jour de semaine, sur 8 derniers mêmes jours.
        # PAS la moyenne 4j (qui surévaluait l'autonomie : un jour creux/0 tirait la conso
        # vers le bas → SP affiché 3j2h alors que la réalité terrain était 2j16h).
        # Le P75 est robuste aux jours à 0 et pessimiste par construction : c'est ce qu'un
        # garde-fou anti-rupture doit être.
        # GARDE-FOU : conso de référence par jour de semaine, sur jours SAINS.
        # Les jours de rupture (connus du journal evenements.cfg) sont EXCLUS par
        # carburant → la conso reflète la vraie demande, pas les ventes bridées.
        # Moyenne simple sur données saines (le P75 n'était qu'un contournement
        # tant qu'on ne savait pas identifier les jours pourris ; maintenant si).
        avg_wd=conso_garde_fou_par_jour_semaine(hist_data or [],nb_derniers=8) if hist_data else {}
        result={}
        today=date.today()
        for carb in ["sp","go","gnr"]:
            stock_brut=sf(pv.get(carb))
            plancher=PLANCHER_PHYSIQUE.get(carb,500)
            # GARDE-FOU INCOHÉRENCE (terrain Bidou 19/05) : un stock NÉGATIF est
            # physiquement impossible (une cuve ne contient pas -989 L). Cela ne
            # signifie JAMAIS une rupture : c'est le signe que les entrées ne sont
            # pas synchronisées (ex : GEST_PISTE pas encore passé à la caisse du
            # jour alors que les ventes du jour sont déjà comptées ailleurs).
            # Dans ce cas on renvoie un marqueur "incoherent" — surtout PAS 0, qui
            # déclencherait à tort l'alerte rupture critique. L'appelant (alerte)
            # doit ignorer ce carburant tant que les données ne sont pas cohérentes.
            if stock_brut<0:
                result[carb]="incoherent"
                continue
            # Stock VENDABLE = stock total - plancher physique. Si déjà sous le
            # plancher (mais >=0), autonomie = 0 = vraie rupture commerciale.
            stock=max(0,stock_brut-plancher)
            if stock<=0:
                result[carb]=0
                continue
            days=0
            # Itération sur jours futurs jusqu'à épuisement du stock vendable.
            # Cap dur à 90 jours pour éviter boucle infinie en cas de conso=0.
            for offset in range(90):
                wd=(today+timedelta(days=offset)).weekday()
                # Conso ATTENDUE : moyenne historique du jour de semaine (jamais 0 sauf si pas de données)
                if avg_wd and wd in avg_wd:
                    conso=sf(avg_wd[wd].get(carb,0))
                else:
                    # Fallback dur si aucune donnée historique
                    conso={"sp":14000,"go":12000,"gnr":700}.get(carb,1000)
                if conso<=0: conso=1
                if stock>=conso:
                    stock-=conso; days+=1
                else:
                    days+=stock/conso; stock=0; break
            result[carb]=round(days,3)
        return result

    def projection_14j(self,pv,hist_data,partial_today=None):
        """Projection stock jour par jour sur 14j avec :
        - ventes estimées = moyenne 4 derniers mêmes jours (hors fériés)
        - livraisons prévues = col E Pre_vision
        - détection anomalie ventes du jour vs moyenne
        Retourne dict avec projection[], anomalies{}, alertes[]."""
        if pv.get("st")!="ok": return {"ok":True,"projection":[],"anomalies":{},"alertes":[]}
        forecasts=pv.get("forecasts",[])
        # Moyennes historiques par jour de semaine
        avg_wd=avg_ventes_par_jour_semaine(hist_data,nb_derniers=4)
        # Jour pivot : premier jour où une commande SARA est encore actionnable.
        # Avant 11h ET aucune caisse saisie → today (= aujourd'hui)
        # Après 11h OU C1 saisie → J+1 ouvré (la journée d'aujourd'hui est figée pour les commandes)
        # IMPORTANT : ce jour pivot sert UNIQUEMENT à savoir QUAND afficher les commandes à passer.
        # Pour le CALCUL de stock, on part TOUJOURS d'aujourd'hui (date.today()) pour intégrer
        # tous les jours et ne pas sauter par dessus des ponts.
        jour_pivot=get_jour_pivot(partial_today=partial_today)
        today=date.today()
        # === DEBUG : log les forecasts construits par _read_pv ===
        debug_log("="*60)
        debug_log(f"projection_14j : aujourd'hui={today}, jour_pivot={jour_pivot}")
        debug_log(f"forecasts ({len(forecasts)} jours) :")
        for fc in forecasts:
            debug_log(f"  {fc.get('date')} sheet={fc.get('sheet','?')} stm_sp={fc.get('stm_sp')} stm_go={fc.get('stm_go')} stm_gnr={fc.get('stm_gnr')} D={fc.get('sp')}/{fc.get('go')}/{fc.get('gnr')} E={fc.get('livr_sp')}/{fc.get('livr_go')}/{fc.get('livr_gnr')}")
        proj={"ok":True,"projection":[],"anomalies":{},"alertes":[],"avg_wd":avg_wd,"jour_pivot":jour_pivot}
        # Stocks de départ : on PRIORISE les stocks matin Pre_vision d'AUJOURD'HUI (vérité Bidou).
        # Sinon fallback sur les stocks "temps réel" calculés par le hub depuis le bilan matière.
        stocks_pv=None
        for fc in forecasts:
            if fc.get("date")==today:
                stm_sp=sf(fc.get("stm_sp",0))
                stm_go=sf(fc.get("stm_go",0))
                stm_gnr=sf(fc.get("stm_gnr",0))
                if stm_sp>0 or stm_go>0 or stm_gnr>0:
                    stocks_pv={"sp":stm_sp,"go":stm_go,"gnr":stm_gnr}
                break
        if stocks_pv:
            stocks=stocks_pv
            debug_log(f"stocks INITIAUX (Pre_vision matin du pivot) : {stocks}")
        else:
            stocks={"sp":sf(pv.get("sp")),"go":sf(pv.get("go")),"gnr":sf(pv.get("gnr"))}
            debug_log(f"stocks INITIAUX (fallback bilan matière) : {stocks}")
        seuil_jours=1.5  # sous 1,5j de stock = alerte
        # Horizon Pre_vision dynamique : ne PAS projeter au-delà du dernier jour couvert
        # par le fichier actuel. Au-delà, on n'a ni ventes prévues ni livraisons fiables
        # → projeter avec 0 livraison créerait des "ponts" fantômes (ex: week-end du
        # cycle suivant signalé en rupture alors que les commandes ne sont pas encore
        # saisies). On s'arrête à l'horizon ; le prochain cycle prendra le relais.
        horizon_pv=get_pre_vision_horizon()
        for offset in range(14):
            target=today+timedelta(days=offset)
            if horizon_pv is not None and target>horizon_pv:
                break
            wd=target.weekday()
            avg_day=avg_wd.get(wd,{"sp":9000,"go":7000,"gnr":300})
            # Livraison prévue (du forecast si disponible)
            livr_prev={"sp":0,"go":0,"gnr":0}
            fc_match=None
            for fc in forecasts:
                fc_d=fc.get("date")
                if fc_d and fc_d==target:
                    fc_match=fc
                    livr_prev={"sp":sf(fc.get("livr_sp",0)),"go":sf(fc.get("livr_go",0)),"gnr":sf(fc.get("livr_gnr",0))}
                    break
            # Ventes estimées : PRIORITÉ aux saisies D7/D8/D9 du fichier Pre_vision si > 0,
            # sinon fallback moyenne historique du jour de semaine.
            # Pour férié sans saisie : profil dimanche (Option A validée par Bidou).
            ventes_est={"sp":avg_day["sp"],"go":avg_day["go"],"gnr":avg_day["gnr"]}
            for carb in ("sp","go","gnr"):
                saisi=sf(fc_match.get(carb,0)) if fc_match else 0
                if saisi>0:
                    ventes_est[carb]=saisi  # saisie pompiste prioritaire
            # Calculer stock fin de jour pour chaque carburant
            day_data={"date":target,"wd":wd,"offset":offset,"ferie":is_ferie(target),
                      "ventes_perdues":{"sp":0,"go":0,"gnr":0}}
            for carb in ["sp","go","gnr"]:
                # Logique de choix de la valeur de vente (D) :
                # 1. Si utilisateur a saisi explicitement (même 0) → respecter sa saisie
                #    (cellule vide ≠ saisie 0 ; un 0 saisi signifie "rupture prévue, je vends 0")
                # 2. Sinon (cellule vide) ET férié → fallback profil dimanche (Option A)
                # 3. Sinon (cellule vide) → fallback moyenne historique du jour de semaine
                if fc_match and fc_match.get(f"{carb}_saisi"):
                    v_voulu=sf(fc_match.get(carb,0))  # saisie explicite, même 0
                elif is_ferie(target):
                    avg_dim=avg_wd.get(6,{"sp":8487,"go":5350,"gnr":253})
                    v_voulu=avg_dim[carb]
                else:
                    v_voulu=ventes_est[carb]
                # PLAFONNEMENT PHYSIQUE : on ne peut pas vendre plus que (stock matin + livraison) - plancher.
                # En dessous du plancher (500L SP/GO, 250L GNR), les pompes ne distribuent plus.
                # Si la vente prévue dépasse cette limite, on plafonne et on stocke le manque à gagner.
                stock_dispo=stocks[carb]+livr_prev[carb]
                plancher=PLANCHER_PHYSIQUE.get(carb,500)
                v_max_physique=max(0,stock_dispo-plancher)
                if v_voulu>v_max_physique:
                    v_reelle=v_max_physique
                    day_data["ventes_perdues"][carb]=int(v_voulu-v_reelle)
                else:
                    v_reelle=v_voulu
                stocks[carb]=stocks[carb]-v_reelle+livr_prev[carb]
                day_data[f"stock_{carb}"]=stocks[carb]
                day_data[f"ventes_{carb}"]=v_reelle
                day_data[f"ventes_voulues_{carb}"]=v_voulu  # ce qui était prévu avant plafonnement
                day_data[f"livr_{carb}"]=livr_prev[carb]
            proj["projection"].append(day_data)
        # === DEBUG : log la projection complète ===
        debug_log("Projection 14j calculée (avec plafonnement physique 500L SP/GO, 250L GNR) :")
        for d in proj["projection"]:
            vp=d.get("ventes_perdues",{"sp":0,"go":0,"gnr":0})
            perdues_str=""
            if any(vp.values()):
                perdues_str=f"  ⚠ VENTES PERDUES SP={vp['sp']} GO={vp['go']} GNR={vp['gnr']}"
            debug_log(f"  offset {d['offset']:>2} {d['date']} : stocks SP/GO/GNR = {d['stock_sp']:.0f}/{d['stock_go']:.0f}/{d['stock_gnr']:.0f}  (ventes {d['ventes_sp']:.0f}/{d['ventes_go']:.0f}/{d['ventes_gnr']:.0f}, livr {d['livr_sp']:.0f}/{d['livr_go']:.0f}/{d['livr_gnr']:.0f}){perdues_str}")
        # Détecter les alertes : RUPTURE RÉELLE uniquement (stock ≤ 0)
        # Distinguer : rupture DANS le cycle Pre_vision (alerte forte) vs APRÈS (fin de cycle à remplir)
        # Dernier jour avec livraison prévue dans Pre_vision
        last_livr_date=None
        for day_data in proj["projection"]:
            has_livr=any(day_data.get(f"livr_{c}",0)>0 for c in ["sp","go","gnr"])
            if has_livr:
                last_livr_date=day_data["date"]
        for carb in ["sp","go","gnr"]:
            rupture_date=None
            stock_min=999999;stock_min_date=None
            for day_data in proj["projection"]:
                s=day_data[f"stock_{carb}"]
                if s<stock_min:
                    stock_min=s;stock_min_date=day_data["date"]
                if s<=0 and not rupture_date:
                    rupture_date=day_data["date"]
            proj[f"stock_min_{carb}"]=stock_min
            proj[f"stock_min_date_{carb}"]=stock_min_date
            if rupture_date:
                # Déterminer la sévérité selon la position dans le cycle
                # Si la rupture est APRÈS le dernier jour avec livraison prévue → fin de cycle (info)
                # Sinon → vraie alerte (rupture au milieu du cycle)
                JOURS_COURTS=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
                if last_livr_date and rupture_date>last_livr_date:
                    # Info : fin de cycle, cycle Prévision à remplir
                    alerte={
                        "carburant":carb.upper(),
                        "severity":"fin_cycle",
                        "date":rupture_date,
                        "date_str":f"{JOURS_COURTS[rupture_date.weekday()]} {rupture_date.strftime('%d/%m')}",
                        "last_livr_date":last_livr_date,
                        "last_livr_date_str":f"{JOURS_COURTS[last_livr_date.weekday()]} {last_livr_date.strftime('%d/%m')}",
                        "stock_min":stock_min,
                        "stock_min_date":stock_min_date,
                    }
                else:
                    # Vraie rupture dans le cycle : calculer deadline commande
                    proj["ok"]=False
                    d=rupture_date-timedelta(days=1)
                    while d.weekday()>=5 or is_ferie(d):
                        d-=timedelta(days=1)
                    alerte={
                        "carburant":carb.upper(),
                        "severity":"rupture",
                        "date":rupture_date,
                        "date_str":f"{JOURS_COURTS[rupture_date.weekday()]} {rupture_date.strftime('%d/%m')}",
                        "deadline":d,
                        "deadline_str":f"{JOURS_COURTS[d.weekday()]} {d.strftime('%d/%m')} (11h)",
                        "stock_min":stock_min,
                        "stock_min_date":stock_min_date,
                    }
                proj["alertes"].append(alerte)
        # Détection anomalie ventes du jour : comparer les ventes cumulées sur N caisses saisies
        # à la moyenne des mêmes N premières caisses sur les 4 derniers mêmes jours de semaine.
        # Évite le bug "proportion horaire" : les caisses se saisient par lots (fin de shift), pas en continu.
        if partial_today:
            nb_c=int(partial_today.get("nb_caisses",0))
            if nb_c>0:
                ref=avg_ventes_n_caisses(hist_data,today.weekday(),nb_c,nb_derniers=4)
                if ref:
                    stage=f"{nb_c}/3 de la journée"
                    for carb in ["sp","go","gnr"]:
                        ventes_partielles=sf(partial_today.get(carb,0))
                        ventes_attendues=ref[carb]
                        if ventes_attendues>0:
                            ratio=ventes_partielles/ventes_attendues
                            ecart_pct=(ratio-1)*100
                            if abs(ecart_pct)>20:
                                proj["anomalies"][carb]={
                                    "ventes_partielles":ventes_partielles,
                                    "ventes_attendues":round(ventes_attendues),
                                    "ratio":ratio,
                                    "ecart_pct":round(ecart_pct),
                                    "stage":stage,
                                    "nb_caisses":nb_c,
                                    "impact":"hausse" if ecart_pct>0 else "baisse",
                                }
        return proj

    def analyze_antirupture(self,pv,proj14,hist_data):
        """Analyse anti-rupture exploitant projection_14j déjà calculée.
        Retourne dict avec :
          - feries_a_venir : liste des fériés J+15 à J+30 (Information)
          - trous : liste des trous de jours non-livrables (≥ 2 jours consécutifs)
          - incoherences_jour_non_livrable : commandes saisies sur jour weekend/férié
          - ventes_irrealistes : saisies D7/D8/D9 vs moyenne historique (écart >30%)
          - ruptures_dans_trou : par carburant, manque calculé
          - plan_lisse : suggestion de répartition des commandes sous contraintes
          - severite_max : 'info' / 'vigilance' / 'critique' pour pilotage UI
        Capacités cuves : SP 40k, GO 40k, GNR 10k. Marge sécurité fichier Bidou :
        1000L SP/GO, 500L GNR. Limite camion : 36 m³/jour."""
        result={
            "feries_a_venir":[],
            "trous":[],
            "incoherences_jour_non_livrable":[],
            "livraisons_non_conformes":[],
            "ventes_irrealistes":[],
            "saisies_physiquement_impossibles":[],
            "livraisons_a_reporter":[],  # livraisons prévues qui dépassent la capacité cuve à 6h
            "ruptures_projetees":[],  # jours où le hub a plafonné les ventes (rupture projetée)
            "ruptures_dans_trou":[],
            "plan_lisse":None,
            "severite_max":"info",  # info / vigilance / critique
        }
        if pv.get("st")!="ok": return result
        # Réutiliser le jour pivot calculé par projection_14j pour cohérence
        today=proj14.get("jour_pivot") or get_jour_pivot()
        result["jour_pivot"]=today
        forecasts=pv.get("forecasts",[]) or []
        avg_wd=proj14.get("avg_wd") or avg_ventes_par_jour_semaine(hist_data,nb_derniers=4)
        # === Constantes (validées par Bidou) ===
        CAPA={"sp":40000,"go":40000,"gnr":10000}
        MARGE_SECU={"sp":1000,"go":1000,"gnr":500}
        CAMION_MAX=36000  # 1 camion = 36 m³ max
        SEUIL_SECU_JOURS=0.5  # 0,5 jour de stock résiduel post-trou
        # === Contexte partagé entre sections (pattern Context Object) ===
        # Permet d'extraire chaque section en méthode sans devoir passer 10 paramètres.
        # Chaque section LIT des entrées via ctx et ÉCRIT ses sorties soit dans result,
        # soit dans ctx lui-même (ex: section 2 écrit `trous` et `projection` dans ctx
        # pour que les sections suivantes y aient accès).
        ctx={"today":today,"forecasts":forecasts,"avg_wd":avg_wd,"proj14":proj14,
             "hist_data":hist_data,"CAPA":CAPA,"MARGE_SECU":MARGE_SECU,
             "CAMION_MAX":CAMION_MAX,"SEUIL_SECU_JOURS":SEUIL_SECU_JOURS,"pv":pv}
        # === Sections 1 & 2 extraites (modularisation progressive) ===
        self._ar_compute_feries_a_venir(result,ctx)
        self._ar_compute_trous(result,ctx)
        self._ar_compute_feries_isoles_imminents(result,ctx)
        # Récupération dans le scope local : ces variables sont utilisées par les sections
        # 3 et 6 qui restent inline pour cette étape du refacto.
        projection=ctx["projection"]
        trous=ctx["trous"]
        JC=ctx["JC"]
        # === Section 3 extraite (modularisation progressive) ===
        # 3a : cohérences jour non-livrable + 3b : conformité SARA
        # 3c-3e : saisies physiquement impossibles, livraisons à reporter, ruptures projetées
        self._ar_compute_coherences_jour_non_livrable(result,ctx)
        self._ar_compute_conformite_sara(result,ctx)
        self._ar_compute_saisies_et_reports(result,ctx)
        # === Sections 4 & 5 extraites (modularisation progressive) ===
        self._ar_compute_ventes_irrealistes(result,ctx)
        self._ar_compute_ruptures_dans_trou(result,ctx)
        # === Section 6 extraite (modularisation progressive) ===
        # 6a : Plan de commande (lignes par carburant selon manques + capa cuve)
        # 6b : Acquittements par pont + détection deadline dépassée (utilise result["ruptures_dans_trou"])
        # 6c : Ruptures imminentes (autonomie < 24h aujourd'hui, indépendant des ponts)
        self._ar_compute_plan_commande(result,ctx)
        self._ar_compute_acquittements(result,ctx)
        self._ar_compute_ruptures_imminentes(result,ctx)
        # ack_status est calculé par _ar_compute_acquittements et exposé dans ctx
        # === Sections finales (7 sévérité, 7bis tendances) — déjà extraites ===
        # Le ctx porte today/proj14/forecasts/avg_wd/CAPA/... + trous/projection/ack_status
        # ajoutés au fil de l'eau par les sections 2/6.
        self._ar_compute_severite(result,ctx)
        self._ar_compute_tendances_fortes(result,ctx)
        return result

    def _ar_compute_plan_commande(self,result,ctx):
        """Section 6a d'analyze_antirupture — Plan de commande aligné sur les manques annoncés.
        
        Règles :
          1. On commande UNIQUEMENT les carburants signalés en manque (section 5)
          2. Volume = manque arrondi à la tranche 1000 L SUPÉRIEURE
          3. Minimum 2000 L par ligne (contrainte SARA)
          4. Date de commande = veille jour ouvré du jour de livraison cible (juste avant le pont)
          5. Si commande > capa cuve dispo → flag "3e tour conseillé" (ventes C1 libèrent la place)
        """
        if not result["ruptures_dans_trou"]: return
        projection=ctx["projection"]
        hist_data=ctx["hist_data"]
        CAPA=ctx["CAPA"]
        MARGE_SECU=ctx["MARGE_SECU"]
        ruptures_par_trou={}
        for r in result["ruptures_dans_trou"]:
            key=(r["trou_start"],r["trou_end"])
            ruptures_par_trou.setdefault(key,[]).append(r)
        plans=[]
        for (ts,te),ruptures in ruptures_par_trou.items():
            # Trouver le jour de livraison cible = jour livrable juste avant le trou
            last_livrable_idx=None
            for idx,day in enumerate(projection):
                if day["date"]<ts and day["date"].weekday()<5 and not is_ferie(day["date"]):
                    last_livrable_idx=idx
            if last_livrable_idx is None: continue
            jour_livraison=projection[last_livrable_idx]
            # Date de commande = veille jour ouvré du jour de livraison
            jour_commande=jour_livraison["date"]-timedelta(days=1)
            while jour_commande.weekday()>=5 or is_ferie(jour_commande):
                jour_commande-=timedelta(days=1)
            # Construire les lignes : un carburant = une entrée si manque > 0
            lignes_carb=[]
            infaisable_carbs=[]
            for r in ruptures:
                carb=r["carburant"].lower()
                manque=r["manque"]
                # Volume à commander = arrondi tranche 1000 L SUPÉRIEURE, min 2000 L SARA
                volume=int(((manque+999)//1000)*1000)
                if volume<2000: volume=2000
                # Vérifier capa cuve : stock_matin du jour de livraison + ce qui est déjà livré
                stock_matin_livr=sf(jour_livraison.get(f"stock_{carb}",0))-sf(jour_livraison.get(f"livr_{carb}",0))+sf(jour_livraison.get(f"ventes_{carb}",0))
                livr_deja_prevue=sf(jour_livraison.get(f"livr_{carb}",0))
                capa_cuve=CAPA.get(carb,40000)
                marge=MARGE_SECU.get(carb,0)
                place_dispo=max(0,capa_cuve-stock_matin_livr-livr_deja_prevue-marge)
                tour_3=False
                if volume>place_dispo:
                    # 3e tour possible si ventes C1 libèrent assez de place d'ici midi
                    wd=jour_livraison["date"].weekday()
                    if hist_data:
                        ref_c1=avg_ventes_n_caisses(hist_data,wd,1,nb_derniers=4) or {}
                        ventes_c1=sf(ref_c1.get(carb,0))
                    else:
                        ventes_c1=0
                    if volume<=place_dispo+ventes_c1:
                        tour_3=True
                    else:
                        infaisable_carbs.append(carb.upper())
                        continue
                lignes_carb.append({
                    "carburant":carb.upper(),
                    "volume":volume,
                    "manque_initial":int(manque),
                    "tour_3":tour_3,
                })
            if not lignes_carb and not infaisable_carbs: continue
            JC2=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
            plans.append({
                "trou_start":ts,
                "trou_end":te,
                "trou_str":f"{ts.strftime('%d/%m')} \u2192 {te.strftime('%d/%m')}",
                "jour_commande":jour_commande,
                "jour_commande_str":f"{JC2[jour_commande.weekday()]} {jour_commande.strftime('%d/%m')} avant 11h",
                "jour_livraison":jour_livraison["date"],
                "jour_livraison_str":f"{JC2[jour_livraison['date'].weekday()]} {jour_livraison['date'].strftime('%d/%m')}",
                "lignes_carb":lignes_carb,
                "infaisable_carbs":infaisable_carbs,
                "infaisable":bool(infaisable_carbs),
            })
        result["plan_lisse"]=plans

    def _ar_compute_acquittements(self,result,ctx):
        """Section 6b d'analyze_antirupture — Acquittements par pont + détection deadline dépassée.
        
        Charge les acquittements existants (ANTIRUPTURE_ACK_FILE). Un acquittement reste valide
        selon son type :
          - rupture_acceptee : toujours acquitté (Bidou a accepté de perdre le pont)
          - snooze : acquitté tant que snooze_until_iso > maintenant
          - controle (défaut) : tant qu'aucun manque n'a augmenté de +20% et qu'aucun nouveau
                                carburant n'est tombé en manque
        Nettoie aussi les acks des ponts disparus (passés ou corrigés).
        Expose `ack_status` dans le ctx pour les sections 6c, 7, 7bis.
        """
        try: acks=load_json(ANTIRUPTURE_ACK_FILE) or {}
        except Exception: acks={}
        # Nettoyage : retirer les acks des ponts qui n'existent plus dans les manques actuels
        ponts_actuels={f"pont_{r['trou_start'].strftime('%d%m%Y')}" for r in result["ruptures_dans_trou"]}
        ack_changed=False
        for pont_id in list(acks.keys()):
            if pont_id not in ponts_actuels:
                del acks[pont_id]
                ack_changed=True
        # Grouper les manques actuels par pont_id pour comparaison avec ack
        manques_par_pont={}
        for r in result["ruptures_dans_trou"]:
            pont_id=f"pont_{r['trou_start'].strftime('%d%m%Y')}"
            manques_par_pont.setdefault(pont_id,{})
            manques_par_pont[pont_id][r["carburant"].lower()]=r["manque"]
        now=datetime.now()
        ack_status={}
        for pont_id,manques_actuels in manques_par_pont.items():
            ack=acks.get(pont_id)
            acquitte=False;raison_inv=None
            type_ack=ack.get("type_ack","controle") if ack else None
            extra={}
            if ack:
                if type_ack=="rupture_acceptee":
                    # Bidou a accepté de perdre le pont : jamais réinvalidé
                    acquitte=True
                    extra["cause"]=ack.get("cause","")
                elif type_ack=="snooze":
                    # Acquitté tant que snooze_until_iso > maintenant
                    try:
                        until_dt=datetime.fromisoformat(ack.get("snooze_until_iso",""))
                        if now<until_dt:
                            acquitte=True
                            extra["snooze_until_iso"]=ack.get("snooze_until_iso")
                        else:
                            raison_inv="snooze_expire"
                    except Exception:
                        raison_inv="snooze_invalide"
                else:
                    # Type "controle" classique : invalide si manque +20% OU nouveau carburant
                    # Plus d'expiration arbitraire 24h (Bidou : "j'ai acquitté = j'ai acquitté")
                    ack_manques=ack.get("manques",{}) or {}
                    invalide=False
                    for carb,m_now in manques_actuels.items():
                        m_ack=sf(ack_manques.get(carb,0))
                        if m_ack<=0:
                            invalide=True;raison_inv="nouveau_carburant_en_manque";break
                        if m_now>m_ack*1.2:
                            invalide=True;raison_inv=f"manque_{carb}_aggravee";break
                    if not invalide:
                        acquitte=True
            # Deadline dépassée = deadline du 1er rupture du pont est passée
            ruptures_pont=[r for r in result["ruptures_dans_trou"] if f"pont_{r['trou_start'].strftime('%d%m%Y')}"==pont_id]
            deadline_depassee=False
            if ruptures_pont:
                deadline_date=ruptures_pont[0]["deadline"]
                deadline_dt=datetime.combine(deadline_date,datetime.min.time()).replace(hour=11)
                if now>deadline_dt:
                    deadline_depassee=True
            status_entry={"acquitte":acquitte,"deadline_depassee":deadline_depassee,
                          "raison_invalidation":raison_inv,"type_ack":type_ack}
            status_entry.update(extra)
            ack_status[pont_id]=status_entry
        # Sauvegarder acks nettoyés
        if ack_changed:
            try: save_json(ANTIRUPTURE_ACK_FILE,acks)
            except Exception as _e: _log_silent_err(exc=_e)
        result["ack_status"]=ack_status
        result["acks_actifs"]=acks
        ctx["ack_status"]=ack_status  # exposé pour sections suivantes (6c, 7, 7bis)

    def _ar_compute_ruptures_imminentes(self,result,ctx):
        """Section 6c d'analyze_antirupture — Ruptures imminentes (autonomie < 24h aujourd'hui).
        
        Distincte de `ruptures_dans_trou` (alerte sur un PONT FUTUR, peut être acquitté) :
        ici on alerte sur MAINTENANT (autonomie réelle aujourd'hui, indépendant des ponts).
        
        Règles :
          - Auto < 8h : alerter TOUJOURS, même si livraison prévue (rien ne garantit qu'elle arrive)
          - Auto 8h-24h : alerter sauf si une livraison est prévue dans la fenêtre d'autonomie
          - Auto >= 24h : pas d'alerte
        
        Conserve aussi un log debug temporaire (debug_rupture_imm.log) pour diagnostic terrain.
        """
        result["ruptures_imminentes"]=[]
        pv=ctx["pv"]
        hist_data=ctx["hist_data"]
        forecasts=ctx["forecasts"]
        trous=ctx["trous"]
        # DEBUG : log temporaire pour comprendre pourquoi la popup ne se déclenche pas
        _dbg_lines=[]
        _dbg_lines.append(f"=== {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
        _dbg_lines.append(f"pv.st={pv.get('st')!r}, pv.sp={pv.get('sp')!r}, pv.go={pv.get('go')!r}, pv.gnr={pv.get('gnr')!r}")
        try:
            auto_calc=self._calc_autonomie(pv,hist_data) or {}
            _dbg_lines.append(f"auto_calc={auto_calc}")
            today_d=date.today()
            _dbg_lines.append(f"nb_forecasts={len(forecasts)}")
            for i,f in enumerate(forecasts[:5]):
                _dbg_lines.append(f"  forecast[{i}] date={f.get('date')} livr_sp={f.get('livr_sp')} livr_go={f.get('livr_go')} livr_gnr={f.get('livr_gnr')}")
            prochain_trou_iso=trous[0]["start_date"].isoformat() if trous else None
            for carb in ("sp","go","gnr"):
                # GARDE-FOU INCOHÉRENCE : si _calc_autonomie a renvoyé "incoherent"
                # (stock négatif = données pas synchronisées, ex. GEST_PISTE pas
                # encore passé à la caisse du jour), on NE déclenche PAS d'alerte
                # rupture sur ce carburant. Un stock impossible n'est pas une
                # rupture : alerter ici serait un faux positif critique (constat
                # terrain Bidou 19/05 : SP -989 L → fausse "RUPTURE CRITIQUE").
                if auto_calc.get(carb)=="incoherent":
                    _dbg_lines.append(f"[{carb}] SKIP : stock incoherent (donnees non synchronisees)")
                    continue
                auto_jours=sf(auto_calc.get(carb,0))
                auto_h=auto_jours*24
                _dbg_lines.append(f"[{carb}] auto_jours={auto_jours} auto_h={auto_h:.1f}")
                if auto_h>=24:
                    _dbg_lines.append(f"[{carb}] SKIP : auto_h>=24"); continue
                # Chercher livraison du jour et livraison dans fenêtre autonomie
                livr_aujourdhui=0
                livr_dans_fenetre=False
                for f in forecasts:
                    f_d=f.get("date")
                    if not f_d: continue
                    jours_avant=(f_d-today_d).days
                    livr_val=sf(f.get(f"livr_{carb}",0))
                    if jours_avant==0:
                        livr_aujourdhui=livr_val
                    if jours_avant>=0 and jours_avant<=auto_jours+1 and livr_val>0:
                        livr_dans_fenetre=True
                _dbg_lines.append(f"[{carb}] livr_aujourdhui={livr_aujourdhui} livr_dans_fenetre={livr_dans_fenetre}")
                # Règle : sous 8h on alerte toujours, au-dessus on masque si livraison prévue
                if auto_h>=8 and livr_dans_fenetre:
                    _dbg_lines.append(f"[{carb}] SKIP : auto>=8h ET livraison pr\u00e9vue"); continue
                # Prochain jour livrable
                # Cohérence inter-modules (fix 30/05/2026) : un jour n'est réellement
                # livrable que si la commande est ENCORE possible pour ce jour-là, càd
                # si la deadline de commande (11h le jour_de_commande) n'est pas passée.
                # Avant ce fix, ce calcul était purement calendaire (1er jour ouvré à
                # partir d'aujourd'hui) et pouvait afficher "livrable aujourd'hui" alors
                # que la deadline de commande était dépassée depuis longtemps. On réutilise
                # jour_de_commande (core) + la deadline 11h déjà appliquée partout dans le
                # HUB (cf compute_moments_cles, "Deadline TEMAG : 11h").
                prochain_livr=None
                _now_pl=datetime.now()
                for offset in range(15):
                    d_test=today_d+timedelta(days=offset)
                    if is_ferie(d_test) or d_test.weekday()>=5:
                        continue
                    jc=jour_de_commande(d_test)
                    if jc is None:
                        prochain_livr=d_test; break  # fallback cas extrême (inchangé)
                    if _now_pl<=datetime.combine(jc,dt_time(11,0)):
                        prochain_livr=d_test; break
                    # sinon : deadline de commande 11h passée pour ce jour → jour suivant
                result["ruptures_imminentes"].append({
                    "carburant":carb.upper(),
                    "stock_actuel":int(sf(pv.get(carb,0))),
                    "autonomie_h":int(auto_h),
                    "autonomie_jours":round(auto_jours,2),
                    "livr_aujourdhui":int(livr_aujourdhui),
                    "prochain_livr":prochain_livr.isoformat() if prochain_livr else None,
                    "prochain_trou":prochain_trou_iso,
                })
                _dbg_lines.append(f"[{carb}] \u2713 AJOUT\u00c9 (stock={int(sf(pv.get(carb,0)))} auto_h={int(auto_h)} livr_jour={int(livr_aujourdhui)})")
        except Exception as e:
            _dbg_lines.append(f"ERREUR: {e}")
            print(f"[ruptures_imminentes] {e}")
        _dbg_lines.append(f"total ruptures_imminentes={len(result['ruptures_imminentes'])}")
        try:
            _log_path=os.path.join(os.path.dirname(os.path.abspath(__file__)),"debug_rupture_imm.log")
            with open(_log_path,"a",encoding="utf-8") as _f:
                _f.write("\n".join(_dbg_lines)+"\n\n")
        except Exception as _e: _log_silent_err(exc=_e)

    def _ar_compute_severite(self,result,ctx):
        """Section 7 d'analyze_antirupture — Sévérité max pour pilotage UI.
        
        Sévérité critique uniquement si AU MOINS UN pont N'EST PAS acquitté
        OU s'il y a une rupture imminente en jour livrable (autonomie < 24h sans livraison prévue).
        Ponts acquittés (et toujours valides) → on continue d'afficher dans le détail PRÉVISION,
        mais plus de popup tant que la situation ne s'aggrave pas.
        
        NB : `saisies_physiquement_impossibles` a sa PROPRE popup avec son propre système de
        silence. Ne PAS l'inclure ici, sinon la popup anti-rupture s'ouvre AUSSI à chaque refresh.
        """
        ack_status=ctx["ack_status"]
        ponts_non_acquittes=[pid for pid,st in ack_status.items() if not st["acquitte"]]
        if ponts_non_acquittes or result["incoherences_jour_non_livrable"] or result["ruptures_imminentes"]:
            result["severite_max"]="critique"
        elif result["ventes_irrealistes"] and any(v["severity"]=="rouge" for v in result["ventes_irrealistes"]):
            result["severite_max"]="vigilance"
        elif result["feries_a_venir"] or result["ventes_irrealistes"] or result["ruptures_dans_trou"]:
            result["severite_max"]="info"

    def _ar_compute_tendances_fortes(self,result,ctx):
        """Section 7bis d'analyze_antirupture — Alertes tendance forte.
        
        Croise anomalies de ventes en cours (calculées dans projection_14j) avec ponts à venir.
        Déclenche une alerte tendance pour un carburant si :
          (a) anomalie >= +30% détectée aujourd'hui (HAUSSE seulement — une baisse allège le pont)
          ET
          (b) carburant tendu sur un pont à venir (manque signalé OU stock fin pont < 1j ventes lendemain)
        Permet de prévenir qu'une tendance forte du jour aggrave un pont déjà tendu.
        """
        result["tendance_alertes"]=[]
        proj14=ctx["proj14"]
        trous=ctx["trous"]
        projection=ctx["projection"]
        anomalies=proj14.get("anomalies",{}) or {}
        if not anomalies: return
        avg_wd=ctx.get("avg_wd",{}) or {}
        # Pour chaque pont, calculer le stock projeté en sortie + ventes lendemain
        # pour identifier les ponts "tendus" même sans manque officiel
        ponts_tendus_par_carb={}  # {carb -> [{pont_id, trou_str, manque_arrondi, ...}]}
        for trou in trous:
            if trou["last_livrable_idx"] is None: continue
            day_avant=projection[trou["last_livrable_idx"]]
            next_idx=trou["next_livrable_idx"]
            pont_id=f"pont_{trou['start_date'].strftime('%d%m%Y')}"
            trou_str=f"{trou['start_date'].strftime('%d/%m')} \u2192 {trou['end_date'].strftime('%d/%m')}"
            # Terme correct : "Weekend" (sam/dim seuls) ou "Pont" (contient un férié).
            # Avant, "Pont" était écrit en dur dans la popup → faux pour un simple weekend.
            terme=qualifier_trou({"start_date":trou["start_date"],
                                   "duree":(trou["end_date"]-trou["start_date"]).days+1})
            extension_jours=trou.get("extension_jours",[]) or []
            for carb in ["sp","go","gnr"]:
                stock_apres=sf(day_avant.get(f"stock_{carb}",0))
                # Ventes pendant le trou : projection (jours dans Pre_vision) + extension (avg_wd pour jours hors Pre_vision)
                ventes_trou_c=sum(sf(projection[idx].get(f"ventes_{carb}",0))
                                  for idx in range(trou["start_idx"],trou["end_idx"]+1))
                ventes_trou_c+=sum(avg_wd.get(d.weekday(),{}).get(carb,0) for d in extension_jours)
                stock_fin_pont=stock_apres-ventes_trou_c
                # Ventes lendemain : depuis projection si dispo, sinon avg_wd (jour livrable hors Pre_vision)
                if next_idx is not None:
                    ventes_lendemain=sf(projection[next_idx].get(f"ventes_{carb}",0))
                else:
                    ventes_lendemain=avg_wd.get((trou["end_date"]+timedelta(days=1)).weekday(),{}).get(carb,0)
                # Tendu = stock fin pont < 1 jour de ventes lendemain
                tendu_relatif=ventes_lendemain>0 and stock_fin_pont<ventes_lendemain
                # OU bien manque déjà détecté (cas (a))
                manque_existant=any(r["trou_start"]==trou["start_date"] and r["carburant"].lower()==carb
                                     for r in result["ruptures_dans_trou"])
                if tendu_relatif or manque_existant:
                    manque_obj=next((r for r in result["ruptures_dans_trou"]
                                     if r["trou_start"]==trou["start_date"] and r["carburant"].lower()==carb),None)
                    manque_arr=max(2000,int(((manque_obj["manque"]+999)//1000)*1000)) if manque_obj else 0
                    ponts_tendus_par_carb.setdefault(carb,[]).append({
                        "pont_id":pont_id,
                        "trou_str":trou_str,
                        "terme":terme,
                        "trou_start":trou["start_date"],
                        "stock_fin_pont":int(stock_fin_pont),
                        "ventes_lendemain":int(ventes_lendemain),
                        "manque_arrondi":manque_arr,
                        "deja_manque":manque_obj is not None,
                    })
        # Pour chaque carburant en HAUSSE >= 30% qui touche un pont tendu : alerte tendance.
        # NB : seule une HAUSSE des ventes aggrave un pont. Une BAISSE allège (stock dure plus).
        for carb,anom in anomalies.items():
            if anom["ecart_pct"]<30: continue
            ponts_concernes=ponts_tendus_par_carb.get(carb,[])
            if not ponts_concernes: continue
            result["tendance_alertes"].append({
                "carburant":carb.upper(),
                "ecart_pct":anom["ecart_pct"],
                "stage":anom["stage"],
                "ventes_partielles":anom["ventes_partielles"],
                "ventes_attendues":anom["ventes_attendues"],
                "impact":anom["impact"],
                "ponts":ponts_concernes,
            })

    def _ar_compute_feries_isoles_imminents(self,result,ctx):
        """Détection des fériés ISOLÉS imminents (1 seul jour férié entre 2 jours ouvrés).
        
        Pour un férié isolé (ex: Ascension jeudi entre mercredi et vendredi ouvrés), la SARA
        ne livre PAS ce jour-là. Donc si l'utilisateur veut une livraison le jour J+1 du férié,
        il doit commander AUJOURD'HUI (deadline = veille du jour ouvré post-férié, avant 11h).
        
        La logique standard des "trous" exige ≥2 jours non-livrables consécutifs et ignore donc
        ces fériés isolés. Cette méthode comble ce trou métier.
        
        Critères d'alerte :
          - Férié isolé (jour-1 ET jour+1 sont des jours ouvrés livrables)
          - Aujourd'hui est le dernier jour ouvré AVANT le férié = deadline de commande
        
        Renseigne `result["feries_isoles_imminents"]` avec la date, le nom du férié, et le jour
        ouvré post-férié pour lequel la commande doit être passée maintenant.
        """
        today=ctx["today"]
        result["feries_isoles_imminents"]=[]
        # On regarde les 7 prochains jours pour repérer un férié isolé imminent
        for offset in range(1,8):
            d_ferie=today+timedelta(days=offset)
            if not is_ferie(d_ferie): continue
            # Vérifier que c'est ISOLÉ : jour-1 ET jour+1 sont ouvrés livrables
            d_avant=d_ferie-timedelta(days=1)
            d_apres=d_ferie+timedelta(days=1)
            if d_avant.weekday()>=5 or is_ferie(d_avant): continue  # avant = non livrable → pas isolé
            if d_apres.weekday()>=5 or is_ferie(d_apres): continue  # après = non livrable → c'est un trou ≥2j, géré ailleurs
            # On n'alerte QUE si aujourd'hui est le dernier jour ouvré avant le férié
            # (= la deadline de commande est aujourd'hui avant 11h pour livraison post-férié)
            if d_avant!=today: continue
            JC=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
            result["feries_isoles_imminents"].append({
                "date_ferie":d_ferie,
                "date_ferie_str":f"{JC[d_ferie.weekday()]} {d_ferie.strftime('%d/%m/%Y')}",
                "nom_ferie":_nom_ferie(d_ferie),
                "date_post_ferie":d_apres,
                "date_post_ferie_str":f"{JC[d_apres.weekday()]} {d_apres.strftime('%d/%m/%Y')}",
                "deadline_str":f"aujourd'hui ({JC[today.weekday()]} {today.strftime('%d/%m')}) avant 11h",
            })

    def _ar_compute_feries_a_venir(self,result,ctx):
        """Section 1 d'analyze_antirupture — Fériés à venir (J+15 à J+30).
        
        Fériés au-delà du cycle Pre_vision (14j) pour anticiper sereinement.
        Pour chaque férié détecté, calcule la durée du trou autour (jours consécutifs non-livrables).
        Initialise aussi la constante JC (libellés jours semaine) dans le ctx pour les sections suivantes.
        """
        today=ctx["today"]
        JC=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
        ctx["JC"]=JC  # exposé pour les sections suivantes (3, 4)
        for offset in range(15,31):
            d=today+timedelta(days=offset)
            if not is_ferie(d): continue
            # Calculer la durée du trou autour de ce férié (jours consécutifs non-livrables)
            trou_start=d
            while trou_start>today and (trou_start-timedelta(days=1)).weekday()>=5 or (trou_start>today and is_ferie(trou_start-timedelta(days=1))):
                trou_start-=timedelta(days=1)
                if trou_start<=today: break
            trou_end=d
            while trou_end.weekday()>=5 or is_ferie(trou_end):
                nxt=trou_end+timedelta(days=1)
                if nxt.weekday()>=5 or is_ferie(nxt):
                    trou_end=nxt
                else: break
            duree=(trou_end-trou_start).days+1
            result["feries_a_venir"].append({
                "date":d,
                "date_str":f"{JC[d.weekday()]} {d.strftime('%d/%m')}",
                "nom":_nom_ferie(d),
                "j_restants":offset,
                "duree_trou":duree,
            })
        result["feries_a_venir"].sort(key=lambda x:x["date"])  # tri chronologique

    def _ar_compute_trous(self,result,ctx):
        """Section 2 d'analyze_antirupture — Détection des trous dans les 14 jours du cycle.
        
        Un "trou" = >= 2 jours consécutifs non-livrables (weekend, weekend+férié, double férié, etc.)
        Pour chaque trou : indices début/fin, dates, durée, dernier jour livrable AVANT, premier APRÈS.
        Expose `trous` et `projection` dans le ctx pour les sections 3, 5, 6, 7bis qui en dépendent.
        
        FIL ROUGE : un jour FORCÉ par Bidou (déclaration explicite via AntiRuptureDlg que la
        livraison aura bien lieu ce jour-là malgré que ce soit normalement non-livrable) est
        RELU ici et exclu du trou. La décision de forçage écrite dans FORCAGE_FILE est donc
        respectée par toutes les alertes en aval.

        EXTRACTION CHIRURGICALE (Étape 3, 27/05/2026) : la logique de détection des trous est
        désormais dans `districarb_core.trous.detecter_trous()`. Comportement strictement
        identique. Les objets `Trou` retournés supportent l'accès par clé (compatibilité dict),
        donc les ~5 sites de consommation en aval (sections 3, 5, 6, 7bis) restent inchangés.
        """
        proj14=ctx["proj14"]
        projection=proj14.get("projection",[])
        trous=detecter_trous(projection,is_date_forcee=is_date_forcee)
        result["trous"]=trous
        # Expose pour sections suivantes (pattern Context Object)
        ctx["trous"]=trous
        ctx["projection"]=projection

    def _ar_compute_coherences_jour_non_livrable(self,result,ctx):
        """Section 3a d'analyze_antirupture — Cohérences : commandes saisies sur jour non-livrable.
        
        Détecte les forecasts qui contiennent une livraison sur un samedi/dimanche/férié.
        Cas opérationnel : SARA ne livre PAS ces jours-là, donc une commande saisie sur ces dates
        ne peut pas être honorée par le transporteur. Bidou doit décaler vers un jour ouvré.
        """
        forecasts=ctx["forecasts"]
        JC=ctx["JC"]
        for fc in forecasts:
            fc_d=fc.get("date")
            if not fc_d: continue
            total=sf(fc.get("livr_sp",0))+sf(fc.get("livr_go",0))+sf(fc.get("livr_gnr",0))
            if total>0 and (fc_d.weekday()>=5 or is_ferie(fc_d)):
                # Si Bidou a déclaré cette date comme exception ASSUMÉE (livraison
                # exceptionnelle volontaire, ex. décalage SARA Pentecôte→samedi),
                # ce n'est plus une incohérence : on ne la signale pas.
                if is_date_forcee(fc_d):
                    continue
                raison="weekend" if fc_d.weekday()>=5 and not is_ferie(fc_d) else (
                    f"f\u00e9ri\u00e9 ({_nom_ferie(fc_d)})" if is_ferie(fc_d) else "weekend")
                result["incoherences_jour_non_livrable"].append({
                    "date":fc_d,
                    "date_str":f"{JC[fc_d.weekday()]} {fc_d.strftime('%d/%m')}",
                    "raison":raison,
                    "volume_total":total,
                    "sp":sf(fc.get("livr_sp",0)),
                    "go":sf(fc.get("livr_go",0)),
                    "gnr":sf(fc.get("livr_gnr",0)),
                })

    def _ar_compute_conformite_sara(self,result,ctx):
        """Section 3b d'analyze_antirupture — Conformité commande SARA (tranches 1000 L, min 2000 L).
        
        SARA n'accepte pas les commandes qui ne sont pas en tranches de 1000 L, ni en dessous de
        2000 L par carburant. Signale les anomalies pour que Bidou ajuste avant la deadline 11h.
        """
        forecasts=ctx["forecasts"]
        JC=ctx["JC"]
        for fc in forecasts:
            fc_d=fc.get("date")
            if not fc_d: continue
            anomalies_carb=[]
            for carb,key in [("SP","livr_sp"),("GO","livr_go"),("GNR","livr_gnr")]:
                v=sf(fc.get(key,0))
                if v<=0: continue  # pas de commande, on ignore
                if 0<v<2000:
                    anomalies_carb.append(f"{carb} {int(v)}L (< minimum 2000L)")
                elif int(v)%1000!=0:
                    anomalies_carb.append(f"{carb} {int(v)}L (non multiple de 1000L)")
            if anomalies_carb:
                result["livraisons_non_conformes"].append({
                    "date":fc_d,
                    "date_str":f"{JC[fc_d.weekday()]} {fc_d.strftime('%d/%m')}",
                    "anomalies":anomalies_carb,
                })

    def _ar_compute_saisies_et_reports(self,result,ctx):
        """Sections 3c, 3d, 3e d'analyze_antirupture — Vérifications physiques sur stocks et livraisons.
        
        3c : Saisies physiquement impossibles
            Vente prévue D7/D8/D9 > (stock matin + livraison) - plancher physique
            Plancher : 500L SP/GO, 250L GNR (sous ce seuil la pompe ne distribue plus).
        3d : Livraisons à reporter
            Cuve déborde à 6h (stock matin + livraison > capacité). On calcule l'heure recommandée
            de report basée sur la vitesse de vente C1 (6h-13h).
            Réglementaire critique : livrer une cuve qui déborde = très grave côté Douane.
        3e : Ruptures projetées
            La projection_14j a déjà plafonné les ventes par le plancher physique. On agrège ici
            les jours où ce plafonnement signale une rupture, avec manque à gagner cumulé.
        """
        forecasts=ctx["forecasts"]
        today=ctx["today"]
        JC=ctx["JC"]
        projection=ctx["projection"]
        CAPA=ctx["CAPA"]
        hist_data=ctx["hist_data"]
        # 3c — Saisies physiquement impossibles
        for fc in forecasts:
            fc_d=fc.get("date")
            if not fc_d: continue
            if fc_d<today: continue
            stm_sp=sf(fc.get("stm_sp",0))
            stm_go=sf(fc.get("stm_go",0))
            stm_gnr=sf(fc.get("stm_gnr",0))
            if stm_sp==0 and stm_go==0 and stm_gnr==0: continue
            for carb,stm_v,d_key in [("sp",stm_sp,"sp"),("go",stm_go,"go"),("gnr",stm_gnr,"gnr")]:
                d_saisi=sf(fc.get(d_key,0))
                e_saisi=sf(fc.get(f"livr_{carb}",0))
                dispo=stm_v+e_saisi
                plancher=PLANCHER_PHYSIQUE[carb]
                vente_max=max(0,dispo-plancher)
                if d_saisi>vente_max+1:  # +1 tolérance arrondi
                    result["saisies_physiquement_impossibles"].append({
                        "date":fc_d,
                        "date_str":f"{JC[fc_d.weekday()]} {fc_d.strftime('%d/%m')}",
                        "carburant":carb.upper(),
                        "stock_matin":int(stm_v),
                        "livraison":int(e_saisi),
                        "dispo":int(dispo),
                        "plancher":plancher,
                        "vente_saisie":int(d_saisi),
                        "vente_max_possible":int(vente_max),
                        "exces":int(d_saisi-vente_max),
                    })
        # 3d — Livraisons à reporter (cuve déborde à 6h)
        # AUTO-RÉSOLUTION 21/05/2026 : si la livraison RÉELLE pour cette date a déjà été
        # saisie dans LIVRAISON_FILE, l'alerte n'a plus de sens (le camion est arrivé/parti,
        # le risque de débordement est passé/géré). Sans ce filtre, l'alerte revient à chaque
        # refresh tant que Pre_vision contient une ligne livraison prévue pour cette date.
        try: livrs_saisies=load_json(LIVRAISON_FILE) or {}
        except Exception: livrs_saisies={}
        for fc in forecasts:
            fc_d=fc.get("date")
            if not fc_d: continue
            if fc_d<today: continue
            # Si livraison du jour déjà saisie pour cette date, skip (l'alerte est résolue de facto)
            try:
                key_livr=fc_d.strftime("%d/%m/%y")
                if key_livr in livrs_saisies:
                    entry=livrs_saisies[key_livr]
                    # Saisie effective (non vide) : skip
                    if isinstance(entry,dict) and (sf(entry.get("sp",0))+sf(entry.get("go",0))+sf(entry.get("gnr",0)))>0:
                        continue
            except Exception as _e: _log_silent_err(exc=_e)
            for carb in ("sp","go","gnr"):
                stm_v=sf(fc.get(f"stm_{carb}",0))
                livr=sf(fc.get(f"livr_{carb}",0))
                if livr<=0: continue
                if stm_v==0: continue  # pas de stock matin saisi = pas d'info fiable
                capa=CAPA.get(carb,40000)
                surplus=stm_v+livr-capa
                if surplus<=0: continue  # tout rentre, OK
                wd=fc_d.weekday()
                ref_c1=avg_ventes_n_caisses(hist_data,wd,1,nb_derniers=4) or {} if hist_data else {}
                ventes_c1_total=sf(ref_c1.get(carb,0))
                if ventes_c1_total<=0:
                    heures_attente=None;heure_recommandee=None;report_au_lendemain=False
                else:
                    vitesse_lh=ventes_c1_total/7.0  # C1 = 6h→13h = 7h
                    heures_attente=surplus/vitesse_lh if vitesse_lh>0 else None
                    if heures_attente is not None:
                        # Arrondi à la 30 min supérieure pour marge
                        heures_attente=((heures_attente*2)+0.5)//1/2
                        # SI > 7h d'attente, la C1 entière ne suffit pas à libérer assez de place.
                        # Inutile d'afficher une heure aberrante du genre "118h30" : on signale
                        # plutôt qu'il faut décaler la livraison au lendemain (ou commander moins).
                        if heures_attente>7:
                            heure_recommandee=None
                            report_au_lendemain=True
                        else:
                            heure_recommandee=6.0+heures_attente
                            report_au_lendemain=False
                    else:
                        heure_recommandee=None;report_au_lendemain=False
                result["livraisons_a_reporter"].append({
                    "date":fc_d,
                    "date_str":f"{JC[wd]} {fc_d.strftime('%d/%m')}",
                    "carburant":carb.upper(),
                    "stock_matin":int(stm_v),
                    "livraison":int(livr),
                    "capacite":capa,
                    "surplus":int(surplus),
                    "ventes_c1_moy":int(ventes_c1_total),
                    "heures_attente":heures_attente,
                    "heure_recommandee":heure_recommandee,
                    "report_au_lendemain":report_au_lendemain,
                })
        # 3d-bis — Livraisons MARGE TENDUE (cuve presque pleine mais ça rentre).
        # Bidou m'a indiqué que dans son fichier Excel, il marque "Attention" dès que la marge
        # de cuve restante après livraison est < 4 000 L. C'est plus prudent que le seuil de
        # dépassement strict (3d) : ça anticipe le risque qu'une C1 plus calme que prévu laisse
        # la cuve trop pleine.
        SEUIL_MARGE_TENDUE=4000
        # FIX timing : si la livraison du jour est DÉJÀ saisie dans LIVRAISON_FILE, l'événement
        # est passé, on ne re-pop plus l'alerte préventive. Le bilan a posteriori est géré
        # séparément par BilanLivraisonDlg déclenchée après LivraisonDialog.
        try: _livrs_saisies=load_json(LIVRAISON_FILE) or {}
        except Exception: _livrs_saisies={}
        for fc in forecasts:
            fc_d=fc.get("date")
            if not fc_d: continue
            if fc_d<today: continue
            # Skip si livraison déjà saisie pour ce jour
            day_key_fr=fc_d.strftime("%d/%m/%y")
            saisie=_livrs_saisies.get(day_key_fr,{})
            if isinstance(saisie,dict) and not saisie.get("none") and sum(
                sf(saisie.get(k,0)) for k in ("sp","go","gnr"))>0:
                continue
            # FIX MIX A+B (25/05/2026) : skip si jour non-livrable ET non forcé.
            # C'est un résidu probable de Pre_vision (cycle S1/S2 pas encore basculé,
            # date oubliée dans le fichier, etc.) — l'anti-rupture s'en charge déjà via
            # incoherences_jour_non_livrable. Pas besoin de double signal contradictoire :
            # "tu seras tendu" (marge) vs "ce jour est non-livrable, replanifie" (anti-rupture).
            if (fc_d.weekday()>=5 or is_ferie(fc_d)) and not is_date_forcee(fc_d):
                continue
            for carb in ("sp","go","gnr"):
                stm_v=sf(fc.get(f"stm_{carb}",0))
                livr=sf(fc.get(f"livr_{carb}",0))
                if livr<=0: continue
                if stm_v==0: continue
                capa=CAPA.get(carb,40000)
                marge_restante=capa-(stm_v+livr)
                if marge_restante<=0: continue  # déjà dans 3d (dépassement strict)
                if marge_restante>=SEUIL_MARGE_TENDUE: continue  # marge confortable, on n'alerte pas
                wd=fc_d.weekday()
                result.setdefault("livraisons_marge_tendue",[]).append({
                    "date":fc_d,
                    "date_str":f"{JC[wd]} {fc_d.strftime('%d/%m')}",
                    "carburant":carb.upper(),
                    "stock_matin":int(stm_v),
                    "livraison":int(livr),
                    "capacite":capa,
                    "marge_restante":int(marge_restante),
                })
        # 3e — Ruptures projetées (ventes plafonnées par plancher physique)
        for d_data in projection:
            vp=d_data.get("ventes_perdues",{}) or {}
            if not any(vp.values()): continue
            for carb in ["sp","go","gnr"]:
                vp_c=vp.get(carb,0)
                if vp_c<=0: continue
                fc_d=d_data["date"]
                wd=fc_d.weekday()
                result["ruptures_projetees"].append({
                    "date":fc_d,
                    "date_str":f"{JC[wd]} {fc_d.strftime('%d/%m')}",
                    "carburant":carb.upper(),
                    "vente_voulue":int(d_data.get(f"ventes_voulues_{carb}",0)),
                    "vente_reelle":int(d_data.get(f"ventes_{carb}",0)),
                    "manque_a_gagner_l":int(vp_c),
                })

    def _ar_compute_ventes_irrealistes(self,result,ctx):
        """Section 4 d'analyze_antirupture — Ventes prévisionnelles irréalistes (D7/D8/D9).
        
        Compare D7/D8/D9 saisies vs moyenne historique du jour de semaine. Seuils ADAPTATIFS :
          - SP/GO en période normale : +/-30% jaune, +/-50% rouge
          - SP/GO en période Observatoire des Prix (25 -> 5) : +/-50% jaune, +/-80% rouge
            (les annonces de prix au 1er créent des variations atypiques de ventes)
          - GNR en toutes périodes : +/-60% jaune, +/-100% rouge
            (carburant à faible volume, naturellement très volatil — pic 3x moyenne possible)
        """
        forecasts=ctx["forecasts"]
        avg_wd=ctx["avg_wd"]
        JC=ctx["JC"]
        for fc in forecasts:
            fc_d=fc.get("date")
            if not fc_d: continue
            wd=fc_d.weekday()
            ref=avg_wd.get(wd)
            if not ref: continue
            obs_period=is_periode_observatoire(fc_d)
            for carb in ["sp","go","gnr"]:
                saisi=sf(fc.get(carb,0))
                if saisi<=0: continue
                attendu=ref.get(carb,0)
                if attendu<=0: continue
                ecart_pct=(saisi-attendu)/attendu*100
                # Choix des seuils selon carburant et contexte
                if carb=="gnr":
                    seuil_jaune,seuil_rouge=60,100
                elif obs_period:
                    seuil_jaune,seuil_rouge=50,80
                else:
                    seuil_jaune,seuil_rouge=30,50
                if abs(ecart_pct)>seuil_jaune:
                    result["ventes_irrealistes"].append({
                        "date":fc_d,
                        "date_str":f"{JC[wd]} {fc_d.strftime('%d/%m')}",
                        "carburant":carb.upper(),
                        "saisi":saisi,
                        "attendu":round(attendu),
                        "ecart_pct":round(ecart_pct),
                        "severity":"rouge" if abs(ecart_pct)>seuil_rouge else "jaune",
                        "contexte":"observatoire" if obs_period else ("gnr_volatil" if carb=="gnr" else "normal"),
                    })

    def _ar_compute_ruptures_dans_trou(self,result,ctx):
        """Section 5 d'analyze_antirupture — Calcul des manques pendant chaque trou.
        
        Pour chaque trou (>= 2j non-livrables) et chaque carburant :
          stock fin du dernier jour livrable - ventes pendant le trou - sécurité (0.5j post-trou)
          = manque (si positif, alerte commande).
        Ajoute aussi une deadline de commande (veille du dernier jour livrable avant 11h).
        """
        trous=ctx["trous"]
        projection=ctx["projection"]
        avg_wd=ctx["avg_wd"]
        SEUIL_SECU_JOURS=ctx["SEUIL_SECU_JOURS"]
        debug_log(f"Trous d\u00e9tect\u00e9s : {len(trous)}")
        for trou in trous:
            if trou["last_livrable_idx"] is None: continue
            day_avant=projection[trou["last_livrable_idx"]]
            debug_log(f"  Trou {trou['start_date']} \u2192 {trou['end_date']} : last_livrable_idx={trou['last_livrable_idx']} ({day_avant['date']}), next_idx={trou['next_livrable_idx']}")
            extension_jours=trou.get("extension_jours",[]) or []
            for carb in ["sp","go","gnr"]:
                # Stock fin de journée du dernier jour livrable
                stock_apres_dernier=sf(day_avant.get(f"stock_{carb}",0))
                # Ventes pendant le trou (la projection a déjà calculé avec profil dimanche pour fériés)
                # + ventes des jours d'extension hors projection via pattern hebdomadaire avg_wd
                ventes_trou=sum(sf(projection[idx].get(f"ventes_{carb}",0))
                                for idx in range(trou["start_idx"],trou["end_idx"]+1))
                ventes_trou+=sum(avg_wd.get(d.weekday(),{}).get(carb,0) for d in extension_jours)
                # Besoin sécurité : 0,5 jour de ventes au sortir du trou
                if trou["next_livrable_idx"] is not None:
                    ventes_lendemain=sf(projection[trou["next_livrable_idx"]].get(f"ventes_{carb}",0))
                else:
                    ventes_lendemain=avg_wd.get((trou["end_date"]+timedelta(days=1)).weekday(),{}).get(carb,0)
                secu=ventes_lendemain*SEUIL_SECU_JOURS
                stock_min_requis=ventes_trou+secu
                manque=stock_min_requis-stock_apres_dernier
                debug_log(f"    {carb.upper()}: stock_fin_jour_avant={stock_apres_dernier:.0f}, ventes_trou={ventes_trou:.0f}, secu={secu:.0f}, besoin={stock_min_requis:.0f}, MANQUE={manque:.0f}")
                if manque>0:
                    JOURS=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
                    # Deadline commande = veille du dernier jour livrable avant 11h
                    deadline=day_avant["date"]-timedelta(days=1)
                    while deadline.weekday()>=5 or is_ferie(deadline):
                        deadline-=timedelta(days=1)
                    result["ruptures_dans_trou"].append({
                        "carburant":carb.upper(),
                        "trou_str":f"{trou['start_date'].strftime('%d/%m')} \u2192 {trou['end_date'].strftime('%d/%m')}",
                        "trou_duree":trou["duree"],
                        "terme":qualifier_trou({"start_date":trou["start_date"],"duree":trou["duree"]}),
                        "trou_start":trou["start_date"],
                        "trou_end":trou["end_date"],
                        "stock_apres_dernier":round(stock_apres_dernier),
                        "ventes_trou":round(ventes_trou),
                        "secu":round(secu),
                        "manque":round(manque),
                        "deadline":deadline,
                        "deadline_str":f"{JOURS[deadline.weekday()]} {deadline.strftime('%d/%m')} avant 11h",
                        "last_livr_jour":day_avant["date"],
                        "last_livr_jour_str":f"{JOURS[day_avant['date'].weekday()]} {day_avant['date'].strftime('%d/%m')}",
                    })

    def _read_ob(self):
        wb=self._open("objectif")
        if not wb: return {"st":"miss"}
        o={"st":"ok"}
        try:
            ws=wb["Objectif mensuel"]
            o["obj_ca"]=sf(self._c(ws,"E2"));o["enc_ca"]=sf(self._c(ws,"E14"))
            o["enc_dc"]=sf(self._c(ws,"E11"));o["taux"]=sf(self._c(ws,"E17"))
            o["ca_jour"]=sf(self._c(ws,"E16"));o["balance"]=sf(self._c(ws,"E7"))
        except Exception as e:
            o["st"]="err"
        finally:
            try:wb.close()
            except Exception as _e: _log_silent_err(exc=_e)
        return o

    def _read_li(self):
        wb=self._open("litrage")
        if not wb: return {"st":"miss"}
        o={"st":"ok"}
        try: ws=wb["Feuil1"];o["rows"]=ws.max_row
        except: o["st"]="err"
        finally:
            try:wb.close()
            except Exception as _e: _log_silent_err(exc=_e)
        return o

    def _read_hist(self,full=False):
        """Lit l'historique LITRAGE.
        - full=False (défaut) : 1500 dernières lignes ≈ 55 jours pour démarrage rapide
        - full=True : tout le fichier (155k+ lignes) pour génération de rapports historiques"""
        wb=self._open("litrage")
        if not wb: return []
        days=[]
        try:
            ws=wb["Feuil1"];rows={}
            if full:
                # Lecture complète : tout le fichier
                for i,row in enumerate(ws.iter_rows(values_only=True),1):
                    rows[i]=list(row)
            else:
                # Lecture partielle : 1500 dernières lignes
                for i,row in enumerate(ws.iter_rows(values_only=True),1):
                    if i>=max(1,ws.max_row-1500): rows[i]=list(row)
            ttal_rows=[i for i,row in rows.items() if len(row)>11 and row[11] and 'TTAL' in str(row[11])]
            for tr in ttal_rows:
                day={};r1=rows.get(tr+1,[]);r2=rows.get(tr+2,[]);r3=rows.get(tr+3,[])
                r4=rows.get(tr+4,[]);r5=rows.get(tr+5,[]);r7=rows.get(tr+7,[])
                day["piste"]=sf(r1[4] if len(r1)>4 else 0)
                day["sp"]=sf(r2[1] if len(r2)>1 else 0);day["cb"]=sf(r2[4] if len(r2)>4 else 0)
                day["esp"]=sf(r2[10] if len(r2)>10 else 0)
                day["go"]=sf(r3[1] if len(r3)>1 else 0);day["cp"]=sf(r3[4] if len(r3)>4 else 0)
                day["gnr"]=sf(r4[1] if len(r4)>1 else 0)
                day["litrage"]=sf(r5[1] if len(r5)>1 else 0);day["bout"]=sf(r5[4] if len(r5)>4 else 0)
                day["total"]=sf(r7[10] if len(r7)>10 else 0)
                for scan in range(tr-1,max(tr-40,0),-1):
                    sr=rows.get(scan,[])
                    if sr and sr[0]:
                        s0=str(sr[0]).lower()
                        if any(d in s0 for d in ['lun','mar','merc','jeu','ven','sam','dim']):
                            day["label"]=str(sr[0]);break
                # Parser caisse par caisse (caisse-1, caisse-2, caisse-3)
                day["caisses"]={}
                for scan in range(tr-1,max(tr-40,0),-1):
                    sr=rows.get(scan,[])
                    if sr and len(sr)>11 and sr[11] and 'caisse-' in str(sr[11]):
                        cnum=str(sr[11]).replace('caisse-','').strip()
                        # Lire les valeurs : +1=header, +2=GNR, +3=SP, +4=GO, +6=TOTAL
                        cgnr=rows.get(scan+2,[]);csp=rows.get(scan+3,[]);cgo=rows.get(scan+4,[])
                        ctotal=rows.get(scan+6,[])
                        c_data={
                            "gnr":sf(cgnr[2] if len(cgnr)>2 else 0),
                            "sp":sf(csp[2] if len(csp)>2 else 0),
                            "go":sf(cgo[2] if len(cgo)>2 else 0),
                            "litrage":sf(ctotal[2] if len(ctotal)>2 else 0),
                            "piste_eur":sf(ctotal[4] if len(ctotal)>4 else 0),
                            "cb":sf(cgnr[6] if len(cgnr)>6 else 0),
                            "cp":sf(cgnr[7] if len(cgnr)>7 else 0),
                            "bout":sf(cgnr[9] if len(cgnr)>9 else 0),
                            # Écarts: piste (différence caisse-théorique piste) et boutique (colonne J = théorique-caisse)
                            "ecart_piste":sf(rows.get(scan+8,[])[4] if len(rows.get(scan+8,[]))>4 else 0),
                            "ecart_bout":sf(cgnr[9] if len(cgnr)>9 else 0)-sf(cgnr[10] if len(cgnr)>10 else 0),
                        }
                        # Écart net cumulé
                        c_data["ecart_net"]=c_data["ecart_piste"]+c_data["ecart_bout"]
                        day["caisses"][cnum]=c_data
                # Détection EN COURS : compter les caisses RÉELLEMENT SAISIES.
                # Le template Excel pré-remplit toujours les 3 blocs caisse-1/2/3 vides,
                # donc la simple présence du label ne prouve pas la saisie.
                # Une caisse est "saisie" = au moins UNE valeur d'activité > 0 :
                #   litrage carburant OU CB OU CP OU boutique OU CA piste.
                # Couvre :
                #   - jour normal (litrage>0)
                #   - rupture carburant / nuit calme (litrage=0 mais CB/CP/bout>0)
                #   - template vierge non encore saisi (tout=0 → PAS saisie)
                caisses_remplies=sum(
                    1 for c in day["caisses"].values()
                    if sf(c.get("litrage",0))>0 or sf(c.get("cb",0))>0
                    or sf(c.get("cp",0))>0 or sf(c.get("bout",0))>0
                    or sf(c.get("piste_eur",0))>0
                )
                day["nb_caisses"]=caisses_remplies
                # En cours si moins de 3 caisses saisies ET au moins 1 caisse présente
                # (ne pas exiger litrage>0 : un jour de rupture carburant avec boutique active
                # est un jour à compter comme en cours).
                if caisses_remplies>0 and caisses_remplies<3:
                    day["en_cours"]=True
                # Activité boutique = somme des CB/CP/boutique des caisses saisies
                act_bout=sum(sf(c.get("cb",0))+sf(c.get("cp",0))+sf(c.get("bout",0)) for c in day["caisses"].values())
                # Ajouter le jour si AU MOINS UN signal d'activité réelle :
                # - litrage carburant > 0 (jour normal)
                # - CA piste > 0
                # - activité boutique > 0 (cas rupture carburant mais boutique active)
                # Pas de "caisses_remplies>0" tout seul : une caisse vide (label présent mais
                # toutes valeurs à 0) ne suffit pas. Il faut une activité réelle quelque part.
                if day["litrage"]>0 or day["piste"]>0 or act_bout>0:
                    days.append(day)
        except Exception as e: print(f"[ERR hist] {e}");traceback.print_exc()
        finally:
            try:wb.close()
            except Exception as _e: _log_silent_err(exc=_e)
        return days

    def _read_alerts(self):
        wb=self._open_styled("objectif")
        if not wb: return {"cp_pending":[],"clients_impayes":[],"cp_total":0,"cli_total":0,"by_client":{}}
        a={"cp_pending":[],"clients_impayes":[],"cp_total":0,"cli_total":0,"by_client":{}}
        try:
            # CP en attente
            ws=wb['Encaissement CB-CP-CS']
            today_d=date.today()
            # Total mois encaissé : lire D3 directement (calculé par Excel via SUMIFS sur date paiement)
            a["mois_total_cbcpcs"]=sf(self._c(ws,"D3"))
            # Ventilation par type : scanner les lignes payé=o avec date PAIEMENT (col L) sur le mois
            mois_cb=0;mois_cp=0;mois_cs=0
            for row in ws.iter_rows(min_row=6,max_row=ws.max_row,values_only=False):
                cells={c.column_letter:c for c in row}
                cv=cells.get('C');dv=cells.get('D');iv=cells.get('I');lv=cells.get('L')
                paye=str(iv.value).strip().lower() if iv and iv.value else ""
                # Date paiement (colonne L) — date où le montant est crédité au compte
                pay_date=lv.value if lv else None
                rd=None
                if hasattr(pay_date,'date'): rd=pay_date.date()
                elif hasattr(pay_date,'year'): rd=pay_date
                in_mois=rd and rd.month==today_d.month and rd.year==today_d.year
                if in_mois and paye=="o":
                    typ=str(cv.value).strip().upper() if cv and cv.value else ""
                    mt=sf(dv.value if dv else 0)
                    if typ=="CB": mois_cb+=mt
                    elif typ=="CP": mois_cp+=mt
                    elif typ=="CS": mois_cs+=mt
                # Alertes CP en attente : date facture col B
                bv=cells.get('B').value if cells.get('B') else None
                if cv and cv.value=='CP' and paye=='n':
                    mt=sf(cells.get('D').value if cells.get('D') else 0)
                    ech=cells.get('L').value if cells.get('L') else None
                    dt_s=bv.strftime("%d/%m") if hasattr(bv,'strftime') else str(bv)[:5]
                    ret=0
                    if hasattr(ech,'date'): ret=(date.today()-ech.date()).days
                    elif hasattr(ech,'toordinal'): ret=(date.today()-ech).days
                    a["cp_pending"].append({"date":dt_s,"montant":mt,"retard":ret})
            a["cp_total"]=sum(c["montant"] for c in a["cp_pending"])
            a["mois_cb"]=mois_cb;a["mois_cp"]=mois_cp;a["mois_cs"]=mois_cs
            # Encaissements à venir : scan CB/CS/CP payé=n avec date col J
            a["enc_pending"]=[]
            for row in ws.iter_rows(min_row=6,max_row=ws.max_row,values_only=False):
                cells={c.column_letter:c for c in row}
                cv=cells.get('C');dv=cells.get('D');iv=cells.get('I');jv=cells.get('J');lv=cells.get('L')
                paye=str(iv.value).strip().lower() if iv and iv.value else ""
                if paye!="n": continue
                mt=sf(dv.value if dv else 0)
                if mt<=0: continue
                typ=str(cv.value).strip().upper() if cv and cv.value else "?"
                prel=jv.value if jv else None
                if prel is None and lv: prel=lv.value
                prel_d=None
                if hasattr(prel,'date'): prel_d=prel.date()
                elif hasattr(prel,'year'): prel_d=prel
                reste=(prel_d-date.today()).days if prel_d else None
                a["enc_pending"].append({"cat":typ,"nom":typ,"montant":mt,"ech":prel_d,"reste":reste})
            # Espèces du mois (Encaissements dep. Exp) — lire D3 directement
            mois_esp=0
            try:
                ws_e=wb['Encaissements dep. Exp']
                mois_esp=sf(self._c(ws_e,"D3"))
                # Scan dépôts espèces payé=n (non encore crédités)
                for row in ws_e.iter_rows(min_row=6,max_row=ws_e.max_row,values_only=False):
                    cells={c.column_letter:c for c in row}
                    cv=cells.get('C');dv=cells.get('D');iv=cells.get('I');bv=cells.get('B');jv=cells.get('J');lv=cells.get('L')
                    paye=str(iv.value).strip().lower() if iv and iv.value else ""
                    if paye!="n": continue
                    mt=sf(dv.value if dv else 0)
                    if mt<=0: continue
                    nom=str(cv.value)[:40] if cv and cv.value else "D\u00e9p\u00f4t express"
                    # Date prélèvement : J puis L puis B (date dépôt)
                    prel=jv.value if jv else None
                    if prel is None and lv: prel=lv.value
                    if prel is None and bv: prel=bv.value
                    prel_d=None
                    if hasattr(prel,'date'): prel_d=prel.date()
                    elif hasattr(prel,'year'): prel_d=prel
                    reste=(prel_d-date.today()).days if prel_d else None
                    a["enc_pending"].append({"cat":"ESP","nom":nom,"montant":mt,"ech":prel_d,"reste":reste})
            except Exception as _e: _log_silent_err(exc=_e)
            a["mois_esp"]=mois_esp
            # Clients impayés — BUG FIX: lire D3 pour le total au lieu de recalculer
            ws2=wb['Clients en compte']
            # D3 = total clients encaissés au compte ce mois (calculé par Excel)
            a["mois_cli"]=sf(self._c(ws2,"D3"))
            # Scanner les lignes payé=n pour alimenter les alertes "clients en retard"
            for row in ws2.iter_rows(min_row=6,max_row=ws2.max_row,values_only=False):
                cells={c.column_letter:c for c in row}
                cv=cells.get('C');iv=cells.get('I')
                paye=str(iv.value).strip().lower() if iv and iv.value else ""
                if cv and cv.value and paye=='n':
                    dt=cells.get('B').value if cells.get('B') else None
                    nm=str(cv.value);mt=sf(cells.get('D').value if cells.get('D') else 0)
                    age=0
                    if hasattr(dt,'date'): age=(date.today()-dt.date()).days
                    elif hasattr(dt,'toordinal'): age=(date.today()-dt).days
                    a["clients_impayes"].append({"nom":nm,"montant":mt,"age":age})
            # cli_total = total clients impayés (somme lignes payé=n)
            a["cli_total"]=sum(c["montant"] for c in a["clients_impayes"])
            # === DÉCAISSEMENTS ===
            # Totaux mensuels (lecture directe des D3 calculés par Excel sur date paiement + payé=o)
            dec_tabs=[("dec_car","D\u00e9caissement Total C."),
                      ("dec_fourn","R\u00e8glements fourniss."),
                      ("dec_fg","Frais g\u00e9n\u00e9raux"),
                      ("dec_div","D\u00e9c.divers"),
                      ("dec_soc","Charges soc - fisc")]
            a["dec_pending"]=[]  # lignes payé=n toutes catégories
            for key,sname in dec_tabs:
                if sname not in wb.sheetnames:
                    a[key]=0;continue
                ws_d=wb[sname]
                a[key]=sf(self._c(ws_d,"D3"))
                # Scanner les lignes payé=n pour alertes
                for row in ws_d.iter_rows(min_row=6,max_row=ws_d.max_row,values_only=False):
                    cells={c.column_letter:c for c in row}
                    iv=cells.get('I');dv=cells.get('D');cv=cells.get('C')
                    bv=cells.get('B');jv=cells.get('J');lv=cells.get('L')
                    paye=str(iv.value).strip().lower() if iv and iv.value else ""
                    if paye!="n": continue
                    mt=sf(dv.value if dv else 0)
                    if mt<=0: continue
                    nom=str(cv.value)[:40] if cv and cv.value else "\u2014"
                    # Date de prélèvement réel : col J en priorité (date paiement effective),
                    # fallback sur col L (échéance) si J vide
                    prel=jv.value if jv else None
                    if prel is None and lv: prel=lv.value
                    prel_d=None
                    if hasattr(prel,'date'): prel_d=prel.date()
                    elif hasattr(prel,'year'): prel_d=prel
                    reste=None
                    if prel_d: reste=(prel_d-date.today()).days
                    a["dec_pending"].append({"cat":sname,"nom":nom,"montant":mt,"ech":prel_d,"reste":reste})
            # Balance D/E calculée par Excel (E7 de l'onglet Objectif mensuel)
            try:
                ws_ob=wb['Objectif mensuel']
                a["balance_de"]=sf(self._c(ws_ob,"E7"))
                a["encours_dc"]=sf(self._c(ws_ob,"B21"))
                a["encours_ec"]=sf(self._c(ws_ob,"H21"))
            except:
                a["balance_de"]=0;a["encours_dc"]=0;a["encours_ec"]=0
            bc={}
            for c in a["clients_impayes"]:
                n=c["nom"]
                if n not in bc: bc[n]={"total":0,"count":0,"max_age":0}
                bc[n]["total"]+=c["montant"];bc[n]["count"]+=1;bc[n]["max_age"]=max(bc[n]["max_age"],c["age"])
            a["by_client"]=bc
            # Scan mots-clés urgents dans les onglets NON déjà traités
            keywords_alert=["rejet","rejet\u00e9","impay\u00e9","saisie","huissier","urgent","erreur","r\u00e9gularisation"]
            a["keyword_alerts"]=[]
            skip_sheets={'Encaissement CB-CP-CS','Clients en compte','Objectif mensuel'}
            for sn in wb.sheetnames:
                if sn in skip_sheets: continue
                ws_k=wb[sn]
                for row in ws_k.iter_rows(min_row=1,max_row=min(ws_k.max_row,200),values_only=False):
                    for cell in row:
                        if cell.value and isinstance(cell.value,str):
                            val_low=cell.value.lower()
                            for kw in keywords_alert:
                                if kw in val_low and cell.column_letter in ('C','E','F','H','K'):
                                    row_cells={c.column_letter:c.value for c in row if c.value is not None}
                                    mt=sf(row_cells.get('D',0))
                                    nom=row_cells.get('C','')
                                    mode=str(row_cells.get('G','') or '').strip()
                                    info=str(row_cells.get('K','') or '').strip()
                                    date_b=row_cells.get('B','')
                                    okko=str(row_cells.get('F','') or '').strip().lower()
                                    paye=str(row_cells.get('I','') or '').strip().lower()
                                    if mt>0 or nom:
                                        a["keyword_alerts"].append({"onglet":sn,"mot":kw,"nom":str(nom)[:25],"montant":mt,"cell":cell.value[:30],"mode":mode,"info":info,"date":date_b,"okko":okko,"paye":paye})
                                    break
        except Exception as e: print(f"[ERR alerts] {e}")
        finally:
            try:wb.close()
            except Exception as _e: _log_silent_err(exc=_e)
        return a

# =============================================================================
class Vignette(ctk.CTkFrame):
    def __init__(self,parent,fd,on_click,on_detail):
        super().__init__(parent,fg_color=C["card"],corner_radius=14,border_width=1,border_color=C["border"],height=180)
        self.fd=fd;self.on_click=on_click;self.on_detail=on_detail;self.pack_propagate(False)
        # Liseré gauche dynamique PLEINE HAUTEUR (plus structurant visuellement)
        self.liseret=ctk.CTkFrame(self,fg_color=fd["color"],corner_radius=0,width=5)
        self.liseret.pack(side="left",fill="y")
        content=ctk.CTkFrame(self,fg_color="transparent");content.pack(side="left",fill="both",expand=True,padx=18,pady=14)
        header=ctk.CTkFrame(content,fg_color="transparent");header.pack(fill="x")
        ctk.CTkLabel(header,text=fd["icon"],font=("Segoe UI Emoji",26),text_color=fd["color"]).pack(side="left",padx=(0,10))
        tf=ctk.CTkFrame(header,fg_color="transparent");tf.pack(side="left",fill="x",expand=True)
        ctk.CTkLabel(tf,text=fd["label"],font=("Segoe UI",16,"bold"),text_color=fd["color"],anchor="w").pack(fill="x")
        ctk.CTkLabel(tf,text=fd["sub"],font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(fill="x")
        # Pastille d'état (plus visible)
        self.pastille=ctk.CTkLabel(header,text="\u25cf",font=("Segoe UI",16),text_color=C["t3"])
        self.pastille.pack(side="right",padx=(4,2))
        self.chevron=ctk.CTkLabel(header,text="\u22ef",font=("Segoe UI",16,"bold"),text_color=C["t3"],cursor="hand2")
        self.chevron.pack(side="right",padx=(4,0))
        self.ind=ctk.CTkFrame(content,fg_color="transparent");self.ind.pack(fill="both",expand=True,pady=(10,0))
        self._bind_click(self)
        # Hover : change fond ET bordure (prend la couleur du liseré → retour visuel franc)
        self.bind("<Enter>",lambda _:self._on_hover(True))
        self.bind("<Leave>",lambda _:self._on_hover(False))
    def _on_hover(self,entering):
        if entering:
            self.configure(fg_color=C["card_h"],border_color=self.liseret.cget("fg_color"))
        else:
            self.configure(fg_color=C["card"],border_color=C["border"])
    def set_status(self,status):
        """status: 'ok' (vert), 'warn' (ambre), 'alert' (rouge)"""
        self._last_status=status  # mémorisé pour exposition API web
        col={"ok":C["green"],"warn":C["amber"],"alert":C["red"]}.get(status,self.fd["color"])
        self.liseret.configure(fg_color=col)
        self.pastille.configure(text_color=col)
    def _bind_click(self,w):
        w.bind("<Button-1>",self._show_menu)
        for ch in w.winfo_children():
            if not isinstance(ch,ctk.CTkButton):
                self._bind_click(ch)
    def _show_menu(self,event):
        popup=ctk.CTkToplevel(self)
        popup.overrideredirect(True);popup.attributes("-topmost",True)
        popup.configure(fg_color=C["card_h"])
        # Position : décalée légèrement pour éviter de masquer le clic
        x=event.x_root-20;y=event.y_root-10
        popup.geometry(f"280x180+{x}+{y}")
        # Cadre principal avec ombre et bordure colorée
        frame=ctk.CTkFrame(popup,fg_color=C["card"],corner_radius=14,border_width=2,border_color=self.fd["color"])
        frame.pack(fill="both",expand=True,padx=2,pady=2)
        # En-tête : icône ronde + label
        hdr=ctk.CTkFrame(frame,fg_color="transparent");hdr.pack(fill="x",padx=14,pady=(14,8))
        ctk.CTkLabel(hdr,text=self.fd["icon"],width=36,height=36,font=("Segoe UI",18,"bold"),text_color="#FFFFFF",fg_color=self.fd["color"],corner_radius=18).pack(side="left")
        txt=ctk.CTkFrame(hdr,fg_color="transparent");txt.pack(side="left",padx=(10,0),fill="both",expand=True)
        ctk.CTkLabel(txt,text=self.fd["label"],font=("Segoe UI",13,"bold"),text_color=C["t1"],anchor="w").pack(anchor="w")
        ctk.CTkLabel(txt,text=self.fd["sub"],font=("Segoe UI",10),text_color=C["t3"],anchor="w").pack(anchor="w")
        # Séparateur fin
        ctk.CTkFrame(frame,fg_color=C["border"],height=1).pack(fill="x",padx=14,pady=(2,8))
        def do_excel():
            popup.destroy();self.on_click(self.fd["key"])
        def do_detail():
            popup.destroy();self.on_detail(self.fd["key"])
        # Boutons en pilules avec rond coloré (style cohérent header)
        for icon,txt2,cmd,col in [
            ("\U0001f4ca","Voir les détails",do_detail,self.fd["color"]),
            ("\U0001f4c2","Ouvrir dans Excel",do_excel,"#5C6BC0")]:
            pill=ctk.CTkFrame(frame,fg_color=C["card_h"],corner_radius=20,height=36)
            pill.pack(fill="x",padx=12,pady=3);pill.pack_propagate(False)
            circ=ctk.CTkLabel(pill,text=icon,width=26,height=26,font=("Segoe UI",12,"bold"),text_color="#FFFFFF",fg_color=col,corner_radius=13)
            circ.pack(side="left",padx=(5,8),pady=5)
            lbl=ctk.CTkLabel(pill,text=txt2,font=("Segoe UI",11,"bold"),text_color=C["t1"])
            lbl.pack(side="left")
            for w in [pill,circ,lbl]:
                w.bind("<Button-1>",lambda e,c=cmd:c())
                w.bind("<Enter>",lambda e,p=pill:p.configure(fg_color=C["panel"]))
                w.bind("<Leave>",lambda e,p=pill:p.configure(fg_color=C["card_h"]))
        popup.bind("<FocusOut>",lambda _:popup.destroy())
        popup.after(100,popup.focus_set)
    def set_data(self,lines):
        # Mémoriser les lines pour exposition via API web (capturé sans toucher la logique).
        # Format conservé tel quel : list of (label, value, color_hex).
        self._last_lines=list(lines) if lines else []
        for w in self.ind.winfo_children(): w.destroy()
        for label,value,color in lines:
            r=ctk.CTkFrame(self.ind,fg_color="transparent");r.pack(fill="x",pady=2)
            # Label : petit, uppercase, spaced — style éditorial qui laisse la vedette à la valeur
            ctk.CTkLabel(r,text=str(label).upper(),font=("Segoe UI",10,"bold"),text_color=C["t2"],anchor="w").pack(side="left")
            # Valeur : taille doublée, devient la star visuelle de la ligne
            ctk.CTkLabel(r,text=value,font=(FONT_NUM,17,"bold"),text_color=color or C["t1"],anchor="e").pack(side="right")
            self._bind_click(r)

class Section(ctk.CTkFrame):
    def __init__(self,parent,title,accent):
        super().__init__(parent,fg_color="transparent")
        # Mémorisation pour exposition API web (pas de re-calcul, simple miroir)
        self._title=title
        self._accent=accent
        self._rows=[]  # list of dicts {kind, label, value, color, big, tr_arrow, tr_color}
        h=ctk.CTkFrame(self,fg_color="transparent");h.pack(fill="x",pady=(0,5))
        ctk.CTkFrame(h,fg_color=accent,width=3,height=14,corner_radius=0).pack(side="left",padx=(0,7))
        ctk.CTkLabel(h,text=title.upper(),font=("Segoe UI",10,"bold"),text_color=C["t2"]).pack(side="left")
        self.box=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=10,border_width=1,border_color=C["border"])
        self.box.pack(fill="x",pady=(0,12))
    def clear(self):
        self._rows=[]  # reset miroir
        for w in self.box.winfo_children(): w.destroy()
    def row(self,label,value,color=None,big=False,tr=""):
        # Capture pour API web
        tr_arrow=None;tr_color=None
        if tr:
            try: tr_arrow,tr_color=tr
            except Exception: pass
        self._rows.append({"kind":"row","label":str(label),"value":str(value),
                           "color":str(color) if color else None,"big":bool(big),
                           "tr_arrow":str(tr_arrow) if tr_arrow else None,
                           "tr_color":str(tr_color) if tr_color else None})
        r=ctk.CTkFrame(self.box,fg_color="transparent");r.pack(fill="x",padx=12,pady=(5 if big else 3,5 if big else 3))
        ctk.CTkLabel(r,text=label,font=("Segoe UI",11 if not big else 12),text_color=C["t2"],anchor="w").pack(side="left")
        right=ctk.CTkFrame(r,fg_color="transparent");right.pack(side="right")
        if tr:
            arr,tc=tr
            if arr: ctk.CTkLabel(right,text=arr,font=("Segoe UI",10),text_color=tc).pack(side="left",padx=(0,4))
        ctk.CTkLabel(right,text=value,font=(FONT_NUM,13 if not big else 16,"bold"),text_color=color or C["t1"],anchor="e").pack(side="right")
    def sep(self):
        self._rows.append({"kind":"sep"})
        ctk.CTkFrame(self.box,fg_color=C["border"],height=1).pack(fill="x",padx=12,pady=3)
    def bar(self,label,value,max_val,color):
        self._rows.append({"kind":"bar","label":str(label),"value":feur(value,d=0),
                           "max_val":max_val,"color":str(color) if color else None})
        r=ctk.CTkFrame(self.box,fg_color="transparent");r.pack(fill="x",padx=12,pady=(5,2))
        ctk.CTkLabel(r,text=label,font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(side="left")
        ctk.CTkLabel(r,text=feur(value,d=0),font=(FONT_NUM,13,"bold"),text_color=color,anchor="e").pack(side="right")
        bf=ctk.CTkFrame(self.box,fg_color=C["border"],height=6,corner_radius=3);bf.pack(fill="x",padx=12,pady=(0,6));bf.pack_propagate(False)
        try: pct=min(1.0,float(value)/float(max_val)) if max_val and float(max_val)>0 else 0
        except: pct=0
        ctk.CTkFrame(bf,fg_color=color,corner_radius=3).place(x=0,y=0,relwidth=max(0.005,pct),relheight=1.0)
    def alert(self,text,color=None):
        self._rows.append({"kind":"alert","text":str(text),"color":str(color) if color else None})
        r=ctk.CTkFrame(self.box,fg_color=C["alert_bg"],corner_radius=6,border_width=1,border_color=C["alert_border"])
        r.pack(fill="x",padx=10,pady=3)
        ctk.CTkLabel(r,text=text,font=("Segoe UI",11),text_color=color or C["amber"],wraplength=300,justify="left",anchor="w").pack(fill="x",padx=10,pady=6)

class LivraisonDialog(ctk.CTkToplevel):
    def __init__(self,parent):
        super().__init__(parent)
        # Référence EXPLICITE au hub : self.master est la racine Tk (pas le parent logique).
        # On stocke `parent` ici pour pouvoir appeler hub.refresh() de manière fiable
        # depuis _save / _no_livr, peu importe qui a créé cette popup (ask_livraison ou tableau).
        self.hub=parent
        self.title("Livraison du jour");self.geometry("520x500");self.minsize(440,480)
        self.configure(fg_color=C["bg"]);self.resizable(False,False);self.transient(parent);self.result=None
        # Fermeture par la croix X = comportement "Pas encore" (snooze 4h) pour ne pas harceler
        self.protocol("WM_DELETE_WINDOW",self._snooze_4h)
        ctk.CTkLabel(self,text="\U0001f69a  Livraison",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(pady=(20,4))
        ctk.CTkLabel(self,text=f"{jour_fr()} {date.today().strftime('%d/%m/%Y')}",font=("Segoe UI",13),text_color=C["gold"]).pack(pady=(0,16))
        self.entries={}
        for carb,color in [("SP",C["blue"]),("GO",C["amber"]),("GNR",C["teal"])]:
            f=ctk.CTkFrame(self,fg_color="transparent");f.pack(fill="x",padx=40,pady=4)
            ctk.CTkLabel(f,text=carb,font=("Segoe UI",14,"bold"),text_color=color,width=60,anchor="w").pack(side="left")
            e=ctk.CTkEntry(f,height=36,width=150,fg_color=C["card"],border_color=C["border"],text_color=C["t1"],font=("Segoe UI",14),placeholder_text="0")
            e.pack(side="left",padx=(10,0));ctk.CTkLabel(f,text="litres",font=("Segoe UI",12),text_color=C["t3"]).pack(side="left",padx=(8,0))
            self.entries[carb.lower()]=e
        # Champ transporteur (texte libre — Bidou demande d'éviter d'avoir à retourner dans Livraisons
        # pour le renseigner après coup).
        ft=ctk.CTkFrame(self,fg_color="transparent");ft.pack(fill="x",padx=40,pady=(12,4))
        ctk.CTkLabel(ft,text="Transporteur",font=("Segoe UI",11),text_color=C["t2"],width=90,anchor="w").pack(side="left")
        self.entry_transporteur=ctk.CTkEntry(ft,height=32,fg_color=C["card"],border_color=C["border"],
                                              text_color=C["t1"],font=("Segoe UI",12),placeholder_text="Nom du transporteur (optionnel)")
        self.entry_transporteur.pack(side="left",fill="x",expand=True,padx=(10,0))
        # Pré-remplir si déjà saisi aujourd'hui
        try:
            livrs=load_json(LIVRAISON_FILE) or {}
            today_key=date.today().strftime("%d/%m/%y")
            if today_key in livrs and isinstance(livrs[today_key],dict):
                tr_existing=livrs[today_key].get("transporteur","")
                if tr_existing: self.entry_transporteur.insert(0,tr_existing)
                for k in ("sp","go","gnr"):
                    v=livrs[today_key].get(k,0)
                    if v and k in self.entries:
                        self.entries[k].insert(0,str(int(v) if v==int(v) else v))
        except Exception as _e: _log_silent_err(exc=_e)
        btns=ctk.CTkFrame(self,fg_color="transparent");btns.pack(fill="x",padx=30,pady=(20,4))
        ctk.CTkButton(btns,text="\u2713 Enregistrer",width=140,height=40,fg_color=C["green"],hover_color="#258A3E",
                      text_color="#FFF",font=("Segoe UI",13,"bold"),corner_radius=8,command=self._save).pack(side="right")
        # Libellés MÉTIER (point 3 liste vivante) : dire ce que ça fait avec un MOMENT MÉTIER,
        # pas un délai opaque. "Après le prochain tour" est calculé selon l'heure courante :
        # 1er tour finit 9h, 2e tour finit 12h, 3e tour finit ~18h (fermeture SARA).
        now=datetime.now();today=now.date()
        if now.hour<9:    next_tour=datetime.combine(today,dt_time(9, 0));tour_lbl="le 1er tour (9h)"
        elif now.hour<12: next_tour=datetime.combine(today,dt_time(12, 0));tour_lbl="le 2e tour (12h)"
        elif now.hour<18: next_tour=datetime.combine(today,dt_time(18, 0));tour_lbl="le 3e tour (~18h)"
        else:             next_tour=datetime.combine(today+timedelta(days=1),dt_time(6, 0));tour_lbl="demain matin (6h)"
        self._snooze_target=next_tour
        ctk.CTkButton(btns,text=f"\u23f1 Pas encore arriv\u00e9e \u2014 me redemander apr\u00e8s {tour_lbl}",width=370,height=40,fg_color=C["card"],hover_color=C["card_h"],
                      border_width=1,border_color=C["border2"],text_color=C["amber"],corner_radius=8,command=self._snooze_tour).pack(side="right",padx=(0,8))
        ctk.CTkButton(btns,text="Aucune livraison aujourd'hui",width=200,height=40,fg_color=C["card"],hover_color=C["card_h"],
                      border_width=1,border_color=C["border2"],text_color=C["t2"],corner_radius=8,command=self._no_livr).pack(side="left")
        # Ligne d'aide : explique sans ambiguïté ce que chaque bouton va faire.
        ctk.CTkLabel(self,text="\u00ab Enregistrer \u00bb : je note la livraison re\u00e7ue.   \u00b7   "
                              f"\u00ab Pas encore \u00bb : je redemande apr\u00e8s {tour_lbl}.   \u00b7   "
                              "\u00ab Aucune livraison \u00bb : je ne redemande plus aujourd'hui.",
                     font=("Segoe UI",9),text_color=C["t3"]).pack(pady=(0,14))
        # Grab modal posé APRÈS la construction complète du contenu, et seulement quand
        # la fenêtre est "viewable". Cause du bug "fenêtre Livraison du jour VIDE"
        # (signalé 21/05/2026) : grab_set() appelé trop tôt (avant l'affichage) lève
        # TclError "window not viewable", ce qui interrompait __init__ juste après le
        # titre → le contenu n'était jamais construit. En le posant ici avec retry,
        # le contenu se construit toujours et le grab s'installe dès que possible.
        def _safe_grab(attempt=0):
            if not self.winfo_exists(): return
            try: self.grab_set()
            except Exception:
                if attempt<20: self.after(50,lambda:_safe_grab(attempt+1))
        self.after(0,_safe_grab)

    def _save(self):
        self.result={}
        for k,e in self.entries.items():
            try: self.result[k]=float(e.get() or 0)
            except: self.result[k]=0
        # Récupère transporteur (texte libre, peut être vide)
        try: transporteur=self.entry_transporteur.get().strip()
        except Exception: transporteur=""
        if sum(self.result.values())>0:
            try:
                livrs=load_json(LIVRAISON_FILE) or {}
                payload=dict(self.result)
                if transporteur: payload["transporteur"]=transporteur
                livrs[date.today().strftime("%d/%m/%y")]=payload
                save_json(LIVRAISON_FILE,livrs)
            except Exception as _e: _log_silent_err(exc=_e)
            # Journal d'événements : capture la livraison reçue
            try:
                evt_data={
                    "jour":date.today().isoformat(),
                    "sp":int(self.result.get("sp",0)),
                    "go":int(self.result.get("go",0)),
                    "gnr":int(self.result.get("gnr",0)),
                }
                if transporteur: evt_data["transporteur"]=transporteur
                add_evenement("livraison",evt_data)
            except Exception as _e: _log_silent_err(exc=_e)
        # Lever le silence "livraison_jour" : la situation est résolue, ne doit plus apparaître
        try: clear_popup_silence("livraison_jour")
        except Exception as _e: _log_silent_err(exc=_e)
        # === BILAN A POSTERIORI ===
        # Si la livraison saisie correspond à des alertes marge_tendue actives aujourd'hui,
        # on demande à Bidou comment ça s'est passé (OK / Tendu / Problème) pour traçabilité.
        marges_a_bilan=self._check_marges_tendues_a_bilan()
        # Rafraîchir le hub via la référence EXPLICITE (self.master ne marche pas pour CTkToplevel)
        try:
            if hasattr(self.hub,"refresh"): self.hub.refresh()
        except Exception as _e: _log_silent_err(exc=_e)
        if marges_a_bilan:
            # Lancer la popup bilan AVANT destroy, en passant le hub comme parent
            try: BilanLivraisonDlg(self.hub,marges_a_bilan)
            except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _check_marges_tendues_a_bilan(self):
        """Cherche dans le journal les événements marge_tendue de TODAY qui n'ont pas encore
        reçu de bilan (champ 'bilan' absent dans data). Retourne la liste à proposer."""
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[]) or []
            today_iso=date.today().isoformat()
            out=[]
            for evt in events:
                if evt.get("type")!="marge_tendue": continue
                data=evt.get("data",{}) or {}
                # Date de l'événement marge_tendue = date de la livraison concernée
                d_evt=(data.get("date","") or "")[:10]
                if d_evt!=today_iso: continue
                if data.get("bilan"): continue  # déjà bilanté
                out.append({
                    "evt_id":evt.get("id"),
                    "carburant":data.get("carburant","?"),
                    "marge_prevue":int(sf(data.get("marge_restante",0))),
                })
            return out
        except Exception as _e:
            _log_silent_err(exc=_e);return []

    def _no_livr(self):
        try:
            livrs=load_json(LIVRAISON_FILE) or {}
            livrs[date.today().strftime("%d/%m/%y")]={"none":True}
            save_json(LIVRAISON_FILE,livrs)
        except Exception as _e: _log_silent_err(exc=_e)
        try: clear_popup_silence("livraison_jour")
        except Exception as _e: _log_silent_err(exc=_e)
        try:
            if hasattr(self.hub,"refresh"): self.hub.refresh()
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()
    def _snooze_tour(self):
        """Snooze jusqu'au PROCHAIN TOUR (moment métier) au lieu d'un +4h arbitraire.
        self._snooze_target est calculé à __init__ selon l'heure courante :
          - now < 9h  → fin 1er tour (9h)
          - now < 12h → fin 2e tour (12h)
          - now < 18h → fin 3e tour (~18h, fermeture SARA)
          - sinon     → demain matin 6h.
        Cohérent avec popup_silence.cfg (centre de gestion unique des alertes différées)."""
        try:
            today_fp=date.today().strftime("%Y-%m-%d")
            until_iso=self._snooze_target.isoformat()
            silence_popup("livraison_jour",[today_fp],until_iso,{today_fp:0})
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

# =============================================================================
# DIALOGUE BILAN A POSTERIORI D'UNE LIVRAISON À MARGE TENDUE
# Déclenché après LivraisonDialog quand une ou plusieurs alertes marge_tendue
# étaient actives pour aujourd'hui. Demande à Bidou comment ça s'est passé sur
# place (OK / Tendu / Problème), avec commentaire OBLIGATOIRE en cas de Problème.
# Le bilan est inscrit dans l'événement marge_tendue (data.bilan + commentaire)
# pour traçabilité historique — utile notamment en cas de litige SARA.
# =============================================================================
class BilanLivraisonDlg(ctk.CTkToplevel):
    def __init__(self,parent,marges_tendues):
        super().__init__(parent)
        self.hub=parent
        self.marges=marges_tendues  # liste [{evt_id, carburant, marge_prevue}, ...]
        nb=len(self.marges)
        self.title("Bilan livraison \u2014 DISTRICARB HUB")
        h=460 if nb<=1 else 460+min(nb-1,3)*40
        self.geometry(f"640x{h}");self.minsize(560,h)
        self.configure(fg_color=C["bg"]);self.resizable(False,False)
        self.transient(parent);self.grab_set()
        # Fermeture croix X : on n'enregistre pas de bilan, on garde l'événement marge_tendue
        # en attente (resté sans bilan) → la popup pourra se reproposer plus tard ou être
        # vue depuis le journal. Pas de fermeture forcée d'écriture pour ne pas piéger Bidou.
        self.protocol("WM_DELETE_WINDOW",self._on_cancel)
        # Header
        head=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=0)
        head.pack(fill="x")
        ctk.CTkLabel(head,text="\u2713 Bilan de la livraison",font=("Segoe UI",17,"bold"),
                     text_color=C["green"]).pack(pady=(16,2))
        msg=("Cette livraison \u00e9tait signal\u00e9e en marge tendue avant r\u00e9ception. "
             "Comment \u00e7a s'est r\u00e9ellement pass\u00e9 sur place ?")
        ctk.CTkLabel(head,text=msg,font=("Segoe UI",11),text_color=C["t2"],
                     wraplength=580,justify="center").pack(pady=(0,14),padx=24)
        # Récap des carburants concernés
        recap=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8,
                            border_width=1,border_color=C["amber"])
        recap.pack(fill="x",padx=18,pady=(12,8))
        ctk.CTkLabel(recap,text="\u26a0  Carburants signal\u00e9s",font=("Segoe UI",10,"bold"),
                     text_color=C["amber"]).pack(anchor="w",padx=14,pady=(10,2))
        for m in self.marges:
            txt=f"  \u2022 {m['carburant']} : marge cuve pr\u00e9vue {m['marge_prevue']:,} L"
            ctk.CTkLabel(recap,text=txt.replace(",","\u202f"),font=("Segoe UI",11),
                         text_color=C["t1"]).pack(anchor="w",padx=14,pady=1)
        ctk.CTkLabel(recap,text="",font=("Segoe UI",4)).pack(pady=(0,8))
        # 3 choix radio
        self.bilan_var=ctk.StringVar(value="")
        choices=ctk.CTkFrame(self,fg_color="transparent")
        choices.pack(fill="x",padx=24,pady=(8,6))
        rb_opts=[
            ("ok","\u2713 OK \u2014 tout est rentr\u00e9, marge respect\u00e9e",C["green"]),
            ("tendu","\u26a0 Tendu \u2014 marge r\u00e9elle plus serr\u00e9e que pr\u00e9vu",C["amber"]),
            ("probleme","\u2717 Probl\u00e8me \u2014 d\u00e9bordement / refus / report",C["red"]),
        ]
        for v,lbl,col in rb_opts:
            rb=ctk.CTkRadioButton(choices,text=lbl,variable=self.bilan_var,value=v,
                                   font=("Segoe UI",11),text_color=C["t1"],
                                   fg_color=col,hover_color=col,border_color=C["border2"],
                                   radiobutton_width=18,radiobutton_height=18,
                                   command=self._on_choice_change)
            rb.pack(anchor="w",pady=4)
        # Commentaire (statut obligatoire selon choix)
        cmt_frame=ctk.CTkFrame(self,fg_color="transparent")
        cmt_frame.pack(fill="x",padx=24,pady=(6,4))
        self.cmt_label=ctk.CTkLabel(cmt_frame,text="Commentaire (optionnel)",
                                     font=("Segoe UI",10),text_color=C["t2"],anchor="w")
        self.cmt_label.pack(anchor="w",pady=(0,4))
        self.cmt_entry=ctk.CTkTextbox(cmt_frame,height=58,fg_color=C["card"],
                                       border_color=C["border"],text_color=C["t1"],
                                       font=("Segoe UI",11))
        self.cmt_entry.pack(fill="x")
        # Hint pour problème
        self.hint_lbl=ctk.CTkLabel(self,text="",font=("Segoe UI",9),text_color=C["red"])
        self.hint_lbl.pack(pady=(2,0))
        # Boutons
        btns=ctk.CTkFrame(self,fg_color="transparent")
        btns.pack(fill="x",padx=24,pady=(8,16))
        ctk.CTkButton(btns,text="Sauter",width=100,height=36,fg_color=C["card"],
                      hover_color=C["card_h"],border_width=1,border_color=C["border2"],
                      text_color=C["t2"],corner_radius=8,
                      font=("Segoe UI",11),command=self._on_cancel).pack(side="left")
        self.btn_ok=ctk.CTkButton(btns,text="\u2713 Enregistrer bilan",width=180,height=36,
                                   fg_color=C["green"],hover_color="#258A3E",text_color="#FFF",
                                   font=("Segoe UI",12,"bold"),corner_radius=8,
                                   command=self._on_save)
        self.btn_ok.pack(side="right")

    def _on_choice_change(self):
        """Met à jour le label commentaire selon le choix (obligatoire si Problème)."""
        v=self.bilan_var.get()
        if v=="probleme":
            self.cmt_label.configure(text="Commentaire OBLIGATOIRE (qu'est-ce qui s'est pass\u00e9 ?)",
                                      text_color=C["red"])
            self.hint_lbl.configure(text="\u2192 indispensable en cas de litige avec SARA")
        elif v=="tendu":
            self.cmt_label.configure(text="Commentaire recommand\u00e9 (d\u00e9tails sur la tension)",
                                      text_color=C["amber"])
            self.hint_lbl.configure(text="")
        else:
            self.cmt_label.configure(text="Commentaire (optionnel)",text_color=C["t2"])
            self.hint_lbl.configure(text="")

    def _on_save(self):
        choix=self.bilan_var.get()
        if not choix:
            self.hint_lbl.configure(text="\u2192 Choisis OK / Tendu / Probl\u00e8me avant d'enregistrer",
                                     text_color=C["red"])
            return
        cmt=self.cmt_entry.get("1.0","end").strip()
        if choix=="probleme" and not cmt:
            self.hint_lbl.configure(text="\u2192 Le commentaire est obligatoire en cas de probl\u00e8me",
                                     text_color=C["red"])
            return
        # Mettre à jour les événements marge_tendue concernés : ajouter data.bilan + commentaire
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[]) or []
            ts_iso=datetime.now().isoformat()
            ts_human=datetime.now().strftime("%d/%m/%Y %Hh%M")
            choix_human={"ok":"\u2713 OK","tendu":"\u26a0 Tendu","probleme":"\u2717 Probl\u00e8me"}.get(choix,choix)
            for m in self.marges:
                for evt in events:
                    if evt.get("id")!=m["evt_id"]: continue
                    data=evt.get("data",{}) or {}
                    data["bilan"]=choix
                    data["bilan_ts"]=ts_iso
                    data["statut"]="resolu"
                    data["lu"]=True
                    if cmt:
                        existing_cmt=evt.get("commentaire","")
                        nouveau=f"[{ts_human}] Bilan livraison : {choix_human} \u2014 {cmt}"
                        evt["commentaire"]=(existing_cmt+"\n"+nouveau).strip() if existing_cmt else nouveau
                    else:
                        existing_cmt=evt.get("commentaire","")
                        nouveau=f"[{ts_human}] Bilan livraison : {choix_human}"
                        evt["commentaire"]=(existing_cmt+"\n"+nouveau).strip() if existing_cmt else nouveau
                    evt["data"]=data
                    break
            all_evt["events"]=events
            save_json(EVENEMENTS_FILE,all_evt)
        except Exception as _e: _log_silent_err(exc=_e)
        # Refresh hub
        try:
            if hasattr(self.hub,"refresh"): self.hub.refresh()
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _on_cancel(self):
        """Fermeture sans bilan : on laisse les événements marge_tendue tels quels.
        Bidou pourra les bilanter plus tard depuis le journal s'il veut."""
        self.destroy()

# =============================================================================
# =============================================================================
# DIALOGUE GÉNÉRATION RAPPORT MENSUEL
# Permet de choisir un mois, générer le snapshot + rapport PDF/HTML, et l'ouvrir.
# =============================================================================
class RapportMensuelDlg(ctk.CTkToplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Rapport \u2014 DISTRICARB HUB")
        self.geometry("1180x820");self.minsize(960,680)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.parent_app=parent
        self._force_regen=False
        # Charger options sauvées (cases à cocher)
        saved_opts=load_rapport_options()
        self.section_vars={}
        for key,label,default in RAPPORT_SECTIONS:
            self.section_vars[key]=ctk.BooleanVar(value=saved_opts.get(key,default))
        # Header
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=60);hdr.pack(fill="x",padx=20,pady=(16,4));hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f4c4  Rapport",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w")
        ctk.CTkLabel(hdr,text="Choisis le type de p\u00e9riode \u00e0 traiter, puis g\u00e9n\u00e8re le PDF/HTML",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(anchor="w",pady=(1,0))
        # Body : 2 colonnes
        body=ctk.CTkFrame(self,fg_color="transparent");body.pack(fill="both",expand=True,padx=20,pady=4)
        # ---- Colonne gauche : choix période + cases à cocher ----
        left=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=8,width=420);left.pack(side="left",fill="y",padx=(0,10))
        left.pack_propagate(False)
        # === SÉLECTEUR DE FORMAT DE SORTIE (HTML/PDF) ===
        # Préférence non contraignante : par défaut 'Auto' (PDF prioritaire avec fallback HTML).
        # L'utilisateur peut imposer HTML uniquement ou PDF uniquement. Le choix est mémorisé
        # entre sessions via load_rapport_format_pref/save_rapport_format_pref.
        ctk.CTkLabel(left,text="Format de sortie",font=("Segoe UI",11,"bold"),text_color=C["t1"]).pack(anchor="w",padx=14,pady=(12,4))
        _fmt_labels={"auto":"Auto (PDF si possible, sinon HTML)","html":"HTML uniquement","pdf":"PDF uniquement"}
        _fmt_pref_saved=load_rapport_format_pref()
        self._fmt_label_to_key={v:k for k,v in _fmt_labels.items()}
        self.format_var=ctk.StringVar(value=_fmt_labels.get(_fmt_pref_saved,_fmt_labels["auto"]))
        self.format_menu=ctk.CTkOptionMenu(left,values=list(_fmt_labels.values()),
                                            variable=self.format_var,
                                            font=("Segoe UI",10),width=390,height=32,
                                            fg_color=C["panel"],button_color=C["border2"],
                                            button_hover_color=C["border"],text_color=C["t1"],
                                            dropdown_fg_color=C["card"],dropdown_text_color=C["t1"])
        self.format_menu.pack(anchor="w",padx=14,pady=(0,8))
        # === SÉLECTEUR DE MODE ===
        ctk.CTkLabel(left,text="Type de rapport",font=("Segoe UI",11,"bold"),text_color=C["t1"]).pack(anchor="w",padx=14,pady=(12,4))
        self.mode_var=ctk.StringVar(value="Mois")
        self.mode_menu=ctk.CTkOptionMenu(left,values=["Mois","Trimestre","Semestre","Année","Plage personnalisée"],
                                          variable=self.mode_var,
                                          font=("Segoe UI",10),width=390,height=32,
                                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                                          text_color=C["t1"],
                                          command=self._on_mode_change)
        self.mode_menu.pack(anchor="w",padx=14,pady=(0,8))
        # === CADRE PARAMÈTRES (change selon le mode) ===
        self.params_frame=ctk.CTkFrame(left,fg_color="transparent")
        self.params_frame.pack(fill="x",padx=14,pady=(4,0))
        # Construction des contrôles initiaux (mode Mois par défaut)
        self._build_params_mois()
        # Statut snapshot / aperçu période
        self.snap_lbl=ctk.CTkLabel(left,text="",font=("Segoe UI",9),text_color=C["t3"],wraplength=390,justify="left")
        self.snap_lbl.pack(anchor="w",padx=14,pady=(8,8))
        # Cases à cocher
        ctk.CTkLabel(left,text="Sections \u00e0 inclure",font=("Segoe UI",11,"bold"),text_color=C["t1"]).pack(anchor="w",padx=14,pady=(8,2))
        self.sections_hint_lbl=ctk.CTkLabel(left,text="Coche celles que tu veux dans le rapport :",font=("Segoe UI",9),text_color=C["t3"])
        self.sections_hint_lbl.pack(anchor="w",padx=14,pady=(0,4))
        cbx_scroll=ctk.CTkScrollableFrame(left,fg_color=C["panel"],scrollbar_button_color=C["border2"])
        cbx_scroll.pack(fill="both",expand=True,padx=10,pady=(0,8))
        for key,label,default in RAPPORT_SECTIONS:
            cb=ctk.CTkCheckBox(cbx_scroll,text=label,variable=self.section_vars[key],
                                font=("Segoe UI",10),text_color=C["t1"],
                                fg_color=C["amber"],hover_color="#C4811D",
                                border_color=C["border2"],
                                command=self._on_change)
            cb.pack(anchor="w",padx=8,pady=3,fill="x")
        # ---- Colonne droite : aperçu ----
        right=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=8);right.pack(side="left",fill="both",expand=True)
        ctk.CTkLabel(right,text="Aper\u00e7u du rapport",font=("Segoe UI",11,"bold"),text_color=C["t1"]).pack(anchor="w",padx=14,pady=(12,2))
        ctk.CTkLabel(right,text="Mise \u00e0 jour automatique selon tes choix \u2014 v\u00e9rifie avant de g\u00e9n\u00e9rer.",
                     font=("Segoe UI",9),text_color=C["t3"]).pack(anchor="w",padx=14,pady=(0,4))
        self.preview_box=ctk.CTkTextbox(right,fg_color=C["bg"],text_color=C["t1"],
                                         font=("Consolas",10),
                                         border_width=1,border_color=C["border2"],wrap="word")
        self.preview_box.pack(fill="both",expand=True,padx=10,pady=(0,8))
        # Footer
        footer=ctk.CTkFrame(self,fg_color="transparent",height=64);footer.pack(side="bottom",fill="x",padx=20,pady=10);footer.pack_propagate(False)
        self.status_lbl=ctk.CTkLabel(footer,text="",font=("Segoe UI",10),text_color=C["t2"])
        self.status_lbl.pack(side="left",pady=12)
        # Bouton Générer (devient "Régénérer" si rapport existe déjà)
        self.btn_generer=ctk.CTkButton(footer,text="\U0001f4c4 G\u00e9n\u00e9rer PDF/HTML",width=210,height=40,
                       fg_color="#E48B2A",hover_color="#C77519",text_color="#FFF",
                       font=("Segoe UI",11,"bold"),corner_radius=8,
                       command=self._generer)
        self.btn_generer.pack(side="right")
        # Bouton Lire (visible seulement si rapport déjà généré)
        self.btn_lire=ctk.CTkButton(footer,text="\U0001f4d6 Lire",width=120,height=40,
                       fg_color="#1F7FD4",hover_color="#1864A8",text_color="#FFF",
                       font=("Segoe UI",11,"bold"),corner_radius=8,
                       command=self._lire)
        self.btn_lire.pack(side="right",padx=(0,8))
        ctk.CTkButton(footer,text="Fermer",width=100,height=40,
                       fg_color=C["panel"],hover_color=C["border2"],text_color=C["t1"],
                       font=("Segoe UI",10),corner_radius=8,
                       command=self._on_close).pack(side="right",padx=(0,8))
        # Initial refresh
        self._on_change()

    # =========================================================================
    # CONSTRUCTION DES CONTRÔLES SELON LE MODE
    # Chaque _build_params_X recrée le contenu de self.params_frame.
    # =========================================================================
    def _clear_params(self):
        """Vide self.params_frame avant de reconstruire le contenu pour un autre mode."""
        for w in self.params_frame.winfo_children():
            try: w.destroy()
            except Exception as _e: _log_silent_err(exc=_e)
    def _build_params_mois(self):
        """Mode Mois : dropdown des 12 derniers mois (comportement historique)."""
        self._clear_params()
        ctk.CTkLabel(self.params_frame,text="Mois \u00e0 traiter",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,4))
        today=date.today()
        self.options=[]
        mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin","juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        for k in range(12):
            yr=today.year if today.month-k>0 else today.year-1
            mo=today.month-k if today.month-k>0 else today.month-k+12
            lbl=f"{mois_noms[mo-1].capitalize()} {yr}"
            if k==0: lbl+=" (en cours)"
            elif k==1: lbl+=" (pr\u00e9c\u00e9dent)"
            self.options.append((yr,mo,lbl))
        default_idx=1 if len(self.options)>1 else 0
        self.choice_var=ctk.StringVar(value=self.options[default_idx][2])
        ctk.CTkOptionMenu(self.params_frame,values=[o[2] for o in self.options],variable=self.choice_var,
                          font=("Segoe UI",10),width=390,height=32,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],
                          command=lambda *a:self._on_change()).pack(anchor="w",pady=(0,4))

    def _years_available(self):
        """Liste des années disponibles. LITRAGE remonte à 2014 ; on propose 2014 → année courante."""
        return list(range(2014,date.today().year+1))

    def _build_params_trimestre(self):
        """Mode Trimestre : dropdown année + radios T1/T2/T3/T4."""
        self._clear_params()
        ctk.CTkLabel(self.params_frame,text="Ann\u00e9e",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,2))
        years=[str(y) for y in reversed(self._years_available())]
        self.year_var=ctk.StringVar(value=str(date.today().year))
        ctk.CTkOptionMenu(self.params_frame,values=years,variable=self.year_var,
                          font=("Segoe UI",10),width=180,height=32,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],
                          command=lambda *a:self._on_change()).pack(anchor="w",pady=(0,8))
        ctk.CTkLabel(self.params_frame,text="Trimestre",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,2))
        self.quart_var=ctk.StringVar(value="T1")
        row=ctk.CTkFrame(self.params_frame,fg_color="transparent");row.pack(anchor="w",pady=(0,4))
        for q in ["T1","T2","T3","T4"]:
            rb=ctk.CTkRadioButton(row,text=q,variable=self.quart_var,value=q,
                                   font=("Segoe UI",10),text_color=C["t1"],
                                   fg_color=C["amber"],hover_color="#C4811D",
                                   border_color=C["border2"],
                                   command=self._on_change)
            rb.pack(side="left",padx=(0,12))

    def _build_params_semestre(self):
        """Mode Semestre : dropdown année + radios S1/S2."""
        self._clear_params()
        ctk.CTkLabel(self.params_frame,text="Ann\u00e9e",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,2))
        years=[str(y) for y in reversed(self._years_available())]
        self.year_var=ctk.StringVar(value=str(date.today().year))
        ctk.CTkOptionMenu(self.params_frame,values=years,variable=self.year_var,
                          font=("Segoe UI",10),width=180,height=32,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],
                          command=lambda *a:self._on_change()).pack(anchor="w",pady=(0,8))
        ctk.CTkLabel(self.params_frame,text="Semestre",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,2))
        self.sem_var=ctk.StringVar(value="S1")
        row=ctk.CTkFrame(self.params_frame,fg_color="transparent");row.pack(anchor="w",pady=(0,4))
        for s in ["S1","S2"]:
            rb=ctk.CTkRadioButton(row,text=s,variable=self.sem_var,value=s,
                                   font=("Segoe UI",10),text_color=C["t1"],
                                   fg_color=C["amber"],hover_color="#C4811D",
                                   border_color=C["border2"],
                                   command=self._on_change)
            rb.pack(side="left",padx=(0,12))

    def _build_params_annee(self):
        """Mode Année : juste un dropdown année."""
        self._clear_params()
        ctk.CTkLabel(self.params_frame,text="Ann\u00e9e compl\u00e8te",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,2))
        years=[str(y) for y in reversed(self._years_available())]
        self.year_var=ctk.StringVar(value=str(date.today().year))
        ctk.CTkOptionMenu(self.params_frame,values=years,variable=self.year_var,
                          font=("Segoe UI",10),width=180,height=32,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],
                          command=lambda *a:self._on_change()).pack(anchor="w",pady=(0,4))

    def _build_params_custom(self):
        """Mode Plage personnalisée : 2 paires (mois + année) début / fin."""
        self._clear_params()
        mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin","juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        years=[str(y) for y in reversed(self._years_available())]
        # Début
        ctk.CTkLabel(self.params_frame,text="D\u00e9but",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,2))
        row1=ctk.CTkFrame(self.params_frame,fg_color="transparent");row1.pack(anchor="w",pady=(0,8),fill="x")
        # Default: 1er janvier de l'année courante
        self.start_month_var=ctk.StringVar(value=mois_noms[0].capitalize())
        self.start_year_var=ctk.StringVar(value=str(date.today().year))
        ctk.CTkOptionMenu(row1,values=[m.capitalize() for m in mois_noms],variable=self.start_month_var,
                          font=("Segoe UI",10),width=140,height=30,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],command=lambda *a:self._on_change()).pack(side="left",padx=(0,8))
        ctk.CTkOptionMenu(row1,values=years,variable=self.start_year_var,
                          font=("Segoe UI",10),width=110,height=30,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],command=lambda *a:self._on_change()).pack(side="left")
        # Fin
        ctk.CTkLabel(self.params_frame,text="Fin",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(anchor="w",pady=(0,2))
        row2=ctk.CTkFrame(self.params_frame,fg_color="transparent");row2.pack(anchor="w",pady=(0,4),fill="x")
        today=date.today()
        self.end_month_var=ctk.StringVar(value=mois_noms[today.month-1].capitalize())
        self.end_year_var=ctk.StringVar(value=str(today.year))
        ctk.CTkOptionMenu(row2,values=[m.capitalize() for m in mois_noms],variable=self.end_month_var,
                          font=("Segoe UI",10),width=140,height=30,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],command=lambda *a:self._on_change()).pack(side="left",padx=(0,8))
        ctk.CTkOptionMenu(row2,values=years,variable=self.end_year_var,
                          font=("Segoe UI",10),width=110,height=30,
                          fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                          text_color=C["t1"],command=lambda *a:self._on_change()).pack(side="left")

    def _on_mode_change(self,_=None):
        """Reconstruit les paramètres quand le mode change."""
        m=self.mode_var.get()
        if m=="Mois": self._build_params_mois()
        elif m=="Trimestre": self._build_params_trimestre()
        elif m=="Semestre": self._build_params_semestre()
        elif m=="Année" or m=="Ann\u00e9e": self._build_params_annee()
        else: self._build_params_custom()
        self._on_change()

    # =========================================================================
    # RÉSOLUTION DE LA PÉRIODE SÉLECTIONNÉE → start_date, end_date, label
    # =========================================================================
    def _resolve_period(self):
        """Retourne (start_date, end_date, label_humain) selon le mode actif.
        Pour le mode Mois, end_date = dernier jour du mois.
        Retourne (None,None,msg) si entrée invalide."""
        m=self.mode_var.get()
        mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin","juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        def _last_day(y,mo):
            if mo==12: return date(y,12,31)
            return date(y,mo+1,1)-timedelta(days=1)
        if m=="Mois":
            sel=self._selected()
            if not sel: return (None,None,"Aucun mois sélectionné")
            yr,mo,lbl=sel
            return (date(yr,mo,1),_last_day(yr,mo),lbl)
        if m=="Trimestre":
            try: yr=int(self.year_var.get())
            except Exception as _e: _log_silent_err(exc=_e); return (None,None,"Année invalide")
            q=getattr(self,"quart_var",None)
            qv=q.get() if q else "T1"
            quarters={"T1":(1,3),"T2":(4,6),"T3":(7,9),"T4":(10,12)}
            m1,m2=quarters.get(qv,(1,3))
            return (date(yr,m1,1),_last_day(yr,m2),f"{qv} {yr}")
        if m=="Semestre":
            try: yr=int(self.year_var.get())
            except Exception as _e: _log_silent_err(exc=_e); return (None,None,"Année invalide")
            s=getattr(self,"sem_var",None)
            sv=s.get() if s else "S1"
            if sv=="S1": return (date(yr,1,1),date(yr,6,30),f"S1 {yr}")
            return (date(yr,7,1),date(yr,12,31),f"S2 {yr}")
        if m=="Année" or m=="Ann\u00e9e":
            try: yr=int(self.year_var.get())
            except Exception as _e: _log_silent_err(exc=_e); return (None,None,"Année invalide")
            return (date(yr,1,1),date(yr,12,31),f"Année {yr}")
        # Custom
        try:
            sm_lbl=self.start_month_var.get().lower()
            em_lbl=self.end_month_var.get().lower()
            sm=mois_noms.index(sm_lbl)+1
            em=mois_noms.index(em_lbl)+1
            sy=int(self.start_year_var.get())
            ey=int(self.end_year_var.get())
        except Exception as _e: _log_silent_err(exc=_e); return (None,None,"Plage personnalisée invalide")
        sd=date(sy,sm,1);ed=_last_day(ey,em)
        if sd>ed: return (None,None,"La date de début doit être antérieure à la date de fin")
        return (sd,ed,f"{sm_lbl.capitalize()} {sy} → {em_lbl.capitalize()} {ey}")

    def _on_close(self):
        # Sauver les options avant de fermer
        try:
            opts={key:var.get() for key,var in self.section_vars.items()}
            save_rapport_options(opts)
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()
    def _selected_sections(self):
        return {key:var.get() for key,var in self.section_vars.items()}
    def _on_change(self):
        """Appelé quand mois change ou case (dé)cochée. Met à jour snap_lbl + preview + boutons."""
        self._refresh_snap_lbl()
        self._refresh_preview()
        self._refresh_buttons()
    def _existing_report_path(self,year,month):
        """Retourne le chemin du rapport mensuel déjà généré pour ce mois (.pdf prioritaire), ou None."""
        out_dir=RAPPORTS_DIR/f"{year:04d}"
        mn_capit=["Janvier","F\u00e9vrier","Mars","Avril","Mai","Juin","Juillet","Ao\u00fbt","Septembre","Octobre","Novembre","D\u00e9cembre"][month-1]
        base=f"Rapport_mensuel_{year:04d}_{month:02d}_{mn_capit}"
        pdf=out_dir/f"{base}.pdf"
        if pdf.exists(): return pdf
        html=out_dir/f"{base}.html"
        if html.exists(): return html
        return None
    def _existing_period_report_path(self,start_date,end_date,label):
        """Retourne le chemin du rapport période déjà généré (PDF prioritaire), ou None.
        Construction du nom alignée sur generate_period_report_html/pdf."""
        safe_label=label.replace(" ","_").replace("/","-").replace("→","a")
        for src,dst in [("é","e"),("è","e"),("ê","e"),("à","a"),("â","a"),("ô","o"),("û","u"),("ç","c"),("É","E"),("È","E")]:
            safe_label=safe_label.replace(src,dst)
        out_dir=RAPPORTS_DIR/f"{start_date.year:04d}"
        sd=start_date.strftime("%Y-%m-%d");ed=end_date.strftime("%Y-%m-%d")
        pdf=out_dir/f"Rapport_periode_{sd}_{ed}_{safe_label}.pdf"
        if pdf.exists(): return pdf
        html=out_dir/f"Rapport_periode_{sd}_{ed}_{safe_label}.html"
        if html.exists(): return html
        return None
    def _refresh_buttons(self):
        """Active/désactive le bouton Lire et adapte le label du bouton Générer."""
        if self.mode_var.get()=="Mois":
            sel=self._selected()
            if not sel:
                self.btn_lire.configure(state="disabled");return
            yr,mo,_=sel
            existing=self._existing_report_path(yr,mo)
        else:
            sd,ed,lbl=self._resolve_period()
            if sd is None: self.btn_lire.configure(state="disabled");return
            existing=self._existing_period_report_path(sd,ed,lbl)
        if existing:
            self.btn_lire.configure(state="normal")
            self.btn_generer.configure(text="\U0001f504 R\u00e9g\u00e9n\u00e9rer")
        else:
            self.btn_lire.configure(state="disabled")
            self.btn_generer.configure(text="\U0001f4c4 G\u00e9n\u00e9rer PDF/HTML")
    def _lire(self):
        """Ouvre le rapport déjà généré dans l'application système, sans rien recalculer."""
        if self.mode_var.get()=="Mois":
            sel=self._selected()
            if not sel: return
            yr,mo,_=sel
            path=self._existing_report_path(yr,mo)
        else:
            sd,ed,lbl=self._resolve_period()
            if sd is None: return
            path=self._existing_period_report_path(sd,ed,lbl)
        if not path:
            self.status_lbl.configure(text="\u26a0 Aucun rapport \u00e0 lire pour cette p\u00e9riode",text_color=C["red"])
            return
        try:
            if sys.platform=="win32": os.startfile(str(path))
            elif sys.platform=="darwin": subprocess.Popen(["open",str(path)])
            else: subprocess.Popen(["xdg-open",str(path)])
            self.status_lbl.configure(text=f"\u2713 {path.name} ouvert",text_color=C["green"])
        except Exception as e:
            self.status_lbl.configure(text=f"\u26a0 {e}",text_color=C["red"])
    def _refresh_snap_lbl(self):
        if self.mode_var.get()=="Mois":
            sel=self._selected()
            if not sel: return
            yr,mo,_=sel
            snap=load_month_snapshot(yr,mo)
            if snap:
                t=snap["totaux"]
                jrs=snap.get("jours_complets",0)
                self.snap_lbl.configure(text=f"\u2713 Snapshot d\u00e9j\u00e0 g\u00e9n\u00e9r\u00e9 \u2014 {jrs} jours, {_format_l(t['litrage_l'])}, CA {_format_eur(t['ca_total_eur'])}")
            else:
                self.snap_lbl.configure(text="Aucun snapshot \u2014 sera cr\u00e9\u00e9 \u00e0 la g\u00e9n\u00e9ration (lecture LITRAGE 10-30s)")
        else:
            sd,ed,lbl=self._resolve_period()
            if sd is None:
                self.snap_lbl.configure(text=f"\u26a0 {lbl}");return
            nb_jours=(ed-sd).days+1
            niveau="Lite" if sd<date(2025,1,1) else "Complet"
            self.snap_lbl.configure(text=f"\u2192 {lbl} \u2014 {nb_jours} jours calendaires \u2014 niveau {niveau}\nLa g\u00e9n\u00e9ration lit LITRAGE.xlsx en entier (10-30s).")
    def _refresh_preview(self):
        """Construit un aperçu textuel du rapport selon les sections cochées."""
        # En mode période, on n'affiche qu'un résumé minimal (pas de snapshot pré-calculé)
        if self.mode_var.get()!="Mois":
            sd,ed,lbl=self._resolve_period()
            self.preview_box.delete("1.0","end")
            if sd is None:
                self.preview_box.insert("1.0",f"\u26a0 {lbl}");return
            nb_jours=(ed-sd).days+1
            niveau="Lite" if sd<date(2025,1,1) else "Complet"
            mono_mois=(sd.year==ed.year and sd.month==ed.month)
            out=[]
            out.append(f"P\u00e9riode : {lbl}")
            out.append(f"  D\u00e9but : {sd.strftime('%d/%m/%Y')}")
            out.append(f"  Fin    : {ed.strftime('%d/%m/%Y')}")
            out.append(f"  Dur\u00e9e  : {nb_jours} jours calendaires")
            out.append("")
            out.append(f"Niveau de d\u00e9tail : {niveau}")
            if niveau=="Lite":
                out.append("  La p\u00e9riode touche du pr\u00e9-2025.")
                out.append("  Contenu limit\u00e9 \u00e0 : litrage, CA piste, CA boutique, marge, d\u00e9tail mensuel.")
            else:
                out.append("  P\u00e9riode enti\u00e8rement \u00e0 partir de 2025.")
                out.append("  Contenu complet : Lite + d\u00e9tail mensuel par carburant + top mois/jours.")
            out.append("")
            if mono_mois:
                out.append("Top 3 sur les meilleures et pires JOURN\u00c9ES.")
                out.append("D\u00e9tail jour par jour inclus.")
            else:
                out.append("Top 3 sur les meilleurs et pires MOIS.")
                out.append("Pas de d\u00e9tail jour par jour (trop volumineux).")
            out.append("")
            out.append("\u26a0 Les sections \u00e0 cocher (panneau gauche) ne s'appliquent")
            out.append("   pas encore aux rapports p\u00e9riode (toutes incluses par d\u00e9faut).")
            self.preview_box.insert("1.0","\n".join(out))
            return
        sel=self._selected()
        if not sel:
            self.preview_box.delete("1.0","end");return
        yr,mo,_=sel
        snap=load_month_snapshot(yr,mo)
        sections=self._selected_sections()
        nb_cochees=sum(1 for v in sections.values() if v)
        out=[]
        if not snap:
            out.append(f"\u26a0 Pas encore de snapshot pour {self._selected()[2]}.")
            out.append("La g\u00e9n\u00e9ration lira LITRAGE.xlsx en entier (10-30s).")
            out.append("")
            out.append(f"Sections coch\u00e9es ({nb_cochees}/{len(RAPPORT_SECTIONS)}) :")
            for key,label,default in RAPPORT_SECTIONS:
                mark="[\u2713]" if sections.get(key) else "[ ]"
                out.append(f"  {mark} {label}")
            self.preview_box.delete("1.0","end")
            self.preview_box.insert("1.0","\n".join(out))
            return
        # Snapshot dispo : aperçu détaillé
        t=snap["totaux"];m=snap["moyennes"];nb_j=snap.get("jours_complets",0)
        moy_bout=int(round(t['ca_boutique_eur']/nb_j)) if nb_j else 0
        out.append(f"=== APER\u00c7U RAPPORT \u2014 {snap['month_name'].capitalize()} {yr} ===")
        out.append(f"({nb_cochees}/{len(RAPPORT_SECTIONS)} sections coch\u00e9es)")
        out.append("")
        if sections.get("synthese"):
            out.append(f"\u25c6 SYNTH\u00c8SE DU MOIS ({nb_j} jours saisis)")
            out.append(f"   Total litrage : {_format_l(t['litrage_l'])}")
            out.append(f"   CA piste : {_format_eur(t['ca_piste_eur'])}")
            out.append(f"   CA boutique : {_format_eur(t['ca_boutique_eur'])}")
            out.append(f"   Moy. CA boutique/j : {_format_eur(moy_bout)}")
            out.append(f"   Moy. litrage/j : {_format_l(m['litrage_jour'])}")
            out.append(f"   CA total : {_format_eur(t['ca_total_eur'])}")
            out.append("")
            # Nouvelle section dédiée Marge totale et détaillée
            marge_data=snap.get("marge",{}) or {}
            marge_carb=marge_data.get("total_eur",0)
            marge_bout=marge_data.get("boutique_eur",0)
            marge_bout_taux=marge_data.get("boutique_taux",0.30)
            out.append(f"\u25c6 MARGE TOTALE ET D\u00c9TAILL\u00c9E")
            out.append(f"   Marge carburant : {_format_eur(marge_carb)} ({marge_data.get('moyen_unit',0):.5f} \u20ac/L moyen)".replace(".",","))
            out.append(f"   Marge boutique  : {_format_eur(marge_bout)} ({marge_bout_taux*100:.2f}% du CA)")
            out.append(f"   Marge totale    : {_format_eur(marge_carb+marge_bout)}")
            eff=marge_data.get("effet_speculation")
            if eff and eff.get("total"):
                tot_eff=eff.get("total",0)
                signe="+" if tot_eff>=0 else ""
                out.append(f"   Effet sp\u00e9culation : {signe}{_format_eur(tot_eff)}")
            out.append("")
        if sections.get("carburants"):
            tot=t['litrage_l'] or 1
            out.append("\u25c6 R\u00c9PARTITION PAR CARBURANT")
            out.append(f"   SP  : {_format_l(t['litrage_sp_l'])} ({100*t['litrage_sp_l']/tot:.1f}%)")
            out.append(f"   GO  : {_format_l(t['litrage_go_l'])} ({100*t['litrage_go_l']/tot:.1f}%)")
            out.append(f"   GNR : {_format_l(t['litrage_gnr_l'])} ({100*t['litrage_gnr_l']/tot:.1f}%)")
            out.append("")
        if sections.get("encaissements"):
            out.append("\u25c6 ENCAISSEMENTS")
            out.append(f"   CB       : {_format_eur(t['encaiss_cb_eur'])}")
            out.append(f"   CP       : {_format_eur(t['encaiss_cp_eur'])}")
            out.append(f"   Esp\u00e8ces  : {_format_eur(t['encaiss_esp_eur'])}")
            out.append("")
        if sections.get("admin"):
            ob=snap.get("objectif",{}) or {}
            block=["\u25c6 PILOTAGE ADMINISTRATIF & ALERTES"]
            has_data=False
            bal=ob.get("balance_de_eur",0)
            if bal: block.append(f"   Balance D/E : {_format_eur(bal)}");has_data=True
            taux=ob.get("taux_avancement",0)
            if taux: block.append(f"   Avancement objectif CA : {taux*100:.1f}% ({_format_eur(ob.get('obj_ca_eur',0))})");has_data=True
            if ob.get("cp_retard_count",0)>0:
                block.append(f"   \u26a0 CP en retard : {_format_eur(ob['cp_retard_total'])} ({ob['cp_retard_count']} op.)");has_data=True
            if ob.get("dec_pending_count",0)>0:
                block.append(f"   D\u00e9caissements \u00e0 venir : {_format_eur(ob['dec_pending_total'])} ({ob['dec_pending_count']})");has_data=True
            if ob.get("enc_retard_count",0)>0:
                block.append(f"   \u26a0 Encaissements en retard : {_format_eur(ob['enc_retard_total'])} ({ob['enc_retard_count']})");has_data=True
            if ob.get("clients_impayes_count",0)>0:
                block.append(f"   Clients impay\u00e9s : {_format_eur(ob['clients_impayes_total'])} ({ob['clients_impayes_count']})");has_data=True
            if has_data: out.extend(block);out.append("")
            else: out.append("\u25c6 PILOTAGE ADMIN. (aucune alerte active)");out.append("")
        if sections.get("detail_jours"):
            out.append(f"\u25c6 D\u00c9TAIL JOUR PAR JOUR ({nb_j} lignes)")
            out.append("")
        if sections.get("top3_piste_meilleures"):
            out.append("\u25c6 TOP 3 MEILLEURES (CA piste)")
            for j in snap.get("top3_meilleures",[]):
                out.append(f"   {j['label']} : {_format_l(j['litrage'])}, CA {_format_eur(j['ca_piste'])}")
            out.append("")
        if sections.get("top3_piste_pires"):
            out.append("\u25c6 TOP 3 PLUS FAIBLES (CA piste)")
            for j in snap.get("top3_pires",[]):
                out.append(f"   {j['label']} : {_format_l(j['litrage'])}, CA {_format_eur(j['ca_piste'])}")
            out.append("")
        if sections.get("top3_bout_meilleures"):
            out.append("\u25c6 TOP 3 MEILLEURES (CA boutique)")
            for j in snap.get("top3_bout_meilleures",[]):
                out.append(f"   {j['label']} : bout. {_format_eur(j['ca_boutique'])}, piste {_format_eur(j['ca_piste'])}")
            out.append("")
        if sections.get("top3_bout_pires"):
            out.append("\u25c6 TOP 3 PLUS FAIBLES (CA boutique)")
            for j in snap.get("top3_bout_pires",[]):
                out.append(f"   {j['label']} : bout. {_format_eur(j['ca_boutique'])}, piste {_format_eur(j['ca_piste'])}")
            out.append("")
        if sections.get("anomalies"):
            anoms=snap.get("anomalies_tendance",[])
            out.append(f"\u25c6 ANOMALIES DE TENDANCE ({len(anoms)})")
            for a in anoms[:8]: out.append(f"   \u2022 {a}")
            out.append("")
        if sections.get("ponts"):
            ponts=snap.get("ponts_traverses",[])
            out.append(f"\u25c6 PONTS TRAVERS\u00c9S ({len(ponts)})")
            for p in ponts[:8]: out.append(f"   \u2022 {p}")
            out.append("")
        if nb_cochees==0:
            out.append("\u26a0 Aucune section coch\u00e9e \u2014 le rapport sera vide.")
        self.preview_box.delete("1.0","end")
        self.preview_box.insert("1.0","\n".join(out))
    def _selected(self):
        v=self.choice_var.get()
        for o in self.options:
            if o[2]==v: return o
        return None
    def _generer(self):
        # Aiguillage selon le mode : Mois → flux mensuel historique, autre → flux période
        if self.mode_var.get()!="Mois":
            return self._generer_periode()
        sel=self._selected()
        if not sel: return
        yr,mo,_=sel
        # Récupérer les sections cochées + sauver pour la prochaine fois
        sections=self._selected_sections()
        if not any(sections.values()):
            self.status_lbl.configure(text="\u26a0 Coche au moins une section",text_color=C["red"])
            return
        try: save_rapport_options(sections)
        except Exception as _e: _log_silent_err(exc=_e)
        self.status_lbl.configure(text="G\u00e9n\u00e9ration en cours...",text_color=C["amber"])
        self.update_idletasks()
        try:
            # 1. Lecture COMPLÈTE de LITRAGE pour avoir tous les jours du mois demandé
            self.status_lbl.configure(text="Lecture compl\u00e8te de LITRAGE.xlsx (peut prendre 10-30s)...",text_color=C["amber"])
            self.update_idletasks()
            cfg=load_json(CONFIG_FILE) or {}
            # ====================================================
            # RÉSOLUTION DU FICHIER OBJECTIF DU MOIS DEMANDÉ
            # Le path par défaut pointe vers le mois en cours. Pour générer le rapport
            # d'un mois passé, il faut lire le fichier Objectif de CE mois-là.
            # ====================================================
            current_obj=cfg.get("objectif")
            target_obj_path=resolve_objectif_path_for_month(yr,mo,current_obj)
            if not target_obj_path:
                # Demander à l'utilisateur de sélectionner le fichier
                from tkinter import filedialog
                mois_fr=["Janvier","F\u00e9vrier","Mars","Avril","Mai","Juin","Juillet","Ao\u00fbt","Septembre","Octobre","Novembre","D\u00e9cembre"]
                mn=mois_fr[mo-1]
                self.status_lbl.configure(text=f"Fichier Objectif de {mn} {yr} introuvable \u2014 s\u00e9lection manuelle...",text_color=C["amber"])
                self.update_idletasks()
                # Path de départ : dossier du fichier Objectif actuel
                init_dir=str(Path(current_obj).parent) if current_obj else str(Path.home())
                target_obj_path=filedialog.askopenfilename(
                    parent=self,
                    title=f"S\u00e9lectionner le fichier Objectif de {mn} {yr}",
                    initialdir=init_dir,
                    filetypes=[("Excel","*.xlsx *.xlsm"),("Tous","*.*")]
                )
                if not target_obj_path:
                    self.status_lbl.configure(text="Annul\u00e9 (aucun fichier s\u00e9lectionn\u00e9)",text_color=C["t3"])
                    return
                # Mémoriser pour ne plus redemander
                remember_objectif_path(yr,mo,target_obj_path)
            # Construire une config temporaire avec le bon Objectif (les autres fichiers restent inchangés)
            cfg_for_report=dict(cfg)
            cfg_for_report["objectif"]=target_obj_path
            reader=DataReader(cfg_for_report)
            hist_data=reader._read_hist(full=True)
            if not hist_data:
                self.status_lbl.configure(text="\u26a0 Impossible de lire LITRAGE.xlsx",text_color=C["red"])
                return
            # Lire alertes et objectif depuis le fichier Objectif du bon mois
            self.status_lbl.configure(text=f"Lecture du fichier Objectif de {Path(target_obj_path).stem}...",text_color=C["amber"])
            self.update_idletasks()
            try:
                alerts_data=reader._read_alerts() or {}
                objectif_data=reader._read_ob() or {}
            except Exception as e:
                print(f"[rapport] erreur lecture objectif : {e}")
                alerts_data={};objectif_data={}
            # Récupérer les prix mémorisés pour le mois demandé (depuis prix_historique.cfg).
            # Si le mois est en cours, c'est mis à jour à chaque refresh. Si c'est un mois
            # passé, on a la valeur figée à la fin du mois.
            prix_for_month=get_prix_for_month(yr,mo) or {}
            # 2. Construire et sauvegarder le snapshot
            snap=build_month_snapshot(hist_data,yr,mo,alerts=alerts_data,objectif=objectif_data,prix_data=prix_for_month)
            if snap.get("jours_complets",0)==0:
                self.status_lbl.configure(text=f"\u26a0 Aucun jour saisi dans LITRAGE pour ce mois",text_color=C["red"])
                return
            save_path=save_month_snapshot(snap)
            # 3. Générer le rapport avec les sections cochées et le format choisi
            fmt_pref=self._fmt_label_to_key.get(self.format_var.get(),"auto")
            save_rapport_format_pref(fmt_pref)
            path,fmt,status=generate_monthly_report(snap,force_regenerate=getattr(self,"_force_regen",False),sections=sections,format_pref=fmt_pref)
            if not path:
                self.status_lbl.configure(text="\u26a0 Erreur lors de la g\u00e9n\u00e9ration du rapport",text_color=C["red"])
                return
            # Si rapport existant déjà en place (mois passé) → demander à l'utilisateur
            if status=="existing":
                resp=messagebox.askyesnocancel(
                    "Rapport d\u00e9j\u00e0 g\u00e9n\u00e9r\u00e9",
                    f"Le rapport de {snap['month_name']} {yr} a d\u00e9j\u00e0 \u00e9t\u00e9 g\u00e9n\u00e9r\u00e9.\n\n"
                    f"Fichier : {path.name}\n\n"
                    f"\u2022 OUI : Ouvrir le rapport existant (recommand\u00e9)\n"
                    f"\u2022 NON : Re-g\u00e9n\u00e9rer (l'ancien sera archiv\u00e9, pas \u00e9cras\u00e9)\n"
                    f"\u2022 ANNULER : Ne rien faire",
                    parent=self
                )
                if resp is None:
                    self.status_lbl.configure(text="Annul\u00e9",text_color=C["t3"])
                    return
                if resp is False:
                    # Re-génération demandée → relancer avec force_regenerate
                    self._force_regen=True
                    self.status_lbl.configure(text="Re-g\u00e9n\u00e9ration en cours (ancien archiv\u00e9)...",text_color=C["amber"])
                    self.update_idletasks()
                    self._generer()
                    self._force_regen=False
                    return
                # resp=True → ouvrir l'existant, on continue le flow normal
            # 4. Mémoriser que le rapport a été généré
            try:
                mem=load_json(RAPPORT_PROMPT_FILE) or {}
                mem[f"{yr:04d}-{mo:02d}"]={"statut":"genere","date":datetime.now().isoformat(),"path":str(path)}
                save_json(RAPPORT_PROMPT_FILE,mem)
            except Exception as _e: _log_silent_err(exc=_e)
            # 5. Ouvrir le fichier dans l'application système
            status_msg={
                "new":f"\u2713 Rapport {fmt.upper()} g\u00e9n\u00e9r\u00e9 : {path.name}",
                "existing":f"\u2713 Ouverture du rapport existant : {path.name}",
                "regenerated":f"\u2713 Rapport re-g\u00e9n\u00e9r\u00e9 (ancien archiv\u00e9) : {path.name}",
            }.get(status,f"\u2713 {path.name}")
            self.status_lbl.configure(text=status_msg,text_color=C["green"])
            try:
                if sys.platform=="win32": os.startfile(str(path))
                elif sys.platform=="darwin": subprocess.Popen(["open",str(path)])
                else: subprocess.Popen(["xdg-open",str(path)])
            except Exception as e:
                print(f"[ouverture rapport] {e}")
            # 6. Refresh snapshot label + aperçu après génération
            self._on_change()
        except Exception as e:
            traceback.print_exc()
            self.status_lbl.configure(text=f"\u26a0 Erreur : {e}",text_color=C["red"])

    def _generer_periode(self):
        """Génère un rapport sur une plage libre (Trimestre, Semestre, Année, Custom).
        Utilise build_period_snapshot + generate_period_report_html. Pas de PDF en mode
        période pour l'instant, uniquement HTML autonome (imprimable Ctrl+P)."""
        sd,ed,lbl=self._resolve_period()
        if sd is None:
            self.status_lbl.configure(text=f"\u26a0 {lbl}",text_color=C["red"]);return
        # Bornes physiques : LITRAGE remonte au 19 juin 2014
        if sd<date(2014,6,19):
            self.status_lbl.configure(text="\u26a0 La date de début est antérieure aux données disponibles (19 juin 2014).",text_color=C["red"]);return
        if ed>date.today():
            ed=date.today()
        self.status_lbl.configure(text="G\u00e9n\u00e9ration en cours...",text_color=C["amber"])
        self.update_idletasks()
        try:
            # 1. Lecture COMPLÈTE de LITRAGE
            self.status_lbl.configure(text="Lecture compl\u00e8te de LITRAGE.xlsx (peut prendre 10-30s)...",text_color=C["amber"])
            self.update_idletasks()
            cfg=load_json(CONFIG_FILE) or {}
            reader=DataReader(cfg)
            hist_data=reader._read_hist(full=True)
            if not hist_data:
                self.status_lbl.configure(text="\u26a0 Impossible de lire LITRAGE.xlsx",text_color=C["red"]);return
            # 2. Construire le snapshot période (alertes/objectif laissés à None : pas pertinents
            # pour une plage multi-mois ; le snapshot les neutralise automatiquement)
            self.status_lbl.configure(text=f"Construction du rapport pour {lbl}...",text_color=C["amber"])
            self.update_idletasks()
            snap=build_period_snapshot(hist_data,sd,ed)
            if snap.get("jours_complets",0)==0:
                self.status_lbl.configure(text=f"\u26a0 Aucun jour saisi dans LITRAGE pour cette p\u00e9riode",text_color=C["red"]);return
            # 3. Générer en respectant la préférence de format choisie par l'utilisateur
            fmt_pref=self._fmt_label_to_key.get(self.format_var.get(),"auto")
            save_rapport_format_pref(fmt_pref)
            path=None;fmt=""
            if fmt_pref=="html":
                path=generate_period_report_html(snap);fmt="HTML"
            elif fmt_pref=="pdf":
                path=generate_period_report_pdf(snap);fmt="PDF"
            else:  # auto : PDF prioritaire, fallback HTML (comportement historique)
                path=generate_period_report_pdf(snap);fmt="PDF"
                if not path:
                    path=generate_period_report_html(snap);fmt="HTML"
            if not path:
                self.status_lbl.configure(text="\u26a0 Erreur lors de la g\u00e9n\u00e9ration du rapport",text_color=C["red"]);return
            # 4. Ouvrir le fichier dans l'application système
            self.status_lbl.configure(text=f"\u2713 Rapport {fmt} g\u00e9n\u00e9r\u00e9 : {path.name}",text_color=C["green"])
            try:
                if sys.platform=="win32": os.startfile(str(path))
                elif sys.platform=="darwin": subprocess.Popen(["open",str(path)])
                else: subprocess.Popen(["xdg-open",str(path)])
            except Exception as e:
                print(f"[ouverture rapport] {e}")
            # 5. Refresh boutons (le rapport existe maintenant → bouton Lire actif)
            self._on_change()
        except Exception as e:
            traceback.print_exc()
            self.status_lbl.configure(text=f"\u26a0 Erreur : {e}",text_color=C["red"])

# =============================================================================
# JOURNAL DES ÉVÉNEMENTS (Sujet E)
# Affiche tous les événements horodatés capturés par le hub : ponts, anomalies,
# ruptures, livraisons. Permet de filtrer par période et par type.
# =============================================================================
class PrixCarburantDlg(ctk.CTkToplevel):
    """Affiche les prix de vente, prix d'achat et marges unitaires des 3 carburants.
    Source : prix_historique.cfg pour le mois en cours. Utile pour répondre rapidement
    à un client ou collaborateur sans devoir ouvrir Prévision compte.xlsx."""
    def __init__(self,parent):
        super().__init__(parent)
        self.title("\U0001f4b0 Prix carburant — DISTRICARB HUB")
        self.geometry("680x480");self.minsize(560,420)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        # Header
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=64);hdr.pack(fill="x",padx=20,pady=(16,4));hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f4b0  Prix carburant",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w")
        # Mois en cours
        today=date.today()
        mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin",
                   "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        self.subtitle_lbl=ctk.CTkLabel(hdr,text=f"Tarifs en vigueur — {mois_noms[today.month-1].capitalize()} {today.year}",
                                        font=("Segoe UI",10),text_color=C["t3"])
        self.subtitle_lbl.pack(anchor="w",pady=(1,0))
        # Body : tableau prix
        body=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8);body.pack(fill="both",expand=True,padx=20,pady=8)
        # En-têtes du tableau
        head=ctk.CTkFrame(body,fg_color="transparent");head.pack(fill="x",padx=14,pady=(14,4))
        for col,(txt,w) in enumerate([("Carburant",110),("Prix vente",120),("Prix achat",120),("Marge unit.",120),("",0)]):
            lbl=ctk.CTkLabel(head,text=txt,font=("Segoe UI",10,"bold"),text_color=C["t3"],width=w,anchor="w")
            lbl.grid(row=0,column=col,sticky="w",padx=(0,8))
        # Séparateur
        sep=ctk.CTkFrame(body,fg_color=C["border2"],height=1);sep.pack(fill="x",padx=14,pady=(2,8))
        # Lignes (1 par carburant)
        self.rows={}
        for carb,couleur,full in [("sp",C["vig_red"],"Sans plomb (SP95)"),
                                   ("go",C["vig_blue"],"Gazole (GO)"),
                                   ("gnr",C["vig_gold"],"GNR")]:
            row=ctk.CTkFrame(body,fg_color=C["panel"],corner_radius=6,height=56);row.pack(fill="x",padx=14,pady=4)
            row.pack_propagate(False)
            inner=ctk.CTkFrame(row,fg_color="transparent");inner.pack(fill="both",expand=True,padx=12,pady=8)
            # Pastille couleur + nom
            pill=ctk.CTkFrame(inner,fg_color=couleur,width=12,height=36,corner_radius=3)
            pill.grid(row=0,column=0,padx=(0,10),sticky="ns");pill.grid_propagate(False)
            ctk.CTkLabel(inner,text=full,font=("Segoe UI",12,"bold"),text_color=C["t1"],
                         width=110,anchor="w").grid(row=0,column=1,sticky="w",padx=(0,8))
            pv_lbl=ctk.CTkLabel(inner,text="—",font=(FONT_NUM,15,"bold"),text_color=C["green"],
                                width=120,anchor="w")
            pv_lbl.grid(row=0,column=2,sticky="w",padx=(0,8))
            pa_lbl=ctk.CTkLabel(inner,text="—",font=(FONT_NUM,13),text_color=C["t2"],
                                width=120,anchor="w")
            pa_lbl.grid(row=0,column=3,sticky="w",padx=(0,8))
            mg_lbl=ctk.CTkLabel(inner,text="—",font=(FONT_NUM,13),text_color=C["amber"],
                                width=120,anchor="w")
            mg_lbl.grid(row=0,column=4,sticky="w",padx=(0,8))
            self.rows[carb]={"pv":pv_lbl,"pa":pa_lbl,"mg":mg_lbl}
        # Note bas + bouton fermer
        self.note_lbl=ctk.CTkLabel(body,text="",font=("Segoe UI",9),text_color=C["t3"],wraplength=600,justify="left")
        self.note_lbl.pack(anchor="w",padx=14,pady=(8,4))
        # Footer
        footer=ctk.CTkFrame(self,fg_color="transparent",height=56);footer.pack(side="bottom",fill="x",padx=20,pady=10);footer.pack_propagate(False)
        ctk.CTkButton(footer,text="Fermer",width=120,height=36,
                      fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                      border_width=1,border_color=C["border2"],
                      font=("Segoe UI",11),corner_radius=8,
                      command=self.destroy).pack(side="right")
        # Bouton "Modifier passage de mois" : permet de revenir sur les ventes 0h-6h
        # saisies pour le mois en cours (utile en cas d'erreur).
        ctk.CTkButton(footer,text="\U0001f4c5 Modifier passage de mois",width=240,height=36,
                      fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                      border_width=1,border_color=C["border2"],
                      font=("Segoe UI",10),corner_radius=8,
                      command=self._open_passage_mois_modif).pack(side="right",padx=(0,8))
        # Bouton "Modifier prix" : garde-fou pour corriger manuellement les prix au cas
        # où la lecture auto Pre_vision donnerait une mauvaise valeur. Le hub re-écrit
        # automatiquement à chaque refresh depuis Pre_vision, donc l'utilisateur doit
        # corriger Pre_vision en parallèle pour que la modif tienne dans la durée.
        ctk.CTkButton(footer,text="\u270f Modifier prix",width=160,height=36,
                      fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                      border_width=1,border_color=C["border2"],
                      font=("Segoe UI",10),corner_radius=8,
                      command=self._open_prix_modif).pack(side="right",padx=(0,8))
        # Charger les prix
        self._load_prix()

    def _open_prix_modif(self):
        """Ouvre la fenêtre de saisie manuelle des prix pour le mois en cours."""
        parent=self.master
        self.destroy()
        try:
            dlg=PrixModifDlg(parent)
            parent.wait_window(dlg)
        except Exception as e: print(f"[prix modif] {e}")
    def _open_passage_mois_modif(self):
        """Ouvre PassageMoisDlg en mode modification pour le mois en cours."""
        today=date.today()
        # Fermer cette fenêtre Prix avant d'ouvrir PassageMoisDlg pour éviter
        # 2 modales empilées
        parent=self.master
        self.destroy()
        try:
            dlg=PassageMoisDlg(parent,target_year=today.year,target_month=today.month,edit_mode=True)
            parent.wait_window(dlg)
        except Exception as e: print(f"[passage mois modif] {e}")
    def _load_prix(self):
        """Charge les prix du mois en cours depuis prix_historique.cfg et remplit le tableau."""
        today=date.today()
        prix=get_prix_for_month(today.year,today.month) or {}
        if not prix:
            self.note_lbl.configure(text="\u26a0  Aucun prix mémorisé pour le mois en cours. Les prix sont normalement mis à jour automatiquement à chaque refresh du hub si Prévision compte.xlsx est accessible.")
            return
        for carb in ("sp","go","gnr"):
            pv=sf(prix.get(f"pv_{carb}",0))
            pa=sf(prix.get(f"pa_{carb}",0))
            mg=sf(prix.get(f"marge_{carb}",pv-pa if (pv and pa) else 0))
            # Précision : 2 décimales pour PV (= prix pompe affiché aux clients), 5 décimales
            # pour PA et marge (la marge réglementaire est définie à 5 décimales : 0,14892 €/L).
            # Format français : virgule décimale.
            self.rows[carb]["pv"].configure(text=(f"{pv:.2f} €/L".replace(".",",")) if pv else "—")
            self.rows[carb]["pa"].configure(text=(f"{pa:.5f} €/L".replace(".",",")) if pa else "—")
            self.rows[carb]["mg"].configure(text=(f"{mg:.5f} €/L".replace(".",",")) if mg else "—")
        # Source + horodatage
        last_seen=prix.get("last_seen","")
        src=prix.get("source","")
        try:
            if last_seen:
                d=datetime.fromisoformat(last_seen)
                last_seen=d.strftime("%d/%m/%Y à %Hh%M")
        except Exception as _e: _log_silent_err(exc=_e)
        note=f"Source : {src or 'manuel'}"
        if last_seen: note+=f" — Dernière mise à jour le {last_seen}"
        note+="\nLes prix de vente sont fixés mensuellement par la préfecture de Martinique (Observatoire des Prix). Le prix d'achat est déduit (PV - marge réglementaire applicable au palier en cours)."
        self.note_lbl.configure(text=note)


class PrixModifDlg(ctk.CTkToplevel):
    """Garde-fou : permet de corriger manuellement les prix du mois en cours dans
    prix_historique.cfg si l'auto-import Pre_vision donne une mauvaise valeur.

    Attention : le hub écrase prix_historique.cfg à chaque refresh avec ce qu'il lit
    dans Prévision compte. Donc une correction ici ne tient que tant que Pre_vision
    n'a pas été re-lu. Pour qu'elle tienne dans la durée, l'utilisateur doit aussi
    corriger Prévision compte en parallèle. Un avertissement est affiché.
    """
    def __init__(self,parent):
        super().__init__(parent)
        self.parent_app=parent
        self.title("\u270f Modifier prix carburant \u2014 DISTRICARB HUB")
        self.geometry("640x540");self.minsize(560,460)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        # Header
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=64);hdr.pack(fill="x",padx=20,pady=(16,4));hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\u270f  Modifier les prix carburant",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w")
        today=date.today()
        mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin",
                   "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        ctk.CTkLabel(hdr,text=f"Mois en cours : {mois_noms[today.month-1]} {today.year}",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(anchor="w",pady=(1,0))
        # Avertissement
        warn=ctk.CTkFrame(self,fg_color="#2A1F0D",corner_radius=6)
        warn.pack(fill="x",padx=20,pady=(4,8))
        ctk.CTkLabel(warn,text="\u26a0  Garde-fou : ces valeurs sont ré-écrasées à chaque refresh par la lecture auto de Prévision compte. Pour une correction durable, modifie aussi Prévision compte.xlsx.",
                     font=("Segoe UI",9),text_color=C["amber"],
                     justify="left",wraplength=560,anchor="w").pack(anchor="w",padx=12,pady=8)
        # Body
        body=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8)
        body.pack(fill="both",expand=True,padx=20,pady=4)
        # En-têtes
        hdr_grid=ctk.CTkFrame(body,fg_color="transparent")
        hdr_grid.pack(fill="x",padx=14,pady=(14,4))
        ctk.CTkLabel(hdr_grid,text="Carburant",font=("Segoe UI",10,"bold"),text_color=C["t3"],width=130,anchor="w").pack(side="left")
        ctk.CTkLabel(hdr_grid,text="Prix vente (€/L)",font=("Segoe UI",10,"bold"),text_color=C["t3"],width=160,anchor="center").pack(side="left",padx=(0,8))
        ctk.CTkLabel(hdr_grid,text="Prix achat (€/L)",font=("Segoe UI",10,"bold"),text_color=C["t3"],width=160,anchor="center").pack(side="left")
        # 3 lignes carburants
        self.entries={}
        prix=get_prix_for_month(today.year,today.month) or {}
        for carb,nom,couleur in [("sp","Sans plomb",C["vig_red"]),
                                  ("go","Gazole",C["vig_blue"]),
                                  ("gnr","GNR",C["vig_gold"])]:
            row=ctk.CTkFrame(body,fg_color=C["panel"],corner_radius=6)
            row.pack(fill="x",padx=14,pady=4)
            row_in=ctk.CTkFrame(row,fg_color="transparent")
            row_in.pack(fill="x",padx=12,pady=10)
            pill=ctk.CTkFrame(row_in,fg_color=couleur,width=10,height=28,corner_radius=2)
            pill.pack(side="left",padx=(0,10));pill.pack_propagate(False)
            ctk.CTkLabel(row_in,text=nom,font=("Segoe UI",11,"bold"),text_color=C["t1"],
                         width=100,anchor="w").pack(side="left")
            pv_e=ctk.CTkEntry(row_in,placeholder_text="0,00",width=140,height=32,
                                fg_color=C["bg"],border_color=C["border2"],
                                font=(FONT_NUM,13),justify="right")
            pv_e.pack(side="left",padx=(8,8))
            pv_val=sf(prix.get(f"pv_{carb}",0))
            if pv_val: pv_e.insert(0,f"{pv_val:.4f}".replace(".",","))
            pa_e=ctk.CTkEntry(row_in,placeholder_text="0,00000",width=140,height=32,
                                fg_color=C["bg"],border_color=C["border2"],
                                font=(FONT_NUM,13),justify="right")
            pa_e.pack(side="left",padx=(8,0))
            pa_val=sf(prix.get(f"pa_{carb}",0))
            if pa_val: pa_e.insert(0,f"{pa_val:.5f}".replace(".",","))
            self.entries[carb]={"pv":pv_e,"pa":pa_e}
        # Marge unitaire (commune aux 3 carburants depuis sept 2024)
        marge_frame=ctk.CTkFrame(body,fg_color=C["panel"],corner_radius=6)
        marge_frame.pack(fill="x",padx=14,pady=(8,4))
        marge_in=ctk.CTkFrame(marge_frame,fg_color="transparent")
        marge_in.pack(fill="x",padx=12,pady=10)
        ctk.CTkLabel(marge_in,text="Marge réglementaire (€/L)",font=("Segoe UI",11,"bold"),
                     text_color=C["t1"],anchor="w").pack(side="left")
        self.marge_e=ctk.CTkEntry(marge_in,placeholder_text="0,14892",width=140,height=32,
                                    fg_color=C["bg"],border_color=C["border2"],
                                    font=(FONT_NUM,13),justify="right")
        self.marge_e.pack(side="right")
        m_val=sf(prix.get("marge_unit",0))
        if m_val: self.marge_e.insert(0,f"{m_val:.5f}".replace(".",","))
        # Status
        self.status_lbl=ctk.CTkLabel(self,text="",font=("Segoe UI",10),text_color=C["t2"],anchor="w",justify="left")
        self.status_lbl.pack(side="bottom",fill="x",padx=20,pady=(0,4))
        # Footer
        footer=ctk.CTkFrame(self,fg_color="transparent",height=52);footer.pack(side="bottom",fill="x",padx=20,pady=(4,10));footer.pack_propagate(False)
        ctk.CTkButton(footer,text="Annuler",width=110,height=36,
                      fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                      border_width=1,border_color=C["border2"],
                      font=("Segoe UI",11),corner_radius=8,
                      command=self.destroy).pack(side="right",padx=(0,8))
        ctk.CTkButton(footer,text="\u2713 Enregistrer",width=160,height=36,
                      fg_color="#2DA84A",hover_color="#1F7A35",text_color="#FFF",
                      font=("Segoe UI",11,"bold"),corner_radius=8,
                      command=self._enregistrer).pack(side="right")

    def _parse(self,e):
        txt=e.get().strip().replace(" ","").replace("\u202f","").replace(",",".")
        try: return float(txt) if txt else 0
        except Exception as _e: _log_silent_err(exc=_e); return 0

    def _enregistrer(self):
        """Mute prix_historique.cfg avec les valeurs saisies. Préserve les autres champs
        (passage_mois, marge_boutique_taux, etc.) grâce à la mutation in-place."""
        try:
            today=date.today()
            histo=_load_prix_histo()
            prix=histo.setdefault("prix",{})
            key=f"{today.year:04d}-{today.month:02d}"
            existing=prix.get(key,{})
            for carb in ("sp","go","gnr"):
                pv=self._parse(self.entries[carb]["pv"])
                pa=self._parse(self.entries[carb]["pa"])
                if pv: existing[f"pv_{carb}"]=round(pv,4)
                if pa: existing[f"pa_{carb}"]=round(pa,5)
            mu=self._parse(self.marge_e)
            if mu:
                existing["marge_unit"]=round(mu,5)
                existing["marge_sp"]=round(mu,5)
                existing["marge_go"]=round(mu,5)
                existing["marge_gnr"]=round(mu,5)
            existing["last_seen"]=datetime.now().isoformat()
            existing["source"]="manuel"
            prix[key]=existing
            _save_prix_histo(histo)
            self.status_lbl.configure(text="\u2713 Prix enregistrés",text_color=C["green"])
            self.after(800,self.destroy)
        except Exception as e:
            self.status_lbl.configure(text=f"\u26a0 Erreur : {e}",text_color=C["red"])


class PassageMoisDlg(ctk.CTkToplevel):
    """Dialogue de saisie du passage de mois. À chaque 1er du mois (à partir de 6h),
    on capture les ventes 0h-6h pour calculer correctement l'effet spéculation :

    - Le stock 6h le 1er du mois est lu automatiquement depuis Prévision compte
      (cellules C7/C8/C9 de l'onglet du jour de la semaine, selon le cycle Sem 1/2)
    - Les ventes 0h-6h du 1er sont saisies par l'utilisateur (depuis ticket caisse)
    - Le hub reconstitue le stock à minuit : stock_minuit = stock_6h + ventes_0h_6h
      (les ventes 0h-6h sont déjà au nouveau prix, donc font partie du stock pivot)
    - Effet = stock_minuit × (PV_nouveau - PV_ancien)

    Args:
        parent: fenêtre Hub principale
        target_year/target_month: mois CIBLE (ex: 2026, 6 = passage mai → juin)
        edit_mode: si True, charge les valeurs existantes pour modification
    """
    def __init__(self,parent,target_year=None,target_month=None,edit_mode=False):
        super().__init__(parent)
        self.parent_app=parent
        self.target_year=target_year or date.today().year
        self.target_month=target_month or date.today().month
        self.edit_mode=edit_mode
        self.title(f"\U0001f4c5 Passage de mois \u2014 DISTRICARB HUB")
        self.geometry("720x680");self.minsize(640,580)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        # Header
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=64);hdr.pack(fill="x",padx=20,pady=(16,4));hdr.pack_propagate(False)
        title_txt="Modifier le passage de mois" if edit_mode else "Saisie du passage de mois"
        ctk.CTkLabel(hdr,text=f"\U0001f4c5  {title_txt}",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w")
        mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin",
                   "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        prev_mo=self.target_month-1 if self.target_month>1 else 12
        prev_yr=self.target_year if self.target_month>1 else self.target_year-1
        ctk.CTkLabel(hdr,text=f"Passage de {mois_noms[prev_mo-1]} {prev_yr} \u2192 {mois_noms[self.target_month-1]} {self.target_year}",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(anchor="w",pady=(1,0))
        # Body : panneau scrollable
        body=ctk.CTkScrollableFrame(self,fg_color=C["card"],corner_radius=8,
                                     scrollbar_button_color=C["border2"])
        body.pack(fill="both",expand=True,padx=20,pady=8)
        # === Section 1 : Stock 6h (auto-lu depuis Pre_vision si possible, éditable) ===
        ctk.CTkLabel(body,text="\U0001f4e6  Stock matin 6h00",font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(anchor="w",padx=14,pady=(14,4))
        ctk.CTkLabel(body,text="Pré-rempli automatiquement depuis Prévision compte si on est le 1er du mois et que le fichier est à jour.\nSinon, saisis manuellement le stock cuve à 6h00 le 1er.",
                     font=("Segoe UI",9),text_color=C["t3"],justify="left").pack(anchor="w",padx=14,pady=(0,8))
        self.stock_entries={}
        stock_frame=ctk.CTkFrame(body,fg_color=C["panel"],corner_radius=6)
        stock_frame.pack(fill="x",padx=14,pady=(0,12))
        stock_inner=ctk.CTkFrame(stock_frame,fg_color="transparent")
        stock_inner.pack(fill="x",padx=12,pady=10)
        for carb,nom,couleur in [("sp","Sans plomb",C["vig_red"]),
                                  ("go","Gazole",C["vig_blue"]),
                                  ("gnr","GNR",C["vig_gold"])]:
            row=ctk.CTkFrame(stock_inner,fg_color="transparent")
            row.pack(fill="x",pady=2)
            pill=ctk.CTkFrame(row,fg_color=couleur,width=10,height=24,corner_radius=2)
            pill.pack(side="left",padx=(0,10));pill.pack_propagate(False)
            ctk.CTkLabel(row,text=nom,font=("Segoe UI",11,"bold"),text_color=C["t1"],
                         width=120,anchor="w").pack(side="left")
            entry=ctk.CTkEntry(row,placeholder_text="0",width=130,height=30,
                                fg_color=C["bg"],border_color=C["border2"],
                                font=(FONT_NUM,13),justify="right")
            entry.pack(side="left",padx=(0,4))
            ctk.CTkLabel(row,text="L",font=("Segoe UI",10),text_color=C["t3"]).pack(side="left")
            self.stock_entries[carb]=entry
            entry.bind("<KeyRelease>",lambda ev:self._refresh_apercu())
        # === Section 2 : Ventes 0h-6h (saisie manuelle) ===
        ctk.CTkLabel(body,text="\U0001f4dd  Ventes 0h00 \u2014 6h00",font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(anchor="w",padx=14,pady=(8,4))
        ctk.CTkLabel(body,text="Lis ton ticket caisse de 6h00 et saisis les litres écoulés entre 0h et 6h.\nCes litres sont déjà au NOUVEAU prix : on les ajoute au stock 6h pour reconstituer le stock à minuit.",
                     font=("Segoe UI",9),text_color=C["t3"],justify="left").pack(anchor="w",padx=14,pady=(0,8))
        ventes_frame=ctk.CTkFrame(body,fg_color=C["panel"],corner_radius=6)
        ventes_frame.pack(fill="x",padx=14,pady=(0,12))
        ventes_inner=ctk.CTkFrame(ventes_frame,fg_color="transparent")
        ventes_inner.pack(fill="x",padx=12,pady=10)
        self.ventes_entries={}
        for carb,nom,couleur in [("sp","Sans plomb",C["vig_red"]),
                                  ("go","Gazole",C["vig_blue"]),
                                  ("gnr","GNR",C["vig_gold"])]:
            row=ctk.CTkFrame(ventes_inner,fg_color="transparent")
            row.pack(fill="x",pady=2)
            pill=ctk.CTkFrame(row,fg_color=couleur,width=10,height=24,corner_radius=2)
            pill.pack(side="left",padx=(0,10));pill.pack_propagate(False)
            ctk.CTkLabel(row,text=nom,font=("Segoe UI",11,"bold"),text_color=C["t1"],
                         width=120,anchor="w").pack(side="left")
            entry=ctk.CTkEntry(row,placeholder_text="0",width=130,height=30,
                                fg_color=C["bg"],border_color=C["border2"],
                                font=(FONT_NUM,13),justify="right")
            entry.pack(side="left",padx=(0,4))
            ctk.CTkLabel(row,text="L",font=("Segoe UI",10),text_color=C["t3"]).pack(side="left")
            self.ventes_entries[carb]=entry
            entry.bind("<KeyRelease>",lambda ev:self._refresh_apercu())
        # Touche Entrée : passer au champ suivant. SP stock → GO stock → GNR stock →
        # SP ventes → GO ventes → GNR ventes (s'arrête sur le dernier).
        order=[self.stock_entries["sp"],self.stock_entries["go"],self.stock_entries["gnr"],
               self.ventes_entries["sp"],self.ventes_entries["go"],self.ventes_entries["gnr"]]
        for i,e in enumerate(order):
            if i<len(order)-1:
                nxt=order[i+1]
                e.bind("<Return>",lambda ev,n=nxt:(n.focus_set(),"break"))
        # === Section 3 : Aperçu du calcul ===
        ctk.CTkLabel(body,text="\u2728  Aperçu effet spéculation",font=("Segoe UI",12,"bold"),text_color=C["green"]).pack(anchor="w",padx=14,pady=(8,4))
        self.apercu_box=ctk.CTkFrame(body,fg_color=C["panel"],corner_radius=6)
        self.apercu_box.pack(fill="x",padx=14,pady=(0,12))
        self.apercu_lbl=ctk.CTkLabel(self.apercu_box,text="Saisis les ventes 0h-6h pour voir le calcul.",
                                      font=("Segoe UI",10),text_color=C["t3"],
                                      justify="left",wraplength=620)
        self.apercu_lbl.pack(anchor="w",padx=14,pady=12)
        # Status au-dessus du footer (pleine largeur, peut être long)
        self.status_lbl=ctk.CTkLabel(self,text="",font=("Segoe UI",10),text_color=C["t2"],
                                      anchor="w",justify="left")
        self.status_lbl.pack(side="bottom",fill="x",padx=20,pady=(0,4))
        # Footer : boutons uniquement
        footer=ctk.CTkFrame(self,fg_color="transparent",height=52);footer.pack(side="bottom",fill="x",padx=20,pady=(4,10));footer.pack_propagate(False)
        # En mode auto : bouton "Me rappeler dans..." (snooze) + bouton "Plus tard"
        # En mode modif : bouton "Annuler"
        if not edit_mode:
            # Dropdown "Me rappeler dans..." en DATES ABSOLUES via make_snooze_options.
            self._snooze_labels,self._snooze_mapping=make_snooze_options([1,2,4,6])
            self.snooze_var=ctk.StringVar(value="Me rappeler dans...")
            ctk.CTkOptionMenu(footer,values=self._snooze_labels,
                              variable=self.snooze_var,
                              font=("Segoe UI",10),width=180,height=38,
                              fg_color=C["panel"],button_color=C["border2"],
                              button_hover_color=C["border"],text_color=C["t1"],
                              command=self._snooze).pack(side="right",padx=(0,8))
            ctk.CTkButton(footer,text="Plus tard",width=100,height=38,
                          fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                          border_width=1,border_color=C["border2"],
                          font=("Segoe UI",11),corner_radius=8,
                          command=self.destroy).pack(side="right",padx=(0,8))
        else:
            ctk.CTkButton(footer,text="Annuler",width=100,height=38,
                          fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                          border_width=1,border_color=C["border2"],
                          font=("Segoe UI",11),corner_radius=8,
                          command=self.destroy).pack(side="right",padx=(0,8))
        ctk.CTkButton(footer,text="\u2713 Enregistrer",width=160,height=38,
                      fg_color="#2DA84A",hover_color="#1F7A35",text_color="#FFF",
                      font=("Segoe UI",11,"bold"),corner_radius=8,
                      command=self._enregistrer).pack(side="right")
        # Charger stock 6h auto + valeurs existantes si modif
        self._charger_stock_6h()
        if edit_mode:
            self._charger_valeurs_existantes()
        self._refresh_apercu()

    def _snooze(self,choix):
        """Stocke un horodatage de rappel : le hub ne ré-ouvrira pas la popup
        avant cette heure. Persistant via passage_mois_snooze.cfg."""
        heures=self._snooze_mapping.get(choix,2)
        rappel=datetime.now()+timedelta(hours=heures)
        try:
            snooze_path=APP_DIR/"passage_mois_snooze.cfg"
            snooze=load_json(snooze_path) or {}
            key=f"{self.target_year:04d}-{self.target_month:02d}"
            snooze[key]=rappel.isoformat()
            save_json(snooze_path,snooze)
        except Exception as e: print(f"[snooze] {e}")
        self.destroy()

    def _charger_stock_6h(self):
        """Lit C7/C8/C9 dans Prévision compte selon l'onglet du jour ET pré-remplit les entries.
        Si on n'est PAS le 1er du mois, on n'écrase pas les entries (le stock 6h actuel
        Pre_vision n'est plus le stock du 1er, il faut une saisie manuelle)."""
        # Si on n'est pas le 1er ET pas en mode modif → laisser les entries vides
        # pour que l'utilisateur saisisse manuellement.
        # En mode modif, on charge les valeurs existantes via _charger_valeurs_existantes.
        today=date.today()
        if today.day!=1 and not self.edit_mode:
            self.status_lbl.configure(
                text=f"\u2139 Saisis manuellement le stock 6h du 01/{self.target_month:02d} "
                     f"(on n'est pas le 1er, lecture auto désactivée)",
                text_color=C["amber"])
            return
        try:
            cfg=load_json(CONFIG_FILE) or {}
            # Le hub stocke le path sous la clé "prevision" (cf _open_livraisons)
            pv_path=cfg.get("prevision","") or cfg.get("pre_vision_path","")
            if not pv_path or not Path(pv_path).exists():
                self.status_lbl.configure(text="\u26a0 Prévision compte.xlsx introuvable \u2014 saisie manuelle",text_color=C["amber"])
                return
            import openpyxl
            # Copie en temporaire pour contourner verrous OneDrive/Excel ouvert
            tmp=copy_to_temp(pv_path)
            wb=openpyxl.load_workbook(tmp,data_only=True)
            # Quel onglet ? Selon le cycle de la semaine + jour de la semaine
            # On utilise la même logique que ailleurs dans le code
            today=date.today()
            wd=today.weekday()  # 0=lundi
            week=get_cycle_week()
            # Onglets jours de la semaine
            sheets_w1=["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
            sheets_w2=["Lundi2","Mardi2","Merc2","Jeudi 2","Vend 2","Sam 2","Dim2"]
            sheets=sheets_w1 if week==1 else sheets_w2
            sheet_name=sheets[wd] if wd<len(sheets) else None
            if not sheet_name or sheet_name not in wb.sheetnames:
                self.status_lbl.configure(text=f"\u26a0 Onglet '{sheet_name}' introuvable dans Pre_vision",text_color=C["red"])
                return
            ws=wb[sheet_name]
            sp=sf(ws["C7"].value or 0)
            go=sf(ws["C8"].value or 0)
            gnr=sf(ws["C9"].value or 0)
            # Pré-remplir les entries (l'utilisateur peut corriger si besoin)
            for carb,val in [("sp",sp),("go",go),("gnr",gnr)]:
                self.stock_entries[carb].delete(0,"end")
                self.stock_entries[carb].insert(0,str(int(val)))
            self.status_lbl.configure(text=f"\u2713 Stock 6h pré-rempli depuis l'onglet '{sheet_name}'",text_color=C["green"])
        except Exception as e:
            self.status_lbl.configure(text=f"\u26a0 Erreur lecture Pre_vision : {e} \u2014 saisis manuellement",text_color=C["amber"])

    def _charger_valeurs_existantes(self):
        """Pré-remplit les champs (stock + ventes 0h-6h) avec les valeurs déjà saisies (mode modif)."""
        passage=get_passage_mois(self.target_year,self.target_month)
        if not passage: return
        sp=passage["stock_pivot"];vb=passage["ventes_avant_6h"]
        for carb in ("sp","go","gnr"):
            self.stock_entries[carb].delete(0,"end")
            self.stock_entries[carb].insert(0,str(int(sp.get(carb,0))))
            self.ventes_entries[carb].delete(0,"end")
            self.ventes_entries[carb].insert(0,str(int(vb.get(carb,0))))

    def _get_stock(self):
        """Retourne dict stock 6h depuis les entries (saisi ou pré-rempli auto)."""
        out={}
        for carb in ("sp","go","gnr"):
            txt=self.stock_entries[carb].get().strip().replace(" ","").replace("\u202f","").replace(",",".")
            try: out[carb]=float(txt) if txt else 0
            except Exception: out[carb]=0
        return out

    def _get_ventes(self):
        """Retourne dict ventes 0h-6h depuis les entries."""
        out={}
        for carb in ("sp","go","gnr"):
            txt=self.ventes_entries[carb].get().strip().replace(" ","").replace("\u202f","").replace(",",".")
            try: out[carb]=float(txt) if txt else 0
            except Exception: out[carb]=0
        return out

    def _refresh_apercu(self):
        """Recalcule et affiche l'effet spéculation prévisionnel."""
        ventes=self._get_ventes()
        stock=self._get_stock()
        eff=calc_effet_speculation(
            self.target_year,self.target_month,
            stock["sp"],stock["go"],stock["gnr"],
            ventes["sp"],ventes["go"],ventes["gnr"],
        )
        if not eff:
            self.apercu_lbl.configure(text="\u26a0 Calcul impossible : prix manquants pour le mois actuel ou précédent dans prix_historique.cfg.",text_color=C["amber"])
            return
        # Construire un récap propre
        prev_pv=eff["prix_avant"];curr_pv=eff["prix_apres"]
        lines=[]
        lines.append(f"Stock pivot effectif (= stock 6h + ventes 0h-6h, reconstitution stock minuit) :")
        for carb,nom in [("sp","SP"),("go","GO"),("gnr","GNR")]:
            sp_v=int(stock[carb]+ventes[carb])
            lines.append(f"  {nom} : {int(stock[carb]):,} + {int(ventes[carb]):,} = {sp_v:,} L".replace(",","\u202f"))
        lines.append("")
        lines.append(f"Évolution des PV ce mois :")
        for carb,nom in [("sp","SP"),("go","GO"),("gnr","GNR")]:
            delta=sf(curr_pv.get(carb,0))-sf(prev_pv.get(carb,0))
            signe="+" if delta>=0 else ""
            lines.append(f"  {nom} : {sf(prev_pv.get(carb,0)):.4f} \u2192 {sf(curr_pv.get(carb,0)):.4f} \u20ac/L  ({signe}{delta:.4f})".replace(".",","))
        lines.append("")
        lines.append(f"Effet par carburant :")
        lines.append(f"  SP  : {eff['sp']:+,.2f} \u20ac".replace(",","\u202f").replace(".",","))
        lines.append(f"  GO  : {eff['go']:+,.2f} \u20ac".replace(",","\u202f").replace(".",","))
        lines.append(f"  GNR : {eff['gnr']:+,.2f} \u20ac".replace(",","\u202f").replace(".",","))
        lines.append("")
        signe_t="+" if eff["total"]>=0 else ""
        col=C["green"] if eff["total"]>=0 else C["red"]
        lines.append(f"\u2192 Effet total : {signe_t}{eff['total']:,.2f} \u20ac".replace(",","\u202f").replace(".",","))
        self.apercu_lbl.configure(text="\n".join(lines),text_color=C["t1"])

    def _enregistrer(self):
        """Sauve le passage de mois dans prix_historique.cfg + journal d'événements."""
        ventes=self._get_ventes()
        stock=self._get_stock()
        if stock["sp"]+stock["go"]+stock["gnr"]==0:
            self.status_lbl.configure(text="\u26a0 Stock 6h vide \u2014 saisis-le ou attends que Pre_vision soit à jour",text_color=C["red"]);return
        ok=save_passage_mois(self.target_year,self.target_month,stock,ventes)
        if not ok:
            self.status_lbl.configure(text="\u26a0 Erreur lors de l'enregistrement",text_color=C["red"]);return
        # Calculer effet pour le journal
        eff=calc_effet_speculation(
            self.target_year,self.target_month,
            stock["sp"],stock["go"],stock["gnr"],
            ventes["sp"],ventes["go"],ventes["gnr"],
        )
        # Hook journal d'événements
        try:
            mois_noms=["janvier","f\u00e9vrier","mars","avril","mai","juin",
                       "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
            prev_mo=self.target_month-1 if self.target_month>1 else 12
            prev_yr=self.target_year if self.target_month>1 else self.target_year-1
            data={
                "year":self.target_year,"month":self.target_month,
                "stock_pivot":stock,"ventes_avant_6h":ventes,
                "effet_total":eff["total"] if eff else 0,
                "label":f"Passage {mois_noms[prev_mo-1]} {prev_yr} \u2192 {mois_noms[self.target_month-1]} {self.target_year}",
            }
            add_evenement("passage_mois",data,
                          commentaire=f"Effet spéculation : {(eff['total'] if eff else 0):+.2f} \u20ac")
        except Exception as e:
            print(f"[passage mois] hook journal : {e}")
        self.status_lbl.configure(text="\u2713 Passage de mois enregistré",text_color=C["green"])
        self.after(800,self.destroy)


class EventActionDlg(ctk.CTkToplevel):
    """Popup d'action rapide sur un événement 'non_traite' depuis le journal.
    Permet de : marquer Résolu, Annulé, ou Ajouter un commentaire daté.
    Modifie l'événement directement dans evenements.cfg puis ferme."""
    def __init__(self,parent,evt_id):
        super().__init__(parent)
        self.title("Action sur \u00e9v\u00e9nement \u2014 DISTRICARB HUB")
        self.geometry("460x320");self.minsize(420,280)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.evt_id=evt_id
        self.modified=False  # Indique si une modif a été faite (pour refresh parent)
        # Charger l'événement
        all_evt=load_json(EVENEMENTS_FILE) or {}
        events=all_evt.get("events",[])
        evt=next((e for e in events if e.get("id")==evt_id),None)
        if not evt:
            self.destroy();return
        body=ctk.CTkFrame(self,fg_color="transparent")
        body.pack(fill="both",expand=True,padx=24,pady=20)
        ctk.CTkLabel(body,text="Action sur \u00e9v\u00e9nement non trait\u00e9",
                     font=("Segoe UI",14,"bold"),text_color=C["t1"]).pack(anchor="w")
        # Description courte de l'événement (rappel contexte)
        type_evt=evt.get("type","")
        data=evt.get("data",{}) or {}
        if type_evt=="livraison_reporter":
            carb=data.get("carburant","")
            jour=data.get("date","") or data.get("jour","")
            surplus=data.get("surplus","")
            desc=f"Livraison {carb} \u00e0 reporter le {jour} \u2014 surplus cuve {surplus} L"
        elif type_evt=="rupture":
            desc=f"Rupture de {data.get('carburant','')} le {data.get('jour','')}"
        elif type_evt=="anomalie":
            desc=f"Anomalie litrage le {data.get('jour','')} \u2014 \u00e9cart {data.get('ecart_pct','')}%"
        elif type_evt=="pont":
            d_deb=data.get("date_debut","");d_fin=data.get("date_fin","")
            try:
                ddt=datetime.fromisoformat(d_deb).date() if d_deb else None
                dft=datetime.fromisoformat(d_fin).date() if d_fin else ddt
                JC_short=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
                d_deb_lisible=f"{JC_short[ddt.weekday()]} {ddt.strftime('%d/%m')}" if ddt else d_deb
                d_fin_lisible=f"{JC_short[dft.weekday()]} {dft.strftime('%d/%m')}" if dft else d_fin
                # Factorisé (Étape 3, 27/05/2026) via districarb_core.trous.qualifier_trou
                if ddt and dft:
                    terme=qualifier_trou({"start_date":ddt,"duree":(dft-ddt).days+1})
                else:
                    terme="Pont"
                desc=f"{terme} du {d_deb_lisible} au {d_fin_lisible}"
            except Exception:
                desc=f"Pont du {d_deb} au {d_fin}"
        else:
            desc=type_evt
        ctk.CTkLabel(body,text=desc,font=("Segoe UI",10),text_color=C["t3"],
                     wraplength=400,justify="left",anchor="w").pack(anchor="w",pady=(6,18))
        ctk.CTkLabel(body,text="Que voulez-vous faire ?",font=("Segoe UI",11),
                     text_color=C["t2"]).pack(anchor="w",pady=(0,8))
        # Trois boutons d'action principaux côte à côte
        btns=ctk.CTkFrame(body,fg_color="transparent")
        btns.pack(fill="x",pady=(4,0))
        ctk.CTkButton(btns,text="\u2713 R\u00e9solu",fg_color=C["green"],hover_color="#5BB54F",
                      command=self._mark_resolu,width=120,height=36,
                      font=("Segoe UI",11,"bold")).pack(side="left",padx=(0,8))
        ctk.CTkButton(btns,text="\u2717 Annul\u00e9",fg_color=C["red"],hover_color="#C42B3A",
                      command=self._mark_annule,width=120,height=36,
                      font=("Segoe UI",11,"bold")).pack(side="left",padx=(0,8))
        ctk.CTkButton(btns,text="\U0001f4ac Commenter",fg_color=C["amber"],hover_color="#C7741B",
                      command=self._add_comment,width=140,height=36,
                      font=("Segoe UI",11,"bold")).pack(side="left")
        # Bouton secondaire : fermer sans action (modale légère)
        ctk.CTkButton(body,text="Fermer sans action",fg_color="transparent",
                      border_width=1,border_color=C["border2"],text_color=C["t2"],
                      hover_color=C["card"],command=self.destroy,width=180).pack(anchor="e",pady=(24,0))

    def _mark_resolu(self):
        """Marque l'événement comme résolu : statut=resolu, lu=True.
        Synchronise aussi le tableau de notifications en silenciant l'alerte correspondante
        dans popup_silence.cfg (sinon elle reviendrait au prochain refresh)."""
        if self._update_evt(statut="resolu",lu=True,note="\u2713 Marqu\u00e9 r\u00e9solu"):
            self.modified=True
            self._silence_corresponding_alert()
        self.destroy()

    def _mark_annule(self):
        """Marque l'événement comme annulé : statut=annule, lu=True.
        Synchronise aussi le tableau de notifications en silenciant l'alerte correspondante."""
        if self._update_evt(statut="annule",lu=True,note="\u2717 Marqu\u00e9 annul\u00e9"):
            self.modified=True
            self._silence_corresponding_alert()
        self.destroy()

    def _silence_corresponding_alert(self):
        """Quand l'utilisateur marque Résolu/Annulé un événement non_traité depuis le journal,
        on doit aussi silencer l'alerte côté `popup_silence.cfg` (qui pilote le tableau de
        notifications et les popups). Sinon l'alerte reviendrait au prochain refresh.

        Mapping (type_evenement → popup_type + fingerprint) :
          - livraison_reporter → 'livr_report' + '{date}_{carburant}'
          - rupture            → 'rupture_imminente' + '{jour}_{carburant}'
          - tendance           → 'tendance' + '{date}_{carburant}'
          - saisies_irregulieres → 'saisies_irr' + '{date}_{carburant}'
          - pont               → 'antirupture' + 'pont_{DDMMYYYY}'

        Silence long (30 jours) car l'utilisateur a explicitement tranché.
        """
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[])
            evt=next((e for e in events if e.get("id")==self.evt_id),None)
            if not evt: return
            type_evt=evt.get("type","")
            data=evt.get("data",{}) or {}
            until_iso=(datetime.now()+timedelta(days=30)).isoformat()
            popup_type=None;fp=None
            if type_evt=="livraison_reporter":
                popup_type="livr_report"
                fp=f"{data.get('date','')}_{data.get('carburant','')}"
            elif type_evt=="rupture":
                popup_type="rupture_imminente"
                fp=f"{data.get('jour','')}_{data.get('carburant','')}"
            elif type_evt=="tendance":
                popup_type="tendance"
                fp=f"{data.get('date','')}_{data.get('carburant','')}"
            elif type_evt=="saisies_irregulieres":
                popup_type="saisies_irr"
                fp=f"{data.get('date','')}_{data.get('carburant','')}"
            elif type_evt=="pont":
                popup_type="antirupture"
                # Reconstruire pont_id depuis date_debut au format DDMMYYYY
                try:
                    d_debut=data.get("date_debut","")
                    if d_debut:
                        dt=datetime.fromisoformat(d_debut) if "T" in d_debut else datetime.strptime(d_debut,"%Y-%m-%d")
                        fp=f"pont_{dt.strftime('%d%m%Y')}"
                except Exception as _e: _log_silent_err(exc=_e)
            elif type_evt=="marge_tendue":
                popup_type="marge_tendue"
                fp=f"{data.get('date','')}_{data.get('carburant','')}"
            elif type_evt=="ferie_isole":
                popup_type="ferie_isole"
                fp=data.get("date_ferie","")
            if popup_type and fp:
                silence_popup(popup_type,[fp],until_iso,{fp:0})
        except Exception as e: print(f"[silence corresponding alert] {e}")

    def _add_comment(self):
        """Ouvre une mini-popup de saisie de commentaire daté (concaténé à l'existant)."""
        from tkinter import simpledialog
        new_comm=simpledialog.askstring("Commentaire",
                                         "Ajouter un commentaire (visible dans le journal) :",
                                         parent=self)
        if new_comm and new_comm.strip():
            if self._update_evt(commentaire=new_comm.strip()):
                self.modified=True
            self.destroy()

    def _update_evt(self,statut=None,lu=None,commentaire=None,note=None):
        """Modifie l'événement dans evenements.cfg. Retourne True si succès."""
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[])
            found=False
            for evt in events:
                if evt.get("id")==self.evt_id:
                    data=evt.setdefault("data",{})
                    if statut is not None: data["statut"]=statut
                    if lu is not None: data["lu"]=lu
                    if commentaire:
                        # Concaténer au commentaire existant avec timestamp dat\u00e9
                        old=evt.get("commentaire","")
                        ts_str=datetime.now().strftime("%d/%m/%Y %Hh%M")
                        new_entry=f"[{ts_str}] {commentaire}"
                        evt["commentaire"]=f"{old}\n{new_entry}" if old else new_entry
                    # Trace d'action automatique (note interne)
                    if note:
                        old=evt.get("commentaire","")
                        ts_str=datetime.now().strftime("%d/%m/%Y %Hh%M")
                        trace=f"[{ts_str}] {note}"
                        evt["commentaire"]=f"{old}\n{trace}" if old else trace
                    found=True;break
            if found:
                all_evt["events"]=events
                save_json(EVENEMENTS_FILE,all_evt)
                return True
        except Exception as e: print(f"[update evt action] {e}")
        return False


class JournalEvenementsDlg(ctk.CTkToplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("\U0001f4cb Journal des \u00e9v\u00e9nements \u2014 DISTRICARB HUB")
        self.geometry("980x700");self.minsize(820,560)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        # État UI : ensemble des clés de situation/événement actuellement développées (expanded).
        # Par défaut, les situations résolues sont affichées en mode condensé (1 ligne) pour
        # alléger l'écran. L'utilisateur clique pour développer une ligne et voir la timeline.
        self.expanded_keys=set()
        # MASQUAGE 21/05/2026 (validé Bidou) : flag d'affichage des événements masqués.
        # False par défaut → on ne voit QUE les events non masqués (journal propre).
        # Toggle en bas → afficher aussi les masqués (pour revoir l'historique, démasquer).
        self.show_masques=False
        # Header
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=64);hdr.pack(fill="x",padx=20,pady=(16,4));hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f4cb  Journal des \u00e9v\u00e9nements",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w")
        ctk.CTkLabel(hdr,text="Historique de ce que le hub a d\u00e9tect\u00e9 : tensions, ruptures, livraisons, commandes",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(anchor="w",pady=(1,0))
        # Filtres : période + type
        filtres=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8);filtres.pack(fill="x",padx=20,pady=(4,8))
        # Période
        ctk.CTkLabel(filtres,text="P\u00e9riode :",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(side="left",padx=(14,6),pady=10)
        self.periode_options=[
            ("mois_courant","Mois en cours"),
            ("mois_prec","Mois pr\u00e9c\u00e9dent"),
            ("3_mois","3 derniers mois"),
            ("annee","Ann\u00e9e en cours"),
            ("tout","Tout")
        ]
        self.periode_var=ctk.StringVar(value=self.periode_options[0][1])
        self.periode_menu=ctk.CTkOptionMenu(filtres,values=[o[1] for o in self.periode_options],
                                              variable=self.periode_var,width=180,height=30,
                                              font=("Segoe UI",10),
                                              fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                                              text_color=C["t1"],
                                              command=lambda v:self._refresh())
        self.periode_menu.pack(side="left",padx=(0,16),pady=10)
        # Type
        ctk.CTkLabel(filtres,text="Type :",font=("Segoe UI",10,"bold"),text_color=C["t1"]).pack(side="left",padx=(0,6),pady=10)
        self.type_options=[("tout","Tous"),("pont","Ponts"),("anomalie","Anomalies"),("rupture","Ruptures"),("livraison","Livraisons"),("commande","Commandes"),("livraison_reporter","Livraisons à reporter"),("passage_mois","Passages de mois")]
        self.type_var=ctk.StringVar(value="Tous")
        self.type_menu=ctk.CTkOptionMenu(filtres,values=[o[1] for o in self.type_options],
                                            variable=self.type_var,width=140,height=30,
                                            font=("Segoe UI",10),
                                            fg_color=C["panel"],button_color=C["border2"],button_hover_color=C["border"],
                                            text_color=C["t1"],
                                            command=lambda v:self._refresh())
        self.type_menu.pack(side="left",padx=(0,16),pady=10)
        # Compteur
        self.count_lbl=ctk.CTkLabel(filtres,text="",font=("Segoe UI",10),text_color=C["t3"])
        self.count_lbl.pack(side="right",padx=14,pady=10)
        # Liste scrollable
        self.list_frame=ctk.CTkScrollableFrame(self,fg_color=C["card"],corner_radius=8,
                                                 scrollbar_button_color=C["border2"])
        self.list_frame.pack(fill="both",expand=True,padx=20,pady=4)
        # Footer
        footer=ctk.CTkFrame(self,fg_color="transparent",height=56);footer.pack(side="bottom",fill="x",padx=20,pady=10);footer.pack_propagate(False)
        ctk.CTkButton(footer,text="Fermer",width=110,height=36,
                       fg_color=C["panel"],hover_color=C["border2"],text_color=C["t1"],
                       font=("Segoe UI",10),corner_radius=8,
                       command=self.destroy).pack(side="right",pady=10)
        # Bouton Tableau d'alertes : ouvre la vue todo-list des alertes actives
        ctk.CTkButton(footer,text="\U0001f514  Tableau d'alertes",width=180,height=36,
                       fg_color="#D9544D",hover_color="#B5403B",text_color="#FFFFFF",
                       font=("Segoe UI",10,"bold"),corner_radius=8,
                       command=self._open_alertes).pack(side="right",padx=(0,8),pady=10)
        # Toggle "Afficher événements masqués" (validé Bidou 21/05/2026)
        self.btn_show_masques=ctk.CTkButton(footer,text="\U0001f441 Afficher masqu\u00e9s",width=180,height=36,
                       fg_color=C["card"],hover_color=C["card_h"],border_width=1,border_color=C["border2"],
                       text_color=C["t2"],font=("Segoe UI",10),corner_radius=8,
                       command=self._toggle_show_masques)
        self.btn_show_masques.pack(side="left",pady=10)
        self._refresh()

    def _toggle_show_masques(self):
        """Bascule l'affichage des événements masqués."""
        self.show_masques=not getattr(self,"show_masques",False)
        if self.show_masques:
            self.btn_show_masques.configure(text="\U0001f441 Masquer les masqu\u00e9s",
                                             fg_color=C["amber"],text_color="#141417")
        else:
            self.btn_show_masques.configure(text="\U0001f441 Afficher masqu\u00e9s",
                                             fg_color=C["card"],text_color=C["t2"])
        self._refresh()

    def _open_alertes(self):
        """Ouvre le tableau de notifications (alertes actives)."""
        try:
            dlg=AlertesDashboardDlg(self.master)
            self.wait_window(dlg)
        except Exception as e: print(f"[open alertes] {e}")

    def _get_period_dates(self):
        """Retourne (start_date, end_date) selon le choix de période."""
        today=date.today()
        choice_label=self.periode_var.get()
        choice_key=next((k for k,v in self.periode_options if v==choice_label),"mois_courant")
        if choice_key=="mois_courant":
            start=date(today.year,today.month,1)
            end=today
        elif choice_key=="mois_prec":
            if today.month==1: yr,mo=today.year-1,12
            else: yr,mo=today.year,today.month-1
            import calendar
            last=calendar.monthrange(yr,mo)[1]
            start=date(yr,mo,1);end=date(yr,mo,last)
        elif choice_key=="3_mois":
            start=today-timedelta(days=90);end=today
        elif choice_key=="annee":
            start=date(today.year,1,1);end=today
        else:  # tout
            start=date(2020,1,1);end=date(2099,12,31)
        return start,end
    def _refresh(self):
        """Recharge la liste des événements selon les filtres.
        
        REGROUPEMENT PAR SITUATION : pour éviter le bruit visuel, les événements
        appartenant à la même situation (ex: tous les événements pour SP livraison du 21/05,
        peu importe les jours de création) sont regroupés en UNE seule carte avec timeline.
        Les types ponctuels (livraison saisie, passage_mois, ack) restent affichés isolément."""
        # Vider la liste actuelle
        for w in self.list_frame.winfo_children(): w.destroy()
        # Charger événements de la période
        start,end=self._get_period_dates()
        events=load_evenements_period(start,end)
        # MASQUAGE : par défaut, exclure les events avec masque=true (toggle "Afficher masqués"
        # en bas réactive l'affichage si Bidou veut revoir/démasquer un événement archivé).
        if not getattr(self,"show_masques",False):
            events=[e for e in events if not e.get("masque",False)]
        # Filtrer par type
        type_label=self.type_var.get()
        type_key=next((k for k,v in self.type_options if v==type_label),"tout")
        if type_key!="tout":
            events=[e for e in events if e.get("type")==type_key]
        # Regrouper par sit_key : 1 carte = 1 situation (avec timeline interne)
        groups={}
        order=[]  # ordre d'apparition des groupes (= ordre du dernier événement)
        for evt in events:
            sk=self._sit_key(evt)
            if sk is None:
                # Types ponctuels : chaque événement = sa propre carte
                sk=f"_evt_{evt.get('id','?')}"
            if sk not in groups:
                groups[sk]=[]
                order.append(sk)
            groups[sk].append(evt)
        # Trier chaque groupe en chronologique pour la timeline (ancien -> récent)
        for sk in groups: groups[sk].sort(key=lambda e:e.get("ts",""))
        # Trier les groupes par TS du dernier événement (descendant = plus récent en haut)
        order.sort(key=lambda sk:groups[sk][-1].get("ts",""),reverse=True)
        # Compteur (groupes + total)
        total_evts=len(events);nb_groups=len(order)
        if nb_groups==total_evts:
            self.count_lbl.configure(text=f"{total_evts} \u00e9v\u00e9nement(s)")
        else:
            self.count_lbl.configure(text=f"{nb_groups} situation(s) \u00b7 {total_evts} \u00e9v\u00e9nement(s)")
        if not order:
            ctk.CTkLabel(self.list_frame,text="Aucun \u00e9v\u00e9nement pour cette p\u00e9riode.",
                          font=("Segoe UI",11),text_color=C["t3"]).pack(pady=40)
            return
        # Afficher chaque groupe avec SÉPARATEURS DE JOUR pour une vue calendaire.
        # Le tri ci-dessus est par TS du dernier événement (option b validée par Bidou) :
        # une situation reste au jour de sa dernière activité réelle, pas à sa date d'origine.
        # On insère un header de jour avant chaque changement de date.
        # Pré-comptage des situations par jour (pour le compteur du bandeau).
        sit_per_day={}
        for sk in order:
            grp=groups[sk]
            try:
                last_ts=grp[-1].get("ts","")
                day_key=datetime.fromisoformat(last_ts).date().isoformat()
                sit_per_day[day_key]=sit_per_day.get(day_key,0)+1
            except Exception as _e: _log_silent_err(exc=_e)
        last_day_key=None
        for sk in order:
            grp=groups[sk]
            try:
                last_ts=grp[-1].get("ts","")
                day_dt=datetime.fromisoformat(last_ts).date()
                day_key=day_dt.isoformat()
            except Exception:
                day_dt=None;day_key=None
            if day_dt is not None and day_key!=last_day_key:
                self._render_day_separator(day_dt,sit_per_day.get(day_key,0))
                last_day_key=day_key
            # Critère compact = situation "non actionnable" qui peut être condensée :
            #  - statut == "resolu" (clôturé par l'utilisateur)
            #  - type ∈ {livraison, passage_mois, ack} = traces historiques, pas des alertes
            # On NE condense PAS les ponts acquittés (Bidou : "à surveiller, pas résolu"),
            # ni les snooze, ni les non_traite : ils restent en grosse carte.
            # EXCEPTION : si l'utilisateur a activé un filtre type (≠ "tout"), on désactive
            # le mode compact car il a déjà filtré → il veut voir les détails du sous-ensemble.
            filter_active=(type_key!="tout")
            last_evt=grp[-1]
            statut_last=(last_evt.get("data",{}) or {}).get("statut","")
            type_last=last_evt.get("type","")
            is_historique=(not filter_active) and ((statut_last=="resolu") or (type_last in ("livraison","passage_mois","ack")))
            if len(grp)==1:
                evt_key=f"evt_{grp[0].get('id','?')}"
                if is_historique and evt_key not in self.expanded_keys:
                    self._render_event_compact(grp[0],evt_key)
                else:
                    self._render_event(grp[0],expand_key=evt_key if is_historique else None)
            else:
                if is_historique and sk not in self.expanded_keys:
                    self._render_situation_group_compact(grp,sk)
                else:
                    self._render_situation_group(grp,expand_key=sk if is_historique else None)

    def _render_day_separator(self,day_dt,n_situations=0):
        """Bandeau de séparation par jour dans le journal pour créer une vue 'agenda'.
        
        Format des labels selon proximité avec aujourd'hui :
          - aujourd'hui  : "Aujourd'hui — mercredi 13 mai 2026"
          - hier         : "Hier — mardi 12 mai 2026"
          - autres jours : "lundi 11 mai 2026"
        
        Layout : bandeau teinté (fond legèrement bleuté) avec date centrée et compteur de
        situations à droite. Style assumé "header de section" sans être agressif.
        """
        today_d=date.today()
        JC_FR=["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
        MOIS_FR=["janvier","f\u00e9vrier","mars","avril","mai","juin",
                 "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        date_long=f"{JC_FR[day_dt.weekday()]} {day_dt.day} {MOIS_FR[day_dt.month-1]} {day_dt.year}"
        if day_dt==today_d:
            label=f"\U0001f4c5  Aujourd'hui \u2014 {date_long}"
        elif day_dt==today_d-timedelta(days=1):
            label=f"\U0001f4c5  Hier \u2014 {date_long}"
        else:
            label=f"\U0001f4c5  {date_long}"
        cnt_txt=f"{n_situations} situation{'s' if n_situations>1 else ''}"
        # Bandeau teinté
        sep_box=ctk.CTkFrame(self.list_frame,fg_color="#182536",corner_radius=8,height=44,
                              border_width=1,border_color=C["border"])
        sep_box.pack(fill="x",padx=6,pady=(14,6))
        sep_box.pack_propagate(False)
        # Layout grid 3 colonnes : spacer gauche / label centré / compteur droite
        sep_box.grid_columnconfigure(0,weight=1)
        sep_box.grid_columnconfigure(1,weight=0)
        sep_box.grid_columnconfigure(2,weight=1)
        ctk.CTkLabel(sep_box,text=label,font=("Segoe UI",12,"bold"),
                     text_color=C["gold"]).grid(row=0,column=1,padx=14,pady=10)
        ctk.CTkLabel(sep_box,text=cnt_txt,font=("Segoe UI",10),
                     text_color=C["t2"]).grid(row=0,column=2,sticky="e",padx=14,pady=10)

    def _toggle_expand(self,key):
        """Bascule l'état développé/condensé d'une situation ou d'un événement résolu."""
        if key in self.expanded_keys: self.expanded_keys.discard(key)
        else: self.expanded_keys.add(key)
        try: self._refresh()
        except Exception as _e: _log_silent_err(exc=_e)

    def _render_situation_group_compact(self,events,sit_key):
        """Mode CONDENSÉ pour situation historique (résolue / trace livraison / passage_mois).
        
        Design assumé : la ligne doit être bien lisible mais clairement différenciée des
        grosses cartes (qui demandent action). Choix :
          - hauteur 44px (entre carte 80+px et bandeau 32px) : présent sans dominer
          - barre couleur 6px à gauche : repère visuel fort du type d'événement
          - icône statut 16pt bold avec couleur sémantique (✓ vert résolu, 📦 vert livraison)
          - titre 11pt BOLD en crème C["t1"] : 100% lisible (pas le t2/t3 illisible précédent)
          - date 10pt en C["t2"] : lisible mais secondaire
          - chevron 14pt à droite : indique cliquable
          - fond C["card_h"] (un cran plus clair que le bg) : la ligne sort visuellement
        Cliquer la ligne développe la situation complète (timeline + commentaires).
        """
        last_evt=events[-1]
        type_evt=last_evt.get("type","")
        type_colors={"pont":"#C99A5B","anomalie":"#8E7BA6","tendance":"#8E7BA6","rupture":"#C5544D","commande":"#5B9E92","livraison":"#7FA86E","ack":"#6B8FB5","passage_mois":"#4A8A63","livraison_reporter":"#C99A5B","marge_tendue":"#C7A95B","ferie_isole":"#C7A95B","livraison_attendue":"#C5544D"}
        col=type_colors.get(type_evt,C["t2"])
        sit_label=self._situation_label(events[0])
        # Choix de l'icône statut : ✓ vert pour résolu, 📦 pour livraison saisie, info pour autres
        statut_last=(last_evt.get("data",{}) or {}).get("statut","")
        if statut_last=="resolu":
            icon_txt="\u2713";icon_col=C["green"]
        elif type_evt=="livraison":
            icon_txt="\U0001f4e6";icon_col=C["green"]  # 📦
        elif type_evt=="passage_mois":
            icon_txt="\U0001f4c5";icon_col=C["gold"]  # 📅
        elif type_evt=="ack":
            icon_txt="\u2713";icon_col="#5C9DDA"
        else:
            icon_txt="\u25cf";icon_col=C["t2"]
        # Date dernière action au format "mar. 13/05 10h36"
        try:
            last_ts=last_evt.get("ts","")
            last_dt=datetime.fromisoformat(last_ts)
            date_compact=f"{self._fmt_date_jour(last_dt.date())} {last_dt.strftime('%Hh%M')}"
        except Exception: date_compact=""
        # Indication "N apparitions" si situation regroupée
        nb_apps=f" \u00b7 {len(events)} apparitions" if len(events)>1 else ""
        # Ligne compacte
        line=ctk.CTkFrame(self.list_frame,fg_color=C["card_h"],corner_radius=8,height=44,
                           border_width=1,border_color=C["border2"],cursor="hand2")
        line.pack(fill="x",padx=6,pady=3);line.pack_propagate(False)
        # Barre couleur 6px pleine hauteur (gauche)
        ctk.CTkFrame(line,fg_color=col,width=6,corner_radius=0).pack(side="left",fill="y")
        # Icône statut bien visible
        ctk.CTkLabel(line,text=icon_txt,font=("Segoe UI Emoji",16,"bold"),
                     text_color=icon_col,width=32).pack(side="left",padx=(8,4))
        # Titre BOLD en crème (lisible)
        ctk.CTkLabel(line,text=sit_label+nb_apps,font=("Segoe UI",11,"bold"),
                     text_color=C["t1"],anchor="w").pack(side="left",padx=(4,8))
        # Chevron à droite (clickable affordance)
        ctk.CTkLabel(line,text="\u203a",font=("Segoe UI",14,"bold"),
                     text_color=C["t2"],width=24).pack(side="right",padx=(0,12))
        # Date en C["t2"] - lisible, secondaire
        ctk.CTkLabel(line,text=date_compact,font=("Segoe UI",10),
                     text_color=C["t2"]).pack(side="right",padx=(0,4))
        # Bouton masquer (validé Bidou 21/05/2026) : un clic ouvre confirmation, puis flag
        # `masque=true` sur TOUS les events de la situation. CTkButton avec command (pas bind)
        # pour que le clic ne se propage pas au bind de la ligne (qui développe).
        evt_ids=[e.get("id") for e in events if e.get("id")]
        # Bouton MASQUER (réversible) : flag masque=true, demeure dans le fichier
        btn_masq=ctk.CTkButton(line,text="\U0001f441\u20e0",width=28,height=26,
                                fg_color="transparent",hover_color=C["card"],text_color=C["t3"],
                                font=("Segoe UI Emoji",13),corner_radius=4,border_width=0,
                                command=lambda ids=evt_ids,lbl=sit_label:self._masquer_situation(ids,lbl))
        btn_masq.pack(side="right",padx=(0,4))
        # Bouton EFFACER (irréversible, validé Bidou 25/05/2026) — uniquement pour types d'alerte
        btn_del=None
        TYPES_ALERTE_COMPACT={"pont","anomalie","rupture","marge_tendue","livraison_reporter","ferie_isole","saisies_irregulieres"}
        if type_evt in TYPES_ALERTE_COMPACT:
            btn_del=ctk.CTkButton(line,text="\U0001f5d1",width=28,height=26,
                                   fg_color="transparent",hover_color="#2A1518",text_color=C["t3"],
                                   font=("Segoe UI Emoji",13),corner_radius=4,border_width=0,
                                   command=lambda ids=evt_ids,lbl=sit_label:self._effacer_situation(ids,lbl))
            btn_del.pack(side="right",padx=(0,4))
        # Clic n'importe où sur la ligne (et ses enfants) → développer
        # SAUF le bouton masquer (déjà géré par command)
        line.bind("<Button-1>",lambda e,k=sit_key:self._toggle_expand(k))
        for child in line.winfo_children():
            if child is btn_masq or (btn_del is not None and child is btn_del): continue
            try: child.bind("<Button-1>",lambda e,k=sit_key:self._toggle_expand(k))
            except Exception as _e: _log_silent_err(exc=_e)

    def _render_event_compact(self,evt,evt_key):
        """Mode CONDENSÉ pour un événement seul historique. Même design que _situation_group_compact."""
        type_evt=evt.get("type","")
        type_colors={"pont":"#C99A5B","anomalie":"#8E7BA6","tendance":"#8E7BA6","rupture":"#C5544D","commande":"#5B9E92","livraison":"#7FA86E","ack":"#6B8FB5","passage_mois":"#4A8A63","livraison_reporter":"#C99A5B","marge_tendue":"#C7A95B","ferie_isole":"#C7A95B","livraison_attendue":"#C5544D"}
        col=type_colors.get(type_evt,C["t2"])
        data=evt.get("data",{}) or {}
        statut_last=data.get("statut","")
        # Icône statut
        if statut_last=="resolu":
            icon_txt="\u2713";icon_col=C["green"]
        elif type_evt=="livraison":
            icon_txt="\U0001f4e6";icon_col=C["green"]
        elif type_evt=="passage_mois":
            icon_txt="\U0001f4c5";icon_col=C["gold"]
        elif type_evt=="ack":
            icon_txt="\u2713";icon_col="#5C9DDA"
        else:
            icon_txt="\u25cf";icon_col=C["t2"]
        # Label situation (utilise _situation_label aussi pour les types unitaires)
        label=self._situation_label(evt)
        if label==type_evt:
            # Fallback : si pas de label dédié, fabriquer un truc lisible
            if type_evt=="livraison":
                d_iso=data.get("date","")
                try:
                    d_obj=datetime.fromisoformat(d_iso).date()
                    label=f"Livraison saisie {self._fmt_date_jour(d_obj)}"
                except Exception: label=f"Livraison saisie {d_iso}"
            elif type_evt=="passage_mois":
                label="Passage de mois"
            elif type_evt=="ack":
                label="Acquittement"
        # Date dernière action
        try:
            last_ts=evt.get("ts","")
            last_dt=datetime.fromisoformat(last_ts)
            date_compact=f"{self._fmt_date_jour(last_dt.date())} {last_dt.strftime('%Hh%M')}"
        except Exception: date_compact=""
        # Sous-titre : résumé court et humain (commentaire de l'événement, nettoyé/tronqué).
        # C'est le 2e niveau de la hiérarchie de lecture (titre fort / détail gris).
        sous_titre=(evt.get("commentaire","") or "").strip()
        for pref in ("\U0001f4ac ","\u2713 ","\u26a0 "):
            if sous_titre.startswith(pref): sous_titre=sous_titre[len(pref):].strip()
        if len(sous_titre)>80: sous_titre=sous_titre[:77]+"\u2026"
        # Frise de lecture : fond neutre, pas de bordure criarde, hauteur adaptative
        # (1 ou 2 lignes selon présence d'un résumé). Distinct des cartes-action du Tableau.
        line=ctk.CTkFrame(self.list_frame,fg_color=C["bg"],corner_radius=8,cursor="hand2")
        line.pack(fill="x",padx=6,pady=2)
        # Barre de couleur FINE : simple repère de type, pas une décoration
        ctk.CTkFrame(line,fg_color=col,width=3,corner_radius=2).pack(side="left",fill="y",pady=6)
        ctk.CTkLabel(line,text=icon_txt,font=("Segoe UI Emoji",14),
                     text_color=icon_col,width=28).pack(side="left",padx=(8,2))
        # Colonne titre + résumé (hiérarchie typographique)
        col_txt=ctk.CTkFrame(line,fg_color="transparent")
        col_txt.pack(side="left",fill="both",expand=True,padx=(4,8),pady=(7,7))
        ctk.CTkLabel(col_txt,text=label,font=("Segoe UI",11,"bold"),
                     text_color=C["t1"],anchor="w").pack(anchor="w")
        if sous_titre:
            ctk.CTkLabel(col_txt,text=sous_titre,font=("Segoe UI",9),
                         text_color=C["t3"],anchor="w").pack(anchor="w",pady=(1,0))
        # Date discrète à droite
        ctk.CTkLabel(line,text=date_compact,font=("Segoe UI",9),
                     text_color=C["t3"]).pack(side="right",padx=(0,10))
        ctk.CTkLabel(line,text="\u203a",font=("Segoe UI",13),
                     text_color=C["t3"],width=16).pack(side="right",padx=(0,4))
        # Boutons d'action — discrets (le Journal est lecture, pas action)
        evt_id_one=evt.get("id")
        btn_masq=ctk.CTkButton(line,text="\U0001f441\u20e0",width=26,height=24,
                                fg_color="transparent",hover_color=C["card_h"],text_color=C["t3"],
                                font=("Segoe UI Emoji",12),corner_radius=4,border_width=0,
                                command=lambda i=evt_id_one,lbl=label:self._masquer_situation([i],lbl))
        btn_masq.pack(side="right",padx=(0,2))
        # Bouton EFFACER (irréversible, validé Bidou 25/05/2026) — uniquement types d'alerte
        btn_del=None
        TYPES_ALERTE_COMPACT={"pont","anomalie","tendance","rupture","marge_tendue","livraison_reporter","ferie_isole","saisies_irregulieres","livraison_attendue"}
        if type_evt in TYPES_ALERTE_COMPACT:
            btn_del=ctk.CTkButton(line,text="\U0001f5d1",width=26,height=24,
                                   fg_color="transparent",hover_color="#2A1518",text_color=C["t3"],
                                   font=("Segoe UI Emoji",12),corner_radius=4,border_width=0,
                                   command=lambda i=evt_id_one,lbl=label:self._effacer_situation([i],lbl))
            btn_del.pack(side="right",padx=(0,2))
        # Clic n'importe où (sauf boutons) pour déplier — binding récursif (le titre et le
        # résumé sont dans une sous-colonne, il faut descendre dans l'arbre des widgets).
        def _bind_expand(w):
            if w is btn_masq or (btn_del is not None and w is btn_del): return
            try: w.bind("<Button-1>",lambda e,k=evt_key:self._toggle_expand(k))
            except Exception: pass
            for ch in w.winfo_children():
                _bind_expand(ch)
        _bind_expand(line)

    def _masquer_situation(self,evt_ids,label):
        """Masque tous les événements d'une situation (ou un événement isolé).
        Confirmation avant action. Trace au journal via masquer_evenement()."""
        try:
            if not evt_ids: return
            nb=len(evt_ids)
            msg=(f"Masquer cette situation ({nb} apparition{'s' if nb>1 else ''}) ?\n\n"
                 f"\u00ab {label} \u00bb\n\n"
                 f"L'\u00e9v\u00e9nement n'appara\u00eetra plus dans le journal, mais ses donn\u00e9es restent stock\u00e9es. "
                 f"Tu peux le revoir en activant \u00ab \U0001f441 Afficher masqu\u00e9s \u00bb en bas.")
            if not messagebox.askyesno("Masquer l'\u00e9v\u00e9nement",msg):
                return
            nb_ok=0
            for eid in evt_ids:
                if eid and masquer_evenement(eid):
                    nb_ok+=1
            if nb_ok>0:
                self._refresh()
        except Exception as _e: _log_silent_err(exc=_e)

    def _effacer_situation(self,evt_ids,label):
        """Supprime DÉFINITIVEMENT tous les événements d'une situation (ou un événement
        isolé) du fichier evenements.cfg. Action IRRÉVERSIBLE. Validée Bidou 25/05/2026
        pour nettoyer les pollutions du journal (résidus Pre_vision, fausses alertes
        héritées, événements obsolètes).
        Confirmation explicite avant action. Refresh du journal après suppression."""
        try:
            if not evt_ids: return
            nb=len(evt_ids)
            msg=(f"Effacer d\u00e9finitivement cette situation ({nb} apparition{'s' if nb>1 else ''}) ?\n\n"
                 f"\u00ab {label} \u00bb\n\n"
                 f"\u26a0 ATTENTION : cette action est IRR\u00c9VERSIBLE.\n"
                 f"L'\u00e9v\u00e9nement sera supprim\u00e9 du fichier journal. "
                 f"Aucun moyen de le r\u00e9cup\u00e9rer.\n\n"
                 f"Si tu veux juste le cacher temporairement, utilise plut\u00f4t \u00ab Masquer \u00bb.")
            if not messagebox.askyesno("Effacer d\u00e9finitivement",msg):
                return
            nb_ok=0
            for eid in evt_ids:
                if eid and delete_evenement(eid):
                    nb_ok+=1
            if nb_ok>0:
                self._refresh()
        except Exception as _e: _log_silent_err(exc=_e)

    def _sit_key(self,evt):
        """Calcule la clé de regroupement (= situation persistante) pour un événement.
        Retourne None si l'événement est ponctuel et ne doit pas être regroupé."""
        t=evt.get("type","")
        d=evt.get("data",{}) or {}
        if t=="livraison_reporter":
            try:
                date_str=(d.get("date","") or "")[:10]
                return f"livr_report:{d.get('carburant','')}:{date_str}"
            except Exception: return None
        elif t=="pont":
            try:
                dd=d.get("date_debut","")
                if dd:
                    dt=datetime.fromisoformat(dd).date()
                    return f"antirupture:pont_{dt.strftime('%d%m%Y')}"
            except Exception: return None
        elif t=="rupture":
            try:
                j=(d.get("jour","") or "")[:10]
                return f"rupture:{d.get('carburant','')}:{j}"
            except Exception: return None
        elif t=="anomalie":
            try:
                j=(d.get("jour","") or "")[:10]
                return f"anomalie:{d.get('carburant','')}:{j}"
            except Exception: return None
        elif t=="marge_tendue":
            # 1 situation = 1 (carburant × date livraison). Bug signalé Bidou 20/05 20h05 :
            # 4 cartes "Marge tendue SP — livraison du 23/05/2026" affichées séparément
            # alors que c'est physiquement la MÊME alerte (acquittée + fermée X plusieurs
            # fois dans la journée). Maintenant : 1 seule carte "X apparitions" pliable.
            try:
                date_str=(d.get("date","") or "")[:10]
                return f"marge_tendue:{d.get('carburant','')}:{date_str}"
            except Exception: return None
        elif t=="ferie_isole":
            # 1 situation = 1 date fériée concernée
            try:
                date_str=(d.get("date","") or "")[:10]
                return f"ferie_isole:{date_str}"
            except Exception: return None
        # livraison, passage_mois, ack : ponctuels, pas de regroupement
        return None

    def _render_situation_group(self,events,expand_key=None):
        """Rend une carte 'situation' regroupant plusieurs événements liés sur la même timeline.
        
        Args:
          events: liste d'événements de la situation, triés du plus ancien au plus récent.
          expand_key: si fourni, c'est que la situation est résolue mais affichée en mode
                      développé suite à clic utilisateur. Permet d'afficher un bouton "Réduire"
                      pour revenir au mode condensé.
        """
        """Affiche une carte unique pour une situation, avec timeline interne des événements.
        Le STATUT et la couleur viennent du DERNIER événement (le plus récent)."""
        last_evt=events[-1]
        type_evt=last_evt.get("type","")
        type_colors={"pont":"#C99A5B","anomalie":"#8E7BA6","tendance":"#8E7BA6","rupture":"#C5544D","commande":"#5B9E92","livraison":"#7FA86E","ack":"#6B8FB5","passage_mois":"#4A8A63","livraison_reporter":"#C99A5B","marge_tendue":"#C7A95B","ferie_isole":"#C7A95B","livraison_attendue":"#C5544D"}
        col=type_colors.get(type_evt,C["t2"])
        type_labels={"pont":"Pont","anomalie":"Anomalie","rupture":"Rupture","commande":"Commande",
                     "livraison":"Livraison","ack":"Acquittement","passage_mois":"Passage mois",
                     "livraison_reporter":"Livr. \u00e0 reporter"}
        type_label=type_labels.get(type_evt,type_evt)
        # Distinction Pont (contient un férié) vs Weekend (sam+dim sans férié).
        # Validé Bidou 25/05/2026 : "weekend devient pont" était trompeur.
        # Pour Weekend : couleur sobre gris-bleu (#7B8896) au lieu de l'orange "alerte".
        if type_evt=="pont":
            try:
                d_pont=last_evt.get("data",{}) or {}
                dd=d_pont.get("date_debut","");df=d_pont.get("date_fin","")
                if dd:
                    ddt=datetime.fromisoformat(dd).date()
                    dft=datetime.fromisoformat(df).date() if df else ddt
                    cur=ddt;has_ferie=False
                    while cur<=dft:
                        if is_ferie(cur): has_ferie=True;break
                        cur+=timedelta(days=1)
                    if has_ferie:
                        type_label="Pont"
                    else:
                        type_label="Weekend"
                        col="#7B8896"  # gris-bleu sobre (palette Apple/Samsung-like)
            except Exception as _e: _log_silent_err(exc=_e)
        # Détecter si la situation est résolue (= statut courant du dernier événement)
        is_resolu=(last_evt.get("data",{}) or {}).get("statut","")=="resolu"
        # Cadre — style atténué si résolu (fond plus discret, bordure plus pâle)
        frame_fg=C["panel"] if is_resolu else C["bg"]
        frame_border=C["border"] if is_resolu else C["border2"]
        title_color=C["t3"] if is_resolu else C["t1"]
        frame=ctk.CTkFrame(self.list_frame,fg_color=frame_fg,corner_radius=8,border_width=1,border_color=frame_border)
        frame.pack(fill="x",padx=6,pady=4)
        # Header : badge type + label de situation
        head=ctk.CTkFrame(frame,fg_color="transparent");head.pack(fill="x",padx=12,pady=(8,4))
        badge=ctk.CTkFrame(head,fg_color=col,corner_radius=4,width=12,height=12)
        badge.pack(side="left",padx=(0,8));badge.pack_propagate(False)
        # Label métier de la situation (pas juste le type)
        sit_label=self._situation_label(events[0])
        ctk.CTkLabel(head,text=sit_label,font=("Segoe UI",11,"bold"),text_color=title_color).pack(side="left")
        # Compteur d'événements à droite
        ctk.CTkLabel(head,text=f"{len(events)} apparitions",font=("Segoe UI",9),text_color=C["t3"]).pack(side="right")
        # Bouton "Réduire" si la situation est résolue et actuellement développée
        if expand_key is not None:
            ctk.CTkButton(head,text="\u2bc6 R\u00e9duire",width=80,height=24,
                          fg_color="transparent",hover_color=C["card_h"],text_color=C["t2"],
                          border_width=1,border_color=C["border2"],
                          font=("Segoe UI",9),corner_radius=4,
                          command=lambda k=expand_key:self._toggle_expand(k)).pack(side="right",padx=(0,8))
        # === RÉSUMÉ UNE LIGNE (toujours visible) ===
        # Montre l'essentiel en un coup d'œil : statut courant + date la plus récente
        # + nombre d'apparitions. C'est ce que Bidou regarde par défaut.
        # La timeline complète n'est dépliée que sur action explicite (point de friction
        # identifié sur le terrain : "16 lignes de 'Popup fermée sans action' = illisible").
        d_last=last_evt.get("data",{}) or {}
        statut_courant=d_last.get("statut","")
        statut_txt={"non_traite":"\u00c0 traiter","resolu":"R\u00e9glé","annule":"Annul\u00e9","snooze":"Snooze\u00e9","ack":"Pris en compte"}.get(statut_courant,statut_courant or "?")
        statut_col={"non_traite":C["red"],"resolu":C["green"],"annule":C["t3"],"snooze":C["amber"],"ack":"#5C9DDA"}.get(statut_courant,C["t2"])
        try:
            last_dt=datetime.fromisoformat(last_evt.get("ts",""))
            depuis=f"depuis {self._fmt_date_jour(last_dt.date())} {last_dt.strftime('%Hh%M')}"
        except Exception: depuis=""
        resume_row=ctk.CTkFrame(frame,fg_color="transparent");resume_row.pack(fill="x",padx=12,pady=(0,6))
        ctk.CTkLabel(resume_row,text=f"\u25cf {statut_txt}",font=("Segoe UI",11,"bold"),text_color=statut_col).pack(side="left")
        if depuis:
            ctk.CTkLabel(resume_row,text=f"  \u00b7  {depuis}",font=("Segoe UI",10),text_color=C["t2"]).pack(side="left")
        # Bouton pour voir/cacher la timeline complète (seulement si historique > 1 événement)
        timeline_visible=(sit_key in self.expanded_keys) if (sit_key:=f"timeline_{events[0].get('id','?')}_{len(events)}") else False
        if len(events)>1:
            chevron="\u2bc6 Masquer historique" if timeline_visible else f"\u2bc8 Voir historique ({len(events)} apparitions)"
            ctk.CTkButton(resume_row,text=chevron,width=240,height=24,
                          fg_color="transparent",hover_color=C["card_h"],text_color=C["t3"],
                          border_width=1,border_color=C["border2"],
                          font=("Segoe UI",9),corner_radius=4,
                          command=lambda k=sit_key:self._toggle_expand(k)).pack(side="right")
        # Timeline complète (affichée seulement si l'utilisateur a déplié)
        # Si len==1 OU si la clé est dans expanded_keys, on déplie.
        show_timeline=(len(events)==1) or timeline_visible
        if show_timeline:
            # Timeline : 1 ligne par événement, MAIS on regroupe les snoozes consécutifs
            # pour ne pas empiler 14 lignes "snooze — snooze — snooze" illisibles
            # (constat terrain Bidou). Un bloc de snoozes successifs = 1 ligne résumée
            # "↻ N rappels reportés (du JJ/MM au JJ/MM)". Les événements porteurs de sens
            # (non_traite, resolu, livraison, ack) restent affichés individuellement.
            timeline=ctk.CTkFrame(frame,fg_color="transparent");timeline.pack(fill="x",padx=12,pady=(0,8))
        else:
            timeline=None
        def _is_snooze(e):
            return (e.get("data",{}) or {}).get("statut","")=="snooze"
        if show_timeline and timeline is not None:
            i=0;n=len(events)
            while i<n:
                evt=events[i]
                if _is_snooze(evt):
                    # Accumuler la série de snoozes consécutifs
                    j=i
                    while j<n and _is_snooze(events[j]): j+=1
                    bloc=events[i:j]
                    if len(bloc)==1:
                        self._render_timeline_line(timeline,bloc[0],col)
                    else:
                        # Résumé compact du bloc de snoozes
                        def _d(e):
                            try:
                                dt=datetime.fromisoformat(e.get("ts",""))
                                return self._fmt_date_jour(dt.date())
                            except Exception: return ""
                        d0=_d(bloc[0]);d1=_d(bloc[-1])
                        plage=f"du {d0} au {d1}" if d0!=d1 else f"le {d0}"
                        line=ctk.CTkFrame(timeline,fg_color="transparent")
                        line.pack(fill="x",pady=2)
                        ctk.CTkLabel(line,text="\u21bb",font=("Segoe UI",11,"bold"),
                                     text_color=C["amber"],width=20).pack(side="left")
                        ctk.CTkLabel(line,
                            text=f"{len(bloc)} rappels report\u00e9s ({plage})",
                            font=("Segoe UI",10),text_color=C["t2"],anchor="w",
                            justify="left").pack(side="left",padx=(4,0))
                    i=j
                else:
                    self._render_timeline_line(timeline,evt,col)
                    i+=1
        # Statut final (pour les non_traite, on autorise l'action sur le dernier)
        d_last=last_evt.get("data",{}) or {}
        # === Zone d'actions à droite : Action requise (si non_traite) + Marquer résolu (toujours sauf si déjà résolu) ===
        # Les vieilles situations qui sont restées en 'snooze' / 'non_traite' figé peuvent être
        # nettoyées d'un clic sans avoir à passer par la pastille rouge + popup EventActionDlg.
        statut_courant=d_last.get("statut","")
        # Actions row : TOUJOURS visible pour permettre l'effacement (validé Bidou 25/05/2026).
        # Le bouton "Effacer" est restreint aux TYPES D'ALERTE uniquement (pas sur les
        # livraisons/commandes/passage_mois qui sont des traces, pas des alertes).
        actions_row=ctk.CTkFrame(frame,fg_color="transparent");actions_row.pack(anchor="e",padx=12,pady=(0,8))
        all_evt_ids=[e.get("id") for e in events if e.get("id")]
        sit_label_short=self._situation_label(events[0]) if events else "cette situation"
        TYPES_ALERTE={"pont","anomalie","rupture","marge_tendue","livraison_reporter","ferie_isole","saisies_irregulieres"}
        if statut_courant!="resolu":
            if statut_courant=="non_traite" and not d_last.get("lu",True):
                ctk.CTkButton(actions_row,text="\u26a0 Action requise",
                    fg_color="#3D1F1F",hover_color="#5A2A2A",text_color="#FF6B6B",
                    border_width=1,border_color=C["red"],font=("Segoe UI",10,"bold"),
                    corner_radius=6,height=28,
                    command=lambda eid=last_evt.get("id"): self._open_event_action(eid)).pack(side="left",padx=(0,6))
            ctk.CTkButton(actions_row,text="\u2713 Marquer comme r\u00e9gl\u00e9",
                fg_color=C["panel"],hover_color=C["card_h"],text_color=C["green"],
                border_width=1,border_color=C["green"],font=("Segoe UI",10),
                corner_radius=6,height=28,
                command=lambda eid=last_evt.get("id"): self._marquer_resolu_evt(eid)).pack(side="left",padx=(0,6))
        # Bouton Effacer (discret) uniquement pour les types d'ALERTE
        if type_evt in TYPES_ALERTE:
            ctk.CTkButton(actions_row,text="\U0001f5d1",width=28,height=28,
                fg_color="transparent",hover_color="#2A1518",text_color=C["t3"],
                border_width=0,font=("Segoe UI Emoji",13),corner_radius=6,
                command=lambda ids=all_evt_ids,lbl=sit_label_short: self._effacer_situation(ids,lbl)).pack(side="left")

    def _render_timeline_line(self,parent,evt,col):
        """Une ligne dans la timeline d'une situation : horodatage + statut + commentaire.
        Format date 'mar. 11/05 07h01' pour lisibilité (avec jour de semaine).
        Police 10 pt en C["t1"] (lisible, plus le t2 illisible précédent)."""
        ts=evt.get("ts","")
        try:
            ts_dt=datetime.fromisoformat(ts)
            ts_str=f"{self._fmt_date_jour(ts_dt.date())} {ts_dt.strftime('%Hh%M')}"
        except Exception: ts_str=ts[:16]
        d=evt.get("data",{}) or {}
        statut=d.get("statut","")
        comment=evt.get("commentaire","")
        # Icône selon statut
        icon_map={"non_traite":"\u2757","resolu":"\u2713","annule":"\u2715","snooze":"\u23f1","ack":"\u2713"}
        icon=icon_map.get(statut,"\u2022")
        icon_col={"non_traite":C["red"],"resolu":C["green"],"annule":C["t3"],"snooze":C["amber"],"ack":"#5C9DDA"}.get(statut,col)
        # Mapping statut technique → libellé métier (cohérence avec le résumé de la carte)
        statut_label_map={"non_traite":"\u00c0 traiter","resolu":"R\u00e9gl\u00e9","annule":"Annul\u00e9","snooze":"Snooz\u00e9","ack":"Pris en compte"}
        statut_pretty=statut_label_map.get(statut,statut)
        # Texte
        parts=[ts_str]
        if statut_pretty: parts.append(statut_pretty)
        if comment: parts.append(comment[:80]+("\u2026" if len(comment)>80 else ""))
        line=ctk.CTkFrame(parent,fg_color="transparent")
        line.pack(fill="x",pady=2)
        ctk.CTkLabel(line,text=icon,font=("Segoe UI",11,"bold"),text_color=icon_col,width=20).pack(side="left")
        ctk.CTkLabel(line,text=" \u2014 ".join(parts),font=("Segoe UI",10),text_color=C["t1"],anchor="w",justify="left").pack(side="left",padx=(4,0))

    def _fmt_date_jour(self,d_obj):
        """Format date avec jour de semaine abrégé : 'mar. 19/05'. Lisibilité prioritaire."""
        JC_FR=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
        try: return f"{JC_FR[d_obj.weekday()]} {d_obj.strftime('%d/%m')}"
        except Exception: return ""

    def _situation_label(self,evt):
        """Label humain d'une situation à partir d'un de ses événements.
        Reformulation 23/05/2026 (validée Bidou) : ton plus naturel/conseiller,
        moins télégraphique. Dates au format complet 'samedi 23 mai' pour les types
        principaux ; format court 'mar. 19/05' conservé en fallback."""
        t=evt.get("type","")
        d=evt.get("data",{}) or {}
        # Helpers locaux pour formatage de date long ("samedi 23 mai")
        JOURS_LONG=["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
        MOIS_LONG=["janvier","f\u00e9vrier","mars","avril","mai","juin",
                   "juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        def _fmt_long(d_obj):
            return f"{JOURS_LONG[d_obj.weekday()]} {d_obj.day} {MOIS_LONG[d_obj.month-1]}"
        if t=="livraison_reporter":
            try:
                date_str=(d.get("date","") or "")[:10]
                if date_str:
                    d_obj=datetime.fromisoformat(date_str).date()
                    return f"Livraison {d.get('carburant','?')} du {JOURS_LONG[d_obj.weekday()]} \u00e0 d\u00e9caler \u2014 la cuve d\u00e9borderait"
            except Exception as _e: _log_silent_err(exc=_e)
        elif t=="pont":
            try:
                dd=d.get("date_debut","")
                df=d.get("date_fin","")
                if dd:
                    ddt=datetime.fromisoformat(dd).date()
                    dft=datetime.fromisoformat(df).date() if df else ddt
                    # Détection : un VRAI pont contient au moins un jour férié.
                    # Sinon c'est un simple weekend (sam+dim), pas un pont.
                    # Validé Bidou 25/05/2026 : "weekend devient pont" était trompeur.
                    cur=ddt;contient_ferie=False
                    while cur<=dft:
                        if is_ferie(cur):
                            contient_ferie=True;break
                        cur+=timedelta(days=1)
                    JC_short=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
                    dd_str=f"{JC_short[ddt.weekday()]} {ddt.strftime('%d/%m')}"
                    df_str=f"{JC_short[dft.weekday()]} {dft.strftime('%d/%m')}"
                    if contient_ferie:
                        return f"Pont sans livraison \u2014 du {dd_str} au {df_str}"
                    if ddt==dft:
                        return f"Weekend sans livraison \u2014 {dd_str}"
                    return f"Weekend sans livraison \u2014 du {dd_str} au {df_str}"
            except Exception as _e: _log_silent_err(exc=_e)
        elif t=="rupture":
            try:
                j=(d.get("jour","") or "")[:10]
                if j:
                    d_obj=datetime.fromisoformat(j).date()
                    return f"Rupture {d.get('carburant','?')} constat\u00e9e {_fmt_long(d_obj)}"
            except Exception as _e: _log_silent_err(exc=_e)
        elif t=="anomalie":
            try:
                j=(d.get("jour","") or "")[:10]
                if j:
                    d_obj=datetime.fromisoformat(j).date()
                    return f"Consommation {d.get('carburant','?')} anormale {_fmt_long(d_obj)}"
            except Exception as _e: _log_silent_err(exc=_e)
        elif t=="livraison":
            try:
                date_str=d.get("date","")
                if date_str:
                    d_obj=datetime.fromisoformat(date_str).date()
                    return f"Livraison re\u00e7ue \u2014 {_fmt_long(d_obj)}"
            except Exception as _e: _log_silent_err(exc=_e)
        elif t=="marge_tendue":
            try:
                date_str=d.get("date","")
                if date_str:
                    d_obj=datetime.fromisoformat(date_str).date()
                    carb=d.get("carburant","?")
                    jour_long=JOURS_LONG[d_obj.weekday()]
                    marge=int(d.get("marge_restante",0) or 0)
                    if marge>0:
                        return f"Livraison {jour_long} : {marge:,} L de marge sur le {carb}".replace(",",".")
                    return f"Marge {carb} serr\u00e9e \u00e0 la livraison de {jour_long}"
            except Exception as _e: _log_silent_err(exc=_e)
        elif t=="ferie_isole":
            try:
                date_str=d.get("date_ferie","")
                if date_str:
                    d_obj=datetime.fromisoformat(date_str).date()
                    nom=d.get("nom_ferie","f\u00e9ri\u00e9")
                    return f"{nom} {JOURS_LONG[d_obj.weekday()]} \u2014 v\u00e9rifier la commande"
            except Exception as _e: _log_silent_err(exc=_e)
        elif t=="commande":
            try:
                date_str=(d.get("jour","") or d.get("date","") or "")[:10]
                sp=int(d.get("sp",0) or 0);go=int(d.get("go",0) or 0);gnr=int(d.get("gnr",0) or 0)
                tour=int(d.get("tour",1) or 1)
                tour_str={1:"1\u02b3\u1d49 tour",2:"2\u1d49 tour",3:"3\u1d49 tour"}.get(tour,f"tour {tour}")
                vols=[]
                if sp>0: vols.append(f"SP {sp:,} L".replace(",","."))
                if go>0: vols.append(f"GO {go:,} L".replace(",","."))
                if gnr>0: vols.append(f"GNR {gnr:,} L".replace(",","."))
                vols_str=" \u00b7 ".join(vols) if vols else "volumes \u00e0 saisir"
                if date_str:
                    d_obj=datetime.fromisoformat(date_str).date()
                    jour_long=JOURS_LONG[d_obj.weekday()]
                    return f"Commande pass\u00e9e pour {jour_long} \u2014 {vols_str} ({tour_str})"
                return f"Commande pass\u00e9e \u2014 {vols_str} ({tour_str})"
            except Exception as _e: _log_silent_err(exc=_e)
        return f"{t}"
    def _open_event_action(self,evt_id):
        """Ouvre la mini-popup d'actions sur un événement non_traite (clic sur pastille rouge).
        Si une modification a été effectuée, rafraîchit l'affichage du journal."""
        dlg=EventActionDlg(self,evt_id)
        self.wait_window(dlg)
        if getattr(dlg,"modified",False):
            try: self._refresh()
            except Exception as e: print(f"[refresh after action] {e}")

    def _marquer_resolu_evt(self,evt_id):
        """Marque un événement comme résolu directement depuis le journal, sans passer par la
        popup EventActionDlg. Permet de nettoyer rapidement les vieilles situations qui sont
        restées en 'snooze' ou 'non_traite' figé (ex: anciennes alertes GNR snoozées dont la
        cause a disparu de Pre_vision).
        
        Effet : statut=resolu, lu=True, silence 30 jours sur l'alerte correspondante, refresh.
        """
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[])
            evt=None
            for e in events:
                if e.get("id")==evt_id:
                    data=e.setdefault("data",{})
                    data["statut"]="resolu";data["lu"]=True
                    old_comm=e.get("commentaire","")
                    ts_str=datetime.now().strftime("%d/%m/%Y %Hh%M")
                    trace=f"[{ts_str}] \u2713 Marqu\u00e9 r\u00e9solu manuellement depuis le journal"
                    e["commentaire"]=f"{old_comm}\n{trace}" if old_comm else trace
                    evt=e;break
            if not evt: return
            all_evt["events"]=events
            save_json(EVENEMENTS_FILE,all_evt)
            # Silence de l'alerte correspondante (réutilise le mapping de EventActionDlg)
            try:
                type_evt=evt.get("type","");data=evt.get("data",{}) or {}
                until_iso=(datetime.now()+timedelta(days=30)).isoformat()
                popup_type=None;fp=None
                if type_evt=="livraison_reporter":
                    popup_type="livr_report";fp=f"{data.get('date','')}_{data.get('carburant','')}"
                elif type_evt=="rupture":
                    popup_type="rupture_imminente";fp=f"{data.get('jour','')}_{data.get('carburant','')}"
                elif type_evt=="tendance":
                    popup_type="tendance";fp=f"{data.get('date','')}_{data.get('carburant','')}"
                elif type_evt=="saisies_irregulieres":
                    popup_type="saisies_irr";fp=f"{data.get('date','')}_{data.get('carburant','')}"
                elif type_evt=="pont":
                    popup_type="antirupture"
                    try:
                        d_debut=data.get("date_debut","")
                        if d_debut:
                            dt=datetime.fromisoformat(d_debut) if "T" in d_debut else datetime.strptime(d_debut,"%Y-%m-%d")
                            fp=f"pont_{dt.strftime('%d%m%Y')}"
                    except Exception as _e: _log_silent_err(exc=_e)
                elif type_evt=="marge_tendue":
                    popup_type="marge_tendue";fp=f"{data.get('date','')}_{data.get('carburant','')}"
                elif type_evt=="ferie_isole":
                    popup_type="ferie_isole";fp=data.get("date_ferie","")
                if popup_type and fp:
                    silence_popup(popup_type,[fp],until_iso,{fp:0})
            except Exception as _e: _log_silent_err(exc=_e)
            self._refresh()
        except Exception as e: print(f"[marquer resolu evt] {e}")

    def _render_event(self,evt,expand_key=None):
        """Affiche une entrée du journal."""
        type_evt=evt.get("type","")
        ts=evt.get("ts","")
        try:
            ts_dt=datetime.fromisoformat(ts)
            JC_FR=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
            ts_str=f"{JC_FR[ts_dt.weekday()]} {ts_dt.strftime('%d/%m/%y')} \u00e0 {ts_dt.strftime('%Hh%M')}"
        except: ts_str=ts
        # Couleur selon type
        type_colors={"pont":"#C99A5B","anomalie":"#8E7BA6","tendance":"#8E7BA6","rupture":"#C5544D","commande":"#5B9E92","livraison":"#7FA86E","ack":"#6B8FB5","passage_mois":"#4A8A63","livraison_reporter":"#C99A5B","marge_tendue":"#C7A95B","ferie_isole":"#C7A95B","livraison_attendue":"#C5544D"}
        # Cas spécial passage_mois : rouge si l'effet est négatif
        if type_evt=="passage_mois":
            try:
                eff_t=float(evt.get("data",{}).get("effet_total",0))
                if eff_t<0: type_colors={"pont":"#C99A5B","anomalie":"#8E7BA6","tendance":"#8E7BA6","rupture":"#C5544D","commande":"#5B9E92","livraison":"#7FA86E","ack":"#6B8FB5","passage_mois":"#4A8A63","livraison_reporter":"#C99A5B","marge_tendue":"#C7A95B","ferie_isole":"#C7A95B","livraison_attendue":"#C5544D"}
            except Exception as _e: _log_silent_err(exc=_e)
        col=type_colors.get(type_evt,C["t2"])
        type_labels={"pont":"Pont","anomalie":"Anomalie","rupture":"Rupture","commande":"Commande","livraison":"Livraison","ack":"Acquittement","passage_mois":"Passage mois","livraison_reporter":"Livr. à reporter","marge_tendue":"Marge tendue","ferie_isole":"Férié isolé"}
        type_label=type_labels.get(type_evt,type_evt)
        # Distinction Pont (avec férié) vs Weekend (sans férié). Validé Bidou 25/05/2026.
        # Pour Weekend : couleur sobre gris-bleu (palette Apple/Samsung-like).
        if type_evt=="pont":
            try:
                d_pont=evt.get("data",{}) or {}
                dd=d_pont.get("date_debut","");df=d_pont.get("date_fin","")
                if dd:
                    ddt=datetime.fromisoformat(dd).date()
                    dft=datetime.fromisoformat(df).date() if df else ddt
                    cur=ddt;has_ferie=False
                    while cur<=dft:
                        if is_ferie(cur): has_ferie=True;break
                        cur+=timedelta(days=1)
                    if has_ferie:
                        type_label="Pont"
                    else:
                        type_label="Weekend"
                        col="#7B8896"  # gris-bleu sobre
            except Exception as _e: _log_silent_err(exc=_e)
        # Détecter résolu pour style atténué (cohérent avec _render_situation_group)
        is_resolu=(evt.get("data",{}) or {}).get("statut","")=="resolu"
        frame_fg=C["panel"] if is_resolu else C["bg"]
        frame_border=C["border"] if is_resolu else C["border2"]
        # Cadre événement
        frame=ctk.CTkFrame(self.list_frame,fg_color=frame_fg,corner_radius=8,border_width=1,border_color=frame_border)
        frame.pack(fill="x",padx=6,pady=4)
        # Header : badge type + timestamp
        head=ctk.CTkFrame(frame,fg_color="transparent");head.pack(fill="x",padx=12,pady=(8,4))
        badge=ctk.CTkLabel(head,text=type_label,font=("Segoe UI",9,"bold"),text_color="#FFF",
                            fg_color=col,corner_radius=10,width=100,height=20)
        badge.pack(side="left",padx=(0,10))
        ctk.CTkLabel(head,text=ts_str,font=("Segoe UI",10),text_color=C["t3"]).pack(side="left")
        # Bouton "Réduire" si résolu et développé suite à clic
        if expand_key is not None:
            ctk.CTkButton(head,text="\u2bc6 R\u00e9duire",width=80,height=24,
                          fg_color="transparent",hover_color=C["card_h"],text_color=C["t2"],
                          border_width=1,border_color=C["border2"],
                          font=("Segoe UI",9),corner_radius=4,
                          command=lambda k=expand_key:self._toggle_expand(k)).pack(side="right",padx=(0,4))
        # Contenu spécifique au type
        data=evt.get("data",{}) or {}
        body=ctk.CTkFrame(frame,fg_color="transparent");body.pack(fill="x",padx=12,pady=(0,8))
        if type_evt=="pont":
            duree=data.get("duree","?")
            d_deb=data.get("date_debut","");d_fin=data.get("date_fin","")
            # Distinction Pont (contient un férié) vs Weekend (sans férié) + dates lisibles
            # Format dates : "sam. 30/05" au lieu de "2026-05-30" pour lisibilité humaine.
            try:
                ddt=datetime.fromisoformat(d_deb).date() if d_deb else None
                dft=datetime.fromisoformat(d_fin).date() if d_fin else ddt
                JC_short=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
                d_deb_lisible=f"{JC_short[ddt.weekday()]} {ddt.strftime('%d/%m')}" if ddt else d_deb
                d_fin_lisible=f"{JC_short[dft.weekday()]} {dft.strftime('%d/%m')}" if dft else d_fin
                # Factorisé (Étape 3, 27/05/2026) via districarb_core.trous.qualifier_trou
                if ddt and dft:
                    terme=qualifier_trou({"start_date":ddt,"duree":(dft-ddt).days+1})
                else:
                    terme="Pont"
            except Exception as _e:
                _log_silent_err(exc=_e)
                d_deb_lisible=d_deb;d_fin_lisible=d_fin;terme="Pont"
            txt=f"{terme} du {d_deb_lisible} au {d_fin_lisible} ({duree} jour{'s' if duree!=1 else ''}) \u2014 acquitt\u00e9 sous contr\u00f4le"
            ctk.CTkLabel(body,text=txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left",wraplength=860).pack(anchor="w")
            manques=data.get("manques",[])
            if manques:
                manques_str=" | ".join(f"{m.get('carburant','?')} : manque {int(m.get('manque',0))}L" for m in manques)
                ctk.CTkLabel(body,text=manques_str,font=("Segoe UI",9),text_color=C["t3"],anchor="w",wraplength=860).pack(anchor="w",pady=(2,0))
        elif type_evt=="anomalie":
            carb=data.get("carburant","?")
            ecart=data.get("ecart_pct",0)
            jour=data.get("jour","")
            sens="hausse" if ecart>0 else "baisse"
            txt=f"Anomalie de {sens} sur {carb} : {ecart:+.0f}% le {jour}"
            ctk.CTkLabel(body,text=txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",wraplength=860).pack(anchor="w")
        elif type_evt=="rupture":
            carb=data.get("carburant","?")
            # Si libelle riche présent (séquence multi-jours via reconstitute), l'utiliser ;
            # sinon retomber sur le rendu jour unique.
            libelle=data.get("libelle","")
            if libelle:
                txt=libelle
            else:
                jour=data.get("jour","")
                txt=f"Rupture de {carb} le {jour} (niveau sous plancher physique)"
            ctk.CTkLabel(body,text=txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",wraplength=860).pack(anchor="w")
        elif type_evt=="livraison":
            jour=data.get("jour","")
            sp=data.get("sp",0);go=data.get("go",0);gnr=data.get("gnr",0)
            parts=[]
            if sp: parts.append(f"SP {sp:,}L".replace(",","."))
            if go: parts.append(f"GO {go:,}L".replace(",","."))
            if gnr: parts.append(f"GNR {gnr:,}L".replace(",","."))
            txt=f"Livraison du {jour} \u2014 {' / '.join(parts) if parts else 'aucun volume'}"
            # Champs enrichis (depuis Achat_carburant.xlsx) : montant + transporteur si présents
            total_eur=sf(data.get("total_eur",0))
            transporteur=data.get("transporteur","")
            if total_eur>0 or transporteur:
                head_line=ctk.CTkFrame(body,fg_color="transparent");head_line.pack(fill="x")
                ctk.CTkLabel(head_line,text=txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",wraplength=620,justify="left").pack(side="left")
                if total_eur>0:
                    montant=f"\u2192 {total_eur:,.2f} \u20ac".replace(",","\u202f").replace(".",",")
                    ctk.CTkLabel(head_line,text=montant,font=("Segoe UI",11,"bold"),text_color=C["green"],anchor="e").pack(side="right")
                if transporteur:
                    ctk.CTkLabel(body,text=f"Transporteur : {transporteur}",font=("Segoe UI",9),text_color=C["t3"],anchor="w").pack(anchor="w",pady=(2,0))
            else:
                ctk.CTkLabel(body,text=txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",wraplength=860).pack(anchor="w")
        elif type_evt=="passage_mois":
            # Ligne principale : libellé du passage + montant en gras coloré
            label=data.get("label","Passage de mois")
            eff_total=sf(data.get("effet_total",0))
            color_eff=C["green"] if eff_total>=0 else C["red"]
            signe="+" if eff_total>=0 else ""
            montant_str=f"{signe}{eff_total:,.2f} \u20ac".replace(",","\u202f").replace(".",",")
            head_line=ctk.CTkFrame(body,fg_color="transparent");head_line.pack(fill="x")
            ctk.CTkLabel(head_line,text=label,font=("Segoe UI",12,"bold"),text_color=C["t1"],anchor="w").pack(side="left")
            ctk.CTkLabel(head_line,text=montant_str,font=("Segoe UI",13,"bold"),text_color=color_eff,anchor="e").pack(side="right")
            # Ligne secondaire : détail stock pivot par carburant (stock 6h + ventes 0h-6h = stock minuit)
            sp_p=data.get("stock_pivot",{}) or {}
            vb=data.get("ventes_avant_6h",{}) or {}
            def _fmt_pivot(carb,nom):
                s=int(sf(sp_p.get(carb,0)));v=int(sf(vb.get(carb,0)));total=s+v
                return f"{nom} : {s:,} + {v:,} = {total:,} L".replace(",","\u202f")
            detail=" \u2502 ".join([_fmt_pivot("sp","SP"),_fmt_pivot("go","GO"),_fmt_pivot("gnr","GNR")])
            ctk.CTkLabel(body,text=detail,font=("Segoe UI",9),text_color=C["t3"],anchor="w",wraplength=860,justify="left").pack(anchor="w",pady=(2,0))
        elif type_evt=="livraison_reporter":
            carb=data.get("carburant","?")
            try:
                d_iso=data.get("date","")
                d_obj=datetime.fromisoformat(d_iso).date() if d_iso else None
                d_str=d_obj.strftime("%d/%m/%Y") if d_obj else d_iso
            except Exception: d_str=str(data.get("date",""))
            stm=int(sf(data.get("stock_matin",0)))
            livr=int(sf(data.get("livraison",0)))
            surplus=int(sf(data.get("surplus",0)))
            heure_rec=data.get("heure_recommandee")
            statut=data.get("statut","")
            lu=data.get("lu",True)
            fmt=lambda v:f"{v:,}".replace(",","\u202f")
            # Ligne principale avec pastille rouge si statut=non_traite et non lu
            head_line=ctk.CTkFrame(body,fg_color="transparent");head_line.pack(fill="x")
            if statut=="non_traite" and not lu:
                pastille=ctk.CTkFrame(head_line,fg_color=C["red"],width=10,height=10,corner_radius=5)
                pastille.pack(side="left",padx=(0,8),pady=4);pastille.pack_propagate(False)
                # Clic sur la pastille → mini-popup d'actions rapides (Résolu / Annulé / Commenter).
                # Curseur main pour signaler que c'est cliquable.
                pastille.configure(cursor="hand2")
                pastille.bind("<Button-1>",lambda e,eid=evt.get("id"):self._open_event_action(eid))
            txt=f"Livraison {carb} \u00e0 reporter le {d_str} \u2014 surplus cuve {fmt(surplus)} L"
            ctk.CTkLabel(head_line,text=txt,font=("Segoe UI",11,"bold"),text_color=C["t1"],anchor="w").pack(side="left")
            sub=f"Stock matin {fmt(stm)} L + livraison {fmt(livr)} L = d\u00e9passement de {fmt(surplus)} L"
            if data.get("report_au_lendemain"):
                # Cas vitesse vente C1 insuffisante : pas d'heure réaliste dans la journée
                sub+="  \u2192 \u00e0 reporter au lendemain (vente C1 insuffisante pour lib\u00e9rer la place \u00e0 temps)"
            elif heure_rec is not None:
                try:
                    h=float(heure_rec)
                    h_h=int(h);h_m=int((h-h_h)*60)
                    sub+=f"  \u2192 heure recommand\u00e9e : {h_h:02d}h{h_m:02d}"
                except Exception as _e: _log_silent_err(exc=_e)
            ctk.CTkLabel(body,text=sub,font=("Segoe UI",9),text_color=C["t3"],anchor="w",wraplength=860,justify="left").pack(anchor="w",pady=(2,0))
            # Indication du statut (si défini : resolu/snooze/non_traite)
            if statut:
                statut_labels={"resolu":"\u2713 R\u00e9gl\u00e9","snooze":"\u23f1 Rappel demand\u00e9","non_traite":"\u26a0 \u00c0 traiter","annule":"\u2717 Annul\u00e9","ack":"\u2713 Pris en compte"}
                statut_colors={"resolu":C["green"],"snooze":C["amber"],"non_traite":C["red"],"annule":C["t3"],"ack":"#5C9DDA"}
                statut_txt=statut_labels.get(statut,statut)
                tour=data.get("tour","")
                if tour: statut_txt+=f" ({tour})"
                ctk.CTkLabel(body,text=statut_txt,font=("Segoe UI",10,"bold"),
                             text_color=statut_colors.get(statut,C["t2"]),
                             anchor="w").pack(anchor="w",pady=(2,0))
        elif type_evt=="marge_tendue":
            # Format propre pour l'événement marge tendue (cuve presque pleine).
            carb=data.get("carburant","?")
            try:
                d_iso=data.get("date","")
                d_obj=datetime.fromisoformat(d_iso).date() if d_iso else None
                d_str=d_obj.strftime("%d/%m/%Y") if d_obj else d_iso
            except Exception: d_str=str(data.get("date",""))
            marge=int(sf(data.get("marge_restante",0)))
            statut=data.get("statut","")
            lu=data.get("lu",True)
            fmt=lambda v:f"{v:,}".replace(",","\u202f")
            head_line=ctk.CTkFrame(body,fg_color="transparent");head_line.pack(fill="x")
            if statut=="non_traite" and not lu:
                pastille=ctk.CTkFrame(head_line,fg_color=C["red"],width=10,height=10,corner_radius=5)
                pastille.pack(side="left",padx=(0,8),pady=4);pastille.pack_propagate(False)
                pastille.configure(cursor="hand2")
                pastille.bind("<Button-1>",lambda e,eid=evt.get("id"):self._open_event_action(eid))
            txt=f"Marge cuve tendue {carb} \u2014 livraison du {d_str}"
            ctk.CTkLabel(head_line,text=txt,font=("Segoe UI",11,"bold"),text_color=C["t1"],anchor="w").pack(side="left")
            sub=f"Marge restante apr\u00e8s livraison : {fmt(marge)} L (seuil d'alerte : 4 000 L)"
            ctk.CTkLabel(body,text=sub,font=("Segoe UI",9),text_color=C["t3"],anchor="w",wraplength=860,justify="left").pack(anchor="w",pady=(2,0))
            if statut:
                statut_labels={"resolu":"\u2713 R\u00e9gl\u00e9","snooze":"\u23f1 Rappel demand\u00e9","non_traite":"\u26a0 \u00c0 traiter","annule":"\u2717 Annul\u00e9","ack":"\u2713 Pris en compte"}
                statut_colors={"resolu":C["green"],"snooze":C["amber"],"non_traite":C["red"],"annule":C["t3"],"ack":"#5C9DDA"}
                statut_txt=statut_labels.get(statut,statut)
                ctk.CTkLabel(body,text=statut_txt,font=("Segoe UI",10,"bold"),
                             text_color=statut_colors.get(statut,C["t2"]),
                             anchor="w").pack(anchor="w",pady=(2,0))
        elif type_evt=="ferie_isole":
            # Format propre pour l'événement férié isolé imminent.
            try:
                d_iso=data.get("date_ferie","")
                d_obj=datetime.fromisoformat(d_iso).date() if d_iso else None
                JC_FR2=["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
                d_str=f"{JC_FR2[d_obj.weekday()]} {d_obj.strftime('%d/%m/%Y')}" if d_obj else d_iso
            except Exception: d_str=str(data.get("date_ferie",""))
            nom=data.get("nom_ferie","?")
            statut=data.get("statut","")
            lu=data.get("lu",True)
            head_line=ctk.CTkFrame(body,fg_color="transparent");head_line.pack(fill="x")
            if statut=="non_traite" and not lu:
                pastille=ctk.CTkFrame(head_line,fg_color=C["red"],width=10,height=10,corner_radius=5)
                pastille.pack(side="left",padx=(0,8),pady=4);pastille.pack_propagate(False)
                pastille.configure(cursor="hand2")
                pastille.bind("<Button-1>",lambda e,eid=evt.get("id"):self._open_event_action(eid))
            txt=f"F\u00e9ri\u00e9 isol\u00e9 imminent : {nom} \u2014 {d_str}"
            ctk.CTkLabel(head_line,text=txt,font=("Segoe UI",11,"bold"),text_color=C["t1"],anchor="w").pack(side="left")
            sub="Pas de livraison SARA ce jour-l\u00e0. Commande \u00e0 anticiper la veille avant 11h."
            ctk.CTkLabel(body,text=sub,font=("Segoe UI",9),text_color=C["t3"],anchor="w",wraplength=860,justify="left").pack(anchor="w",pady=(2,0))
            if statut:
                statut_labels={"resolu":"\u2713 R\u00e9gl\u00e9","snooze":"\u23f1 Rappel demand\u00e9","non_traite":"\u26a0 \u00c0 traiter","annule":"\u2717 Annul\u00e9","ack":"\u2713 Pris en compte"}
                statut_colors={"resolu":C["green"],"snooze":C["amber"],"non_traite":C["red"],"annule":C["t3"],"ack":"#5C9DDA"}
                statut_txt=statut_labels.get(statut,statut)
                ctk.CTkLabel(body,text=statut_txt,font=("Segoe UI",10,"bold"),
                             text_color=statut_colors.get(statut,C["t2"]),
                             anchor="w").pack(anchor="w",pady=(2,0))
        elif type_evt=="ack" and data.get("sujet")=="livraison_exceptionnelle_forcee":
            # Forçage d'une livraison sur jour non-livrable assumé par Bidou
            try:
                ds=(data.get("date","") or "")[:10]
                d_obj=datetime.fromisoformat(ds).date()
                d_disp=self._fmt_date_jour(d_obj)+d_obj.strftime(" /%Y")
            except Exception:
                d_disp=str(data.get("date",""))[:10]
            vol=int(data.get("volume",0) or 0)
            txt=f"\u2713 Livraison exceptionnelle forc\u00e9e \u2014 {d_disp}"
            ctk.CTkLabel(body,text=txt,font=("Segoe UI",11,"bold"),
                         text_color=C["t1"],anchor="w").pack(side="left")
            sub=(f"Livraison sur jour non-livrable valid\u00e9e volontairement"
                 f"{f' ({vol:,} L)'.replace(',','.') if vol else ''}. "
                 f"Le hub ne la signalera plus comme incoh\u00e9rence.")
            ctk.CTkLabel(body,text=sub,font=("Segoe UI",9),text_color=C["t3"],
                         anchor="w",wraplength=860,justify="left").pack(anchor="w",pady=(2,0))
        elif type_evt=="commande":
            # Format propre pour l'événement commande passée. Avant ajout de cette branche
            # (27/05/2026), les events de type "commande" tombaient dans le fallback générique
            # qui affichait du dict brut type :
            #   "jour: 2026-05-28 · sp: 14000 · go: 10000 · gnr: 0 · tour: 1 · premier voyage: False"
            # Pattern aligné sur les branches voisines (marge_tendue, ferie_isole) :
            # head_line + pastille rouge si non traité + sous-texte avec volumes + statut.
            try:
                d_iso=(data.get("jour","") or data.get("date","") or "")[:10]
                d_obj=datetime.fromisoformat(d_iso).date() if d_iso else None
            except Exception: d_obj=None
            if d_obj:
                JC_FR_CMD=["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
                d_str=f"{JC_FR_CMD[d_obj.weekday()]} {d_obj.strftime('%d/%m/%Y')}"
            else:
                d_str=str(data.get("jour","") or data.get("date",""))
            sp=int(data.get("sp",0) or 0)
            go=int(data.get("go",0) or 0)
            gnr=int(data.get("gnr",0) or 0)
            tour=int(data.get("tour",1) or 1)
            premier_voyage=bool(data.get("premier_voyage",False))
            statut=data.get("statut","")
            lu=data.get("lu",True)
            fmt=lambda v:f"{v:,}".replace(",","\u202f")
            tour_str={1:"1\u1d49\u02b3 tour",2:"2\u1d49 tour",3:"3\u1d49 tour"}.get(tour,f"tour {tour}")
            vols=[]
            if sp>0: vols.append(f"SP {fmt(sp)} L")
            if go>0: vols.append(f"GO {fmt(go)} L")
            if gnr>0: vols.append(f"GNR {fmt(gnr)} L")
            vols_str=" \u00b7 ".join(vols) if vols else "volumes \u00e0 saisir"
            head_line=ctk.CTkFrame(body,fg_color="transparent");head_line.pack(fill="x")
            if statut=="non_traite" and not lu:
                pastille=ctk.CTkFrame(head_line,fg_color=C["red"],width=10,height=10,corner_radius=5)
                pastille.pack(side="left",padx=(0,8),pady=4);pastille.pack_propagate(False)
                pastille.configure(cursor="hand2")
                pastille.bind("<Button-1>",lambda e,eid=evt.get("id"):self._open_event_action(eid))
            txt=f"Commande pass\u00e9e \u2014 livraison {d_str}" if d_obj else f"Commande pass\u00e9e \u2014 {d_str}"
            ctk.CTkLabel(head_line,text=txt,font=("Segoe UI",11,"bold"),text_color=C["t1"],anchor="w").pack(side="left")
            suf=" \u00b7 1er voyage" if premier_voyage else ""
            sub=f"{vols_str} ({tour_str}{suf})"
            ctk.CTkLabel(body,text=sub,font=("Segoe UI",9),text_color=C["t3"],anchor="w",wraplength=860,justify="left").pack(anchor="w",pady=(2,0))
            if statut:
                statut_labels={"resolu":"\u2713 R\u00e9gl\u00e9","snooze":"\u23f1 Rappel demand\u00e9","non_traite":"\u26a0 \u00c0 traiter","annule":"\u2717 Annul\u00e9","ack":"\u2713 Pris en compte"}
                statut_colors={"resolu":C["green"],"snooze":C["amber"],"non_traite":C["red"],"annule":C["t3"],"ack":"#5C9DDA"}
                statut_txt=statut_labels.get(statut,statut)
                ctk.CTkLabel(body,text=statut_txt,font=("Segoe UI",10,"bold"),
                             text_color=statut_colors.get(statut,C["t2"]),
                             anchor="w").pack(anchor="w",pady=(2,0))
        elif type_evt=="tendance":
            carb=(data.get("carburant","?") or "?").upper()
            ec=data.get("ecart_pct",0)
            sg=data.get("stage","")
            signe="+" if (isinstance(ec,(int,float)) and ec>0) else ""
            sens="hausse" if (isinstance(ec,(int,float)) and ec>0) else "baisse"
            txt=(f"{carb} en {sens} : ventes {signe}{ec}% \u00e0 {sg} vs moyenne" if sg
                 else f"{carb} en {sens} : ventes {signe}{ec}% vs moyenne")
            ctk.CTkLabel(body,text=txt,font=("Segoe UI",11),text_color=C["t1"],
                         anchor="w",justify="left",wraplength=860).pack(anchor="w")
        elif type_evt=="livraison_attendue":
            tl=data.get("tour_label","Tour")
            hl=data.get("heure_limite","")
            txt=f"{tl} d\u00e9pass\u00e9{(' ('+str(hl)+'h)') if hl else ''} \u2014 camion attendu non confirm\u00e9 arriv\u00e9"
            ctk.CTkLabel(body,text=txt,font=("Segoe UI",11),text_color=C["t1"],
                         anchor="w",justify="left",wraplength=860).pack(anchor="w")
            ctk.CTkLabel(body,text="Action : v\u00e9rifier aupr\u00e8s de TotalEnergies.",font=("Segoe UI",9),
                         text_color=C["t3"],anchor="w").pack(anchor="w",pady=(2,0))
        else:
            # Fallback générique LISIBLE (jamais de dict brut à l'écran) :
            # on liste les paires clé→valeur en clair plutôt que str(data).
            try:
                if isinstance(data,dict) and data:
                    pretty=" \u00b7 ".join(
                        f"{k.replace('_',' ')}: {v}"
                        for k,v in data.items()
                        if k not in ("lu",) and v not in (None,"",[]))
                else:
                    pretty=str(data)
            except Exception:
                pretty=str(data)
            ctk.CTkLabel(body,text=pretty or "\u2014",font=("Segoe UI",10),
                         text_color=C["t2"],anchor="w",wraplength=860,
                         justify="left").pack(anchor="w")
        # Commentaire (si présent)
        comm=evt.get("commentaire")
        if comm:
            ctk.CTkLabel(body,text=f"\U0001f4ac {comm}",font=("Segoe UI",10,"italic"),text_color=C["amber"],
                          anchor="w",justify="left",wraplength=860).pack(anchor="w",pady=(4,0))
        # Bouton "Marquer résolu" + "Effacer" sur les cartes événement SEUL (non regroupées).
        # Bouton Effacer (discret, validé Bidou 25/05/2026) restreint aux types d'ALERTE.
        # Création conditionnelle du frame actions_row (validé Bidou 27/05/2026) : sinon
        # CTkFrame vide garde sa hauteur par défaut ~200px, créant un vide géant sur les
        # cartes Livraison/Commande qui n'ont ni statut actif ni bouton Effacer.
        statut_evt=(evt.get("data",{}) or {}).get("statut","")
        type_evt_curr=evt.get("type","")
        TYPES_ALERTE={"pont","anomalie","rupture","marge_tendue","livraison_reporter","ferie_isole","saisies_irregulieres"}
        show_resolve=bool(statut_evt and statut_evt!="resolu")
        show_erase=type_evt_curr in TYPES_ALERTE
        if show_resolve or show_erase:
            actions_row=ctk.CTkFrame(body,fg_color="transparent");actions_row.pack(anchor="e",pady=(6,0))
            if show_resolve:
                ctk.CTkButton(actions_row,text="\u2713 Marquer comme r\u00e9gl\u00e9",
                    fg_color=C["panel"],hover_color=C["card_h"],text_color=C["green"],
                    border_width=1,border_color=C["green"],font=("Segoe UI",10),
                    corner_radius=6,height=26,width=140,
                    command=lambda eid=evt.get("id"): self._marquer_resolu_evt(eid)).pack(side="left",padx=(0,6))
            if show_erase:
                evt_id_one=evt.get("id")
                evt_label_short=self._situation_label(evt) if evt else "cet \u00e9v\u00e9nement"
                ctk.CTkButton(actions_row,text="\U0001f5d1",width=28,height=26,
                    fg_color="transparent",hover_color="#2A1518",text_color=C["t3"],
                    border_width=0,font=("Segoe UI Emoji",13),corner_radius=6,
                    command=lambda i=evt_id_one,lbl=evt_label_short: self._effacer_situation([i],lbl)).pack(side="left")


class LivraisonAttendueDlg(ctk.CTkToplevel):
    """Brique 2 — popup DÉDIÉ du matin. Une livraison est attendue aujourd'hui
    (pilotée par commandes.cfg). Demande si le camion est arrivé.
      - "Oui" → flux d'ajout livraison EXISTANT (aucun nouveau formulaire,
        aucune logique de save dupliquée).
      - "Pas encore" / croix X → silence du jour ; l'alerte CALIBRÉE tombe à
        l'heure LIMITE du tour visé (escalade dans _check_time_alerts).
    Ne touche JAMAIS LivraisonDialog ni son "Plus tard (4h)"."""
    def __init__(self,parent,cmd):
        super().__init__(parent)
        self.hub=parent
        self.cmd=cmd or {}
        self.title("Livraison attendue \u2014 DISTRICARB HUB")
        self.geometry("560x500");self.minsize(520,460)
        self.configure(fg_color=C["bg"]);self.resizable(False,False)
        self.transient(parent);self.grab_set()
        self.protocol("WM_DELETE_WINDOW",self._pas_encore)
        tour=int(self.cmd.get("tour",1) or 1)
        info=TOURS_LIVRAISON.get(tour,TOURS_LIVRAISON[1])
        ctk.CTkLabel(self,text="\U0001f4e6  Livraison attendue aujourd'hui",
                     font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(pady=(22,2))
        ctk.CTkLabel(self,text=f"{jour_fr()} {date.today().strftime('%d/%m/%Y')}",
                     font=("Segoe UI",13),text_color=C["gold"]).pack(pady=(0,14))
        box=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8);box.pack(fill="x",padx=30,pady=(0,16))
        sp=int(self.cmd.get("sp",0));go=int(self.cmd.get("go",0));gnr=int(self.cmd.get("gnr",0))
        ctk.CTkLabel(box,text=f"Commande pass\u00e9e : SP {sp} L \u00b7 GO {go} L \u00b7 GNR {gnr} L",
                     font=("Segoe UI",13,"bold"),text_color=C["t1"],
                     anchor="w",justify="left",wraplength=470).pack(fill="x",padx=14,pady=(12,4))
        ctk.CTkLabel(box,text=f"{info['label']} ({info['plage']})",
                     font=("Segoe UI",12),text_color=C["t2"],anchor="w").pack(fill="x",padx=14,pady=(0,2))
        if self.cmd.get("premier_voyage"):
            ctk.CTkLabel(box,text="\u26a1 Premier voyage demand\u00e9 (priorit\u00e9 file SARA)",
                         font=("Segoe UI",12,"bold"),text_color=C["gold"],anchor="w").pack(fill="x",padx=14,pady=(0,2))
        ctk.CTkLabel(box,text=f"Sans saisie, alerte \u00e0 {info['alerte']}h (fin du {info['label']}).",
                     font=("Segoe UI",11),text_color=C["t3"],anchor="w").pack(fill="x",padx=14,pady=(0,12))
        ctk.CTkLabel(self,text="Le camion est-il arriv\u00e9 ?",font=("Segoe UI",14,"bold"),
                     text_color=C["t1"]).pack(pady=(2,8))
        btns=ctk.CTkFrame(self,fg_color="transparent");btns.pack(side="bottom",fill="x",padx=30,pady=(10,18))
        ctk.CTkButton(btns,text="\u2713 Oui \u2014 saisir la livraison",width=240,height=44,
                      fg_color=C["green"],hover_color="#258A3E",text_color="#FFF",
                      font=("Segoe UI",13,"bold"),corner_radius=8,command=self._oui).pack(side="right")
        ctk.CTkButton(btns,text="Pas encore",width=140,height=44,fg_color=C["card"],
                      hover_color=C["card_h"],border_width=1,border_color=C["border2"],
                      text_color=C["amber"],corner_radius=8,command=self._pas_encore).pack(side="left")

    def _oui(self):
        try: clear_popup_silence("livraison_attendue")
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()
        try:
            if hasattr(self.hub,"_open_livraisons"): self.hub._open_livraisons()
        except Exception as _e: _log_silent_err(exc=_e)

    def _pas_encore(self):
        # Silence du jour : on a demandé, Bidou sait. L'alerte calibrée sur
        # l'heure limite du tour prend le relais (pas de "+4h" arbitraire).
        try:
            today_fp=date.today().strftime("%Y-%m-%d")
            until_iso=datetime.combine(date.today(),datetime.min.time()).replace(hour=23,minute=59).isoformat()
            silence_popup("livraison_attendue",[today_fp],until_iso,{today_fp:0})
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()


class CommandeDialog(ctk.CTkToplevel):
    """Fenêtre DÉDIÉE à la commande du jour (Brique 1). Une seule chose :
    enregistrer la commande passée chez Total. AUCUNE saisie de livraison reçue
    ici (ça reste sur "+ Ajouter une livraison"), AUCUN "Plus tard 4h", AUCUN
    "Aucune prévue". Le hub PROPOSE (attendu Pre_vision + criticité réelle), il
    ne décide pas, il n'invente pas. parent = le HUB (pour lire last_data et
    rafraîchir)."""
    def __init__(self,parent):
        super().__init__(parent)
        self.hub=parent
        self.result=None
        self.title("Commande du jour \u2014 DISTRICARB HUB")
        self.geometry("560x600");self.minsize(520,560)
        self.configure(fg_color=C["bg"]);self.resizable(False,False)
        # FIX 21/05/2026 : popup qui se cachait derrière le hub principal (bug récurrent
        # forçant fermeture sauvage du hub via taskkill). Cause racine : `grab_set()` peut
        # échouer silencieusement sur Windows si la fenêtre n'est pas encore viewable.
        # Solution : update_idletasks() force le rendu, puis lift+focus_force+topmost
        # temporaire (200ms) garantit que la popup arrive devant et capture le focus.
        # `after(200, ...)` retire le topmost après pour ne pas bloquer Alt+Tab.
        self.transient(parent)
        try: self.update_idletasks()
        except Exception as _e: _log_silent_err(exc=_e)
        try: self.grab_set()
        except Exception as _e: _log_silent_err(exc=_e)
        try:
            self.lift();self.focus_force()
            self.attributes("-topmost",True)
            self.after(200,lambda:self.attributes("-topmost",False))
        except Exception as _e: _log_silent_err(exc=_e)
        ctx=self._ctx()
        self.cmd_target=ctx.get("target")
        a=ctx.get("attendu",{}) or {}
        # En-tête
        ctk.CTkLabel(self,text="\U0001f4e6  Commande du jour",font=("Segoe UI",18,"bold"),
                     text_color=C["t1"]).pack(pady=(20,2))
        # Date de livraison : pré-remplie INTELLIGEMMENT (forçage-aware via _ctx),
        # mais ÉDITABLE — porte de secours discrète si la proposition est fausse
        # (popup d'alerte ratée, cas tordu). Réutilise EXACTEMENT le sélecteur de
        # "+ Ajouter une livraison" (tkcalendar si dispo, sinon JJ/MM/AA).
        default_dt=self.cmd_target or date.today()
        fdate=ctk.CTkFrame(self,fg_color="transparent");fdate.pack(pady=(0,14))
        ctk.CTkLabel(fdate,text="Livraison du :",font=("Segoe UI",13),
                     text_color=C["gold"]).pack(side="left",padx=(0,8))
        self._tkcal_ok=False
        try:
            from tkcalendar import DateEntry
            self.e_date=DateEntry(fdate,width=12,locale="fr_FR",date_pattern="dd/mm/yy",
                                   font=("Segoe UI",13),background=C["card"],foreground=C["t1"],
                                   borderwidth=2,headersbackground=C["card_h"],headersforeground=C["t1"],
                                   selectbackground=C["red"],selectforeground="#FFFFFF",
                                   normalbackground=C["panel"],normalforeground=C["t1"],
                                   weekendbackground=C["panel"],weekendforeground=C["t2"])
            self.e_date.set_date(default_dt)
            self.e_date.pack(side="left",ipady=4)
            self._tkcal_ok=True
            ctk.CTkLabel(fdate,text="\U0001f4c5",font=("Segoe UI",11),text_color=C["t3"]).pack(side="left",padx=(8,0))
        except ImportError:
            self.e_date=ctk.CTkEntry(fdate,height=34,width=130,fg_color=C["card"],border_color=C["border"],
                                      text_color=C["t1"],font=("Segoe UI",13),placeholder_text="JJ/MM/AA")
            self.e_date.pack(side="left");self.e_date.insert(0,default_dt.strftime("%d/%m/%y"))
            ctk.CTkLabel(fdate,text="JJ/MM/AA",font=("Segoe UI",9),text_color=C["amber"]).pack(side="left",padx=(8,0))
        # --- Le hub propose (lecture seule) ---
        box=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8);box.pack(fill="x",padx=30,pady=(0,14))
        if ctx.get("attendu_found"):
            att=(f"Attendu Pre_vision : SP {a.get('sp',0)} L \u00b7 "
                 f"GO {a.get('go',0)} L \u00b7 GNR {a.get('gnr',0)} L")
        else:
            att="Pre_vision : aucune livraison pr\u00e9vue pour cette date."
        ctk.CTkLabel(box,text=att,font=("Segoe UI",12),text_color=C["t2"],
                     anchor="w",justify="left",wraplength=470).pack(fill="x",padx=14,pady=(10,4))
        if ctx.get("tendu"):
            # FIX 21/05/2026 (chemin B) : afficher la marge restante en L pour chaque
            # carburant tendu (quand l'info est disponible via livraisons_marge_tendue).
            # Donne à Bidou une info concrète "SP marge cuve 2508 L" au lieu d'un simple
            # "SP", cohérent avec ce que MargeTendueDlg lui montre par ailleurs.
            mt_target=ctx.get("marges_tendues_target",{}) or {}
            parts=[]
            for c in ctx["tendu"]:
                if c in mt_target:
                    parts.append(f"{c} (marge cuve {mt_target[c]:,} L)".replace(",","."))
                else:
                    parts.append(c)
            crit=f"\u26a0 Tension : {', '.join(parts)}"
            if ctx.get("deadline_str"): crit+=f" \u2014 commander avant {ctx['deadline_str']}"
            ctk.CTkLabel(box,text=crit,font=("Segoe UI",12,"bold"),text_color=C["red"],
                         anchor="w",justify="left",wraplength=470).pack(fill="x",padx=14,pady=(0,4))
        else:
            ctk.CTkLabel(box,text="\u2713 Pas de tension d\u00e9tect\u00e9e par le moteur sur cette date.",
                         font=("Segoe UI",12),text_color=C["green"],
                         anchor="w").pack(fill="x",padx=14,pady=(0,4))
        # Tensions sur les JOURS PROCHES (target+1 à +7) : on les remonte aussi en orange,
        # car la popup commande gagne à anticiper. Si le moteur dit "pas de tension le 21/05"
        # mais qu'il y en a une le 22/05, c'est important pour Bidou de le voir ici plutôt
        # qu'à devoir aller fouiller le tableau de notifications. Aligne le hub avec le
        # principe "fil rouge — modules qui se relisent".
        jp=ctx.get("tensions_jours_proches",[]) or []
        if jp:
            # Grouper par date
            par_date={}
            for t in jp:
                d=t.get("date")
                par_date.setdefault(d,[]).append(t)
            # Formater
            JC=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
            lignes=[]
            for d in sorted(par_date.keys()):
                items=par_date[d]
                carbs=", ".join(sorted({t["carburant"] for t in items}))
                jr=f"{JC[d.weekday()]} {d.strftime('%d/%m')}" if hasattr(d,"weekday") else str(d)
                lignes.append(f"{jr} : {carbs}")
            txt="\u26a0 Tensions \u00e0 venir : "+" \u00b7 ".join(lignes)
            ctk.CTkLabel(box,text=txt,font=("Segoe UI",11),text_color=C["amber"],
                         anchor="w",justify="left",wraplength=470).pack(fill="x",padx=14,pady=(0,4))
        reco=ctx.get("reco")
        if reco:
            r=f"Reco moteur : {' \u00b7 '.join(reco['lignes'])}"
            if reco.get("tour_3"): r+="  (3e tour conseill\u00e9)"
            if reco.get("infaisable"): r+="  \u26a0 vol. infaisable sur ce cycle"
            ctk.CTkLabel(box,text=r,font=("Segoe UI",11),text_color=C["gold"],
                         anchor="w",justify="left",wraplength=470).pack(fill="x",padx=14,pady=(0,10))
        else:
            ctk.CTkLabel(box,text=" ",font=("Segoe UI",2),fg_color="transparent").pack(pady=(0,6))
        # --- Ta commande réellement passée ---
        ctk.CTkLabel(self,text="Ta commande r\u00e9ellement pass\u00e9e chez Total :",
                     font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(anchor="w",padx=32,pady=(0,6))
        existing=get_commande(self.cmd_target) if self.cmd_target else None
        rowf=ctk.CTkFrame(self,fg_color="transparent");rowf.pack(fill="x",padx=30,pady=(0,8))
        self.cmd_entries={}
        for carb,color in [("SP",C["blue"]),("GO",C["amber"]),("GNR",C["teal"])]:
            cell=ctk.CTkFrame(rowf,fg_color="transparent");cell.pack(side="left",expand=True,fill="x",padx=4)
            ctk.CTkLabel(cell,text=carb,font=("Segoe UI",12,"bold"),text_color=color).pack(anchor="w")
            # FIX 21/05/2026 (Bug B) : on PRÉ-REMPLIT le champ au lieu d'utiliser un
            # placeholder gris. Avant : la valeur Pre_vision s'affichait en gris (placeholder)
            # mais .get() retournait "" tant que Bidou n'avait pas explicitement re-tapé.
            # Conséquence : clic Enregistrer sans modifier → csp+cgo+cgnr=0 → la popup
            # se fermait silencieusement sans rien enregistrer (cas samedi 23/05 où Bidou
            # voulait commander exactement les valeurs Pre_vision suggérées).
            # Maintenant : pré-remplissage réel → si Bidou accepte la suggestion telle
            # quelle, l'enregistrement passe. S'il modifie, il modifie. S'il efface
            # complètement à 0, il a pris une décision consciente.
            v_attendu=int(a.get(carb.lower(),0)) if ctx.get("attendu_found") else 0
            e=ctk.CTkEntry(cell,height=36,fg_color=C["card"],border_color=C["border"],
                           text_color=C["t1"],font=("Segoe UI",14),placeholder_text="0")
            e.pack(fill="x")
            # Priorité de pré-remplissage : commande existante > attendu Pre_vision > vide
            if existing and existing.get(carb.lower()):
                e.insert(0,str(int(existing.get(carb.lower(),0))))
            elif v_attendu>0:
                e.insert(0,str(v_attendu))
            self.cmd_entries[carb.lower()]=e
        self.cmd_tour=int(existing["tour"]) if (existing and existing.get("tour")) else 1
        tline=ctk.CTkFrame(self,fg_color="transparent");tline.pack(fill="x",padx=32,pady=(6,2))
        ctk.CTkLabel(tline,text="Tour vis\u00e9 :",font=("Segoe UI",12),
                     text_color=C["t2"],width=82,anchor="w").pack(side="left")
        self._tour_btns={}
        for n in (1,2,3):
            b=ctk.CTkButton(tline,text=TOURS_LIVRAISON[n]["label"],width=96,height=32,
                            corner_radius=6,font=("Segoe UI",11,"bold"),
                            command=lambda nn=n:self._set_tour(nn))
            b.pack(side="left",padx=4);self._tour_btns[n]=b
        self._set_tour(self.cmd_tour)
        self.cmd_premier=ctk.CTkCheckBox(self,text="Premier voyage (priorit\u00e9 en t\u00eate de file SARA)",
                                         font=("Segoe UI",11),text_color=C["t2"],
                                         fg_color=C["gold"],hover_color=C["gold"])
        self.cmd_premier.pack(anchor="w",padx=32,pady=(10,2))
        if existing and existing.get("premier_voyage"): self.cmd_premier.select()
        # --- Boutons : Enregistrer / Annuler UNIQUEMENT ---
        btns=ctk.CTkFrame(self,fg_color="transparent");btns.pack(side="bottom",fill="x",padx=30,pady=(14,16))
        ctk.CTkButton(btns,text="\u2713 Enregistrer",width=160,height=42,fg_color=C["green"],
                      hover_color="#258A3E",text_color="#FFF",font=("Segoe UI",13,"bold"),
                      corner_radius=8,command=self._save).pack(side="right")
        ctk.CTkButton(btns,text="Annuler",width=120,height=42,fg_color=C["card"],
                      hover_color=C["card_h"],border_width=1,border_color=C["border2"],
                      text_color=C["t2"],corner_radius=8,command=self._cancel).pack(side="left")

    def _ctx(self):
        """Date cible (prochain jour livrable strict) + attendu Pre_vision +
        criticité RÉELS lus dans hub.last_data. Aucune invention ; si une donnée
        manque, le champ est marqué absent et l'UI le dit honnêtement."""
        ctx={"target":None,"attendu":{"sp":0,"go":0,"gnr":0},"attendu_found":False,
             "tendu":[],"deadline_str":"","reco":None}
        try:
            t=date.today()+timedelta(days=1)
            # Le hub LIT le forçage déjà posé (résolution de l'alerte Prévision
            # via AntiRuptureDlg → FORCAGE_FILE). Un jour non-livrable mais DÉJÀ
            # forcé est une cible valide : décision déjà prise ailleurs, on la
            # RELIT au lieu de la re-demander (modules interconnectés, fil rouge).
            while (t.weekday()>=5 or is_ferie(t)) and not is_date_forcee(t):
                t+=timedelta(days=1)
            ctx["target"]=t
        except Exception as _e:
            _log_silent_err(exc=_e); return ctx
        ld=getattr(self.hub,"last_data",None) or {}
        try:
            for day in ((ld.get("proj14") or {}).get("projection") or []):
                if day.get("date")==ctx["target"]:
                    av={"sp":int(sf(day.get("livr_sp",0))),
                        "go":int(sf(day.get("livr_go",0))),
                        "gnr":int(sf(day.get("livr_gnr",0)))}
                    ctx["attendu"]=av
                    ctx["attendu_found"]=any(v>0 for v in av.values())
                    break
        except Exception as _e: _log_silent_err(exc=_e)
        try:
            for al in ((ld.get("proj14") or {}).get("alertes") or []):
                if al.get("severity")=="rupture":
                    c=al.get("carburant","?")
                    if c not in ctx["tendu"]: ctx["tendu"].append(c)
                    if not ctx["deadline_str"]:
                        ctx["deadline_str"]=al.get("deadline_str","")
        except Exception as _e: _log_silent_err(exc=_e)
        # FIL ROUGE — modules qui se relisent : la popup commande lit AUSSI le
        # moteur anti-rupture, pas seulement proj14. Constat terrain Bidou (20/05) :
        # Pre_vision Excel calcule "Impossible -556L" sur SP jeu. 21/05, le tableau
        # notif affiche "SP — livraison 21/05 · dépassement 556L · heure recommandée
        # 07h00", mais cette même popup commande dit "✓ Pas de tension détectée".
        # Bug : _ctx() ne lisait QUE proj14.alertes (alerte d'autonomie courte), pas
        # les dépassements de capacité cuve qui sont dans antirupture.livraisons_a_reporter
        # ni les saisies impossibles dans antirupture.saisies_physiquement_impossibles.
        # Maintenant on lit ces 2 sources et on ajoute leurs carburants à ctx["tendu"].
        # En plus, on capture aussi les TENSIONS SUR LES JOURS PROCHES (target+1 à +7)
        # pour les afficher en avertissement orange — ainsi même si target ne match
        # pas pile, Bidou voit qu'il y a un problème à venir.
        ctx["tensions_jours_proches"]=[]  # liste de (date, carb, surplus)
        try:
            ar=ld.get("antirupture") or {}
            target_d=ctx["target"]
            for lr in (ar.get("livraisons_a_reporter",[]) or []):
                d=lr.get("date")
                if d==target_d:
                    c=lr.get("carburant","?")
                    if c not in ctx["tendu"]: ctx["tendu"].append(c)
                    if not ctx["deadline_str"]:
                        # Heure recommandée de livraison (calculée via vente C1)
                        h=lr.get("heure_recommandee")
                        if h is not None:
                            h_h=int(h);h_m=int((h-h_h)*60)
                            ctx["deadline_str"]=f"livrer apr\u00e8s {h_h:02d}h{h_m:02d}"
                elif target_d is not None and d is not None and hasattr(d,"toordinal"):
                    # Tension sur un jour PROCHE (1 à 7 jours après target)
                    delta=(d-target_d).days
                    if 0<delta<=7:
                        ctx["tensions_jours_proches"].append({
                            "date":d,"carburant":lr.get("carburant","?"),
                            "surplus":lr.get("surplus",0),"type":"livr_report"})
            for s in (ar.get("saisies_physiquement_impossibles",[]) or []):
                d=s.get("date")
                if d==target_d:
                    c=s.get("carburant","?")
                    if c not in ctx["tendu"]: ctx["tendu"].append(c)
                elif target_d is not None and d is not None and hasattr(d,"toordinal"):
                    delta=(d-target_d).days
                    if 0<delta<=7:
                        ctx["tensions_jours_proches"].append({
                            "date":d,"carburant":s.get("carburant","?"),
                            "exces":s.get("exces",0),"type":"saisies_irr"})
            # FIX 21/05/2026 (chemin B Bidou) : lecture de livraisons_marge_tendue
            # — la 3e source de tension que _ctx ignorait. C'est la même liste qui
            # alimente MargeTendueDlg (ligne ~11050). Sans cette lecture, la popup
            # commande disait "Pas de tension détectée" pour samedi 23/05 alors que
            # la popup marge tendue criait au même moment (marge cuve 2508 L < 4000 L
            # seuil "Attention" aligné Excel Bidou). Cohérence rétablie : les deux
            # popups parlent maintenant au même moteur antirupture, plus de divergence.
            # On stocke aussi marge_restante (en L) pour enrichir le label affiché.
            ctx["marges_tendues_target"]={}  # carb -> marge_restante pour target
            for mt in (ar.get("livraisons_marge_tendue",[]) or []):
                d=mt.get("date")
                if d==target_d:
                    c=mt.get("carburant","?")
                    if c not in ctx["tendu"]: ctx["tendu"].append(c)
                    ctx["marges_tendues_target"][c]=int(sf(mt.get("marge_restante",0)))
                elif target_d is not None and d is not None and hasattr(d,"toordinal"):
                    delta=(d-target_d).days
                    if 0<delta<=7:
                        ctx["tensions_jours_proches"].append({
                            "date":d,"carburant":mt.get("carburant","?"),
                            "marge_restante":int(sf(mt.get("marge_restante",0))),
                            "type":"marge_tendue"})
        except Exception as _e: _log_silent_err(exc=_e)
        try:
            for pl in ((ld.get("antirupture") or {}).get("plan_lisse") or []):
                if pl.get("jour_livraison")==ctx["target"]:
                    lignes=[];tour3=False
                    for lg in pl.get("lignes_carb",[]):
                        lignes.append(f"{lg.get('carburant')} {int(lg.get('volume',0))} L")
                        if lg.get("tour_3"): tour3=True
                    if lignes:
                        ctx["reco"]={"lignes":lignes,"tour_3":tour3,
                                     "infaisable":pl.get("infaisable",False)}
                    break
        except Exception as _e: _log_silent_err(exc=_e)
        return ctx

    def _set_tour(self,n):
        self.cmd_tour=int(n)
        for k,b in getattr(self,"_tour_btns",{}).items():
            if k==int(n):
                b.configure(fg_color=C["gold"],text_color="#1A1A1A",hover_color=C["gold"])
            else:
                b.configure(fg_color=C["card"],text_color=C["t2"],hover_color=C["card_h"])

    def _chosen_date(self):
        """Date de livraison choisie : proposée par défaut (forçage-aware),
        ÉDITABLE. Reproduit le parse de '+ Ajouter une livraison' (cohérence).
        Retourne un date, ou None si invalide (erreur affichée, _save annulé)."""
        try:
            if getattr(self,"_tkcal_ok",False):
                try: return self.e_date.get_date()
                except Exception:
                    messagebox.showerror("Date invalide","Choisis une date valide")
                    return None
            d_str=self.e_date.get().strip()
            p=d_str.split("/")
            assert len(p)==3
            return date(2000+int(p[2]),int(p[1]),int(p[0]))
        except Exception:
            messagebox.showerror("Date invalide","Format attendu : JJ/MM/AA (ex : 04/04/26)")
            return None

    def _save(self):
        try:
            def _g(k):
                try: return float(self.cmd_entries[k].get() or 0)
                except Exception: return 0
            csp,cgo,cgnr=_g("sp"),_g("go"),_g("gnr")
            if (csp+cgo+cgnr)<=0:
                # Rien saisi : ne rien enregistrer, juste fermer (pas de coquille vide)
                self.destroy(); return
            target=self._chosen_date()
            if target is None:
                return  # date invalide : on reste sur la fenêtre (erreur déjà affichée)
            tour_val=getattr(self,"cmd_tour",1)
            try: pv_val=bool(self.cmd_premier.get())
            except Exception: pv_val=False
            add_commande(target,sp=csp,go=cgo,gnr=cgnr,tour=tour_val,premier_voyage=pv_val)
            # Trace + confirmation dans le journal de notifications EXISTANT
            # (réutilise evenements.cfg — pas de nouveau système de To-Do).
            try:
                add_evenement("commande",{
                    "jour": target.isoformat(),
                    "sp": int(round(csp)),"go": int(round(cgo)),"gnr": int(round(cgnr)),
                    "tour": int(tour_val),
                    "premier_voyage": pv_val,
                })
            except Exception as _e: _log_silent_err(exc=_e)
            self.result=True
            try:
                if hasattr(self.hub,"refresh"): self.hub.refresh()
            except Exception as _e: _log_silent_err(exc=_e)
            # FIX 22/05/2026 : si la fenêtre Livraisons & Commandes est ouverte en
            # parallèle, on déclenche aussi son refresh local pour que la commande
            # apparaisse instantanément dans la timeline (sans qu'il faille fermer
            # et rouvrir). Symétrise le comportement des livraisons (qui s'ajoutent
            # via un bouton interne à la fenêtre et profitent du refresh natif).
            # Pattern winfo_children+isinstance (cohérent avec anti-empilement MargeTendueDlg).
            try:
                for w in self.hub.winfo_children():
                    if isinstance(w,LivraisonsHistDlg) and w.winfo_exists():
                        try: w._refresh_table()
                        except Exception as _e: _log_silent_err(exc=_e)
                        break
            except Exception as _e: _log_silent_err(exc=_e)
        except Exception as _e:
            _log_silent_err(exc=_e)
        self.destroy()

    def _cancel(self):
        self.destroy()


class LivraisonsHistDlg(ctk.CTkToplevel):
    """Historique des livraisons : visualiser, ajouter, modifier, supprimer.
    Détecte aussi les écarts entre livraisons.json (logiciel) et Prévision compte (Excel)."""
    def __init__(self,parent,prevision_path=None):
        super().__init__(parent)
        self.title("Livraisons & Commandes \u2014 DISTRICARB HUB")
        self.geometry("1200x720");self.minsize(960,580)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.parent_app=parent
        self.prevision_path=prevision_path
        self.filter_mois=False
        self.changed=False
        # === HEADER ===
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=70);hdr.pack(fill="x",padx=24,pady=(20,10));hdr.pack_propagate(False)
        left=ctk.CTkFrame(hdr,fg_color="transparent");left.pack(side="left",fill="y")
        ctk.CTkLabel(left,text="\U0001f4cb  Livraisons & Commandes",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w")
        self.lbl_total=ctk.CTkLabel(left,text="",font=("Segoe UI",11),text_color=C["t3"])
        self.lbl_total.pack(anchor="w",pady=(2,0))
        right=ctk.CTkFrame(hdr,fg_color="transparent");right.pack(side="right",fill="y")
        # Accès DIRECT à la commande du jour (Brique 1) : fenêtre dédiée
        # CommandeDialog. parent = le HUB (self.parent_app) pour lire last_data
        # (attendu/criticité) et rafraîchir. AUCUN lien avec la livraison reçue.
        ctk.CTkButton(right,text="\U0001f4e6 Commande du jour",width=200,height=34,
                       fg_color=C["teal"],hover_color="#1F8F7F",text_color="#FFF",
                       font=("Segoe UI",11,"bold"),corner_radius=8,command=self._commande_jour).pack(side="left",padx=(0,8))
        # Toggle filtre
        self.btn_filtre=ctk.CTkButton(right,text="\U0001f4c5 Mois en cours",width=140,height=34,
                                       fg_color=C["card"],hover_color=C["card_h"],border_width=1,border_color=C["border2"],
                                       text_color=C["t1"],font=("Segoe UI",11),corner_radius=8,command=self._toggle_filtre)
        self.btn_filtre.pack(side="left",padx=(0,8))
        ctk.CTkButton(right,text="+ Ajouter une livraison",width=170,height=34,
                       fg_color=C["blue"],hover_color="#1E78B8",text_color="#FFF",
                       font=("Segoe UI",11,"bold"),corner_radius=8,command=self._ajouter).pack(side="left",padx=(0,8))
        ctk.CTkButton(right,text="\u2715 Fermer",width=90,height=34,
                       fg_color=C["card"],hover_color=C["card_h"],border_width=1,border_color=C["border2"],
                       text_color=C["t2"],font=("Segoe UI",11),corner_radius=8,command=self._close).pack(side="left")
        # === BODY : tableau scrollable ===
        body=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,
                                     scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        body.pack(fill="both",expand=True,padx=24,pady=(0,12))
        self.body=body
        # Footer info
        ftr=ctk.CTkFrame(self,fg_color="transparent",height=24);ftr.pack(fill="x",padx=24,pady=(0,12));ftr.pack_propagate(False)
        ctk.CTkLabel(ftr,text="\U0001f4a1 Toutes les modifications sont conserv\u00e9es. Les chiffres se mettent \u00e0 jour \u00e0 la prochaine actualisation.",
                     font=("Segoe UI",9),text_color=C["t3"]).pack(side="left")
        self._refresh_table()

    def _toggle_filtre(self):
        self.filter_mois=not self.filter_mois
        if self.filter_mois:
            self.btn_filtre.configure(fg_color=C["amber"],text_color="#141417",text="\U0001f4c5 Tout l'historique")
        else:
            self.btn_filtre.configure(fg_color=C["card"],text_color=C["t1"],text="\U0001f4c5 Mois en cours")
        self._refresh_table()

    def _close(self):
        if self.changed: self.parent_app.refresh()
        self.destroy()

    def _commande_jour(self):
        """Ouvre la fenêtre DÉDIÉE Commande du jour (CommandeDialog). Parent =
        le HUB (self.parent_app), pas cette fenêtre, pour que l'attendu/criticité
        (last_data) et le refresh fonctionnent. Au retour, si une commande a été
        enregistrée, on marque le changement (le hub sera rafraîchi par _close)."""
        try:
            dlg=CommandeDialog(self.parent_app)
            self.wait_window(dlg)
            if getattr(dlg,"result",None):
                self.changed=True
        except Exception as _e: _log_silent_err(exc=_e)

    def _read_prevision_livraisons(self):
        """Lit UNIQUEMENT la livraison prévue pour AUJOURD'HUI dans Prévision compte.
        Utilise get_sheet_for_day(0) pour tenir compte du cycle 14j (semaine 1 / semaine 2).
        Passe par copy_to_temp pour contourner le lock Excel dans la plupart des cas.
        Retours possibles :
          - {date_str(dd/mm/yy): {sp,go,gnr,onglet}} : lecture OK avec prévision non nulle
          - {} : pas de prévision aujourd'hui (cas normal, ventes prévues saisies à 0)
          - {"_error":"notfound"} : chemin configuré mais fichier absent
          - {"_error":"locked"} : copy ou ouverture impossible (lock Excel, sync OneDrive, etc.)
          - {"_error":"read"} : fichier ouvert mais format/contenu inattendu"""
        if not self.prevision_path: return {}
        if not os.path.exists(self.prevision_path): return {"_error":"notfound"}
        # Étage 1 : copie temporaire pour contourner le lock Excel
        tmp=copy_to_temp(self.prevision_path)
        if not tmp: return {"_error":"locked"}
        try:
            import openpyxl
            wb=openpyxl.load_workbook(tmp,data_only=True,read_only=True)
        except PermissionError:
            return {"_error":"locked"}
        except Exception as e:
            print(f"[LivHist] erreur ouverture Prévision : {e}")
            return {"_error":"read"}
        try:
            today=date.today()
            # Utiliser le cycle 14j pour choisir le bon onglet (Mardi vs Mardi2 selon la semaine)
            sn=get_sheet_for_day(0)
            # Fallback sur l'onglet simple si le cycle n'est pas configuré
            if not sn or sn not in wb.sheetnames:
                jours_ws=["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
                sn=jours_ws[today.weekday()]
            if sn not in wb.sheetnames:
                wb.close();return {}
            ws=wb[sn]
            try:
                sp=float(ws["E7"].value or 0)
                go=float(ws["E8"].value or 0)
                gnr=float(ws["E9"].value or 0)
            except:
                wb.close();return {"_error":"read"}
            wb.close()
            if sp+go+gnr==0: return {}
            return {today.strftime("%d/%m/%y"):{"sp":sp,"go":go,"gnr":gnr,"onglet":sn}}
        except Exception as e:
            print(f"[LivHist] erreur lecture Prévision : {e}")
            try: wb.close()
            except Exception as _e: _log_silent_err(exc=_e)
            return {"_error":"read"}

    def _refresh_table(self):
        # Vider
        for w in self.body.winfo_children(): w.destroy()
        # === FUSION 21/05/2026 : LIVRAISONS + COMMANDES dans la même timeline ===
        # Avant : seulement livraisons (LIVRAISON_FILE). Maintenant : on ajoute aussi les
        # commandes (commandes.cfg) pour avoir l'historique complet "commande → livraison".
        # Chaque item est typé "L" (livraison) ou "C" (commande) pour le rendu différencié.
        # Charger livraisons et "à plat" : (date_str_JJ/MM/AA, idx, type, dict)
        livrs=load_json(LIVRAISON_FILE)
        items=[]
        for d_str,raw in livrs.items():
            for idx,liv in enumerate(normalize_livr_day(raw)):
                items.append((d_str,idx,"L",liv))
        # Charger commandes (clés ISO YYYY-MM-DD → convertir en JJ/MM/AA pour homogénéité)
        cmds=load_commandes()
        for k_iso,cmd in cmds.items():
            try:
                # Convertir clé ISO en JJ/MM/AA
                p=k_iso.split("-")
                if len(p)==3:
                    d_str=f"{int(p[2]):02d}/{int(p[1]):02d}/{p[0][-2:]}"
                else:
                    continue
                items.append((d_str,0,"C",cmd))
            except Exception as _e: _log_silent_err(exc=_e)
        # Tri par date décroissante puis idx (livraisons puis commandes le même jour)
        def _parse_date(ds):
            try:
                p=ds.split("/")
                return date(2000+int(p[2]),int(p[1]),int(p[0]))
            except Exception as _e: _log_silent_err(exc=_e); return date(1900,1,1)
        # Tri : date décroissante, puis livraisons (L) avant commandes (C) le même jour
        items.sort(key=lambda x:(-_parse_date(x[0]).toordinal(),0 if x[2]=="L" else 1,x[1]))
        # Filtre mois en cours
        today=date.today()
        if self.filter_mois:
            items=[t for t in items if _parse_date(t[0]).month==today.month and _parse_date(t[0]).year==today.year]
        # Lecture Prévision pour détecter les écarts
        prev_data=self._read_prevision_livraisons()
        # Index "livraison existe pour cette date ?" (pour statut des commandes)
        livraisons_par_date=set()
        for d_str,idx,t,v in items:
            if t=="L" and sum(v.get(k,0) for k in ("sp","go","gnr"))>0:
                livraisons_par_date.add(d_str)
        # OPTION A (validée Bidou 23/05/2026) : masquer les commandes dont la livraison
        # a été reçue. Une seule ligne par événement (la livraison effective domine).
        # Les commandes restent visibles UNIQUEMENT tant qu'elles ne sont pas livrées
        # (statut "Attente" ou "MANQUANTE"). Évite le doublon visuel commande+livraison
        # pour la même date qui créait de la confusion dans la timeline.
        items=[t for t in items if not (t[2]=="C" and t[0] in livraisons_par_date)]
        # Stats : nb de livraisons + commandes + total volumes livrés
        nb_livr=sum(1 for _,_,t,_ in items if t=="L")
        nb_cmd=sum(1 for _,_,t,_ in items if t=="C")
        tot_l=sum((v.get("sp",0)+v.get("go",0)+v.get("gnr",0)) for _,_,t,v in items if t=="L")
        scope="ce mois-ci" if self.filter_mois else "au total"
        self.lbl_total.configure(text=f"{nb_livr} livraison(s) \u2022 {nb_cmd} commande(s) \u2502 {tot_l:,.0f} L livr\u00e9s {scope}".replace(",","."))
        # Bandeau d'alerte si Prévision inaccessible (toujours visible, au-dessus du header)
        err=prev_data.get("_error") if isinstance(prev_data,dict) else None
        if err:
            msg={"locked":"\u26a0 Prévision compte.xlsx est verrouill\u00e9 (ouvert dans Excel ou en cours de synchro OneDrive) \u2014 fermer puis \u21bb Actualiser",
                 "notfound":"\u26a0 Fichier Pr\u00e9vision introuvable \u2014 v\u00e9rifier \u2699 Chemins",
                 "read":"\u26a0 Erreur de lecture du fichier Pr\u00e9vision (voir console)"}.get(err,"\u26a0 Pr\u00e9vision indisponible")
            alert=ctk.CTkFrame(self.body,fg_color="#1A1215",corner_radius=8,border_width=1,border_color="#3A1520")
            alert.pack(fill="x",pady=(0,6),padx=4)
            ctk.CTkLabel(alert,text=msg,font=("Segoe UI",11,"bold"),text_color=C["red"],wraplength=900,justify="left",anchor="w").pack(fill="x",padx=12,pady=8)
        # Headers
        hdr=ctk.CTkFrame(self.body,fg_color=C["panel"],corner_radius=6,height=36)
        hdr.pack(fill="x",pady=(0,4));hdr.pack_propagate(False)
        cols=[("DATE",105),("TRANSPORTEUR",95),("SP (L)",75),("GO (L)",75),("GNR (L)",75),("TOTAL",85),("PRÉVU SP",70),("PRÉVU GO",70),("PRÉVU GNR",70),("ÉCART",80),("ACTIONS / NOTE",300)]
        for txt,w in cols:
            cell=ctk.CTkFrame(hdr,fg_color="transparent",width=w);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            anchor="w" if txt in ("DATE","ACTIONS / NOTE") else "e"
            ctk.CTkLabel(cell,text=txt,font=("Segoe UI",9,"bold"),text_color=C["t3"],anchor=anchor).pack(fill="both",expand=True,padx=8)
        # Lignes
        if not items:
            empty=ctk.CTkFrame(self.body,fg_color=C["card"],corner_radius=8,height=80);empty.pack(fill="x",pady=20);empty.pack_propagate(False)
            ctk.CTkLabel(empty,text="Aucune livraison ni commande enregistrée"+(" pour ce mois" if self.filter_mois else ""),
                          font=("Segoe UI",12),text_color=C["t3"]).pack(expand=True)
            return
        JOURS_FR_COURT=["Lun.","Mar.","Mer.","Jeu.","Ven.","Sam.","Dim."]
        # Pour la détection d'écart Prévision : somme des LIVRAISONS par jour (pas les commandes)
        sum_by_day={}
        for d_str,idx,t,v in items:
            if t!="L": continue
            sum_by_day.setdefault(d_str,{"sp":0,"go":0,"gnr":0})
            sum_by_day[d_str]["sp"]+=v.get("sp",0)
            sum_by_day[d_str]["go"]+=v.get("go",0)
            sum_by_day[d_str]["gnr"]+=v.get("gnr",0)
        # Compter les livraisons par jour (pour afficher "1/2" si plusieurs)
        count_by_day={}
        for d_str,idx,t,v in items:
            if t!="L": continue
            count_by_day[d_str]=count_by_day.get(d_str,0)+1
        for d_str,idx,t,v in items:
            d_obj=_parse_date(d_str)
            is_today=(d_obj==today)
            if t=="C":
                # === LIGNE COMMANDE ===
                self._render_commande_row(d_str,d_obj,is_today,v,livraisons_par_date,JOURS_FR_COURT,today,prev_data)
                continue
            # === LIGNE LIVRAISON (rendu existant) ===
            sp=v.get("sp",0);go=v.get("go",0);gnr=v.get("gnr",0);tot=sp+go+gnr
            nb_day=count_by_day.get(d_str,1)
            is_first_of_day=(idx==0)
            # Données Prévision : on compare à la SOMME du jour (pas à chaque livraison)
            pv=prev_data.get(d_str,{}) if is_first_of_day else {}
            day_sum=sum_by_day[d_str]
            psp=pv.get("sp",0);pgo=pv.get("go",0);pgnr=pv.get("gnr",0)
            ecart_sp=day_sum["sp"]-psp if pv else None
            ecart_go=day_sum["go"]-pgo if pv else None
            ecart_gnr=day_sum["gnr"]-pgnr if pv else None
            has_ecart=pv and (abs(ecart_sp)>1 or abs(ecart_go)>1 or abs(ecart_gnr)>1)
            row_bg="#0F1B12" if is_today else (C["card"] if not has_ecart else "#1A1612")
            row_brd=C["green"] if is_today else (C["amber"] if has_ecart else C["border"])
            row=ctk.CTkFrame(self.body,fg_color=row_bg,corner_radius=8,border_width=1,border_color=row_brd,height=46)
            row.pack(fill="x",pady=2);row.pack_propagate(False)
            # DATE
            jour_lbl=JOURS_FR_COURT[d_obj.weekday()]+f" {d_obj.strftime('%d/%m/%Y')}"
            if nb_day>1:
                jour_lbl+=f"  ({idx+1}/{nb_day})"
            cell=ctk.CTkFrame(row,fg_color="transparent",width=105);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            prefix="\u25cf "if is_today and is_first_of_day else ""
            ctk.CTkLabel(cell,text=prefix+jour_lbl,font=("Segoe UI",10,"bold" if is_today else "normal"),
                          text_color=C["green"] if is_today else C["t1"],anchor="w").pack(fill="both",expand=True,padx=8)
            # TRANSPORTEUR
            transp=v.get("transporteur","\u2014") or "\u2014"
            cell=ctk.CTkFrame(row,fg_color="transparent",width=95);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            ctk.CTkLabel(cell,text=transp[:12],font=("Segoe UI",10),text_color=C["t2"],anchor="w").pack(fill="both",expand=True,padx=6)
            # SP, GO, GNR
            for val,col_w in [(sp,75),(go,75),(gnr,75)]:
                cell=ctk.CTkFrame(row,fg_color="transparent",width=col_w);cell.pack(side="left",fill="y");cell.pack_propagate(False)
                ctk.CTkLabel(cell,text=f"{val:,.0f}".replace(",","."),font=("Segoe UI",11),text_color=C["t1"],anchor="e").pack(fill="both",expand=True,padx=8)
            # TOTAL (de cette livraison ou du jour si fragmenté)
            cell=ctk.CTkFrame(row,fg_color="transparent",width=85);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            ctk.CTkLabel(cell,text=f"{tot:,.0f} L".replace(",","."),font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="e").pack(fill="both",expand=True,padx=8)
            # PRÉV (uniquement sur 1re ligne du jour, sinon vide)
            for val,delta,col_w in [(psp,ecart_sp,70),(pgo,ecart_go,70),(pgnr,ecart_gnr,70)]:
                cell=ctk.CTkFrame(row,fg_color="transparent",width=col_w);cell.pack(side="left",fill="y");cell.pack_propagate(False)
                if pv:
                    txt_col=C["red"] if (delta is not None and abs(delta)>1) else C["t2"]
                    ctk.CTkLabel(cell,text=f"{val:,.0f}".replace(",","."),font=("Segoe UI",11),text_color=txt_col,anchor="e").pack(fill="both",expand=True,padx=8)
                else:
                    ctk.CTkLabel(cell,text="",font=("Segoe UI",11),text_color=C["t3"],anchor="e").pack(fill="both",expand=True,padx=8)
            # ÉCART
            cell=ctk.CTkFrame(row,fg_color="transparent",width=80);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            if pv:
                if has_ecart:
                    deltas=[abs(ecart_sp),abs(ecart_go),abs(ecart_gnr)]
                    max_d=max(deltas)
                    ctk.CTkLabel(cell,text=f"\u26a0 {max_d:,.0f} L".replace(",","."),font=("Segoe UI",11,"bold"),text_color=C["amber"],anchor="e").pack(fill="both",expand=True,padx=8)
                else:
                    ctk.CTkLabel(cell,text="\u2713 OK",font=("Segoe UI",11,"bold"),text_color=C["green"],anchor="e").pack(fill="both",expand=True,padx=8)
            else:
                ctk.CTkLabel(cell,text="",font=("Segoe UI",11),text_color=C["t3"],anchor="e").pack(fill="both",expand=True,padx=8)
            # ACTIONS / NOTE (Modifier + Suppr + 💬 + aperçu note)
            note_v=v.get("note","").strip() if v.get("note") else ""
            cell=ctk.CTkFrame(row,fg_color="transparent",width=300);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            # Bouton Modifier avec texte
            ctk.CTkButton(cell,text="\u270f Modifier",width=78,height=28,fg_color=C["blue"],hover_color="#1E78B8",
                           text_color="#FFF",font=("Segoe UI",10,"bold"),corner_radius=4,
                           command=lambda ds=d_str,ix=idx,vv=v:self._modifier(ds,ix,vv)).pack(side="left",padx=(6,2),pady=8)
            # Bouton Suppr avec texte
            ctk.CTkButton(cell,text="\u2716 Supprimer",width=92,height=28,fg_color="#3F1A1F",hover_color="#5C2A30",
                           text_color=C["red"],border_width=1,border_color="#5C2A30",
                           font=("Segoe UI",10,"bold"),corner_radius=4,
                           command=lambda ds=d_str,ix=idx,vv=v:self._supprimer(ds,ix,vv)).pack(side="left",padx=2,pady=8)
            # Icône 💬 cliquable uniquement si note présente (pas de placeholder grisé)
            if note_v:
                ctk.CTkButton(cell,text="\U0001f4ac",width=28,height=28,fg_color=C["card_h"],
                               hover_color=C["amber"],text_color=C["amber"],font=("Segoe UI Emoji",13),
                               corner_radius=4,command=lambda nt=note_v,ds=d_str:self._show_note(ds,nt)).pack(side="left",padx=(4,4),pady=8)
                # Aperçu de la note (30 premiers caractères, tronqué avec ...)
                preview=note_v[:30]+("\u2026" if len(note_v)>30 else "")
                ctk.CTkLabel(cell,text=preview,font=("Segoe UI",10,"italic"),text_color=C["amber"],anchor="w").pack(side="left",fill="x",expand=True,padx=(2,8),pady=8)

    def _show_note(self,date_str,note_text):
        """Affiche la note d'une livraison dans un petit popup."""
        pop=ctk.CTkToplevel(self)
        pop.title(f"Note — Livraison {date_str}");pop.geometry("420x280");pop.resizable(False,False)
        pop.configure(fg_color=C["bg"]);pop.transient(self);pop.grab_set()
        ctk.CTkLabel(pop,text=f"\U0001f4ac Note du {date_str}",font=("Segoe UI",14,"bold"),text_color=C["t1"]).pack(pady=(20,12))
        txt=ctk.CTkTextbox(pop,fg_color=C["card"],border_color=C["border"],border_width=1,
                            text_color=C["t1"],font=("Segoe UI",11),wrap="word")
        txt.pack(fill="both",expand=True,padx=20,pady=(0,12))
        txt.insert("1.0",note_text);txt.configure(state="disabled")
        ctk.CTkButton(pop,text="Fermer",width=100,height=32,fg_color=C["card"],hover_color=C["card_h"],
                       border_width=1,border_color=C["border2"],text_color=C["t1"],corner_radius=8,
                       command=pop.destroy).pack(pady=(0,16))

    def _render_commande_row(self,d_str,d_obj,is_today,cmd,livraisons_par_date,JOURS_FR_COURT,today,prev_data):
        """Rendu d'UNE ligne commande dans la timeline fusionnée Livraisons & Commandes.
        Réutilise la même grille de colonnes que les livraisons pour cohérence visuelle,
        mais avec différentiation : icône 📦 + fond légèrement bleuté (vs vert pour livraisons).
        Le statut de la commande (Attente / Livrée / Manquante) remplace la colonne ÉCART."""
        # Statut commande
        cmd_sp=int(cmd.get("sp",0));cmd_go=int(cmd.get("go",0));cmd_gnr=int(cmd.get("gnr",0))
        cmd_tot=cmd_sp+cmd_go+cmd_gnr
        is_future=(d_obj>today)
        is_past=(d_obj<today)
        livr_existe=d_str in livraisons_par_date
        if livr_existe:
            statut_txt="\u2713 Livr\u00e9e";statut_col=C["green"];row_brd=C["border"];row_bg=C["card"]
        elif is_future:
            statut_txt="\u23f3 Attente";statut_col=C["teal"];row_brd="#1F3A5C";row_bg="#0F1620"
        elif is_today:
            statut_txt="\u23f3 Aujourd'hui";statut_col=C["amber"];row_brd=C["amber"];row_bg="#1A1612"
        else:
            # Date passée sans livraison correspondante → PAS REÇUE
            statut_txt="\u26a0 Pas re\u00e7ue";statut_col=C["red"];row_brd=C["red"];row_bg="#1A1215"
        row=ctk.CTkFrame(self.body,fg_color=row_bg,corner_radius=8,border_width=1,border_color=row_brd,height=46)
        row.pack(fill="x",pady=2);row.pack_propagate(False)
        # DATE avec icône commande
        jour_lbl=JOURS_FR_COURT[d_obj.weekday()]+f" {d_obj.strftime('%d/%m/%Y')}"
        cell=ctk.CTkFrame(row,fg_color="transparent",width=105);cell.pack(side="left",fill="y");cell.pack_propagate(False)
        ctk.CTkLabel(cell,text=f"\U0001f4e6 {jour_lbl}",font=("Segoe UI",10),
                      text_color=C["t1"],anchor="w").pack(fill="both",expand=True,padx=8)
        # TRANSPORTEUR (= type commande : "1er tour" / "2e tour" / "3e tour" + "· 1er voyage" si premier voyage)
        tour=int(cmd.get("tour",1))
        pv_mark=" \u00b7 1er voyage" if cmd.get("premier_voyage") else ""
        tour_lbl={1:"1er tour",2:"2e tour",3:"3e tour"}.get(tour,f"tour {tour}")
        transp_txt=f"{tour_lbl}{pv_mark}"
        cell=ctk.CTkFrame(row,fg_color="transparent",width=95);cell.pack(side="left",fill="y");cell.pack_propagate(False)
        ctk.CTkLabel(cell,text=transp_txt,font=("Segoe UI",10,"italic"),text_color=C["teal"],anchor="w").pack(fill="both",expand=True,padx=6)
        # SP, GO, GNR (commandés)
        for val,col_w in [(cmd_sp,75),(cmd_go,75),(cmd_gnr,75)]:
            cell=ctk.CTkFrame(row,fg_color="transparent",width=col_w);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            ctk.CTkLabel(cell,text=f"{val:,.0f}".replace(",","."),font=("Segoe UI",11),text_color=C["t1"],anchor="e").pack(fill="both",expand=True,padx=8)
        # TOTAL
        cell=ctk.CTkFrame(row,fg_color="transparent",width=85);cell.pack(side="left",fill="y");cell.pack_propagate(False)
        ctk.CTkLabel(cell,text=f"{cmd_tot:,.0f} L".replace(",","."),font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="e").pack(fill="both",expand=True,padx=8)
        # PRÉV. SP/GO/GNR — référence Pre_vision pour comparer cohérence commande vs prévision
        pv=prev_data.get(d_str,{}) if isinstance(prev_data,dict) else {}
        psp=pv.get("sp",0);pgo=pv.get("go",0);pgnr=pv.get("gnr",0)
        for val,col_w in [(psp,70),(pgo,70),(pgnr,70)]:
            cell=ctk.CTkFrame(row,fg_color="transparent",width=col_w);cell.pack(side="left",fill="y");cell.pack_propagate(False)
            if pv:
                ctk.CTkLabel(cell,text=f"{val:,.0f}".replace(",","."),font=("Segoe UI",11),text_color=C["t2"],anchor="e").pack(fill="both",expand=True,padx=8)
            else:
                ctk.CTkLabel(cell,text="",font=("Segoe UI",11),text_color=C["t3"],anchor="e").pack(fill="both",expand=True,padx=8)
        # STATUT (à la place de ÉCART pour les commandes)
        cell=ctk.CTkFrame(row,fg_color="transparent",width=80);cell.pack(side="left",fill="y");cell.pack_propagate(False)
        ctk.CTkLabel(cell,text=statut_txt,font=("Segoe UI",11,"bold"),text_color=statut_col,anchor="e").pack(fill="both",expand=True,padx=8)
        # ACTIONS : Modifier / Supprimer (réutilise les helpers existants avec wrappers commande)
        cell=ctk.CTkFrame(row,fg_color="transparent",width=300);cell.pack(side="left",fill="y");cell.pack_propagate(False)
        ctk.CTkButton(cell,text="\u270f Modifier",width=78,height=28,fg_color=C["blue"],hover_color="#1E78B8",
                       text_color="#FFF",font=("Segoe UI",10,"bold"),corner_radius=4,
                       command=lambda d=d_obj,c=cmd:self._editer_commande(d,c)).pack(side="left",padx=(6,2),pady=8)
        ctk.CTkButton(cell,text="\u2716 Supprimer",width=92,height=28,fg_color=C["red"],hover_color="#9A2722",
                       text_color="#FFF",font=("Segoe UI",10,"bold"),corner_radius=4,
                       command=lambda d=d_obj:self._supprimer_commande(d)).pack(side="left",padx=(2,2),pady=8)

    def _editer_commande(self,d_obj,cmd):
        """Ouvre la popup CommandeDialog en mode édition (date pré-remplie sur d_obj,
        valeurs préremplies depuis cmd). Si Bidou confirme, add_commande écrase l'entrée."""
        try:
            dlg=CommandeDialog(self.parent_app)
            # Pré-remplir la date et les valeurs après instanciation
            try:
                if hasattr(dlg,"e_date"):
                    if getattr(dlg,"_tkcal_ok",False):
                        dlg.e_date.set_date(d_obj)
                    else:
                        dlg.e_date.delete(0,"end")
                        dlg.e_date.insert(0,d_obj.strftime("%d/%m/%y"))
                for k in ("sp","go","gnr"):
                    if k in dlg.cmd_entries:
                        dlg.cmd_entries[k].delete(0,"end")
                        v=cmd.get(k,0)
                        if v: dlg.cmd_entries[k].insert(0,str(int(v)))
                # Tour pré-rempli (1/2/3)
                if hasattr(dlg,"cmd_tour"):
                    try: dlg.cmd_tour=int(cmd.get("tour",1))
                    except Exception: pass
                if hasattr(dlg,"cmd_premier") and cmd.get("premier_voyage"):
                    try: dlg.cmd_premier.set(True)
                    except Exception: pass
            except Exception as _e: _log_silent_err(exc=_e)
            self.wait_window(dlg)
            if getattr(dlg,"result",False):
                self.changed=True
                self._refresh_table()
        except Exception as _e: _log_silent_err(exc=_e)

    def _supprimer_commande(self,d_obj):
        """Supprime la commande pour d_obj avec confirmation. Trace dans le journal."""
        try:
            if not messagebox.askyesno("Supprimer la commande",
                f"Confirmer la suppression de la commande pour le {d_obj.strftime('%d/%m/%Y')} ?\n\nCette action sera trac\u00e9e dans le journal des \u00e9v\u00e9nements."):
                return
            if delete_commande(d_obj):
                self.changed=True
                self._refresh_table()
                try:
                    if hasattr(self.parent_app,"refresh"): self.parent_app.refresh()
                except Exception as _e: _log_silent_err(exc=_e)
        except Exception as _e: _log_silent_err(exc=_e)

    def _log(self,action,date_livr,**kwargs):
        log=load_json(LIVRAISONS_LOG_FILE)
        if not isinstance(log,list): log=[]
        entry={"action":action,"date_livraison":date_livr,"horodatage":datetime.now().isoformat(timespec="seconds")}
        entry.update(kwargs)
        log.append(entry)
        save_json(LIVRAISONS_LOG_FILE,log)

    def _ajouter(self):
        dlg=LivraisonEditDlg(self,mode="add")
        self.wait_window(dlg)
        if dlg.result:
            d_str=dlg.result["date"]
            new_liv={k:dlg.result[k] for k in ("sp","go","gnr")}
            new_liv["transporteur"]=dlg.result.get("transporteur","")
            new_liv["note"]=dlg.result.get("note","")
            livrs=load_json(LIVRAISON_FILE)
            # Toujours stocker en LISTE (nouveau format multi-livraisons)
            existing=normalize_livr_day(livrs.get(d_str))
            existing.append(new_liv)
            livrs[d_str]=existing
            save_json(LIVRAISON_FILE,livrs)
            self._log("add",d_str,valeurs=new_liv,position=len(existing)-1)
            # Journal d'événements (Sujet E)
            try:
                # d_str est en format dd/mm/yy → convertir en YYYY-MM-DD
                day_iso=datetime.strptime(d_str,"%d/%m/%y").date().isoformat()
                add_evenement("livraison",{
                    "jour":day_iso,
                    "sp":int(new_liv.get("sp",0)),
                    "go":int(new_liv.get("go",0)),
                    "gnr":int(new_liv.get("gnr",0)),
                })
            except Exception as _e: _log_silent_err(exc=_e)
            self.changed=True
            self._refresh_table()

    def _modifier(self,d_str,idx,v):
        dlg=LivraisonEditDlg(self,mode="edit",date_str=d_str,values=v)
        self.wait_window(dlg)
        if dlg.result:
            new_vals={k:dlg.result[k] for k in ("sp","go","gnr")}
            new_vals["transporteur"]=dlg.result.get("transporteur","")
            new_vals["note"]=dlg.result.get("note","")
            new_date=dlg.result["date"]
            old_vals={k:v.get(k,0) for k in ("sp","go","gnr")}
            old_vals["transporteur"]=v.get("transporteur","")
            old_vals["note"]=v.get("note","")
            livrs=load_json(LIVRAISON_FILE)
            # Récupère la liste du jour source
            src_list=normalize_livr_day(livrs.get(d_str))
            if idx>=len(src_list):
                # Sécurité : si idx hors borne, on rafraîchit sans toucher
                self._refresh_table();return
            if new_date==d_str:
                # Même jour : on remplace à l'index
                src_list[idx]=new_vals
                livrs[d_str]=src_list
            else:
                # Changement de date : retire de la source, ajoute en fin de la cible
                src_list.pop(idx)
                if src_list: livrs[d_str]=src_list
                else: livrs.pop(d_str,None)
                tgt_list=normalize_livr_day(livrs.get(new_date))
                tgt_list.append(new_vals)
                livrs[new_date]=tgt_list
            save_json(LIVRAISON_FILE,livrs)
            self._log("modify",new_date,avant=old_vals,apres=new_vals,
                       date_avant=d_str if new_date!=d_str else None,position=idx)
            self.changed=True
            self._refresh_table()

    def _supprimer(self,d_str,idx,v):
        transp=v.get("transporteur","")
        transp_str=f"\nTransporteur : {transp}" if transp else ""
        confirm=ConfirmDlg(self,
            title="Supprimer cette livraison ?",
            message=f"Date : {d_str}{transp_str}\nSP : {v.get('sp',0):,.0f} L\nGO : {v.get('go',0):,.0f} L\nGNR : {v.get('gnr',0):,.0f} L\n\nCette action est tracée dans le journal.\nElle est définitive.".replace(",","."))
        self.wait_window(confirm)
        if confirm.result:
            livrs=load_json(LIVRAISON_FILE)
            src_list=normalize_livr_day(livrs.get(d_str))
            if idx<len(src_list):
                old=src_list.pop(idx)
                if src_list: livrs[d_str]=src_list
                else: livrs.pop(d_str,None)
                save_json(LIVRAISON_FILE,livrs)
                self._log("delete",d_str,valeurs={k:old.get(k,0) for k in ("sp","go","gnr","transporteur","note")},position=idx)
                self.changed=True
            self._refresh_table()

# Petit dialogue de confirmation
class ConfirmDlg(ctk.CTkToplevel):
    def __init__(self,parent,title,message):
        super().__init__(parent)
        self.title(title);self.geometry("440x280");self.resizable(False,False)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set();self.result=False
        ctk.CTkLabel(self,text="\u26a0",font=("Segoe UI",36),text_color=C["amber"]).pack(pady=(20,4))
        ctk.CTkLabel(self,text=title,font=("Segoe UI",14,"bold"),text_color=C["t1"]).pack(pady=(0,12))
        ctk.CTkLabel(self,text=message,font=("Segoe UI",11),text_color=C["t2"],justify="left").pack(padx=30,pady=(0,16))
        btns=ctk.CTkFrame(self,fg_color="transparent");btns.pack(side="bottom",fill="x",padx=30,pady=20)
        ctk.CTkButton(btns,text="Annuler",width=120,height=36,fg_color=C["card"],hover_color=C["card_h"],
                       border_width=1,border_color=C["border2"],text_color=C["t1"],corner_radius=8,
                       command=self.destroy).pack(side="left")
        ctk.CTkButton(btns,text="Confirmer",width=120,height=36,fg_color=C["red"],hover_color="#C41E24",
                       text_color="#FFF",font=("Segoe UI",12,"bold"),corner_radius=8,
                       command=self._yes).pack(side="right")
    def _yes(self): self.result=True;self.destroy()

# Dialogue d'édition / création d'une livraison
class LivraisonEditDlg(ctk.CTkToplevel):
    TRANSPORTEURS=["TMDM","Soatrans","Autre"]
    def __init__(self,parent,mode="add",date_str=None,values=None):
        super().__init__(parent)
        title="Ajouter une livraison" if mode=="add" else "Modifier la livraison"
        self.title(title);self.geometry("520x620");self.resizable(False,False)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set();self.result=None
        ctk.CTkLabel(self,text=("\u002b" if mode=="add" else "\u270f")+"  "+title,
                      font=("Segoe UI",16,"bold"),text_color=C["t1"]).pack(pady=(20,16))
        # === DATE (tkcalendar si dispo, sinon champ texte) ===
        f=ctk.CTkFrame(self,fg_color="transparent");f.pack(fill="x",padx=40,pady=4)
        ctk.CTkLabel(f,text="Date",font=("Segoe UI",12,"bold"),text_color=C["t1"],width=80,anchor="w").pack(side="left")
        # Date par défaut
        if date_str:
            try:
                p=date_str.split("/");default_dt=date(2000+int(p[2]),int(p[1]),int(p[0]))
            except: default_dt=date.today()
        else: default_dt=date.today()
        self._tkcal_ok=False
        try:
            from tkcalendar import DateEntry
            self.e_date=DateEntry(f,width=14,locale="fr_FR",date_pattern="dd/mm/yy",
                                   font=("Segoe UI",13),background=C["card"],foreground=C["t1"],
                                   borderwidth=2,headersbackground=C["card_h"],headersforeground=C["t1"],
                                   selectbackground=C["red"],selectforeground="#FFFFFF",
                                   normalbackground=C["panel"],normalforeground=C["t1"],
                                   weekendbackground=C["panel"],weekendforeground=C["t2"])
            self.e_date.set_date(default_dt)
            self.e_date.pack(side="left",padx=(10,0),ipady=4)
            self._tkcal_ok=True
            ctk.CTkLabel(f,text="\U0001f4c5 calendrier",font=("Segoe UI",9),text_color=C["t3"]).pack(side="left",padx=(8,0))
        except ImportError:
            self.e_date=ctk.CTkEntry(f,height=34,width=180,fg_color=C["card"],border_color=C["border"],
                                      text_color=C["t1"],font=("Segoe UI",13),placeholder_text="JJ/MM/AA")
            self.e_date.pack(side="left",padx=(10,0))
            self.e_date.insert(0,default_dt.strftime("%d/%m/%y"))
            ctk.CTkLabel(f,text="format JJ/MM/AA",font=("Segoe UI",9),text_color=C["amber"]).pack(side="left",padx=(8,0))
        # === TRANSPORTEUR ===
        f=ctk.CTkFrame(self,fg_color="transparent");f.pack(fill="x",padx=40,pady=4)
        ctk.CTkLabel(f,text="Transporteur",font=("Segoe UI",12,"bold"),text_color=C["t1"],width=80,anchor="w").pack(side="left")
        cur_transp=(values or {}).get("transporteur","TMDM")
        # Si valeur existante hors liste = "Autre"
        if cur_transp and cur_transp not in self.TRANSPORTEURS:
            self._cur_other=cur_transp
            cur_select="Autre"
        else:
            self._cur_other=""
            cur_select=cur_transp if cur_transp in self.TRANSPORTEURS else "TMDM"
        self.transp_var=ctk.StringVar(value=cur_select)
        self.transp_menu=ctk.CTkOptionMenu(f,values=self.TRANSPORTEURS,variable=self.transp_var,
                                             width=140,height=34,fg_color=C["card"],button_color=C["card_h"],
                                             button_hover_color=C["border2"],dropdown_fg_color=C["card"],
                                             text_color=C["t1"],font=("Segoe UI",12),
                                             command=self._on_transp_change)
        self.transp_menu.pack(side="left",padx=(10,8))
        # Champ "Autre" — nom libre, visible seulement si Autre sélectionné
        self.e_autre=ctk.CTkEntry(f,height=34,width=180,fg_color=C["card"],border_color=C["border"],
                                    text_color=C["t1"],font=("Segoe UI",12),placeholder_text="Nom du transporteur")
        if cur_select=="Autre":
            self.e_autre.pack(side="left")
            self.e_autre.insert(0,self._cur_other)
        # === SP, GO, GNR ===
        self.entries={}
        for carb,color,key in [("SP",C["blue"],"sp"),("GO",C["amber"],"go"),("GNR",C["teal"],"gnr")]:
            f=ctk.CTkFrame(self,fg_color="transparent");f.pack(fill="x",padx=40,pady=4)
            ctk.CTkLabel(f,text=carb,font=("Segoe UI",13,"bold"),text_color=color,width=80,anchor="w").pack(side="left")
            e=ctk.CTkEntry(f,height=34,width=180,fg_color=C["card"],border_color=C["border"],
                            text_color=C["t1"],font=("Segoe UI",13),placeholder_text="0")
            e.pack(side="left",padx=(10,0));ctk.CTkLabel(f,text="litres",font=("Segoe UI",11),text_color=C["t3"]).pack(side="left",padx=(8,0))
            if values and values.get(key) is not None:
                e.insert(0,str(int(values[key])))
            self.entries[key]=e
        # === NOTE LIBRE ===
        f=ctk.CTkFrame(self,fg_color="transparent");f.pack(fill="x",padx=40,pady=(8,4))
        ctk.CTkLabel(f,text="Note (facultatif)",font=("Segoe UI",12,"bold"),text_color=C["t1"],anchor="w").pack(anchor="w")
        self.txt_note=ctk.CTkTextbox(self,height=80,fg_color=C["card"],border_color=C["border"],border_width=1,
                                       text_color=C["t1"],font=("Segoe UI",11),wrap="word")
        self.txt_note.pack(fill="x",padx=40,pady=(2,4))
        if values and values.get("note"):
            self.txt_note.insert("1.0",str(values["note"]))
        # === BOUTONS ===
        btns=ctk.CTkFrame(self,fg_color="transparent");btns.pack(side="bottom",fill="x",padx=30,pady=20)
        ctk.CTkButton(btns,text="Annuler",width=120,height=36,fg_color=C["card"],hover_color=C["card_h"],
                       border_width=1,border_color=C["border2"],text_color=C["t1"],corner_radius=8,
                       command=self.destroy).pack(side="left")
        self.btn_save=ctk.CTkButton(btns,text="\u2713 "+("Ajouter" if mode=="add" else "Enregistrer"),width=140,height=36,
                       fg_color=C["green"],hover_color="#258A3E",text_color="#FFF",font=("Segoe UI",12,"bold"),
                       corner_radius=8,command=self._save)
        self.btn_save.pack(side="right")
        # === NAVIGATION CLAVIER : Tab + Entrée ===
        # Ordre de tab : date → transporteur → autre (si visible) → SP → GO → GNR → note → save
        # Bind Entrée sur tous les champs : passe au suivant ou valide
        all_widgets=[]
        if self._tkcal_ok: all_widgets.append(self.e_date)
        else: all_widgets.append(self.e_date)
        all_widgets+=[self.entries["sp"],self.entries["go"],self.entries["gnr"]]
        for i,w in enumerate(all_widgets):
            if i<len(all_widgets)-1:
                next_w=all_widgets[i+1]
                w.bind("<Return>",lambda e,nw=next_w:(nw.focus_set(),"break")[1])
            else:
                w.bind("<Return>",lambda e:self._save())
        # Note : Ctrl+Entrée valide (sinon Entrée ajoute une nouvelle ligne dans le textbox)
        self.txt_note.bind("<Control-Return>",lambda e:self._save())
        # Focus initial sur le premier champ vide
        if mode=="add":
            self.entries["sp"].focus_set()
        else:
            self.entries["sp"].focus_set();self.entries["sp"].select_range(0,"end")

    def _on_transp_change(self,choice):
        if choice=="Autre":
            if not self.e_autre.winfo_ismapped():
                self.e_autre.pack(side="left")
        else:
            if self.e_autre.winfo_ismapped():
                self.e_autre.pack_forget()

    def _save(self):
        # Récupérer la date
        if self._tkcal_ok:
            try:
                dt=self.e_date.get_date()
                d_str=dt.strftime("%d/%m/%y")
            except Exception:
                messagebox.showerror("Date invalide","Veuillez choisir une date valide")
                return
        else:
            d_str=self.e_date.get().strip()
            try:
                p=d_str.split("/");assert len(p)==3
                date(2000+int(p[2]),int(p[1]),int(p[0]))
            except Exception:
                messagebox.showerror("Date invalide","Le format attendu est JJ/MM/AA (ex: 15/04/26)")
                return
        # Transporteur
        sel=self.transp_var.get()
        if sel=="Autre":
            transp=self.e_autre.get().strip()
            if not transp:
                messagebox.showerror("Transporteur manquant","Veuillez saisir le nom du transporteur")
                return
        else:
            transp=sel
        # Volumes
        try:
            vals={k:float(e.get() or 0) for k,e in self.entries.items()}
        except ValueError:
            messagebox.showerror("Volume invalide","Les volumes doivent être numériques")
            return
        if sum(vals.values())==0:
            messagebox.showerror("Volume vide","Au moins un carburant doit avoir un volume > 0")
            return
        # Note
        note=self.txt_note.get("1.0","end").strip()
        self.result={"date":d_str,"transporteur":transp,"note":note,**vals}
        self.destroy()

# =============================================================================
class DetailWindow(ctk.CTkToplevel):
    """Fenêtre de détails — raconte une histoire, ne liste pas des chiffres."""
    def __init__(self,parent,key,data):
        super().__init__(parent)
        fd=next((f for f in HUB_FILES if f["key"]==key),{})
        self.title(f"D\u00e9tails \u2014 {fd.get('label',key)}")
        self.geometry("720x780");self.minsize(620,560)
        self.configure(fg_color=C["bg"]);self.resizable(True,True);self.transient(parent)
        self.fd=fd;self.data=data
        # Header avec icône, titre et baseline
        hdr=ctk.CTkFrame(self,fg_color=C["panel"],corner_radius=0,height=80)
        hdr.pack(fill="x");hdr.pack_propagate(False)
        ctk.CTkFrame(hdr,fg_color=fd.get("color",C["t1"]),width=4,corner_radius=0).pack(side="left",fill="y")
        hcontent=ctk.CTkFrame(hdr,fg_color="transparent");hcontent.pack(side="left",fill="both",expand=True,padx=20,pady=14)
        top_row=ctk.CTkFrame(hcontent,fg_color="transparent");top_row.pack(fill="x")
        ctk.CTkLabel(top_row,text=fd.get("icon",""),font=("Segoe UI Emoji",28),text_color=fd.get("color",C["t1"])).pack(side="left",padx=(0,12))
        title_col=ctk.CTkFrame(top_row,fg_color="transparent");title_col.pack(side="left",fill="x",expand=True)
        ctk.CTkLabel(title_col,text=fd.get("label",""),font=("Segoe UI",22,"bold"),text_color=C["t1"],anchor="w").pack(fill="x")
        ctk.CTkLabel(title_col,text=fd.get("sub",""),font=("Segoe UI",11),text_color=C["t3"],anchor="w").pack(fill="x")
        # Scroll content
        scroll=ctk.CTkScrollableFrame(self,fg_color="transparent",scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        scroll.pack(fill="both",expand=True,padx=16,pady=12)
        # Préparer les données communes
        hist=data.get("hist",[])
        self.complete=[h for h in hist if not h.get("en_cours")]
        # Jour en cours = dernier jour récent (< 3j) avec caisses partielles, ignore les orphelins
        self.today_data=get_current_partial(hist)
        self.j7=self._find_j7(hist)
        # Aiguillage selon hub
        if key=="gest_piste": self._detail_gestpiste(scroll,data)
        elif key=="cartes": self._detail_cartes(scroll,data)
        elif key=="prevision": self._detail_prevision(scroll,data)
        elif key=="objectif": self._detail_objectif(scroll,data)
        elif key=="litrage": self._detail_litrage(scroll,data)

    def _find_j7(self,hist):
        """Trouve le jour J-7 (même jour la semaine dernière)."""
        target=date.today()-timedelta(days=7)
        target_dd=target.strftime("%d/%m")
        for h in hist:
            if target_dd in h.get("label","") and not h.get("en_cours"):
                return h
        return None

    # ---------- Helpers de présentation ----------
    def _hero(self,parent,title,subtitle,color=None):
        """Bandeau en tête avec titre fort + sous-titre explicatif."""
        f=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=12,border_width=1,border_color=color or C["border"])
        f.pack(fill="x",pady=(4,12))
        inner=ctk.CTkFrame(f,fg_color="transparent");inner.pack(fill="x",padx=16,pady=12)
        ctk.CTkLabel(inner,text=title,font=("Segoe UI",13,"bold"),text_color=color or C["t1"],anchor="w").pack(fill="x")
        ctk.CTkLabel(inner,text=subtitle,font=("Segoe UI",10),text_color=C["t3"],anchor="w").pack(fill="x",pady=(2,0))
        return f
    def _hero_kpi(self,parent,kpis):
        """Bandeau avec plusieurs KPI côte à côte (label + value + delta)."""
        f=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=12,border_width=1,border_color=C["border"])
        f.pack(fill="x",pady=(4,12))
        row=ctk.CTkFrame(f,fg_color="transparent");row.pack(fill="x",padx=16,pady=14)
        for i,(label,value,delta,color) in enumerate(kpis):
            col=ctk.CTkFrame(row,fg_color="transparent");col.pack(side="left",fill="x",expand=True,padx=(0,12) if i<len(kpis)-1 else 0)
            ctk.CTkLabel(col,text=label,font=("Segoe UI",10,"bold"),text_color=C["t3"],anchor="w").pack(fill="x")
            ctk.CTkLabel(col,text=value,font=("Segoe UI",18,"bold"),text_color=color or C["t1"],anchor="w").pack(fill="x",pady=(2,0))
            if delta:
                ctk.CTkLabel(col,text=delta,font=("Segoe UI",10),text_color=C["t3"],anchor="w").pack(fill="x",pady=(1,0))
        return f
    def _section(self,parent,title,subtitle=None):
        f=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=10,border_width=1,border_color=C["border"])
        f.pack(fill="x",pady=(8,4))
        head=ctk.CTkFrame(f,fg_color="transparent");head.pack(fill="x",padx=14,pady=(10,6))
        ctk.CTkLabel(head,text=title,font=("Segoe UI",12,"bold"),text_color=C["gold"],anchor="w").pack(fill="x")
        if subtitle:
            ctk.CTkLabel(head,text=subtitle,font=("Segoe UI",9),text_color=C["t3"],anchor="w").pack(fill="x",pady=(1,0))
        return f
    def _row(self,parent,label,value,color=None,sub=None):
        r=ctk.CTkFrame(parent,fg_color="transparent");r.pack(fill="x",padx=14,pady=2)
        left=ctk.CTkFrame(r,fg_color="transparent");left.pack(side="left",fill="x",expand=True)
        ctk.CTkLabel(left,text=label,font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(fill="x")
        if sub:
            ctk.CTkLabel(left,text=sub,font=("Segoe UI",9),text_color=C["t3"],anchor="w").pack(fill="x")
        ctk.CTkLabel(r,text=value,font=("Segoe UI",12,"bold"),text_color=color or C["t1"],anchor="e").pack(side="right",padx=(8,0))
    def _alert_box(self,parent,text,color=None):
        bg="#1A1215" if color==C["red"] else "#1A1812"
        bd="#3A1520" if color==C["red"] else "#3A2A15"
        f=ctk.CTkFrame(parent,fg_color=bg,corner_radius=8,border_width=1,border_color=bd)
        f.pack(fill="x",pady=4,padx=4)
        ctk.CTkLabel(f,text=text,font=("Segoe UI",11),text_color=color or C["amber"],wraplength=620,justify="left",anchor="w").pack(fill="x",padx=12,pady=8)
    def _total_banner(self,parent,label,value,color,sub=None):
        """Bandeau total autonome : fond contrasté, bordure, montant en gros."""
        bg="#2A1519" if color==C["red"] else ("#2A2315" if color==C["amber"] else "#15251A")
        bd="#5A1F2E" if color==C["red"] else ("#5A4820" if color==C["amber"] else "#1F5A2E")
        box=ctk.CTkFrame(parent,fg_color=bg,corner_radius=8,border_width=2,border_color=bd)
        box.pack(fill="x",padx=10,pady=(8,6))
        inner=ctk.CTkFrame(box,fg_color="transparent");inner.pack(fill="x",padx=14,pady=10)
        left=ctk.CTkFrame(inner,fg_color="transparent");left.pack(side="left",fill="x",expand=True)
        ctk.CTkLabel(left,text=label,font=("Segoe UI",11,"bold"),text_color=C["t2"],anchor="w").pack(fill="x")
        if sub:
            ctk.CTkLabel(left,text=sub,font=("Segoe UI",9),text_color=C["t3"],anchor="w").pack(fill="x",pady=(2,0))
        ctk.CTkLabel(inner,text=value,font=("Segoe UI",22,"bold"),text_color=color,anchor="e").pack(side="right",padx=(8,0))
        # Divider sous le bandeau pour séparer du détail
        ctk.CTkFrame(parent,fg_color=C["border"],height=1).pack(fill="x",padx=20,pady=(2,6))
    def _collapsible(self,parent,title,count,total_str=None,color=None):
        """Section pliable fermée par défaut, retourne le conteneur à remplir."""
        wrap=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=10,border_width=1,border_color=C["border"])
        wrap.pack(fill="x",pady=(8,4))
        header=ctk.CTkFrame(wrap,fg_color="transparent",cursor="hand2");header.pack(fill="x",padx=14,pady=(10,8))
        state={"open":False}
        chev=ctk.CTkLabel(header,text="\u25b8",font=("Segoe UI",12,"bold"),text_color=color or C["gold"])
        chev.pack(side="left",padx=(0,8))
        txt=f"{title} \u2014 {count}"
        if total_str: txt+=f" \u2014 {total_str}"
        lbl=ctk.CTkLabel(header,text=txt,font=("Segoe UI",12,"bold"),text_color=color or C["gold"],anchor="w")
        lbl.pack(side="left",fill="x",expand=True)
        content=ctk.CTkFrame(wrap,fg_color="transparent")
        def toggle(_=None):
            state["open"]=not state["open"]
            if state["open"]:
                content.pack(fill="x",padx=4,pady=(0,8))
                chev.configure(text="\u25be")
            else:
                content.pack_forget()
                chev.configure(text="\u25b8")
        for w in (header,chev,lbl): w.bind("<Button-1>",toggle)
        return content
    def _delta_text(self,today_val,ref_val):
        """Génère un texte de comparaison vs J-7 avec flèche et couleur."""
        if not ref_val or ref_val==0: return ("",C["t3"])
        diff_pct=(today_val-ref_val)/ref_val*100
        if diff_pct>2: return (f"\u25b2 {diff_pct:+.0f}% vs J-7",C["green"])
        elif diff_pct<-2: return (f"\u25bc {diff_pct:+.0f}% vs J-7",C["red"])
        else: return (f"\u25cf stable vs J-7",C["t3"])

    # ---------- DÉTAIL GEST PISTE ----------
    def _detail_gestpiste(self,parent,d):
        td=self.today_data;j7=self.j7;complete=self.complete
        # Récupérer tous les écarts anormaux sur l'historique disponible
        # Tri caisses C3 -> C2 -> C1 dans chaque jour (la C3 étant la plus récente du jour)
        ecarts=[]
        # Construire la liste des jours à analyser : tous les jours complets +
        # le jour en cours si dispo (ses caisses déjà saisies seulement, le filtre
        # "ecart_net > seuil" exclut naturellement les caisses encore vides).
        # Permet de voir les écarts dès qu'une caisse est saisie sans attendre fin
        # de journée (la C3 nuit complétée le lendemain matin).
        jours_a_analyser=list(complete)
        if td and td.get("caisses"):
            jours_a_analyser.append(td)
        for h in jours_a_analyser:
            lbl=h.get("label","")
            caisses_dict=h.get("caisses",{}) or {}
            # Trier par numéro décroissant (C3, C2, C1) en gérant les clés str/int
            sorted_caisses=sorted(caisses_dict.items(),key=lambda kv:-int(str(kv[0]).strip() or 0) if str(kv[0]).strip().isdigit() else 0)
            for cnum,c in sorted_caisses:
                net=sf(c.get("ecart_net",0))
                if abs(net)>10:
                    ep=sf(c.get("ecart_piste",0));eb=sf(c.get("ecart_bout",0))
                    ecarts.append({"key":f"{lbl}_C{cnum}","jour":lbl,"caisse":cnum,"piste":ep,"bout":eb,"net":net})
        resolus=load_json(ECARTS_FILE) or {}
        # Détecter les écarts dont la valeur a changé depuis le marquage "résolu" :
        # un écart résolu reste résolu seulement si la valeur n'a pas bougé (à 1€ près
        # ET sans changement de signe). Sinon on l'extrait pour l'afficher en alerte
        # avec la mention "Modifié depuis marquage" — l'utilisateur sait que c'est
        # l'ancien écart qui a évolué, pas un nouveau.
        # Format stockage :
        #   - ancien : {"key": "2026-05-04T..." (string ISO date)}
        #   - nouveau : {"key": {"date": "2026-05-04T...", "valeur": -167.57}}
        # Migration douce : entrée ancien format = pas de valeur de référence,
        # on garde résolu (pas de faux positifs sur les marquages d'avant ce fix).
        non_resolus=[]
        modifies=[]  # écarts résolus dont la valeur a changé : à réafficher avec badge
        for e in ecarts:
            entry=resolus.get(e["key"])
            if not entry:
                non_resolus.append(e);continue
            # Entry existe : vérifier si format nouveau
            if isinstance(entry,dict) and "valeur" in entry:
                ancienne_val=sf(entry.get("valeur",0))
                # Considérer modifié si écart > 1€ OU changement de signe
                if abs(e["net"]-ancienne_val)>1 or (ancienne_val*e["net"]<0):
                    e2=dict(e)
                    e2["modifie"]=True
                    e2["valeur_marquage"]=ancienne_val
                    e2["date_marquage"]=entry.get("date","")
                    modifies.append(e2)
                # sinon : valeur stable, reste résolu (pas dans non_resolus)
            # else : ancien format, on garde résolu (pas de référence pour comparer)
        non_resolus=non_resolus+modifies  # les modifiés s'ajoutent en tête de liste
        nb_jours=min(30,len(complete))
        net_30=sum(e["net"] for e in ecarts[-nb_jours*3:] if not resolus.get(e["key"]) or any(m["key"]==e["key"] for m in modifies))
        proj_an=(net_30/nb_jours)*365 if nb_jours>0 else 0

        # === HERO POSITIF : performance du jour en cours ===
        if td:
            piste_t=sf(td.get("piste",0));bout_t=sf(td.get("bout",0));litrage_t=sf(td.get("litrage",0))
            cb_t=sf(td.get("cb",0));cp_t=sf(td.get("cp",0));esp_t=sf(td.get("esp",0))
            nb=int(td.get("nb_caisses",0))
            # Comparaison J-7 piste cumulé sur les caisses fermées
            j7_piste=0
            if j7 and nb>0:
                for i in range(1,nb+1):
                    j7_piste+=sf(j7.get("caisses",{}).get(str(i),{}).get("piste_eur",0))
            dlt=self._delta_text(piste_t,j7_piste)
            self._hero_kpi(parent,[
                ("PISTE JOUR",feur(piste_t,d=0),dlt[0] or f"Caisses {nb}/3",dlt[1] if dlt[1]!=C["t3"] else C["gold"]),
                ("BOUTIQUE",feur(bout_t,d=0),f"{fnum(litrage_t,'L')} servis",C["gold"]),
                ("CB + CP",feur(cb_t+cp_t,d=0),f"Esp\u00e8ces {feur(esp_t,d=0)}",C["t1"]),
            ])
        # === Performance boutique cumulée ===
        bout_vals=[h.get("bout",0) for h in complete if h.get("bout",0)>0]
        if bout_vals:
            avg=sum(bout_vals)/len(bout_vals)
            best_b=max(complete,key=lambda x:x.get("bout",0))
            s=self._section(parent,"\U0001f4ca Performance boutique",f"Sur {len(bout_vals)} jours")
            self._row(s,"Moyenne /jour",feur(avg,d=0),C["t1"])
            self._row(s,"\u2b06 Meilleur jour",feur(best_b.get('bout'),d=0),C["green"],sub=best_b.get("label",""))
            if td and td.get("bout",0)>0:
                self._row(s,"Aujourd'hui",feur(td.get('bout'),d=0),C["gold"],sub=f"Caisse(s) {td.get('nb_caisses',0)}/3")
        # === Performance piste cumulée ===
        piste_vals=[h.get("piste",0) for h in complete if h.get("piste",0)>0]
        if piste_vals:
            avg_p=sum(piste_vals)/len(piste_vals)
            best_p=max(complete,key=lambda x:x.get("piste",0))
            sp=self._section(parent,"\u26fd Performance piste",f"Sur {len(piste_vals)} jours")
            self._row(sp,"Moyenne /jour",feur(avg_p,d=0),C["t1"])
            self._row(sp,"\u2b06 Meilleur jour",feur(best_p.get('piste'),d=0),C["green"],sub=best_p.get("label",""))
        # === Suivi écarts caisse — bloc visuel marqué (en bas, après le positif) ===
        if non_resolus or abs(proj_an)>50:
            if proj_an<-200:
                col=C["red"];msg="Fuite caisse \u00e0 traiter"
            elif proj_an<0:
                col=C["amber"];msg="\u00c0 surveiller"
            elif proj_an>200:
                col=C["amber"];msg="\u00c9carts POSITIFS anormaux \u2014 v\u00e9rifier saisies pompistes (clients revenus, oublis livraison, sur-d\u00e9clarations)"
            else:
                col=C["t3"];msg="Dans la norme"
            proj_txt=f"{proj_an:+,.0f} \u20ac/an".replace(",",".")
            box=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=12,border_width=2,border_color=col)
            box.pack(fill="x",pady=(12,4))
            inner=ctk.CTkFrame(box,fg_color="transparent");inner.pack(fill="x",padx=20,pady=16)
            ctk.CTkLabel(inner,text="\u00c9CARTS DE CAISSE \u2014 PROJECTION ANNUELLE",font=("Segoe UI",10,"bold"),text_color=C["t3"],anchor="w").pack(fill="x")
            ctk.CTkLabel(inner,text=proj_txt,font=("Segoe UI",26,"bold"),text_color=col,anchor="w").pack(fill="x",pady=(4,4))
            ctk.CTkLabel(inner,text=msg,font=("Segoe UI",11),text_color=col,anchor="w",wraplength=620,justify="left").pack(fill="x")
            ctk.CTkLabel(inner,text=f"Bas\u00e9 sur {len(non_resolus)} \u00e9cart(s) non r\u00e9solu(s) sur {nb_jours} derniers jours",font=("Segoe UI",9),text_color=C["t3"],anchor="w").pack(fill="x",pady=(6,0))
        # ÉCARTS — Liste chronologique avec toggle résolu
        if ecarts:
            s2=self._section(parent,f"\u26a0 \u00c9carts de caisse \u2014 historique",f"Cliquer sur une ligne pour marquer comme r\u00e9solu")
            # Tri pour affichage : jours récents en haut + dans chaque jour C3 -> C2 -> C1
            # On parse la date du label pour trier proprement
            import re as _re_sort
            def _sort_ecart(e):
                m=_re_sort.search(r'(\d{1,2})/(\d{1,2})/(\d{2})',str(e['jour']))
                if m:
                    try:
                        day_ord=date(2000+int(m.group(3)),int(m.group(2)),int(m.group(1))).toordinal()
                    except: day_ord=0
                else: day_ord=0
                try: caisse_num=int(str(e['caisse']).strip())
                except: caisse_num=0
                # Plus récent d'abord (-day_ord), puis C3 -> C2 -> C1 (-caisse_num)
                return (-day_ord,-caisse_num)
            ecarts_affiches=sorted(ecarts,key=_sort_ecart)[:20]
            for e in ecarts_affiches:
                # 3 états possibles :
                #   - non résolu (rouge/ambre)
                #   - résolu et stable (vert sombre, grisé)
                #   - résolu mais valeur a changé depuis (badge "Modifié", couleur active)
                entry=resolus.get(e["key"])
                if not entry:
                    is_resolu=False;is_modifie=False
                else:
                    if isinstance(entry,dict) and "valeur" in entry:
                        ancienne=sf(entry.get("valeur",0))
                        if abs(e["net"]-ancienne)>1 or (ancienne*e["net"]<0):
                            is_resolu=False;is_modifie=True
                            valeur_marquage=ancienne
                            date_marquage=entry.get("date","")
                        else:
                            is_resolu=True;is_modifie=False
                    else:
                        is_resolu=True;is_modifie=False  # ancien format
                # Couleur du fond : modifié = teinte spéciale (orange foncé), sinon comme avant
                if is_modifie: bg="#2A1F0D"  # ambre sombre pour signaler "à revoir"
                elif is_resolu: bg="#0F1B12"
                else: bg="#1A0F12" if e["net"]<0 else "#1A1912"
                r=ctk.CTkFrame(s2,fg_color=bg,corner_radius=6)
                r.pack(fill="x",padx=10,pady=2)
                row_inner=ctk.CTkFrame(r,fg_color="transparent");row_inner.pack(fill="x",padx=10,pady=6)
                # Colonne gauche : date + caisse
                left=ctk.CTkFrame(row_inner,fg_color="transparent");left.pack(side="left",fill="x",expand=True)
                txt_color=C["t3"] if is_resolu else C["t1"]
                prefix="\u2713 " if is_resolu else ("\u26a0 " if is_modifie else "")
                ctk.CTkLabel(left,text=f"{prefix}{e['jour']} \u2014 Caisse {e['caisse']}",font=("Segoe UI",11,"bold"),text_color=txt_color,anchor="w").pack(fill="x")
                # Sous-ligne : détail piste/boutique + mention "Modifié" si applicable
                if is_modifie:
                    # Format date marquage en jj/mm
                    date_h=""
                    try:
                        dt=datetime.fromisoformat(date_marquage)
                        date_h=dt.strftime("%d/%m")
                    except Exception as _e: _log_silent_err(exc=_e)
                    sub=f"Piste {e['piste']:+.2f}\u20ac \u2502 Bout. {e['bout']:+.2f}\u20ac \u2502 Modifi\u00e9 depuis marquage{(' du '+date_h) if date_h else ''} (ancien : {valeur_marquage:+.2f}\u20ac)".replace(".",",")
                    ctk.CTkLabel(left,text=sub,font=("Segoe UI",9),text_color=C["amber"],anchor="w").pack(fill="x")
                else:
                    ctk.CTkLabel(left,text=f"Piste {e['piste']:+.2f}\u20ac \u2502 Boutique {e['bout']:+.2f}\u20ac".replace(".",","),font=("Segoe UI",9),text_color=C["t3"],anchor="w").pack(fill="x")
                # Colonne droite : montant net
                net_col=C["t3"] if is_resolu else (C["red"] if e["net"]<0 else C["amber"])
                net_txt=f"{e['net']:+.2f} \u20ac".replace(".",",")
                ctk.CTkLabel(row_inner,text=net_txt,font=("Segoe UI",14,"bold"),text_color=net_col,anchor="e").pack(side="right",padx=(10,0))
                # Clic pour toggle résolu : on stocke maintenant la valeur en plus de la date,
                # pour pouvoir détecter une évolution à la prochaine ouverture.
                def toggle(ev,key=e["key"],valeur=e["net"],win=self):
                    res=load_json(ECARTS_FILE) or {}
                    if res.get(key): del res[key]
                    else: res[key]={"date":datetime.now().isoformat(),"valeur":round(valeur,2)}
                    save_json(ECARTS_FILE,res)
                    win.destroy()
                    win.master.show_details("gest_piste")
                for w in (r,row_inner,left):
                    w.bind("<Button-1>",toggle)
                for w in left.winfo_children():
                    w.bind("<Button-1>",toggle)

    # ---------- DÉTAIL CARTES ----------
    def _detail_cartes(self,parent,d):
        td=self.today_data;j7=self.j7;complete=self.complete
        # HERO : KPIs jour en cours
        if td:
            nb=td.get("nb_caisses",0)
            cb_t=td.get("cb",0);cp_t=td.get("cp",0)
            # Comparaison J-7 caisse par caisse
            j7_cb_p=0;j7_cp_p=0
            if j7 and nb>0:
                for i in range(1,nb+1):
                    j7_cb_p+=sf(j7.get("caisses",{}).get(str(i),{}).get("cb",0))
                    j7_cp_p+=sf(j7.get("caisses",{}).get(str(i),{}).get("cp",0))
            cb_delta=self._delta_text(cb_t,j7_cb_p)
            cp_delta=self._delta_text(cp_t,j7_cp_p)
            self._hero_kpi(parent,[
                ("CB AUJOURD'HUI",feur(cb_t,d=0),cb_delta[0],cb_delta[1] if cb_delta[1]!=C["t3"] else C["t1"]),
                ("CP AUJOURD'HUI",feur(cp_t,d=0),cp_delta[0],cp_delta[1] if cp_delta[1]!=C["t3"] else C["t1"]),
                ("CAISSES",f"{nb}/3","Jour en cours",C["amber"]),
            ])
        # SECTION : Télécollectes du mois en cours
        if complete:
            import re as _re3
            _today=date.today();_m=_today.month;_y=_today.year%100
            _mnoms=["janvier","f\u00e9vrier","mars","avril","mai","juin","juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
            def _im2(lbl):
                mm=_re3.search(r'(\d{1,2})/(\d{1,2})/(\d{2})',str(lbl or ""))
                return mm and int(mm.group(2))==_m and int(mm.group(3))==_y
            _mois=[h for h in complete if _im2(h.get("label",""))]
            cb_m=sum(h.get("cb",0) for h in _mois)
            cp_m=sum(h.get("cp",0) for h in _mois)
            tot_m=cb_m+cp_m
            s=self._section(parent,"\U0001f4b3 T\u00e9l\u00e9collectes "+_mnoms[_m-1]+f" {_today.year}",f"Sur {len(_mois)} jour(s) du mois")
            self._row(s,"CB total",feur(cb_m,d=0),C["t1"])
            self._row(s,"CP total",feur(cp_m,d=0),C["t1"])
            self._row(s,"\u2014 Total cartes \u2014",feur(tot_m,d=0),C["green"])
            if tot_m>0:
                self._row(s,"Part CB",f"{cb_m/tot_m*100:.0f}%",C["teal"])
            # SECTION : Télécollectes cumulées (historique complet disponible)
            cb_c=sum(h.get("cb",0) for h in complete)
            cp_c=sum(h.get("cp",0) for h in complete)
            tot_c=cb_c+cp_c
            sc=self._section(parent,"\U0001f4b3 T\u00e9l\u00e9collectes cumul\u00e9es",f"Sur {len(complete)} jours complets")
            self._row(sc,"CB total",feur(cb_c,d=0),C["t1"])
            self._row(sc,"CP total",feur(cp_c,d=0),C["t1"])
            self._row(sc,"\u2014 Total cartes \u2014",feur(tot_c,d=0),C["green"])
            if tot_c>0:
                self._row(sc,"Part CB",f"{cb_c/tot_c*100:.0f}%",C["teal"])
        # SECTION : Alertes opérations > 100€ non pointées
        ca=d.get("ca",{})
        critical=ca.get("critical",[])
        if critical:
            total_crit=sum(c["montant"] for c in critical)
            sc=self._section(parent,f"\U0001f6a8 {len(critical)} op\u00e9rations > 100\u20ac en retard",f"Total : {feur(total_crit,d=0)} \u2014 Non point\u00e9es depuis + 3 jours")
            for op in sorted(critical,key=lambda x:x.get("age") or 0,reverse=True):
                age=op.get("age")
                age_txt=f"{age} jours de retard" if age else "date inconnue"
                date_txt=op["date"].strftime("%d/%m/%Y") if op.get("date") else "\u2014"
                col=C["red"] if age and age>7 else C["amber"]
                self._row(sc,date_txt,feur(op["montant"],d=0),col,sub=age_txt)
        # SECTION : Télécollectes en attente de pointage (collapsible avec détail)
        rose=ca.get("rose",0)
        if rose>0:
            # Collecter toutes les TC non pointées avec date + montant
            pending_tcs=[]
            for day in ca.get("days",[]):
                for t in day.get("tcs",[]):
                    if not t.get("jaune"):
                        pending_tcs.append({"date":day["date"],"montant":t["montant"]})
            total_pending=sum(p["montant"] for p in pending_tcs)
            coll=self._collapsible(parent,"\U0001f4cd T\u00e9l\u00e9collectes en attente de pointage banque",len(pending_tcs),feur(total_pending,d=0),C["amber"])
            # Tri par date décroissante (plus récent en premier)
            for p in sorted(pending_tcs,key=lambda x:x["date"],reverse=True):
                age=(date.today()-p["date"]).days
                sub=f"{age}j" if age>0 else "aujourd'hui"
                col=C["red"] if age>3 else C["amber"]
                self._row(coll,p["date"].strftime("%d/%m/%Y"),feur(p["montant"],d=0),col,sub=sub)
        # SECTION : Détail des derniers jours
        s3=self._section(parent,"\U0001f4c5 Derniers jours \u2014 d\u00e9tail",None)
        # En-têtes
        h_row=ctk.CTkFrame(s3,fg_color=C["panel"]);h_row.pack(fill="x",padx=10,pady=(4,2))
        for i,t in enumerate(["Jour","CB","CP","Total cartes"]):
            w=140 if i==0 else 120
            ctk.CTkLabel(h_row,text=t,font=("Segoe UI",9,"bold"),text_color=C["t3"],width=w,anchor="e" if i>0 else "w").pack(side="left",padx=2,pady=4)
        for day in reversed(complete[-7:]):
            r=ctk.CTkFrame(s3,fg_color="transparent");r.pack(fill="x",padx=10,pady=1)
            cb_v=sf(day.get("cb",0));cp_v=sf(day.get("cp",0))
            for i,val in enumerate([day.get("label",""),feur(cb_v,d=0),feur(cp_v,d=0),feur(cb_v+cp_v,d=0)]):
                w=140 if i==0 else 120
                col=C["gold"] if i==0 else C["green"] if i==3 else C["t1"]
                ctk.CTkLabel(r,text=val,font=("Segoe UI",10),text_color=col,width=w,anchor="e" if i>0 else "w").pack(side="left",padx=2,pady=2)

    # ---------- DÉTAIL PRÉVISION ----------
    def _detail_prevision(self,parent,d):
        pv=d.get("pv",{});auto=d.get("auto",{});partial=d.get("partial")
        # HERO : Stocks temps réel
        sp_real=pv.get("sp",0);go_real=pv.get("go",0);gnr_real=pv.get("gnr",0)
        # Hero : tous les carburants avec autonomie, regroupés par niveau de criticité
        carbs=[("SP","sp",auto.get("sp",0)),("GO","go",auto.get("go",0)),("GNR","gnr",auto.get("gnr",0))]
        min_j=min(c[2] for c in carbs)
        # Tous les carburants à <= min+0.1j sont co-critiques
        critiques=[c for c in carbs if c[2]<=min_j+0.1]
        def _col(j): return C["green"] if j>2 else C["amber"] if j>1 else C["red"]
        if len(critiques)==1 and min_j<2:
            c=critiques[0]
            self._hero_kpi(parent,[(f"\u26a0 {c[0]} CRITIQUE",fmt_autonomie(c[2]),"Autonomie minimum",_col(c[2]))])
        elif len(critiques)>1 and min_j<2:
            label="\u26a0 "+" + ".join(c[0] for c in critiques)+" CRITIQUES"
            self._hero_kpi(parent,[(label,fmt_autonomie(min_j),f"{len(critiques)} carburants \u00e0 surveiller",_col(min_j))])
        else:
            # Tout est OK : afficher les 3 ensemble
            self._hero_kpi(parent,[(c[0],fmt_autonomie(c[2]),f"{fnum(pv.get(c[1],0),'L')}",_col(c[2])) for c in carbs])
        # ===== ANTI-RUPTURE : sections selon sévérité =====
        ar=d.get("antirupture",{}) or {}
        sev=ar.get("severite_max","info")
        # Bandeau d'alerte rouge si critique (résumé condensé, le détail est dans le popup)
        if sev=="critique":
            ruptures=ar.get("ruptures_dans_trou",[])
            incoh=ar.get("incoherences_jour_non_livrable",[])
            msg_parts=[]
            if ruptures:
                carbs_rupt=", ".join(sorted(set(r["carburant"] for r in ruptures)))
                msg_parts.append(f"Risque de rupture {carbs_rupt}")
            if incoh:
                msg_parts.append(f"{len(incoh)} commande(s) sur jour non-livrable")
            self._alert_box(parent,f"\U0001f6a8 ALERTE ANTI-RUPTURE \u2014 {' \u2502 '.join(msg_parts)}. Voir popup au prochain refresh.",C["red"])
        # Section : ruptures détaillées dans un trou (manques arrondis tranche 500 L sup)
        ruptures=ar.get("ruptures_dans_trou",[])
        ack_status=ar.get("ack_status",{}) or {}
        if ruptures:
            sec=self._section(parent,f"\U0001f4a5 Risque de manque pendant un weekend / pont",f"{len(ruptures)} carburant(s) en risque")
            for r in ruptures:
                pont_id=f"pont_{r['trou_start'].strftime('%d%m%Y')}"
                st=ack_status.get(pont_id,{})
                acquitte=st.get("acquitte",False)
                manque_arrondi=max(2000,int(((r["manque"]+999)//1000)*1000))
                # Couleur de la box selon acquittement
                box_bg=C["card"] if acquitte else C["alert_bg"]
                box_border=C["green"] if acquitte else C["alert_border"]
                line_box=ctk.CTkFrame(sec,fg_color=box_bg,corner_radius=8,border_width=1,border_color=box_border)
                line_box.pack(fill="x",padx=14,pady=(4,6))
                badge_ack=" \u2713 sous contr\u00f4le" if acquitte else ""
                _terme_r=r.get("terme","Pont").lower()
                ctk.CTkLabel(line_box,text=f"{r['carburant']} \u2014 {_terme_r} {r['trou_str']} ({r['trou_duree']}j non-livrable){badge_ack}",
                             font=("Segoe UI",12,"bold"),text_color=C["green"] if acquitte else C["red"],anchor="w").pack(anchor="w",padx=12,pady=(8,3))
                detail=(f"Il te manque environ {manque_arrondi:,}L (calcul bas\u00e9 sur tes saisies Pre_vision, marge demi-journ\u00e9e incluse)\n"
                        f"\U0001f550 Derni\u00e8re commande utile : {r['deadline_str']}").replace(",",".")
                ctk.CTkLabel(line_box,text=detail,font=("Segoe UI",11),text_color=C["t2"],anchor="w",justify="left").pack(anchor="w",padx=12,pady=(0,8))
        # Section : incohérences (commande sur jour non-livrable)
        incoh=ar.get("incoherences_jour_non_livrable",[])
        if incoh:
            sec=self._section(parent,"\u26a0 Commandes sur jour non-livrable",f"{len(incoh)} \u00e0 corriger")
            for i in incoh:
                txt=f"{i['date_str']} ({i['raison']}) \u2014 SP {int(i['sp']):,}L | GO {int(i['go']):,}L | GNR {int(i['gnr']):,}L".replace(",",".")
                self._row(sec,txt,"\u2192 replanifier",C["red"])
        # Section : SAISIES PHYSIQUEMENT IMPOSSIBLES (priorité haute, affichée avant ventes_irrealistes)
        saisies_irr=ar.get("saisies_physiquement_impossibles",[])
        if saisies_irr:
            sec=self._section(parent,"\U0001f6ab Saisies physiquement impossibles",f"{len(saisies_irr)} jour(s)/carburant(s) en exc\u00e8s")
            for s in saisies_irr:
                txt=f"{s['date_str']} {s['carburant']} : pr\u00e9vu {s['vente_saisie']:,}L vs max possible {s['vente_max_possible']:,}L".replace(",",".")
                detail=f"\u2192 r\u00e9duire de {s['exces']:,}L (cuve {s['dispo']:,}L \u2212 plancher {s['plancher']}L)".replace(",",".")
                self._row(sec,txt,detail,C["red"])
        # Section : ruptures projetées (issues du plafonnement physique)
        ruptures_proj=ar.get("ruptures_projetees",[])
        if ruptures_proj:
            # Calculer le total manque à gagner par carburant pour le résumé
            total_par_carb={"SP":0,"GO":0,"GNR":0}
            for r in ruptures_proj:
                total_par_carb[r["carburant"]]+=r["manque_a_gagner_l"]
            total_str_parts=[]
            for c in ["SP","GO","GNR"]:
                if total_par_carb[c]>0:
                    total_str_parts.append(f"{c} {total_par_carb[c]:,}L".replace(",","."))
            sub_str=f"{len(ruptures_proj)} rupture(s) projet\u00e9e(s) \u2014 manque \u00e0 gagner total : {', '.join(total_str_parts)}"
            sec=self._section(parent,"\U0001f4c9 Ruptures projet\u00e9es \u2014 ventes perdues",sub_str)
            for r in ruptures_proj:
                txt=f"{r['date_str']} {r['carburant']} : pr\u00e9vu {r['vente_voulue']:,}L \u2192 vendable {r['vente_reelle']:,}L".replace(",",".")
                detail=f"Manque \u00e0 gagner : {r['manque_a_gagner_l']:,}L (rupture en cours de journ\u00e9e, plancher cuve atteint)".replace(",",".")
                self._row(sec,txt,detail,C["amber"])
        # Section : ventes prévues irréalistes (vigilance)
        ventes_irr=ar.get("ventes_irrealistes",[])
        if ventes_irr:
            sec=self._section(parent,"\u26a0 Ventes pr\u00e9visionnelles \u00e0 v\u00e9rifier",f"{len(ventes_irr)} \u00e9cart(s) anormal(aux)")
            for v in ventes_irr:
                col=C["red"] if v["severity"]=="rouge" else C["amber"]
                signe="+" if v["ecart_pct"]>0 else ""
                txt=f"{v['date_str']} {v['carburant']} : saisi {v['saisi']:,}L vs attendu {v['attendu']:,}L".replace(",",".")
                self._row(sec,txt,f"{signe}{v['ecart_pct']}%",col)
        # Section informative : fériés à venir J+15 → J+30
        feries=ar.get("feries_a_venir",[])
        if feries:
            sec=self._section(parent,"\U0001f4c5 F\u00e9ri\u00e9s \u00e0 anticiper",f"{len(feries)} f\u00e9ri\u00e9(s) dans 15 \u00e0 30 jours")
            for f in feries:
                txt=f"{f['date_str']} \u2014 {f['nom']}"
                detail=f"Dans {f['j_restants']}j \u2502 trou de {f['duree_trou']}j non-livrable"
                self._row(sec,txt,detail,C["t2"])
        # Alerte fraîcheur du bilan
        fresh=pv.get("bilan_freshness","unknown")
        if fresh=="today":
            self._alert_box(parent,f"\u2713 Bilan mati\u00e8re \u00e0 jour \u2014 contient les ventes de la veille ({pv.get('bilan_date_raw','')}). Stocks affich\u00e9s = ce matin {jour_fr()} {date.today().strftime('%d/%m/%y')}.",C["green"])
            # Afficher les écarts du bilan
            es=pv.get("ecart_sp");eg=pv.get("ecart_go");eg2=pv.get("ecart_gnr")
            if es is not None or eg is not None:
                sec=self._section(parent,"\U0001f4cb \u00c9carts du bilan mati\u00e8re","Diff\u00e9rence comptable / physique")
                for nm,e in [("SP",es),("GO",eg),("GNR",eg2)]:
                    if e is not None:
                        col=C["green"] if abs(e)<10 else C["amber"] if abs(e)<50 else C["red"]
                        self._row(sec,nm,f"{e:+.0f} L",col)
        else:
            self._alert_box(parent,f"\u26a0 Bilan mati\u00e8re non actualis\u00e9 aujourd'hui. Stocks reconstitu\u00e9s \u00e0 partir de la veille.",C["amber"])
        # SECTION : Détail temps réel par carburant
        s=self._section(parent,"\u26fd Stocks par carburant",None)
        for nm,k in [("SP","sp"),("GO","go"),("GNR","gnr")]:
            matin=sf(pv.get(f"{k}_matin",0))
            reel=sf(pv.get(k,0))
            j=auto.get(k,0)
            jc=C["green"] if j>2 else C["amber"] if j>1 else C["red"]
            sub=f"Stock matin {fnum(matin,'L')}"
            if partial:
                p_vendu=sf(partial.get(k,0))
                if p_vendu>0: sub+=f" \u2212 {fnum(p_vendu,'L')} vendu"
            if pv.get("livr_recu"):
                livr_raw=load_json(LIVRAISON_FILE).get(date.today().strftime("%d/%m/%y"),{})
                livr_agg=aggregate_livr_day(livr_raw)
                p_livr=sf(livr_agg.get(k,0))
                if p_livr>0: sub+=f" + {fnum(p_livr,'L')} livr\u00e9"
            self._row(s,nm,f"{fnum(reel,'L')}  \u2192  {fmt_autonomie(j)}",jc,sub=sub)
        self._row(s,"Valorisation",feur(pv.get("valo"),d=0),C["gold"])
        # Statut livraison (logique weekend)
        is_weekend=date.today().weekday()>=5
        if pv.get("livr_recu"):
            self._alert_box(s,"\u2713 Livraison saisie pour aujourd'hui (incluse dans les stocks)",C["green"])
        elif is_weekend:
            self._alert_box(s,"\u2713 Pas de livraison pr\u00e9vue (week-end). Bouton Livraison disponible si exception.",C["green"])
        else:
            self._alert_box(s,"\u26a0 Aucune livraison saisie pour aujourd'hui.",C["amber"])
        # SECTION : Prévisions de vente — fix du bug (forecasts est maintenant une liste)
        forecasts=pv.get("forecasts",[])
        if forecasts:
            s2=self._section(parent,"\U0001f4c8 Pr\u00e9visions de vente \u2014 7 prochains jours",f"Source : tes onglets Pr\u00e9vision (cycle 14 jours)")
            # En-têtes
            h_row=ctk.CTkFrame(s2,fg_color=C["panel"]);h_row.pack(fill="x",padx=10,pady=(4,2))
            for i,t in enumerate(["Jour","SP","GO","GNR","Total","Onglet"]):
                w=110 if i==0 else 75 if i<4 else 90
                ctk.CTkLabel(h_row,text=t,font=("Segoe UI",9,"bold"),text_color=C["t3"],width=w,anchor="e" if 0<i<5 else "w").pack(side="left",padx=2,pady=4)
            for i,fc in enumerate(forecasts[:7]):
                r=ctk.CTkFrame(s2,fg_color="transparent");r.pack(fill="x",padx=10,pady=1)
                day_lbl=fc.get("day","")
                if i==0: day_lbl=f"\u25b6 {day_lbl}"
                tot=sf(fc.get("sp"))+sf(fc.get("go"))+sf(fc.get("gnr"))
                cells=[(day_lbl,"label"),(fnum(fc.get("sp"),"L"),"sp"),(fnum(fc.get("go"),"L"),"go"),(fnum(fc.get("gnr"),"L"),"gnr"),(fnum(tot,"L"),"tot"),(fc.get("sheet",""),"sheet")]
                for j,(val,k) in enumerate(cells):
                    w=110 if j==0 else 75 if j<4 else 90
                    col=C["amber"] if j==0 and i==0 else C["gold"] if j==0 else C["t3"] if j==5 else C["t1"]
                    ctk.CTkLabel(r,text=val,font=("Segoe UI",10,"bold" if j==0 and i==0 else "normal"),text_color=col,width=w,anchor="e" if 0<j<5 else "w").pack(side="left",padx=2,pady=2)
        # SECTION : COHÉRENCE PRÉVISION 14J
        proj14=self.data.get("proj14",{})
        if proj14:
            JOURS_COURTS_P=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
            alertes=proj14.get("alertes",[])
            anomalies=proj14.get("anomalies",{})
            ruptures=[a for a in alertes if a["severity"]=="rupture"]
            fin_cycles=[a for a in alertes if a["severity"]=="fin_cycle"]
            if proj14.get("ok"):
                # Affichage simplifié : juste les stocks minimums sur le cycle 14j (info synthétique).
                # Les RUPTURES détaillées et les COMMANDES sont gérées par le module anti-rupture
                # ci-dessus (pas de doublon ici).
                sp=self._section(parent,"\U0001f4ca Stocks minimums sur le cycle 14j","Niveau le plus bas pr\u00e9vu")
                for carb,nm in [("sp","SP"),("go","GO"),("gnr","GNR")]:
                    smin=proj14.get(f"stock_min_{carb}")
                    smin_d=proj14.get(f"stock_min_date_{carb}")
                    if smin is not None and smin_d:
                        d_str=f"{JOURS_COURTS_P[smin_d.weekday()]} {smin_d.strftime('%d/%m')}"
                        col=C["green"] if smin>3000 else C["amber"] if smin>1000 else C["red"]
                        self._row(sp,f"Stock min {nm}",f"{fnum(smin,'L')} ({d_str})",col)
            if fin_cycles:
                last_d=fin_cycles[0].get("last_livr_date_str","")
                carbs_str=", ".join(a["carburant"] for a in fin_cycles)
                self._alert_box(parent,f"\U0001f4c5 Fin du cycle Prévision le {last_d} — penser à remplir le cycle suivant ({carbs_str} concerné(s))",C["t2"])
            if anomalies:
                sa=self._section(parent,"\U0001f4c8 Anomalie de ventes aujourd'hui","Ventes en cours vs moyenne")
                for carb,anom in anomalies.items():
                    signe="+" if anom["ecart_pct"]>0 else ""
                    col=C["red"] if abs(anom["ecart_pct"])>40 else C["amber"]
                    lbl="hausse" if anom["ecart_pct"]>0 else "baisse"
                    self._row(sa,f"{carb.upper()} {signe}{anom['ecart_pct']}%",
                              f"{fnum(anom['ventes_partielles'],'L')} vs {fnum(anom['ventes_attendues'],'L')} attendus",col,
                              sub=f"À {anom['stage']} \u2014 {lbl} anormale")
                if ruptures:
                    self._alert_box(parent,"Si les ventes anormales se confirment, les dates de rupture pourraient être avancées. Surveillez de près et ajustez vos commandes.",C["amber"])
            elif not anomalies and proj14.get("ok") and not fin_cycles:
                self._alert_box(parent,"\u2713 Ventes conformes aux moyennes \u2014 votre stratégie de commande tient la route.",C["green"])

    # ---------- DÉTAIL OBJECTIF ----------
    def _detail_objectif(self,parent,d):
        alerts=d.get("alerts",{});ob=d.get("ob",{})
        hist=d.get("hist",[]);complete=[h for h in hist if not h.get("en_cours")]
        today_d=date.today()
        mois_nom=["janvier","février","mars","avril","mai","juin","juillet","août","septembre","octobre","novembre","décembre"][today_d.month-1]
        # Données consolidées
        enc_mois=sf(alerts.get("mois_total_cbcpcs",0))+sf(alerts.get("mois_esp",0))+sf(alerts.get("mois_cli",0))
        dec_mois=sf(alerts.get("dec_car",0))+sf(alerts.get("dec_fourn",0))+sf(alerts.get("dec_fg",0))+sf(alerts.get("dec_div",0))+sf(alerts.get("dec_soc",0))
        bal=sf(alerts.get("balance_de",0))
        # === 1. SYNTHÈSE MOIS (3 chiffres) ===
        bc=C["green"] if bal>=0 else C["red"]
        self._hero_kpi(parent,[
            ("📥 ENCAISSÉ "+mois_nom,feur(enc_mois,d=0),"Crédité au compte",C["green"]),
            ("⚖ BALANCE",feur(bal,d=0),"Trésorerie "+("positive" if bal>=0 else "négative"),bc),
            ("📤 DÉCAISSÉ "+mois_nom,feur(dec_mois,d=0),"Sorti du compte",C["amber"]),
        ])
        # === 2. ALERTES CRITIQUES ===
        cp_urg=[c for c in alerts.get("cp_pending",[]) if c.get("retard",0)>0]
        enc_retard=[e for e in alerts.get("enc_pending",[]) if e.get("reste") is not None and e["reste"]<0]
        dec_retard=[dd for dd in alerts.get("dec_pending",[]) if dd.get("reste") is not None and dd["reste"]<0]
        bc_imp=alerts.get("by_client",{})
        cli_old=[(n,dt) for n,dt in bc_imp.items() if dt.get("max_age",0)>30]
        kw=alerts.get("keyword_alerts",[])
        has_critical=bool(kw or cp_urg or enc_retard or dec_retard or cli_old)
        if has_critical:
            sa=self._section(parent,"⚠ Alertes critiques","À traiter en priorité")
            if kw:
                # Compter les opérations distinctes (pas les mots-clés)
                ops_distinct=set()
                for k in kw:
                    if k.get('nom'): ops_distinct.add((k['onglet'],k['nom'],k['montant']))
                nb_ops=len(ops_distinct) if ops_distinct else len(kw)
                self._alert_box(sa,f"🚨 {nb_ops} opération(s) à vérifier — détail plus bas",C["red"])
            if dec_retard:
                tot=sum(dd["montant"] for dd in dec_retard);oldest=max(-dd["reste"] for dd in dec_retard)
                self._row(sa,f"📤 Décaissements en retard",feur(tot,d=0),C["red"],sub=f"{len(dec_retard)} opération(s) · plus ancien {oldest}j · détail plus bas")
            if enc_retard:
                tot=sum(e["montant"] for e in enc_retard);oldest=max(-e["reste"] for e in enc_retard)
                self._row(sa,f"📥 Encaissements en retard",feur(tot,d=0),C["red"],sub=f"{len(enc_retard)} opération(s) · plus ancien {oldest}j · détail plus bas")
            if cp_urg:
                tot=sum(c["montant"] for c in cp_urg);oldest=max(c["retard"] for c in cp_urg)
                self._row(sa,f"💳 CP en retard",feur(tot,d=0),C["red"],sub=f"{len(cp_urg)} opération(s) · plus ancien {oldest}j · détail plus bas")
            if cli_old:
                tot=sum(dt["total"] for _,dt in cli_old);oldest=max(dt["max_age"] for _,dt in cli_old)
                self._row(sa,f"👥 Règlements en attente >30j",feur(tot,d=0),C["red"],sub=f"{len(cli_old)} client(s) · plus ancien {oldest}j · détail plus bas")
        # Section pliable : opérations à surveiller (groupées par ligne réelle)
        if kw:
            grouped={}
            for ka in kw:
                if not ka.get('nom'): continue
                key=(ka['onglet'],ka['nom'],ka['montant'])
                if key not in grouped:
                    grouped[key]={'onglet':ka['onglet'],'nom':ka['nom'],'montant':ka['montant'],'mots':set(),'cell':ka.get('cell',''),'mode':ka.get('mode',''),'info':ka.get('info',''),'date':ka.get('date',''),'okko':ka.get('okko',''),'paye':ka.get('paye','')}
                grouped[key]['mots'].add(ka['mot'])
                # Compléter mode/info/date/okko/paye si vide
                if not grouped[key]['mode'] and ka.get('mode'): grouped[key]['mode']=ka['mode']
                if not grouped[key]['info'] and ka.get('info'): grouped[key]['info']=ka['info']
                if not grouped[key]['date'] and ka.get('date'): grouped[key]['date']=ka['date']
                if not grouped[key]['okko'] and ka.get('okko'): grouped[key]['okko']=ka['okko']
                if not grouped[key]['paye'] and ka.get('paye'): grouped[key]['paye']=ka['paye']
            uniq_kw=list(grouped.values())
            # Classification entrée/sortie par onglet
            ENTREE_ONGLETS={"Encaissements dep. Exp","Encaissement CB-CP-CS","Clients en compte","Recouvrements","Divers","CA Mensuel"}
            SORTIE_ONGLETS={"Décaissement Total C.","Règlements fourniss.","Frais généraux","Déc.divers","Charges soc - fisc"}
            def _classify(ka):
                """Retourne (signe, couleur). Red flag si ko/n même sur entrée."""
                is_entree=ka['onglet'] in ENTREE_ONGLETS
                # 'retour saisie' = remboursement, opération positive remarquable (PAS un red flag)
                # On regarde dans la cellule matchée, le mode et l'info
                is_retour=any('retour' in str(ka.get(f,'')).lower() for f in ('cell','mode','info'))
                # Détection problème : ko, ou paye=n, ou mode contient impayé/rejeté
                is_problem=(ka.get('okko')=='ko' or ka.get('paye')=='n' or
                            any(p in ka.get('mode','').lower() for p in ['impay','rejet']))
                # Override : 'retour saisie' n'est jamais un problème
                if is_retour and ka.get('mot')=='saisie':
                    is_problem=False
                if is_problem:
                    # Red flag : rouge, signe selon sens (entrée problématique = manque à gagner = -)
                    return ('-' if is_entree else '-',C["red"])
                # OK : couleur selon sens
                if is_entree: return ('+',C["green"])
                else: return ('-',C["red"])
            # Tri par date la plus récente en premier (date manquante en bas)
            def _sort_key(ka):
                d=ka.get('date')
                if hasattr(d,'toordinal'): return (0,-d.toordinal())
                return (1,0)  # sans date à la fin
            uniq_kw.sort(key=_sort_key)
            if uniq_kw:
                coll=self._collapsible(parent,"🚨 Opérations à surveiller",len(uniq_kw),None,C["red"])
                for ka in uniq_kw:
                    mots_str=", ".join(sorted(ka['mots']))
                    signe,couleur=_classify(ka)
                    # Construction du sub : date + mode + info
                    sub_parts=[]
                    d=ka.get('date')
                    if d:
                        if hasattr(d,'strftime'): sub_parts.append(d.strftime("%d/%m/%Y"))
                        else: sub_parts.append(str(d).replace(" 00:00:00","")[:10])
                    if ka.get('mode'): sub_parts.append(ka['mode'])
                    info_val=ka.get('info','')
                    if info_val:
                        if hasattr(info_val,'strftime'):
                            sub_parts.append(info_val.strftime("%d/%m/%Y"))
                        else:
                            info_str=str(info_val).strip()
                            if " 00:00:00" in info_str: info_str=info_str.replace(" 00:00:00","")
                            if info_str: sub_parts.append(info_str[:50])
                    cell_txt=str(ka.get('cell','')).strip()
                    is_just_keyword=cell_txt.lower() in {m.lower() for m in ka['mots']}
                    if cell_txt and not is_just_keyword and cell_txt not in sub_parts:
                        sub_parts.append(cell_txt[:50])
                    sub=" · ".join(sub_parts) if sub_parts else None
                    # Montant signé
                    montant_str=f"{signe}{feur(ka['montant'],d=0)}"
                    self._row(coll,f"{ka['onglet']} — {ka['nom']} [{mots_str}]",montant_str,couleur,sub=sub)
        # === 3. FLUX ENTRANT À VENIR ===
        enc_pending=alerts.get("enc_pending",[])
        cp_pending=alerts.get("cp_pending",[])
        all_in=[]
        for e in enc_pending: all_in.append({"cat":e["cat"],"nom":e["nom"],"montant":e["montant"],"reste":e.get("reste")})
        # cp_pending est un sous-ensemble de enc_pending (scan des mêmes lignes), déjà inclus
        if all_in:
            tot_in=sum(x["montant"] for x in all_in)
            in_retard=[x for x in all_in if x.get("reste") is not None and x["reste"]<0]
            sf2=self._section(parent,"📥 Flux entrant à venir",None)
            in_col=C["red"] if in_retard else C["amber"]
            in_sub=f"Total à encaisser"+(f" · dont {feur(sum(x['montant'] for x in in_retard),d=0)} en retard" if in_retard else "")
            self._total_banner(sf2,in_sub,feur(tot_in,d=0),in_col,sub=f"{len(all_in)} opération(s)")
            # Ventilation par type
            by_type={}
            for x in all_in: by_type.setdefault(x["cat"],0);by_type[x["cat"]]+=x["montant"]
            for typ,mt in sorted(by_type.items(),key=lambda kv:-kv[1]):
                self._row(sf2,typ,feur(mt,d=0),C["t1"])
            # Liste pliable
            coll=self._collapsible(parent,"Voir détail opérations entrantes",len(all_in),feur(tot_in,d=0),C["amber"])
            for x in sorted(all_in,key=lambda y:(y.get("reste") is None, y.get("reste") if y.get("reste") is not None else 0)):
                r=x.get("reste")
                if r is None: sub="date inconnue"
                elif r<0: sub=f"\u00c9ch\u00e9ance d\u00e9pass\u00e9e de {-r}j"
                elif r==0: sub="Aujourd'hui"
                else: sub=f"Dans {r}j"
                col=C["red"] if r is not None and r<0 else (C["amber"] if r is not None and r<=3 else C["t1"])
                self._row(coll,f"{x['cat']} — {x['nom']}",feur(x["montant"],d=0),col,sub=sub)
        # === 4. FLUX SORTANT À VENIR ===
        dec_pending=alerts.get("dec_pending",[])
        if dec_pending:
            tot_out=sum(x["montant"] for x in dec_pending)
            out_retard=[x for x in dec_pending if x.get("reste") is not None and x["reste"]<0]
            sf3=self._section(parent,"📤 Flux sortant à venir",None)
            out_col=C["red"] if out_retard else C["amber"]
            out_sub=f"Total à décaisser"+(f" · dont {feur(sum(x['montant'] for x in out_retard),d=0)} en retard" if out_retard else "")
            self._total_banner(sf3,out_sub,feur(tot_out,d=0),out_col,sub=f"{len(dec_pending)} opération(s)")
            by_cat={}
            for x in dec_pending:
                k=x["cat"].replace("Décaissement Total C.","Carburant").replace("Règlements fourniss.","Fournisseurs").replace("Frais généraux","Frais généraux").replace("Charges soc - fisc","Charges soc/fisc").replace("Déc.divers","Divers")
                by_cat.setdefault(k,0);by_cat[k]+=x["montant"]
            for cat,mt in sorted(by_cat.items(),key=lambda kv:-kv[1]):
                self._row(sf3,cat,feur(mt,d=0),C["t1"])
            coll=self._collapsible(parent,"Voir détail opérations sortantes",len(dec_pending),feur(tot_out,d=0),C["amber"])
            for x in sorted(dec_pending,key=lambda y:(y.get("reste") is None, y.get("reste") if y.get("reste") is not None else 0)):
                r=x.get("reste")
                if r is None: sub="date inconnue"
                elif r<0: sub=f"\u00c9ch\u00e9ance d\u00e9pass\u00e9e de {-r}j"
                elif r==0: sub="Aujourd'hui"
                else: sub=f"Dans {r}j"
                col=C["red"] if r is not None and r<0 else (C["amber"] if r is not None and r<=3 else C["t1"])
                cat_s=x["cat"].replace("Décaissement Total C.","Carb").replace("Règlements fourniss.","Fourn").replace("Frais généraux","FG").replace("Charges soc - fisc","Soc/Fisc").replace("Déc.divers","Div")
                self._row(coll,f"{cat_s} — {x['nom']}",feur(x["montant"],d=0),col,sub=sub)
        # === 5. ENCAISSÉ CE MOIS (détail) ===
        cb_m=sf(alerts.get("mois_cb",0));cp_m=sf(alerts.get("mois_cp",0));cs_m=sf(alerts.get("mois_cs",0))
        esp_m=sf(alerts.get("mois_esp",0));cli_m=sf(alerts.get("mois_cli",0))
        if enc_mois>0:
            coll=self._collapsible(parent,f"💰 Encaissé en {mois_nom} — détail par type",5,feur(enc_mois,d=0),C["green"])
            self._row(coll,"CB",feur(cb_m,d=0),C["green"])
            self._row(coll,"CP",feur(cp_m,d=0),C["green"])
            self._row(coll,"CS (chèques service)",feur(cs_m,d=0),C["green"])
            self._row(coll,"Espèces (dépôts express)",feur(esp_m,d=0),C["green"])
            self._row(coll,"Clients en compte",feur(cli_m,d=0),C["green"])
        # === 6. DÉCAISSÉ CE MOIS (détail) ===
        if dec_mois>0:
            coll=self._collapsible(parent,f"💸 Décaissé en {mois_nom} — détail par type",5,feur(dec_mois,d=0),C["amber"])
            self._row(coll,"Carburant",feur(alerts.get("dec_car",0),d=0),C["t1"])
            self._row(coll,"Fournisseurs",feur(alerts.get("dec_fourn",0),d=0),C["t1"])
            self._row(coll,"Frais généraux",feur(alerts.get("dec_fg",0),d=0),C["t1"])
            self._row(coll,"Charges soc/fisc",feur(alerts.get("dec_soc",0),d=0),C["t1"])
            self._row(coll,"Divers",feur(alerts.get("dec_div",0),d=0),C["t1"])
        # === 7. AVANCEMENT OBJECTIF CA ===
        if ob.get("st")=="ok":
            obj=sf(ob.get("obj_ca"));enc_ca=sf(ob.get("enc_ca"));taux=sf(ob.get("taux"))
            proj=enc_ca/today_d.day*30 if today_d.day>0 and enc_ca>0 else 0
            tc=C["green"] if taux>0.25 else C["amber"] if taux>0.10 else C["red"]
            sp=self._section(parent,"📊 Avancement objectif CA",f"Objectif mensuel {feur(obj,d=0)}")
            self._row(sp,"Encours CA",feur(enc_ca,d=0),tc,sub=f"{taux*100:.1f}% de l'objectif".replace(".",","))
            self._row(sp,"Projection 30j",feur(proj,d=0),C["green"] if proj>=obj else C["amber"])
        # === 8. CLIENTS IMPAYÉS (pliable) ===
        if bc_imp:
            tot_imp=sf(alerts.get("cli_total",0))
            coll=self._collapsible(parent,"👥 Règlements en attente — détail par client",len(bc_imp),feur(tot_imp,d=0),C["red"])
            for nom,dt in sorted(bc_imp.items(),key=lambda x:x[1]["max_age"],reverse=True):
                col=C["red"] if dt["max_age"]>60 else C["amber"]
                self._row(coll,nom,feur(dt["total"],d=0),col,sub=f"{dt['count']} facture(s) · plus ancienne {dt['max_age']}j")

    # ---------- DÉTAIL LITRAGE ----------
    def _detail_litrage(self,parent,d):
        td=self.today_data;j7=self.j7;complete=self.complete
        if not complete: return
        # HERO : Aujourd'hui en cours ou dernier jour complet
        if td:
            nb=td.get("nb_caisses",0)
            j7_lit=0;j7_piste=0
            if j7 and nb>0:
                for i in range(1,nb+1):
                    j7_lit+=sf(j7.get("caisses",{}).get(str(i),{}).get("litrage",0))
                    j7_piste+=sf(j7.get("caisses",{}).get(str(i),{}).get("piste_eur",0))
            lit_delta=self._delta_text(td.get("litrage",0),j7_lit)
            piste_delta=self._delta_text(td.get("piste",0),j7_piste)
            self._hero_kpi(parent,[
                ("LITRAGE AUJOURD'HUI",fnum(td.get("litrage"),"L"),lit_delta[0],lit_delta[1] if lit_delta[1]!=C["t3"] else C["gold"]),
                ("CA PISTE",feur(td.get("piste"),d=0),piste_delta[0],piste_delta[1] if piste_delta[1]!=C["t3"] else C["teal"]),
                ("CAISSES",f"{nb}/3","Jour en cours",C["amber"]),
            ])
        # SECTION : Performance générale (vignette LITRAGE = focus litres)
        avg=sum(h.get("total",0) for h in complete)/len(complete)
        avg_lit=sum(h.get("litrage",0) for h in complete)/len(complete)
        # Best/worst sur le LITRAGE (pas le CA, pour cohérence avec la vignette)
        best=max(complete,key=lambda x:x.get("litrage",0))
        worst=min(complete,key=lambda x:x.get("litrage",0))
        s=self._section(parent,"\U0001f4ca Performance g\u00e9n\u00e9rale",f"Sur {len(complete)} jours complets")
        self._row(s,"Litrage moyen /jour",fnum(avg_lit,"L"),C["t1"])
        self._row(s,"CA total moyen /jour",feur(avg,d=0),C["t3"])
        self._row(s,"\u2b06 Meilleur jour",fnum(best.get('litrage'),"L"),C["green"],sub=f"{best.get('label','')} \u00b7 {feur(best.get('total'),d=0)}")
        self._row(s,"\u2b07 Plus faible",fnum(worst.get('litrage'),"L"),C["red"],sub=f"{worst.get('label','')} \u00b7 {feur(worst.get('total'),d=0)}")
        # SECTION : Détail par carburant (3 cards sectorisées avec liseré couleur)
        s2=self._section(parent,"\u26fd D\u00e9tail par carburant",None)
        carbs_cfg=[("SP","sp",C["green"]),("GO","go",C["blue"]),("GNR","gnr",C["amber"])]
        for carb,k,col_carb in carbs_cfg:
            vals=[h.get(k,0) for h in complete if h.get(k,0)>0]
            if not vals: continue
            cmax=max(complete,key=lambda x:x.get(k,0))
            cmin=min((h for h in complete if h.get(k,0)>0),key=lambda x:x.get(k,0))
            moy=sum(vals)/len(vals)
            # Card compacte : liseré couleur + contenu en grille logique
            card=ctk.CTkFrame(s2,fg_color=C["card_h"],corner_radius=8,border_width=1,border_color=C["border2"])
            card.pack(fill="x",padx=14,pady=(4,6))
            # Liseré couleur (5px) sur toute la hauteur à gauche
            ctk.CTkFrame(card,fg_color=col_carb,corner_radius=0,width=5).pack(side="left",fill="y")
            # Body : grille 4 colonnes (label carb | Moyenne | Max | Min)
            body=ctk.CTkFrame(card,fg_color="transparent")
            body.pack(side="left",fill="both",expand=True,padx=(16,16),pady=14)
            body.grid_columnconfigure(0,weight=0,minsize=60)
            for ci in (1,2,3): body.grid_columnconfigure(ci,weight=1)
            # Colonne 0 : label carburant (gros, couleur du liseré)
            ctk.CTkLabel(body,text=carb,font=("Segoe UI",20,"bold"),text_color=col_carb,anchor="w").grid(row=0,column=0,rowspan=3,sticky="w",padx=(0,12))
            # Colonne 1 : Moyenne (label en clair + valeur en gros blanc)
            ctk.CTkLabel(body,text="Moyenne",font=("Segoe UI",11,"bold"),text_color=C["t1"],anchor="w").grid(row=0,column=1,sticky="w")
            ctk.CTkLabel(body,text=fnum(moy,"L"),font=("Segoe UI",15,"bold"),text_color=C["t1"],anchor="w").grid(row=1,column=1,sticky="w",pady=(2,0))
            # (pas de date pour moyenne, laisser rowspan 2)
            # Colonne 2 : Max (label vert + valeur verte + date en clair)
            ctk.CTkLabel(body,text="\u2b06 Max",font=("Segoe UI",11,"bold"),text_color=C["green"],anchor="w").grid(row=0,column=2,sticky="w")
            ctk.CTkLabel(body,text=fnum(cmax.get(k),"L"),font=("Segoe UI",15,"bold"),text_color=C["green"],anchor="w").grid(row=1,column=2,sticky="w",pady=(2,0))
            ctk.CTkLabel(body,text=cmax.get("label",""),font=("Segoe UI",10),text_color=C["t2"],anchor="w").grid(row=2,column=2,sticky="w",pady=(2,0))
            # Colonne 3 : Min (label rouge + valeur rouge + date en clair)
            ctk.CTkLabel(body,text="\u2b07 Min",font=("Segoe UI",11,"bold"),text_color=C["red"],anchor="w").grid(row=0,column=3,sticky="w")
            ctk.CTkLabel(body,text=fnum(cmin.get(k),"L"),font=("Segoe UI",15,"bold"),text_color=C["red"],anchor="w").grid(row=1,column=3,sticky="w",pady=(2,0))
            ctk.CTkLabel(body,text=cmin.get("label",""),font=("Segoe UI",10),text_color=C["t2"],anchor="w").grid(row=2,column=3,sticky="w",pady=(2,0))

# =============================================================================
# =============================================================================
# Vignette RACCOURCIS — accès rapide aux outils externes
# =============================================================================
class RaccourciVignette(ctk.CTkFrame):
    """Vignette spéciale avec 6 cercles cliquables vers outils externes."""
    def __init__(self,parent,raccourcis):
        super().__init__(parent,fg_color=C["card"],corner_radius=14,border_width=1,border_color=C["border"],height=180)
        self.pack_propagate(False)
        # Liseré gauche ambre désaturé (signature unique des raccourcis, plus en doublon avec Prévision)
        ctk.CTkFrame(self,fg_color=C["vig_amber"],corner_radius=0,width=4).pack(side="left",fill="y")
        # Contenu
        body=ctk.CTkFrame(self,fg_color="transparent");body.pack(side="left",fill="both",expand=True,padx=14,pady=12)
        # Header
        hdr=ctk.CTkFrame(body,fg_color="transparent");hdr.pack(fill="x",pady=(0,8))
        ctk.CTkLabel(hdr,text="\U0001f517 RACCOURCIS",font=("Segoe UI",13,"bold"),text_color=C["vig_amber"]).pack(side="left")
        ctk.CTkLabel(hdr,text="Outils & sites",font=("Segoe UI",10),text_color=C["t3"]).pack(side="left",padx=(8,0))
        # Grille 3x2 de cercles
        grid=ctk.CTkFrame(body,fg_color="transparent");grid.pack(fill="both",expand=True)
        for i in range(3): grid.grid_columnconfigure(i,weight=1)
        for i in range(2): grid.grid_rowconfigure(i,weight=1)
        for i,r in enumerate(raccourcis):
            row,col=i//3,i%3
            cell=ctk.CTkFrame(grid,fg_color="transparent");cell.grid(row=row,column=col,sticky="nsew",padx=4,pady=4)
            if PIL_OK:
                # Bulles Pillow : dégradé bombé + reflet + ombre portée (vraies images)
                img_normal=make_bubble(r["color"],r["icon"],size=62,state="normal")
                img_hover=make_bubble(r["color"],r["icon"],size=62,state="hover")
                img_press=make_bubble(r["color"],r["icon"],size=62,state="press")
                ci_normal=ctk.CTkImage(light_image=img_normal,dark_image=img_normal,size=(62,62))
                ci_hover=ctk.CTkImage(light_image=img_hover,dark_image=img_hover,size=(62,62))
                ci_press=ctk.CTkImage(light_image=img_press,dark_image=img_press,size=(62,62))
                bubble_btn=ctk.CTkLabel(cell,image=ci_normal,text="",cursor="hand2")
                bubble_btn.pack(pady=(2,2))
                lbl=ctk.CTkLabel(cell,text=r["label"],font=("Segoe UI",9),text_color=C["t2"])
                lbl.pack(pady=(6,0))
                def _hover_on(e,b=bubble_btn,img=ci_hover):
                    b.configure(image=img)
                def _hover_off(e,b=bubble_btn,img=ci_normal):
                    b.configure(image=img)
                def _press(e,b=bubble_btn,img=ci_press):
                    b.configure(image=img)
                def _release(e,b=bubble_btn,img=ci_normal,a=r["action"]):
                    b.configure(image=img);a()
                for w in [cell,bubble_btn,lbl]:
                    w.bind("<Enter>",_hover_on)
                    w.bind("<Leave>",_hover_off)
                    w.bind("<ButtonPress-1>",_press)
                    w.bind("<ButtonRelease-1>",_release)
            else:
                # Fallback sans Pillow : bulle customtkinter simple
                circle_wrap=ctk.CTkFrame(cell,fg_color="transparent",width=56,height=56)
                circle_wrap.pack(pady=(2,2));circle_wrap.pack_propagate(False)
                shadow=ctk.CTkFrame(circle_wrap,width=50,height=50,fg_color="#000814",corner_radius=25)
                shadow.place(x=3,y=6)
                circle=ctk.CTkFrame(circle_wrap,width=50,height=50,fg_color=r["color"],corner_radius=25)
                circle.place(x=3,y=2);circle.pack_propagate(False)
                circle_txt=ctk.CTkLabel(circle,text=r["icon"],font=("Segoe UI Symbol",20,"bold"),text_color="#FFFFFF")
                circle_txt.place(relx=0.5,rely=0.5,anchor="center")
                lbl=ctk.CTkLabel(cell,text=r["label"],font=("Segoe UI",9),text_color=C["t2"])
                lbl.pack(pady=(6,0))
                def _press(e,c=circle,a=r["action"]):
                    c.place(x=3,y=5)
                def _release(e,c=circle,a=r["action"]):
                    c.place(x=3,y=2);a()
                for w in [cell,circle_wrap,circle,circle_txt,lbl]:
                    w.bind("<ButtonPress-1>",_press)
                    w.bind("<ButtonRelease-1>",_release)

# =============================================================================
# Lanceurs de raccourcis
# =============================================================================
def launch_chrome(url,incognito=False):
    """Lance Chrome (incognito optionnel) sur l'URL donnée."""
    chrome=r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    if not Path(chrome).exists():
        chrome=r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    args=[chrome]
    if incognito: args.append("--incognito")
    args.append(url)
    try: subprocess.Popen(args)
    except Exception as e: messagebox.showerror("Erreur",f"Impossible de lancer Chrome :\n{e}")

def open_onedrive(cfg_holder):
    """Ouvre le dossier OneDrive local. Demande le chemin si pas configuré."""
    cfg=load_json(CONFIG_FILE)
    od=cfg.get("onedrive_path","")
    # Vérifier que le chemin existe
    if not od or not Path(od).exists():
        # Tenter le chemin par défaut
        default=Path.home()/"OneDrive"
        if default.exists():
            od=str(default);cfg["onedrive_path"]=od;save_json(CONFIG_FILE,cfg)
        else:
            # Demander à l'utilisateur
            od=filedialog.askdirectory(title="Sélectionnez votre dossier OneDrive")
            if not od: return
            cfg["onedrive_path"]=od;save_json(CONFIG_FILE,cfg)
    try: os.startfile(od)
    except Exception as e: messagebox.showerror("Erreur",f"Impossible d'ouvrir OneDrive :\n{e}")

def open_orange_mail_tool():
    """Outil de récupération de pièces jointes Orange Mail (à intégrer en V0.5)."""
    messagebox.showinfo("Orange Mail Factures","\U0001f527 Outil de récupération de pièces jointes Orange Mail\n\nCet outil sera intégré dans la prochaine version (V0.5).\nIl utilisera l'API OX (/find?action=query) avec :\n  • Date pickers début/fin\n  • Recherche par fournisseur\n  • Téléchargement vers Desktop ou dossier choisi\n  • Sous-dossiers Factures_<term>/ par fournisseur\n\nEn attendant : la version HTML standalone reste utilisable.")


# =============================================================================
# MINI SERVEUR WEB INTÉGRÉ (validé Bidou 25/05/2026)
# =============================================================================
# Architecture Option A hybride : le hub Python continue de tourner normalement.
# En parallèle, un mini-serveur HTTP basé sur la stdlib expose :
#   - GET /api/dashboard → JSON des données du dernier refresh (last_data)
#   - GET /              → page HTML moderne consommant l'API
#
# Avantages :
#   - Zéro dépendance externe (pas de pip install fastapi/flask/uvicorn).
#   - Démarre dans un thread daemon, n'interfère pas avec le hub principal.
#   - Multi-utilisateur (ThreadingHTTPServer).
#   - Accessible depuis localhost ET le réseau local (ports 8765 par défaut).
#   - Préfigure l'accès distant via Cloudflare Tunnel (étape ultérieure).
#
# Source de vérité unique : self.last_data du hub. L'interface web ne fait QUE
# lire ces données, jamais recalculer. Le moteur métier reste intouché.

_HUB_INSTANCE_REF=None  # Référence globale au hub pour accès depuis le handler HTTP

def _json_safe(obj):
    """Convertit récursivement un objet pour qu'il soit JSON-serializable.
    Gère datetime, date, set, et autres types non natifs JSON.
    Indispensable car self.last_data contient des objets date/datetime un peu partout."""
    if isinstance(obj,(datetime,date)):
        return obj.isoformat()
    if isinstance(obj,set):
        return list(obj)
    if isinstance(obj,dict):
        return {str(k):_json_safe(v) for k,v in obj.items()}
    if isinstance(obj,(list,tuple)):
        return [_json_safe(v) for v in obj]
    if hasattr(obj,"__dict__"):
        # Objet custom : on tente d'extraire ses attributs sérialisables
        try: return _json_safe({k:v for k,v in obj.__dict__.items() if not k.startswith("_")})
        except Exception: return str(obj)
    return obj

# Page HTML servie à la racine — design sobre, cohérent avec la maquette journal validée.
# Volontairement minimaliste pour cette première itération : on affiche le pilotage jour.
# On enrichira progressivement (journal, livraisons, etc.).
_DASHBOARD_HTML=r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<title>DISTRICARB HUB</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  :root {
    --bg: #0E1117;
    --bg-card: #15191F;
    --bg-panel: #11151A;
    --border: rgba(255, 255, 255, 0.07);
    --text-1: #E6E8EB;
    --text-2: rgba(230, 232, 235, 0.65);
    --text-3: rgba(230, 232, 235, 0.42);
    --red: #C94A52;
    --blue: #5B92D4;
    --gold: #E4BC4D;
    --green: #4FAE5F;
    --teal: #3FB5A3;
    --amber: #D4934A;
  }
  html, body { height: 100%; }
  body {
    background: var(--bg); color: var(--text-1);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Helvetica Neue", sans-serif;
    font-size: 12px; line-height: 1.4;
    -webkit-font-smoothing: antialiased;
    padding: 8px 12px; overflow-x: hidden;
  }
  /* Layout dense type cockpit, sans scroll vertical sur desktop standard */
  .topbar {
    display: flex; align-items: center; justify-content: space-between;
    padding: 4px 4px 8px; border-bottom: 1px solid var(--border);
    margin-bottom: 8px;
  }
  .topbar h1 {
    font-size: 16px; font-weight: 700; letter-spacing: -0.2px;
    display: flex; align-items: center; gap: 8px;
  }
  .topbar h1::before {
    content: ""; width: 8px; height: 8px; border-radius: 50%;
    background: var(--red); flex-shrink: 0;
  }
  .topbar .right { display: flex; align-items: center; gap: 12px; }
  .topbar .meta { color: var(--text-3); font-size: 11px; }
  .btn {
    background: transparent; color: var(--text-2);
    border: 1px solid var(--border); border-radius: 6px;
    padding: 4px 10px; cursor: pointer; font-family: inherit; font-size: 11px;
  }
  .btn:hover { background: rgba(255,255,255,0.04); color: var(--text-1); }
  .btn-primary {
    background: rgba(75, 174, 95, 0.12); color: var(--green);
    border-color: rgba(75, 174, 95, 0.4);
  }
  .btn-primary:hover { background: rgba(75, 174, 95, 0.2); }

  /* Grille principale : 3 colonnes. Vignettes | Alertes actives | Pilotage Jour */
  .main {
    display: grid;
    grid-template-columns: minmax(0, 1fr) 280px 320px;
    gap: 10px;
    align-items: start;
  }
  /* Colonne gauche : 5 vignettes en 2 rangées */
  .vignettes {
    display: grid; grid-template-columns: 1fr 1fr 1fr;
    gap: 8px;
  }
  .card {
    background: var(--bg-card); border: 1px solid var(--border);
    border-radius: 8px; padding: 10px 12px;
    border-left: 3px solid var(--red);
  }
  .card[data-status="ok"] { border-left-color: var(--green); }
  .card[data-status="warn"] { border-left-color: var(--amber); }
  .card[data-status="alert"] { border-left-color: var(--red); }
  .card-head {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 4px;
  }
  .card-title {
    font-size: 10px; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.6px; color: var(--text-3);
  }
  .card-dot { width: 6px; height: 6px; border-radius: 50%; background: var(--text-3); }
  .card-dot.ok { background: var(--green); }
  .card-dot.warn { background: var(--amber); }
  .card-dot.alert { background: var(--red); }
  .card-sub { font-size: 10px; color: var(--text-2); margin-bottom: 8px; }
  .row {
    display: flex; justify-content: space-between; align-items: baseline;
    padding: 4px 0; gap: 8px;
    border-bottom: 1px solid var(--border);
  }
  .row:last-child { border-bottom: none; }
  .row-label {
    color: var(--text-2); font-size: 9.5px;
    text-transform: uppercase; letter-spacing: 0.4px; font-weight: 600;
  }
  .row-value {
    color: var(--text-1); font-weight: 700; font-size: 12px;
    font-variant-numeric: tabular-nums; white-space: nowrap;
    text-align: right;
  }

  /* Colonne milieu : Alertes actives */
  .alerts-panel {
    background: var(--bg-panel); border: 1px solid var(--border);
    border-radius: 8px; padding: 10px;
  }
  .panel-title {
    color: var(--red); font-size: 10px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.6px; margin-bottom: 8px;
    display: flex; align-items: center; justify-content: space-between;
  }
  .panel-title.gold { color: var(--gold); }
  .panel-title .count {
    background: rgba(201, 74, 82, 0.15); color: var(--red);
    padding: 1px 6px; border-radius: 10px; font-size: 10px;
  }
  .alert-item {
    background: var(--bg-card); border: 1px solid var(--border);
    border-left: 2px solid var(--amber);
    border-radius: 6px; padding: 6px 8px; margin-bottom: 6px;
    display: flex; flex-direction: column; gap: 4px;
  }
  .alert-item.pont { border-left-color: var(--amber); }
  .alert-item.weekend { border-left-color: #7B8896; }
  .alert-item.rupture { border-left-color: var(--red); }
  .alert-item.marge_tendue { border-left-color: var(--amber); }
  .alert-item.livraison_reporter { border-left-color: var(--blue); }
  .alert-item.anomalie { border-left-color: var(--red); }
  .alert-item.ferie_isole { border-left-color: var(--amber); }
  .alert-item.saisies_irregulieres { border-left-color: var(--red); }
  .alert-text { font-size: 11px; color: var(--text-1); line-height: 1.3; }
  .alert-actions { display: flex; gap: 4px; align-items: center; }
  .alert-statut {
    font-size: 9px; padding: 1px 5px; border-radius: 3px;
    background: rgba(201, 74, 82, 0.15); color: var(--red);
    text-transform: uppercase; letter-spacing: 0.4px; font-weight: 600;
  }
  .alert-statut.snooze { background: rgba(212, 147, 74, 0.15); color: var(--amber); }
  .alert-statut.ack { background: rgba(91, 146, 212, 0.15); color: var(--blue); }
  .btn-resolve {
    background: rgba(75, 174, 95, 0.1); color: var(--green);
    border: 1px solid rgba(75, 174, 95, 0.3);
    border-radius: 4px; padding: 2px 8px; font-size: 10px;
    cursor: pointer; font-family: inherit; margin-left: auto;
  }
  .btn-resolve:hover { background: rgba(75, 174, 95, 0.2); }
  .empty-state {
    color: var(--text-3); font-size: 11px; font-style: italic;
    padding: 12px 4px; text-align: center;
  }

  /* Colonne droite : Pilotage Jour */
  .pilotage {
    background: var(--bg-panel); border: 1px solid var(--border);
    border-radius: 8px; padding: 10px 12px;
  }
  .pilotage-head { margin-bottom: 8px; }
  .pilotage-date {
    font-size: 15px; font-weight: 700; color: var(--text-1);
  }
  .pilotage-sub {
    font-size: 10px; color: var(--text-3); margin-top: 1px;
  }
  .section { margin-top: 8px; }
  .section-head {
    display: flex; align-items: center; gap: 6px; margin-bottom: 4px;
  }
  .section-bar {
    width: 3px; height: 10px; border-radius: 1px; background: var(--red);
  }
  .section-bar.accent-blue { background: var(--blue); }
  .section-bar.accent-amber { background: var(--amber); }
  .section-bar.accent-gold { background: var(--gold); }
  .section-bar.accent-green { background: var(--green); }
  .section-bar.accent-teal { background: var(--teal); }
  .section-title {
    font-size: 9.5px; font-weight: 700; color: var(--text-2);
    text-transform: uppercase; letter-spacing: 0.5px;
  }
  .section-body {
    background: var(--bg-card); border: 1px solid var(--border);
    border-radius: 6px; padding: 2px 10px;
  }
  .sec-row {
    display: flex; justify-content: space-between; align-items: baseline;
    padding: 4px 0; font-size: 11px;
  }
  .sec-row-label { color: var(--text-2); font-size: 10.5px; }
  .sec-row-value {
    color: var(--text-1); font-weight: 700; font-size: 11.5px;
    font-variant-numeric: tabular-nums; white-space: nowrap;
  }
  .sec-row-tr { font-size: 10px; margin-right: 4px; }
  .sec-sep { height: 1px; background: var(--border); margin: 1px 0; }
  .sec-alert {
    background: rgba(212, 147, 74, 0.08);
    border: 1px solid rgba(212, 147, 74, 0.3);
    border-radius: 4px; padding: 5px 8px; margin: 3px 0;
    color: var(--amber); font-size: 11px;
  }

  /* Bandeau historique en bas, compact */
  .historique {
    margin-top: 10px;
    background: var(--bg-panel); border: 1px solid var(--border);
    border-radius: 8px; padding: 10px;
  }
  .hist-title {
    color: var(--gold); font-size: 10px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.6px; margin-bottom: 6px;
  }
  .hist-table {
    display: grid;
    grid-template-columns: 110px repeat(9, 1fr);
    gap: 1px;
    background: var(--border);
    border-radius: 6px; overflow: hidden;
  }
  .hist-cell {
    background: var(--bg-card); padding: 4px 8px;
    font-size: 11px; font-variant-numeric: tabular-nums;
    text-align: right; white-space: nowrap; overflow: hidden;
    text-overflow: ellipsis;
  }
  .hist-cell.label { text-align: left; color: var(--gold); font-weight: 600; }
  .hist-cell.head {
    background: var(--bg-panel); font-size: 9.5px; font-weight: 700;
    color: var(--text-2); text-transform: uppercase; letter-spacing: 0.4px;
  }
  .hist-cell.en-cours-bg { background: #142030; }
  .hist-cell.en-cours-bg.label { color: var(--amber); }
  .hist-cell.en-cours-bg:not(.label) { color: #9DD7FF; }
  .hist-cell.best { background: #1B4A2A; color: #FFF; font-weight: 700; }
  .hist-cell.worst { background: #4A1B1B; color: #FFF; font-weight: 700; }

  footer {
    margin-top: 8px; padding-top: 6px;
    border-top: 1px solid var(--border);
    display: flex; justify-content: space-between;
    color: var(--text-3); font-size: 10px;
  }
  .loading { text-align: center; padding: 40px; color: var(--text-3); }
  .error {
    text-align: center; padding: 16px;
    background: rgba(201, 74, 82, 0.08);
    border: 1px solid rgba(201, 74, 82, 0.3);
    border-radius: 8px; color: var(--red);
    max-width: 600px; margin: 24px auto;
  }
  .toast {
    position: fixed; bottom: 16px; right: 16px;
    background: var(--bg-card); border: 1px solid var(--green);
    color: var(--green); padding: 8px 14px; border-radius: 6px;
    font-size: 12px; opacity: 0; transition: opacity 0.3s;
    pointer-events: none; z-index: 1000;
  }
  .toast.show { opacity: 1; }
  .toast.err { border-color: var(--red); color: var(--red); }

  /* Resserrement à largeur moyenne */
  @media (max-width: 1400px) {
    .main { grid-template-columns: minmax(0, 1fr) 260px 300px; }
    .vignettes { grid-template-columns: 1fr 1fr; }
  }
  @media (max-width: 1000px) {
    .main { grid-template-columns: 1fr; }
    .vignettes { grid-template-columns: 1fr 1fr; }
  }
</style>
</head>
<body>
<div class="topbar">
  <h1>DISTRICARB HUB <span style="font-size:10px;color:var(--text-3);font-weight:400;margin-left:6px;">v0.5</span></h1>
  <div class="right">
    <span class="meta" id="topMeta">B. DISTRICARB SARL — Place d'Armes, Le Lamentin</span>
    <button class="btn" onclick="loadDashboard()">Actualiser</button>
  </div>
</div>
<div id="app" class="loading">Chargement…</div>
<footer><span id="footerLeft">—</span><span id="footerRight">DISTRICARB HUB v0.5</span></footer>
<div id="toast" class="toast"></div>
<script>
function escHtml(s) {
  return String(s).replace(/[&<>"']/g, c => ({
    '&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'
  }[c]));
}
function showToast(msg, err) {
  const t = document.getElementById("toast");
  t.textContent = msg;
  t.className = "toast show" + (err ? " err" : "");
  setTimeout(() => { t.className = "toast"; }, 2500);
}
async function loadDashboard() {
  try {
    const r = await fetch("/api/dashboard");
    if (!r.ok) throw new Error("HTTP " + r.status);
    const data = await r.json();
    render(data);
  } catch (e) {
    document.getElementById("app").outerHTML =
      '<div id="app" class="error">Erreur : ' + escHtml(e.message) + '</div>';
  }
}
async function resolveEvent(evtId) {
  if (!confirm("Marquer cet événement comme réglé ?")) return;
  try {
    const r = await fetch("/api/event/resolve", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({evt_id: evtId})
    });
    const res = await r.json();
    if (res.ok) {
      showToast("Événement marqué comme réglé");
      setTimeout(loadDashboard, 600);
    } else {
      showToast("Erreur : " + (res.error || "inconnue"), true);
    }
  } catch (e) {
    showToast("Erreur réseau : " + e.message, true);
  }
}
function renderVignettes(vignettes) {
  let html = '<div class="vignettes">';
  for (const v of vignettes) {
    html += '<div class="card" data-status="'+escHtml(v.status||"")+'">';
    html +=   '<div class="card-head"><div class="card-title">'+escHtml(v.label)+'</div>';
    html +=   '<span class="card-dot '+escHtml(v.status||"")+'"></span></div>';
    html +=   '<div class="card-sub">'+escHtml(v.sub)+'</div>';
    if (!v.lines || v.lines.length === 0) {
      html += '<div style="color:var(--text-3); font-size:11px; padding:4px 0; font-style:italic;">Aucune donnée</div>';
    } else {
      for (const line of v.lines) {
        const style = line.color ? ' style="color: '+escHtml(line.color)+'"' : '';
        html += '<div class="row"><span class="row-label">'+escHtml(line.label)+'</span>';
        html +=   '<span class="row-value"'+style+'>'+escHtml(line.value)+'</span></div>';
      }
    }
    html += '</div>';
  }
  html += '</div>';
  return html;
}
function renderAlertesActives(events) {
  let html = '<aside class="alerts-panel">';
  html +=   '<div class="panel-title">Alertes actives <span class="count">'+(events?events.length:0)+'</span></div>';
  if (!events || events.length === 0) {
    html += '<div class="empty-state">Rien à signaler ✓</div>';
  } else {
    for (const e of events) {
      const cls = escHtml(e.type || "");
      const statutCls = escHtml(e.statut || "non_traite");
      const statutLabel = ({
        "non_traite": "À traiter",
        "snooze": "Snoozé",
        "ack": "Pris en compte"
      })[e.statut] || e.statut;
      html += '<div class="alert-item '+cls+'">';
      html +=   '<div class="alert-text">'+escHtml(e.label || e.type)+'</div>';
      html +=   '<div class="alert-actions">';
      html +=     '<span class="alert-statut '+statutCls+'">'+escHtml(statutLabel)+'</span>';
      html +=     '<button class="btn-resolve" onclick="resolveEvent('+JSON.stringify(e.id)+')">✓ Régler</button>';
      html +=   '</div>';
      html += '</div>';
    }
  }
  html += '</aside>';
  return html;
}
function renderPilotage(p) {
  if (!p) return '';
  let html = '<aside class="pilotage">';
  html +=   '<div class="pilotage-head"><div class="panel-title gold" style="margin-bottom:4px;">Pilotage Jour</div>';
  html +=   '<div class="pilotage-date">'+escHtml(p.date||"—")+'</div>';
  html +=   '<div class="pilotage-sub">'+escHtml(p.sub||"")+'</div></div>';
  for (const sec of (p.sections||[])) {
    if (!sec.rows || sec.rows.length === 0) continue;
    html += '<div class="section">';
    html +=   '<div class="section-head"><span class="section-bar accent-'+escHtml(sec.accent||"red")+'"></span>';
    html +=   '<span class="section-title">'+escHtml(sec.title)+'</span></div>';
    html +=   '<div class="section-body">';
    for (const row of sec.rows) {
      if (row.kind === "sep") {
        html += '<div class="sec-sep"></div>';
      } else if (row.kind === "alert") {
        const acol = row.color ? ' style="color: '+escHtml(row.color)+'"' : '';
        html += '<div class="sec-alert"'+acol+'>'+escHtml(row.text)+'</div>';
      } else {
        const vstyle = row.color ? ' style="color: '+escHtml(row.color)+'"' : '';
        const tr = row.tr_arrow ? '<span class="sec-row-tr" style="color:'+escHtml(row.tr_color||"")+'">'+escHtml(row.tr_arrow)+'</span>' : '';
        html += '<div class="sec-row"><span class="sec-row-label">'+escHtml(row.label)+'</span>';
        html +=   '<span>'+tr+'<span class="sec-row-value"'+vstyle+'>'+escHtml(row.value)+'</span></span></div>';
      }
    }
    html +=   '</div>';
    html += '</div>';
  }
  html += '</aside>';
  return html;
}
function renderHistorique(h) {
  if (!h || !h.rows || h.rows.length === 0) return '';
  let html = '<section class="historique">';
  html +=   '<div class="hist-title">Historique des derniers jours</div>';
  html +=   '<div class="hist-table">';
  for (let i=0; i<h.cols.length; i++) {
    const align = i===0 ? 'text-align:left' : '';
    html += '<div class="hist-cell head" style="'+align+'">'+escHtml(h.cols[i])+'</div>';
  }
  for (const row of h.rows) {
    for (const c of (row.cells||[])) {
      let cls = 'hist-cell';
      if (c.is_label) cls += ' label';
      if (row.is_en_cours) cls += ' en-cours-bg';
      if (c.is_best) cls += ' best';
      if (c.is_worst) cls += ' worst';
      html += '<div class="'+cls+'">'+escHtml(c.value)+'</div>';
    }
  }
  html +=   '</div>';
  html += '</section>';
  return html;
}
function render(data) {
  const app = document.getElementById("app");
  app.className = "";
  let html = '<div class="main">';
  html +=     renderVignettes(data.vignettes||[]);
  html +=     renderAlertesActives(data.events_actifs||[]);
  html +=     renderPilotage(data.pilotage);
  html +=   '</div>';
  html += renderHistorique(data.historique);
  app.innerHTML = html;
  document.getElementById("topMeta").textContent =
    "B. DISTRICARB SARL — Place d'Armes, Le Lamentin · "
    + new Date(data._meta.refresh_ts).toLocaleString("fr-FR");
  document.getElementById("footerLeft").textContent = data.footer || "—";
}
loadDashboard();
setInterval(loadDashboard, 60000);
</script>
</body>
</html>
"""

class ThreadingHTTPServer(ThreadingMixIn,HTTPServer):
    """Serveur HTTP multi-thread pour gérer plusieurs requêtes en parallèle.
    daemon_threads=True : les threads sont tués quand le hub se ferme."""
    daemon_threads=True
    allow_reuse_address=True

class HubHTTPHandler(BaseHTTPRequestHandler):
    """Handler HTTP minimaliste pour exposer les données du hub.
    Routes :
      - GET /              → page HTML dashboard
      - GET /api/dashboard → JSON des données last_data
      - POST /api/event/resolve → marque un événement comme résolu (body JSON {evt_id})
    """
    def log_message(self,format,*args):
        # Désactive les logs verbeux par défaut (sinon console pleine)
        pass
    def do_GET(self):
        try:
            if self.path=="/" or self.path.startswith("/?") or self.path=="/index.html":
                self._send_html(_DASHBOARD_HTML)
                return
            if self.path=="/api/dashboard":
                self._send_dashboard_json()
                return
            self.send_error(404,"Not found")
        except Exception as e:
            try: self.send_error(500,f"Internal error: {e}")
            except Exception: pass
    def do_POST(self):
        try:
            if self.path=="/api/event/resolve":
                self._handle_resolve_event()
                return
            self.send_error(404,"Not found")
        except Exception as e:
            try: self.send_error(500,f"Internal error: {e}")
            except Exception: pass
    def _handle_resolve_event(self):
        """Marque un événement comme résolu. Reproduit la logique de _marquer_resolu_evt
        sans la partie 'silence alerte' (suffisant pour l'action web)."""
        hub=_HUB_INSTANCE_REF
        if hub is None:
            self._send_json({"ok":False,"error":"Hub non initialisé"},status=503)
            return
        try:
            length=int(self.headers.get("Content-Length","0"))
            raw=self.rfile.read(length) if length>0 else b"{}"
            body=json.loads(raw.decode("utf-8") or "{}")
            evt_id=body.get("evt_id","")
            if not evt_id:
                self._send_json({"ok":False,"error":"evt_id manquant"},status=400)
                return
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[])
            found=False
            for e in events:
                if e.get("id")==evt_id:
                    data=e.setdefault("data",{})
                    data["statut"]="resolu";data["lu"]=True
                    old_comm=e.get("commentaire","")
                    ts_str=datetime.now().strftime("%d/%m/%Y %Hh%M")
                    trace=f"[{ts_str}] \u2713 Marqu\u00e9 r\u00e9solu depuis le web"
                    e["commentaire"]=f"{old_comm}\n{trace}" if old_comm else trace
                    found=True;break
            if not found:
                self._send_json({"ok":False,"error":"Événement introuvable"},status=404)
                return
            all_evt["events"]=events
            save_json(EVENEMENTS_FILE,all_evt)
            # Demander au hub de se rafraîchir pour répercuter immédiatement la modif
            try: hub.after(0,hub.refresh)
            except Exception: pass
            self._send_json({"ok":True})
        except Exception as e:
            self._send_json({"ok":False,"error":f"{type(e).__name__}: {e}"},status=500)
    def _send_html(self,html):
        body=html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type","text/html; charset=utf-8")
        self.send_header("Content-Length",str(len(body)))
        self.send_header("Cache-Control","no-cache")
        self.end_headers()
        self.wfile.write(body)
    def _send_dashboard_json(self):
        hub=_HUB_INSTANCE_REF
        if hub is None:
            self._send_json({"_error":"Hub non initialisé"},status=503)
            return
        # ---- Vignettes principales (5 vignettes du panneau gauche)
        vignettes=[]
        vig_meta=[
            ("gest_piste","GEST PISTE","Boutique & caisse","red"),
            ("cartes","CARTES","Télécollectes CB/CP","blue"),
            ("prevision","PRÉVISION","Stocks & commandes","gold"),
            ("objectif","OBJECTIF","Alertes & pilotage","green"),
            ("litrage","LITRAGE","Performance & historique","teal"),
        ]
        for key,label,sub,base_color in vig_meta:
            vig=hub.vigs.get(key)
            if vig is None:
                vignettes.append({"key":key,"label":label,"sub":sub,"base_color":base_color,
                                  "status":"unknown","lines":[]})
                continue
            lines_raw=getattr(vig,"_last_lines",[]) or []
            status=getattr(vig,"_last_status","unknown")
            lines=[{"label":str(l[0]),"value":str(l[1]),"color":str(l[2]) if l[2] else None}
                   for l in lines_raw]
            vignettes.append({"key":key,"label":label,"sub":sub,"base_color":base_color,
                              "status":status,"lines":lines})
        # ---- Panneau Pilotage Jour (droite)
        sections_pilotage=[]
        for attr_name,base_color in [("s_carb","red"),("s_enc","blue"),("s_bout","amber"),
                                       ("s_stock","gold"),("s_obj","green"),("s_alert","red")]:
            sec=getattr(hub,attr_name,None)
            if sec is None: continue
            sections_pilotage.append({
                "title":getattr(sec,"_title",""),
                "accent":base_color,
                "rows":getattr(sec,"_rows",[]) or [],
            })
        # ---- Historique tableau du bas
        hist_rows=getattr(hub,"_hist_rows",[]) or []
        # ---- Événements actifs du journal (non résolus, non masqués)
        events_actifs=[]
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            for e in (all_evt.get("events",[]) or []):
                data=e.get("data",{}) or {}
                if data.get("statut","")=="resolu": continue
                if data.get("masque",False): continue
                t=e.get("type","")
                if t in ("livraison","commande","passage_mois","ack"): continue
                # Libellé court via _situation_label si dispo
                try: label=hub._situation_label(e)
                except Exception: label=t
                events_actifs.append({
                    "id":e.get("id",""),
                    "type":t,
                    "label":label,
                    "statut":data.get("statut","non_traite"),
                    "ts":e.get("ts",""),
                })
            # Tri par timestamp décroissant, top 12
            events_actifs.sort(key=lambda x:x.get("ts",""),reverse=True)
            events_actifs=events_actifs[:12]
        except Exception: events_actifs=[]
        # ---- Bandeaux & footer
        pilotage_date=getattr(hub,"_b_date_text","—")
        pilotage_sub=getattr(hub,"_b_sub_text","—")
        footer_text=getattr(hub,"_ft_lbl_text","")
        payload={
            "_meta":{
                "refresh_ts":datetime.now().isoformat(),
                "station":"B. DISTRICARB SARL",
                "version":"v0.5",
            },
            "vignettes":vignettes,
            "pilotage":{
                "date":pilotage_date,
                "sub":pilotage_sub,
                "sections":sections_pilotage,
            },
            "historique":{
                "cols":["Jour","SP","GO","GNR","Total L","CA Piste","CB","CP","Boutique","Total"],
                "rows":hist_rows,
            },
            "events_actifs":events_actifs,
            "footer":footer_text,
        }
        self._send_json(payload)
    def _send_json(self,obj,status=200):
        try:
            body=json.dumps(obj,ensure_ascii=False).encode("utf-8")
        except Exception as e:
            body=json.dumps({"_error":f"Serialization failed: {e}"},ensure_ascii=False).encode("utf-8")
            status=500
        self.send_response(status)
        self.send_header("Content-Type","application/json; charset=utf-8")
        self.send_header("Content-Length",str(len(body)))
        self.send_header("Access-Control-Allow-Origin","*")  # CORS pour dev local
        self.send_header("Cache-Control","no-cache")
        self.end_headers()
        self.wfile.write(body)

def start_web_server(hub,port=8765):
    """Démarre le serveur HTTP dans un thread daemon. Non bloquant.
    Le serveur écoute sur toutes les interfaces (0.0.0.0) pour permettre
    l'accès depuis le réseau local en plus de localhost.
    
    Logs explicites dans ~/.districarb_hub/web_server.log pour diagnostic
    (les print() peuvent être avalés selon le mode de lancement Windows).
    """
    global _HUB_INSTANCE_REF
    _HUB_INSTANCE_REF=hub
    log_path=APP_DIR/"web_server.log"
    def _log(msg):
        try:
            with open(log_path,"a",encoding="utf-8") as f:
                f.write(f"[{datetime.now().isoformat()}] {msg}\n")
        except Exception: pass
        try: print(f"[web] {msg}")
        except Exception: pass
    def _run():
        try:
            _log(f"Démarrage du serveur HTTP sur 0.0.0.0:{port}…")
            srv=ThreadingHTTPServer(("0.0.0.0",port),HubHTTPHandler)
            _log(f"✓ Serveur prêt — http://localhost:{port}/ (et http://<IP-LAN>:{port}/)")
            srv.serve_forever()
        except OSError as e:
            # Port déjà utilisé OU permissions refusées
            _log(f"✗ OSError au démarrage : {e}")
            _log(f"  → Le port {port} est peut-être déjà pris. Essaie un autre port ou ferme l'app qui l'utilise.")
        except Exception as e:
            _log(f"✗ Exception inattendue : {type(e).__name__}: {e}")
            _log(f"  → Traceback : {traceback.format_exc()}")
    t=threading.Thread(target=_run,daemon=True,name="HubWebServer")
    t.start()
    _log(f"Thread serveur lancé (id={t.ident}). Log : {log_path}")
    return t


# =============================================================================
class Hub(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.withdraw()  # Cacher pendant le chargement
        # Détection police Space Grotesk (après création racine Tk, avant build UI)
        _detect_num_font()
        self.title("DISTRICARB HUB \u2014 B. DISTRICARB SARL")
        self.geometry("1440x900");self.minsize(1200,780);self.configure(fg_color=C["bg"])
        self.protocol("WM_DELETE_WINDOW",self._on_close)
        self.cfg=load_json(CONFIG_FILE);self.reader=DataReader(self.cfg);self.vigs={};self.last_data={}
        try: start_web_server(self,port=8765)
        except Exception as _e: print(f"[web] Init failed: {_e}")
        self.alert_shown={"1030":False,"1100":False}
        # Splash screen
        splash=ctk.CTkToplevel(self);splash.overrideredirect(True)
        sw,sh=520,280;splash.geometry(f"{sw}x{sh}+{(splash.winfo_screenwidth()-sw)//2}+{(splash.winfo_screenheight()-sh)//2}")
        splash.configure(fg_color=C["bg"])
        ctk.CTkLabel(splash,text="\u25cf",font=("Segoe UI",36),text_color=C["red"]).pack(pady=(30,0))
        ctk.CTkLabel(splash,text="DISTRICARB HUB",font=("Segoe UI",26,"bold"),text_color=C["t1"]).pack(pady=(6,0))
        ctk.CTkLabel(splash,text="B. DISTRICARB SARL \u2014 Place d'Armes",font=("Segoe UI",11),text_color=C["t3"]).pack(pady=(4,0))
        sp_status=ctk.CTkLabel(splash,text="Chargement...",font=("Segoe UI",11),text_color=C["gold"])
        sp_status.pack(pady=(20,6))
        sp_bar=ctk.CTkProgressBar(splash,width=320,height=5,fg_color=C["border"],progress_color=C["red"])
        sp_bar.pack();sp_bar.set(0);splash.update()
        # Build
        sp_status.configure(text="Construction de l'interface...");sp_bar.set(0.2);splash.update()
        self._build()
        sp_status.configure(text="V\u00e9rification des fichiers...");sp_bar.set(0.4);splash.update()
        self._first_check()
        # Reconstitution unique du journal 2026 (ruptures auto + livraisons depuis Achat_carburant.xlsx)
        # Ne s'exécute que la première fois (drapeau dans journal_reconstitue.cfg)
        sp_status.configure(text="V\u00e9rification du journal d'\u00e9v\u00e9nements...");sp_bar.set(0.5);splash.update()
        try:
            reconstitute_journal_2026()
        except Exception as e: print(f"[reconstitute journal] {e}")
        sp_status.configure(text="Lecture des donn\u00e9es...");sp_bar.set(0.6);splash.update()
        self.refresh()
        sp_status.configure(text="Pr\u00eat !");sp_bar.set(1.0);splash.update()
        splash.after(700,splash.destroy)
        self.after(800,self.deiconify)
        # Post-splash : vérification cycle/livraison du jour
        # NB : la même méthode est rappelée périodiquement dans _check_time_alerts pour que
        # ces popups apparaissent aussi quand le hub tourne H24 (pas seulement au démarrage).
        self.after(1200,self._check_cycle_and_livraison)
        self._refresh_timer=self.after(REFRESH_MS,self._loop)
        self.after(60000,self._check_time_alerts)
        # Suggestion automatique du rapport mensuel au premier lancement du mois
        self.after(4000,self._check_rapport_suggestion)
        # ============================================================
        # FIX FENÊTRE INVISIBLE (introduit 11/05/2026)
        # Bug intermittent : le hub se retrouve minimisé/invisible et impossible à
        # ramener au premier plan (uniquement via Gestionnaire de tâches → Fin de tâche).
        # 3 niveaux complémentaires :
        #   N1 - Raccourci Ctrl+Alt+H : force la restauration de la fenêtre principale.
        #        Marche globalement, depuis n'importe où dans le hub.
        #   N2 - Détection d'état suspect (chaque 60s) : si la fenêtre est iconic SANS
        #        popup ouverte (état anormal), trace dans errors.log. On N'AUTO-RESTAURE
        #        PAS — Bidou peut avoir minimisé volontairement, on ne va pas le contrarier.
        #   N3 - Tracing des transitions Map/Unmap/Visibility/Focus dans errors.log.
        #        La prochaine fois que le bug arrive, on aura l'historique exact pour
        #        comprendre la cause (quelle popup était ouverte juste avant, etc.).
        # ============================================================
        self.bind_all("<Control-Alt-h>",lambda e:self._emergency_restore_window())
        self.bind_all("<Control-Alt-H>",lambda e:self._emergency_restore_window())
        # FIX 21/05/2026 (Bug C) : on filtre `event.widget is self` pour ne logger
        # QUE les transitions de la fenêtre principale, pas celles des widgets enfants.
        # Avant : <Unmap> sur la fenêtre principale se propageait par bubbling à TOUS les
        # widgets enfants (centaines de Frames/Labels/Boutons). Chaque widget déclenchait
        # _log_window_transition → des centaines d'écritures dans errors.log en moins
        # d'une seconde à chaque minimisation. Cause majeure du fichier de 99 Mo.
        self.bind("<Map>",lambda e: (e.widget is self) and self._log_window_transition("Map"))
        self.bind("<Unmap>",lambda e: (e.widget is self) and self._log_window_transition("Unmap"))
        self.bind("<FocusIn>",lambda e: (e.widget is self) and self._log_window_transition("FocusIn"))
        self.bind("<FocusOut>",lambda e: (e.widget is self) and self._log_window_transition("FocusOut"))
        # Démarrer la surveillance santé fenêtre 90s après le launch (le temps que tout soit stable)
        self.after(90000,self._check_window_health)
        # Surveillance fichier flag RESTORE.flag : mécanisme de récupération indépendant
        # de l'état tkinter. Quand Ctrl+Alt+H tkinter ne marche pas (focus perdu hors hub),
        # un raccourci Windows natif peut créer ce fichier flag → hub le détecte et restaure.
        # Voir restore_hub.bat livré avec le hub.
        self.after(2000,self._check_restore_flag)

    # ============================================================
    # FIX FENÊTRE INVISIBLE — méthodes des 3 niveaux
    # ============================================================
    def _emergency_restore_window(self,event=None):
        """N1 — Restauration forcée de la fenêtre principale au premier plan.
        Déclenchée par Ctrl+Alt+H globalement. Combine plusieurs techniques :
          - deiconify() : sort d'un état iconic/withdrawn
          - lift() : remonte la fenêtre dans le z-order
          - attributes('-topmost', True/False) : force temporairement au premier plan
            puis re-désactive (sinon la fenêtre resterait toujours devant les autres)
          - focus_force() : récupère le focus clavier
        """
        try:
            self.deiconify()
            self.lift()
            self.attributes("-topmost",True)
            self.update_idletasks()
            self.focus_force()
            # Re-désactiver topmost après 200ms pour ne pas figer la fenêtre devant les autres
            self.after(200,lambda:self.attributes("-topmost",False))
            self._log_window_transition("EMERGENCY_RESTORE (Ctrl+Alt+H)")
        except Exception as _e: _log_silent_err(exc=_e)

    def _check_window_health(self):
        """N2 — Surveillance périodique de l'état de la fenêtre.
        Détecte un état suspect (iconic sans aucune popup ouverte → potentiellement
        le bug fenêtre invisible). On NE FAIT QUE TRACER : pas d'auto-restauration
        agressive (Bidou peut avoir minimisé volontairement le hub pour faire autre
        chose, on ne va pas le contrarier). La trace permet de comprendre la cause
        au prochain incident.
        Reprogramme la prochaine vérification dans 60s.
        """
        try:
            try: state=self.state()
            except Exception: state="?"
            if state=="iconic":
                # Y a-t-il une popup Toplevel ouverte ?
                has_toplevel_visible=False
                try:
                    for w in self.winfo_children():
                        if isinstance(w,ctk.CTkToplevel) and w.winfo_exists() and w.winfo_viewable():
                            has_toplevel_visible=True;break
                except Exception: pass
                if not has_toplevel_visible:
                    # État suspect : iconic sans aucune popup → potentiellement le bug
                    self._log_window_transition("SUSPECT_iconic_no_popup")
        except Exception as _e: _log_silent_err(exc=_e)
        # Reprogrammer
        try: self.after(60000,self._check_window_health)
        except Exception as _e: _log_silent_err(exc=_e)

    def _log_window_transition(self,event_type):
        """N3 — Trace une transition d'état de la fenêtre principale dans errors.log.
        Aide à diagnostiquer les bugs d'état de fenêtre.
        
        FIX 21/05/2026 : dédoublonnage temporel — si le MÊME event_type a été loggué
        dans la dernière seconde, on skip. Évite que des rafales d'événements tkinter
        (notamment Unmap déclenché plusieurs fois en cascade lors d'une minimisation)
        n'inondent le fichier. Couplé au filtre event.widget is self (au bind),
        ramène le volume de log de ~1000 lignes/min à quelques lignes par jour.
        """
        try:
            # Dédoublonnage 1 sec par event_type
            now=datetime.now()
            last=getattr(self,"_last_window_log",{}) or {}
            prev=last.get(event_type)
            if prev is not None and (now-prev).total_seconds()<1.0:
                return  # spam : skip silencieusement
            last[event_type]=now
            self._last_window_log=last
            _init_errors_log()
            try: state=self.state()
            except Exception: state="?"
            # Compter les popups visibles pour contexte
            popup_count=0
            try:
                for w in self.winfo_children():
                    if isinstance(w,ctk.CTkToplevel) and w.winfo_exists() and w.winfo_viewable():
                        popup_count+=1
            except Exception: pass
            msg=f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] [window] {event_type} - state={state} popups={popup_count}\n"
            try:
                with open(_ERRORS_LOG_PATH,"a",encoding="utf-8") as f:
                    f.write(msg)
            except Exception: pass
        except Exception: pass  # ne JAMAIS casser le flux

    def _check_restore_flag(self):
        """Polling toutes les 2 secondes du fichier RESTORE.flag à côté du hub.
        Mécanisme de récupération indépendant de l'état tkinter : quand le bug
        'fenêtre invisible' frappe, Ctrl+Alt+H (bind_all tkinter) ne marche pas
        parce que tkinter n'a plus le focus Windows. Mais le polling `after` continue
        de tourner tant que Python tourne (vérifiable via Gestionnaire de tâches).
        
        Si RESTORE.flag existe → le supprime + restaure la fenêtre. L'utilisateur
        crée ce fichier via un raccourci Windows natif (restore_hub.bat ou un .lnk
        avec touche de raccourci globale assignée).
        """
        try:
            flag_path=Path(__file__).parent/"RESTORE.flag"
            if flag_path.exists():
                try: flag_path.unlink()
                except Exception as _e: _log_silent_err(exc=_e)
                self._emergency_restore_window()
                self._log_window_transition("RESTORE_FLAG_TRIGGERED")
        except Exception as _e: _log_silent_err(exc=_e)
        # Reprogrammer (2 secondes pour réactivité acceptable, charge CPU négligeable)
        try: self.after(2000,self._check_restore_flag)
        except Exception as _e: _log_silent_err(exc=_e)

    def _ask_cycle(self):
        """Demande à l'utilisateur quelle semaine du cycle on est."""
        # Garde-fou anti-empilement (lue par _check_cycle_and_livraison)
        self._cycle_dlg_open=True
        dlg=ctk.CTkToplevel(self)
        dlg.title("Cycle 14 jours")
        w,h=480,300;sw=dlg.winfo_screenwidth();sh=dlg.winfo_screenheight()
        dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        dlg.configure(fg_color=C["bg"]);dlg.resizable(False,False)
        dlg.transient(self);dlg.grab_set()
        # Reset le flag quand la popup se ferme (par bouton OU par croix X)
        def _reset_flag():
            self._cycle_dlg_open=False
        dlg.protocol("WM_DELETE_WINDOW",lambda:(_reset_flag(),dlg.destroy()))
        ctk.CTkLabel(dlg,text="\U0001f504  Cycle de pr\u00e9vision",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(pady=(24,4))
        ctk.CTkLabel(dlg,text="Quelle semaine du cycle de 14 jours sommes-nous ?",font=("Segoe UI",12),text_color=C["t2"]).pack(pady=(0,4))
        # Suggestion automatique basée sur le cycle existant
        suggested=get_cycle_week()
        if suggested:
            ctk.CTkLabel(dlg,text=f"(Suggestion automatique : Semaine {suggested})",font=("Segoe UI",10,"italic"),text_color=C["gold"]).pack(pady=(0,16))
        else:
            ctk.CTkLabel(dlg,text="(Premier lancement, pas de suggestion)",font=("Segoe UI",10,"italic"),text_color=C["t3"]).pack(pady=(0,16))
        info=ctk.CTkFrame(dlg,fg_color=C["card"],corner_radius=8);info.pack(fill="x",padx=30,pady=(0,16))
        ctk.CTkLabel(info,text="Semaine 1 = onglets Lundi, Mardi, ...\nSemaine 2 = onglets Lundi2, Mardi2, ...",
                     font=("Segoe UI",10),text_color=C["t2"],justify="left").pack(padx=14,pady=10)
        def choose(w):
            set_cycle_week(w)
            cycle=load_json(CYCLE_FILE)
            cycle["last_asked"]=date.today().strftime("%Y-%m-%d")
            save_json(CYCLE_FILE,cycle)
            _reset_flag()
            dlg.destroy()
            self.refresh()
            # Enchaîner avec la livraison si pas encore demandée (sauf weekend ou férié)
            today_str=date.today().strftime("%d/%m/%y")
            if today_str not in load_json(LIVRAISON_FILE) and date.today().weekday()<5 and not is_ferie(date.today()):
                self.after(300,self._ask_livraison)
        btns=ctk.CTkFrame(dlg,fg_color="transparent");btns.pack(fill="x",padx=30,pady=(0,20))
        ctk.CTkButton(btns,text="Semaine 1",width=180,height=44,fg_color=C["blue"],hover_color="#3A82E0",
                      text_color="#FFF",font=("Segoe UI",13,"bold"),corner_radius=8,
                      command=lambda:choose(1)).pack(side="left",padx=(0,8))
        ctk.CTkButton(btns,text="Semaine 2",width=180,height=44,fg_color=C["gold"],hover_color="#D9A02E",
                      text_color="#000",font=("Segoe UI",13,"bold"),corner_radius=8,
                      command=lambda:choose(2)).pack(side="right")

    def _on_close(self):
        if messagebox.askokcancel("Fermer","Voulez-vous fermer DISTRICARB HUB ?"):
            try: shutil.rmtree(TEMP_DIR,ignore_errors=True)
            except Exception as _e: _log_silent_err(exc=_e)
            self.destroy()
    def _ask_livraison(self):
        # Garde-fou anti-empilement (lue par _check_cycle_and_livraison)
        self._livr_dlg_open=True
        try:
            # ANTI-EMPILEMENT BLINDÉ (signalé Bidou 21/05/2026 7h14 : 4 LivraisonDialog
            # empilées vides — le flag _livr_dlg_open seul n'a pas suffi, cause racine
            # non identifiée encore). On scanne les enfants top-level et on détruit
            # toute LivraisonDialog existante AVANT d'en créer une nouvelle. Même
            # pattern que celui qu'on a appliqué hier aux autres popups (winfo_children
            # + isinstance) — référence-attribut pas fiable, énumération si.
            try:
                for w in list(self.winfo_children()):
                    if isinstance(w,LivraisonDialog):
                        try: w.destroy()
                        except Exception: pass
            except Exception as _e: _log_silent_err(exc=_e)
            dlg=LivraisonDialog(self);self.wait_window(dlg)
            if dlg.result: self.refresh()
        finally:
            self._livr_dlg_open=False
    def _ask_livraison_attendue(self,cmd):
        """Brique 2 — popup DÉDIÉ 'Livraison attendue' piloté par commandes.cfg.
        Ne touche pas LivraisonDialog. 'Oui' → flux d'ajout existant. 'Pas
        encore' → silence du jour ; l'alerte calibrée tombe à l'heure limite du
        tour (gérée par _check_livraison_attendue_escalade)."""
        self._livr_dlg_open=True
        try:
            dlg=LivraisonAttendueDlg(self,cmd);self.wait_window(dlg)
        finally:
            self._livr_dlg_open=False
    def _check_livraison_attendue_escalade(self,now):
        """Brique 2 — l'alerte CALIBRÉE sur le tour (= le 'réveil' du doc). Si une
        commande pilote l'attendu d'aujourd'hui, qu'aucune livraison n'est saisie
        et que l'heure LIMITE du tour visé est dépassée → alerte rouge
        actionnable, une seule fois par jour."""
        try:
            today=now.date()
            cmd=get_commande(today)
            if not cmd: return  # pas de commande pour aujourd'hui → rien à escalader
            # Pas de filtre weekday/férié ici : si une commande vise EXPLICITEMENT
            # aujourd'hui (ex : samedi forcé déjà résolu via l'alerte Prévision),
            # l'attendu est réel. On suit la décision déjà prise.
            livrs=load_json(LIVRAISON_FILE) or {}
            if today.strftime("%d/%m/%y") in livrs:
                return  # livraison saisie → camion arrivé, rien à escalader
            tour=int(cmd.get("tour",1) or 1)
            info=TOURS_LIVRAISON.get(tour,TOURS_LIVRAISON[1])
            if now.hour < info["alerte"]:
                return  # pas encore l'heure limite du tour visé
            if getattr(self,"_escalade_attendue_day",None)==today:
                return  # déjà escaladé aujourd'hui
            self._escalade_attendue_day=today
            self._show_toast(
                f"\u26a0 {info['label']} d\u00e9pass\u00e9 ({info['alerte']}h) \u2014 "
                f"livraison attendue non arriv\u00e9e \u2014 appelle TotalEnergies maintenant",
                C["red"])
            # Trace durable dans le centre de notifications (validée Bidou 02/06/2026).
            # Le toast ci-dessus s'évapore en quelques secondes : si Bidou est absent au
            # moment du tour, il rate l'info et rien ne reste. On enregistre donc aussi
            # un événement non_traite qui atterrit dans le Tableau de notifications, qu'il
            # retrouve à son retour et écarte manuellement (geste Résolu). Idempotence :
            # 1 par jour (fingerprint livratt:date), aucun doublon.
            try:
                add_evenement("livraison_attendue",{
                    "date":today.strftime("%Y-%m-%d"),
                    "tour":tour,
                    "tour_label":info["label"],
                    "heure_limite":info["alerte"],
                    "statut":"non_traite",
                    "lu":False,
                },commentaire=f"{info['label']} d\u00e9pass\u00e9 ({info['alerte']}h) \u2014 livraison attendue non confirm\u00e9e arriv\u00e9e \u2014 appeler TotalEnergies")
            except Exception as _e: _log_silent_err(exc=_e)
        except Exception as _e: _log_silent_err(exc=_e)
    def _open_livraisons(self):
        # Récupère le chemin de Prévision compte depuis la config
        cfg=load_json(CONFIG_FILE)
        prev_path=cfg.get("prevision","")
        dlg=LivraisonsHistDlg(self,prevision_path=prev_path if prev_path else None)
        self.wait_window(dlg)
    def _open_rapport_menu(self):
        """Ouvre la dialogue de génération de rapport mensuel.
        L'utilisateur choisit un mois (par défaut : mois précédent), génère et le hub
        ouvre le PDF/HTML produit dans l'application système."""
        dlg=RapportMensuelDlg(self)
        self.wait_window(dlg)
    def _open_journal_evenements(self):
        """Ouvre le journal des événements (Sujet E).
        Affiche les ponts traversés, anomalies confirmées, ruptures et livraisons
        archivés au fil du temps. Permet de filtrer par période et par type."""
        dlg=JournalEvenementsDlg(self)
        self.wait_window(dlg)

    def _open_alertes_dashboard(self):
        """Ouvre le tableau de notifications : toutes les alertes actives en un coup d'œil
        (livraisons à reporter, saisies impossibles, ponts non-acquittés, anomalies, ruptures
        projetées, événements `non_traite` du journal). Vue todo-list opérationnelle."""
        dlg=AlertesDashboardDlg(self)
        self.wait_window(dlg)

    def _count_alertes_actives(self):
        """Compte les alertes actives toutes catégories confondues. Source unique du nombre
        affiché dans le badge ET dans le tableau (pour cohérence)."""
        n=0
        ar=(self.last_data or {}).get("antirupture",{}) or {}
        # Anti-rupture : ponts non-acquittés
        try:
            ack_status=ar.get("ack_status",{}) or {}
            n+=sum(1 for st in ack_status.values() if not st.get("acquitte"))
        except Exception as _e: _log_silent_err(exc=_e)
        # Saisies physiquement impossibles (à corriger dans Pre_vision)
        n+=len(ar.get("saisies_physiquement_impossibles",[]) or [])
        # Livraisons à reporter (capacité cuve dépassée)
        n+=len(ar.get("livraisons_a_reporter",[]) or [])
        # Incohérences jour non-livrable (ex : livraison saisie sur weekend)
        n+=len(ar.get("incoherences_jour_non_livrable",[]) or [])
        # Ruptures projetées
        n+=len(ar.get("ruptures_projetees",[]) or [])
        # Événements `non_traite` du journal (pastille rouge non lue)
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            for evt in all_evt.get("events",[]):
                d=evt.get("data",{}) or {}
                if d.get("statut")=="non_traite" and not d.get("lu",True):
                    n+=1
        except Exception as _e: _log_silent_err(exc=_e)
        return n

    def _refresh_alertes_badge(self):
        """Met à jour le badge numérique du bouton Alertes 🔔. Caché si 0, sinon affiché
        en superposition sur le cercle d'icône (place avec relx=1.0, rely=0.0)."""
        if not self.alertes_badge: return
        try:
            count=self._count_alertes_actives()
            if count>0:
                txt=str(count) if count<10 else "9+"
                self.alertes_badge.configure(text=txt)
                # place sur le coin supérieur droit du cercle parent
                self.alertes_badge.place(relx=1.0,rely=0.0,anchor="ne",x=4,y=-4)
            else:
                self.alertes_badge.place_forget()
        except Exception as e: print(f"[badge update] {e}")

    def _open_prix_carburant(self):
        """Ouvre la fenêtre Prix carburant : PV, PA et marge unitaire SP/GO/GNR
        pour le mois en cours. Évite d'avoir à ouvrir Prévision compte.xlsx."""
        dlg=PrixCarburantDlg(self)
        self.wait_window(dlg)

    def _open_passage_mois(self,target_year=None,target_month=None,edit_mode=False):
        """Ouvre la fenêtre passage de mois (saisie ventes 0h-6h + calcul effet
        spéculation). Utilisable en mode auto (popup début de mois) ou en mode
        modif (depuis bouton dans fenêtre Prix)."""
        dlg=PassageMoisDlg(self,target_year=target_year,target_month=target_month,edit_mode=edit_mode)
        self.wait_window(dlg)

    def _maybe_open_passage_mois_dlg(self):
        """Ouvre PassageMoisDlg automatiquement si :
          - On est le 1er du mois ou plus tard, et il est ≥ 6h00
          - Le passage du mois en cours n'a pas encore été saisi
          - Les prix du mois courant ET précédent sont dispo (sinon calcul impossible)
          - Le snooze utilisateur (s'il existe) est expiré

        Le snooze est persistant : stocké dans ~/.districarb_hub/passage_mois_snooze.cfg
        sous forme {"YYYY-MM": "ISO datetime"}. Survit aux fermetures du hub."""
        now=datetime.now()
        target_year,target_month=now.year,now.month
        # Avant le 1er du mois 6h, ne rien déclencher
        if now.day==1 and now.hour<6:
            return
        # Si le passage de ce mois est déjà saisi, plus rien à faire
        if get_passage_mois(target_year,target_month):
            return
        # Vérifier que les prix sont dispo pour calcul
        if not get_prix_for_month(target_year,target_month): return
        if not get_prix_previous_month(target_year,target_month): return
        # Snooze utilisateur : si une heure de "rappel" a été posée et qu'elle n'est
        # pas encore atteinte, on n'ouvre pas
        snooze_path=APP_DIR/"passage_mois_snooze.cfg"
        snooze=load_json(snooze_path) or {}
        key=f"{target_year:04d}-{target_month:02d}"
        snooze_iso=snooze.get(key)
        if snooze_iso:
            try:
                rappel=datetime.fromisoformat(snooze_iso)
                if now<rappel: return  # snooze actif, on attend
            except Exception as _e: _log_silent_err(exc=_e)
        # Ne s'ouvre qu'une fois par session pour éviter ré-ouverture en boucle
        # quand l'utilisateur ferme avec Plus tard sans poser de snooze
        if getattr(self,"_passage_mois_shown_session",False):
            return
        self._passage_mois_shown_session=True
        self._open_passage_mois(target_year=target_year,target_month=target_month,edit_mode=False)

    def _maybe_open_observatoire_dlg(self,ar_data):
        """Ouvre ObservatoireDlg automatiquement si :
          - On est dans la période Observatoire (25 du mois → 5 du suivant)
          - AUCUNE vente atypique détectée dans Pre_vision sur cette période
            (= utilisateur n'a pas anticipé l'effet Observatoire → popup utile en rappel)
          - La popup n'a pas déjà été acquittée pour cette période
          - Le snooze (s'il existe) est expiré
          - Pas déjà montrée dans cette session
        
        Si l'utilisateur a déjà saisi des ventes atypiques sur la période (= a anticipé),
        la popup ne s'ouvre PAS car le travail métier est déjà fait. Le moteur tague ces
        ventes via ventes_irrealistes[].contexte = "observatoire"."""
        today=date.today()
        if not is_periode_observatoire(today): return
        # Identifier la période : si on est entre 25-31, c'est la période du mois en cours.
        # Si on est entre 1-5, c'est la période démarrée le mois précédent.
        if today.day>=25:
            periode_key=f"{today.year:04d}-{today.month:02d}"
        else:
            prev_year=today.year-1 if today.month==1 else today.year
            prev_month=12 if today.month==1 else today.month-1
            periode_key=f"{prev_year:04d}-{prev_month:02d}"
        # Lecture du snooze persistant
        try:
            obs_path=APP_DIR/"observatoire_snooze.cfg"
            data=load_json(obs_path) or {}
            entry=data.get(periode_key,{})
            if entry.get("acquitte"): return
            snooze_iso=entry.get("snooze_until")
            if snooze_iso:
                try:
                    rappel=datetime.fromisoformat(snooze_iso)
                    if datetime.now()<rappel: return
                except Exception as _e: _log_silent_err(exc=_e)
        except Exception as e: print(f"[observatoire check] {e}")
        # Vérifier qu'aucune vente atypique n'est déjà détectée sur la période.
        # Le moteur tague les ventes_irrealistes avec contexte="observatoire" quand on
        # est dans la période. Si des entrées existent → utilisateur a déjà anticipé.
        ventes_irr=ar_data.get("ventes_irrealistes",[]) if ar_data else []
        atypiques_observatoire=[v for v in ventes_irr if v.get("contexte")=="observatoire"]
        if atypiques_observatoire: return  # déjà anticipé, pas besoin de rappeler
        # Pas déjà montrée dans cette session (évite ré-ouverture en boucle)
        if getattr(self,"_observatoire_shown_session",False): return
        self._observatoire_shown_session=True
        try:
            ObservatoireDlg(self,periode_key=periode_key)
        except Exception as e: print(f"[observatoire dlg] {e}")
    def _check_rapport_suggestion(self):
        """Au premier lancement du mois, propose de générer le rapport du mois précédent.
        Mémorise le choix de l'utilisateur (génération ok / plus tard / ne plus demander)."""
        try:
            now=date.today()
            # Calcul du mois précédent
            if now.month==1: prev_y,prev_m=now.year-1,12
            else: prev_y,prev_m=now.year,now.month-1
            key=f"{prev_y:04d}-{prev_m:02d}"
            mem=load_json(RAPPORT_PROMPT_FILE) or {}
            entry=mem.get(key,{})
            # Si déjà généré OU "ne plus demander" → on ne propose plus
            if entry.get("statut") in ("genere","jamais_plus"): return
            # Proposer une seule fois par jour pour ne pas être intrusif
            last_prompt=entry.get("last_prompt_date","")
            if last_prompt==now.strftime("%Y-%m-%d"): return
            # Mettre à jour la date de prompt
            mem[key]={**entry,"last_prompt_date":now.strftime("%Y-%m-%d")}
            save_json(RAPPORT_PROMPT_FILE,mem)
            # Afficher la proposition (non-bloquante)
            self.after(2500,lambda:self._show_rapport_suggestion(prev_y,prev_m,key))
        except Exception as e:
            print(f"[suggestion rapport] {e}")
    def _show_rapport_suggestion(self,year,month,key):
        """Affiche un toast non-bloquant proposant de générer le rapport du mois précédent."""
        try:
            mn=["janvier","f\u00e9vrier","mars","avril","mai","juin","juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"][month-1]
            self._show_toast(f"\U0001f4c4 G\u00e9n\u00e8re le rapport de {mn} {year} \u2014 bouton Rapport",C["amber"])
        except Exception as _e: _log_silent_err(exc=_e)
    def _check_cycle_and_livraison(self):
        """Vérifie l'état du cycle 14j et de la saisie livraison du jour, ouvre la popup
        appropriée si nécessaire. Appelée au démarrage ET périodiquement (via _check_time_alerts)
        pour que ces popups apparaissent aussi quand le hub tourne H24.
        Garde-fous :
          - ne pas relancer si une popup cycle/livraison est déjà ouverte
          - anti-spam : pas de re-déclenchement avant 30 min après une tentative
        """
        try:
            # Garde-fou : ne pas empiler les popups
            if getattr(self,"_cycle_dlg_open",False) or getattr(self,"_livr_dlg_open",False):
                return
            # Anti-spam : 30 min entre 2 tentatives (évite le harcèlement si Bidou ferme par X)
            last_check=getattr(self,"_last_cycle_livr_check",0)
            if (datetime.now().timestamp()-last_check)<1800:
                return
            today_str=date.today().strftime("%d/%m/%y")
            livrs=load_json(LIVRAISON_FILE) or {}
            cycle=load_json(CYCLE_FILE) or {}
            ask_cycle=False
            if not cycle.get("anchor_date"):
                ask_cycle=True  # premier lancement
            elif date.today().weekday()==0:  # Lundi
                last_ask=cycle.get("last_asked","")
                if last_ask!=date.today().strftime("%Y-%m-%d"):
                    ask_cycle=True
            if ask_cycle:
                self._last_cycle_livr_check=datetime.now().timestamp()
                self._ask_cycle()
            elif today_str not in livrs:
                today_d=date.today()
                today_fp=today_d.strftime("%Y-%m-%d")
                # Commande explicite pour AUJOURD'HUI → popup dédié, QUEL QUE SOIT
                # le jour (samedi forcé inclus : la commande dit explicitement ce
                # jour ; on relit cette décision au lieu de l'ignorer).
                cmd_today=None
                try: cmd_today=get_commande(today_d)
                except Exception as _e: _log_silent_err(exc=_e)
                if cmd_today:
                    if is_popup_silenced("livraison_attendue",[today_fp],{today_fp:0}):
                        return  # déjà demandé aujourd'hui ("Pas encore")
                    self._last_cycle_livr_check=datetime.now().timestamp()
                    self._ask_livraison_attendue(cmd_today)
                    return
                # Aucune commande pour aujourd'hui : chemin LEGACY strictement
                # inchangé — uniquement Lun-Ven NON fériés (on ne casse rien).
                if today_d.weekday()>=5 or is_ferie(today_d):
                    return
                # Respect du silence "livraison_jour" écrit par LivraisonDialog (Plus tard 4h / croix X).
                if is_popup_silenced("livraison_jour",[today_fp],{today_fp:0}):
                    return  # encore en snooze, ne rien faire
                self._last_cycle_livr_check=datetime.now().timestamp()
                self._ask_livraison()
        except Exception as e: print(f"[check_cycle_livr] {e}")

    def _check_time_alerts(self):
        now=datetime.now();h,m=now.hour,now.minute;wd=now.weekday()
        # RESET QUOTIDIEN des flags toast 10h30/11h. self.alert_shown était initialisé
        # UNE seule fois dans __init__ et jamais remis à False : le hub tournant en
        # continu (H24, jamais redémarré), le toast 11h ne partait qu'UNE fois dans
        # toute la vie du process puis plus jamais (= "popup 11h disparue").
        _today=now.date()
        if getattr(self,"_alert_day",None)!=_today:
            self._alert_day=_today
            self.alert_shown={"1030":False,"1100":False}
        # Vérification cycle 14j + livraison du jour : déclenche aussi en cours de journée
        # (pas seulement au démarrage), avec garde-fou anti-spam de 30 min
        try: self._check_cycle_and_livraison()
        except Exception as e: print(f"[check_cycle_livr in time_alerts] {e}")
        # BRIQUE 2 : alerte CALIBRÉE sur l'heure limite du tour (escalade).
        try: self._check_livraison_attendue_escalade(now)
        except Exception as _e: _log_silent_err(exc=_e)
        # Rappels uniquement lundi(0) à vendredi(4)
        if wd>4: self.after(60000,self._check_time_alerts);return
        today_str=date.today().strftime("%d/%m/%y")
        livrs=load_json(LIVRAISON_FILE)
        livr_done=today_str in livrs
        # Rappels remplacés par des toasts non-bloquants (plus de popup système intrusive)
        if h==10 and m>=30 and not self.alert_shown["1030"]:
            self.alert_shown["1030"]=True
            if not livr_done:
                self._show_toast("\u23f0 10h30 — V\u00e9rifie ta livraison du jour, rectifie ta commande si elle tarde",C["amber"])
        # 11h : seulement si on n'a pas dépassé l'heure (pas de toast à 11h15)
        if h==11 and m<5 and not self.alert_shown["1100"]:
            self.alert_shown["1100"]=True
            # Le toast 11h est désormais CONSCIENT de la commande. Cible = prochain
            # jour livrable strict (= ce pour quoi la deadline TEMAG 11h compte).
            try:
                _cible=date.today()+timedelta(days=1)
                while _cible.weekday()>=5 or is_ferie(_cible):
                    _cible+=timedelta(days=1)
                _cmd=get_commande(_cible)
                _cib_str=f"{JOURS_FR[_cible.weekday()]} {_cible.strftime('%d/%m')}"
            except Exception as _e:
                _log_silent_err(exc=_e); _cmd=None; _cib_str=""
            if _cmd:
                # Commande déjà saisie → on ÉTEINT le bruit (rien à rappeler).
                pass
            else:
                # Commande NON saisie + deadline maintenant → rappel précis et utile.
                self._show_toast(f"\u26a0 11h00 — Commande non saisie pour {_cib_str} \u2014 deadline TEMAG maintenant",C["red"])
        self.after(60000,self._check_time_alerts)

    def _add_tooltip(self,widget,text,delay=500):
        """Ajoute un tooltip qui apparaît au survol du widget après un court délai.
        Utilisé pour les boutons icon-only qui n'ont pas de label texte explicite."""
        state={"win":None,"after_id":None}
        def show():
            state["after_id"]=None
            if state["win"] is not None: return
            try:
                x=widget.winfo_rootx()+widget.winfo_width()//2
                y=widget.winfo_rooty()+widget.winfo_height()+6
                tip=ctk.CTkToplevel(widget)
                tip.overrideredirect(True)
                tip.attributes("-topmost",True)
                tip.configure(fg_color="#1E2736")
                ctk.CTkLabel(tip,text=text,font=("Segoe UI",10),text_color=C["t1"]).pack(padx=10,pady=4)
                tip.update_idletasks()
                w=tip.winfo_width()
                tip.geometry(f"+{x-w//2}+{y}")
                state["win"]=tip
            except Exception as _e: _log_silent_err(exc=_e)
        def hide(_=None):
            if state["after_id"]:
                try: widget.after_cancel(state["after_id"])
                except Exception as _e: _log_silent_err(exc=_e)
                state["after_id"]=None
            if state["win"] is not None:
                try: state["win"].destroy()
                except Exception as _e: _log_silent_err(exc=_e)
                state["win"]=None
        def on_enter(_):
            hide()
            state["after_id"]=widget.after(delay,show)
        widget.bind("<Enter>",on_enter,add="+")
        widget.bind("<Leave>",hide,add="+")
        widget.bind("<ButtonPress-1>",hide,add="+")

    def _build(self):
        self.grid_columnconfigure(0,weight=1);self.grid_columnconfigure(1,weight=0,minsize=400);self.grid_rowconfigure(2,weight=1)
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=72);hdr.grid(row=0,column=0,columnspan=2,sticky="ew",padx=24,pady=(18,10));hdr.grid_propagate(False);hdr.grid_columnconfigure(0,weight=1)
        left=ctk.CTkFrame(hdr,fg_color="transparent");left.grid(row=0,column=0,sticky="w")
        tr=ctk.CTkFrame(left,fg_color="transparent");tr.pack(anchor="w")
        ctk.CTkLabel(tr,text="\u25cf",font=("Segoe UI",20),text_color=C["red"]).pack(side="left",padx=(0,8))
        ctk.CTkLabel(tr,text="DISTRICARB HUB",font=("Segoe UI",22,"bold"),text_color=C["t1"]).pack(side="left")
        ctk.CTkLabel(tr,text="v0.5",font=("Segoe UI",9,"bold"),text_color=C["gold"],fg_color=C["card"],corner_radius=4,padx=6,pady=2).pack(side="left",padx=(10,0))
        # Pilule date du jour (format long "Samedi 9 mai 2026").
        # Le bug d'affichage de la date qui rétrécissait selon le contenu a été corrigé en
        # imposant des largeurs explicites aux boutons de droite (cf. header_btn_width plus haut) :
        # sans ça, chaque CTkFrame outer gardait une largeur par défaut ~200 px qui étranglait
        # la zone gauche du header. Plus besoin de forcer la date à une largeur fixe.
        # sur certains postes Windows — la pilule peut tronquer la fin de la date. Voir
        # BUG_DATE_PILL_RECAP.md pour le diagnostic complet et les pistes restantes.
        _jours_fr=["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
        _mois_fr=["janvier","f\u00e9vrier","mars","avril","mai","juin","juillet","ao\u00fbt","septembre","octobre","novembre","d\u00e9cembre"]
        _td=date.today()
        _date_str=f"{_jours_fr[_td.weekday()]} {_td.day} {_mois_fr[_td.month-1]} {_td.year}"
        date_pill=ctk.CTkFrame(tr,fg_color="#1E2736",corner_radius=18,height=36)
        date_pill.pack(side="left",padx=(14,0))
        ctk.CTkLabel(date_pill,text="\U0001f4c5",width=26,height=26,font=("Segoe UI Emoji",11),text_color=C["t2"],fg_color="transparent",corner_radius=0).pack(side="left",padx=(5,8),pady=5)
        # tk.Label NATIF, texte propre sans bouclier d'espaces (la pilule n'est plus clippée).
        self.lbl_date_header=tk.Label(date_pill,text=_date_str,font=("Segoe UI",12),fg=C["t1"],bg="#1E2736",bd=0,anchor="w")
        self.lbl_date_header.pack(side="left",padx=(0,14),pady=5)
        self._jours_fr_head=_jours_fr;self._mois_fr_head=_mois_fr
        ctk.CTkLabel(left,text="B. DISTRICARB SARL \u2014 Place d'Armes, Le Lamentin",font=("Segoe UI",11),text_color=C["t3"]).pack(anchor="w",pady=(2,0))
        right=ctk.CTkFrame(hdr,fg_color="transparent");right.grid(row=0,column=1,sticky="e")
        # Boutons en pilules avec rond coloré - effet relief + hover + clic
        self.refresh_pill=None;self.refresh_circle=None;self.refresh_label=None
        # PATCH ChatGPT : largeur explicite par bouton calibrée sur le contenu réel.
        # Sans cette contrainte, chaque CTkFrame outer garde sa largeur CTk par défaut (~200 px),
        # ce qui fait ~1200 px à droite et étrangle la zone gauche (titre + pilule date).
        header_btn_width={"Actualiser":135,"Livraisons":135,"Journal":118,"Alertes":118,"Prix":95,"Rapport":120,"Param\u00e8tres":145}
        for icon,txt,cmd,col,key in [
            ("\u21bb","Actualiser",self.refresh,"#1F7FD4","refresh"),
            ("\U0001f69a","Livraisons",self._open_livraisons,"#7DD36F",None),
            ("\U0001f4cb","Journal",self._open_journal_evenements,"#A47BC4",None),
            ("\U0001f6a8","Alertes",self._open_alertes_dashboard,"#E54E5C",None),
            ("\U0001f4b0","Prix",self._open_prix_carburant,"#2DA84A",None),
            ("\U0001f4c4","Rapport",self._open_rapport_menu,"#E48B2A",None),
            ("\u2699","Param\u00e8tres",self.settings,"#9B59D4",None)]:
            # Conteneur externe pour effet relief (bordure plus foncée en bas)
            outer=ctk.CTkFrame(right,fg_color="transparent",width=header_btn_width.get(txt,112),height=42)
            outer.pack(side="right",padx=(12,0));outer.pack_propagate(False)
            pill=ctk.CTkFrame(outer,fg_color=C["card"],corner_radius=22,border_width=1,border_color=C["border2"],height=40)
            pill.pack(fill="x",pady=(0,2))  # 2px de marge en bas = effet relief
            circle=ctk.CTkLabel(pill,text=icon,width=30,height=30,font=("Segoe UI",13,"bold"),text_color="#FFFFFF",fg_color=col,corner_radius=15)
            circle.pack(side="left",padx=(5,8),pady=5)
            label=ctk.CTkLabel(pill,text=txt,font=("Segoe UI",12),text_color=C["t1"])
            label.pack(side="left",padx=(0,8))
            if key=="refresh":
                self.refresh_pill=pill;self.refresh_circle=circle;self.refresh_label=label
            # Effets : hover = légère teinte de la couleur du bouton, clic = enfoncement
            def _hover_on(e,p=pill,c=col):
                p.configure(border_color=c)
            def _hover_off(e,p=pill):
                p.configure(border_color=C["border2"])
            def _press(e,p=pill,o=outer):
                p.pack_configure(pady=(2,0))  # enfoncé (relief disparaît)
            def _release(e,p=pill,c=cmd):
                p.pack_configure(pady=(0,2));c()
            for w in [pill,circle,label]:
                w.bind("<Enter>",_hover_on)
                w.bind("<Leave>",_hover_off)
                w.bind("<ButtonPress-1>",_press)
                w.bind("<ButtonRelease-1>",_release)
        # Bannière d'alerte Prévision : visible UNIQUEMENT quand Pre_vision_compte.xlsx est inaccessible
        # (verrou Excel, fichier manquant, ou erreur de lecture). Masquée par défaut.
        # Sans ça, le hub utilisait silencieusement des données vides → risque de passer à côté
        # d'alertes anti-rupture critiques sans s'en apercevoir.
        self.banner_prev=ctk.CTkFrame(self,fg_color="#3A1520",corner_radius=8,border_width=1,border_color=C["red"])
        self.banner_prev_lbl=ctk.CTkLabel(self.banner_prev,text="",font=("Segoe UI",11,"bold"),
                                          text_color="#FFB0B0",wraplength=1300,justify="left",anchor="w")
        self.banner_prev_lbl.pack(fill="x",padx=14,pady=10)
        self.banner_prev.grid(row=1,column=0,columnspan=2,sticky="ew",padx=24,pady=(0,8))
        self.banner_prev.grid_remove()  # caché tant que pas d'erreur
        self.lp=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        self.lp.grid(row=2,column=0,sticky="nsew",padx=(24,12),pady=(0,16));self.lp.grid_columnconfigure(0,weight=1);self.lp.grid_columnconfigure(1,weight=1)
        for i,fd in enumerate(HUB_FILES):
            v=Vignette(self.lp,fd,on_click=self.open_file,on_detail=self.show_details);v.grid(row=i//2,column=i%2,sticky="nsew",padx=6,pady=6);self.vigs[fd["key"]]=v
        # 6e vignette : RACCOURCIS vers outils externes
        raccourcis=[
            {"icon":"\u2601","label":"OneDrive","color":"#4E6D92","action":lambda:open_onedrive(self)},
            {"icon":"\u20ac","label":"Crédit Agricole","color":"#5A7552","action":lambda:launch_chrome("https://www.ca-martinique.fr",incognito=True)},
            {"icon":"\u2709","label":"Mail Factures","color":"#9B7048","action":open_orange_mail_tool},
            {"icon":"\u26fd","label":"TotalEnergies","color":"#8B4F4F","action":lambda:launch_chrome("https://dealer.fleet.totalenergies.com/mq/logout-success")},
            {"icon":"\u25c6","label":"Nepting","color":"#5F5F85","action":lambda:launch_chrome("https://nepsa1.nepting.com/alladmin/?xpdevalue=sunadmin#MainPlace:default")},
            {"icon":"\u25c9","label":"Worldline","color":"#4A7579","action":lambda:launch_chrome("https://prepaidservicesfrance.com")},
        ]
        rv=RaccourciVignette(self.lp,raccourcis);rv.grid(row=len(HUB_FILES)//2,column=len(HUB_FILES)%2,sticky="nsew",padx=6,pady=6)
        self.hist_frame=ctk.CTkFrame(self.lp,fg_color="transparent");self.hist_frame.grid(row=3,column=0,columnspan=2,sticky="ew",padx=6,pady=(12,6))
        self.rp=ctk.CTkScrollableFrame(self,fg_color=C["panel"],corner_radius=14,border_width=1,border_color=C["border"],scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        self.rp.grid(row=2,column=1,sticky="nsew",padx=(12,24),pady=(0,16))
        ctk.CTkLabel(self.rp,text="PILOTAGE JOUR",font=("Segoe UI",11,"bold"),text_color=C["gold"]).pack(anchor="w",padx=16,pady=(14,2))
        self.b_date=ctk.CTkLabel(self.rp,text="\u2014",font=("Segoe UI",20,"bold"),text_color=C["t1"]);self.b_date.pack(anchor="w",padx=16,pady=(0,1))
        self.b_sub=ctk.CTkLabel(self.rp,text="\u2014",font=("Segoe UI",10),text_color=C["t3"]);self.b_sub.pack(anchor="w",padx=16,pady=(0,14))
        self.s_carb=Section(self.rp,"Carburants",C["red"]);self.s_carb.pack(fill="x",padx=14)
        self.s_enc=Section(self.rp,"Encaissements",C["blue"]);self.s_enc.pack(fill="x",padx=14)
        self.s_bout=Section(self.rp,"Boutique",C["amber"]);self.s_bout.pack(fill="x",padx=14)
        self.s_stock=Section(self.rp,"Stocks carburant",C["gold"]);self.s_stock.pack(fill="x",padx=14)
        self.s_obj=Section(self.rp,"Objectif mensuel",C["green"]);self.s_obj.pack(fill="x",padx=14)
        self.s_alert=Section(self.rp,"\U0001f6a8 Alertes",C["red"]);self.s_alert.pack(fill="x",padx=14)
        ft=ctk.CTkFrame(self,fg_color="transparent",height=24);ft.grid(row=3,column=0,columnspan=2,sticky="ew",padx=24,pady=(0,10));ft.grid_propagate(False)
        self.ft_lbl=ctk.CTkLabel(ft,text="Pr\u00eat",font=("Segoe UI",10),text_color=C["t3"]);self.ft_lbl.pack(side="left")
        ctk.CTkLabel(ft,text="DISTRICARB HUB v0.5",font=("Segoe UI",9),text_color=C["t3"]).pack(side="right")

    def _first_check(self):
        missing=[f["key"] for f in HUB_FILES if not self.cfg.get(f["key"])]
        if missing: self.after(400,self.settings)
    def settings(self):
        dlg=SettingsDlg(self,self.cfg);self.wait_window(dlg)
        if dlg.result:
            try:
                self.cfg=dict(dlg.result)
                save_json(CONFIG_FILE,self.cfg)
                # Vérification que l'écriture a bien pris (utile si l'antivirus bloque)
                relu=load_json(CONFIG_FILE) or {}
                if relu.get("objectif")!=self.cfg.get("objectif"):
                    messagebox.showwarning("Sauvegarde incomplète",
                        f"La sauvegarde n'a pas pris effet.\n\n"
                        f"En mémoire : {self.cfg.get('objectif','(vide)')}\n"
                        f"Sur disque : {relu.get('objectif','(vide)')}\n\n"
                        f"Fichier : {CONFIG_FILE}")
                self.reader=DataReader(self.cfg)
                self.refresh()
            except Exception as e:
                messagebox.showerror("Erreur sauvegarde",f"Impossible d'enregistrer la configuration :\n\n{e}\n\nFichier visé : {CONFIG_FILE}")
    def open_file(self,key):
        p=self.cfg.get(key)
        if not p: return
        if not os.path.exists(p): messagebox.showerror("Introuvable",p);return
        try:
            if sys.platform=="win32": os.startfile(p)
            elif sys.platform=="darwin": os.system(f'open "{p}"')
            else: os.system(f'xdg-open "{p}"')
        except Exception as e: messagebox.showerror("Erreur",str(e))
    def show_details(self,key):
        if self.last_data: DetailWindow(self,key,self.last_data)
    def _open_antirupture_dlg(self,ar):
        """Ouvre la popup anti-rupture en gardant une référence singleton.
        Si une popup est déjà ouverte, on ne fait rien (évite l'empilement timer 15min)."""
        existing=getattr(self,"_antirupture_dlg",None)
        if existing is not None and existing.winfo_exists():
            try: existing.lift();existing.focus_force()
            except Exception as _e: _log_silent_err(exc=_e)
            return
        self._antirupture_dlg=AntiRuptureDlg(self,ar)
    def _maybe_open_tendance_dlg(self,tendance_alertes):
        """Décide d'ouvrir la popup tendance forte selon les acquittements existants.
        Logique :
         - Pour chaque alerte (= un carburant), on calcule sa clé d'acquittement {date}_{stage}_{carb}
         - Si la clé n'existe pas dans le fichier ack → alerte non acquittée → afficher popup
         - Si la clé existe MAIS l'écart actuel a augmenté de plus de 20 points → alerte aggravée → réafficher popup
         - Sinon (acquittée et stable) → on n'affiche rien
        Singleton : si la popup est déjà ouverte, on ne la dédouble pas."""
        try:
            acks=load_json(TENDANCE_ACK_FILE) or {}
        except Exception: acks={}
        d_str=date.today().strftime("%Y-%m-%d")
        alertes_a_afficher=[]
        for al in tendance_alertes:
            key=f"{d_str}_{al['stage'].replace(' ','_').replace('/','-')}_{al['carburant']}"
            ack=acks.get(key)
            if not ack:
                # Pas acquittée → on la montre
                alertes_a_afficher.append(al)
                continue
            # Acquittée : vérifier si l'écart s'est aggravé de >20 points
            try: ecart_ack=float(ack.get("ecart_pct",0))
            except Exception: ecart_ack=0
            ecart_now=al["ecart_pct"]
            # Aggravation = on s'éloigne davantage de 0 dans le même sens
            if ecart_now>0 and ecart_ack>=0 and (ecart_now-ecart_ack)>20:
                alertes_a_afficher.append(al)
            elif ecart_now<0 and ecart_ack<=0 and (ecart_ack-ecart_now)>20:
                alertes_a_afficher.append(al)
        if not alertes_a_afficher: return
        # Singleton anti-empilement
        existing=getattr(self,"_tendance_dlg",None)
        if existing is not None and existing.winfo_exists():
            try: existing.lift();existing.focus_force()
            except Exception as _e: _log_silent_err(exc=_e)
            return
        self._tendance_dlg=TendanceAlerteDlg(self,alertes_a_afficher)
    def _maybe_open_saisies_irr_dlg(self,saisies_irr):
        """Décide d'ouvrir la popup saisies physiquement impossibles selon les acquittements existants.
        Logique :
         - Charger snapshot précédent (last_ack)
         - Si une saisie irréaliste actuelle n'existe PAS dans le snapshot → nouveau cas → afficher
         - Si une saisie existait mais son excès a augmenté de >1000 L → aggravation → afficher
         - Sinon → ne pas afficher (déjà acquitté)
        Singleton : si la popup est déjà ouverte, on ne la dédouble pas."""
        try:
            acks=load_json(SAISIES_IRR_ACK_FILE) or {}
        except Exception: acks={}
        last_ack=acks.get("last_ack",{}) or {}
        snapshot_prec=last_ack.get("snapshot",{}) or {}
        afficher=False
        for s in saisies_irr:
            key=f"{s['date'].strftime('%Y-%m-%d')}_{s['carburant']}"
            exces_now=s["exces"]
            exces_prec=sf(snapshot_prec.get(key,-1))
            if exces_prec<0:
                # Nouveau cas pas dans le snapshot
                afficher=True;break
            if exces_now>exces_prec+1000:
                # Aggravation significative
                afficher=True;break
        if not afficher: return
        # Singleton anti-empilement
        existing=getattr(self,"_saisies_irr_dlg",None)
        if existing is not None and existing.winfo_exists():
            try: existing.lift();existing.focus_force()
            except Exception as _e: _log_silent_err(exc=_e)
            return
        self._saisies_irr_dlg=SaisiesIrrealistesDlg(self,saisies_irr)
    def refresh(self):
        # Annuler le timer en cours et replanifier dans 15 min à partir de maintenant
        try:
            if hasattr(self,"_refresh_timer") and self._refresh_timer:
                self.after_cancel(self._refresh_timer)
        except Exception as _e: _log_silent_err(exc=_e)
        # Mettre à jour la date du header (utile quand le hub tourne sans redémarrage plusieurs jours)
        if hasattr(self,"lbl_date_header") and hasattr(self,"_jours_fr_head"):
            _td=date.today()
            _date_str=f"{self._jours_fr_head[_td.weekday()]} {_td.day} {self._mois_fr_head[_td.month-1]} {_td.year}"
            try: self.lbl_date_header.configure(text=_date_str)
            except Exception as _e: _log_silent_err(exc=_e)
        # Feedback visuel : cercle grisé pendant chargement (label optionnel s'il existe)
        if self.refresh_circle:
            self.refresh_circle.configure(fg_color="#888888")
        if self.refresh_label:
            self.refresh_label.configure(text="Chargement...",text_color=C["t3"])
        self.ft_lbl.configure(text="Actualisation...")
        self.update_idletasks()  # Force le redessin AVANT le travail bloquant
        try:
            d=self.reader.read_all();self.last_data=d;self._upd_vigs(d);self._upd_banner(d);self._upd_hist(d);self._update_prevision_banner(d)
            # NOTE : detect_recent_events est désormais appelé DANS read_all (Point A),
            # avant analyze_antirupture/_calc_autonomie, pour que la mémoire soit à jour
            # AVANT les calculs du même refresh. Plus d'appel ici (évite double passage).
            ft_text=f"Actualis\u00e9 \u00e0 {d['ts']} \u2502 Refresh : 15 min"
            self.ft_lbl.configure(text=ft_text)
            self._ft_lbl_text=ft_text  # miroir API
            # Restaurer cercle bleu (label optionnel s'il existe) + toast succès
            if self.refresh_circle:
                self.refresh_circle.configure(fg_color="#1F7FD4")
            if self.refresh_label:
                self.refresh_label.configure(text="Actualiser",text_color=C["t1"])
            self._show_toast(f"\u2713 Actualis\u00e9 \u00e0 {d['ts']}",C["green"])
            # ============================================================
            # POPUP : Passage de mois (1er du mois après 6h non saisi)
            # Se déclenche tant que l'utilisateur n'a pas saisi OU cliqué "Plus tard".
            # "Plus tard" = on reporte au prochain refresh (pas de mémorisation
            # définitive : le but est de capturer la donnée tant qu'elle est fraîche).
            # ============================================================
            try:
                self._maybe_open_passage_mois_dlg()
            except Exception as _e: print(f"[passage mois detect] {_e}")
            ar=d.get("antirupture",{}) or {}
            # ============================================================
            # POPUP : Période Observatoire des Prix (mensuel, fin de mois)
            # Se déclenche entre le 25 et le 5 si l'utilisateur n'a pas
            # encore anticipé l'effet via Pre_vision. Cible : Véronique/Nadine
            # qui n'ont pas forcément intégré ce mécanisme métier.
            # ============================================================
            try:
                self._maybe_open_observatoire_dlg(ar)
            except Exception as _e: print(f"[observatoire detect] {_e}")
            # ============================================================
            # POPUP 1 : Saisies physiquement impossibles
            # Fingerprint = "{date}_{carburant}" ; extra = dépassement en L
            # Aggravation : +1000L de dépassement
            # ============================================================
            saisies_irr=ar.get("saisies_physiquement_impossibles",[]) or []
            if saisies_irr:
                fps=[f"{s.get('date','')}_{s.get('carburant','')}" for s in saisies_irr]
                extra={f"{s.get('date','')}_{s.get('carburant','')}":s.get("depassement_l",0) for s in saisies_irr}
                def aggrav_irr(fp,old,new):
                    try: return (new.get(fp,0)-old.get(fp,0))>1000
                    except Exception as _e: _log_silent_err(exc=_e); return False
                if not is_popup_silenced("saisies_irr",fps,extra,aggrav_irr):
                    self._maybe_open_saisies_irr_dlg(saisies_irr)
            # ============================================================
            # POPUP 1bis : Livraisons à reporter (capacité cuve dépassée à 6h)
            # Fingerprint = "{date}_{carburant}" ; extra = surplus en L
            # Aggravation : +500L de surplus
            # ============================================================
            livr_report=ar.get("livraisons_a_reporter",[]) or []
            # FENÊTRE D'ACTION (anti-harcèlement) : une livraison à reporter ne doit
            # solliciter Bidou que lorsqu'on entre dans les 3 JOURS LIVRABLES avant
            # sa date. Trop tôt = il ne peut rien faire (pas encore commandé) et se
            # fait harceler (14 snoozes constatés). L'événement reste connu/visible
            # dans le tableau de notifs, mais ne POP plus prématurément.
            FENETRE_LIVR_REPORT=3
            livr_report_fenetre=[]
            for l in livr_report:
                try:
                    dc=l["date"]
                    dc=dc if isinstance(dc,date) else dc.date()
                    if nb_jours_livrables_avant(dc)<=FENETRE_LIVR_REPORT:
                        livr_report_fenetre.append(l)
                except Exception as _e:
                    _log_silent_err(exc=_e)
                    livr_report_fenetre.append(l)  # en cas de doute, on garde
            if livr_report_fenetre:
                fps=[f"{l['date'].strftime('%Y-%m-%d')}_{l['carburant']}" for l in livr_report_fenetre]
                extra={f"{l['date'].strftime('%Y-%m-%d')}_{l['carburant']}":l.get("surplus",0) for l in livr_report_fenetre}
                def aggrav_lr(fp,old,new):
                    try: return (new.get(fp,0)-old.get(fp,0))>500
                    except Exception as _e: _log_silent_err(exc=_e); return False
                if not is_popup_silenced("livr_report",fps,extra,aggrav_lr):
                    # SINGLETON ANTI-EMPILEMENT BLINDÉ — voir MargeTendueDlg
                    already_open=False
                    try:
                        for w in self.winfo_children():
                            if isinstance(w,LivraisonsAReporterDlg) and w.winfo_exists():
                                already_open=True
                                try: w.lift();w.focus_force()
                                except Exception: pass
                                break
                    except Exception as _e: _log_silent_err(exc=_e)
                    if not already_open:
                        try:
                            dlg=LivraisonsAReporterDlg(self,livr_report_fenetre)
                            self.wait_window(dlg)
                        except Exception as e: print(f"[popup livr report] {e}")
            # ============================================================
            # POPUP 1bis : Rupture imminente JOUR LIVRABLE
            # Cas non couvert par la popup anti-rupture (qui ne traite que les ruptures DANS un trou).
            # Ici on alerte quand autonomie < 24h en jour livrable et aucune livraison prévue.
            # Fingerprint = "{date}_{carburant}" ; extra = autonomie en heures
            # Aggravation : autonomie chute de 2h ou plus (la situation empire vraiment)
            # ============================================================
            ruptures_imm=ar.get("ruptures_imminentes",[]) or []
            if ruptures_imm:
                d_str=date.today().strftime("%Y-%m-%d")
                fps=[f"{d_str}_{r.get('carburant','')}" for r in ruptures_imm]
                extra={f"{d_str}_{r.get('carburant','')}":r.get("autonomie_h",0) for r in ruptures_imm}
                def aggrav_ri(fp,old,new):
                    # Aggravation = l'autonomie h a diminué de 2h ou plus depuis le dernier ack
                    # (situation qui empire, justifie une re-popup avant fin du silence)
                    try: return (old.get(fp,99)-new.get(fp,99))>=2
                    except Exception as _e: _log_silent_err(exc=_e); return False
                if not is_popup_silenced("rupture_imminente",fps,extra,aggrav_ri):
                    # SINGLETON ANTI-EMPILEMENT BLINDÉ — voir MargeTendueDlg
                    already_open=False
                    try:
                        for w in self.winfo_children():
                            if isinstance(w,RuptureImminenteDlg) and w.winfo_exists():
                                already_open=True
                                try: w.lift();w.focus_force()
                                except Exception: pass
                                break
                    except Exception as _e: _log_silent_err(exc=_e)
                    if not already_open:
                        try:
                            dlg=RuptureImminenteDlg(self,ruptures_imm)
                            self.wait_window(dlg)
                        except Exception as e: print(f"[popup rupture imm] {e}")
            # ============================================================
            # POPUP 1bis : Férié isolé imminent (jour férié seul entre 2 jours ouvrés)
            # Cas typique : demain = Ascension/8 mai/etc. → deadline commande aujourd'hui 11h
            # pour livraison après-demain. Pas géré par la logique "trou" qui exige ≥ 2 jours.
            # Fingerprint = date ISO du férié. Pas d'aggravation (situation statique).
            # ============================================================
            feries_isoles=ar.get("feries_isoles_imminents",[]) or []
            if feries_isoles:
                fps_fi=[f["date_ferie"].isoformat() for f in feries_isoles]
                if not is_popup_silenced("ferie_isole",fps_fi,{}):
                    # SINGLETON ANTI-EMPILEMENT BLINDÉ — voir MargeTendueDlg
                    already_open=False
                    try:
                        for w in self.winfo_children():
                            if isinstance(w,FerieIsoleDlg) and w.winfo_exists():
                                already_open=True
                                try: w.lift();w.focus_force()
                                except Exception: pass
                                break
                    except Exception as _e: _log_silent_err(exc=_e)
                    if not already_open:
                        try:
                            dlg=FerieIsoleDlg(self,feries_isoles)
                            self.wait_window(dlg)
                        except Exception as e: print(f"[popup ferie isole] {e}")
            # ============================================================
            # POPUP 1ter : Livraisons à marge tendue (cuve presque pleine).
            # Seuil < 4000 L de marge cuve après livraison (aligné fichier Excel "Attention" maison).
            # Pas un blocage : la livraison rentre, mais surveillance recommandée.
            # ============================================================
            livr_tendues=ar.get("livraisons_marge_tendue",[]) or []
            if livr_tendues:
                fps_mt=[]
                for l in livr_tendues:
                    d_iso=l["date"].isoformat() if hasattr(l.get("date"),"isoformat") else str(l.get("date",""))
                    fps_mt.append(f"{d_iso}_{l['carburant']}")
                if not is_popup_silenced("marge_tendue",fps_mt,{}):
                    # SINGLETON ANTI-EMPILEMENT BLINDÉ (bug 20/05 18h37 : 5 popups
                    # empilées malgré le premier fix par référence-attribut).
                    # Le premier fix utilisait self._marge_tendue_dlg, mais cette
                    # référence pouvait être perdue/désynchronisée. Approche robuste :
                    # demander à Tk la liste de SES enfants Toplevel et vérifier
                    # s'il y en a déjà un de classe MargeTendueDlg.
                    already_open=False
                    try:
                        for w in self.winfo_children():
                            if isinstance(w,MargeTendueDlg) and w.winfo_exists():
                                already_open=True
                                try: w.lift();w.focus_force()
                                except Exception: pass
                                break
                    except Exception as _e: _log_silent_err(exc=_e)
                    if not already_open:
                        try:
                            dlg=MargeTendueDlg(self,livr_tendues)
                            self.wait_window(dlg)
                        except Exception as e: print(f"[popup marge tendue] {e}")
            # ============================================================
            # POPUP 2 : Anti-rupture (ponts non-acquittés ou saisies impossibles)
            # Fingerprint = pont_id ; extra = manques en L par carburant
            # Aggravation : un manque augmente de +20%
            # ============================================================
            if ar.get("severite_max")=="critique":
                # Garde-fou : ne pas ouvrir la popup si elle n'aurait rien à afficher.
                has_content=(ar.get("ruptures_dans_trou") or ar.get("incoherences_jour_non_livrable"))
                # FENÊTRE D'ACTION sur le PONT (anti-harcèlement) : un pont à anticiper
                # ne doit solliciter Bidou que lorsqu'on approche du JOUR DE COMMANDE
                # (3 jours livrables avant). Trop tôt = harcèlement inutile (cf. les
                # 14 snoozes constatés). MAIS sécurité absolue : on ne masque JAMAIS
                # une incohérence jour non-livrable (urgente par nature), ni un pont
                # dont le jour de commande est déjà là ou passé.
                FENETRE_PONT=3
                incoh=ar.get("incoherences_jour_non_livrable") or []
                plan_lisse=ar.get("plan_lisse",[]) or []
                pont_dans_fenetre=False
                if not plan_lisse:
                    # Pas d'info de plan → on ne prend pas le risque de masquer
                    pont_dans_fenetre=True
                else:
                    for p in plan_lisse:
                        jc=p.get("jour_commande")
                        if jc is None:
                            pont_dans_fenetre=True;break
                        try:
                            jcd=jc if isinstance(jc,date) else jc.date()
                            if nb_jours_livrables_avant(jcd)<=FENETRE_PONT:
                                pont_dans_fenetre=True;break
                        except Exception as _e:
                            _log_silent_err(exc=_e);pont_dans_fenetre=True;break
                # On ouvre si : du contenu existe ET (une incohérence urgente OU un
                # pont dont le jour de commande est dans la fenêtre).
                if not has_content:
                    pass  # rien à afficher → on n'ouvre pas
                elif not (incoh or pont_dans_fenetre):
                    pass  # pont encore lointain, aucune urgence → pas de pop (reste au tableau)
                else:
                    fps,extra=_antirupture_fps_extra(ar)
                    def aggrav_ar(fp,old,new):
                        try:
                            old_m=old.get(fp,{}) or {};new_m=new.get(fp,{}) or {}
                            for carb,nv in new_m.items():
                                ov=old_m.get(carb,0)
                                if ov<=0:
                                    if nv>0: return True
                                elif (nv-ov)/ov>0.20: return True
                            return False
                        except Exception as _e: _log_silent_err(exc=_e); return False
                    if fps and is_popup_silenced("antirupture",fps,extra,aggrav_ar):
                        pass  # silence actif : on n'ouvre pas la popup
                    else:
                        existing=getattr(self,"_antirupture_dlg",None)
                        already_open=existing is not None and existing.winfo_exists()
                        if not already_open:
                            self.after(200,lambda ar=ar:self._open_antirupture_dlg(ar))
            # ============================================================
            # POPUP 3 : Tendance forte (ventes anormales du jour)
            # Fingerprint = "{date}_{carburant}" ; extra = écart_pct
            # Aggravation : écart +20 points
            # ============================================================
            tendance_alertes=ar.get("tendance_alertes",[]) or []
            if tendance_alertes:
                d_str=date.today().strftime("%Y-%m-%d")
                fps=[f"{d_str}_{al.get('carburant','')}" for al in tendance_alertes]
                extra={f"{d_str}_{al.get('carburant','')}":al.get("ecart_pct",0) for al in tendance_alertes}
                def aggrav_td(fp,old,new):
                    try:
                        ov=old.get(fp,0);nv=new.get(fp,0)
                        if ov>0 and nv>0: return (nv-ov)>20
                        if ov<0 and nv<0: return (ov-nv)>20
                        return False  # sens opposé : pas d'aggravation, c'est une autre alerte
                    except Exception as _e: _log_silent_err(exc=_e); return False
                if not is_popup_silenced("tendance",fps,extra,aggrav_td):
                    self._maybe_open_tendance_dlg(tendance_alertes)
        except Exception as e:
            traceback.print_exc();self.ft_lbl.configure(text=f"Erreur: {e}")
            if self.refresh_circle:
                self.refresh_circle.configure(fg_color="#1F7FD4")
            if self.refresh_label:
                self.refresh_label.configure(text="Actualiser",text_color=C["t1"])
            self._show_toast(f"\u26a0 Erreur : {str(e)[:50]}",C["red"])
        # Replanifier le prochain refresh dans 15 min à partir de MAINTENANT
        self._refresh_timer=self.after(REFRESH_MS,self._loop)

    def _show_toast(self,msg,color):
        """Toast en bas à droite, disparaît après 3s."""
        try:
            toast=ctk.CTkToplevel(self)
            toast.overrideredirect(True);toast.attributes("-topmost",True)
            toast.configure(fg_color=C["card"])
            # Position en bas à droite
            self.update_idletasks()
            wx=self.winfo_x()+self.winfo_width()-340
            wy=self.winfo_y()+self.winfo_height()-90
            toast.geometry(f"320x50+{wx}+{wy}")
            frame=ctk.CTkFrame(toast,fg_color=C["card"],corner_radius=12,border_width=2,border_color=color)
            frame.pack(fill="both",expand=True,padx=2,pady=2)
            ctk.CTkLabel(frame,text=msg,font=("Segoe UI",12,"bold"),text_color=color).pack(padx=14,pady=12)
            toast.after(3000,toast.destroy)
        except Exception as _e: _log_silent_err(exc=_e)
    def _loop(self): self.refresh()

    def _compare_jminus7(self,hist,today_data,key):
        """Compare la valeur du jour en cours avec le même jour il y a 7 jours, à équivalent de caisses."""
        if not today_data or not today_data.get("en_cours"):
            return None
        nb_caisses=today_data.get("nb_caisses",0)
        if nb_caisses==0: return None
        # Trouver le jour J-7 (même jour de la semaine, semaine précédente)
        target_date=date.today()-timedelta(days=7)
        target_dd=target_date.strftime("%d/%m")
        ref_day=None
        for h in hist:
            if target_dd in h.get("label","") and not h.get("en_cours"):
                ref_day=h;break
        if not ref_day or not ref_day.get("caisses"): return None
        # Sommer les N premières caisses du jour J-7
        ref_caisses=ref_day.get("caisses",{})
        ref_sum=0
        for i in range(1,nb_caisses+1):
            c=ref_caisses.get(str(i),{})
            ref_sum+=sf(c.get(key,0))
        return ref_sum

    def _upd_vigs(self,d):
        gp=d.get("gp",{});hist=d.get("hist",[]);alerts=d.get("alerts",{});auto=d.get("auto",{})
        complete=[h for h in hist if not h.get("en_cours")];j1=complete[-1] if complete else {};j2=complete[-2] if len(complete)>=2 else {}
        today_data=get_current_partial(hist) or {}
        # === CALCUL DES ÉCARTS DE CAISSE (7 derniers jours + jour en cours) ===
        # On inclut le jour en cours pour ne pas attendre la fin de journée pour
        # voir un écart anormal apparaître. Les caisses non saisies (tout à 0) ne
        # passent pas le filtre |net|>10€ et sont donc naturellement ignorées.
        resolus=load_json(ECARTS_FILE) or {}
        ecarts_anormaux=[]
        jours_a_analyser=list(complete[-7:])
        if today_data and today_data.get("caisses"):
            jours_a_analyser.append(today_data)
        for h in jours_a_analyser:
            lbl=h.get("label","")
            for cnum,c in (h.get("caisses",{}) or {}).items():
                net=sf(c.get("ecart_net",0))
                key=f"{lbl}_C{cnum}"
                if abs(net)<=10: continue
                # Vérifier statut résolu : si résolu mais valeur a changé > 1€ ou
                # changement de signe, on remet en alerte (cohérent avec le détail).
                entry=resolus.get(key)
                if entry:
                    if isinstance(entry,dict) and "valeur" in entry:
                        ancienne=sf(entry.get("valeur",0))
                        if abs(net-ancienne)>1 or (ancienne*net<0):
                            # Modifié depuis marquage : redevient alerte
                            ecarts_anormaux.append({"jour":lbl,"caisse":cnum,"net":net})
                        # sinon : stable, reste résolu
                    # ancien format : reste résolu
                else:
                    ecarts_anormaux.append({"jour":lbl,"caisse":cnum,"net":net})
        # Statut GEST PISTE
        max_ecart=max((abs(e["net"]) for e in ecarts_anormaux),default=0)
        if max_ecart>15 or len(ecarts_anormaux)>3: gp_status="alert"
        elif max_ecart>10 or ecarts_anormaux: gp_status="warn"
        else: gp_status="ok"

        # GEST PISTE → utiliser LITRAGE comme source de vérité pour jour/caisse
        if gp.get("st")=="ok":
            # Jour affiché = jour RÉEL des caisses en cours (extrait du label LITRAGE)
            # Pas le jour calendaire — sinon "Mercredi C2/3" alors que les caisses sont mardi
            if today_data and today_data.get("label"):
                _raw=today_data["label"].split()[0].lower() if today_data["label"].split() else ""
                _map={"lun":"Lundi","lund":"Lundi","lundi":"Lundi",
                      "mar":"Mardi","mard":"Mardi","mardi":"Mardi",
                      "mer":"Mercredi","merc":"Mercredi","mercredi":"Mercredi",
                      "jeu":"Jeudi","jeud":"Jeudi","jeudi":"Jeudi",
                      "ven":"Vendredi","vend":"Vendredi","vendredi":"Vendredi",
                      "sam":"Samedi","samedi":"Samedi",
                      "dim":"Dimanche","dimanche":"Dimanche"}
                li_jour=_map.get(_raw,jour_fr().capitalize())
            else:
                li_jour=jour_fr().capitalize()
            nb_today=int(today_data.get("nb_caisses",0)) if today_data else 0
            li_caisse=str(nb_today)
            cross_alert=False
            # Croiser avec gest_piste pour détecter fiche décalée
            gp_caisse=str(gp.get('caisse','?'))
            try: gp_caisse_int=int(gp_caisse)
            except: gp_caisse_int=-1
            if nb_today>0 and gp_caisse_int>=0 and gp_caisse_int!=nb_today:
                cross_alert=True
            if nb_today==0:
                label=f"{li_jour} C0/3 \u2014 en attente"
                col=C["amber"]
            else:
                label=f"{li_jour} C{nb_today}/3"
                col=C["amber"] if cross_alert else C["t1"]
            lines=[("Jour / Caisse",label,col)]
            if cross_alert:
                lines.append(("\u26a0 Fiche d\u00e9cal\u00e9e",f"GP={gp_caisse}",C["amber"]))
            if today_data and today_data.get("bout",0)>0:
                lines.append(("Boutique jour",feur(today_data.get("bout"),d=0),C["gold"]))
            elif j1 and j1.get("bout",0)>0:
                lines.append(("Boutique J-1",feur(j1.get("bout"),d=0),C["t1"]))
            if ecarts_anormaux:
                pire=max(ecarts_anormaux,key=lambda e:abs(e["net"]))
                lines.append((f"\u26a0 \u00c9cart C{pire['caisse']} {pire['jour']}",feur(pire["net"],d=0),C["red"] if abs(pire["net"])>15 else C["amber"]))
            else:
                lines.append(("\u00c9carts caisse","\u2713 OK",C["green"]))
            self.vigs["gest_piste"].set_data(lines)
            self.vigs["gest_piste"].set_status(gp_status)
        else:
            self.vigs["gest_piste"].set_data([("\u00c9tat","Non disponible",C["t3"])])
            self.vigs["gest_piste"].set_status("ok")

        # CARTES → filtrage > 100€ et > 3j basé sur le vrai parsing
        ca=d.get("ca",{})
        if ca.get("st")=="ok":
            lines=[]
            if complete:
                import re as _re2
                _today=date.today();_m=_today.month;_y=_today.year%100
                def _im(lbl):
                    mm=_re2.search(r'(\d{1,2})/(\d{1,2})/(\d{2})',str(lbl or ""))
                    return mm and int(mm.group(2))==_m and int(mm.group(3))==_y
                _mois=[h for h in complete if _im(h.get("label",""))]
                cb_mois=sum(h.get("cb",0) for h in _mois)
                cp_mois=sum(h.get("cp",0) for h in _mois)
                lines.append((f"CB mois ({len(_mois)}j)",feur(cb_mois,d=0),C["t1"]))
                lines.append(("CP mois",feur(cp_mois,d=0),C["t1"]))
            critical=ca.get("critical",[])
            nb_critique=len(critical)
            if nb_critique>3: ca_status="alert"
            elif nb_critique>0: ca_status="warn"
            else: ca_status="ok"
            if nb_critique>0:
                lines.append((f"\u26a0 Op. > 100\u20ac en retard",f"{nb_critique}",C["red"] if nb_critique>3 else C["amber"]))
                # Afficher le plus ancien pour donner un indice
                if critical:
                    oldest=max((c for c in critical if c.get("age")),key=lambda x:x["age"],default=None)
                    if oldest:
                        lines.append((f"Plus anc. : {oldest['age']}j",feur(oldest['montant'],d=0),C["red"]))
            else:
                lines.append(("Pointage banque","\u2713 OK",C["green"]))
            self.vigs["cartes"].set_data(lines)
            self.vigs["cartes"].set_status(ca_status)
        else:
            self.vigs["cartes"].set_data([("\u00c9tat","Non disponible",C["t3"])])
            self.vigs["cartes"].set_status("ok")

        # PRÉVISION
        pv=d.get("pv",{})
        if pv.get("st")=="ok":
            lines=[]
            # Alerte fraîcheur du bilan
            fresh=pv.get("bilan_freshness","unknown")
            if fresh!="today":
                lines.append(("\u26a0 Bilan non actualis\u00e9","veille",C["amber"]))
            # Détails par carburant SANS total agrégé
            sp_j=auto.get("sp",0);go_j=auto.get("go",0);gnr_j=auto.get("gnr",0)
            min_auto=min(sp_j,go_j,gnr_j)
            if min_auto<1: pv_status="alert"
            elif min_auto<2: pv_status="warn"
            else: pv_status="ok"
            for nm,k,j in [("SP","sp",sp_j),("GO","go",go_j),("GNR","gnr",gnr_j)]:
                jc=C["green"] if j>2 else C["amber"] if j>1 else C["red"]
                lines.append((nm,f"{fnum(sf(pv.get(k)),'L')} \u00b7 {fmt_autonomie(j)}",jc))
            self.vigs["prevision"].set_data(lines)
            self.vigs["prevision"].set_status(pv_status)
        else:
            self.vigs["prevision"].set_data([("\u00c9tat","Non disponible",C["t3"])])
            self.vigs["prevision"].set_status("ok")

        # OBJECTIF
        ob=d.get("ob",{})
        cp_urg=len([c for c in alerts.get("cp_pending",[]) if c.get("retard",0)>0])
        cp_n=len(alerts.get("cp_pending",[]));cli_total=alerts.get("cli_total",0)
        cli_n=len(alerts.get("clients_impayes",[]))
        bal_de=sf(alerts.get("balance_de",0))
        dec_pending=alerts.get("dec_pending",[])
        dec_pend_tot=sum(d2["montant"] for d2 in dec_pending)
        dec_urgent=[d2 for d2 in dec_pending if d2.get("reste") is not None and d2["reste"]<=7]
        if cp_urg>2 or cli_total>10000 or bal_de<0: ob_status="alert"
        elif cp_urg>0 or cli_n>0 or dec_urgent: ob_status="warn"
        else: ob_status="ok"
        if ob.get("st")=="ok":
            lines=[]
            # Balance D/E en tête (indicateur pilotage principal)
            bc=C["green"] if bal_de>0 else C["red"]
            lines.append(("Balance D/E",feur(bal_de,d=0),bc))
            if cp_urg>0: lines.append((f"\u26a0 {cp_urg} CP en retard",feur(sum(c["montant"] for c in alerts.get("cp_pending",[]) if c.get("retard",0)>0),d=0),C["red"]))
            if dec_pend_tot>0:
                dec_col=C["red"] if dec_urgent else C["amber"]
                lbl=f"\u26a0 D\u00e9c. \u00e0 venir ({len(dec_pending)})" if dec_urgent else f"D\u00e9c. \u00e0 venir ({len(dec_pending)})"
                lines.append((lbl,feur(dec_pend_tot,d=0),dec_col))
            # Encaissements en retard
            enc_pending=alerts.get("enc_pending",[])
            enc_retard=[e for e in enc_pending if e.get("reste") is not None and e["reste"]<0]
            if enc_retard:
                enc_ret_tot=sum(e["montant"] for e in enc_retard)
                lines.append((f"\u26a0 Enc. en retard ({len(enc_retard)})",feur(enc_ret_tot,d=0),C["red"]))
            if cli_n>0: lines.append((f"R\u00e8glements en attente ({cli_n})",feur(cli_total,d=0),C["red"]))
            t=sf(ob.get("taux"));tc=C["green"] if t>0.25 else C["amber"] if t>0.10 else C["red"]
            lines.append(("Avancement CA",fpct(t),tc))
            self.vigs["objectif"].set_data(lines)
            self.vigs["objectif"].set_status(ob_status)
        else:
            self.vigs["objectif"].set_data([("\u00c9tat","Non disponible",C["t3"])])
            self.vigs["objectif"].set_status("ok")

        # LITRAGE (en litres — c'est le sujet de cette vignette)
        li=d.get("li",{})
        if li.get("st")=="ok" and complete:
            best=max(complete,key=lambda x:x.get("litrage",0));worst=min(complete,key=lambda x:x.get("litrage",0))
            avg=sum(h.get("litrage",0) for h in complete)/len(complete) if complete else 0
            # Tendance 7j vs 7j précédents (en L)
            last7=complete[-7:] if len(complete)>=7 else complete
            prev7=complete[-14:-7] if len(complete)>=14 else []
            li_status="ok"
            if prev7:
                a1=sum(h.get("litrage",0) for h in last7)/len(last7)
                a2=sum(h.get("litrage",0) for h in prev7)/len(prev7)
                if a2>0 and (a1-a2)/a2<-0.10: li_status="warn"
            self.vigs["litrage"].set_data([("Moy. /jour",fnum(avg,"L"),C["t1"]),
                ("\u2b06 Meilleur",f"{best.get('label','')}  \u2022  {fnum(best.get('litrage'),'L')}",C["green"]),
                ("\u2b07 Plus faible",f"{worst.get('label','')}  \u2022  {fnum(worst.get('litrage'),'L')}",C["amber"])])
            self.vigs["litrage"].set_status(li_status)
        else:
            self.vigs["litrage"].set_data([("\u00c9tat","Non disponible",C["t3"])])
            self.vigs["litrage"].set_status("ok")
        # Sauvegarder pour le détail
        self.last_data["ecarts_anormaux"]=ecarts_anormaux

    def _update_prevision_banner(self,d):
        """Affiche ou cache la bannière d'erreur Prévision selon le code d'erreur remonté par _read_pv.
        Sans cette bannière, le hub continuait silencieusement avec des données vides quand
        Pre_vision_compte.xlsx était verrouillé, et l'utilisateur ne savait pas que ses calculs
        anti-rupture étaient partiels."""
        pv=d.get("pv",{}) or {}
        err=pv.get("_error") if isinstance(pv,dict) else None
        if err in ("locked","notfound","read"):
            msg_map={
                "locked":"\u26a0 Pre_vision_compte.xlsx est verrouill\u00e9 (ouvert dans Excel ou en cours de synchro OneDrive). Les calculs d'autonomie et la d\u00e9tection anti-rupture utilisent des donn\u00e9es incompl\u00e8tes. Ferme Excel puis \u21bb Actualiser.",
                "notfound":"\u26a0 Pre_vision_compte.xlsx introuvable. V\u00e9rifie le chemin dans \u2699 Param\u00e8tres.",
                "read":"\u26a0 Erreur de lecture du fichier Pr\u00e9vision (format ou contenu inattendu). Voir errors.log pour le d\u00e9tail.",
            }
            try:
                self.banner_prev_lbl.configure(text=msg_map.get(err,"\u26a0 Pr\u00e9vision indisponible"))
                self.banner_prev.grid()
            except Exception as _e: _log_silent_err(exc=_e)
        else:
            try: self.banner_prev.grid_remove()
            except Exception as _e: _log_silent_err(exc=_e)

    def _upd_hist(self,d):
        for w in self.hist_frame.winfo_children(): w.destroy()
        hist=d.get("hist",[])
        # Reset miroir API
        self._hist_rows=[]
        if not hist: return
        ctk.CTkLabel(self.hist_frame,text="HISTORIQUE DES DERNIERS JOURS",font=("Segoe UI",11,"bold"),text_color=C["gold"]).pack(anchor="w",pady=(0,8))
        table=ctk.CTkFrame(self.hist_frame,fg_color=C["card"],corner_radius=10,border_width=1,border_color=C["border"]);table.pack(fill="x")
        cols=["Jour","SP","GO","GNR","Total L","CA Piste","CB","CP","Boutique","Total"]
        data_keys=["","sp","go","gnr","litrage","piste","cb","cp","bout","total"]
        widths=[130,80,80,80,95,95,95,95,95,95]
        # Header row - mêmes cadres que les cellules pour alignement parfait
        hf=ctk.CTkFrame(table,fg_color=C["panel"],corner_radius=0);hf.pack(fill="x",pady=(0,2))
        for i,col in enumerate(cols):
            w=widths[i]
            cell=ctk.CTkFrame(hf,fg_color="transparent",width=w,height=30)
            cell.pack(side="left",padx=2,pady=4);cell.pack_propagate(False)
            ctk.CTkLabel(cell,text=col,font=("Segoe UI",9,"bold"),text_color=C["t2"],anchor="e" if i>0 else "w").pack(fill="both",expand=True,padx=4)
        # Préparer les jours : EN COURS d'abord, puis 7 derniers complets en ordre antichrono
        _ec=get_current_partial(hist)
        en_cours_today=[_ec] if _ec else []
        complets_recents=list(reversed([h for h in hist if not h.get("en_cours")][-7:]))
        shown=en_cours_today+complets_recents
        # Calculer min/max par colonne parmi les jours complets
        col_max={};col_min={}
        if len(complets_recents)>1:
            for ki in range(1,len(data_keys)):
                k=data_keys[ki]
                vals=[d.get(k,0) for d in complets_recents]
                if len(vals)>1:
                    col_max[ki]=max(vals);col_min[ki]=min(vals)
                    if col_max[ki]==col_min[ki]: col_max[ki]=None;col_min[ki]=None
        for j,day in enumerate(shown):
            is_en_cours=day.get("en_cours",False)
            # EN COURS : fond bleu très subtil pour la distinguer
            if is_en_cours: bg="#142030"
            else: bg=C["card"] if j%2==0 else C["panel"]
            rf=ctk.CTkFrame(table,fg_color=bg,corner_radius=0);rf.pack(fill="x")
            lbl=day.get("label","\u2014")
            if is_en_cours:
                nb=day.get("nb_caisses",0)
                lbl=f"\u25cf {lbl} (C{nb}/3)"
            vals=[lbl,fnum(day.get("sp")),fnum(day.get("go")),fnum(day.get("gnr")),fnum(day.get("litrage"),"L"),
                  feur(day.get("piste"),d=0),feur(day.get("cb"),d=0),feur(day.get("cp"),d=0),feur(day.get("bout"),d=0),feur(day.get("total"),d=0)]
            is_complete=not is_en_cours and day.get("total",0)>0
            # Capture API : cellules avec meta best/worst
            api_cells=[]
            for i,val in enumerate(vals):
                is_best=is_complete and i in col_max and col_max.get(i) is not None and day.get(data_keys[i],0)==col_max[i]
                is_worst=is_complete and i in col_min and col_min.get(i) is not None and day.get(data_keys[i],0)==col_min[i]
                api_cells.append({"value":val,"is_best":bool(is_best and i>0),"is_worst":bool(is_worst and i>0),"is_label":i==0})
            self._hist_rows.append({"is_en_cours":bool(is_en_cours),"cells":api_cells})
            for i,val in enumerate(vals):
                w=widths[i]
                is_best=is_complete and i in col_max and col_max.get(i) is not None and day.get(data_keys[i],0)==col_max[i]
                is_worst=is_complete and i in col_min and col_min.get(i) is not None and day.get(data_keys[i],0)==col_min[i]
                if is_best and i>0:
                    cell_bg="#1B4A2A";txt_color="#FFFFFF"
                elif is_worst and i>0:
                    cell_bg="#4A1B1B";txt_color="#FFFFFF"
                else:
                    cell_bg=bg
                    if i==0:
                        txt_color=C["amber"] if is_en_cours else C["gold"]
                    else:
                        txt_color="#9DD7FF" if is_en_cours else C["t1"]
                cell=ctk.CTkFrame(rf,fg_color=cell_bg,corner_radius=4,width=w,height=28)
                cell.pack(side="left",padx=2,pady=2);cell.pack_propagate(False)
                ctk.CTkLabel(cell,text=val,font=("Segoe UI",10,"bold" if (is_best or is_worst) and i>0 else "normal"),text_color=txt_color,anchor="e" if i>0 else "w").pack(fill="both",expand=True,padx=4)

    def _upd_banner(self,d):
        gp=d.get("gp",{});pv=d.get("pv",{});ob=d.get("ob",{});alerts=d.get("alerts",{});auto=d.get("auto",{})
        hist=d.get("hist",[]);complete=[h for h in hist if not h.get("en_cours")]
        j1=complete[-1] if len(complete)>=1 else {};j2=complete[-2] if len(complete)>=2 else {}
        today_data=get_current_partial(hist) or {}
        dt=gp.get("date");jour=gp.get("jour","")
        ds=dt.strftime("%d/%m/%Y") if isinstance(dt,datetime) else str(dt)[:10] if dt else "\u2014"
        date_text=f"{jour} {ds}".strip()
        self.b_date.configure(text=date_text)
        self._b_date_text=date_text  # miroir API
        j1_l=j1.get("label","?") if j1 else "?";caisse=gp.get("caisse","\u2014")
        sub=f"Caisse : {caisse}"
        if today_data:
            if today_data.get("en_cours"): sub+=f" \u2502 Jour EN COURS ({today_data.get('litrage',0):,.0f} L)".replace(","," ")
            else: sub+=f" \u2502 Jour COMPLET ({today_data.get('litrage',0):,.0f} L)".replace(","," ")
        sub+=f" \u2502 J-1 : {j1_l}"
        self.b_sub.configure(text=sub)
        self._b_sub_text=sub  # miroir API
        # Carburants
        self.s_carb.clear()
        src=today_data if today_data and today_data.get("litrage",0)>0 else j1
        is_today_partial=src==today_data and today_data.get("en_cours",False)
        # Quand jour EN COURS : comparer à J-7 même nb de caisses. Sinon : comparer à J-1.
        def cmp_for(key):
            if is_today_partial:
                ref=self._compare_jminus7(hist,today_data,key)
                return trend(src.get(key),ref) if ref is not None else ("",C["t3"])
            else:
                cmp=j1 if src==today_data else j2
                return trend(src.get(key),cmp.get(key) if cmp else None)
        if is_today_partial:
            tag=f"jour C{today_data.get('nb_caisses',0)}"
        else:
            tag="jour" if src==today_data else "J-1"
        if src:
            self.s_carb.row(f"SP ({tag})",fnum(src.get("sp"),"L"),C["t1"],tr=cmp_for("sp"))
            self.s_carb.row(f"GO ({tag})",fnum(src.get("go"),"L"),C["t1"],tr=cmp_for("go"))
            self.s_carb.row(f"GNR ({tag})",fnum(src.get("gnr"),"L"),C["t1"],tr=cmp_for("gnr"))
            self.s_carb.sep()
            self.s_carb.row(f"Total {tag}",fnum(src.get("litrage"),"L"),C["gold"],big=True,tr=cmp_for("litrage"))
            self.s_carb.row(f"CA piste {tag}",feur(src.get("piste"),d=0),C["teal"],big=True,tr=cmp_for("piste"))
            if is_today_partial:
                self.s_carb.row("vs m\u00eame jour J-7","",C["t3"])
        # Encaissements
        self.s_enc.clear()
        if src:
            self.s_enc.row(f"CB ({tag})",feur(src.get("cb"),d=0),C["t1"],tr=cmp_for("cb"))
            self.s_enc.row(f"CP ({tag})",feur(src.get("cp"),d=0),C["t1"],tr=cmp_for("cp"))
            self.s_enc.row(f"Esp\u00e8ces ({tag})",feur(src.get("esp"),d=0),C["t1"])
            self.s_enc.sep()
            self.s_enc.row(f"Total {tag}",feur(src.get("total"),d=0),C["green"],big=True,tr=cmp_for("total"))
        # Boutique
        self.s_bout.clear()
        if src and src.get("bout",0)>0:
            self.s_bout.row(f"CA Boutique {tag}",feur(src.get("bout"),d=0),C["gold"],big=True,tr=cmp_for("bout"))
        # Stocks avec autonomie intelligente
        self.s_stock.clear()
        if pv.get("st")=="ok":
            for nm,k in [("SP","sp"),("GO","go"),("GNR","gnr")]:
                s=sf(pv.get(k));j=auto.get(k,0)
                jc=C["green"] if j>2 else C["amber"] if j>1 else C["red"]
                self.s_stock.row(nm,f"{fnum(s,'L')}  ({fmt_autonomie(j)})",jc)
            self.s_stock.sep();self.s_stock.bar("Valorisation",sf(pv.get("valo")),80000,C["gold"])
            if pv.get("livr_recu"): self.s_stock.row("\U0001f69a Livraison","\u2713 Prise en compte",C["green"])
        # Objectif
        self.s_obj.clear()
        if ob.get("st")=="ok":
            obj_ca=sf(ob.get("obj_ca"));enc_ca=sf(ob.get("enc_ca"));t=sf(ob.get("taux"))
            tc=C["green"] if t>0.25 else C["amber"] if t>0.10 else C["red"]
            self.s_obj.bar("Encours CA",enc_ca,obj_ca,tc)
            self.s_obj.row("Reste \u00e0 faire",feur(obj_ca-enc_ca,d=0),C["t1"])
            today_d=date.today()
            if today_d.day>0 and enc_ca>0:
                proj=enc_ca/today_d.day*30;pc=C["green"] if proj>=obj_ca else C["amber"]
                self.s_obj.row("Projection 30j",feur(proj,d=0),pc,big=True)
            bal=sf(ob.get("balance"));bc=C["green"] if bal>0 else C["red"]
            self.s_obj.sep();self.s_obj.row("Balance D/E",feur(bal,d=0),bc,big=True)
        # Alertes
        self.s_alert.clear()
        cp_list=alerts.get("cp_pending",[]);cp_urg=[c for c in cp_list if c.get("retard",0)>0]
        if cp_urg:
            total_urg=sum(c["montant"] for c in cp_urg)
            self.s_alert.alert(f"\u26a0 {len(cp_urg)} CP en retard pour {feur(total_urg,d=0)}\n"+"\n".join(f"  \u2022 {c['date']} : {feur(c['montant'],d=0)} ({c['retard']}j)" for c in cp_urg),C["red"])
        cp_ec=[c for c in cp_list if c.get("retard",0)<=0]
        if cp_ec: self.s_alert.alert(f"\U0001f4b3 {len(cp_ec)} CP \u00e0 venir pour {feur(sum(c['montant'] for c in cp_ec),d=0)}",C["amber"])
        bc=alerts.get("by_client",{})
        if bc:
            for nom,data in sorted(bc.items(),key=lambda x:x[1]["max_age"],reverse=True)[:5]:
                age=data["max_age"];mois=age//30;age_s=f"{mois} mois" if mois>0 else f"{age}j"
                self.s_alert.alert(f"\U0001f464 {nom} : {feur(data['total'],d=0)} ({data['count']} fact., {age_s})",C["red"] if age>90 else C["amber"])
        # Alertes mots-clés (rejet, saisie, urgent, etc.)
        kw_alerts=alerts.get("keyword_alerts",[])
        # Dédupliquer par nom+montant
        seen=set()
        for ka in kw_alerts:
            key=f"{ka['nom']}_{ka['montant']}"
            if key not in seen and ka['nom']:
                seen.add(key)
                mode_str=f" {ka['mode']}" if ka.get('mode') else ""
                # Tag affiché : si la cellule contient "retour saisie", afficher ça plutôt que juste "saisie"
                tag=ka['mot']
                if ka['mot']=='saisie':
                    for f in ('cell','info'):
                        v=str(ka.get(f,'')).lower()
                        if 'retour saisie' in v: tag='retour saisie';break
                self.s_alert.alert(f"\u26a0 {ka['onglet']} : {ka['nom']} {feur(ka['montant'],d=0)}{mode_str} [{tag}]",C["red"])
        # === ALERTES PROJECTION 14J (rupture stock + anomalie ventes) ===
        proj14=d.get("proj14",{})
        if proj14.get("alertes"):
            # Séparer ruptures réelles et fin de cycle
            ruptures=[a for a in proj14["alertes"] if a["severity"]=="rupture"]
            fin_cycles=[a for a in proj14["alertes"] if a["severity"]=="fin_cycle"]
            # Afficher ruptures (rouge)
            for al in ruptures:
                txt=f"\u26fd {al['carburant']} : rupture {al['date_str']} \u2014 commander avant {al['deadline_str']}"
                self.s_alert.alert(txt,C["red"])
            # Afficher fin de cycle (groupé, gris/bleu = info)
            if fin_cycles:
                carbs=", ".join(a["carburant"] for a in fin_cycles)
                last_d=fin_cycles[0].get("last_livr_date_str","")
                self.s_alert.alert(f"\U0001f4c5 Cycle Prévision se termine {last_d} \u2014 remplir le suivant ({carbs})",C["t2"])
        # Anomalies de ventes du jour
        if proj14.get("anomalies"):
            _d_anom=date.today().strftime("%Y-%m-%d")
            for carb,anom in proj14["anomalies"].items():
                signe="+" if anom["ecart_pct"]>0 else ""
                col=C["amber"] if abs(anom["ecart_pct"])<40 else C["red"]
                self.s_alert.alert(f"\U0001f4ca {carb.upper()} : ventes {signe}{anom['ecart_pct']}% \u00e0 {anom['stage']} vs moyenne",col)
                # Trace durable (validée Bidou 31/05/2026) : chaque anomalie affichée dans
                # la vignette laisse une trace dans le journal + tableau de notifications,
                # via le type "tendance" déjà géré en aval. Idempotence native du fingerprint
                # (carburant + date) → 1 trace par carburant par jour, aucun doublon au refresh.
                try:
                    add_evenement("tendance",{
                        "carburant":carb,
                        "date":_d_anom,
                        "ecart_pct":anom["ecart_pct"],
                        "stage":anom.get("stage",""),
                        "statut":"non_traite",
                        "lu":False,
                    },commentaire=f"{carb.upper()} : ventes {signe}{anom['ecart_pct']}% \u00e0 {anom.get('stage','')} vs moyenne")
                except Exception as _e: _log_silent_err(exc=_e)

# =============================================================================
# POPUP D'ALERTE ANTI-RUPTURE — déclenché si severite_max == 'critique'
# Réapparaît à chaque refresh tant que la rupture / incohérence n'est pas corrigée.
class AntiRuptureDlg(ctk.CTkToplevel):
    def __init__(self,parent,antirupture):
        super().__init__(parent)
        self.title("\U0001f6a8 Alerte anti-rupture \u2014 DISTRICARB HUB")
        self.geometry("780x680");self.minsize(700,560)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.ar=antirupture
        # Header rouge
        hdr=ctk.CTkFrame(self,fg_color=C["alert_bg"],corner_radius=0,height=80,border_width=0);hdr.pack(fill="x");hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f6a8",font=("Segoe UI Emoji",32),text_color=C["red"]).pack(side="left",padx=(24,12),pady=18)
        title_box=ctk.CTkFrame(hdr,fg_color="transparent");title_box.pack(side="left",fill="y",pady=14)
        ctk.CTkLabel(title_box,text="ALERTE ANTI-RUPTURE",font=("Segoe UI",16,"bold"),text_color=C["red"],anchor="w").pack(anchor="w")
        ctk.CTkLabel(title_box,text="Action requise sur ton fichier Pr\u00e9vision compte.xlsx",font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(anchor="w")
        # Corps scrollable
        body=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        body.pack(fill="both",expand=True,padx=20,pady=(12,0))
        # Encart d'explication (toujours visible en haut de la popup)
        info_box=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=8,border_width=1,border_color=C["border2"])
        info_box.pack(fill="x",padx=4,pady=(4,12))
        ctk.CTkLabel(info_box,text="\u2139\ufe0f Comment lire cette alerte",font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="w").pack(anchor="w",padx=12,pady=(8,2))
        ctk.CTkLabel(info_box,text="Les manques affich\u00e9s sont calcul\u00e9s APR\u00c8S int\u00e9gration de toutes tes commandes\nd\u00e9j\u00e0 saisies dans Pre_vision. Si un manque appara\u00eet, c'est qu'il faut commander\nDU CARBURANT EN PLUS de ce que tu as d\u00e9j\u00e0 pr\u00e9vu pour couvrir le pont.",
                     font=("Segoe UI",10),text_color=C["t2"],anchor="w",justify="left").pack(anchor="w",padx=12,pady=(0,10))
        # Section 1 : Incohérences (commande sur jour non-livrable)
        incoh=self.ar.get("incoherences_jour_non_livrable",[])
        if incoh:
            self._section_title(body,"\u26a0 Commandes saisies sur jour non-livrable",C["red"])
            # Mention "résidu probable" (Fix A 25/05/2026) : ces commandes sont presque toujours
            # des dates oubliées dans Pre_vision après bascule de cycle S1/S2. Préciser
            # le bon réflexe : corriger le fichier source plutôt que forcer dans le hub.
            ctk.CTkLabel(body,
                text="\U0001f4a1 Ces commandes sont souvent des r\u00e9sidus de ton fichier Pr\u00e9vision compte.xlsx "
                     "(date oubli\u00e9e apr\u00e8s bascule de cycle, dimanche/f\u00e9ri\u00e9 non corrig\u00e9). "
                     "Le bon r\u00e9flexe : corriger directement dans le fichier source plut\u00f4t que forcer dans le hub.",
                font=("Segoe UI",10,"italic"),text_color=C["t3"],anchor="w",
                justify="left",wraplength=580).pack(anchor="w",padx=18,pady=(0,8))
            for i in incoh:
                txt=f"{i['date_str']} ({i['raison']}) \u2014 SP {int(i['sp']):,}L | GO {int(i['go']):,}L | GNR {int(i['gnr']):,}L".replace(",",".")
                box=ctk.CTkFrame(body,fg_color=C["alert_bg"],corner_radius=8,
                                 border_width=1,border_color=C["alert_border"])
                box.pack(fill="x",padx=4,pady=(2,8))
                ctk.CTkLabel(box,text=txt,font=("Segoe UI",12,"bold"),text_color=C["t1"],
                             anchor="w",justify="left").pack(anchor="w",padx=14,pady=(10,4))
                ctk.CTkLabel(box,
                    text="La SARA ne livre pas ce jour. Replanifie sur un jour ouvr\u00e9, "
                         "OU force-la si cette livraison est volontaire et assum\u00e9e.",
                    font=("Segoe UI",11),text_color=C["t2"],anchor="w",
                    justify="left",wraplength=560).pack(anchor="w",padx=14,pady=(0,8))
                btnrow=ctk.CTkFrame(box,fg_color="transparent")
                btnrow.pack(anchor="w",padx=14,pady=(0,10))
                ctk.CTkButton(btnrow,text="\u2713 Forcer cette livraison (exceptionnel)",
                    width=260,height=32,
                    fg_color="#2A2315",hover_color="#3A3020",text_color=C["amber"],
                    border_width=1,border_color=C["amber"],font=("Segoe UI",10,"bold"),
                    corner_radius=6,
                    command=lambda dd=i.get("date"),vol=i.get("volume_total",0):self._forcer_livraison_exceptionnelle(dd,vol)
                    ).pack(side="left")
        # Section 2 : Ruptures par pont avec ACQUITTEMENT INDIVIDUEL
        ruptures=self.ar.get("ruptures_dans_trou",[])
        plans=self.ar.get("plan_lisse",[]) or []
        ack_status=self.ar.get("ack_status",{}) or {}
        if ruptures:
            # Grouper ruptures + plan par pont_id
            par_pont={}
            for r in ruptures:
                pont_id=f"pont_{r['trou_start'].strftime('%d%m%Y')}"
                # Qualification Pont (contient férié) vs Weekend (sam+dim sans férié).
                # Factorisée (Étape 3, 27/05/2026) via districarb_core.trous.qualifier_trou.
                par_pont.setdefault(pont_id,{
                    "trou_str":r["trou_str"],"trou_duree":r["trou_duree"],
                    "trou_start":r["trou_start"],"deadline_str":r["deadline_str"],
                    "deadline":r["deadline"],"manques":[],"plan":None,
                    "terme":qualifier_trou({"start_date":r["trou_start"],"duree":r["trou_duree"]}),
                })
                manque_arrondi=max(2000,int(((r["manque"]+999)//1000)*1000))
                par_pont[pont_id]["manques"].append((r["carburant"],manque_arrondi,r["manque"]))
            for p in plans:
                pont_id=f"pont_{p['trou_start'].strftime('%d%m%Y')}"
                if pont_id in par_pont:
                    par_pont[pont_id]["plan"]=p
            # === HIÉRARCHIE VISUELLE ===
            # Trier les ponts pour mettre les NON ACQUITTÉS en premier (à régler maintenant),
            # puis les acquittés (sous contrôle) en bas avec un séparateur visuel.
            # Évite que les gros encarts verts "sous contrôle" éclipsent les ponts rouges
            # qui demandent vraiment une action.
            pont_ids_a_regler=[]
            pont_ids_sous_controle=[]
            for pont_id in par_pont:
                if ack_status.get(pont_id,{}).get("acquitte",False):
                    pont_ids_sous_controle.append(pont_id)
                else:
                    pont_ids_a_regler.append(pont_id)
            pont_ids_ordered=pont_ids_a_regler+pont_ids_sous_controle
            # Afficher chaque pont dans son propre cadre avec ses boutons
            for pont_id in pont_ids_ordered:
                info=par_pont[pont_id]
                # Insérer le séparateur "Sous contrôle" juste AVANT le premier pont acquitté
                if pont_ids_sous_controle and pont_id==pont_ids_sous_controle[0] and pont_ids_a_regler:
                    sep_box=ctk.CTkFrame(body,fg_color="transparent");sep_box.pack(fill="x",padx=4,pady=(16,4))
                    sep_line=ctk.CTkFrame(sep_box,fg_color=C["border2"],height=1);sep_line.pack(fill="x",pady=(0,8))
                    ctk.CTkLabel(sep_box,text=f"\u2713 Sous contr\u00f4le ({len(pont_ids_sous_controle)})",
                                 font=("Segoe UI",11,"italic"),text_color=C["t3"],anchor="w").pack(anchor="w")
                st=ack_status.get(pont_id,{})
                acquitte=st.get("acquitte",False)
                deadline_depassee=st.get("deadline_depassee",False)
                raison_inv=st.get("raison_invalidation")
                # Cadre du pont
                container=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=10,border_width=2,
                                       border_color=C["green"] if acquitte else C["alert_border"])
                container.pack(fill="x",padx=4,pady=(8,8))
                # Titre du pont avec badge état
                titre_top=ctk.CTkFrame(container,fg_color="transparent");titre_top.pack(fill="x",padx=14,pady=(12,4))
                if acquitte:
                    titre_txt=f"\u2713 {info['terme']} du {info['trou_str']} \u2014 sous contr\u00f4le"
                    titre_color=C["green"]
                elif deadline_depassee:
                    titre_txt=f"\u26a0 {info['terme']} du {info['trou_str']} \u2014 deadline d\u00e9pass\u00e9e"
                    titre_color=C["amber"]
                else:
                    titre_txt=f"\U0001f6a8 {info['terme']} du {info['trou_str']} \u2014 {info['trou_duree']} jours sans livraison"
                    titre_color=C["red"]
                ctk.CTkLabel(titre_top,text=titre_txt,font=("Segoe UI",13,"bold"),text_color=titre_color,anchor="w").pack(anchor="w")
                # Si invalidation après ack précédent
                if raison_inv=="manque_aggravee" or (raison_inv and raison_inv.startswith("manque_")):
                    ctk.CTkLabel(container,text="\u26a0 Tu avais acquitt\u00e9 ce pont mais la situation s'est aggrav\u00e9e (>20%). \u00c0 rev\u00e9rifier.",
                                 font=("Segoe UI",10,"italic"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(0,4))
                elif raison_inv=="ack_expire_24h":
                    ctk.CTkLabel(container,text="\u23f0 Acquittement de plus de 24h. \u00c0 reconfirmer.",
                                 font=("Segoe UI",10,"italic"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(0,4))
                elif raison_inv=="nouveau_carburant_en_manque":
                    ctk.CTkLabel(container,text="\u26a0 Un nouveau carburant est en manque depuis ton acquittement.",
                                 font=("Segoe UI",10,"italic"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(0,4))
                # Manques par carburant — COMPACTÉ pour ponts acquittés (1 ligne au lieu de N)
                if acquitte:
                    # Format compact en une seule ligne : "SP 3000L · GO 2000L"
                    manques_compact=" \u2022 ".join(f"{carb} {fnum(m,'L')}".replace(",",".") for carb,m,_ in info["manques"])
                    ctk.CTkLabel(container,text=f"   {manques_compact}",font=("Segoe UI",10),text_color=C["t2"],anchor="w").pack(anchor="w",padx=14,pady=(2,4))
                else:
                    manques_txt="\n".join(f"   \u2022 {carb} : il te manque environ {fnum(m,'L')}".replace(",",".")
                                           for carb,m,_ in info["manques"])
                    ctk.CTkLabel(container,text=manques_txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(2,4))
                # Bloc commande à passer (seulement si plan disponible, deadline pas dépassée, ET pont non acquitté)
                # Pour les ponts acquittés on cache le plan détaillé (gain visuel) — Bidou a déjà acquitté, pas besoin de re-voir.
                if not deadline_depassee and info["plan"] and not acquitte:
                    p=info["plan"]
                    ctk.CTkLabel(container,
                                 text=f"\u2192 Passe la commande {p['jour_commande_str']}, pour livraison {p['jour_livraison_str']} :",
                                 font=("Segoe UI",11),text_color=C["t1"],anchor="w").pack(anchor="w",padx=14,pady=(4,2))
                    for ligne in p.get("lignes_carb",[]):
                        tour_hint=" \u2014 \U0001f4cd demander 3e tour" if ligne.get("tour_3") else ""
                        ligne_txt=f"   \u2022 {ligne['carburant']} : commander {fnum(ligne['volume'],'L')}{tour_hint}".replace(",",".")
                        ctk.CTkLabel(container,text=ligne_txt,font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="w").pack(anchor="w",padx=14,pady=1)
                    if p.get("infaisable_carbs"):
                        inf_txt=f"   \u26a0 IMPOSSIBLE pour : {', '.join(p['infaisable_carbs'])} (cuve d\u00e9j\u00e0 satur\u00e9e par tes saisies)"
                        ctk.CTkLabel(container,text=inf_txt,font=("Segoe UI",10),text_color=C["red"],anchor="w").pack(anchor="w",padx=14,pady=1)
                # Boutons d'acquittement : 3 boutons TOUJOURS présents tant que pont non acquitté
                btn_row=ctk.CTkFrame(container,fg_color="transparent");btn_row.pack(fill="x",padx=14,pady=(8,12))
                if acquitte:
                    statut=st.get("type_ack","controle")
                    if statut=="rupture_acceptee":
                        cause=st.get("cause","")
                        msg=f"\u26a0 Rupture accept\u00e9e \u2014 cause : {cause}" if cause else "\u26a0 Rupture accept\u00e9e."
                        ctk.CTkLabel(btn_row,text=msg,font=("Segoe UI",10,"italic"),text_color=C["red"]).pack(side="left")
                    elif statut=="snooze":
                        try:
                            until_dt=datetime.fromisoformat(st.get("snooze_until_iso",""))
                            until_str=until_dt.strftime("%Hh%M")
                            mins=int((until_dt-datetime.now()).total_seconds()/60)
                            if mins>0:
                                ctk.CTkLabel(btn_row,text=f"\u23f1 Rappel d\u00e9sactiv\u00e9 jusqu'\u00e0 {until_str} (dans {mins} min)",
                                             font=("Segoe UI",10,"italic"),text_color=C["amber"]).pack(side="left")
                            else:
                                ctk.CTkLabel(btn_row,text="\u23f1 Snooze expir\u00e9",
                                             font=("Segoe UI",10,"italic"),text_color=C["t3"]).pack(side="left")
                        except Exception:
                            ctk.CTkLabel(btn_row,text="\u23f1 Snooze actif",font=("Segoe UI",10,"italic"),text_color=C["amber"]).pack(side="left")
                    else:
                        ctk.CTkLabel(btn_row,text="\u2713 Tu as acquitt\u00e9 ce pont. Reviendra si la situation change.",
                                     font=("Segoe UI",10,"italic"),text_color=C["green"]).pack(side="left")
                else:
                    # Stocker manques actuels pour comparaison future
                    manques_dict={carb.lower():m_brut for carb,_,m_brut in info["manques"]}
                    # Bouton 1 : C'est sous contrôle (vert)
                    ctk.CTkButton(btn_row,text="\u2713 C'est sous contr\u00f4le",width=170,height=32,
                                  fg_color=C["green"],hover_color="#15943C",text_color="#FFF",
                                  font=("Segoe UI",11,"bold"),corner_radius=6,
                                  command=lambda pid=pont_id,m=manques_dict:self._acquitter_pont(pid,m)).pack(side="left",padx=(0,6))
                    # Bouton 2 : Snooze (ambre) — TOUJOURS visible (Bidou veut pouvoir snoozer même après deadline)
                    ctk.CTkButton(btn_row,text="\u23f1 Me le rappeler dans...",width=180,height=32,
                                  fg_color=C["amber"],hover_color="#C4811D",text_color="#000",
                                  font=("Segoe UI",11,"bold"),corner_radius=6,
                                  command=lambda pid=pont_id,m=manques_dict:self._snooze_pont(pid,m)).pack(side="left",padx=(0,6))
                    # Bouton 3 : Accepter la rupture (rouge sombre) — TOUJOURS visible
                    # (le pont peut être accepté à l'avance si Bidou sait qu'il ne pourra pas s'en sortir)
                    ctk.CTkButton(btn_row,text="\u26a0 J'accepte la rupture",width=170,height=32,
                                  fg_color="#7A1F22",hover_color="#5C1518",text_color="#FFF",
                                  font=("Segoe UI",11,"bold"),corner_radius=6,
                                  command=lambda pid=pont_id,m=manques_dict:self._accepter_rupture(pid,m)).pack(side="left",padx=(0,6))
        # === Fallback : si aucune section n'a été rendue, l'utilisateur a probablement cliqué
        # sur un pont déjà résolu/acquitté ou silencé depuis le tableau. Plutôt qu'une popup
        # quasi-vide qui paraît cassée, on affiche un encart explicatif clair. ===
        has_content=bool(
            self.ar.get("incoherences_jour_non_livrable") or
            self.ar.get("ruptures_dans_trou") or
            self.ar.get("ruptures_imminentes") or
            self.ar.get("ruptures_projetees") or
            self.ar.get("livraisons_non_conformes") or
            self.ar.get("tendance_alertes")
        )
        if not has_content:
            empty_box=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=10,border_width=1,border_color="#2A4A2A")
            empty_box.pack(fill="x",padx=4,pady=20)
            ctk.CTkLabel(empty_box,text="\u2713",font=("Segoe UI",36),text_color=C["green"]).pack(pady=(24,8))
            ctk.CTkLabel(empty_box,text="Aucune alerte anti-rupture active actuellement",
                         font=("Segoe UI",13,"bold"),text_color=C["t1"]).pack()
            ctk.CTkLabel(empty_box,
                         text=("Le pont ou la situation que tu as ouvert(e) depuis le tableau a probablement\n"
                               "d\u00e9j\u00e0 \u00e9t\u00e9 acquitt\u00e9(e), r\u00e9solu(e), ou les manques ont disparu apr\u00e8s correction de Pre_vision.\n\n"
                               "Si tu veux le retirer du tableau de notifications, utilise le bouton \u2713 R\u00e9solu\n"
                               "sur la carte correspondante."),
                         font=("Segoe UI",10),text_color=C["t2"],justify="center").pack(pady=(4,20),padx=20)
        # Footer : bouton fermer
        footer=ctk.CTkFrame(self,fg_color="transparent",height=64);footer.pack(side="bottom",fill="x",padx=20,pady=14);footer.pack_propagate(False)
        ctk.CTkLabel(footer,text="Les ponts non acquitt\u00e9s r\u00e9appara\u00eetront au prochain refresh.",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(side="left",padx=(4,0),pady=14)
        ctk.CTkButton(footer,text="Fermer",width=140,height=40,fg_color=C["panel"],hover_color=C["card"],
                       text_color=C["t1"],font=("Segoe UI",12,"bold"),corner_radius=8,command=self.destroy).pack(side="right")
    def _pont_info_for_id(self,pont_id):
        """Retrouve les infos d'un pont depuis self.ar à partir de son ID."""
        try:
            ruptures=self.ar.get("ruptures_dans_trou",[]) if hasattr(self,'ar') else []
            for r in ruptures:
                rid=f"pont_{r['trou_start'].strftime('%d%m%Y')}"
                if rid==pont_id: return r
        except Exception as _e: _log_silent_err(exc=_e)
        return None

    def _acquitter_pont(self,pont_id,manques_actuels):
        """Enregistre l'acquittement 'sous contrôle' d'un pont avec snapshot des manques actuels.
        La popup ne réapparaîtra que :
         - si les manques augmentent de plus de 20% sur un carburant
         - OU si un nouveau carburant tombe en manque
         - OU si l'acquittement a plus de 24h."""
        try:
            acks=load_json(ANTIRUPTURE_ACK_FILE) or {}
        except Exception: acks={}
        acks[pont_id]={"ack_at_iso":datetime.now().isoformat(),
                       "manques":manques_actuels,
                       "type_ack":"controle"}
        try: save_json(ANTIRUPTURE_ACK_FILE,acks)
        except Exception as e: print(f"[acquittement] {e}")
        # Journal d'événements (Sujet E) : capture le pont acquitté
        try:
            # Retrouver les infos du pont depuis self.ar
            ruptures=self.ar.get("ruptures_dans_trou",[])
            pont_info=None
            for r in ruptures:
                rid=f"pont_{r['trou_start'].strftime('%d%m%Y')}"
                if rid==pont_id:
                    pont_info=r;break
            if pont_info:
                # Demander un commentaire optionnel (Q2 : option b)
                commentaire=self._ask_commentaire_optionnel("Acquitter le pont","Comment as-tu géré ce pont ? (optionnel)")
                add_evenement("pont",{
                    "date_debut":pont_info["trou_start"].isoformat(),
                    "date_fin":(pont_info["trou_start"]+timedelta(days=pont_info.get("trou_duree",1)-1)).isoformat(),
                    "duree":pont_info.get("trou_duree",1),
                    "manques":[{"carburant":c,"manque":m} for c,m in (manques_actuels or {}).items()],
                    "ack_type":"controle_manuel",
                },commentaire=commentaire)
        except Exception as e: print(f"[evt pont] {e}")
        # Silence global popup anti-rupture : par défaut jusqu'à demain matin 6h.
        # On ne remontre la popup que pour une nouvelle alerte ou aggravation.
        try:
            tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
            ar_now=getattr(self.master,"last_data",{}).get("antirupture",{}) or {}
            fps,extra=_antirupture_fps_extra(ar_now)
            silence_popup("antirupture",fps,tomorrow_6h.isoformat(),extra)
        except Exception as e: print(f"[silence antirupture controle] {e}")
        self.destroy()

    def _ask_commentaire_optionnel(self,title,prompt):
        """Mini-dialogue pour saisir un commentaire optionnel. Retourne la chaîne saisie ou None.
        Conformément à Q2 : champ optionnel, on peut juste skip."""
        dlg=ctk.CTkToplevel(self);dlg.title(title);dlg.geometry("520x240")
        dlg.configure(fg_color=C["bg"]);dlg.transient(self);dlg.grab_set()
        result={"value":None}
        ctk.CTkLabel(dlg,text=prompt,font=("Segoe UI",11),text_color=C["t1"],wraplength=480,justify="left").pack(pady=(20,8),padx=20,anchor="w")
        ctk.CTkLabel(dlg,text="Tu peux laisser vide et cliquer Skip si tu n'as rien \u00e0 ajouter.",
                     font=("Segoe UI",9),text_color=C["t3"]).pack(padx=20,anchor="w")
        entry=ctk.CTkEntry(dlg,width=460,height=34,placeholder_text="Ex : commande ajust\u00e9e tot, livraison renforc\u00e9e, rupture accept\u00e9e, etc.")
        entry.pack(pady=12,padx=20)
        entry.focus()
        btns=ctk.CTkFrame(dlg,fg_color="transparent");btns.pack(pady=10)
        def _ok():
            v=entry.get().strip()
            result["value"]=v if v else None
            dlg.destroy()
        def _skip():
            result["value"]=None
            dlg.destroy()
        ctk.CTkButton(btns,text="Skip",width=100,height=34,fg_color=C["panel"],hover_color=C["border2"],
                      text_color=C["t1"],command=_skip).pack(side="left",padx=6)
        ctk.CTkButton(btns,text="OK",width=100,height=34,fg_color=C["green"],hover_color="#0e7c3a",
                      text_color="#fff",command=_ok).pack(side="left",padx=6)
        entry.bind("<Return>",lambda e:_ok())
        dlg.wait_window()
        return result["value"]
    def _snooze_pont(self,pont_id,manques_actuels):
        """Ouvre une mini-fenêtre pour choisir la durée de snooze (1h, 2h, 4h),
        puis enregistre l'acquittement avec snooze_until_iso."""
        # Mini-popup choix durée — structure pack SIMPLE (la version qui marchait à l'origine).
        # On conserve la logique adaptative (deadline TEMAG si pont J+2) ET le fallback garanti,
        # mais on revient au pack pour les widgets — pas de grid qui casse le rendu.
        dlg=ctk.CTkToplevel(self);dlg.title("Me le rappeler dans...")
        dlg.geometry("560x340");dlg.configure(fg_color=C["bg"]);dlg.transient(self);dlg.grab_set()
        ctk.CTkLabel(dlg,text="Dans combien de temps je dois te le rappeler ?",
                     font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(pady=(20,8))
        ctk.CTkLabel(dlg,text="L'alerte sera silencieuse pendant la dur\u00e9e choisie,\npuis r\u00e9appara\u00eetra au prochain refresh.",
                     font=("Segoe UI",10),text_color=C["t2"],justify="center").pack(pady=(0,16))
        btns=ctk.CTkFrame(dlg,fg_color="transparent");btns.pack(pady=8)
        def choisir(minutes):
            try:
                acks=load_json(ANTIRUPTURE_ACK_FILE) or {}
            except Exception: acks={}
            until=datetime.now()+timedelta(minutes=minutes)
            acks[pont_id]={"ack_at_iso":datetime.now().isoformat(),
                           "manques":manques_actuels,
                           "type_ack":"snooze",
                           "snooze_until_iso":until.isoformat()}
            try: save_json(ANTIRUPTURE_ACK_FILE,acks)
            except Exception as e: print(f"[snooze] {e}")
            # Silence global popup anti-rupture : capture l'état actuel de TOUTES les alertes
            # (pas seulement ce pont) pour éviter que la popup revienne pour une autre cause.
            try:
                ar_now=getattr(self.master,"last_data",{}).get("antirupture",{}) or {}
                fps,extra=_antirupture_fps_extra(ar_now)
                # Note : les saisies impossibles ont leur propre silence (popup différente).
                # Le snooze pont ne couvre QUE l'anti-rupture, pas les saisies.
                silence_popup("antirupture",fps,until.isoformat(),extra)
            except Exception as e: print(f"[silence antirupture] {e}")
            # Journal d'événements : trace la décision "snooze" pour traçabilité
            try:
                pont_info=self._pont_info_for_id(pont_id)
                if pont_info:
                    add_evenement("pont",{
                        "date_debut":pont_info["trou_start"].isoformat(),
                        "date_fin":(pont_info["trou_start"]+timedelta(days=pont_info.get("trou_duree",1)-1)).isoformat(),
                        "duree":pont_info.get("trou_duree",1),
                        "manques":[{"carburant":c,"manque":m} for c,m in (manques_actuels or {}).items()],
                        "ack_type":"snooze",
                        "snooze_minutes":minutes,
                        "snooze_until":until.isoformat(),
                    })
            except Exception as e: print(f"[evt pont snooze] {e}")
            dlg.destroy();self.destroy()
        # Calcul du prochain jour ouvré (skip weekends et fériés)
        def next_business_day(d):
            d=d+timedelta(days=1)
            while d.weekday()>=5 or is_ferie(d): d=d+timedelta(days=1)
            return d
        # === OPTIONS ADAPTATIVES SELON DISTANCE AU PONT ===
        # Pour un pont à J+11 (genre 22/05 vu le 11/05), proposer "30 min" ou "1h" est absurde.
        # On adapte les options et le bouton principal selon la distance temporelle :
        #   - Pont à J+8 et plus  → options en jours/semaine, rappel principal = pont - 2j matin
        #   - Pont à J+3 à J+7    → mix heures/jours
        #   - Pont à J+0 à J+2    → DEADLINE TEMAG aujourd'hui à respecter
        pont_info=self._pont_info_for_id(pont_id)
        jours_jusqu_au_pont=(pont_info["trou_start"]-date.today()).days if pont_info else 0
        JC=["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
        if jours_jusqu_au_pont>=8 and pont_info:
            # Pont/weekend lointain : caler les rappels sur le JOUR DE COMMANDE (cohérent
            # avec l'alerte qui dit "passe la commande tel jour avant 11h"), au lieu de
            # durées génériques (1j/3j/1sem) déconnectées de la deadline. Le rappel doit
            # tomber quand Bidou peut AGIR sur la commande : la veille au soir, ou le matin
            # même avant la deadline 11h. C'est la cohérence inter-modules : le rappel
            # utilise la même règle de commande que le reste du HUB.
            now_dt=datetime.now()
            # Jour de livraison cible = dernier jour livrable avant le trou
            jour_livr=pont_info["trou_start"]-timedelta(days=1)
            while jour_livr.weekday()>=5 or is_ferie(jour_livr): jour_livr-=timedelta(days=1)
            # Jour de commande = veille ouvrée, deadline 11h (via le core)
            jour_cmd=jour_de_commande(jour_livr) or (jour_livr-timedelta(days=1))
            # Veille ouvrée du jour de commande
            veille_cmd=jour_cmd-timedelta(days=1)
            while veille_cmd.weekday()>=5 or is_ferie(veille_cmd): veille_cmd-=timedelta(days=1)
            moments=[
                datetime.combine(veille_cmd,dt_time(18,0)),  # veille de la commande, le soir
                datetime.combine(jour_cmd,dt_time(8,0)),      # matin du jour de commande
                datetime.combine(jour_cmd,dt_time(10,30)),    # juste avant la deadline 11h
            ]
            # Labels laissés vides : convertis en dates absolues lisibles plus bas.
            snooze_options=[("",int((mt-now_dt).total_seconds()/60)) for mt in moments
                            if mt>now_dt+timedelta(minutes=30)]
            matin_cmd=datetime.combine(jour_cmd,dt_time(8,0))
            mins_until_next=max(1,int((matin_cmd-now_dt).total_seconds()/60))
            nxt_label=f"\u2192 {JC[jour_cmd.weekday()]} {jour_cmd.day}/{jour_cmd.month:02d} matin (jour de commande)"
        elif jours_jusqu_au_pont>=8:
            # Sécurité : pont lointain mais infos manquantes → durées génériques (fallback)
            snooze_options=[("1 jour",24*60),("3 jours",3*24*60),("1 semaine",7*24*60)]
            mins_until_next=24*60
            nxt_label="\u2192 demain"
        elif jours_jusqu_au_pont>=3:
            # Pont moyen terme : mix
            snooze_options=[("4 h",240),("1 jour",24*60),("3 jours",3*24*60)]
            nxt=next_business_day(date.today())
            until_dt=datetime.combine(nxt,datetime.min.time()).replace(hour=6)
            mins_until_next=int((until_dt-datetime.now()).total_seconds()/60)
            nxt_label=f"\u2192 {JC[nxt.weekday()]} matin"
        else:
            # Pont proche (0-2 jours) : LA COMMANDE EST CRITIQUE.
            # Si jours_jusqu_au_pont == 2 et aujourd'hui avant 11h → DEADLINE TEMAG CE MATIN
            # pour livraison demain qui couvrira le pont. Toutes les options doivent
            # respecter cette deadline et NE PAS proposer un rappel à demain matin.
            today_11h=datetime.combine(date.today(),dt_time(11, 0))
            today_1045=today_11h-timedelta(minutes=15)  # marge 15 min pour saisir
            now_dt=datetime.now()
            deadline_aujourdhui=(jours_jusqu_au_pont==2 and now_dt<today_11h)
            if deadline_aujourdhui:
                if now_dt<today_1045:
                    # Avant 10h45 : options courtes capées à 10h45
                    cand=[]
                    for lbl,m in [("\u23f1 Dans 30 min",30),("\u23f1 Dans 1h",60),("\u23f1 Dans 2h",120)]:
                        if now_dt+timedelta(minutes=m)<today_1045:
                            cand.append((lbl,m))
                    mins_to_1045=int((today_1045-now_dt).total_seconds()/60)
                    cand.append(("\u23f1 \u00c0 10h45 (juste avant deadline)",mins_to_1045))
                    snooze_options=cand
                    mins_until_next=mins_to_1045
                    nxt_label="\u26a0 Passe la commande avant 10h45 (deadline TEMAG 11h)"
                else:
                    # Entre 10h45 et 11h : urgence absolue
                    cand=[("\u23f1 Dans 5 min",5),("\u23f1 Dans 10 min",10)]
                    snooze_options=cand
                    mins_until_next=5
                    nxt_label="\u26a0 Deadline TEMAG dans <15 min \u2014 commande MAINTENANT"
            else:
                # Pas de deadline aujourd'hui (jours_jusqu_au_pont != 2, ou après 11h)
                snooze_options=[("30 min",30),("1 h",60),("2 h",120),("4 h",240)]
                nxt=next_business_day(date.today())
                until_dt=datetime.combine(nxt,datetime.min.time()).replace(hour=6)
                mins_until_next=int((until_dt-datetime.now()).total_seconds()/60)
                nxt_label=f"\u2192 {JC[nxt.weekday()]} matin"
        # Cap de sécurité : ne jamais snoozer au-delà de la veille du pont à 5h30
        if pont_info:
            cap_dt=datetime.combine(pont_info["trou_start"]-timedelta(days=1),datetime.min.time()).replace(hour=5,minute=30)
            cap_mins=int((cap_dt-datetime.now()).total_seconds()/60)
            if cap_mins>0:
                snooze_options=[(lbl,min(m,cap_mins)) for lbl,m in snooze_options]
                mins_until_next=min(mins_until_next,cap_mins)
        # Conversion : si label MÉTIER (commence par ⏱ ou ⚠), garder tel quel ; sinon timestamp absolu
        def _is_metier_label(lbl):
            return lbl.startswith("\u23f1") or lbl.startswith("\u26a0")
        snooze_options=[(lbl if _is_metier_label(lbl) else fmt_rappel_dt(datetime.now()+timedelta(minutes=m)),m) for lbl,m in snooze_options]
        # FALLBACK garanti : si la liste est vide pour une raison X, ajouter au moins "Dans 1h"
        if not snooze_options:
            snooze_options=[("\u23f1 Dans 1h",60)]
        # Boutons options : sur 1 ligne si ≤3, sur 2 lignes si plus
        if len(snooze_options)<=3:
            for label,mins in snooze_options:
                ctk.CTkButton(btns,text=label,width=180,height=36,fg_color=C["amber"],hover_color="#C4811D",
                               text_color="#000",font=("Segoe UI",10,"bold"),corner_radius=6,
                               command=lambda m=mins:choisir(m)).pack(side="left",padx=4)
        else:
            row1=ctk.CTkFrame(btns,fg_color="transparent");row1.pack(fill="x",pady=(0,4))
            row2=ctk.CTkFrame(btns,fg_color="transparent");row2.pack(fill="x")
            half=(len(snooze_options)+1)//2
            for label,mins in snooze_options[:half]:
                ctk.CTkButton(row1,text=label,width=200,height=36,fg_color=C["amber"],hover_color="#C4811D",
                               text_color="#000",font=("Segoe UI",10,"bold"),corner_radius=6,
                               command=lambda m=mins:choisir(m)).pack(side="left",padx=4)
            for label,mins in snooze_options[half:]:
                ctk.CTkButton(row2,text=label,width=200,height=36,fg_color=C["amber"],hover_color="#C4811D",
                               text_color="#000",font=("Segoe UI",10,"bold"),corner_radius=6,
                               command=lambda m=mins:choisir(m)).pack(side="left",padx=4)
        # Bouton principal (row 3)
        is_today_non_actionable=date.today().weekday()>=5 or is_ferie(date.today())
        nxt_color=C["green"] if is_today_non_actionable else "#5A7CC4"
        nxt_hover="#15943C" if is_today_non_actionable else "#3F60A5"
        # Si nxt_label est un warning métier (commence par ⚠), affichage rouge urgent
        if nxt_label.startswith("\u26a0"):
            nxt_color=C["red"];nxt_hover="#A82820"
        ctk.CTkButton(dlg,text=nxt_label,width=420,height=44,fg_color=nxt_color,hover_color=nxt_hover,
                       text_color="#FFF",font=("Segoe UI",12,"bold"),corner_radius=8,
                       command=lambda m=mins_until_next:choisir(m)).grid(row=3,column=0,pady=(14,4))
        # Aide WE/férié (row 4, conditionnel)
        if is_today_non_actionable:
            ctk.CTkLabel(dlg,text="\U0001f4a1 La SARA est ferm\u00e9e aujourd'hui, tu ne peux rien commander avant le prochain jour ouvr\u00e9.",
                         font=("Segoe UI",9,"italic"),text_color=C["gold"],wraplength=540,justify="center").grid(row=4,column=0,pady=(0,4))
        # Bouton Annuler (row 5)
        ctk.CTkButton(dlg,text="Annuler",width=100,height=28,fg_color=C["panel"],hover_color=C["card"],
                       text_color=C["t2"],font=("Segoe UI",10),corner_radius=6,
                       command=dlg.destroy).grid(row=5,column=0,pady=(8,12))
    def _accepter_rupture(self,pont_id,manques_actuels):
        """Ouvre une mini-fenêtre demandant la cause de la rupture acceptée,
        puis enregistre l'acquittement avec type_ack=rupture_acceptee + cause.
        Cette information sera archivée pour le rapport mensuel."""
        dlg=ctk.CTkToplevel(self);dlg.title("Accepter la rupture")
        dlg.geometry("520x280");dlg.configure(fg_color=C["bg"]);dlg.transient(self);dlg.grab_set()
        ctk.CTkLabel(dlg,text="\u26a0 Tu acceptes la rupture sur ce pont",
                     font=("Segoe UI",13,"bold"),text_color=C["red"]).pack(pady=(20,4))
        ctk.CTkLabel(dlg,text="Cette d\u00e9cision sera archiv\u00e9e dans le rapport mensuel.\nIndique bri\u00e8vement la cause :",
                     font=("Segoe UI",10),text_color=C["t2"],justify="center").pack(pady=(0,10))
        cause_var=ctk.StringVar()
        entry=ctk.CTkEntry(dlg,width=460,height=36,textvariable=cause_var,
                            placeholder_text="Ex : livraison partielle Total, deadline rat\u00e9e, autre cause...",
                            font=("Segoe UI",11))
        entry.pack(pady=(4,4),padx=20);entry.focus_set()
        # Suggestions rapides cliquables
        sug_row=ctk.CTkFrame(dlg,fg_color="transparent");sug_row.pack(pady=(2,12))
        for label in ["Livraison partielle Total","Deadline rat\u00e9e","Autre"]:
            ctk.CTkButton(sug_row,text=label,height=24,fg_color=C["panel"],hover_color=C["card"],
                           text_color=C["t2"],font=("Segoe UI",9),corner_radius=4,
                           command=lambda l=label:cause_var.set(l)).pack(side="left",padx=3)
        def valider():
            cause=cause_var.get().strip()
            if not cause:
                cause="Non pr\u00e9cis\u00e9e"
            try:
                acks=load_json(ANTIRUPTURE_ACK_FILE) or {}
            except Exception: acks={}
            acks[pont_id]={"ack_at_iso":datetime.now().isoformat(),
                           "manques":manques_actuels,
                           "type_ack":"rupture_acceptee",
                           "cause":cause}
            try: save_json(ANTIRUPTURE_ACK_FILE,acks)
            except Exception as e: print(f"[accepter rupture] {e}")
            # Silence global popup anti-rupture : jusqu'à demain matin 6h.
            # La rupture est acceptée donc on ne veut plus être harcelé pour ça.
            try:
                tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
                ar_now=getattr(self.master,"last_data",{}).get("antirupture",{}) or {}
                fps,extra=_antirupture_fps_extra(ar_now)
                silence_popup("antirupture",fps,tomorrow_6h.isoformat(),extra)
            except Exception as e: print(f"[silence antirupture rupture] {e}")
            # Journal d'événements : trace la décision "rupture acceptée" + cause
            try:
                pont_info=self._pont_info_for_id(pont_id)
                if pont_info:
                    add_evenement("pont",{
                        "date_debut":pont_info["trou_start"].isoformat(),
                        "date_fin":(pont_info["trou_start"]+timedelta(days=pont_info.get("trou_duree",1)-1)).isoformat(),
                        "duree":pont_info.get("trou_duree",1),
                        "manques":[{"carburant":c,"manque":m} for c,m in (manques_actuels or {}).items()],
                        "ack_type":"rupture_acceptee",
                        "cause":cause,
                    },commentaire=f"Cause : {cause}")
            except Exception as e: print(f"[evt pont rupture] {e}")
            dlg.destroy();self.destroy()
        btn_row=ctk.CTkFrame(dlg,fg_color="transparent");btn_row.pack(pady=10)
        ctk.CTkButton(btn_row,text="Annuler",width=100,height=34,fg_color=C["panel"],hover_color=C["card"],
                       text_color=C["t1"],font=("Segoe UI",11),corner_radius=6,
                       command=dlg.destroy).pack(side="left",padx=6)
        ctk.CTkButton(btn_row,text="Valider la rupture accept\u00e9e",width=200,height=34,
                       fg_color="#7A1F22",hover_color="#5C1518",text_color="#FFF",
                       font=("Segoe UI",11,"bold"),corner_radius=6,
                       command=valider).pack(side="left",padx=6)
    def _section_title(self,parent,text,color):
        f=ctk.CTkFrame(parent,fg_color="transparent");f.pack(fill="x",pady=(14,6))
        ctk.CTkLabel(f,text=text,font=("Segoe UI",13,"bold"),text_color=color,anchor="w").pack(anchor="w")
    def _forcer_livraison_exceptionnelle(self,d,vol=0):
        """Déclare une livraison sur jour non-livrable comme exception ASSUMÉE.
        Confirmation requise (désactive une alerte de sécurité), persiste le
        forçage, journalise, ferme la popup et déclenche un refresh."""
        if d is None: return
        try:
            import tkinter.messagebox as _mb
            label=d.strftime("%A %d/%m/%Y") if hasattr(d,"strftime") and not isinstance(d,str) else str(d)
            ok=_mb.askyesno(
                "Forcer une livraison exceptionnelle",
                f"Confirmer que la livraison du {label} est VOLONTAIRE et assum\u00e9e ?\n\n"
                f"Le hub ne la signalera plus comme incoh\u00e9rence (jour non-livrable).\n"
                f"\u00c0 utiliser uniquement quand tu sais que la livraison aura bien lieu "
                f"ce jour-l\u00e0 (ex : d\u00e9calage SARA Pentec\u00f4te \u2192 samedi).",
                parent=self)
            if not ok: return
            add_forcage(d,note="exception assum\u00e9e via popup anti-rupture",vol=vol)
            try:
                add_evenement("ack",{
                    "date":d.strftime("%Y-%m-%d") if hasattr(d,"strftime") and not isinstance(d,str) else str(d)[:10],
                    "sujet":"livraison_exceptionnelle_forcee",
                    "volume":int(vol or 0),
                    "statut":"resolu","lu":True,
                })
            except Exception as _e: _log_silent_err(exc=_e)
            try:
                app=self.master
                while app is not None and not hasattr(app,"refresh"):
                    app=getattr(app,"master",None)
                if app is not None: app.refresh()
            except Exception as _e: _log_silent_err(exc=_e)
            try: self.destroy()
            except Exception as _e: _log_silent_err(exc=_e)
        except Exception as _e:
            _log_silent_err(exc=_e)

    def _line_box(self,parent,title,detail):
        box=ctk.CTkFrame(parent,fg_color=C["alert_bg"],corner_radius=8,border_width=1,border_color=C["alert_border"])
        box.pack(fill="x",padx=4,pady=(2,8))
        ctk.CTkLabel(box,text=title,font=("Segoe UI",12,"bold"),text_color=C["t1"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(10,4))
        ctk.CTkLabel(box,text=detail,font=("Segoe UI",11),text_color=C["t2"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(0,10))

# =============================================================================
# POPUP D'ALERTE TENDANCE FORTE — déclenchée si anomalie ≥ 30% sur ventes en cours
# d'un carburant tendu sur un pont à venir. Acquittable par carburant+jour+stage.
# Réapparaît si l'écart s'aggrave de plus de +20 points (ex : C1 +44% → C2 +65%).
class TendanceAlerteDlg(ctk.CTkToplevel):
    def __init__(self,parent,tendance_alertes):
        super().__init__(parent)
        self.title("\U0001f514 Tendance forte d\u00e9tect\u00e9e \u2014 DISTRICARB HUB")
        self.geometry("760x560");self.minsize(680,460)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.alertes=tendance_alertes
        # Header ambre
        hdr=ctk.CTkFrame(self,fg_color="#3D2A1A",corner_radius=0,height=80,border_width=0);hdr.pack(fill="x");hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f514",font=("Segoe UI Emoji",30),text_color=C["amber"]).pack(side="left",padx=(24,12),pady=18)
        title_box=ctk.CTkFrame(hdr,fg_color="transparent");title_box.pack(side="left",fill="y",pady=14)
        ctk.CTkLabel(title_box,text="TENDANCE FORTE D\u00c9TECT\u00c9E",font=("Segoe UI",16,"bold"),text_color=C["amber"],anchor="w").pack(anchor="w")
        ctk.CTkLabel(title_box,text="Tes ventes du jour s'\u00e9cartent fortement de la moyenne sur un carburant tendu",
                     font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(anchor="w")
        # Corps scrollable
        body=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        body.pack(fill="both",expand=True,padx=20,pady=(12,8))
        # Une carte par alerte (= un carburant)
        for al in self.alertes:
            container=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=10,border_width=2,border_color=C["amber"])
            container.pack(fill="x",padx=4,pady=(8,8))
            signe="+" if al["ecart_pct"]>0 else ""
            ctk.CTkLabel(container,text=f"\u26a1 {al['carburant']} : {signe}{al['ecart_pct']}% \u00e0 {al['stage']}",
                         font=("Segoe UI",13,"bold"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(12,4))
            ctk.CTkLabel(container,text=f"   {al['ventes_partielles']:,}L vendus vs {al['ventes_attendues']:,}L attendus".replace(",","."),
                         font=("Segoe UI",11),text_color=C["t1"],anchor="w").pack(anchor="w",padx=14,pady=2)
            # Conséquences sur les ponts/week-ends
            ctk.CTkLabel(container,text="Cons\u00e9quence sur les week-ends / ponts \u00e0 venir :",
                         font=("Segoe UI",11,"italic"),text_color=C["t2"],anchor="w").pack(anchor="w",padx=14,pady=(8,2))
            for p in al["ponts"]:
                _terme=p.get("terme","Pont")
                _terme_min=_terme.lower()
                if p["deja_manque"]:
                    txt=f"   \u2022 {_terme} {p['trou_str']} : d\u00e9j\u00e0 en marge serr\u00e9e ({al['carburant']} manque ~{p['manque_arrondi']:,}L). Si la tendance se confirme \u2192 risque accru.".replace(",",".")
                else:
                    txt=f"   \u2022 {_terme} {p['trou_str']} : stock fin {_terme_min} pr\u00e9vu {p['stock_fin_pont']:,}L = {p['stock_fin_pont']/max(1,p['ventes_lendemain']):.1f} jour. Si la tendance se confirme \u2192 risque de manque.".replace(",",".")
                ctk.CTkLabel(container,text=txt,font=("Segoe UI",10),text_color=C["t1"],anchor="w",justify="left",wraplength=680).pack(anchor="w",padx=14,pady=2)
            ctk.CTkLabel(container,text="",height=8).pack()
        # Footer : bouton acquittement
        footer=ctk.CTkFrame(self,fg_color="transparent",height=70);footer.pack(side="bottom",fill="x",padx=20,pady=14);footer.pack_propagate(False)
        ctk.CTkLabel(footer,text="L'alerte reviendra si la tendance s'aggrave (>20 points) au prochain refresh.",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(side="left",padx=(4,0),pady=18)
        ctk.CTkButton(footer,text="C'est not\u00e9",width=180,height=40,fg_color=C["amber"],hover_color="#C4811D",
                       text_color="#000",font=("Segoe UI",12,"bold"),corner_radius=8,
                       command=self._acquitter).pack(side="right")
    def _acquitter(self):
        """Enregistre l'acquittement des alertes tendance avec snapshot des écarts actuels.
        Format dans le fichier (par jour+stage+carb) :
          {"2026-04-29_1/3_GO": {"ack_at_iso":"...","ecart_pct":44}}
        Réapparition si l'écart s'aggrave de plus de 20 points (passe de +44% à +65%)."""
        try:
            acks=load_json(TENDANCE_ACK_FILE) or {}
        except Exception: acks={}
        now=datetime.now().isoformat()
        d_str=date.today().strftime("%Y-%m-%d")
        for al in self.alertes:
            key=f"{d_str}_{al['stage'].replace(' ','_').replace('/','-')}_{al['carburant']}"
            acks[key]={"ack_at_iso":now,"ecart_pct":al["ecart_pct"]}
        try: save_json(TENDANCE_ACK_FILE,acks)
        except Exception as e: print(f"[acquittement tendance] {e}")
        # Silence global popup tendance : jusqu'à demain matin 6h.
        # Réveil prématuré uniquement si nouveau carburant ou aggravation >20 points.
        try:
            tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
            fps=[f"{d_str}_{al.get('carburant','')}" for al in self.alertes]
            extra={f"{d_str}_{al.get('carburant','')}":al.get("ecart_pct",0) for al in self.alertes}
            silence_popup("tendance",fps,tomorrow_6h.isoformat(),extra)
        except Exception as e: print(f"[silence tendance] {e}")
        self.destroy()

# =============================================================================
# POPUP SAISIES PHYSIQUEMENT IMPOSSIBLES — déclenchée si ventes prévues D7/D8/D9 dépassent
# le stock physique disponible (matin + livraison) - plancher (500/500/250).
# Acquittable globalement par "OK je gère" (snapshot des saisies actuelles).
# Réapparaît si une nouvelle saisie irréaliste apparaît OU si une existante s'aggrave.
class SaisiesIrrealistesDlg(ctk.CTkToplevel):
    def __init__(self,parent,saisies_irr):
        super().__init__(parent)
        self.title("\U0001f6ab Saisies irr\u00e9alistes \u2014 DISTRICARB HUB")
        self.geometry("780x580");self.minsize(700,460)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.saisies=saisies_irr
        # Header rouge (problème primaire à corriger)
        hdr=ctk.CTkFrame(self,fg_color="#3D1A1A",corner_radius=0,height=80,border_width=0);hdr.pack(fill="x");hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f6ab",font=("Segoe UI Emoji",30),text_color=C["red"]).pack(side="left",padx=(24,12),pady=18)
        title_box=ctk.CTkFrame(hdr,fg_color="transparent");title_box.pack(side="left",fill="y",pady=14)
        ctk.CTkLabel(title_box,text="SAISIES PHYSIQUEMENT IMPOSSIBLES",font=("Segoe UI",16,"bold"),text_color=C["red"],anchor="w").pack(anchor="w")
        ctk.CTkLabel(title_box,text="Tes pr\u00e9visions de ventes d\u00e9passent le stock disponible en cuve",
                     font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(anchor="w")
        # Corps scrollable
        body=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        body.pack(fill="both",expand=True,padx=20,pady=(12,8))
        # Encart d'explication
        info_box=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=8,border_width=1,border_color=C["border2"])
        info_box.pack(fill="x",padx=4,pady=(4,12))
        ctk.CTkLabel(info_box,text="\u2139\ufe0f Pourquoi cette alerte",font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="w").pack(anchor="w",padx=12,pady=(8,2))
        ctk.CTkLabel(info_box,
                     text="En dessous du plancher physique (500L SP/GO, 250L GNR), les pompes ne distribuent plus.\n"
                          "Donc tes ventes max d'un jour = (stock matin + livraison du jour) - plancher.\n"
                          "Si une de tes saisies D7/D8/D9 d\u00e9passe ce maximum, c'est physiquement impossible.\n"
                          "Tu dois ajuster tes pr\u00e9visions dans Prévision compte.xlsx.",
                     font=("Segoe UI",10),text_color=C["t2"],anchor="w",justify="left").pack(anchor="w",padx=12,pady=(0,10))
        # Grouper les saisies irréalistes par jour pour un affichage plus clair
        par_jour={}
        for s in self.saisies:
            par_jour.setdefault(s["date_str"],[]).append(s)
        for date_str,items in par_jour.items():
            container=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=10,border_width=2,border_color=C["red"])
            container.pack(fill="x",padx=4,pady=(8,8))
            ctk.CTkLabel(container,text=f"\U0001f6a8 {date_str}",
                         font=("Segoe UI",13,"bold"),text_color=C["red"],anchor="w").pack(anchor="w",padx=14,pady=(12,4))
            for s in items:
                line_txt=(f"   \u2022 {s['carburant']} : tu pr\u00e9vois {s['vente_saisie']:,}L de ventes\n"
                          f"        mais tu n'as que {s['dispo']:,}L disponibles ({s['stock_matin']:,} matin + {s['livraison']:,} livraison).\n"
                          f"        Plancher physique {s['plancher']}L. Vente max possible : {s['vente_max_possible']:,}L.").replace(",",".")
                ctk.CTkLabel(container,text=line_txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=2)
                # Suggestion d'ajustement
                action_txt=f"        \u2192 Ajuste D{['sp','go','gnr'].index(s['carburant'].lower())+7} de cet onglet \u00e0 {s['vente_max_possible']:,}L max (au lieu de {s['vente_saisie']:,}L).".replace(",",".")
                ctk.CTkLabel(container,text=action_txt,font=("Segoe UI",10,"italic"),text_color=C["amber"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(2,4))
            ctk.CTkLabel(container,text="",height=6).pack()
        # Footer : bouton acquittement
        footer=ctk.CTkFrame(self,fg_color="transparent",height=70);footer.pack(side="bottom",fill="x",padx=20,pady=14);footer.pack_propagate(False)
        ctk.CTkLabel(footer,text="L'alerte reviendra si une nouvelle saisie irr\u00e9aliste appara\u00eet ou si une existante s'aggrave.",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(side="left",padx=(4,0),pady=18)
        ctk.CTkButton(footer,text="\u2713 OK je g\u00e8re",width=180,height=40,fg_color=C["green"],hover_color="#15943C",
                       text_color="#FFF",font=("Segoe UI",12,"bold"),corner_radius=8,
                       command=self._acquitter).pack(side="right")
    def _acquitter(self):
        """Enregistre l'acquittement avec snapshot des excès actuels.
        L'alerte reviendra uniquement si :
         - une nouvelle saisie irréaliste apparaît (date+carb non présents dans le snapshot)
         - OU une existante voit son excès augmenter de plus de 1000L"""
        try:
            acks=load_json(SAISIES_IRR_ACK_FILE) or {}
        except Exception: acks={}
        snapshot={}
        for s in self.saisies:
            key=f"{s['date'].strftime('%Y-%m-%d')}_{s['carburant']}"
            snapshot[key]=s["exces"]
        acks["last_ack"]={"ack_at_iso":datetime.now().isoformat(),"snapshot":snapshot}
        try: save_json(SAISIES_IRR_ACK_FILE,acks)
        except Exception as e: print(f"[ack saisies irr] {e}")
        # Silence global popup saisies impossibles : jusqu'à demain matin 6h.
        # Réveil prématuré uniquement si nouvelle saisie ou aggravation >1000L.
        try:
            tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
            fps=[f"{s['date'].strftime('%Y-%m-%d')}_{s['carburant']}" for s in self.saisies]
            extra={f"{s['date'].strftime('%Y-%m-%d')}_{s['carburant']}":s.get("exces",0) for s in self.saisies}
            silence_popup("saisies_irr",fps,tomorrow_6h.isoformat(),extra)
        except Exception as e: print(f"[silence saisies irr] {e}")
        self.destroy()


class MargeTendueDlg(ctk.CTkToplevel):
    """Popup d'alerte pour livraisons à MARGE TENDUE (cuve presque pleine).
    
    Cas couvert : la livraison rentre dans la cuve (pas de dépassement strict) MAIS la marge
    restante après livraison est < 4 000 L. Risque : si les ventes nuit/C1 sont plus calmes
    que prévu, la cuve sera tendue à la livraison du matin.
    
    Alerte plus prudente que LivraisonsAReporterDlg (qui ne se déclenche qu'en dépassement
    strict). Aligne le hub sur le seuil "Attention" du fichier Excel maison de Bidou.
    
    Actions :
      - "C'est noté" : silence jusqu'à demain matin 6h
      - "Plus tard (2h)" : snooze 2h pour revenir avant 11h
    Croix X = non traité (pastille rouge journal).
    """
    def __init__(self,parent,livraisons_tendues):
        super().__init__(parent)
        self.title("\u26a0 Livraisons marge tendue \u2014 DISTRICARB HUB")
        # Tout est packé DIRECTEMENT dans self, comme LivraisonDialog qui marche
        # parfaitement chez Bidou. Pas de header/body/footer containers séparés,
        # pas de CTkScrollableFrame (responsable du bug d'affichage chez Bidou —
        # confirmé par le diagnostic du 20/05 : Tests 1-4 isolés OK, mais dès qu'on
        # ajoute du contenu réel dans un ScrollableFrame, le footer disparaît).
        # Pattern minimaliste : un widget par "ligne", packé séquentiellement.
        # En contrepartie : on n'a plus de scroll. Mais en pratique il y a presque
        # toujours UNE seule card de livraison concernée (cas Bidou 20/05 : sam.23/05
        # SP). Si rare cas multi-cards, la fenêtre s'agrandira en hauteur via la
        # géométrie auto (pas resize=False explicite).
        self.geometry("680x520")
        self.configure(fg_color=C["bg"])
        self.resizable(False,False)
        self.transient(parent);self.grab_set()
        self.livraisons_tendues=livraisons_tendues
        self._user_action_taken=False
        self.protocol("WM_DELETE_WINDOW",self._on_close)
        # ===== HEADER simple (titre + sous-titre, comme LivraisonDialog) =====
        ctk.CTkLabel(self,text="\u26a0  LIVRAISONS \u00c0 MARGE TENDUE",
                     font=("Segoe UI",17,"bold"),text_color=C["amber"]).pack(pady=(18,2))
        ctk.CTkLabel(self,text="La cuve sera presque pleine \u00e0 la livraison \u2014 surveillance recommand\u00e9e",
                     font=("Segoe UI",11),text_color=C["t2"]).pack(pady=(0,12))
        # ===== ENCART explicatif (Pourquoi cette alerte) =====
        info_box=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8)
        info_box.pack(fill="x",padx=24,pady=(0,10))
        ctk.CTkLabel(info_box,text="\u2139\ufe0f Pourquoi cette alerte",
                     font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="w").pack(anchor="w",padx=14,pady=(8,2))
        ctk.CTkLabel(info_box,
                     text=("Stock matin + livraison rentre dans la cuve, MAIS la marge restante est < 4 000 L.\n"
                           "Si les ventes nuit/C1 sont plus calmes que pr\u00e9vu, la cuve sera tr\u00e8s tendue \u00e0 la livraison."),
                     font=("Segoe UI",10),text_color=C["t2"],anchor="w",justify="left",wraplength=600).pack(anchor="w",padx=14,pady=(0,4))
        ctk.CTkLabel(info_box,
                     text=("\u00ab Pris en compte \u00bb : silence jusqu'\u00e0 demain matin 6h.  \u00b7  "
                           "\u00ab Me redemander... \u00bb : choisis un moment m\u00e9tier.  \u00b7  "
                           "Fermer (X) : reste en alerte, pastille rouge journal."),
                     font=("Segoe UI",9,"italic"),text_color=C["t3"],anchor="w",justify="left",wraplength=600).pack(anchor="w",padx=14,pady=(0,8))
        # ===== Cartes de détails (1 par jour de livraison concerné) =====
        par_jour={}
        for l in livraisons_tendues:
            par_jour.setdefault(l["date_str"],[]).append(l)
        for date_str,items in par_jour.items():
            card=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8)
            card.pack(fill="x",padx=24,pady=(0,8))
            ctk.CTkLabel(card,text=f"\U0001f69b {date_str}",
                         font=("Segoe UI",13,"bold"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(10,2))
            for l in items:
                fmt=lambda v:f"{int(v):,}".replace(",","\u202f")
                txt=(f"   \u2022 {l['carburant']} : livraison de {fmt(l['livraison'])} L sur stock matin {fmt(l['stock_matin'])} L\n"
                     f"        \u2192 marge cuve restante : {fmt(l['marge_restante'])} L (capacit\u00e9 {fmt(l['capacite'])} L)")
                ctk.CTkLabel(card,text=txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(2,8))
        # ===== Footer boutons (adaptés au MOMENT MÉTIER courant) =====
        # Validé Bidou 20/05/2026 : on n'affiche que les boutons qui ONT DU SENS au moment
        # où la popup s'ouvre. Plus de "Plus tard 2h" ou "Demain matin 6h" arbitraires.
        # On utilise detecter_moment_courant(date_livraison) qui retourne un code parmi :
        #   anticipation / jour_commande_avant_10h30 / jour_commande_deadline /
        #   jour_commande_apres_11h / entre_commande_et_livraison / matin_livraison.
        # Pour chaque mode, on affiche les boutons utiles métier + un texte d'aide
        # qui explique CE QUE FAIT le bouton (compréhensible par Véronique/Nadine).
        # Date de livraison la plus proche (gouverne le mode)
        dates_livr=[]
        for l in self.livraisons_tendues:
            d=l.get("date")
            if hasattr(d,"isoformat"):
                dates_livr.append(d if isinstance(d,date) else d.date())
        cible_livr=min(dates_livr) if dates_livr else (date.today()+timedelta(days=2))
        mode=detecter_moment_courant(cible_livr)
        # Frame footer
        btns=ctk.CTkFrame(self,fg_color="transparent")
        btns.pack(fill="x",padx=24,pady=(8,4))
        # Texte d'aide (sous les boutons) — change selon le mode
        aide_par_mode={
            "anticipation":         "L'alerte est anticip\u00e9e. Rien \u00e0 faire concr\u00e8tement maintenant : le hub te re-rappellera le jour de la commande pour que tu pr\u00e9viennes le transporteur.",
            "jour_commande_avant_10h30":"\u26a0 C'est AUJOURD'HUI que tu dois pr\u00e9venir le transporteur. Deadline TEMAG : 11h.",
            "jour_commande_deadline":"\u26a0\u26a0 DEADLINE TEMAG DANS MOINS DE 30 MIN. Appelle le transporteur MAINTENANT.",
            "jour_commande_apres_11h":"\u26a0 Deadline TEMAG pass\u00e9e. Si tu n'as pas pr\u00e9venu le transporteur, appelle d'urgence.",
            "entre_commande_et_livraison":"Commande pass\u00e9e. V\u00e9rification finale le matin de la livraison \u00e0 6h.",
            "matin_livraison":      "Matin de la livraison. V\u00e9rifie que les chiffres sont toujours coh\u00e9rents avant l'arriv\u00e9e du camion.",
            "livraison_passee":     "Livraison pass\u00e9e. Cette alerte n'a plus lieu d'\u00eatre, tu peux la marquer r\u00e9solue.",
        }
        # Boutons selon le mode (du plus contextuel au plus universel)
        if mode=="anticipation":
            # 1 seul bouton : OK noté
            tk.Button(btns,text="\u2713 OK not\u00e9 \u2014 rappelle-moi le jour de la commande",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_note).pack(side="left")
        elif mode=="jour_commande_avant_10h30":
            tk.Button(btns,text="\u2713 Transporteur pr\u00e9venu",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_note).pack(side="left",padx=(0,10))
            tk.Button(btns,text="\u23f1 Pas encore \u2014 rappelle-moi \u00e0 10h30",
                      bg="#E6A53D",fg="#000000",font=("Segoe UI",10,"bold"),
                      relief="raised",bd=2,padx=16,pady=8,cursor="hand2",
                      command=self._snooze_avant_deadline).pack(side="left")
        elif mode=="jour_commande_deadline":
            # URGENT : un seul bouton vert + un rouge
            tk.Button(btns,text="\u2713 Transporteur pr\u00e9venu",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_note).pack(side="left",padx=(0,10))
            tk.Button(btns,text="\u26a0 Pas encore — URGENT",
                      bg="#C8362E",fg="#FFFFFF",font=("Segoe UI",10,"bold"),
                      relief="raised",bd=2,padx=14,pady=8,cursor="hand2",
                      command=self._snooze_15min).pack(side="left")
        elif mode=="jour_commande_apres_11h":
            tk.Button(btns,text="\u2713 Transporteur pr\u00e9venu",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_note).pack(side="left")
        elif mode=="entre_commande_et_livraison":
            tk.Button(btns,text="\u2713 OK not\u00e9 \u2014 rappelle-moi le matin de la livraison",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_note).pack(side="left")
        elif mode=="matin_livraison":
            tk.Button(btns,text="\u2713 V\u00e9rifi\u00e9, chiffres coh\u00e9rents",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_note).pack(side="left")
        else:  # livraison_passee ou fallback
            tk.Button(btns,text="\u2713 Marquer comme r\u00e9gl\u00e9",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_note).pack(side="left")
        # Texte d'aide en bas (toujours présent, change selon le mode)
        aide_txt=aide_par_mode.get(mode,"")
        if aide_txt:
            ctk.CTkLabel(self,text=aide_txt,font=("Segoe UI",10,"italic"),
                         text_color=C["t2"],wraplength=620,justify="left").pack(padx=24,pady=(4,16),anchor="w")

    def _fingerprints(self):
        return [f"{l['date'].isoformat() if hasattr(l.get('date'),'isoformat') else l['date_str']}_{l['carburant']}" for l in self.livraisons_tendues]

    def _cest_note(self):
        """'C'est noté' = je gère, mais NE PAS éteindre à jamais.

        Bug corrigé (signalé terrain Bidou 18/05) : avant, ceci silençait jusqu'à
        demain 6h ET marquait l'événement 'resolu' → l'alerte ne revenait JAMAIS.
        Conséquence : une marge tendue sur livraison samedi 23/05, notée lundi 18,
        disparaissait pour toujours. Le jeudi/vendredi (férié, 0 vente), Bidou
        n'était plus rappelé → risque réel de débordement le jour J si le livreur
        n'a pas été prévenu.

        Nouveau comportement : on met en veille JUSQU'À l'avant-veille OUVRÉE de
        la date de livraison la plus proche (dernier moment utile pour prévenir
        le livreur / agir). L'alerte REVIENT d'elle-même à ce moment-là.
        On NE marque PAS 'resolu' (sinon elle ne reviendrait pas) : on trace
        juste l'acquittement avec statut 'ack' pour l'historique.
        """
        self._user_action_taken=True
        try:
            # Date de livraison la plus proche parmi les lignes concernées
            dates_livr=[]
            for l in self.livraisons_tendues:
                d=l.get("date")
                if hasattr(d,"isoformat"): dates_livr.append(d if isinstance(d,date) else d.date())
            cible=min(dates_livr) if dates_livr else (date.today()+timedelta(days=2))
            # Silence intelligent : jusqu'au PROCHAIN moment-clé métier STRICTEMENT futur.
            # Fix 21/05/2026 : on utilisait prochain_moment_cle qui a une marge -1h utile
            # pour DÉTECTER le mode courant mais NÉFASTE pour le silence (à 10h54, retournait
            # 10h30 → silence déjà expiré → popup ré-ouverte au refresh suivant). Maintenant
            # on utilise la variante _strict (>now sans marge) pour le silence.
            mc=prochain_moment_cle_strict(cible)
            if mc is None:
                # Plus aucun moment-clé futur : silence long (24h) — la livraison est passée
                # ou très proche, on évite tout harcèlement.
                until_dt=datetime.now()+timedelta(hours=24)
            else:
                until_dt=mc[0]
            silence_popup("marge_tendue",self._fingerprints(),until_dt.isoformat(),{})
            # Trace : acquittement (PAS resolu → l'alerte pourra revenir)
            for l in self.livraisons_tendues:
                add_evenement("marge_tendue",{
                    "date":l["date"].isoformat() if hasattr(l.get("date"),"isoformat") else str(l.get("date","")),
                    "carburant":l["carburant"],
                    "marge_restante":l["marge_restante"],
                    "statut":"ack","lu":True,
                    "rappel_prevu":until_dt.isoformat(),
                })
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _snooze_avant_deadline(self):
        """Rappel à 10h30 du jour de la commande (juste avant deadline TEMAG 11h)."""
        self._user_action_taken=True
        try:
            dates_livr=[]
            for l in self.livraisons_tendues:
                d=l.get("date")
                if hasattr(d,"isoformat"): dates_livr.append(d if isinstance(d,date) else d.date())
            cible=min(dates_livr) if dates_livr else (date.today()+timedelta(days=2))
            jc=jour_de_commande(cible) or (cible-timedelta(days=1))
            until_dt=datetime.combine(jc,dt_time(10,30))
            # Si 10h30 est déjà passé aujourd'hui, snooze court 15 min
            if until_dt<=datetime.now():
                until_dt=datetime.now()+timedelta(minutes=15)
            silence_popup("marge_tendue",self._fingerprints(),until_dt.isoformat(),{})
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _snooze_15min(self):
        """Snooze court 15 min — utilisé dans le mode URGENT proche deadline TEMAG."""
        self._user_action_taken=True
        try:
            until_iso=(datetime.now()+timedelta(minutes=15)).isoformat()
            silence_popup("marge_tendue",self._fingerprints(),until_iso,{})
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _snooze_metier(self,choice):
        """Snooze MÉTIER : silence jusqu'à un moment qui a du sens (matin de la
        livraison, demain matin avant 1ère caisse...) plutôt qu'un délai arbitraire."""
        self._user_action_taken=True
        target=None
        for lbl,dt in getattr(self,"_snooze_options",[]):
            if lbl==choice: target=dt; break
        if target is None: target=datetime.now()+timedelta(hours=2)
        try:
            silence_popup("marge_tendue",self._fingerprints(),target.isoformat(),{})
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _on_close(self):
        if not self._user_action_taken:
            try:
                for l in self.livraisons_tendues:
                    add_evenement("marge_tendue",{
                        "date":l["date"].isoformat() if hasattr(l.get("date"),"isoformat") else str(l.get("date","")),
                        "carburant":l["carburant"],
                        "marge_restante":l["marge_restante"],
                        "statut":"non_traite","lu":False,
                    })
            except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()


class FerieIsoleDlg(ctk.CTkToplevel):
    """Popup d'alerte pour férié ISOLÉ imminent (1 jour férié entre 2 jours ouvrés).
    
    Cas couvert : demain est férié (ex: Ascension), aujourd'hui = deadline commande pour la
    livraison du surlendemain (jour ouvré post-férié).
    
    Différent d'un "trou" classique (≥2j non livrables consécutifs) géré par AntiRuptureDlg.
    Ici on n'a pas forcément de manque calculé ; c'est une alerte INFORMATIVE qui rappelle
    de vérifier les commandes avant la deadline 11h.
    
    Actions utilisateur :
      - "C'est noté" : ack + silence jusqu'au lendemain du férié
      - "Plus tard (2h)" : snooze 2h pour revenir avant 11h
    Croix X (fermeture sans action) → événement non_traité (pastille rouge journal).
    """
    def __init__(self,parent,feries_isoles):
        super().__init__(parent)
        self.title("\U0001f5d3 F\u00e9ri\u00e9 imminent \u2014 DISTRICARB HUB")
        self.geometry("720x520");self.minsize(640,420)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.feries_isoles=feries_isoles
        self._user_action_taken=False
        self.protocol("WM_DELETE_WINDOW",self._on_close)
        # Header amber (informatif/anticipatoire, pas critique)
        hdr=ctk.CTkFrame(self,fg_color="#3D2F0F",corner_radius=0,height=80,border_width=0)
        hdr.pack(fill="x");hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f5d3",font=("Segoe UI Emoji",32),text_color=C["amber"]).pack(side="left",padx=(24,12),pady=18)
        title_box=ctk.CTkFrame(hdr,fg_color="transparent");title_box.pack(side="left",fill="y",pady=14)
        ctk.CTkLabel(title_box,text="F\u00c9RI\u00c9 IMMINENT",font=("Segoe UI",16,"bold"),text_color=C["amber"],anchor="w").pack(anchor="w")
        ctk.CTkLabel(title_box,text="V\u00e9rifie tes commandes avant la deadline SARA de 11h",
                     font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(anchor="w")
        # Corps
        body=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,
                                     scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        body.pack(fill="both",expand=True,padx=20,pady=(12,0))
        # Encart explicatif
        info_box=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=8,border_width=1,border_color=C["border2"])
        info_box.pack(fill="x",padx=4,pady=(4,12))
        ctk.CTkLabel(info_box,text="\u2139\ufe0f Pourquoi cette alerte",
                     font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="w").pack(anchor="w",padx=12,pady=(8,2))
        ctk.CTkLabel(info_box,
                     text=("Demain est un jour f\u00e9ri\u00e9 isol\u00e9 : la SARA ne livre pas.\n"
                           "Toute commande pour le jour ouvr\u00e9 suivant doit \u00eatre pass\u00e9e AUJOURD'HUI\n"
                           "avant 11h (deadline SARA).\n\n"
                           "C'est un rappel : v\u00e9rifie que tes commandes saisies dans Pre_vision sont OK."),
                     font=("Segoe UI",10),text_color=C["t2"],anchor="w",justify="left").pack(anchor="w",padx=12,pady=(0,10))
        # Détails pour chaque férié isolé
        for f in feries_isoles:
            card=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=10,border_width=2,border_color=C["amber"])
            card.pack(fill="x",padx=4,pady=(8,8))
            ctk.CTkLabel(card,text=f"\U0001f5d3 {f['date_ferie_str']} \u2014 {f['nom_ferie']}",
                         font=("Segoe UI",13,"bold"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(12,4))
            txt=(f"   \u2192 Prochaine livraison SARA possible : {f['date_post_ferie_str']}\n"
                 f"   \u2192 Deadline commande : {f['deadline_str']}")
            ctk.CTkLabel(card,text=txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(2,12))
        # Footer : boutons
        footer=ctk.CTkFrame(self,fg_color="transparent",height=64);footer.pack(side="bottom",fill="x",padx=20,pady=14);footer.pack_propagate(False)
        ctk.CTkLabel(footer,text="Croix X = laiss\u00e9 non trait\u00e9 (pastille rouge dans le journal)",
                     font=("Segoe UI",9),text_color=C["t3"]).pack(side="left",padx=(4,0),pady=20)
        ctk.CTkButton(footer,text="\u23f1 Plus tard (2h)",width=140,height=40,
                      fg_color=C["panel"],hover_color=C["card_h"],text_color=C["amber"],
                      border_width=1,border_color=C["amber"],
                      font=("Segoe UI",11),corner_radius=8,
                      command=self._snooze_2h).pack(side="right",padx=(0,8))
        ctk.CTkButton(footer,text="\u2713 C'est not\u00e9",width=140,height=40,
                      fg_color=C["green"],hover_color="#1F7C36",text_color="#FFF",
                      font=("Segoe UI",11,"bold"),corner_radius=8,
                      command=self._cest_note).pack(side="right",padx=(0,8))

    def _cest_note(self):
        """Bouton C'est noté : silence jusqu'au lendemain du férié + journal résolu."""
        self._user_action_taken=True
        try:
            # Silence jusqu'au lendemain du dernier férié (passe la deadline)
            last_ferie=max(f["date_ferie"] for f in self.feries_isoles)
            until_dt=datetime.combine(last_ferie+timedelta(days=1),datetime.min.time()).replace(hour=6)
            fps=[f["date_ferie"].isoformat() for f in self.feries_isoles]
            silence_popup("ferie_isole",fps,until_dt.isoformat(),{})
            for f in self.feries_isoles:
                add_evenement("ferie_isole",{
                    "date_ferie":f["date_ferie"].isoformat(),
                    "nom_ferie":f["nom_ferie"],
                    "statut":"resolu",
                    "lu":True,
                })
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _snooze_2h(self):
        """Snooze 2h pour revenir à l'alerte avant la deadline 11h."""
        self._user_action_taken=True
        try:
            until_iso=(datetime.now()+timedelta(hours=2)).isoformat()
            fps=[f["date_ferie"].isoformat() for f in self.feries_isoles]
            silence_popup("ferie_isole",fps,until_iso,{})
        except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _on_close(self):
        """Croix X = non traité (cohérent avec les autres popups d'alerte)."""
        if not self._user_action_taken:
            try:
                for f in self.feries_isoles:
                    add_evenement("ferie_isole",{
                        "date_ferie":f["date_ferie"].isoformat(),
                        "nom_ferie":f["nom_ferie"],
                        "statut":"non_traite",
                        "lu":False,
                    })
            except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()


class ObservatoireDlg(ctk.CTkToplevel):
    """Popup informative — Observatoire des Prix Martinique.
    
    Contexte métier : entre le 25 du mois et le 5 du suivant, l'Observatoire annonce
    les nouveaux prix applicables au 1er. Les clients anticipent (font le plein avant
    hausse, attendent si baisse), créant des ventes atypiques.
    
    Déclenchement : popup s'ouvre une fois par période Observatoire SI le moteur n'a
    détecté AUCUNE vente atypique dans les saisies Pre_vision (= utilisateur n'a pas
    encore anticipé). Si saisies déjà ajustées (anticipation détectée), popup ne
    s'ouvre pas car le travail métier est déjà fait.
    
    Cible : Bidou + Véronique + Nadine. Notamment Véronique/Nadine qui n'ont pas
    forcément intégré ce mécanisme métier.
    
    Actions :
      - "J'ai noté" : acquittement définitif pour ce mois Observatoire
      - "Demain 6h" : snooze 1 jour
      - "Dans 3 jours" : snooze pour revoir Pre_vision après quelques jours
      - "Avant le 1er" : snooze ciblé sur le 30 du mois (vérif finale avant changement)
    Croix X = pas d'acquittement, la popup pourra se ré-ouvrir au prochain refresh.
    """
    def __init__(self,parent,periode_key):
        super().__init__(parent)
        self.title("\U0001f4a1 P\u00e9riode Observatoire \u2014 DISTRICARB HUB")
        self.geometry("680x540")
        self.configure(fg_color=C["bg"])
        self.resizable(False,False)
        self.transient(parent);self.grab_set()
        self.periode_key=periode_key  # ex: "2026-05"
        self._user_action_taken=False
        self.protocol("WM_DELETE_WINDOW",self._on_close)
        # ===== HEADER =====
        ctk.CTkLabel(self,text="\U0001f4a1  P\u00c9RIODE OBSERVATOIRE DES PRIX",
                     font=("Segoe UI",17,"bold"),text_color=C["gold"]).pack(pady=(18,2))
        ctk.CTkLabel(self,text="Information utile pour la fin de mois en cours",
                     font=("Segoe UI",11),text_color=C["t2"]).pack(pady=(0,14))
        # ===== BLOC EXPLICATION =====
        box1=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8,
                          border_width=1,border_color=C["border"])
        box1.pack(fill="x",padx=20,pady=(0,10))
        ctk.CTkLabel(box1,text="POURQUOI CETTE ALERTE",
                     font=("Segoe UI",10,"bold"),text_color=C["t3"],
                     anchor="w").pack(anchor="w",padx=14,pady=(10,4))
        ctk.CTkLabel(box1,text=(
            "Entre le 25 et le 5 du mois, l'Observatoire des Prix Martinique "
            "annonce les nouveaux prix du SP, GO et GNR applicables au 1er. "
            "Les clients anticipent et viennent faire le plein avant le changement.\n\n"
            "\u2192 Hausse anticip\u00e9e : affluence + ventes en pic les 28-30\n"
            "\u2192 Baisse anticip\u00e9e : ventes en creux, puis pic apr\u00e8s le 1er"
            ),
            font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left",
            wraplength=600).pack(anchor="w",padx=14,pady=(0,12))
        # ===== BLOC À FAIRE =====
        box2=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8,
                          border_width=1,border_color=C["border"])
        box2.pack(fill="x",padx=20,pady=(0,14))
        ctk.CTkLabel(box2,text="\U0001f3af  \u00c0 FAIRE",
                     font=("Segoe UI",10,"bold"),text_color=C["amber"],
                     anchor="w").pack(anchor="w",padx=14,pady=(10,4))
        ctk.CTkLabel(box2,text=(
            "1. V\u00e9rifie tes pr\u00e9visions D7/D8/D9 dans Pre_vision compte.xlsx\n"
            "2. Si tu attends une hausse : majore le SP/GO sur les 27-30 du mois\n"
            "3. Si tu attends une baisse : minore et pr\u00e9pare le pic du 1er\n\n"
            "Le hub n'alertera pas si tes ventes pr\u00e9vues semblent atypiques sur "
            "cette p\u00e9riode (les seuils d'alerte sont automatiquement assouplis)."
            ),
            font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left",
            wraplength=600).pack(anchor="w",padx=14,pady=(0,12))
        # ===== BOUTONS =====
        btnrow=ctk.CTkFrame(self,fg_color="transparent")
        btnrow.pack(side="bottom",pady=(0,16))
        ctk.CTkButton(btnrow,text="\u2713 J'ai not\u00e9",width=140,height=36,
                      fg_color=C["green"],hover_color="#249239",text_color="#FFFFFF",
                      font=("Segoe UI",11,"bold"),corner_radius=8,
                      command=self._acquitter).pack(side="left",padx=4)
        ctk.CTkButton(btnrow,text="Demain 6h",width=110,height=36,
                      fg_color=C["card_h"],hover_color=C["border2"],text_color=C["t1"],
                      font=("Segoe UI",10),corner_radius=8,
                      command=lambda:self._snooze("demain_6h")).pack(side="left",padx=4)
        ctk.CTkButton(btnrow,text="Dans 3 jours",width=110,height=36,
                      fg_color=C["card_h"],hover_color=C["border2"],text_color=C["t1"],
                      font=("Segoe UI",10),corner_radius=8,
                      command=lambda:self._snooze("3j")).pack(side="left",padx=4)
        ctk.CTkButton(btnrow,text="Avant le 1er",width=110,height=36,
                      fg_color=C["card_h"],hover_color=C["border2"],text_color=C["t1"],
                      font=("Segoe UI",10),corner_radius=8,
                      command=lambda:self._snooze("avant_1er")).pack(side="left",padx=4)
    def _acquitter(self):
        """Acquittement définitif pour la période Observatoire en cours."""
        self._user_action_taken=True
        try:
            obs_path=APP_DIR/"observatoire_snooze.cfg"
            data=load_json(obs_path) or {}
            data[self.periode_key]={"acquitte":True,"ts":datetime.now().isoformat()}
            save_json(obs_path,data)
        except Exception as e: print(f"[observatoire ack] {e}")
        self.destroy()
    def _snooze(self,choix):
        """Reporte l'ouverture de la popup à un moment précis."""
        self._user_action_taken=True
        now=datetime.now()
        today=now.date()
        if choix=="demain_6h":
            rappel=datetime.combine(today+timedelta(days=1),datetime.min.time()).replace(hour=6)
        elif choix=="3j":
            rappel=datetime.combine(today+timedelta(days=3),datetime.min.time()).replace(hour=6)
        elif choix=="avant_1er":
            # Cible : le 30 du mois en cours pour une vérif finale avant changement
            if today.day<30:
                rappel_day=date(today.year,today.month,30)
                rappel=datetime.combine(rappel_day,datetime.min.time()).replace(hour=6)
            else:
                # Déjà passé le 30 : snooze court de 12h pour revenir bientôt
                rappel=now+timedelta(hours=12)
        else:
            rappel=now+timedelta(hours=12)  # fallback défensif
        try:
            obs_path=APP_DIR/"observatoire_snooze.cfg"
            data=load_json(obs_path) or {}
            data[self.periode_key]={"snooze_until":rappel.isoformat()}
            save_json(obs_path,data)
        except Exception as e: print(f"[observatoire snooze] {e}")
        self.destroy()
    def _on_close(self):
        """Fermeture passive (croix X). Pas d'acquittement, pas de snooze : la popup
        pourra se ré-ouvrir au prochain refresh."""
        self.destroy()


class RuptureImminenteDlg(ctk.CTkToplevel):
    """Popup d'alerte pour les ruptures imminentes en JOUR LIVRABLE.

    Cas couvert : autonomie carburant < 24h et aucune livraison prévue dans la fenêtre.
    NB : les ruptures DANS un trou (pont/weekend/férié) sont gérées par AntiRuptureDlg.

    3 actions utilisateur :
      - "Livraison commandée" : ack + silence jusqu'à demain 6h (ou détection livraison)
      - "Rappel dans X heures" : silence court (1/2/4/6h) pour y revenir
      - "Rupture acceptée" : ack + événement journal + silence demain 6h

    Croix X (fermeture sans action) → événement non_traité automatique (pastille rouge journal).
    """
    def __init__(self,parent,ruptures_imm):
        super().__init__(parent)
        self.title("\u26a0 Rupture imminente \u2014 DISTRICARB HUB")
        self.geometry("780x560");self.minsize(680,440)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        self.ruptures_imm=ruptures_imm
        # Flag pour gérer la fermeture par X = non_traité automatique (cohérent avec LivraisonsAReporterDlg)
        self._user_action_taken=False
        self.protocol("WM_DELETE_WINDOW",self._on_close)
        # Header : rouge si au moins un carburant < 6h (critique), amber sinon (urgent)
        any_critique=any(r.get("autonomie_h",24)<6 for r in ruptures_imm)
        if any_critique:
            hdr_color,accent_color,title_color="#3D1F1F",C["red"],C["red"]
            title_text="RUPTURE IMMINENTE \u2014 CRITIQUE"
            subtitle="Stock va atteindre le plancher physique dans moins de 6 heures"
        else:
            hdr_color,accent_color,title_color="#3D2F0F",C["amber"],C["amber"]
            title_text="RUPTURE IMMINENTE"
            subtitle="Stock va atteindre le plancher physique dans moins de 24 heures"
        hdr=ctk.CTkFrame(self,fg_color=hdr_color,corner_radius=0,height=80,border_width=0);hdr.pack(fill="x");hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\u26a0",font=("Segoe UI Emoji",30),text_color=accent_color).pack(side="left",padx=(24,12),pady=18)
        title_box=ctk.CTkFrame(hdr,fg_color="transparent");title_box.pack(side="left",fill="y",pady=14)
        ctk.CTkLabel(title_box,text=title_text,font=("Segoe UI",16,"bold"),text_color=title_color,anchor="w").pack(anchor="w")
        ctk.CTkLabel(title_box,text=subtitle,font=("Segoe UI",11),text_color=C["t2"],anchor="w").pack(anchor="w")
        # Corps scrollable
        body=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        body.pack(fill="both",expand=True,padx=20,pady=(12,8))
        # Encart explication
        info_box=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=8,border_width=1,border_color=C["border2"])
        info_box.pack(fill="x",padx=4,pady=(4,12))
        ctk.CTkLabel(info_box,text="\u2139 Pourquoi cette alerte",font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="w").pack(anchor="w",padx=12,pady=(8,2))
        ctk.CTkLabel(info_box,
                     text="Le stock actuel d'un ou plusieurs carburants permet moins de 24 heures de ventes\n"
                          "selon la moyenne historique du jour, ET aucune livraison n'est pr\u00e9vue avant l'\u00e9puisement.\n"
                          "Plancher physique : 500 L SP/GO, 250 L GNR (les pompes ne distribuent plus en dessous).",
                     font=("Segoe UI",10),text_color=C["t2"],anchor="w",justify="left").pack(anchor="w",padx=12,pady=(0,10))
        # Cartes par carburant en alerte
        jours_fr=["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
        fmt=lambda v:f"{int(v):,}".replace(",","\u202f")
        for r in ruptures_imm:
            carb=r.get("carburant","?")
            stock=r.get("stock_actuel",0)
            auto_h=r.get("autonomie_h",0)
            livr_jour=r.get("livr_aujourdhui",0)
            prochain_livr_iso=r.get("prochain_livr")
            critique=auto_h<6
            border_col=C["red"] if critique else C["amber"]
            card=ctk.CTkFrame(body,fg_color=C["card"],corner_radius=10,border_width=2,border_color=border_col)
            card.pack(fill="x",padx=4,pady=(8,4))
            ctk.CTkLabel(card,text=f"\u26fd {carb}",font=("Segoe UI",14,"bold"),text_color=border_col,anchor="w").pack(anchor="w",padx=14,pady=(12,2))
            niveau="CRITIQUE" if critique else "URGENT"
            ligne_stock=f"Stock actuel : {fmt(stock)} L \u2014 autonomie {auto_h}h ({niveau})"
            ctk.CTkLabel(card,text=ligne_stock,font=("Segoe UI",11),text_color=C["t1"],anchor="w").pack(anchor="w",padx=14,pady=2)
            # Cas 1 : livraison prévue aujourd'hui → message "livraison à surveiller"
            if livr_jour>0:
                msg_livr=f"\U0001f69a Livraison de {fmt(livr_jour)} L pr\u00e9vue aujourd'hui \u2014 doit arriver \u00e0 temps"
                ctk.CTkLabel(card,text=msg_livr,font=("Segoe UI",10,"bold"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(2,2))
                ctk.CTkLabel(card,text="\u26a0 Si la livraison est en retard, rupture commerciale dans la journ\u00e9e.",
                             font=("Segoe UI",10),text_color=C["t2"],anchor="w").pack(anchor="w",padx=14,pady=(0,12))
            # Cas 2 : aucune livraison prévue → message "commander"
            else:
                if prochain_livr_iso:
                    try:
                        plivr=datetime.fromisoformat(prochain_livr_iso).date()
                        jour_str=jours_fr[plivr.weekday()].capitalize()
                        delta_j=(plivr-date.today()).days
                        if delta_j==0:
                            livr_txt=f"\u2192 Aucune livraison pr\u00e9vue \u2014 prochain jour livrable : aujourd'hui ({jour_str} {plivr.strftime('%d/%m')})"
                        elif delta_j==1:
                            livr_txt=f"\u2192 Aucune livraison pr\u00e9vue \u2014 prochain jour livrable : demain ({jour_str} {plivr.strftime('%d/%m')})"
                        else:
                            livr_txt=f"\u2192 Aucune livraison pr\u00e9vue \u2014 prochain jour livrable : {jour_str} {plivr.strftime('%d/%m')} (dans {delta_j}j)"
                        ctk.CTkLabel(card,text=livr_txt,font=("Segoe UI",10),text_color=C["t2"],anchor="w").pack(anchor="w",padx=14,pady=(2,12))
                    except Exception:
                        ctk.CTkLabel(card,text=f"\u2192 Aucune livraison pr\u00e9vue \u2014 prochain jour livrable : {prochain_livr_iso}",font=("Segoe UI",10),text_color=C["t2"],anchor="w").pack(anchor="w",padx=14,pady=(2,12))
                else:
                    ctk.CTkLabel(card,text="\u2192 Aucune livraison pr\u00e9vue \u2014 aucun jour livrable trouv\u00e9 dans les 15 jours",font=("Segoe UI",10),text_color=C["t3"],anchor="w").pack(anchor="w",padx=14,pady=(2,12))
        # Footer avec 3 boutons d'action + bandeau d'info silence
        ctk.CTkLabel(self,text="L'alerte reviendra automatiquement si la situation s'aggrave (autonomie diminue).",
                     font=("Segoe UI",9,"italic"),text_color=C["t3"]).pack(pady=(0,4))
        footer=ctk.CTkFrame(self,fg_color="transparent");footer.pack(fill="x",padx=20,pady=(0,16))
        # Bouton 3 : Rupture acceptée (à droite)
        ctk.CTkButton(footer,text="\u2717 Rupture accept\u00e9e",width=170,height=40,
                      fg_color="transparent",text_color=C["red"],hover_color="#2A1818",
                      border_width=1,border_color=C["red"],
                      font=("Segoe UI",10),corner_radius=8,
                      command=self._rupture_acceptee).pack(side="right",padx=(0,8))
        # Bouton 2 : Rappel dans... (option menu — labels en DATES ABSOLUES via make_snooze_options).
        self._snooze_durations_h=[1,2,4,6]
        self._snooze_labels,self._snooze_mapping=make_snooze_options(self._snooze_durations_h)
        self.snooze_var=ctk.StringVar(value="\u23f1 Rappel dans...")
        ctk.CTkOptionMenu(footer,values=self._snooze_labels,
                          variable=self.snooze_var,
                          font=("Segoe UI",10),width=180,height=40,
                          fg_color=C["panel"],button_color=C["amber"],
                          button_hover_color="#B5721E",text_color=C["t1"],
                          dropdown_fg_color=C["card"],dropdown_text_color=C["t1"],
                          command=self._snooze).pack(side="right",padx=(0,8))
        # Bouton 1 : Livraison commandée (vert, action principale)
        ctk.CTkButton(footer,text="\U0001f69a Livraison command\u00e9e",width=200,height=40,fg_color=C["green"],hover_color="#15943C",
                      text_color="#FFF",font=("Segoe UI",12,"bold"),corner_radius=8,
                      command=self._livraison_commandee).pack(side="right",padx=(0,8))

    def _fingerprints(self):
        """Fingerprints stables pour silence_popup : {date_aujourdhui}_{carburant}."""
        d_str=date.today().strftime("%Y-%m-%d")
        return [f"{d_str}_{r.get('carburant','')}" for r in self.ruptures_imm]

    def _extra(self):
        """Extra pour silence_popup : {fingerprint: autonomie_h} → permet de détecter aggravation."""
        d_str=date.today().strftime("%Y-%m-%d")
        return {f"{d_str}_{r.get('carburant','')}":r.get("autonomie_h",0) for r in self.ruptures_imm}

    def _silence_until(self,until_iso):
        """Silence le popup_type 'rupture_imminente' jusqu'à until_iso."""
        try:
            silence_popup("rupture_imminente",self._fingerprints(),until_iso,self._extra())
        except Exception as e: print(f"[silence rupture imm] {e}")

    def _livraison_commandee(self):
        """Action 1 : livraison commandée → silence jusqu'à demain 6h."""
        self._user_action_taken=True
        tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
        self._silence_until(tomorrow_6h.isoformat())
        try:
            for r in self.ruptures_imm:
                add_evenement("rupture",{
                    "carburant":r.get("carburant",""),
                    "jour":date.today().strftime("%Y-%m-%d"),
                    "stock_actuel":r.get("stock_actuel",0),
                    "autonomie_h":r.get("autonomie_h",0),
                    "statut":"livraison_commandee",
                },commentaire=f"\u2713 Livraison command\u00e9e \u2014 stock {r.get('stock_actuel',0)} L, autonomie {r.get('autonomie_h',0)}h")
        except Exception as e: print(f"[evt rupture imm livr] {e}")
        self.destroy()

    def _snooze(self,choice):
        """Action 2 : rappel à une date absolue choisie → silence court."""
        self._user_action_taken=True
        h=self._snooze_mapping.get(choice,2)
        until=datetime.now()+timedelta(hours=h)
        self._silence_until(until.isoformat())
        self.destroy()

    def _rupture_acceptee(self):
        """Action 3 : rupture acceptée → silence demain 6h + événement journal."""
        self._user_action_taken=True
        tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
        self._silence_until(tomorrow_6h.isoformat())
        try:
            for r in self.ruptures_imm:
                add_evenement("rupture",{
                    "carburant":r.get("carburant",""),
                    "jour":date.today().strftime("%Y-%m-%d"),
                    "stock_actuel":r.get("stock_actuel",0),
                    "autonomie_h":r.get("autonomie_h",0),
                    "statut":"rupture_acceptee",
                },commentaire=f"\u2717 Rupture accept\u00e9e \u2014 stock {r.get('stock_actuel',0)} L, autonomie {r.get('autonomie_h',0)}h")
        except Exception as e: print(f"[evt rupture imm acc] {e}")
        self.destroy()

    def _on_close(self):
        """Croix X = non_traité automatique : silence court + événement journal lu=False."""
        if not self._user_action_taken:
            try:
                # Silence court (1h) pour que ça repopupe rapidement si pas traité
                until=datetime.now()+timedelta(hours=1)
                self._silence_until(until.isoformat())
                for r in self.ruptures_imm:
                    add_evenement("rupture",{
                        "carburant":r.get("carburant",""),
                        "jour":date.today().strftime("%Y-%m-%d"),
                        "stock_actuel":r.get("stock_actuel",0),
                        "autonomie_h":r.get("autonomie_h",0),
                        "statut":"non_traite",
                        "lu":False,
                    },commentaire=f"\u26a0 Popup ferm\u00e9e sans action \u2014 stock {r.get('stock_actuel',0)} L, autonomie {r.get('autonomie_h',0)}h")
            except Exception as e: print(f"[evt rupture imm on_close] {e}")
        self.destroy()


class LivraisonsAReporterDlg(ctk.CTkToplevel):
    """Popup d'alerte pour les livraisons à reporter (capacité cuve dépassée à 6h).
    Affiche pour chaque livraison concernée :
      - le surplus en L (combien dépasse la capacité)
      - l'heure de livraison recommandée (calculée via vitesse de vente C1)
    L'utilisateur peut acquitter en cliquant 'OK je gère'. Re-déclenchement uniquement
    si nouveau cas ou aggravation du surplus >500L.
    """
    def __init__(self,parent,livraisons):
        super().__init__(parent)
        self.title("\u26a0 Livraisons \u00e0 reporter \u2014 DISTRICARB HUB")
        # Pattern minimaliste type LivraisonDialog (cf. commentaire MargeTendueDlg).
        # Tout est packé directement dans self, pas de scrollable frame.
        self.geometry("760x600")
        self.configure(fg_color=C["bg"])
        self.resizable(False,False)
        self.transient(parent);self.grab_set()
        self.livraisons=livraisons
        # Flag pour détecter si l'utilisateur a pris une action explicite.
        # Si la popup est fermée par la croix X sans action, on traite comme "non_traité"
        # via _on_close, pour ne pas perdre l'alerte silencieusement.
        self._user_action_taken=False
        self.protocol("WM_DELETE_WINDOW",self._on_close)
        # ===== HEADER simple =====
        ctk.CTkLabel(self,text="\u26a0  LIVRAISONS \u00c0 REPORTER",
                     font=("Segoe UI",17,"bold"),text_color=C["amber"]).pack(pady=(18,2))
        ctk.CTkLabel(self,text="Une ou plusieurs livraisons d\u00e9passent la capacit\u00e9 cuve si re\u00e7ues \u00e0 6h",
                     font=("Segoe UI",11),text_color=C["t2"]).pack(pady=(0,12))
        # ===== ENCART explicatif =====
        info_box=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8)
        info_box.pack(fill="x",padx=24,pady=(0,8))
        ctk.CTkLabel(info_box,text="\u2139 Pourquoi cette alerte",
                     font=("Segoe UI",11,"bold"),text_color=C["gold"],anchor="w").pack(anchor="w",padx=14,pady=(8,2))
        ctk.CTkLabel(info_box,
                     text=("Capacit\u00e9 cuve : 40 000 L SP/GO, 10 000 L GNR. Si stock matin + livraison la d\u00e9passe,\n"
                           "livrer \u00e0 6h ferait d\u00e9border la cuve (interdit c\u00f4t\u00e9 Douane).\n"
                           "Solution : reporter la livraison de quelques heures, le temps que les ventes C1 lib\u00e8rent de la place."),
                     font=("Segoe UI",10),text_color=C["t2"],anchor="w",justify="left",wraplength=680).pack(anchor="w",padx=14,pady=(0,4))
        ctk.CTkLabel(info_box,
                     text=("\u00ab C'est fait \u00bb : silence jusqu'\u00e0 demain matin 6h.  \u00b7  "
                           "\u00ab Garder en alerte \u00bb : revient \u00e0 chaque refresh.  \u00b7  "
                           "Sinon choisis un rappel m\u00e9tier."),
                     font=("Segoe UI",9,"italic"),text_color=C["t3"],anchor="w",justify="left",wraplength=680).pack(anchor="w",padx=14,pady=(0,8))
        # ===== Cartes par jour =====
        par_jour={}
        for l in self.livraisons:
            par_jour.setdefault(l["date_str"],[]).append(l)
        for date_str,items in par_jour.items():
            card=ctk.CTkFrame(self,fg_color=C["card"],corner_radius=8)
            card.pack(fill="x",padx=24,pady=(0,8))
            ctk.CTkLabel(card,text=f"\U0001f69b {date_str}",
                         font=("Segoe UI",13,"bold"),text_color=C["amber"],anchor="w").pack(anchor="w",padx=14,pady=(10,2))
            for l in items:
                fmt=lambda v:f"{int(v):,}".replace(",","\u202f")
                line_txt=(f"   \u2022 {l['carburant']} : livraison de {fmt(l['livraison'])} L sur stock matin {fmt(l['stock_matin'])} L\n"
                          f"        \u2192 d\u00e9passement cuve : {fmt(l['surplus'])} L (capacit\u00e9 {fmt(l['capacite'])} L)")
                ctk.CTkLabel(card,text=line_txt,font=("Segoe UI",11),text_color=C["t1"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(2,2))
                # Recommandation
                if l.get("report_au_lendemain"):
                    rec_txt=f"        \u2192 Reco : reporter au LENDEMAIN (vente C1 {fmt(l.get('ventes_c1_moy',0))} L/j trop faible)."
                elif l.get("heures_attente") is not None and l.get("heure_recommandee") is not None:
                    h=l["heure_recommandee"]
                    h_h=int(h);h_m=int((h-h_h)*60)
                    h_str=f"{h_h:02d}h{h_m:02d}"
                    rec_txt=f"        \u2192 Reco : livrer \u00e0 partir de {h_str} (~{l['heures_attente']:.1f}h d'attente)."
                else:
                    rec_txt=f"        \u2192 Reco : reporter de quelques heures (donn\u00e9es C1 insuffisantes)."
                ctk.CTkLabel(card,text=rec_txt,font=("Segoe UI",10,"italic"),text_color=C["green"],anchor="w",justify="left").pack(anchor="w",padx=14,pady=(0,8))
        # ===== Footer boutons (adaptés au MOMENT MÉTIER courant) =====
        # Validé Bidou 20/05/2026 : même pattern que MargeTendueDlg.
        # detecter_moment_courant(closest_livr) → boutons utiles + texte d'aide explicite.
        now=datetime.now();today=now.date()
        livr_dates=[l.get("date") for l in self.livraisons if l.get("date") is not None]
        self._closest_livr=min(livr_dates) if livr_dates else (today+timedelta(days=2))
        mode=detecter_moment_courant(self._closest_livr)
        btns=ctk.CTkFrame(self,fg_color="transparent")
        btns.pack(fill="x",padx=24,pady=(8,4))
        aide_par_mode={
            "anticipation":         "L'alerte est anticip\u00e9e. Rien \u00e0 faire concr\u00e8tement maintenant : le hub te re-rappellera le jour de la commande pour que tu pr\u00e9viennes le transporteur de d\u00e9caler l'arriv\u00e9e.",
            "jour_commande_avant_10h30":"\u26a0 C'est AUJOURD'HUI que tu dois pr\u00e9venir le transporteur de d\u00e9caler l'heure d'arriv\u00e9e. Deadline TEMAG : 11h.",
            "jour_commande_deadline":"\u26a0\u26a0 DEADLINE TEMAG DANS MOINS DE 30 MIN. Appelle le transporteur MAINTENANT pour d\u00e9caler l'heure d'arriv\u00e9e.",
            "jour_commande_apres_11h":"\u26a0 Deadline TEMAG pass\u00e9e. Si tu n'as pas pr\u00e9venu le transporteur, appelle d'urgence pour qu'il arrive plus tard.",
            "entre_commande_et_livraison":"Heure d'arriv\u00e9e ajust\u00e9e avec le transporteur. V\u00e9rification finale le matin de la livraison \u00e0 6h (chiffres + heure de pr\u00e9sentation du camion).",
            "matin_livraison":      "Matin de la livraison. V\u00e9rifie que la cuve a la place pour recevoir, et que le camion arrive bien \u00e0 l'heure ajust\u00e9e.",
            "livraison_passee":     "Livraison pass\u00e9e. Cette alerte n'a plus lieu d'\u00eatre, tu peux la marquer r\u00e9solue.",
        }
        if mode=="anticipation":
            tk.Button(btns,text="\u2713 OK not\u00e9 \u2014 rappelle-moi le jour de la commande",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=lambda:self._resolve_direct("\u2713 Anticipation not\u00e9e")).pack(side="left")
        elif mode=="jour_commande_avant_10h30":
            tk.Button(btns,text="\u2713 Transporteur pr\u00e9venu",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_fait).pack(side="left",padx=(0,10))
            tk.Button(btns,text="\u23f1 Pas encore \u2014 rappelle-moi \u00e0 10h30",
                      bg="#E6A53D",fg="#000000",font=("Segoe UI",10,"bold"),
                      relief="raised",bd=2,padx=16,pady=8,cursor="hand2",
                      command=self._snooze_avant_deadline_tk).pack(side="left")
        elif mode=="jour_commande_deadline":
            tk.Button(btns,text="\u2713 Transporteur pr\u00e9venu",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_fait).pack(side="left",padx=(0,10))
            tk.Button(btns,text="\u26a0 Pas encore — URGENT",
                      bg="#C8362E",fg="#FFFFFF",font=("Segoe UI",10,"bold"),
                      relief="raised",bd=2,padx=14,pady=8,cursor="hand2",
                      command=self._snooze_15min_tk).pack(side="left")
        elif mode=="jour_commande_apres_11h":
            tk.Button(btns,text="\u2713 Transporteur pr\u00e9venu",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=self._cest_fait).pack(side="left")
        elif mode=="entre_commande_et_livraison":
            tk.Button(btns,text="\u2713 OK not\u00e9 \u2014 rappelle-moi le matin de la livraison",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=lambda:self._resolve_direct("\u2713 OK not\u00e9 entre commande et livraison")).pack(side="left")
        elif mode=="matin_livraison":
            tk.Button(btns,text="\u2713 V\u00e9rifi\u00e9, prêt à recevoir",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=lambda:self._resolve_direct("\u2713 V\u00e9rifi\u00e9 matin livraison")).pack(side="left")
        else:  # livraison_passee ou fallback
            tk.Button(btns,text="\u2713 Marquer comme r\u00e9gl\u00e9",
                      bg="#2BA84A",fg="#FFFFFF",font=("Segoe UI",11,"bold"),
                      relief="raised",bd=2,padx=20,pady=8,cursor="hand2",
                      command=lambda:self._resolve_direct("\u2713 Marqu\u00e9 r\u00e9solu manuellement")).pack(side="left")
        # Bouton "Garder en alerte" toujours présent à droite (mode harcèlement consenti)
        tk.Button(btns,text="\u26a0 Garder en alerte",
                  bg="#1A2332",fg="#C8362E",font=("Segoe UI",10),
                  relief="raised",bd=2,padx=14,pady=8,cursor="hand2",
                  command=self._non_traite).pack(side="right")
        # Texte d'aide en bas (toujours présent, change selon le mode)
        aide_txt=aide_par_mode.get(mode,"")
        if aide_txt:
            ctk.CTkLabel(self,text=aide_txt,font=("Segoe UI",10,"italic"),
                         text_color=C["t2"],wraplength=680,justify="left").pack(padx=24,pady=(4,16),anchor="w")

    def _silence_until(self,until_iso,cap_5h30=True):
        """Silence la popup jusqu'à until_iso (datetime ISO). Helper commun aux 3 boutons.

        Garde-fou (b) : si cap_5h30=True (snooze ou non_traite), plafonne le silence à
        5h30 du jour de livraison (réveil dernière chance avant l'arrivée du camion à 6h).
        Si cap_5h30=False (clic 'C'est fait' = resolu), respecte le silence demandé sans plafond.
        """
        try:
            cap_until=until_iso
            if cap_5h30:
                try:
                    until_dt=datetime.fromisoformat(until_iso)
                    for l in self.livraisons:
                        livr_d=l.get("date")
                        if not livr_d: continue
                        rappel=datetime.combine(livr_d,datetime.min.time()).replace(hour=5,minute=30)
                        if rappel>datetime.now() and rappel<until_dt:
                            until_dt=rappel
                    cap_until=until_dt.isoformat()
                except Exception as _e: _log_silent_err(exc=_e)
            fps=[f"{l['date'].strftime('%Y-%m-%d')}_{l['carburant']}" for l in self.livraisons]
            extra={f"{l['date'].strftime('%Y-%m-%d')}_{l['carburant']}":l.get("surplus",0) for l in self.livraisons}
            silence_popup("livr_report",fps,cap_until,extra)
        except Exception as e: print(f"[silence livr report] {e}")

    def _save_ack(self):
        """Snapshot des surplus actuels pour détecter les aggravations futures."""
        try:
            acks=load_json(LIVR_REPORT_ACK_FILE) or {}
        except Exception: acks={}
        snapshot={}
        for l in self.livraisons:
            key=f"{l['date'].strftime('%Y-%m-%d')}_{l['carburant']}"
            snapshot[key]=l["surplus"]
        acks["last_ack"]={"ack_at_iso":datetime.now().isoformat(),"snapshot":snapshot}
        try: save_json(LIVR_REPORT_ACK_FILE,acks)
        except Exception as e: print(f"[ack livr report] {e}")

    def _snooze_2h_tk(self):
        """Snooze 2h via bouton tk fallback (CTkOptionMenu invisible chez Bidou)."""
        self._user_action_taken=True
        self._save_ack()
        until=datetime.now()+timedelta(hours=2)
        self._silence_until(until.isoformat())
        try:
            for l in self.livraisons:
                add_evenement("livraison_reporter",{
                    "date":l["date"].isoformat(),
                    "carburant":l["carburant"],
                    "stock_matin":l["stock_matin"],
                    "livraison":l["livraison"],
                    "surplus":l["surplus"],
                    "statut":"snooze",
                    "snooze_heures":2,
                    "snooze_until":until.isoformat(),
                },commentaire=f"\u23f1 Rappel demand\u00e9 \u2192 dans 2h")
        except Exception as e: print(f"[evt livr report snooze 2h tk] {e}")
        self.destroy()

    def _snooze_demain_matin_tk(self):
        """Snooze jusqu'à demain matin 6h via bouton tk fallback."""
        self._user_action_taken=True
        self._save_ack()
        tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),dt_time(6, 0))
        self._silence_until(tomorrow_6h.isoformat())
        try:
            for l in self.livraisons:
                add_evenement("livraison_reporter",{
                    "date":l["date"].isoformat(),
                    "carburant":l["carburant"],
                    "stock_matin":l["stock_matin"],
                    "livraison":l["livraison"],
                    "surplus":l["surplus"],
                    "statut":"snooze",
                    "snooze_until":tomorrow_6h.isoformat(),
                },commentaire="\u23f1 Rappel demain matin 6h")
        except Exception as e: print(f"[evt livr report snooze demain tk] {e}")
        self.destroy()

    def _snooze_avant_deadline_tk(self):
        """Rappel à 10h30 du jour de la commande (juste avant deadline TEMAG 11h).
        Si 10h30 est passé, snooze court 15 min."""
        self._user_action_taken=True
        self._save_ack()
        cible=getattr(self,"_closest_livr",None) or (date.today()+timedelta(days=2))
        jc=jour_de_commande(cible) or (cible-timedelta(days=1))
        until_dt=datetime.combine(jc,dt_time(10,30))
        if until_dt<=datetime.now():
            until_dt=datetime.now()+timedelta(minutes=15)
        self._silence_until(until_dt.isoformat())
        try:
            for l in self.livraisons:
                add_evenement("livraison_reporter",{
                    "date":l["date"].isoformat(),
                    "carburant":l["carburant"],
                    "stock_matin":l["stock_matin"],
                    "livraison":l["livraison"],
                    "surplus":l["surplus"],
                    "statut":"snooze",
                    "snooze_until":until_dt.isoformat(),
                },commentaire=f"\u23f1 Rappel avant deadline TEMAG \u2192 {until_dt.strftime('%H:%M')}")
        except Exception as e: print(f"[evt livr report snooze deadline tk] {e}")
        self.destroy()

    def _snooze_15min_tk(self):
        """Snooze court 15 min — utilisé dans le mode URGENT proche deadline TEMAG."""
        self._user_action_taken=True
        self._save_ack()
        until=datetime.now()+timedelta(minutes=15)
        self._silence_until(until.isoformat())
        try:
            for l in self.livraisons:
                add_evenement("livraison_reporter",{
                    "date":l["date"].isoformat(),
                    "carburant":l["carburant"],
                    "stock_matin":l["stock_matin"],
                    "livraison":l["livraison"],
                    "surplus":l["surplus"],
                    "statut":"snooze",
                    "snooze_heures":0.25,
                    "snooze_until":until.isoformat(),
                },commentaire="\u23f1 Rappel URGENT dans 15 min")
        except Exception as e: print(f"[evt livr report snooze 15min tk] {e}")
        self.destroy()

    def _resolve_direct(self,commentaire="\u2713 R\u00e9solu"):
        """Résolution DIRECTE sans mini-dialog 2e/3e tour.
        
        Utilisé pour les modes où demander un tour n'a pas de sens :
        - anticipation (silence jusqu'au jour de commande)
        - entre_commande_et_livraison (silence jusqu'au matin de la livraison)
        - matin_livraison (livraison en cours/imminente)
        - livraison_passee
        
        Bug fix 21/05/2026 : avant, ces modes pointaient sur _cest_fait qui ouvrait
        un mini-dialog 2e/3e tour totalement inadapté. Bidou annulait, le silence
        n'était jamais posé, la popup revenait à chaque refresh.
        
        Silence calculé via prochain_moment_cle_strict (>now strict) jusqu'au
        prochain moment utile, ou +24h si plus aucun moment-clé futur."""
        self._user_action_taken=True
        self._save_ack()
        cible=getattr(self,"_closest_livr",None) or (date.today()+timedelta(days=2))
        mc=prochain_moment_cle_strict(cible)
        if mc is None:
            until_dt=datetime.now()+timedelta(hours=24)
        else:
            until_dt=mc[0]
        self._silence_until(until_dt.isoformat(),cap_5h30=False)
        try:
            for l in self.livraisons:
                add_evenement("livraison_reporter",{
                    "date":l["date"].isoformat(),
                    "carburant":l["carburant"],
                    "stock_matin":l["stock_matin"],
                    "livraison":l["livraison"],
                    "surplus":l["surplus"],
                    "heure_recommandee":l.get("heure_recommandee"),
                    "heures_attente":l.get("heures_attente"),
                    "statut":"resolu",
                },commentaire=commentaire)
        except Exception as e: print(f"[evt livr report resolve direct] {e}")
        self.destroy()

    def _cest_fait(self):
        """Ouvre un mini-choix 2e tour / 3e tour / Annuler avant validation finale.
        
        Utilisé UNIQUEMENT dans les modes jour_commande_* où Bidou peut effectivement
        décaler au 2e ou 3e tour TEMAG. Dans les autres modes (anticipation, matin
        livraison, livraison passée), on utilise _resolve_direct qui résout sans
        demander de tour (qui n'aurait pas de sens dans ces contextes)."""
        choice_dlg=ctk.CTkToplevel(self)
        choice_dlg.title("\u2713 Confirmer la livraison report\u00e9e")
        choice_dlg.geometry("420x240");choice_dlg.minsize(400,220)
        choice_dlg.configure(fg_color=C["bg"]);choice_dlg.transient(self);choice_dlg.grab_set()
        ctk.CTkLabel(choice_dlg,text="\u00c0 quel tour la livraison est-elle\nreport\u00e9e ?",
                     font=("Segoe UI",13,"bold"),text_color=C["t1"],justify="center").pack(pady=(28,8))
        ctk.CTkLabel(choice_dlg,text="Choisis le moment auquel le camion\nsera autoris\u00e9 \u00e0 d\u00e9verser.",
                     font=("Segoe UI",10),text_color=C["t3"],justify="center").pack(pady=(0,16))
        btns=ctk.CTkFrame(choice_dlg,fg_color="transparent");btns.pack(pady=8)
        def _confirme(tour_label):
            self._user_action_taken=True
            self._save_ack()
            # Silence définitif (resolu) jusqu'à après-demain 6h. PAS de cap à 5h30 du jour J
            # car l'utilisateur a confirmé que c'est géré.
            today=date.today()
            silence_until=datetime.combine(today+timedelta(days=2),datetime.min.time()).replace(hour=6)
            self._silence_until(silence_until.isoformat(),cap_5h30=False)
            try:
                for l in self.livraisons:
                    add_evenement("livraison_reporter",{
                        "date":l["date"].isoformat(),
                        "carburant":l["carburant"],
                        "stock_matin":l["stock_matin"],
                        "livraison":l["livraison"],
                        "surplus":l["surplus"],
                        "heure_recommandee":l.get("heure_recommandee"),
                        "heures_attente":l.get("heures_attente"),
                        "statut":"resolu",
                        "tour":tour_label,
                    },commentaire=f"\u2713 Report\u00e9 au {tour_label}")
            except Exception as e: print(f"[evt livr report fait] {e}")
            choice_dlg.destroy();self.destroy()
        # FALLBACK BOUTONS STANDARDS — voir commentaire dans LivraisonsAReporterDlg
        tk.Button(btns,text="2\u1d49 tour\n(~9h)",
                  bg="#2BA84A",fg="#FFFFFF",
                  font=("Segoe UI",11,"bold"),
                  relief="raised",bd=2,padx=20,pady=14,
                  cursor="hand2",
                  command=lambda:_confirme("2\u1d49 tour")).pack(side="left",padx=8)
        tk.Button(btns,text="3\u1d49 tour\n(~11h)",
                  bg="#2BA84A",fg="#FFFFFF",
                  font=("Segoe UI",11,"bold"),
                  relief="raised",bd=2,padx=20,pady=14,
                  cursor="hand2",
                  command=lambda:_confirme("3\u1d49 tour")).pack(side="left",padx=8)
        tk.Button(choice_dlg,text="Annuler",
                  bg="#1A2332",fg="#A8B5C4",
                  font=("Segoe UI",10),
                  relief="raised",bd=1,padx=14,pady=4,
                  cursor="hand2",
                  command=choice_dlg.destroy).pack(pady=(12,16))

    def _snooze(self,choice):
        """Rappel dans X heures/jours : silence jusqu'à now+X. Le cap_5h30 (dans
        _silence_until) plafonne automatiquement le silence à 5h30 du jour de livraison
        pour les longues durées (ex. '1 semaine' pour une livraison J+3 → cap à J+3 05h30)."""
        self._user_action_taken=True
        # Mapping inverse via labels dynamiques générés à l'ouverture de la popup.
        h=self._snooze_mapping.get(choice,2)
        until=datetime.now()+timedelta(hours=h)
        until_label=fmt_rappel_dt(until)
        self._save_ack()
        self._silence_until(until.isoformat())
        try:
            for l in self.livraisons:
                add_evenement("livraison_reporter",{
                    "date":l["date"].isoformat(),
                    "carburant":l["carburant"],
                    "stock_matin":l["stock_matin"],
                    "livraison":l["livraison"],
                    "surplus":l["surplus"],
                    "heure_recommandee":l.get("heure_recommandee"),
                    "heures_attente":l.get("heures_attente"),
                    "statut":"snooze",
                    "snooze_heures":h,
                    "snooze_until":until.isoformat(),
                },commentaire=f"\u23f1 Rappel demand\u00e9 \u2192 {until_label}")
        except Exception as e: print(f"[evt livr report snooze] {e}")
        self.destroy()

    def _non_traite(self):
        """Non traité : silence jusqu'à demain 6h, événement journal avec statut non_traite + lu=False."""
        self._user_action_taken=True
        self._save_ack()
        tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
        self._silence_until(tomorrow_6h.isoformat())
        try:
            for l in self.livraisons:
                add_evenement("livraison_reporter",{
                    "date":l["date"].isoformat(),
                    "carburant":l["carburant"],
                    "stock_matin":l["stock_matin"],
                    "livraison":l["livraison"],
                    "surplus":l["surplus"],
                    "heure_recommandee":l.get("heure_recommandee"),
                    "heures_attente":l.get("heures_attente"),
                    "statut":"non_traite",
                    "lu":False,  # Pastille rouge dans le journal tant que non lu
                },commentaire="\u26a0 Non trait\u00e9 \u2014 \u00e0 reprendre")
        except Exception as e: print(f"[evt livr report non traite] {e}")
        self.destroy()

    def _on_close(self):
        """Intercepteur de fermeture par la croix X. Si l'utilisateur n'a pris aucune
        action explicite (C'est fait / Rappel / Non traité), on traite la fermeture
        comme un 'non_traité' automatique : silence jusqu'à demain 6h + événement
        journal avec statut=non_traite + lu=False. Évite de perdre l'alerte silencieusement."""
        if not self._user_action_taken:
            try:
                self._save_ack()
                tomorrow_6h=datetime.combine(date.today()+timedelta(days=1),datetime.min.time()).replace(hour=6)
                self._silence_until(tomorrow_6h.isoformat())
                for l in self.livraisons:
                    add_evenement("livraison_reporter",{
                        "date":l["date"].isoformat(),
                        "carburant":l["carburant"],
                        "stock_matin":l["stock_matin"],
                        "livraison":l["livraison"],
                        "surplus":l["surplus"],
                        "heure_recommandee":l.get("heure_recommandee"),
                        "heures_attente":l.get("heures_attente"),
                        "statut":"non_traite",
                        "lu":False,
                    },commentaire="\u26a0 Popup ferm\u00e9e sans action \u2014 marqu\u00e9 non trait\u00e9")
            except Exception as e: print(f"[evt livr report on_close] {e}")
        self.destroy()


# =============================================================================
class AlertesDashboardDlg(ctk.CTkToplevel):
    """Tableau de notifications — modèle SITUATION unifié.

    PRINCIPE : une situation = une carte = une ligne journal qui s'enrichit.
    Plus de doublons entre "alerte active", "alerte en pause" et "non traité journal" :
    les 3 sources sont fusionnées par sit_id stable pour produire UNE carte par situation.

    SOURCES :
      - last_data['antirupture'] : alertes calculées actuellement actives
      - popup_silence.cfg : silences (snoozes) en cours
      - evenements.cfg : historique journal (statut, non_traite, etc.)

    SIT_ID : identifiant stable d'une situation, ex. `livr_report:SP:2026-05-21`.
    Toutes les sources se mappent vers le même sit_id, ce qui permet l'unification.

    SCORE D'URGENCE (escalade visuelle, fond rouge plus appuyé) :
      0 = normal : 1ère apparition, peu d'historique
      1 = escalade : vue 2-3 fois OU non traitée < 24h
      2 = critique : vue 4+ fois OU non traitée > 24h
    Plus le score est élevé, plus la carte ressort visuellement.

    ACTION SUR UNE CARTE :
      - Clic = rouvre la popup d'origine pour traiter la situation
      - Bouton "✓ Résolu" = lève le silence + marque le journal résolu (clôture rapide)
    """
    def __init__(self,parent):
        super().__init__(parent)
        self.parent_app=parent
        self.title("\U0001f514 Tableau de notifications \u2014 DISTRICARB HUB")
        self.geometry("900x700");self.minsize(780,520)
        self.configure(fg_color=C["bg"]);self.transient(parent);self.grab_set()
        # Header
        hdr=ctk.CTkFrame(self,fg_color="transparent",height=70);hdr.pack(fill="x",padx=24,pady=(20,8));hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,text="\U0001f514  Tableau de notifications",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w")
        # Subtitle multi-éléments : nombre d'alertes + ÉTAT DE FRAÎCHEUR
        # (montre que le tableau VIT, qu'il n'est pas figé sur des vieilles données)
        sub_row=ctk.CTkFrame(hdr,fg_color="transparent");sub_row.pack(anchor="w",pady=(2,0),fill="x")
        self.subtitle_lbl=ctk.CTkLabel(sub_row,text="",font=("Segoe UI",11),text_color=C["t3"])
        self.subtitle_lbl.pack(side="left")
        self.freshness_lbl=ctk.CTkLabel(sub_row,text="",font=("Segoe UI",10),text_color=C["t3"])
        self.freshness_lbl.pack(side="left",padx=(12,0))
        # Corps scrollable
        self.body=ctk.CTkScrollableFrame(self,fg_color=C["bg"],corner_radius=0,
                                          scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        self.body.pack(fill="both",expand=True,padx=24,pady=4)
        # Footer
        footer=ctk.CTkFrame(self,fg_color="transparent",height=58);footer.pack(side="bottom",fill="x",padx=24,pady=(8,18));footer.pack_propagate(False)
        ctk.CTkLabel(footer,text="Clique une carte pour rouvrir la popup. Bouton \u2713 R\u00e9solu pour cl\u00f4turer.",
                     font=("Segoe UI",10),text_color=C["t3"]).pack(side="left",padx=(4,0),pady=14)
        ctk.CTkButton(footer,text="Fermer",width=110,height=36,
                      fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                      border_width=1,border_color=C["border2"],
                      font=("Segoe UI",11),corner_radius=8,
                      command=self._close).pack(side="right")
        ctk.CTkButton(footer,text="\U0001f5d1 Tout effacer",width=130,height=36,
                      fg_color=C["panel"],hover_color="#5A2222",text_color=C["t1"],
                      border_width=1,border_color=C["border2"],
                      font=("Segoe UI",11),corner_radius=8,
                      command=self._resolve_all).pack(side="right",padx=(0,10))
        self._render()
        # Tick périodique pour rafraîchir la ligne "prochain refresh dans..."
        # toutes les 30s, sans toucher au reste du tableau.
        self._freshness_tick_id=None
        self._update_freshness()

    def _close(self):
        """Stop le tick avant destruction pour éviter les callbacks fantômes."""
        try:
            if self._freshness_tick_id is not None:
                self.after_cancel(self._freshness_tick_id)
        except Exception: pass
        # Refresh complet du HUB UNE seule fois, à la fermeture, si des clôtures ont eu
        # lieu (au lieu d'un refresh par clôture). Met à jour le badge du bouton Alertes.
        if getattr(self,"_needs_parent_refresh",False):
            try:
                if hasattr(self.parent_app,"refresh"): self.parent_app.refresh()
            except Exception as _e: _log_silent_err(exc=_e)
        self.destroy()

    def _update_freshness(self):
        """Met à jour l'indicateur de fraîcheur (actualisé à X, prochain dans Y).
        Tourne en boucle toutes les 30s tant que le tableau est ouvert.
        Couleur change selon l'âge : vert <2 min, gris <15 min, ambre <30 min, rouge au-delà."""
        try:
            ts=(self.parent_app.last_data or {}).get("ts","")
            if ts:
                # ts est au format "HH:MM:SS" du hub principal
                try:
                    hh,mm,ss=ts.split(":")
                    refresh_dt=datetime.now().replace(hour=int(hh),minute=int(mm),second=int(ss),microsecond=0)
                    # Si refresh_dt est dans le futur, c'est qu'il est d'hier (cas marginal)
                    if refresh_dt>datetime.now()+timedelta(hours=1):
                        refresh_dt-=timedelta(days=1)
                    age_min=int((datetime.now()-refresh_dt).total_seconds()/60)
                    # Prochain refresh = il y a (age) min, donc dans (15 - age) min
                    next_in=max(0,15-age_min)
                    color=C["green"] if age_min<2 else (C["t3"] if age_min<15 else (C["amber"] if age_min<30 else C["red"]))
                    txt=f"\u00b7 Actualis\u00e9 \u00e0 {ts} (il y a {age_min} min, prochain dans {next_in} min)"
                    self.freshness_lbl.configure(text=txt,text_color=color)
                except Exception:
                    self.freshness_lbl.configure(text=f"\u00b7 Actualis\u00e9 \u00e0 {ts}",text_color=C["t3"])
            else:
                self.freshness_lbl.configure(text="\u00b7 Pas encore actualis\u00e9",text_color=C["amber"])
        except Exception: pass
        # Re-planifier dans 30s
        try:
            self._freshness_tick_id=self.after(30000,self._update_freshness)
        except Exception: pass

    # =========================================================================
    # COLLECTE ET UNIFICATION DES SITUATIONS
    # =========================================================================
    def _collect_situations(self):
        """Fusionne les 3 sources (actives + silences + journal) en liste de Situations.
        Une situation = un dict avec sit_id stable + état + historique + score."""
        situations={}  # sit_id -> dict
        ar=(self.parent_app.last_data or {}).get("antirupture",{}) or {}
        # ---- Source 1 : ALERTES ACTIVES ----
        for lr in (ar.get("livraisons_a_reporter",[]) or []):
            sid=self._sit_id_livr_report(lr)
            self._upsert(situations,sid,category="livr_report",
                label=self._label_livr_report(lr),
                sub_details=self._details_livr_report(lr),
                is_active=True,raw=lr)
        for s in (ar.get("saisies_physiquement_impossibles",[]) or []):
            sid=self._sit_id_saisies(s)
            self._upsert(situations,sid,category="saisies_irr",
                label=self._label_saisies(s),
                sub_details=self._details_saisies(s),
                is_active=True,raw=s)
        for r in (ar.get("ruptures_imminentes",[]) or []):
            sid=self._sit_id_rupture_imminente(r)
            self._upsert(situations,sid,category="rupture_imminente",
                label=self._label_rupture_imm(r),
                sub_details=self._details_rupture_imm(r),
                is_active=True,raw=r)
        # Ponts non acquittés (catégorie antirupture)
        ack_status=ar.get("ack_status",{}) or {}
        for pid,st in ack_status.items():
            if st.get("acquitte"): continue
            sid=f"antirupture:{pid}"
            self._upsert(situations,sid,category="antirupture",
                label=self._label_pont(pid),
                sub_details=self._details_pont(st),
                is_active=True,raw={"pont_id":pid,"info":st})
        # Incohérences jour non-livrable (pas de silence applicable)
        for it in (ar.get("incoherences_jour_non_livrable",[]) or []):
            d=it.get("date")
            d_str=d.strftime("%Y-%m-%d") if hasattr(d,"strftime") else str(d)
            sid=f"incoh:{it.get('carburant','')}:{d_str}"
            self._upsert(situations,sid,category="incoh",
                label=f"{it.get('carburant','?')} \u2014 livraison sur jour non-livrable le {d_str}",
                sub_details="\u00e0 v\u00e9rifier",
                is_active=True,raw=it)
        # Ruptures projetées (pas de silence)
        for it in (ar.get("ruptures_projetees",[]) or []):
            d=it.get("date")
            d_str=d.strftime("%Y-%m-%d") if hasattr(d,"strftime") else str(d)
            sid=f"rupt_proj:{it.get('carburant','')}:{d_str}"
            self._upsert(situations,sid,category="rupt_proj",
                label=f"Rupture projet\u00e9e {it.get('carburant','?')} le {d_str}",
                sub_details="",
                is_active=True,raw=it)
        # ---- Source 2 : SILENCES (snoozes en cours) ----
        try:
            sd=_load_silence() or {}
            now=datetime.now()
            for ptype,d in sd.items():
                if not d: continue
                until_iso=d.get("until_iso","")
                if not until_iso: continue
                try: until_dt=datetime.fromisoformat(until_iso)
                except Exception: continue
                if until_dt<=now: continue
                for fp in (d.get("fingerprints",[]) or []):
                    sid=self._sit_id_from_silence(ptype,fp)
                    if not sid: continue
                    if sid in situations:
                        # Situation déjà créée (active) : ajoute juste l'info de silence.
                        # Les chiffres viennent de Source 1 = RECALCULÉS ce refresh. OK.
                        situations[sid]["is_silenced"]=True
                        situations[sid]["silence_until"]=until_dt
                    else:
                        # Situation silencée mais ABSENTE du calcul actuel.
                        # BUG CORRIGÉ (terrain Bidou 18/05) : avant, on recréait une
                        # carte "fossile" depuis le fingerprint figé (_label_from_silence)
                        # → vieux chiffres trompeurs (ex : "dépassement 4453 L" alors que
                        # Pre_vision avait été corrigé et que le vrai dépassement valait
                        # quelques centaines de L, ou avait disparu).
                        #
                        # Nouveau : on cherche d'abord une situation ÉQUIVALENTE déjà
                        # présente (même catégorie + même date cible, fingerprint
                        # différent car les chiffres ont changé). Si trouvée → c'est la
                        # version FRAÎCHE : on lui transfère l'état de pause et on
                        # n'affiche PAS la coquille figée.
                        equiv=self._find_equivalent_active(situations,ptype,fp)
                        if equiv is not None:
                            situations[equiv]["is_silenced"]=True
                            situations[equiv]["silence_until"]=until_dt
                        else:
                            # Aucune situation équivalente active : la situation a
                            # réellement disparu du calcul (résolue de fait : Pre_vision
                            # corrigé, livraison saisie, etc.). On NE la ressuscite PAS
                            # avec de vieux chiffres. Le silence devenu orphelin sera
                            # nettoyé par les invariants de résolution plus bas.
                            pass
        except Exception as _e: _log_silent_err(exc=_e)
        # ---- Source 3 : JOURNAL (historique d'apparitions et non_traite) ----
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            for evt in all_evt.get("events",[]):
                sid=self._sit_id_from_evt(evt)
                if not sid: continue
                evt_data=evt.get("data",{}) or {}
                if sid in situations:
                    s=situations[sid]
                    s["journal_entries"].append(evt)
                    s["nb_apparitions"]+=1
                    if evt_data.get("statut")=="non_traite" and not evt_data.get("lu",True):
                        s["has_non_traite"]=True
                        try:
                            ts=datetime.fromisoformat(evt.get("ts",""))
                            if s["last_non_traite_ts"] is None or ts>s["last_non_traite_ts"]:
                                s["last_non_traite_ts"]=ts
                        except Exception as _e: _log_silent_err(exc=_e)
                else:
                    # Événement journal non lié à une situation active ou silencée :
                    # ne créer une carte QUE si c'est un non_traite (sinon trace pure)
                    if evt_data.get("statut")=="non_traite" and not evt_data.get("lu",True):
                        cat=self._cat_from_evt_type(evt.get("type",""))
                        # FILTRE BUG COHÉRENCE (signalé Bidou 20/05) : certaines catégories
                        # sont INSTANTANÉES par nature — leur état "non traité" du passé
                        # n'a plus de sens si la condition est résolue aujourd'hui.
                        #
                        # Exemple : "Rupture imminente du 19/05 07h31" non traitée. Si
                        # aujourd'hui (20/05) le stock est normal, la rupture est résolue
                        # de fait : il n'y a plus rien à faire dessus. La maintenir comme
                        # "alerte active" dans le tableau est trompeur — Bidou clique
                        # dessus et le hub lui répond "Cette alerte n'est plus d'actualité".
                        #
                        # Catégories instantanées (état présent uniquement) :
                        #   - rupture_imminente : dépend du stock courant
                        #   - rupture : idem
                        # Pour ces catégories, si l'événement n'a PAS de situation active
                        # ce refresh, on ne crée pas de carte fantôme. Le non_traite reste
                        # dans le journal pour la traçabilité, mais pas dans le tableau.
                        #
                        # Catégories à mémoire (gardées même sans situation active) :
                        #   - livr_report : dépend de Pre_vision saisi par Bidou. Si Pre_vision
                        #     n'est plus chargé, l'alerte peut temporairement disparaître du
                        #     calcul sans être réellement résolue.
                        #   - saisies_irr : idem
                        #   - antirupture (ponts) : à long terme
                        if cat in ("rupture_imminente","rupture"):
                            # Saute : on ne crée PAS de carte dans le tableau, mais on
                            # garde l'événement dans le journal pour l'historique.
                            continue
                        lbl=self._label_from_evt(evt)
                        self._upsert(situations,sid,category=cat,label=lbl,
                            has_non_traite=True,raw=None)
                        situations[sid]["journal_entries"].append(evt)
                        situations[sid]["nb_apparitions"]=1
                        try:
                            ts=datetime.fromisoformat(evt.get("ts",""))
                            situations[sid]["last_non_traite_ts"]=ts
                        except Exception as _e: _log_silent_err(exc=_e)
        except Exception as _e: _log_silent_err(exc=_e)
        # ---- Calcul du score d'urgence pour chaque situation ----
        for sid,s in situations.items():
            s["score"]=self._compute_score(s)
        # ---- RÉSOLUTION AUTO PAR INVARIANT MÉTIER ----
        # Certaines situations sont résolues *de fait* par l'état des données, même si
        # le silence n'a pas été levé ni le journal marqué résolu. Exemple :
        # "Livraison du jour" est résolue dès que la livraison du jour est saisie dans
        # livraisons.cfg, sans que l'utilisateur n'ait à cliquer "Résolu".
        # On nettoie en passant : silence levé + journal marqué résolu, automatiquement.
        self._apply_resolution_invariants(situations)
        return list(situations.values())

    def _apply_resolution_invariants(self,situations):
        """Filtre les situations rendues caduques par les données métier.
        Nettoie aussi les sources (silence, journal) en passant."""
        to_remove=[]
        for sid,s in situations.items():
            if self._is_resolved_by_invariant(s):
                self._cleanup_resolved_situation(s)
                to_remove.append(sid)
        for sid in to_remove: del situations[sid]

    def _is_resolved_by_invariant(self,s):
        """Détermine si une situation est résolue par l'état actuel des données.
        Permet d'éviter d'afficher des alertes que l'utilisateur a en réalité déjà traitées.
        
        ATTENTION — on garde UNIQUEMENT des invariants qui s'appuient sur des sources
        de vérité STABLES (fichiers persistants). Pas d'invariant basé sur l'ABSENCE
        d'une alerte dans `ar` (le calcul peut temporairement la louper sans que la
        situation soit réellement résolue → on supprimerait à tort).
        
        Si tu veux clôturer une alerte qui n'est plus pertinente, clique "✓ Résolu"
        manuellement sur sa carte.
        """
        cat=s["category"]
        if cat=="livraison_jour":
            # Résolue dès que la livraison du jour est saisie dans livraisons.cfg.
            # Source de vérité fiable : on écrit dans livraisons.cfg uniquement après
            # une saisie utilisateur explicite (popup ou écran Livraisons).
            try:
                livrs=load_json(LIVRAISON_FILE) or {}
                today_key=date.today().strftime("%d/%m/%y")
                return today_key in livrs
            except Exception as _e: _log_silent_err(exc=_e); return False
        return False

    def _cleanup_resolved_situation(self,s):
        """Nettoie les traces d'une situation résolue par invariant : lève le silence
        et marque le journal résolu. Évite que la situation re-apparaisse au prochain refresh."""
        cat=s["category"]
        if s.get("is_silenced"):
            try: clear_popup_silence(cat)
            except Exception as _e: _log_silent_err(exc=_e)
        try:
            ids_to_resolve={e["id"] for e in s.get("journal_entries",[])
                            if (e.get("data",{}) or {}).get("statut")=="non_traite"}
            if ids_to_resolve:
                all_evt=load_json(EVENEMENTS_FILE) or {}
                events=all_evt.get("events",[])
                for i,evt in enumerate(events):
                    if evt.get("id") in ids_to_resolve:
                        d=evt.get("data",{}) or {}
                        d["statut"]="resolu";d["lu"]=True
                        evt["data"]=d
                        ts=datetime.now().strftime("%d/%m/%Y %Hh%M")
                        cur=evt.get("commentaire","") or ""
                        evt["commentaire"]=(cur+f" \u2014 R\u00e9solu automatiquement le {ts}").strip()
                        events[i]=evt
                all_evt["events"]=events
                save_json(EVENEMENTS_FILE,all_evt)
        except Exception as _e: _log_silent_err(exc=_e)

    def _upsert(self,situations,sid,**kwargs):
        """Crée ou met à jour une situation. Ne remplace pas les champs déjà remplis."""
        if sid not in situations:
            situations[sid]={
                "sit_id":sid,"category":"?","label":"?","sub_details":"",
                "is_active":False,"is_silenced":False,"silence_until":None,
                "journal_entries":[],"nb_apparitions":0,"has_non_traite":False,
                "last_non_traite_ts":None,"score":0,"raw":None,
            }
        for k,v in kwargs.items():
            if v is not None:
                if k in ("is_active","is_silenced","has_non_traite") and v:
                    situations[sid][k]=True
                elif situations[sid].get(k) in (None,"",0,False,"?"):
                    situations[sid][k]=v

    def _compute_score(self,s):
        """Score 0-2 selon historique (escalade visuelle)."""
        nb=s.get("nb_apparitions",0)
        has_nt=s.get("has_non_traite",False)
        nt_ts=s.get("last_non_traite_ts")
        nt_age_h=0
        if nt_ts:
            try: nt_age_h=(datetime.now()-nt_ts).total_seconds()/3600
            except Exception: pass
        if nb>=4 or (has_nt and nt_age_h>24): return 2
        if nb>=2 or has_nt: return 1
        return 0

    # =========================================================================
    # MAPPING SIT_ID (un sit_id stable par situation, peu importe la source)
    # =========================================================================
    def _sit_id_livr_report(self,item):
        d=item.get("date")
        d_str=d.strftime("%Y-%m-%d") if hasattr(d,"strftime") else str(d)
        return f"livr_report:{item.get('carburant','')}:{d_str}"
    def _sit_id_saisies(self,item):
        d=item.get("date")
        d_str=d.strftime("%Y-%m-%d") if hasattr(d,"strftime") else str(d)
        return f"saisies_irr:{item.get('carburant','')}:{d_str}"
    def _sit_id_rupture_imminente(self,item):
        # Ruptures imminentes sont du jour
        return f"rupture_imminente:{item.get('carburant','')}:{date.today().strftime('%Y-%m-%d')}"
    def _find_equivalent_active(self,situations,ptype,fp):
        """Cherche une situation ACTIVE équivalente à un silence donné.

        Le sid est construit comme '{cat}:{carburant}:{date_cible}' — il est STABLE
        même quand les chiffres changent (seul le fingerprint, qui encode les
        valeurs, change). Donc une alerte snoozée puis recalculée avec d'autres
        chiffres a le MÊME sid. On retrouve ainsi la version fraîche pour lui
        transférer l'état de pause, au lieu d'afficher la coquille figée.

        Retourne la clé (sid) de la situation active équivalente, ou None.
        """
        try:
            sid=self._sit_id_from_silence(ptype,fp)
            if sid and sid in situations:
                # Match direct par sid stable (cas le plus fréquent).
                return sid
            # Fallback : même catégorie + même date cible (dernier segment du sid),
            # au cas où léger remaniement d'id. On compare catégorie+carburant+date.
            if not sid: return None
            seg=sid.split(":")
            if len(seg)>=3:
                cat_s,carb_s,date_s=seg[0],seg[1],seg[2]
                for k in situations:
                    ks=k.split(":")
                    if len(ks)>=3 and ks[0]==cat_s and ks[1]==carb_s and ks[2]==date_s:
                        return k
        except Exception as _e:
            _log_silent_err(exc=_e)
        return None

    def _sit_id_from_silence(self,ptype,fp):
        """Mapping silence -> sit_id (clé stable, peu importe la source)."""
        if ptype=="livr_report":
            parts=fp.rsplit("_",1)
            if len(parts)==2: return f"livr_report:{parts[1]}:{parts[0]}"
        elif ptype=="saisies_irr":
            parts=fp.rsplit("_",1)
            if len(parts)==2: return f"saisies_irr:{parts[1]}:{parts[0]}"
        elif ptype=="rupture_imminente":
            parts=fp.rsplit("_",1)
            if len(parts)==2: return f"rupture_imminente:{parts[1]}:{parts[0]}"
        elif ptype=="antirupture":
            return f"antirupture:{fp}"
        elif ptype=="livraison_jour":
            return f"livraison_jour:{fp}"
        elif ptype=="tendance":
            parts=fp.rsplit("_",1)
            if len(parts)==2: return f"tendance:{parts[1]}:{parts[0]}"
        return None
    def _sit_id_from_evt(self,evt):
        """Mapping événement journal -> sit_id."""
        t=evt.get("type","");d=evt.get("data",{}) or {}
        if t=="livraison_reporter":
            try:
                date_str=(d.get("date","") or "")[:10]
                return f"livr_report:{d.get('carburant','')}:{date_str}"
            except Exception: return None
        elif t=="pont":
            try:
                dd=d.get("date_debut","")
                if dd:
                    dt=datetime.fromisoformat(dd).date()
                    return f"antirupture:pont_{dt.strftime('%d%m%Y')}"
            except Exception: return None
        elif t=="rupture":
            try:
                j=(d.get("jour","") or "")[:10]
                return f"rupture_imminente:{d.get('carburant','')}:{j}"
            except Exception: return None
        elif t=="anomalie":
            try:
                j=(d.get("jour","") or "")[:10]
                return f"tendance:{d.get('carburant','')}:{j}"
            except Exception: return None
        elif t=="tendance":
            try:
                j=(d.get("date","") or d.get("jour","") or "")[:10]
                return f"tendance:{d.get('carburant','')}:{j}"
            except Exception: return None
        elif t=="livraison_attendue":
            try:
                j=(d.get("date","") or d.get("jour","") or "")[:10]
                return f"livr_attendue:{j}"
            except Exception: return None
        return None
    def _cat_from_evt_type(self,t):
        return {"livraison_reporter":"livr_report","pont":"antirupture",
                "rupture":"rupture_imminente","anomalie":"tendance","tendance":"tendance",
                "livraison_attendue":"livr_attendue"}.get(t,t)

    # =========================================================================
    # LABELS HUMAINS
    # =========================================================================
    def _label_livr_report(self,it):
        d=it.get("date")
        if hasattr(d,"strftime"):
            JC=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
            d_str=f"{JC[d.weekday()]} {d.strftime('%d/%m')}"
        else: d_str=str(d)
        return f"{it.get('carburant','?')} \u2014 livraison {d_str}"
    def _details_livr_report(self,it):
        parts=[]
        surplus=it.get("surplus",0)
        if surplus: parts.append(f"d\u00e9passement {int(surplus)} L")
        if it.get("report_au_lendemain"):
            parts.append("\u00e0 reporter au lendemain (C1 insuffisante)")
        else:
            heure=it.get("heure_recommandee")
            if heure is not None:
                try:
                    h=float(heure)
                    h_h=int(h);h_m=int((h-h_h)*60)
                    parts.append(f"heure recommand\u00e9e : {h_h:02d}h{h_m:02d}")
                except Exception as _e: _log_silent_err(exc=_e)
        return " \u00b7 ".join(parts)
    def _label_saisies(self,it):
        d=it.get("date")
        if hasattr(d,"strftime"):
            JC=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
            d_str=f"{JC[d.weekday()]} {d.strftime('%d/%m')}"
        else: d_str=str(d)
        return f"Saisie {it.get('carburant','?')} \u2014 {d_str}"
    def _details_saisies(self,it):
        v=it.get("ventes",0);c=it.get("capacite_max",0)
        if v and c: return f"ventes saisies {int(v)} L vs capacit\u00e9 max {int(c)} L"
        return "valeur impossible \u00e0 d\u00e9biter physiquement"
    def _label_rupture_imm(self,it):
        return f"Rupture imminente {it.get('carburant','?')}"
    def _details_rupture_imm(self,it):
        a=it.get("autonomie_h",0)
        if a: return f"autonomie restante : {a:.1f}h"
        return ""
    def _label_pont(self,pid):
        # pid = "pont_DDMMYYYY" → "Weekend du …" ou "Pont du …" selon présence d'un férié
        if pid.startswith("pont_") and len(pid)>=13:
            try:
                dd,mm,yyyy=int(pid[5:7]),int(pid[7:9]),int(pid[9:13])
                d_obj=date(yyyy,mm,dd)
                JC=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
                # Reconstruit la durée du trou (jours non-livrables consécutifs) pour
                # qualifier : "Pont" si un férié est dedans, sinon "Weekend".
                cur=d_obj;contient_ferie=False;n=0
                while (cur.weekday()>=5 or is_ferie(cur)) and n<7:
                    if is_ferie(cur): contient_ferie=True
                    cur+=timedelta(days=1);n+=1
                terme="Pont" if contient_ferie else "Weekend"
                return f"{terme} du {JC[d_obj.weekday()]} {d_obj.strftime('%d/%m')}"
            except Exception: return f"Weekend/pont du {pid[5:7]}/{pid[7:9]}/{pid[9:13]}"
        return f"Weekend/pont {pid}"
    def _details_pont(self,st):
        return st.get("info","") if isinstance(st.get("info"),str) else ""
    def _label_from_silence(self,ptype,fp):
        sid=self._sit_id_from_silence(ptype,fp)
        if not sid: return f"{ptype} \u2014 {fp}"
        # Reconstruire un label à partir du sit_id
        if ptype=="livr_report":
            parts=fp.rsplit("_",1)
            if len(parts)==2:
                try:
                    d=datetime.fromisoformat(parts[0]).date()
                    return f"{parts[1]} \u2014 livraison le {d.strftime('%d/%m/%Y')}"
                except Exception: pass
        elif ptype=="saisies_irr":
            parts=fp.rsplit("_",1)
            if len(parts)==2:
                try:
                    d=datetime.fromisoformat(parts[0]).date()
                    return f"Saisie {parts[1]} \u2014 {d.strftime('%d/%m/%Y')}"
                except Exception: pass
        elif ptype=="rupture_imminente":
            parts=fp.rsplit("_",1)
            if len(parts)==2: return f"Rupture imminente {parts[1]}"
        elif ptype=="antirupture":
            return self._label_pont(fp)
        elif ptype=="livraison_jour":
            try:
                d=datetime.fromisoformat(fp).date()
                return f"Livraison du jour ({d.strftime('%d/%m/%Y')})"
            except Exception: return "Livraison du jour"
        elif ptype=="tendance":
            parts=fp.rsplit("_",1)
            if len(parts)==2: return f"Tendance ventes {parts[1]}"
        return f"{ptype} \u2014 {fp}"
    def _label_from_evt(self,evt):
        t=evt.get("type","");d=evt.get("data",{}) or {}
        ts=evt.get("ts","")
        try: ts_str=datetime.fromisoformat(ts).strftime("%d/%m %Hh%M")
        except Exception: ts_str=ts[:16]
        if t=="livraison_reporter":
            try:
                date_str=(d.get("date","") or "")[:10]
                date_disp=datetime.fromisoformat(date_str).strftime("%d/%m/%Y") if date_str else "?"
                return f"{d.get('carburant','?')} \u2014 livraison le {date_disp}"
            except Exception: return f"Livraison \u00e0 reporter \u2014 {ts_str}"
        if t=="tendance":
            carb=(d.get("carburant","?") or "?").upper()
            ec=d.get("ecart_pct",0)
            sg=d.get("stage","")
            signe="+" if ec>0 else ""
            jour=(d.get("date","") or d.get("jour","") or "")[:10]
            try: jour_disp=datetime.fromisoformat(jour).strftime("%d/%m") if jour else ""
            except Exception: jour_disp=jour
            base=f"{carb} {signe}{ec}% \u00e0 {sg}" if sg else f"{carb} {signe}{ec}%"
            return f"{base} \u2014 {jour_disp}" if jour_disp else base
        if t=="livraison_attendue":
            tl=d.get("tour_label","Tour")
            return f"{tl} d\u00e9pass\u00e9 \u2014 camion non confirm\u00e9 arriv\u00e9"
        return f"{t.capitalize()} \u2014 {ts_str}"

    # =========================================================================
    # RENDU
    # =========================================================================
    def _render(self):
        """Construit le tableau : groupe par catégorie métier, trie par score (urgent en haut)."""
        for w in self.body.winfo_children(): w.destroy()
        situations=self._collect_situations()
        # Filtrer : on n'affiche QUE les situations qui sont actuellement pertinentes
        situations=[s for s in situations
                    if s["is_active"] or s["is_silenced"] or s["has_non_traite"]]
        total=len(situations)
        if total==0:
            self.subtitle_lbl.configure(text="Aucune alerte active. Tout est sous contr\u00f4le.")
            ok=ctk.CTkFrame(self.body,fg_color=C["card"],corner_radius=10,border_width=1,border_color="#2A4A2A")
            ok.pack(fill="x",padx=4,pady=20)
            ctk.CTkLabel(ok,text="\u2713",font=("Segoe UI",36),text_color=C["green"]).pack(pady=(20,8))
            ctk.CTkLabel(ok,text="Aucune alerte \u00e0 traiter",font=("Segoe UI",14,"bold"),text_color=C["t1"]).pack()
            ctk.CTkLabel(ok,text="Tout est sous contr\u00f4le.",font=("Segoe UI",11),text_color=C["t3"]).pack(pady=(4,20))
            return
        self.subtitle_lbl.configure(text=f"{total} alerte{'s' if total>1 else ''} active{'s' if total>1 else ''}")
        # Regrouper par catégorie métier, dans un ordre fixe (criticité décroissante)
        cat_order=["rupture_imminente","livr_attendue","saisies_irr","incoh","rupt_proj",
                   "antirupture","livr_report","livraison_jour","tendance"]
        cat_meta={
            "rupture_imminente":("\u26a0","Rupture imminente",C["red"]),
            "livr_attendue":("\U0001f69a","Livraison attendue non arriv\u00e9e",C["red"]),
            "saisies_irr":("\U0001f6ab","Saisies physiquement impossibles",C["red"]),
            "incoh":("\u26a0","Livraisons sur jour non-livrable",C["red"]),
            "rupt_proj":("\U0001f4c9","Ruptures projet\u00e9es",C["red"]),
            "antirupture":("\U0001f3d7","Week-ends / ponts \u00e0 g\u00e9rer",C["amber"]),
            "livr_report":("\U0001f69b","Livraisons \u00e0 reporter",C["amber"]),
            "livraison_jour":("\U0001f69b","Livraison du jour",C["amber"]),
            "tendance":("\U0001f4ca","Tendance ventes",C["amber"]),
        }
        by_cat={}
        for s in situations: by_cat.setdefault(s["category"],[]).append(s)
        # Tri intra-catégorie : score décroissant, puis sit_id pour stabilité
        for cat in by_cat: by_cat[cat].sort(key=lambda x:(-x["score"],x["sit_id"]))
        for cat in cat_order:
            if cat not in by_cat: continue
            items=by_cat[cat]
            icon,titre,color=cat_meta.get(cat,("\u2022",cat,C["t2"]))
            # Header section : bandeau légèrement teinté, padding généreux, compteur en pastille
            hdr=ctk.CTkFrame(self.body,fg_color=C["panel"],corner_radius=6,height=40,
                             border_width=0)
            hdr.pack(fill="x",padx=4,pady=(16,4));hdr.pack_propagate(False)
            ctk.CTkLabel(hdr,text=icon,font=("Segoe UI Emoji",15),text_color=color,
                         width=32).pack(side="left",padx=(12,0))
            ctk.CTkLabel(hdr,text=titre,font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(side="left",padx=(4,0))
            # Compteur en pastille pleine (style "badge")
            badge=ctk.CTkFrame(hdr,fg_color=color,corner_radius=12,width=26,height=22)
            badge.pack(side="left",padx=(10,0));badge.pack_propagate(False)
            ctk.CTkLabel(badge,text=str(len(items)),font=("Segoe UI",10,"bold"),
                         text_color="#000" if color==C["amber"] else "#FFF").pack(expand=True)
            for s in items: self._render_situation(s)
        # Catégories non listées (sécurité : on n'oublie rien)
        for cat in by_cat:
            if cat in cat_order: continue
            for s in by_cat[cat]: self._render_situation(s)

    def _category_icon(self,cat):
        """Icône métier par catégorie de situation. Sert de repère visuel rapide."""
        return {
            "antirupture":"\U0001f309",       # 🌉 Pont
            "livr_report":"\U0001f69b",       # 🚛 Camion (livraison à reporter)
            "livraison_jour":"\U0001f4e6",    # 📦 Colis (livraison du jour)
            "saisies_irr":"\u26a0",            # ⚠ Triangle (saisies irrégulières)
            "rupture_imminente":"\u26fd",      # ⛽ Pompe (rupture imminente)
            "rupt_proj":"\u26fd",              # ⛽ Pompe (ruptures projetées)
            "incoh":"\u26a0",                  # ⚠ Triangle (incohérences)
            "tendance":"\U0001f4c8",          # 📈 Tendance
            "ferie_isole":"\U0001f5d3",       # 🗓 Calendrier (férié isolé)
            "marge_tendue":"\u26a0",           # ⚠ Triangle (marge tendue)
        }.get(cat,"\u25cf")  # ● par défaut

    def _render_situation(self,s):
        """Carte notification inspirée Linear/Sentry : pastille statut circulaire + titre fort
        + détails + méta + hover effect + actions à droite. Hauteur adaptée au contenu."""
        score=s["score"]
        cat=s.get("category","")
        # Système couleurs sévérité
        if score==2:
            card_bg=C["card"];card_hover="#2A1A1D";border=C["red"];border_w=2
            pastille_fg=C["red"];pastille_text="#FFF"
            title_color="#FFF";chip_color="#FF8B8B";icon_color="#FFF"
            sev_color=C["red"]
        elif score==1:
            card_bg=C["card"];card_hover="#221A12";border=C["amber"];border_w=2
            pastille_fg=C["amber"];pastille_text="#000"
            title_color=C["t1"];chip_color=C["amber"];icon_color="#000"
            sev_color=C["amber"]
        else:
            card_bg=C["card"];card_hover=C["card_h"];border=C["border2"];border_w=1
            pastille_fg=C["border2"];pastille_text=C["t1"]
            title_color=C["t1"];chip_color=C["t2"];icon_color=C["t1"]
            sev_color=C["t2"]
        # Carte
        card=ctk.CTkFrame(self.body,fg_color=card_bg,corner_radius=10,
                          border_width=border_w,border_color=border)
        card.pack(fill="x",padx=4,pady=4)
        # Contenu interne avec padding généreux
        inner=ctk.CTkFrame(card,fg_color="transparent")
        inner.pack(fill="x",padx=14,pady=12)
        # === Pastille statut circulaire à gauche (cercle 32px avec icône au centre) ===
        # Donne un repère visuel fort de sévérité avant même de lire le titre.
        pastille=ctk.CTkFrame(inner,fg_color=pastille_fg,width=34,height=34,corner_radius=17)
        pastille.pack(side="left",padx=(0,14))
        pastille.pack_propagate(False)
        icon=self._category_icon(cat)
        ctk.CTkLabel(pastille,text=icon,font=("Segoe UI Emoji",15),text_color=icon_color).pack(expand=True)
        # === Bouton Résolu à droite (empilé en 1er pour place fixe) ===
        if score==2:
            btn_fg=C["green"];btn_hover="#1F7C36";btn_text="#FFF";btn_border=C["green"]
        else:
            btn_fg="transparent";btn_hover=C["panel"];btn_text=C["green"];btn_border=C["green"]
        btn_resolu=ctk.CTkButton(inner,text="\u2713 R\u00e9solu",width=100,height=34,
            fg_color=btn_fg,hover_color=btn_hover,text_color=btn_text,
            border_width=1,border_color=btn_border,font=("Segoe UI",10,"bold"),
            corner_radius=8,command=lambda sit=s: self._mark_resolved(sit))
        btn_resolu.pack(side="right",padx=(12,0))
        # === Bloc texte central : titre + détails + méta ===
        col_text=ctk.CTkFrame(inner,fg_color="transparent")
        col_text.pack(side="left",fill="both",expand=True)
        # Titre 13 pt bold (plus gros qu'avant pour hiérarchie claire)
        prefix="\u203c " if score==2 else ""
        ctk.CTkLabel(col_text,text=f"{prefix}{s['label']}",font=("Segoe UI",13,"bold"),
                     text_color=title_color,anchor="w",wraplength=580,justify="left").pack(anchor="w")
        # Détails secondaires (10pt en t2)
        if s.get("sub_details"):
            ctk.CTkLabel(col_text,text=s['sub_details'],font=("Segoe UI",10),
                         text_color=C["t2"],anchor="w",wraplength=580,justify="left").pack(anchor="w",pady=(4,0))
        # Méta-chips ligne 3 : silence + apparitions + non_traite
        chips=[]
        if s.get("is_silenced") and s.get("silence_until"):
            try:
                u=s["silence_until"]
                JC_FR=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
                if u.date()==date.today(): chips.append(f"\u23f1 rappel \u00e0 {u.strftime('%Hh%M')}")
                else: chips.append(f"\u23f1 rappel {JC_FR[u.weekday()]} {u.strftime('%d/%m')} \u00e0 {u.strftime('%Hh%M')}")
            except Exception as _e: _log_silent_err(exc=_e)
        nb=s.get("nb_apparitions",0)
        if nb>=2: chips.append(f"\U0001f441 vu {nb} fois")
        if s.get("has_non_traite"):
            nt_ts=s.get("last_non_traite_ts")
            if nt_ts:
                try:
                    age_h=(datetime.now()-nt_ts).total_seconds()/3600
                    if age_h<1: chips.append(f"\u2757 non trait\u00e9e")
                    elif age_h<24: chips.append(f"\u2757 non trait\u00e9e depuis {int(age_h)}h")
                    else: chips.append(f"\u2757 non trait\u00e9e depuis {int(age_h/24)}j")
                except Exception: chips.append("\u2757 non trait\u00e9e")
            else: chips.append("\u2757 non trait\u00e9e")
        if chips:
            ctk.CTkLabel(col_text,text="  \u00b7  ".join(chips),font=("Segoe UI",9),
                         text_color=chip_color,anchor="w").pack(anchor="w",pady=(6,0))
        # === Hover effect : éclaircit le fond quand on survole ===
        def _on_enter(e,c=card,h=card_hover):
            try: c.configure(fg_color=h)
            except Exception: pass
        def _on_leave(e,c=card,bg=card_bg):
            try: c.configure(fg_color=bg)
            except Exception: pass
        def _on_click(e,sit=s): self._open_situation(sit)
        # Bind sur tout sauf le bouton Résolu (qui a son propre clic)
        clickables=[card,inner,col_text,pastille]+list(col_text.winfo_children())
        for w in clickables:
            try:
                w.bind("<Button-1>",_on_click)
                w.bind("<Enter>",_on_enter)
                w.bind("<Leave>",_on_leave)
                w.configure(cursor="hand2")
            except Exception as _e: _log_silent_err(exc=_e)

    # =========================================================================
    # ACTIONS
    # =========================================================================
    def _open_situation(self,s):
        """Clic sur une carte : rouvre la popup d'origine selon la catégorie.
        
        Règle importante : on n'ouvre la popup QUE si la situation est encore active dans
        `ar` (donnée fraîche). On NE reconstruit JAMAIS la popup à partir des chiffres figés
        dans le journal car ces chiffres peuvent être obsolètes (ex: Pre_vision a été corrigé
        entretemps). Une popup avec des chiffres périmés est trompeuse — pire qu'un message
        clair "n'est plus d'actualité, à clôturer".
        """
        cat=s["category"]
        ar=(self.parent_app.last_data or {}).get("antirupture",{}) or {}
        try:
            if cat=="livr_report":
                target=s.get("raw")  # raw n'est posé QUE si la situation est dans `ar` actuel
                if target:
                    dlg=LivraisonsAReporterDlg(self.parent_app,[target])
                    self.wait_window(dlg);self._render()
                else:
                    # Alerte "fantôme" : journal en non_traite mais plus dans ar courant.
                    # On propose directement la clôture (UX : 1 clic au lieu de 3).
                    self._show_resolve_dialog(s)
            elif cat=="saisies_irr":
                target=s.get("raw")
                if target:
                    dlg=SaisiesIrrealistesDlg(self.parent_app,[target])
                    self.wait_window(dlg);self._render()
                else: self._show_resolve_dialog(s)
            elif cat=="rupture_imminente":
                target=s.get("raw")
                if target:
                    dlg=RuptureImminenteDlg(self.parent_app,[target])
                    self.wait_window(dlg);self._render()
                else: self._show_resolve_dialog(s)
            elif cat=="antirupture":
                if ar:
                    dlg=AntiRuptureDlg(self.parent_app,ar)
                    self.wait_window(dlg);self._render()
                else: self._show_resolve_dialog(s)
            elif cat=="tendance":
                # Une tendance de ventes n'a pas de popup d'origine (elle vit dans la
                # vignette de droite, pas dans une fenêtre dédiée). Le clic propose donc
                # directement la clôture, au lieu d'ouvrir la popup anti-rupture (vide).
                self._show_resolve_dialog(s)
            elif cat=="livraison_jour":
                dlg=LivraisonDialog(self.parent_app)
                self.wait_window(dlg);self._render()
            else:
                self._show_msg(f"Cat\u00e9gorie non g\u00e9r\u00e9e pour clic direct : {cat}")
        except Exception as e: print(f"[dashboard open {cat}] {e}")

    def _build_livr_report_item_from_evt(self,d):
        """Reconstruit un item compatible avec LivraisonsAReporterDlg depuis data journal."""
        try: livr_d=datetime.fromisoformat(d.get("date","")).date()
        except Exception: livr_d=date.today()
        JC=["lun.","mar.","mer.","jeu.","ven.","sam.","dim."]
        return {
            "date":livr_d,
            "date_str":f"{JC[livr_d.weekday()]} {livr_d.strftime('%d/%m')}",
            "carburant":d.get("carburant","?"),
            "stock_matin":d.get("stock_matin",0),
            "livraison":d.get("livraison",0),
            "capacite":40000 if d.get("carburant","").upper() in ("SP","GO") else 10000,
            "surplus":d.get("surplus",0),
            "ventes_c1_moy":0,
            "heures_attente":d.get("heures_attente"),
            "heure_recommandee":d.get("heure_recommandee"),
        }

    def _mark_resolved(self,s):
        """Bouton ✓ Résolu : lève le silence ET marque les événements journal liés résolus."""
        cat=s["category"]
        # 1. Lever le silence si actif
        if s.get("is_silenced"):
            try: clear_popup_silence(cat)
            except Exception as _e: _log_silent_err(exc=_e)
        # 2. Marquer non_traite -> resolu dans le journal
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[])
            ids_to_resolve={e["id"] for e in s.get("journal_entries",[]) 
                            if (e.get("data",{}) or {}).get("statut")=="non_traite"}
            if ids_to_resolve:
                for i,evt in enumerate(events):
                    if evt.get("id") in ids_to_resolve:
                        d=evt.get("data",{}) or {}
                        d["statut"]="resolu";d["lu"]=True
                        evt["data"]=d
                        ts=datetime.now().strftime("%d/%m/%Y %Hh%M")
                        cur=evt.get("commentaire","") or ""
                        evt["commentaire"]=(cur+f" \u2014 R\u00e9solu via tableau le {ts}").strip()
                        events[i]=evt
                all_evt["events"]=events
                save_json(EVENEMENTS_FILE,all_evt)
        except Exception as _e: _log_silent_err(exc=_e)
        # 3. Re-render la liste (instantané). On NE relance PAS le tour de calcul complet
        #    du HUB ici : marquer une notification résolue ne change pas l'anti-rupture
        #    (qui dépend des stocks/prévisions). Le refresh complet coûteux est différé à
        #    la fermeture du tableau (_close) et fait UNE seule fois. Avant ce fix, chaque
        #    clôture relançait tout le tour → lenteur insupportable sur plusieurs alertes.
        self._needs_parent_refresh=True
        self._render()

    def _resolve_all(self):
        """Bouton 'Tout effacer' : clôture toutes les notifications affichées en UNE
        seule passe (une seule écriture journal, un seul re-render), avec confirmation.
        Évite de cliquer 'Résolu' carte par carte."""
        situations=self._collect_situations()
        situations=[s for s in situations
                    if s["is_active"] or s["is_silenced"] or s["has_non_traite"]]
        n=len(situations)
        if n==0:
            self._show_msg("Aucune notification \u00e0 effacer.")
            return
        # Confirmation
        m=ctk.CTkToplevel(self);m.title("Tout effacer")
        m.geometry("440x180");m.configure(fg_color=C["bg"]);m.transient(self)
        m.after(0,lambda:(m.winfo_exists() and m.grab_set()))
        ctk.CTkLabel(m,text=f"Cl\u00f4turer les {n} notification{'s' if n>1 else ''} affich\u00e9e{'s' if n>1 else ''} ?",
                     font=("Segoe UI",12,"bold"),text_color=C["t1"],wraplength=400,justify="center").pack(pady=(26,6))
        ctk.CTkLabel(m,text="Elles restent consultables dans le Journal des \u00e9v\u00e9nements.",
                     font=("Segoe UI",10),text_color=C["t3"],wraplength=400,justify="center").pack(pady=(0,16))
        row=ctk.CTkFrame(m,fg_color="transparent");row.pack()
        ctk.CTkButton(row,text="Annuler",width=110,height=36,fg_color=C["panel"],
                      text_color=C["t1"],border_width=1,border_color=C["border2"],
                      corner_radius=8,command=m.destroy).pack(side="left",padx=(0,10))
        def _do():
            m.destroy()
            self._do_resolve_all(situations)
        ctk.CTkButton(row,text="\U0001f5d1 Tout effacer",width=140,height=36,fg_color=C["red"],
                      hover_color="#A01818",text_color="#FFF",corner_radius=8,command=_do).pack(side="left")

    def _do_resolve_all(self,situations):
        """Exécute la clôture de masse : lève les silences + marque tous les non_traite
        résolus en une seule écriture, puis un seul re-render."""
        # 1. Lever les silences des catégories concernées
        for cat in {s["category"] for s in situations if s.get("is_silenced")}:
            try: clear_popup_silence(cat)
            except Exception as _e: _log_silent_err(exc=_e)
        # 2. Marquer tous les non_traite -> resolu en UNE passe
        try:
            all_evt=load_json(EVENEMENTS_FILE) or {}
            events=all_evt.get("events",[])
            ids_to_resolve=set()
            for s in situations:
                for e in s.get("journal_entries",[]):
                    if (e.get("data",{}) or {}).get("statut")=="non_traite":
                        ids_to_resolve.add(e.get("id"))
            if ids_to_resolve:
                ts=datetime.now().strftime("%d/%m/%Y %Hh%M")
                for i,evt in enumerate(events):
                    if evt.get("id") in ids_to_resolve:
                        d=evt.get("data",{}) or {}
                        d["statut"]="resolu";d["lu"]=True
                        evt["data"]=d
                        cur=evt.get("commentaire","") or ""
                        evt["commentaire"]=(cur+f" \u2014 R\u00e9solu via tableau le {ts}").strip()
                        events[i]=evt
                all_evt["events"]=events
                save_json(EVENEMENTS_FILE,all_evt)
        except Exception as _e: _log_silent_err(exc=_e)
        self._needs_parent_refresh=True
        self._render()

    def _show_msg(self,msg):
        """Mini-message info temporaire au-dessus du dashboard."""
        m=ctk.CTkToplevel(self);m.title("Info");m.geometry("420x150")
        m.configure(fg_color=C["bg"]);m.transient(self);m.grab_set()
        ctk.CTkLabel(m,text=msg,font=("Segoe UI",11),text_color=C["t1"],wraplength=380,justify="center").pack(pady=(28,16))
        ctk.CTkButton(m,text="OK",width=100,fg_color=C["panel"],text_color=C["t1"],
                      border_width=1,border_color=C["border2"],command=m.destroy).pack()

    def _show_resolve_dialog(self,situation):
        """Mini-popup actionnable pour les alertes 'fantômes' (raw=None) :
        l'événement existe dans le journal en non_traite, mais ne correspond plus
        à une situation active dans le calcul anti-rupture (probablement résolue
        de fait : Pre_vision corrigé, livraison saisie, etc.).
        
        Au lieu d'afficher juste "n'est plus d'actualité" + OK (frustrant : 
        l'utilisateur doit ensuite chercher comment clôturer), on propose
        DIRECTEMENT le bouton de clôture ici. UX : un clic au lieu de trois.
        """
        m=ctk.CTkToplevel(self);m.title("Alerte sans \u00e9quivalent actif")
        m.geometry("520x230");m.configure(fg_color=C["bg"]);m.transient(self);m.grab_set()
        # Message explicatif
        ctk.CTkLabel(m,text="\u2139 Cette alerte ne correspond plus \u00e0 une situation active",
                     font=("Segoe UI",12,"bold"),text_color=C["gold"]).pack(pady=(20,4))
        ctk.CTkLabel(m,
                     text="Elle vient du journal en \u00e9tat \u00ab non trait\u00e9 \u00bb, mais le calcul anti-rupture\n"
                          "actuel ne d\u00e9tecte plus cette tension (Pre_vision a probablement \u00e9t\u00e9\n"
                          "corrig\u00e9 ou la livraison a \u00e9t\u00e9 saisie depuis).\n\n"
                          "Tu peux la cl\u00f4turer ici si elle est obsol\u00e8te.",
                     font=("Segoe UI",10),text_color=C["t2"],wraplength=460,justify="center").pack(pady=(0,16))
        # Footer : 2 boutons (Garder + Clôturer) — pas de label gauche (largeur OK : 480px sur 520-40=480 dispo)
        footer=ctk.CTkFrame(m,fg_color="transparent",height=50)
        footer.pack(side="bottom",fill="x",padx=20,pady=(0,16));footer.pack_propagate(False)
        ctk.CTkButton(footer,text="\u2713 Cl\u00f4turer maintenant",width=200,height=38,
                      fg_color=C["green"],hover_color="#15943C",text_color="#FFF",
                      font=("Segoe UI",11,"bold"),corner_radius=8,
                      command=lambda:self._resolve_and_close(situation,m)).pack(side="right",padx=(0,8))
        ctk.CTkButton(footer,text="Garder en suspens",width=160,height=38,
                      fg_color=C["panel"],hover_color=C["card_h"],text_color=C["t1"],
                      border_width=1,border_color=C["border2"],
                      font=("Segoe UI",11),corner_radius=8,
                      command=m.destroy).pack(side="right",padx=(0,8))

    def _resolve_and_close(self,situation,dialog):
        """Clôture la situation et ferme le dialogue."""
        try: self._mark_resolved(situation)
        except Exception as _e: _log_silent_err(exc=_e)
        try: dialog.destroy()
        except Exception: pass
        try: self._render()
        except Exception: pass



# =============================================================================
class SettingsDlg(ctk.CTkToplevel):
    def __init__(self,parent,cfg):
        super().__init__(parent);self.title("Chemins \u2014 DISTRICARB HUB");self.geometry("820x640");self.minsize(700,480)
        self.configure(fg_color=C["bg"]);self.resizable(True,True);self.transient(parent);self.grab_set()
        self.result=None;self.cfg=dict(cfg);self.entries={}
        ctk.CTkLabel(self,text="Configuration des chemins",font=("Segoe UI",18,"bold"),text_color=C["t1"]).pack(anchor="w",padx=24,pady=(24,4))
        ctk.CTkLabel(self,text="Fichiers lus en lecture seule.",font=("Segoe UI",11),text_color=C["t3"]).pack(anchor="w",padx=24,pady=(0,12))
        btns=ctk.CTkFrame(self,fg_color="transparent",height=56);btns.pack(side="bottom",fill="x",padx=24,pady=14);btns.pack_propagate(False)
        ctk.CTkButton(btns,text="\u2713  Enregistrer",width=160,height=40,fg_color=C["red"],hover_color="#C41E24",text_color="#FFF",font=("Segoe UI",13,"bold"),corner_radius=8,command=self._save).pack(side="right")
        ctk.CTkButton(btns,text="Annuler",width=110,height=40,fg_color=C["card"],hover_color=C["card_h"],border_width=1,border_color=C["border2"],text_color=C["t1"],corner_radius=8,command=self.destroy).pack(side="right",padx=(0,10))
        ctk.CTkFrame(self,fg_color=C["border"],height=1).pack(side="bottom",fill="x",padx=24)
        scroll=ctk.CTkScrollableFrame(self,fg_color="transparent",scrollbar_fg_color=C["panel"],scrollbar_button_color=C["border2"])
        scroll.pack(side="top",fill="both",expand=True,padx=24,pady=(0,8))
        for fd in HUB_FILES: self._row(scroll,fd)
        # Section Cycle 14 jours
        self._cycle_section(scroll)
        # Section Marges
        self._marges_section(scroll)
    def _cycle_section(self,parent):
        """Affiche et permet de modifier l'ancrage du cycle 14 jours.
        Prévision compte tourne sur 2 semaines : semaine 1 (Lundi/Mardi/.../Dimanche)
        et semaine 2 (Lundi2/Mardi2/.../Dim2). L'ancrage indique quel lundi a démarré
        en semaine 1, et le hub calcule les semaines suivantes par rotation 14 jours."""
        cyc=load_json(CYCLE_FILE) or {}
        anchor_date=cyc.get("anchor_date","(non d\u00e9fini)")
        anchor_week=cyc.get("anchor_week","(non d\u00e9fini)")
        # Calcul de la semaine actuelle
        sem_actuelle="(inconnu)"
        if cyc.get("anchor_date") and cyc.get("anchor_week"):
            try:
                ad=datetime.strptime(anchor_date,"%Y-%m-%d").date()
                today=date.today()
                # Aller au lundi de cette semaine
                lundi=today-timedelta(days=today.weekday())
                lundi_anchor=ad-timedelta(days=ad.weekday())
                diff_jours=(lundi-lundi_anchor).days
                nb_sem=diff_jours//7
                if anchor_week==1:
                    sem_actuelle="Semaine 1" if nb_sem%2==0 else "Semaine 2"
                else:
                    sem_actuelle="Semaine 2" if nb_sem%2==0 else "Semaine 1"
            except Exception as _e: _log_silent_err(exc=_e)
        # Frame
        sec=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=8,border_width=1,border_color=C["border"])
        sec.pack(fill="x",pady=(16,4))
        top=ctk.CTkFrame(sec,fg_color="transparent");top.pack(fill="x",padx=14,pady=(10,4))
        ctk.CTkLabel(top,text="\U0001f4c5",font=("Segoe UI Emoji",16),text_color=C["gold"]).pack(side="left",padx=(0,6))
        ctk.CTkLabel(top,text="CYCLE 14 JOURS (Prévision compte)",font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(side="left")
        # Affichage état actuel
        info=ctk.CTkFrame(sec,fg_color="transparent");info.pack(fill="x",padx=14,pady=(4,4))
        ctk.CTkLabel(info,text=f"Ancrage : lundi {anchor_date} = Semaine {anchor_week}",
                     font=("Segoe UI",11),text_color=C["t2"]).pack(anchor="w")
        ctk.CTkLabel(info,text=f"Cette semaine = {sem_actuelle}",
                     font=("Segoe UI",11,"bold"),
                     text_color=C["green"] if "Semaine" in str(sem_actuelle) else C["amber"]).pack(anchor="w",pady=(2,0))
        # Aide
        ctk.CTkLabel(sec,
                     text="Si le hub te parle de mauvais onglets dans Pre_vision (ex : Jeudi 2 au lieu\n"
                          "de Jeudi), c'est que l'ancrage est faux. Corrige-le ci-dessous :",
                     font=("Segoe UI",10),text_color=C["t3"],justify="left").pack(anchor="w",padx=14,pady=(4,4))
        # Boutons d'ajustement rapide
        btns=ctk.CTkFrame(sec,fg_color="transparent");btns.pack(fill="x",padx=14,pady=(2,12))
        ctk.CTkButton(btns,text="Cette semaine = Semaine 1",width=190,height=32,
                       fg_color=C["green"],hover_color="#15943C",text_color="#FFF",
                       font=("Segoe UI",10,"bold"),corner_radius=6,
                       command=lambda:self._set_cycle_week(1)).pack(side="left",padx=(0,8))
        ctk.CTkButton(btns,text="Cette semaine = Semaine 2",width=190,height=32,
                       fg_color=C["gold"],hover_color="#A87E2C",text_color="#FFF",
                       font=("Segoe UI",10,"bold"),corner_radius=6,
                       command=lambda:self._set_cycle_week(2)).pack(side="left")
    def _marges_section(self,parent):
        """Section Paramètres : taux de marge boutique mensuel.
        S'applique au mois en cours dans prix_historique.cfg. Pour les rapports historiques,
        on garde le taux mémorisé pour ce mois-là."""
        today=date.today()
        cur_month_key=f"{today.year:04d}-{today.month:02d}"
        cur_data=get_prix_for_month(today.year,today.month) or {}
        cur_taux=sf(cur_data.get("marge_boutique_taux",0.30))
        # Frame
        sec=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=8,border_width=1,border_color=C["border"])
        sec.pack(fill="x",pady=(16,4))
        top=ctk.CTkFrame(sec,fg_color="transparent");top.pack(fill="x",padx=14,pady=(10,4))
        ctk.CTkLabel(top,text="\U0001f4b0",font=("Segoe UI Emoji",16),text_color=C["green"]).pack(side="left",padx=(0,6))
        ctk.CTkLabel(top,text="MARGES",font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(side="left")
        # Taux marge boutique
        info=ctk.CTkFrame(sec,fg_color="transparent");info.pack(fill="x",padx=14,pady=(4,4))
        ctk.CTkLabel(info,text=f"Taux de marge boutique pour {cur_month_key}",
                     font=("Segoe UI",11),text_color=C["t2"]).pack(anchor="w")
        ctk.CTkLabel(info,text="Pourcentage du CA boutique qui constitue la marge brute. R\u00e9el 2025 : 30,27%.",
                     font=("Segoe UI",9),text_color=C["t3"]).pack(anchor="w",pady=(2,0))
        # Champ saisie + bouton
        edit=ctk.CTkFrame(sec,fg_color="transparent");edit.pack(fill="x",padx=14,pady=(8,12))
        ctk.CTkLabel(edit,text="Taux :",font=("Segoe UI",11),text_color=C["t1"]).pack(side="left",padx=(0,6))
        self.bout_taux_entry=ctk.CTkEntry(edit,width=80,height=30,fg_color=C["bg"],border_color=C["border"],
                                            text_color=C["t1"],font=("Segoe UI",11),placeholder_text="30.0")
        self.bout_taux_entry.pack(side="left",padx=(0,4))
        self.bout_taux_entry.insert(0,f"{cur_taux*100:.2f}")
        ctk.CTkLabel(edit,text="%",font=("Segoe UI",11),text_color=C["t2"]).pack(side="left",padx=(0,12))
        ctk.CTkButton(edit,text="Enregistrer",width=120,height=30,
                       fg_color=C["green"],hover_color="#15943C",text_color="#FFF",
                       font=("Segoe UI",10,"bold"),corner_radius=6,
                       command=self._save_marge_boutique).pack(side="left")
    def _save_marge_boutique(self):
        """Enregistre le taux de marge boutique pour le mois en cours dans prix_historique.cfg."""
        try:
            txt=self.bout_taux_entry.get().strip().replace(",",".")
            taux=float(txt)/100.0
            if taux<=0 or taux>1:
                messagebox.showerror("Taux invalide","Le taux doit \u00eatre entre 0 et 100%.",parent=self)
                return
            today=date.today()
            key=f"{today.year:04d}-{today.month:02d}"
            histo=_load_prix_histo()
            prix=histo.setdefault("prix",{})
            if key not in prix: prix[key]={}
            prix[key]["marge_boutique_taux"]=round(taux,4)
            _save_prix_histo(histo)
            messagebox.showinfo("Taux mis \u00e0 jour",
                                f"Taux de marge boutique pour {key} : {taux*100:.2f}%\n\n"
                                f"Pour modifier les taux d'autres mois, \u00e9dite directement\n"
                                f"~/.districarb_hub/prix_historique.cfg",
                                parent=self)
        except ValueError:
            messagebox.showerror("Format invalide","Saisis un nombre (ex : 30 ou 30.27).",parent=self)
        except Exception as e:
            messagebox.showerror("Erreur",f"{e}",parent=self)
    def _set_cycle_week(self,week):
        """Force l'ancrage du cycle : lundi de cette semaine = Semaine X."""
        today=date.today()
        lundi=today-timedelta(days=today.weekday())
        cyc={"anchor_date":lundi.strftime("%Y-%m-%d"),"anchor_week":week,"set_at":datetime.now().isoformat()}
        save_json(CYCLE_FILE,cyc)
        messagebox.showinfo("Cycle mis \u00e0 jour",
                            f"Ancrage : lundi {lundi.strftime('%d/%m/%Y')} = Semaine {week}\n\n"
                            f"Le hub utilisera les bons onglets de Prévision compte au prochain refresh.",
                            parent=self)
        self.destroy()
    def _row(self,parent,fd):
        r=ctk.CTkFrame(parent,fg_color=C["card"],corner_radius=8,border_width=1,border_color=C["border"]);r.pack(fill="x",pady=4)
        top=ctk.CTkFrame(r,fg_color="transparent");top.pack(fill="x",padx=14,pady=(8,3))
        ctk.CTkLabel(top,text=fd["icon"],font=("Segoe UI Emoji",16),text_color=fd["color"]).pack(side="left",padx=(0,6))
        ctk.CTkLabel(top,text=fd["label"],font=("Segoe UI",12,"bold"),text_color=C["t1"]).pack(side="left")
        bot=ctk.CTkFrame(r,fg_color="transparent");bot.pack(fill="x",padx=14,pady=(0,8))
        e=ctk.CTkEntry(bot,height=30,fg_color=C["bg"],border_color=C["border"],text_color=C["t1"],font=("Segoe UI",11),placeholder_text="Aucun chemin")
        e.pack(side="left",fill="x",expand=True,padx=(0,6))
        cur=self.cfg.get(fd["key"],"")
        if cur: e.insert(0,cur)
        ctk.CTkButton(bot,text="Parcourir\u2026",width=95,height=30,fg_color=C["card_h"],hover_color=C["border2"],border_width=1,border_color=C["border2"],text_color=C["t1"],command=lambda en=e:self._pick(en)).pack(side="right")
        self.entries[fd["key"]]=e
    def _pick(self,entry):
        p=filedialog.askopenfilename(title="Fichier Excel",filetypes=[("Excel","*.xlsx *.xlsm *.xls"),("Tous","*.*")])
        if p: entry.delete(0,"end");entry.insert(0,p)
    def _save(self):
        result={}
        for k,e in self.entries.items():
            try:
                val=e.get().strip()
                if val: result[k]=val
            except Exception as _e: _log_silent_err(exc=_e)
        # Préserver les autres clés du config qui ne sont pas dans entries (ex: onedrive_path)
        for k,v in self.cfg.items():
            if k not in result and k not in self.entries:
                result[k]=v
        self.result=result
        self.destroy()

def main():
    ctk.set_appearance_mode("dark");ctk.set_default_color_theme("dark-blue");Hub().mainloop()
if __name__=="__main__": main()
